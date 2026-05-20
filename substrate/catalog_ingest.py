"""Atlas substrate — catalog ingest engine.

Reads the catalog XLSX template (the format Roxy/TLG sent on
2026-05-19) and writes:
  - asin_metadata rows (one per ASIN)
  - outcome_events rows (one per ASIN per metric: revenue, units, sessions)
  - cohort_classifications rows (active vs unknown per the Day 1 rules)
  - brand_workspace stats updated (catalog_size_asins, sales period)

Returns a coverage report:
  {
    rows_loaded, rows_skipped, rows_in_sales,
    column_fill_rate: {col: pct},
    skipped_rules: [list of rule_names that can't fire because data
                    needed is missing — e.g., 'reviews_low_count' when
                    no reviews data on file],
    cohort_counts: {active, unknown, ...},
    period_start, period_end,
  }

The contract of this module: best-effort + transparent. We never
silently invent values. If a cell is missing, it's missing — and the
coverage report tells the operator what they don't have.
"""
from __future__ import annotations

import logging
import os
from datetime import date, datetime
from typing import Any, Optional

logger = logging.getLogger("atlas.substrate.catalog_ingest")


# Canonical column names → {position_index, type, applies_to_substrate}
# Maps the 39 columns of the Roxy template to substrate fields.
TEMPLATE_COLUMNS = [
    "ASIN", "Parent ASIN", "SKU", "UPC", "Style #",
    "Model Name", "Title", "Bullet 1", "Bullet 2", "Bullet 3",
    "Bullet 4", "Bullet 5", "Description", "Backend Keywords",
    "Color", "Size", "Variation Theme", "Parent / Child",
    "Main Image URL", "Other Image URLs", "Image Count", "Video Count",
    "Brand", "Category", "Subcategory", "Item Type Keyword",
    "List Price", "Sale Price", "Buy Box Price", "Buy Box Winner",
    "Quantity", "Fabric / Material", "Country of Origin",
    "Care Instructions", "Item Weight", "Package Dimensions",
    "A+ / EBC Status", "Listing Status", "Fulfillment Method",
]

# Cohort classification thresholds (decision #3 from M6 spec)
ACTIVE_THRESHOLD_UNITS = 1
ACTIVE_THRESHOLD_SESSIONS = 50

# Rules that need data we may not have on every ingest. Surface in the
# skipped_rules list when the column is empty across the catalog.
RULE_DATA_DEPENDENCIES = {
    "missing_country_of_origin":  "Country of Origin",
    "missing_care_instructions":  "Care Instructions",
    "title_missing_brand":        "Brand",
    # The following depend on data not in the standard XLSX template:
    "fewer_than_3_reviews":       "_reviews",
    "below_bsr_floor":            "_bsr",
    "out_of_stock_90d":           "_inventory_status",
}


def _coerce_int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _coerce_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _truthy(v: Any) -> bool:
    return v not in (None, "", "N/A", "n/a", "Required", "Optional")


def parse_workbook(filepath: str) -> dict[str, Any]:
    """Read the XLSX into memory. Returns {catalog_rows, sales_rows,
    cat_headers, sales_headers, errors}.

    Pure parse: no DB writes, no substrate calls. Lets us unit-test
    parsing separately from ingest.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"errors": ["openpyxl not installed"]}

    if not os.path.exists(filepath):
        return {"errors": [f"file not found: {filepath}"]}

    wb = load_workbook(filepath, data_only=True, read_only=True)
    sheets = set(wb.sheetnames)

    catalog_sheet = "Catalog" if "Catalog" in sheets else None
    sales_sheet = next(
        (s for s in sheets if s.lower().startswith("sales")), None,
    )

    if catalog_sheet is None:
        return {"errors": ["No 'Catalog' sheet found"]}

    ws_cat = wb[catalog_sheet]
    cat_headers = [c.value for c in ws_cat[1]]

    catalog_rows: list[dict[str, Any]] = []
    for row in ws_cat.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # no ASIN
            continue
        rec = dict(zip(cat_headers, row))
        catalog_rows.append(rec)

    sales_rows: list[dict[str, Any]] = []
    sales_headers: list[str] = []
    period_start, period_end = None, None
    if sales_sheet:
        ws_sales = wb[sales_sheet]
        sales_headers = [c.value for c in ws_sales[1]]
        for row in ws_sales.iter_rows(min_row=2, values_only=True):
            asin = row[0]
            if not asin or asin in ("Required", "Optional"):
                continue
            rec = dict(zip(sales_headers, row))
            sales_rows.append(rec)
            ps, pe = rec.get("Period Start"), rec.get("Period End")
            if isinstance(ps, datetime) and (period_start is None or ps < period_start):
                period_start = ps
            if isinstance(pe, datetime) and (period_end is None or pe > period_end):
                period_end = pe

    return {
        "errors": [],
        "catalog_rows": catalog_rows,
        "sales_rows": sales_rows,
        "cat_headers": cat_headers,
        "sales_headers": sales_headers,
        "period_start": period_start.date() if period_start else None,
        "period_end": period_end.date() if period_end else None,
    }


def compute_coverage(
    catalog_rows: list[dict[str, Any]],
    cat_headers: list[str],
) -> dict[str, float]:
    """Per-column fill rate across the catalog."""
    if not catalog_rows:
        return {}
    n = len(catalog_rows)
    out: dict[str, float] = {}
    for h in cat_headers:
        if not h:
            continue
        filled = sum(1 for r in catalog_rows if _truthy(r.get(h)))
        out[h] = round(100 * filled / n, 1)
    return out


def detect_skipped_rules(
    coverage: dict[str, float],
    threshold: float = 5.0,
) -> list[str]:
    """A rule is skipped when its data dependency is below threshold pct
    fill rate. Threshold default 5% — a column with <5% fill rate is
    treated as missing for catalog-wide rule firing."""
    skipped = []
    for rule, dep in RULE_DATA_DEPENDENCIES.items():
        if dep.startswith("_"):
            # Connector required — never present in standard XLSX.
            skipped.append(rule)
            continue
        if coverage.get(dep, 0) < threshold:
            skipped.append(rule)
    return skipped


def classify_active_cohort(
    asin: str,
    sales_record: Optional[dict[str, Any]],
) -> tuple[str, dict[str, Any], str]:
    """Apply the Day-1 cohort rule:
       Active = TTM units >= 1 OR sessions >= 50.
       Anything else = Unknown (until inventory + BSR connected).

    Returns (cohort_label, inputs_used_dict, rule_applied).
    """
    if sales_record is None:
        return (
            "unknown",
            {"ttm_units": 0, "ttm_sessions": 0, "in_sales_sheet": False,
             "inventory_status": "missing", "bsr": "missing"},
            "no_sales_record",
        )
    units = _coerce_int(sales_record.get("Units")) or 0
    sessions = _coerce_int(sales_record.get("Sessions")) or 0
    if units >= ACTIVE_THRESHOLD_UNITS or sessions >= ACTIVE_THRESHOLD_SESSIONS:
        return (
            "active",
            {"ttm_units": units, "ttm_sessions": sessions,
             "in_sales_sheet": True},
            f"active_threshold(units>={ACTIVE_THRESHOLD_UNITS} OR "
            f"sessions>={ACTIVE_THRESHOLD_SESSIONS})",
        )
    return (
        "unknown",
        {"ttm_units": units, "ttm_sessions": sessions,
         "in_sales_sheet": True,
         "inventory_status": "missing", "bsr": "missing"},
        "below_active_threshold_no_inventory_data",
    )


def ingest_workbook(
    filepath: str,
    workspace_id: str,
    *,
    write_substrate: bool = True,
    ingested_by: str = "devang",
    progress_cb: Optional[Any] = None,
) -> dict[str, Any]:
    """End-to-end ingest of a single workbook.

    Returns a coverage report. If write_substrate=False, runs the parse
    + classification but skips DB writes — useful for previewing what an
    ingest would do.
    """
    parsed = parse_workbook(filepath)
    if parsed.get("errors"):
        return {"ok": False, "errors": parsed["errors"]}

    catalog_rows = parsed["catalog_rows"]
    sales_rows = parsed["sales_rows"]
    cat_headers = parsed["cat_headers"]

    sales_lookup = {r["ASIN"]: r for r in sales_rows if r.get("ASIN")}

    coverage = compute_coverage(catalog_rows, cat_headers)
    skipped_rules = detect_skipped_rules(coverage)

    # Classify each ASIN
    cohort_rows = []
    asin_to_metadata: list[dict[str, Any]] = []
    asin_to_outcomes: list[dict[str, Any]] = []

    for r in catalog_rows:
        asin = r["ASIN"]
        sr = sales_lookup.get(asin)
        cohort, inputs_used, rule_applied = classify_active_cohort(asin, sr)
        cohort_rows.append({
            "asin": asin, "cohort": cohort,
            "inputs_used": inputs_used, "rule_applied": rule_applied,
        })

        # Build asin_metadata payload (subset; ground_truth_fields gets
        # the columns we know map to canonical fields).
        gtf = {
            "title": r.get("Title"),
            "brand": r.get("Brand"),
            "category": r.get("Category"),
            "subcategory": r.get("Subcategory"),
            "color_name": r.get("Color"),
            "size": r.get("Size"),
            "color_map": r.get("Color"),
            "list_price": _coerce_float(r.get("List Price")),
            "image_count": _coerce_int(r.get("Image Count")),
            "video_count": _coerce_int(r.get("Video Count")),
            "a_plus_status": r.get("A+ / EBC Status"),
            "fulfillment_method": r.get("Fulfillment Method"),
            "country_of_origin": r.get("Country of Origin"),
            "care_instructions": r.get("Care Instructions"),
            "material": r.get("Fabric / Material"),
            "model_name": r.get("Model Name"),
            "style_number": r.get("Style #"),
            "sku": r.get("SKU"),
            "upc": r.get("UPC"),
            "variation_theme": r.get("Variation Theme"),
            "item_type_keyword": r.get("Item Type Keyword"),
            "package_dimensions": r.get("Package Dimensions"),
            "item_weight": r.get("Item Weight"),
            # bullet count is computed; bullets themselves stored separately
            "bullet_count": sum(
                1 for b in (r.get("Bullet 1"), r.get("Bullet 2"),
                            r.get("Bullet 3"), r.get("Bullet 4"),
                            r.get("Bullet 5"))
                if _truthy(b)
            ),
            "title_length": len(r.get("Title") or ""),
            "description_present": _truthy(r.get("Description")),
            "bullets": [
                r.get(f"Bullet {i}") for i in range(1, 6)
                if _truthy(r.get(f"Bullet {i}"))
            ],
            "description": r.get("Description"),
        }
        # Drop empty values to avoid noise in substrate
        gtf = {k: v for k, v in gtf.items() if v not in (None, "")}

        asin_to_metadata.append({
            "asin": asin,
            "parent_asin":
                r.get("Parent ASIN")
                if r.get("Parent ASIN") not in (None, "None", "") else None,
            "ground_truth_fields": gtf,
        })

        # Outcome events (only for sales-sheet rows)
        if sr:
            for metric, value in [
                ("revenue", _coerce_float(sr.get("Revenue"))),
                ("units_sold", _coerce_int(sr.get("Units"))),
                ("sessions", _coerce_int(sr.get("Sessions"))),
            ]:
                if value is None:
                    continue
                asin_to_outcomes.append({
                    "asin": asin, "metric": metric, "value": value,
                    "period_start": parsed.get("period_start"),
                    "period_end": parsed.get("period_end"),
                })

    cohort_counts: dict[str, int] = {}
    for c in cohort_rows:
        cohort_counts[c["cohort"]] = cohort_counts.get(c["cohort"], 0) + 1

    if write_substrate:
        from . import asin_metadata as am
        from . import catalog_audit as ca
        from . import brand_workspace as bw

        def _progress(pct: int, msg: str) -> None:
            if progress_cb:
                try:
                    progress_cb(pct, msg)
                except Exception:
                    pass

        # Workspace registration (idempotent)
        _progress(30, "Registering workspace…")
        bw.register_workspace(
            workspace_id,
            display_name=workspace_id.title(),
            brand_role="audit_only",
            sales_period_start=parsed.get("period_start"),
            sales_period_end=parsed.get("period_end"),
            catalog_size_asins=len(catalog_rows),
        )

        # asin_metadata bulk write — one transaction, batched executemany.
        # The old per-row path was ~5 min on Render's Postgres for 38k
        # ASINs; bulk path is ~15-20s.
        _progress(40, f"Writing {len(asin_to_metadata):,} ASIN metadata rows…")
        wrote_metadata = am.set_asin_metadata_bulk(
            workspace_id, asin_to_metadata,
            set_by=ingested_by,
            bump_revision=False,
        )

        # cohort_classifications bulk
        _progress(75, f"Classifying {len(cohort_rows):,} cohorts…")
        wrote_cohort = ca.classify_cohort_bulk(
            workspace_id, cohort_rows, classified_by="catalog_ingest",
        )

        # outcome_events: write directly to substrate_events table via
        # the existing unit_economics module if available; else skip
        _progress(90, f"Writing {len(asin_to_outcomes):,} outcome events…")
        wrote_outcomes = _write_outcome_events(
            workspace_id, asin_to_outcomes,
        )
    else:
        wrote_metadata = 0
        wrote_cohort = 0
        wrote_outcomes = 0

    return {
        "ok": True,
        "workspace_id": workspace_id,
        "rows_loaded": len(catalog_rows),
        "rows_in_sales": len(sales_rows),
        "metadata_written": wrote_metadata,
        "cohorts_classified": wrote_cohort,
        "outcome_events_written": wrote_outcomes,
        "column_fill_rate": coverage,
        "skipped_rules": skipped_rules,
        "cohort_counts": cohort_counts,
        "period_start":
            parsed["period_start"].isoformat()
            if parsed.get("period_start") else None,
        "period_end":
            parsed["period_end"].isoformat()
            if parsed.get("period_end") else None,
        "errors": [],
    }


def _write_outcome_events(
    workspace_id: str,
    rows: list[dict[str, Any]],
) -> int:
    """Best-effort bulk insert into substrate_events / outcome_events.
    The catalog ingest only knows how to write the standard 3 metrics
    (revenue, units_sold, sessions) on a single TTM cut.
    """
    if not rows:
        return 0
    try:
        from .db import get_pool
    except ImportError:
        return 0
    pool = get_pool()
    if pool is None:
        return 0

    # Try outcome_events first (UNIT_ECONOMICS table); fall back gracefully
    try:
        with pool.connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT to_regclass('outcome_events')")
                if not cur.fetchone()[0]:
                    return 0
                import uuid as _uuid
                values = []
                for r in rows:
                    values.append((
                        str(_uuid.uuid4()),
                        workspace_id, r["asin"], r["metric"],
                        float(r["value"]),
                        r.get("period_start"), r.get("period_end"),
                        "catalog_ingest",
                    ))
                cur.executemany(
                    """
                    INSERT INTO outcome_events (
                        outcome_id, workspace_id, asin, metric, value,
                        period_start, period_end, source_kind
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT DO NOTHING
                    """,
                    values,
                )
            conn.commit()
            return len(values)
    except Exception as exc:
        logger.warning("_write_outcome_events failed: %s", exc)
        return 0


__all__ = [
    "parse_workbook",
    "compute_coverage",
    "detect_skipped_rules",
    "classify_active_cohort",
    "ingest_workbook",
    "TEMPLATE_COLUMNS",
    "ACTIVE_THRESHOLD_UNITS",
    "ACTIVE_THRESHOLD_SESSIONS",
    "RULE_DATA_DEPENDENCIES",
]
