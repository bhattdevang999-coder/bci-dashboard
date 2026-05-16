"""Parsers for marketing input files.

Two formats supported in Phase 1:

- PPC bulk file (Amazon Campaign Manager export, US English):
    columns include `Keyword Text`, `Match Type`, `Bid`, `Impressions`,
    `Clicks`, `Spend`, `Sales`, `ACOS`, `Orders`, `ASIN`, `Campaign Name`,
    `Ad Group Name`. Column names vary slightly across Amazon's exports
    (Sponsored Products vs Sponsored Brands, with/without "(USD)" suffix);
    we match flexibly.

- Search Term Report (US English):
    `Customer Search Term`, `Match Type`, `Impressions`, `Clicks`, `CTR`,
    `Cost Per Click (CPC)`, `Spend`, `Sales`, `ACOS`, `Orders`, `ASIN`.

Each parser returns a list of observation dicts shaped for
substrate.marketing.record_keyword_observations():

    {
        "keyword": "yoga pants",
        "match_type": "broad",
        "asin": "B0XXXX",         # may be empty string
        "metrics": {
            "impressions": 1234, "clicks": 56, "acos": 0.31,
            "spend": 12.34, "orders": 3, ...
        },
    }

The parsers tolerate missing columns. They reject rows where the keyword
itself is missing or empty.
"""
from __future__ import annotations

import csv
import io
import logging
import re
from typing import Any, Optional

logger = logging.getLogger("atlas.substrate.marketing_parsers")


# Header aliases. Lowercase match against the file's headers (case-insensitive,
# whitespace-collapsed, "(usd)" / "%" suffixes stripped).
_HEADER_ALIASES: dict[str, list[str]] = {
    # PPC bulk
    "keyword":       ["keyword text", "keyword", "search term", "customer search term"],
    "match_type":    ["match type", "matchtype"],
    "asin":          ["asin", "advertised asin"],
    "campaign":      ["campaign name", "campaign"],
    "ad_group":      ["ad group name", "ad group"],
    "impressions":   ["impressions"],
    "clicks":        ["clicks"],
    "spend":         ["spend", "cost"],
    "sales":         ["sales", "7 day total sales", "14 day total sales", "30 day total sales"],
    "acos":          ["acos", "total acos"],
    "orders":        ["orders", "7 day total orders (#)", "14 day total orders (#)", "30 day total orders (#)"],
    "ctr":           ["click-thru rate (ctr)", "ctr"],
    "cpc":           ["cost per click (cpc)", "cpc"],
    "bid":           ["bid", "max bid"],
}


def _norm_header(h: str) -> str:
    """Normalise a header for fuzzy matching."""
    s = (h or "").strip().lower()
    # strip currency / percent suffixes
    s = re.sub(r"\s*\(usd\)\s*$", "", s)
    s = re.sub(r"\s*\(%\)\s*$", "", s)
    s = re.sub(r"\s+", " ", s)
    return s


def _build_index(headers: list[str]) -> dict[str, int]:
    """Map canonical key → column index, picking the first matching alias."""
    norm = [_norm_header(h) for h in headers]
    idx: dict[str, int] = {}
    for canonical, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            for i, h in enumerate(norm):
                if h == alias:
                    idx[canonical] = i
                    break
            if canonical in idx:
                break
    return idx


def _to_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    s = str(v).strip()
    if not s:
        return None
    # Strip currency / percent
    neg = s.startswith("(") and s.endswith(")")
    s = s.replace("$", "").replace(",", "").replace("%", "").strip("()")
    try:
        f = float(s)
        return -f if neg else f
    except ValueError:
        return None


def _to_int(v: Any) -> Optional[int]:
    f = _to_float(v)
    return int(f) if f is not None else None


def _normalise_match_type(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    s = s.strip().lower()
    if "exact" in s:  return "exact"
    if "phrase" in s: return "phrase"
    if "broad" in s:  return "broad"
    if "auto" in s:   return "auto"
    return s or None


def _parse_csv(content: bytes) -> tuple[list[str], list[list[str]]]:
    """Parse CSV or TSV bytes. Returns (headers, rows).

    Detects separator by sniffing the first non-empty line.
    """
    text = content.decode("utf-8-sig", errors="replace")
    sample = "\n".join(text.split("\n", 5)[:5])
    sep = "\t" if sample.count("\t") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=sep)
    rows = [r for r in reader if any((c or "").strip() for c in r)]
    if not rows:
        return [], []
    headers = rows[0]
    return headers, rows[1:]


def _parse_xlsx(content: bytes) -> tuple[list[str], list[list[str]]]:
    try:
        import openpyxl
    except ImportError:
        return [], []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        headers_tuple = next(rows_iter, None)
        if headers_tuple is None:
            return [], []
        headers = [str(h).strip() if h is not None else "" for h in headers_tuple]
        rows: list[list[str]] = []
        for r in rows_iter:
            rows.append([str(v) if v is not None else "" for v in r])
        return headers, rows
    except Exception as exc:
        logger.warning("xlsx parse failed: %s", exc)
        return [], []


def _extract_observations(
    headers: list[str],
    rows: list[list[str]],
) -> list[dict[str, Any]]:
    """Shared row-shaping logic for both PPC bulk and STR formats."""
    idx = _build_index(headers)
    if "keyword" not in idx:
        return []
    out: list[dict[str, Any]] = []
    for row in rows:
        if not row:
            continue
        def get(key: str) -> Optional[str]:
            i = idx.get(key)
            if i is None or i >= len(row):
                return None
            v = row[i]
            return v if (v is None or isinstance(v, str)) else str(v)
        kw = (get("keyword") or "").strip()
        if not kw:
            continue
        asin = (get("asin") or "").strip()
        # Build metrics dict, only including present numerical fields.
        metrics: dict[str, Any] = {}
        m = {
            "impressions": _to_int(get("impressions")),
            "clicks":      _to_int(get("clicks")),
            "orders":      _to_int(get("orders")),
            "spend":       _to_float(get("spend")),
            "sales":       _to_float(get("sales")),
            "acos":        _to_float(get("acos")),
            "ctr":         _to_float(get("ctr")),
            "cpc":         _to_float(get("cpc")),
            "bid":         _to_float(get("bid")),
        }
        # ACOS in Amazon exports can be expressed as 0.31 or 31. Normalise to 0..1
        # if it looks like a percentage scale.
        if m["acos"] is not None and m["acos"] > 1.5:
            m["acos"] = m["acos"] / 100.0
        if m["ctr"] is not None and m["ctr"] > 1.5:
            m["ctr"] = m["ctr"] / 100.0
        for k, v in m.items():
            if v is not None:
                metrics[k] = v
        if not metrics:
            # Skip rows with no signal at all.
            continue
        out.append({
            "keyword": kw,
            "match_type": _normalise_match_type(get("match_type")),
            "asin": asin,
            "campaign": (get("campaign") or "").strip() or None,
            "ad_group": (get("ad_group") or "").strip() or None,
            "metrics": metrics,
        })
    return out


def parse_ppc_bulk(content: bytes, file_name: Optional[str] = None) -> list[dict[str, Any]]:
    """Parse a Sponsored Products / Sponsored Brands bulk export.

    Returns a list of observation dicts shaped for
    record_keyword_observations(). Empty list on parse failure.
    """
    is_xlsx = (file_name or "").lower().endswith((".xlsx", ".xlsm")) or content[:4] == b"PK\x03\x04"
    headers, rows = _parse_xlsx(content) if is_xlsx else _parse_csv(content)
    return _extract_observations(headers, rows)


def parse_search_term(content: bytes, file_name: Optional[str] = None) -> list[dict[str, Any]]:
    """Parse a Search Term Report. Same shape as PPC bulk output."""
    is_xlsx = (file_name or "").lower().endswith((".xlsx", ".xlsm")) or content[:4] == b"PK\x03\x04"
    headers, rows = _parse_xlsx(content) if is_xlsx else _parse_csv(content)
    return _extract_observations(headers, rows)


__all__ = ["parse_ppc_bulk", "parse_search_term"]
