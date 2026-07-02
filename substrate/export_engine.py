"""Catalog Intel — export engine (CSV / XLSX / PDF).

Produces client-facing exports of findings with the full verifiability
chain intact:
  - Every finding row carries its rule label, category, severity,
    predicate, threshold, severity logic, and proposed fix.
  - Aggregate findings with per-ASIN resolvers include a companion
    'Affected ASINs' table with reason_tag per ASIN.
  - The rules methodology (all 15 rules with SQL predicates and
    verify queries) is included so an exported document is defensible
    outside the dashboard.

Design:
  - build_xlsx  -> BytesIO (openpyxl workbook, 4 sheets)
  - build_pdf   -> BytesIO (reportlab, cover + findings + methodology)
  - build_csv   -> BytesIO (flat findings only — for spreadsheet upload)

All builders take the same input:
  findings:     list of finding dicts (from get_findings)
  snapshot:     dict with snapshot metadata (id, uploaded_at, file_name, ...)
  workspace_id: str
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
)
from reportlab.lib.enums import TA_LEFT

# Import rule specs (verifiability layer)
from substrate.rules_catalog import RULE_SPECS


# ================================================================
# Helpers
# ================================================================

_SEVERITY_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4, "strategic": 5}


def _severity_key(f: dict) -> tuple:
    return (_SEVERITY_ORDER.get(f.get("severity", "medium"), 9),
            -(f.get("priority_score") or 0))


def _sorted_findings(findings: list) -> list:
    return sorted(findings, key=_severity_key)


def _pretty_evidence_num(v) -> str:
    """Format an evidence value for a spreadsheet cell."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, (int, float)):
        # Percentages
        if isinstance(v, float) and (v < 100 or "pct" in ""):
            return f"{v:.2f}" if abs(v) < 100 else f"{v:,.0f}"
        return f"{v:,}"
    if isinstance(v, list):
        return "; ".join(str(x)[:60] for x in v[:5]) + (f" ({len(v)} total)" if len(v) > 5 else "")
    if isinstance(v, dict):
        # Compact key: val pairs
        return "; ".join(f"{k}={v[k]}" for k in list(v.keys())[:5])
    return str(v)[:200]


def _finding_rule_spec(rule_name: str) -> dict:
    return RULE_SPECS.get(rule_name, {})


# ================================================================
# XLSX builder
# ================================================================

def build_xlsx(findings: list, snapshot: Optional[dict], workspace_id: str,
               affected_asins_by_finding: Optional[dict] = None) -> io.BytesIO:
    """Build a multi-sheet XLSX workbook.

    Sheets:
      1. Findings              — one row per finding, rule spec inline
      2. Affected ASINs        — one row per (finding, asin), with reason_tag
      3. Rules methodology     — one row per rule (all 15)
      4. Snapshot info         — metadata + generation info

    affected_asins_by_finding is a dict {finding_id: [{asin, reason_tag}, ...]}
    passed in from the endpoint (resolvers already ran).
    """
    wb = Workbook()

    # ─── Sheet 1: Findings ───────────────────────────────────────
    ws = wb.active
    ws.title = "Findings"
    _write_findings_sheet(ws, _sorted_findings(findings))

    # ─── Sheet 2: Affected ASINs ─────────────────────────────────
    ws2 = wb.create_sheet("Affected ASINs")
    _write_asins_sheet(ws2, findings, affected_asins_by_finding or {})

    # ─── Sheet 3: Rules Methodology ──────────────────────────────
    ws3 = wb.create_sheet("Rules Methodology")
    _write_methodology_sheet(ws3)

    # ─── Sheet 4: Snapshot Info ──────────────────────────────────
    ws4 = wb.create_sheet("Snapshot Info")
    _write_snapshot_info_sheet(ws4, snapshot, workspace_id, len(findings))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _write_findings_sheet(ws, findings: list) -> None:
    header = [
        "Severity", "Priority", "Rule", "Rule ID", "Category", "Data Source",
        "ASIN", "Proposed Fix",
        "Rule Predicate", "Threshold", "Severity Logic", "Fields Inspected",
        "Key Numbers", "Sample ASINs",
        "Finding ID", "Snapshot ID", "Created At",
    ]
    ws.append(header)
    _style_header_row(ws, len(header))

    for f in findings:
        rule = f.get("rule_name") or ""
        spec = _finding_rule_spec(rule)
        ev = f.get("evidence") or {}
        # Compact "key numbers" cell
        key_numbers = _key_numbers_for_finding(ev)
        sample_asins = ev.get("sample_asins") or []
        row = [
            (f.get("severity") or "").upper(),
            round(float(f.get("priority_score") or 0), 2),
            spec.get("label") or rule.replace("_", " "),
            rule,
            spec.get("category", ""),
            spec.get("data_source", ""),
            f.get("asin") or "",
            f.get("proposed_fix") or "",
            spec.get("predicate", ""),
            spec.get("threshold", ""),
            spec.get("severity_logic", ""),
            ", ".join(spec.get("checks_field") or []),
            key_numbers,
            ", ".join(sample_asins[:20]) + (f" (+{len(sample_asins)-20} more)" if len(sample_asins) > 20 else ""),
            f.get("finding_id") or "",
            f.get("snapshot_id") or "",
            f.get("created_at") or "",
        ]
        ws.append(row)

    # Column widths — tuned for readability
    widths = [9, 8, 34, 26, 12, 14, 12, 60, 60, 40, 40, 24, 40, 40, 38, 38, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Wrap text on prose columns
    for col in [8, 9, 10, 11, 13, 14]:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


def _write_asins_sheet(ws, findings: list, resolved_map: dict) -> None:
    header = ["Finding ID", "Rule", "Severity", "ASIN", "Reason Tag"]
    ws.append(header)
    _style_header_row(ws, len(header))

    for f in _sorted_findings(findings):
        fid = f.get("finding_id")
        rule = f.get("rule_name") or ""
        sev = (f.get("severity") or "").upper()
        # Per-ASIN finding: the finding IS one ASIN
        if f.get("asin"):
            ws.append([fid, rule, sev, f.get("asin"), "per-ASIN finding"])
            continue
        # Aggregate finding: use resolver output if available
        asins = resolved_map.get(fid) or []
        if asins:
            for a in asins:
                ws.append([fid, rule, sev, a.get("asin"), a.get("reason_tag") or ""])
        else:
            # Fall back to sample_asins from evidence
            ev = f.get("evidence") or {}
            for asin_str in (ev.get("sample_asins") or [])[:100]:
                ws.append([fid, rule, sev, asin_str, "(from evidence sample)"])

    widths = [38, 32, 10, 14, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _write_methodology_sheet(ws) -> None:
    header = [
        "Rule ID", "Label", "Category", "Data Source",
        "Predicate", "Threshold", "Severity Logic",
        "Fields Inspected", "Minimum Coverage",
        "SQL Predicate (as executed)", "Standalone Verify Query",
    ]
    ws.append(header)
    _style_header_row(ws, len(header))

    for rule_id in sorted(RULE_SPECS.keys()):
        s = RULE_SPECS[rule_id]
        ws.append([
            rule_id, s.get("label", ""), s.get("category", ""), s.get("data_source", ""),
            s.get("predicate", ""), s.get("threshold", ""), s.get("severity_logic", ""),
            ", ".join(s.get("checks_field") or []) or "(no single field)",
            s.get("min_coverage", ""),
            s.get("sql_predicate", ""), s.get("verify_query", ""),
        ])
    widths = [26, 34, 12, 14, 60, 40, 40, 24, 30, 60, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Wrap on all prose columns
    for col in [5, 6, 7, 9, 10, 11]:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


def _write_snapshot_info_sheet(ws, snapshot: Optional[dict], workspace_id: str,
                               n_findings: int) -> None:
    ws.append(["Field", "Value"])
    _style_header_row(ws, 2)
    generated_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    rows = [
        ("Workspace ID", workspace_id or ""),
        ("Snapshot ID", (snapshot or {}).get("snapshot_id") or ""),
        ("File name", (snapshot or {}).get("file_name") or ""),
        ("Uploaded at", (snapshot or {}).get("uploaded_at") or ""),
        ("Period start", (snapshot or {}).get("period_start") or ""),
        ("Period end", (snapshot or {}).get("period_end") or ""),
        ("Catalog rows", (snapshot or {}).get("row_count_catalog") or ""),
        ("Sales rows", (snapshot or {}).get("row_count_sales") or ""),
        ("Findings in this export", n_findings),
        ("Rules registered", len(RULE_SPECS)),
        ("Exported at (UTC)", generated_at),
        ("Exporter", "Atlas Catalog Intel v0.9"),
    ]
    for r in rows:
        ws.append(list(r))
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60


def _style_header_row(ws, ncols: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="F9FAFB", bold=True, name="Calibri", size=11)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _key_numbers_for_finding(ev: dict) -> str:
    """Extract a compact 'key: value' string of the finding's math."""
    keys_of_interest = [
        "total", "total_catalog", "total_asins",
        "dead_count", "dead_pct",
        "missing_description", "with_description", "pct_with_description",
        "filled", "pct_filled",
        "likely_owner", "likely_owner_pct",
        "active_asins", "top_50pct_asins", "top_80pct_asins", "top_90pct_asins",
        "under_5_pct", "under_3_pct", "inconsistent_pct",
    ]
    parts = []
    for k in keys_of_interest:
        if k in ev and not isinstance(ev[k], (dict, list)):
            parts.append(f"{k}={_pretty_evidence_num(ev[k])}")
    return "; ".join(parts)


# ================================================================
# CSV builder — findings only, flat
# ================================================================

def build_csv(findings: list, snapshot: Optional[dict], workspace_id: str) -> io.BytesIO:
    """Flat CSV of findings, one row each, with rule spec inline."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "severity", "priority", "rule_id", "rule_label", "category", "data_source",
        "asin", "proposed_fix",
        "predicate", "threshold", "severity_logic", "fields_inspected",
        "key_numbers", "sample_asins",
        "finding_id", "snapshot_id", "created_at",
    ])
    for f in _sorted_findings(findings):
        rule = f.get("rule_name") or ""
        spec = _finding_rule_spec(rule)
        ev = f.get("evidence") or {}
        sample_asins = ev.get("sample_asins") or []
        writer.writerow([
            f.get("severity") or "",
            round(float(f.get("priority_score") or 0), 2),
            rule,
            spec.get("label") or rule.replace("_", " "),
            spec.get("category", ""),
            spec.get("data_source", ""),
            f.get("asin") or "",
            f.get("proposed_fix") or "",
            spec.get("predicate", ""),
            spec.get("threshold", ""),
            spec.get("severity_logic", ""),
            "|".join(spec.get("checks_field") or []),
            _key_numbers_for_finding(ev),
            "|".join(sample_asins),
            f.get("finding_id") or "",
            f.get("snapshot_id") or "",
            f.get("created_at") or "",
        ])
    out = io.BytesIO(buf.getvalue().encode("utf-8-sig"))  # utf-8-sig for Excel compatibility
    out.seek(0)
    return out


# ================================================================
# PDF builder
# ================================================================

# Palette (matches dashboard's dark theme adapted for print/light)
COL_PRIMARY = colors.HexColor("#0F172A")
COL_ACCENT  = colors.HexColor("#1E3A5F")
COL_MUTED   = colors.HexColor("#64748B")
COL_TEXT    = colors.HexColor("#28251D")
COL_BORDER  = colors.HexColor("#D4D1CA")
COL_BG      = colors.HexColor("#F9F8F5")
SEV_COLORS = {
    "critical": colors.HexColor("#7f1d1d"),
    "high":     colors.HexColor("#A12C7B"),
    "medium":   colors.HexColor("#964219"),
    "low":      colors.HexColor("#437A22"),
    "info":     colors.HexColor("#64748B"),
    "strategic":colors.HexColor("#0891B2"),
}


def build_pdf(findings: list, snapshot: Optional[dict], workspace_id: str,
              affected_asins_by_finding: Optional[dict] = None) -> io.BytesIO:
    """Professional client-facing PDF: cover + findings + methodology appendix."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=0.7*inch, rightMargin=0.7*inch,
        topMargin=0.7*inch, bottomMargin=0.7*inch,
        title="Atlas Catalog Intel — Audit Report",
        author="Perplexity Computer",
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="AtlasH1", fontName="Helvetica-Bold", fontSize=22, leading=28,
        textColor=COL_PRIMARY, spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name="AtlasH2", fontName="Helvetica-Bold", fontSize=14, leading=18,
        textColor=COL_PRIMARY, spaceBefore=14, spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name="AtlasH3", fontName="Helvetica-Bold", fontSize=11, leading=14,
        textColor=COL_ACCENT, spaceBefore=10, spaceAfter=3,
    ))
    styles.add(ParagraphStyle(
        name="AtlasBody", fontName="Helvetica", fontSize=10, leading=14,
        textColor=COL_TEXT, spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        name="AtlasMuted", fontName="Helvetica", fontSize=9, leading=12,
        textColor=COL_MUTED, spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        name="AtlasCode", fontName="Courier", fontSize=8, leading=11,
        textColor=COL_ACCENT, spaceAfter=6, leftIndent=8,
    ))
    styles.add(ParagraphStyle(
        name="AtlasFix", fontName="Helvetica-Oblique", fontSize=10, leading=14,
        textColor=COL_TEXT, leftIndent=12, spaceAfter=6,
    ))

    story = []
    _pdf_cover(story, styles, findings, snapshot, workspace_id)
    _pdf_findings(story, styles, findings, affected_asins_by_finding or {})
    _pdf_methodology(story, styles)

    doc.build(story)
    buf.seek(0)
    return buf


def _pdf_cover(story, styles, findings, snapshot, workspace_id):
    story.append(Paragraph("Catalog Intel — Audit Report", styles["AtlasH1"]))
    subline_parts = []
    if workspace_id:
        subline_parts.append(f"Workspace: <b>{workspace_id}</b>")
    if snapshot:
        if snapshot.get("uploaded_at"):
            subline_parts.append(f"Snapshot: <b>{snapshot['uploaded_at']}</b>")
        if snapshot.get("file_name"):
            subline_parts.append(f"File: <b>{snapshot['file_name']}</b>")
    if subline_parts:
        story.append(Paragraph("&nbsp;·&nbsp;".join(subline_parts), styles["AtlasMuted"]))
    story.append(Spacer(1, 0.15*inch))

    # Severity tally table
    sev_counts = {}
    for f in findings:
        sev_counts[f.get("severity", "medium")] = sev_counts.get(f.get("severity", "medium"), 0) + 1
    tally_rows = [["Severity", "Findings"]]
    for sev in ["critical", "high", "medium", "low", "info", "strategic"]:
        if sev_counts.get(sev, 0):
            tally_rows.append([sev.upper(), str(sev_counts[sev])])
    tally_rows.append(["TOTAL", str(len(findings))])

    tally = Table(tally_rows, colWidths=[2.5*inch, 1.5*inch], hAlign="LEFT")
    tally.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COL_PRIMARY),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.5, COL_BORDER),
        ("BACKGROUND", (0, -1), (-1, -1), COL_BG),
        ("FONTNAME",   (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    story.append(tally)
    story.append(Spacer(1, 0.15*inch))

    story.append(Paragraph(
        "This report is generated by Atlas Catalog Intel. Every finding "
        "below carries its rule label, predicate, threshold, and severity "
        "logic inline. A full methodology appendix follows the findings — "
        "for every rule, it lists the exact SQL predicate as executed plus "
        "a standalone verify query anyone can run against their own database "
        "to independently reproduce the count.",
        styles["AtlasBody"],
    ))
    story.append(PageBreak())


def _pdf_findings(story, styles, findings, resolved_map: dict):
    story.append(Paragraph("Findings", styles["AtlasH2"]))
    story.append(Paragraph(
        "Sorted by severity, then priority. Each finding includes the rule "
        "predicate and the proposed fix.",
        styles["AtlasMuted"],
    ))
    story.append(Spacer(1, 0.1*inch))

    for i, f in enumerate(_sorted_findings(findings), 1):
        rule = f.get("rule_name") or ""
        spec = _finding_rule_spec(rule)
        sev = (f.get("severity") or "medium").lower()
        title_txt = f"{i}. {spec.get('label') or rule}"
        # Severity chip in the heading
        sev_color = SEV_COLORS.get(sev, COL_MUTED).hexval()
        story.append(Paragraph(
            f'<font color="{sev_color}"><b>{sev.upper()}</b></font>  ·  '
            f'{title_txt}',
            styles["AtlasH3"],
        ))
        # Category + data source metadata
        story.append(Paragraph(
            f"<font color='#64748B'>Category: {spec.get('category','')} · "
            f"Data source: {spec.get('data_source','')}"
            f"{' · ASIN: '+f['asin'] if f.get('asin') else ''}"
            f"</font>",
            styles["AtlasMuted"],
        ))
        # Proposed fix (main content the client reads)
        if f.get("proposed_fix"):
            story.append(Paragraph(f["proposed_fix"], styles["AtlasFix"]))
        # Rule spec inline (verifiability)
        if spec.get("predicate"):
            story.append(Paragraph(
                f"<b>Rule predicate:</b> {spec['predicate']}", styles["AtlasBody"]))
        if spec.get("threshold"):
            story.append(Paragraph(
                f"<b>Threshold:</b> {spec['threshold']}", styles["AtlasBody"]))
        if spec.get("severity_logic"):
            story.append(Paragraph(
                f"<b>Severity logic:</b> {spec['severity_logic']}", styles["AtlasBody"]))
        # Key numbers
        ev = f.get("evidence") or {}
        kn = _key_numbers_for_finding(ev)
        if kn:
            story.append(Paragraph(f"<b>Key numbers:</b> {kn}", styles["AtlasBody"]))
        # Affected ASINs sample
        fid = f.get("finding_id")
        resolved = (resolved_map or {}).get(fid) or []
        sample = ev.get("sample_asins") or []
        if resolved:
            preview = ", ".join(x["asin"] for x in resolved[:10])
            more = len(resolved) - 10
            story.append(Paragraph(
                f"<b>Affected ASINs ({len(resolved)}):</b> {preview}"
                f"{f' (+{more} more)' if more > 0 else ''}",
                styles["AtlasBody"],
            ))
        elif sample:
            preview = ", ".join(sample[:10])
            more = len(sample) - 10
            story.append(Paragraph(
                f"<b>Sample ASINs ({len(sample)}):</b> {preview}"
                f"{f' (+{more} more)' if more > 0 else ''}",
                styles["AtlasBody"],
            ))
        story.append(Spacer(1, 0.08*inch))

    story.append(PageBreak())


def _pdf_methodology(story, styles):
    story.append(Paragraph("Methodology Appendix", styles["AtlasH2"]))
    story.append(Paragraph(
        "For every rule in Catalog Intel, this section lists the exact "
        "check as executed. Each rule ships with a standalone verify query "
        "you can paste into your own database — the count you get back "
        "should match the count reported in the findings above.",
        styles["AtlasMuted"],
    ))
    story.append(Spacer(1, 0.1*inch))

    for rule_id in sorted(RULE_SPECS.keys()):
        s = RULE_SPECS[rule_id]
        story.append(Paragraph(f"{s.get('label') or rule_id}", styles["AtlasH3"]))
        story.append(Paragraph(
            f"<font color='#64748B'>ID: {rule_id} · Category: {s.get('category','')} · "
            f"Data source: {s.get('data_source','')}</font>",
            styles["AtlasMuted"],
        ))
        story.append(Paragraph(f"<b>What this rule checks:</b> {s.get('predicate','')}",
                               styles["AtlasBody"]))
        story.append(Paragraph(f"<b>Threshold:</b> {s.get('threshold','')}",
                               styles["AtlasBody"]))
        story.append(Paragraph(f"<b>Severity logic:</b> {s.get('severity_logic','')}",
                               styles["AtlasBody"]))
        story.append(Paragraph(f"<b>Minimum coverage:</b> {s.get('min_coverage','')}",
                               styles["AtlasBody"]))
        if s.get("checks_field"):
            story.append(Paragraph(
                f"<b>Fields inspected:</b> {', '.join(s['checks_field'])}",
                styles["AtlasBody"],
            ))
        if s.get("sql_predicate"):
            story.append(Paragraph("<b>SQL predicate (as executed):</b>", styles["AtlasBody"]))
            story.append(Paragraph(s["sql_predicate"].replace("<", "&lt;"), styles["AtlasCode"]))
        if s.get("verify_query"):
            story.append(Paragraph("<b>Standalone verify query:</b>", styles["AtlasBody"]))
            story.append(Paragraph(s["verify_query"].replace("<", "&lt;"), styles["AtlasCode"]))
        story.append(Spacer(1, 0.08*inch))
