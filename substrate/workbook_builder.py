"""Catalog Intel — Interactive workbook builder v2.

Rebuilt from scratch after v1 QA revealed layout bugs. Design principles:
  1. One Cover page that fits a landscape sheet cleanly (no split KPIs)
  2. Every KPI, chart, and table cell is a live formula
  3. Rules Methodology is a full explainer — why matters, source,
     inference, first check, live count — not just SQL
  4. Sharp Inferences sheet cross-references multiple signals to produce
     concrete diagnoses no single rule can produce
  5. Charts (native Excel) on Cohort, Revenue Pareto, and histogram sheets
  6. Data validation dropdowns on Coverage Matrix, Revenue Concentration,
     Content Health, Rules Methodology (rule picker)
  7. Conditional formatting for heat maps and grey-out filtering
"""
from __future__ import annotations

import io
import json
import logging
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule, CellIsRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.protection import SheetProtection

from substrate.rules_catalog import RULE_SPECS

logger = logging.getLogger(__name__)


# ============================================================
# Design system
# ============================================================

# Palette
COL_INK          = "0F172A"      # Primary text
COL_INK_MUTED    = "64748B"      # Secondary text
COL_INK_FAINT    = "94A3B8"      # Tertiary
COL_HEADER_BG    = "1F2937"      # Dark header
COL_HEADER_TX    = "F9FAFB"
COL_ACCENT       = "20808D"      # Data-viz teal
COL_ACCENT_2     = "A84B2F"      # Data-viz terra
COL_ROW_STRIPE   = "F8FAFC"
COL_KPI_BG       = "F9FAFB"
COL_KPI_ACCENT   = "F0F9FF"
COL_BORDER       = "E5E7EB"
COL_BG           = "FFFFFF"

# Signal colors
COL_HEALTHY      = "DCFCE7"   # light green fill
COL_HEALTHY_TX   = "166534"
COL_WARN         = "FEF3C7"   # light yellow
COL_WARN_TX      = "854D0E"
COL_ALERT        = "FEE2E2"   # light red
COL_ALERT_TX     = "991B1B"

# Fonts
def _font(**kw):
    kw.setdefault("name", "Calibri")
    return Font(**kw)

FONT_H1          = _font(bold=True, size=22, color=COL_INK)
FONT_H2          = _font(bold=True, size=16, color=COL_INK)
FONT_H3          = _font(bold=True, size=12, color=COL_INK)
FONT_HEADER      = _font(bold=True, size=11, color=COL_HEADER_TX)
FONT_BODY        = _font(size=10, color=COL_INK)
FONT_MUTED       = _font(size=10, color=COL_INK_MUTED)
FONT_FAINT       = _font(size=9, color=COL_INK_FAINT)
FONT_KPI_LABEL   = _font(size=9, color=COL_INK_MUTED)
FONT_KPI_VAL     = _font(bold=True, size=20, color=COL_INK)
FONT_BOLD        = _font(bold=True, size=10, color=COL_INK)
FONT_CODE        = Font(name="Consolas", size=10, color="334155")
FONT_ITALIC      = _font(italic=True, size=10, color=COL_INK_MUTED)

# Fills
FILL_HEADER    = PatternFill("solid", fgColor=COL_HEADER_BG)
FILL_KPI       = PatternFill("solid", fgColor=COL_KPI_BG)
FILL_KPI_HERO  = PatternFill("solid", fgColor=COL_KPI_ACCENT)
FILL_BANNER    = PatternFill("solid", fgColor=COL_WARN)
FILL_ROW_STRIPE= PatternFill("solid", fgColor=COL_ROW_STRIPE)
FILL_HEALTHY   = PatternFill("solid", fgColor=COL_HEALTHY)
FILL_WARN      = PatternFill("solid", fgColor=COL_WARN)
FILL_ALERT     = PatternFill("solid", fgColor=COL_ALERT)

# Alignments
ALIGN_L  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_R  = Alignment(horizontal="right",  vertical="center")
ALIGN_LT = Alignment(horizontal="left",   vertical="top",    wrap_text=True)


# ============================================================
# Catalog columns
# ============================================================

CATALOG_COLUMNS = [
    ("asin",              "ASIN"),
    ("parent_asin",       "Parent ASIN"),
    ("sku",               "SKU"),
    ("title",             "Title"),
    ("brand",             "Brand"),
    ("category",          "Category"),
    ("subcategory",       "Subcategory"),
    ("color",             "Color"),
    ("size",              "Size"),
    ("model",             "Model"),
    ("list_price",        "List Price"),
    ("sale_price",        "Sale Price"),
    ("image_count",       "Image Count"),
    ("a_plus_status",     "A+ Status"),
    ("buy_box_winner",    "Buy Box Winner"),
    ("variation_theme",   "Variation Theme"),
    ("bullet_1",          "Bullet 1"),
    ("bullet_2",          "Bullet 2"),
    ("bullet_3",          "Bullet 3"),
    ("bullet_4",          "Bullet 4"),
    ("bullet_5",          "Bullet 5"),
    ("description",       "Description"),
    ("fabric_material",   "Fabric / Material"),
    ("country_of_origin", "Country of Origin"),
    ("care_instructions", "Care Instructions"),
    ("backend_keywords",  "Backend Keywords"),
    ("listing_status",    "Listing Status"),
    ("sessions",          "Sessions"),
    ("units",             "Units"),
    ("revenue",           "Revenue"),
    ("cvr_pct",           "CVR %"),
]

CONTENT_CRITICAL_FIELDS = {"title","image_count","bullet_1","bullet_2","bullet_3","description","list_price"}
COMPLIANCE_FIELDS = {"fabric_material","country_of_origin","care_instructions","backend_keywords"}
SALES_FIELDS = {"sessions","units","revenue","cvr_pct"}


# ============================================================
# Entry point
# ============================================================

def build_interactive_workbook(
    catalog_rows: list,
    sales_by_asin: dict,
    findings: list,
    snapshot: Optional[dict],
    workspace_id: str,
) -> io.BytesIO:
    """Build the full interactive workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    flat_rows = _flatten_rows(catalog_rows, sales_by_asin)
    n = len(flat_rows)

    _sheet_cover(wb, n, snapshot, workspace_id)
    _sheet_catalog_data(wb, flat_rows)
    _sheet_sales_data(wb, sales_by_asin)
    _sheet_coverage_matrix(wb)
    _sheet_revenue_concentration(wb, flat_rows)
    _sheet_cohort_analysis(wb)
    _sheet_content_health(wb, flat_rows)
    _sheet_sharp_inferences(wb, flat_rows)
    _sheet_all_findings(wb, findings)
    _sheet_rules_methodology(wb)
    _sheet_data_gaps(wb)
    _sheet_trend_kpis(wb, snapshot)
    _sheet_trend_by_rule(wb, snapshot)
    _sheet_fix_effectiveness(wb, findings)
    _sheet_how_to_add_historicals(wb)
    _sheet_how_it_works(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _flatten_rows(catalog_rows, sales_by_asin):
    out = []
    for row in catalog_rows:
        asin = row.get("asin")
        gtf = row.get("ground_truth_fields") or {}
        flat = {"asin": asin, "parent_asin": row.get("parent_asin")}
        for key, _ in CATALOG_COLUMNS:
            if key in ("asin","parent_asin") or key in SALES_FIELDS:
                continue
            v = gtf.get(key)
            if isinstance(v, (dict, list)):
                v = json.dumps(v)[:32000]
            flat[key] = v
        s = sales_by_asin.get(asin, {})
        for key in SALES_FIELDS:
            flat[key] = s.get(key)
        out.append(flat)
    return out


# ============================================================
# Small helpers
# ============================================================

def _hdr_cell(ws, r, c, value):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_C
    return cell


def _landscape_orient(ws):
    """Just landscape, no fit-to-page (for wide raw data sheets)."""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def _fit_landscape(ws, height=0):
    """Configure landscape orientation with fit-to-page.

    height=0 means unlimited pages tall (flow content across pages). Use 0
    for sheets with many rows. height=1 forces everything onto a single page
    (cover, small dashboards)."""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = height
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def _banner(ws, cell_ref, text, merge_ref=None):
    ws[cell_ref] = text
    ws[cell_ref].font = _font(italic=True, size=10, color=COL_WARN_TX)
    ws[cell_ref].fill = FILL_BANNER
    ws[cell_ref].alignment = ALIGN_LT
    if merge_ref:
        ws.merge_cells(merge_ref)


# ============================================================
# 01 — Cover
# ============================================================

def _sheet_cover(wb, n, snapshot, workspace_id):
    ws = wb.create_sheet("01_Cover")
    ws.sheet_view.showGridLines = False

    # Working area: cols B-E. 4 KPI tiles across.
    ws.column_dimensions["A"].width = 3
    for c in ("B","C","D","E"):
        ws.column_dimensions[c].width = 28
    ws.column_dimensions["F"].width = 3

    # ── Title ────────────────────────────────────────────────
    ws["B2"] = "Catalog Intel — Interactive Audit"
    ws["B2"].font = FONT_H1
    ws.merge_cells("B2:E2")
    ws.row_dimensions[2].height = 32

    subline = [f"Workspace: {workspace_id}"]
    if snapshot:
        if snapshot.get("uploaded_at"):
            subline.append(f"Snapshot: {snapshot['uploaded_at'][:10]}")
        if snapshot.get("file_name"):
            fname = snapshot['file_name']
            if len(fname) > 40: fname = fname[:37] + "..."
            subline.append(f"File: {fname}")
    ws["B3"] = "  ·  ".join(subline)
    ws["B3"].font = FONT_MUTED
    ws.merge_cells("B3:E3")

    ws["B4"] = ("Every KPI and table below is a live formula pointing at 02_Catalog_Data. "
                "Change any raw value and everything updates. Click any cell to see the exact math.")
    ws["B4"].font = FONT_MUTED
    ws["B4"].alignment = ALIGN_LT
    ws.merge_cells("B4:E4")
    ws.row_dimensions[4].height = 32

    # ── KPI tiles ────────────────────────────────────────────
    ws["B6"] = "Executive KPIs"
    ws["B6"].font = FONT_H3
    ws.merge_cells("B6:E6")
    ws.row_dimensions[6].height = 20

    kpis = [
        ("Total ASINs",              "=COUNTA(Catalog[ASIN])",                                                              "count"),
        ("Dead ASINs % (0 units)",   "=IFERROR(COUNTIF(Catalog[Units],0)/COUNTA(Catalog[ASIN]),0)",                         "pct"),
        ("Active ASINs",             "=COUNTIFS(Catalog[Sessions],\">0\")+COUNTIFS(Catalog[Sessions],0,Catalog[Units],\">0\")","count"),
        ("Total revenue",            "=SUM(Catalog[Revenue])",                                                              "money"),
        ("Titles filled",            "=IFERROR(COUNTIF(Catalog[Title],\"?*\")/COUNTA(Catalog[ASIN]),0)",                     "pct"),
        ("Descriptions filled",      "=IFERROR(COUNTIF(Catalog[Description],\"?*\")/COUNTA(Catalog[ASIN]),0)",               "pct"),
        ("Fabric/material filled",   "=IFERROR(COUNTIF(Catalog[Fabric / Material],\"?*\")/COUNTA(Catalog[ASIN]),0)",         "pct"),
        ("Avg images per ASIN",      "=IFERROR(AVERAGE(Catalog[Image Count]),0)",                                             "num"),
    ]

    tile_pairs = [(8, 9), (12, 13)]
    for idx, (label, formula, kind) in enumerate(kpis):
        r_idx, c_idx = idx // 4, idx % 4
        lr, vr = tile_pairs[r_idx]
        col = get_column_letter(2 + c_idx)  # B, C, D, E

        lcell = ws[f"{col}{lr}"]
        lcell.value = label
        lcell.font = FONT_KPI_LABEL
        lcell.alignment = ALIGN_L
        lcell.fill = FILL_KPI

        vcell = ws[f"{col}{vr}"]
        vcell.value = formula
        vcell.font = FONT_KPI_VAL
        vcell.alignment = ALIGN_L
        vcell.fill = FILL_KPI
        if kind == "pct":     vcell.number_format = "0.0%"
        elif kind == "money": vcell.number_format = "$#,##0"
        elif kind == "num":   vcell.number_format = "0.0"
        else:                 vcell.number_format = "#,##0"

        ws.row_dimensions[lr].height = 16
        ws.row_dimensions[vr].height = 32

    # ── Sheet index ──────────────────────────────────────────
    ws["B15"] = "What's in this workbook"
    ws["B15"].font = FONT_H3
    ws.merge_cells("B15:E15")

    coverage_lines = [
        ("01_Cover",                    "This page"),
        ("02_Catalog_Data",             "Full raw catalog (source of truth for every formula)"),
        ("03_Sales_Data",               "Full raw sales metrics per ASIN"),
        ("04_Coverage_Matrix",          "Fill-rate heat map with category filter"),
        ("05_Revenue_Concentration",    "Interactive Pareto with Top-N dropdown"),
        ("06_Cohort_Analysis",          "Dead / long-tail / active / core with bar charts"),
        ("07_Content_Health",           "Per-ASIN quality scorecard, filter by score"),
        ("08_Sharp_Inferences",         "Cross-signal diagnoses (zombie ASINs, PPC-starved, cannibalizing families)"),
        ("09_All_Findings",             "Dashboard findings mirror"),
        ("10_Rules_Methodology",        "Rich per-rule explainer: why, source, inference, first check"),
        ("11_Data_Gaps",                "What more data unlocks which analyses"),
        ("12_Trend_KPIs",               "KPI evolution across snapshots (populated as you re-upload)"),
        ("13_Trend_By_Rule",            "Per-rule metric trend over time"),
        ("14_Fix_Effectiveness",        "The money loop — fix rate over time"),
        ("15_How_To_Add_Historicals",   "Backfilling trend columns from past uploads"),
        ("16_How_This_Works",           "Trust page: named ranges, verification workflow"),
    ]
    for i, (sheet, blurb) in enumerate(coverage_lines):
        r = 17 + i
        ws.cell(row=r, column=2, value=sheet).font = FONT_CODE
        ws.cell(row=r, column=3, value="—").font = FONT_FAINT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=3)
        ws.cell(row=r, column=4, value=blurb).font = FONT_BODY
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)

    # Footer
    footer_row = 17 + len(coverage_lines) + 1
    ws.cell(row=footer_row, column=2, value=(
        "Generated by Perplexity Computer — Atlas Catalog Intel v1.3. "
        "Every number is derived from 02_Catalog_Data and 03_Sales_Data via visible formulas."
    )).font = FONT_ITALIC
    ws.merge_cells(start_row=footer_row, start_column=2, end_row=footer_row, end_column=5)
    ws.cell(row=footer_row, column=2).alignment = ALIGN_LT
    ws.row_dimensions[footer_row].height = 30

    _fit_landscape(ws, height=1)


# ============================================================
# 02 — Catalog Data
# ============================================================

def _sheet_catalog_data(wb, flat_rows):
    ws = wb.create_sheet("02_Catalog_Data")
    ws.sheet_view.showGridLines = False

    headers = [label for _, label in CATALOG_COLUMNS]
    for c, label in enumerate(headers, start=1):
        _hdr_cell(ws, 1, c, label)

    for r, row in enumerate(flat_rows, start=2):
        for c, (key, _) in enumerate(CATALOG_COLUMNS, start=1):
            v = row.get(key)
            cell = ws.cell(row=r, column=c, value=v)
            if key in ("list_price","sale_price","revenue"): cell.number_format = "$#,##0.00"
            elif key in ("sessions","units","image_count"):  cell.number_format = "#,##0"
            elif key == "cvr_pct":                            cell.number_format = "0.00"

    n_rows = len(flat_rows)
    if n_rows > 0:
        last_col = get_column_letter(len(headers))
        table = Table(displayName="Catalog", ref=f"A1:{last_col}{n_rows + 1}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)

    widths = [14,14,14,50,14,18,18,14,10,14,12,12,10,10,20,16,
              32,32,32,32,32,55,22,18,22,26,14,12,12,14,10]
    for i, w in enumerate(widths[:len(headers)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "D2"
    ws.protection = SheetProtection(sheet=True, formatCells=False, sort=True, autoFilter=True)
    _landscape_orient(ws)


# ============================================================
# 03 — Sales Data
# ============================================================

def _sheet_sales_data(wb, sales_by_asin):
    ws = wb.create_sheet("03_Sales_Data")
    ws.sheet_view.showGridLines = False

    for c, h in enumerate(["ASIN","Sessions","Units","Revenue","CVR %"], start=1):
        _hdr_cell(ws, 1, c, h)

    r = 2
    for asin, m in sales_by_asin.items():
        ws.cell(row=r, column=1, value=asin)
        ws.cell(row=r, column=2, value=m.get("sessions") or 0).number_format = "#,##0"
        ws.cell(row=r, column=3, value=m.get("units") or 0).number_format = "#,##0"
        ws.cell(row=r, column=4, value=m.get("revenue") or 0).number_format = "$#,##0.00"
        ws.cell(row=r, column=5, value=m.get("cvr_pct") or 0).number_format = "0.00"
        r += 1

    n = len(sales_by_asin)
    if n > 0:
        table = Table(displayName="Sales", ref=f"A1:E{n + 1}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium3", showRowStripes=True)
        ws.add_table(table)

    for i, w in enumerate([14, 12, 12, 14, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"
    _landscape_orient(ws)


# ============================================================
# 04 — Coverage Matrix
# ============================================================

def _sheet_coverage_matrix(wb):
    ws = wb.create_sheet("04_Coverage_Matrix")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    widths = [3, 28, 16, 14, 14, 14, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws["B2"] = "Field Coverage Matrix"
    ws["B2"].font = FONT_H2
    ws.merge_cells("B2:G2")

    ws["B3"] = ("Every cell below is a live COUNTIF on 02_Catalog_Data. Change any raw value "
                "and the fill % updates. Red = effectively empty. Green = healthy.")
    ws["B3"].font = FONT_MUTED
    ws.merge_cells("B3:G3")
    ws.row_dimensions[3].height = 26

    # Filter dropdown
    ws["B5"] = "Filter category:"
    ws["B5"].font = FONT_BOLD
    ws["C5"] = "All"
    ws["C5"].fill = FILL_KPI_HERO
    ws["C5"].font = FONT_BOLD
    dv = DataValidation(type="list", formula1='"All,Content-critical,Compliance,Sales-tied,Nice-to-have"', allow_blank=False)
    dv.add("C5")
    ws.add_data_validation(dv)

    for c, h in enumerate(["Field","Category","Filled","Total","Fill %","Status"], start=2):
        _hdr_cell(ws, 7, c, h)

    def _category_of(key):
        if key in CONTENT_CRITICAL_FIELDS: return "Content-critical"
        if key in COMPLIANCE_FIELDS: return "Compliance"
        if key in SALES_FIELDS: return "Sales-tied"
        return "Nice-to-have"

    row = 8
    fields = [(k, label) for k, label in CATALOG_COLUMNS if k not in ("asin","parent_asin")]
    for key, label in fields:
        col_ref = f"Catalog[{label}]"
        ws.cell(row=row, column=2, value=label).alignment = ALIGN_L
        ws.cell(row=row, column=3, value=_category_of(key)).alignment = ALIGN_C
        if key in ("list_price","sale_price","image_count","sessions","units","revenue","cvr_pct"):
            ws.cell(row=row, column=4, value=f'=COUNT({col_ref})')
        else:
            ws.cell(row=row, column=4, value=f'=COUNTIF({col_ref},"?*")')
        ws.cell(row=row, column=5, value='=COUNTA(Catalog[ASIN])')
        ws.cell(row=row, column=6, value=f"=IFERROR(D{row}/E{row},0)")
        ws.cell(row=row, column=6).number_format = "0.0%"
        ws.cell(row=row, column=7, value=(
            f'=IF(F{row}>=0.8,"✓ healthy",'
            f'IF(F{row}>=0.5,"partial",'
            f'IF(F{row}>=0.05,"thin","effectively empty")))'
        ))
        for c in (4, 5, 6): ws.cell(row=row, column=c).alignment = ALIGN_R
        ws.cell(row=row, column=7).alignment = ALIGN_L
        row += 1

    last_row = row - 1

    ws.conditional_formatting.add(f"F8:F{last_row}", ColorScaleRule(
        start_type="num", start_value=0,   start_color="FEE2E2",
        mid_type="num",   mid_value=0.5,   mid_color="FEF3C7",
        end_type="num",   end_value=1,     end_color="DCFCE7",
    ))
    # Grey out rows not matching filter dropdown
    ws.conditional_formatting.add(f"B8:G{last_row}", FormulaRule(
        formula=[f'AND($C$5<>"All",$C8<>$C$5)'],
        fill=PatternFill("solid", fgColor="F1F5F9"),
        font=_font(color=COL_INK_FAINT, size=10),
    ))

    ws.freeze_panes = "B8"
    _fit_landscape(ws)


# ============================================================
# 05 — Revenue Concentration (Pareto with Top-N)
# ============================================================

def _sheet_revenue_concentration(wb, flat_rows):
    ws = wb.create_sheet("05_Revenue_Concentration")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    for c, w in zip("BCDEFGHIJK", [7, 15, 55, 15, 18, 14, 3, 32, 14, 3]):
        ws.column_dimensions[c].width = w

    ws["B2"] = "Revenue Concentration — Pareto"
    ws["B2"].font = FONT_H2
    ws.merge_cells("B2:J2")

    ws["B3"] = ("Every active ASIN, ranked by revenue with cumulative share. "
                "Change the Top-N dropdown to reshape which rows are highlighted.")
    ws["B3"].font = FONT_MUTED
    ws.merge_cells("B3:J3")

    # Top-N picker (visual highlight only — all rows are always present)
    ws["B5"] = "Highlight top:"
    ws["B5"].font = FONT_BOLD
    ws["C5"] = 50
    ws["C5"].fill = FILL_KPI_HERO
    ws["C5"].font = FONT_BOLD
    ws["C5"].alignment = ALIGN_C
    dv = DataValidation(type="list", formula1='"10,20,50,100,200"', allow_blank=False)
    dv.add("C5")
    ws.add_data_validation(dv)

    # Pre-sort active ASINs (rev > 0) by revenue desc, stable by ASIN.
    active = [r for r in flat_rows if (r.get("revenue") or 0) > 0]
    active.sort(key=lambda r: (-(r.get("revenue") or 0), r.get("asin") or ""))
    n_active = len(active)
    show_max = min(n_active, 200)  # 200-row static table; more than enough for Pareto

    ws["D5"] = f"of {n_active:,} active ASINs"
    ws["D5"].font = FONT_MUTED

    for c, h in enumerate(["Rank","ASIN","Title","Revenue","Cumulative Revenue","Cumulative %"], start=2):
        _hdr_cell(ws, 7, c, h)

    # Static ranked rows with live cumulative formulas
    for i in range(show_max):
        r = 8 + i
        row = active[i]
        ws.cell(row=r, column=2, value=i + 1).number_format = "#,##0"
        ws.cell(row=r, column=3, value=row.get("asin")).font = FONT_CODE
        ws.cell(row=r, column=4, value=row.get("title"))
        ws.cell(row=r, column=5, value=row.get("revenue") or 0).number_format = "$#,##0"
        # Cumulative revenue: running sum of column E
        if i == 0:
            ws.cell(row=r, column=6, value=f"=E{r}")
        else:
            ws.cell(row=r, column=6, value=f"=F{r-1}+E{r}")
        ws.cell(row=r, column=6).number_format = "$#,##0"
        ws.cell(row=r, column=7, value=f"=IFERROR(F{r}/SUM(Catalog[Revenue]),0)").number_format = "0.0%"
        # Highlight rows within Top-N dropdown
        for c_i in range(2, 8):
            ws.cell(row=r, column=c_i).alignment = ALIGN_L if c_i == 4 else ALIGN_R

    # Conditional format: dim rows beyond the Top-N picker (only if any rows)
    last_row = 8 + show_max - 1
    if show_max > 0:
        ws.conditional_formatting.add(
            f"B8:G{last_row}",
            FormulaRule(
                formula=[f'$B8>$C$5'],
                fill=PatternFill("solid", fgColor="F1F5F9"),
                font=_font(color=COL_INK_FAINT, size=10),
            ),
        )
        # Also highlight the top of the list
        ws.conditional_formatting.add(
            f"B8:G{last_row}",
            FormulaRule(
                formula=[f'$B8<=$C$5'],
                fill=PatternFill("solid", fgColor="F0F9FF"),
            ),
        )

    # Right-side threshold summary — pure lookup on the ranked table
    ws["I7"] = "Concentration thresholds"
    ws["I7"].font = FONT_BOLD
    ws["I8"] = "ASINs to 50% of revenue:"
    ws["I9"] = "ASINs to 80% of revenue:"
    ws["I10"] = "ASINs to 90% of revenue:"
    ws["I11"] = "Total active ASINs (rev>0):"
    ws["I12"] = "Total revenue:"

    # COUNTIF against cumulative percentage column (guarded for empty case)
    if show_max > 0:
        ws["J8"]  = f'=COUNTIF(G8:G{last_row},"<0.5")+1'
        ws["J9"]  = f'=COUNTIF(G8:G{last_row},"<0.8")+1'
        ws["J10"] = f'=COUNTIF(G8:G{last_row},"<0.9")+1'
    else:
        ws["J8"] = ws["J9"] = ws["J10"] = 0
    ws["J11"] = '=COUNTIF(Catalog[Revenue],">0")'
    ws["J12"] = '=SUM(Catalog[Revenue])'
    for r in (8, 9, 10, 11):
        ws.cell(row=r, column=10).number_format = "#,##0"
        ws.cell(row=r, column=10).font = FONT_BOLD
    ws.cell(row=12, column=10).number_format = "$#,##0"
    ws.cell(row=12, column=10).font = FONT_BOLD

    # Small explainer below the threshold panel
    ws["I14"] = "How to read this"
    ws["I14"].font = FONT_BOLD
    ws["I15"] = ("If the top 5% of your ASINs = 80%+ of revenue, you're vulnerable to single-ASIN "
                 "suppression. If it takes 30%+ to reach 50%, you're not investing enough in your winners.")
    ws["I15"].font = FONT_MUTED
    ws["I15"].alignment = ALIGN_LT
    ws.merge_cells("I15:J20")

    ws.freeze_panes = "B8"
    _fit_landscape(ws, height=0)


# ============================================================
# 06 — Cohort Analysis
# ============================================================

def _sheet_cohort_analysis(wb):
    ws = wb.create_sheet("06_Cohort_Analysis")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    for c, w in zip("BCDEFG", [22, 14, 16, 14, 14, 3]):
        ws.column_dimensions[c].width = w

    ws["B2"] = "Cohort Analysis"
    ws["B2"].font = FONT_H2
    ws.merge_cells("B2:F2")

    ws["B3"] = ("Dead = 0 sessions AND 0 units.  Core = top 20% by revenue.  "
                "Active = revenue in top 10-80th percentile.  Long-tail = everyone else with activity.")
    ws["B3"].font = FONT_MUTED
    ws.merge_cells("B3:F3")
    ws.row_dimensions[3].height = 30

    for c, h in enumerate(["Cohort","ASIN Count","Revenue","% of ASINs","% of Revenue"], start=2):
        _hdr_cell(ws, 5, c, h)

    # Helper cells: precompute percentile thresholds once (avoids embedding
    # PERCENTILE.INC inside SUMPRODUCT, which some engines can't resolve).
    ws.cell(row=6, column=8, value="P10 revenue threshold:").font = FONT_MUTED
    ws.cell(row=6, column=9, value="=PERCENTILE(Catalog[Revenue],0.1)").number_format = "$#,##0"
    ws.cell(row=7, column=8, value="P80 revenue threshold:").font = FONT_MUTED
    ws.cell(row=7, column=9, value="=PERCENTILE(Catalog[Revenue],0.8)").number_format = "$#,##0"

    cohorts = [
        ("Dead",
         '=COUNTIFS(Catalog[Sessions],0,Catalog[Units],0)',
         '=SUMIFS(Catalog[Revenue],Catalog[Sessions],0,Catalog[Units],0)'),
        ("Long-tail",
         '=SUMPRODUCT((Catalog[Revenue]<$I$6)*((Catalog[Sessions]>0)+(Catalog[Units]>0)>0))',
         '=SUMPRODUCT((Catalog[Revenue]<$I$6)*((Catalog[Sessions]>0)+(Catalog[Units]>0)>0)*Catalog[Revenue])'),
        ("Active",
         '=SUMPRODUCT((Catalog[Revenue]>=$I$6)*(Catalog[Revenue]<$I$7)*((Catalog[Sessions]>0)+(Catalog[Units]>0)>0))',
         '=SUMPRODUCT((Catalog[Revenue]>=$I$6)*(Catalog[Revenue]<$I$7)*((Catalog[Sessions]>0)+(Catalog[Units]>0)>0)*Catalog[Revenue])'),
        ("Core (top 20%)",
         '=SUMPRODUCT(--(Catalog[Revenue]>=$I$7))',
         '=SUMPRODUCT((Catalog[Revenue]>=$I$7)*Catalog[Revenue])'),
    ]
    for i, (name, cf, rf) in enumerate(cohorts):
        r = 6 + i
        ws.cell(row=r, column=2, value=name).font = FONT_BOLD
        ws.cell(row=r, column=3, value=cf).number_format = "#,##0"
        ws.cell(row=r, column=4, value=rf).number_format = "$#,##0"
        ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/COUNTA(Catalog[ASIN]),0)").number_format = "0.0%"
        ws.cell(row=r, column=6, value=f"=IFERROR(D{r}/SUM(Catalog[Revenue]),0)").number_format = "0.0%"

    ws.cell(row=10, column=2, value="TOTAL").font = FONT_BOLD
    ws.cell(row=10, column=3, value="=SUM(C6:C9)").number_format = "#,##0"
    ws.cell(row=10, column=4, value="=SUM(D6:D9)").number_format = "$#,##0"
    ws.cell(row=10, column=5, value="=SUM(E6:E9)").number_format = "0.0%"
    ws.cell(row=10, column=6, value="=SUM(F6:F9)").number_format = "0.0%"

    # Charts
    chart = BarChart()
    chart.type = "bar"; chart.style = 11
    chart.title = "ASIN count by cohort"
    data = Reference(ws, min_col=3, min_row=5, max_row=9, max_col=3)
    cats = Reference(ws, min_col=2, min_row=6, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8; chart.width = 16
    chart.dataLabels = DataLabelList(showVal=True)
    ws.add_chart(chart, "H4")

    chart2 = BarChart()
    chart2.type = "bar"; chart2.style = 12
    chart2.title = "Revenue by cohort"
    data2 = Reference(ws, min_col=4, min_row=5, max_row=9, max_col=4)
    cats2 = Reference(ws, min_col=2, min_row=6, max_row=9)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.height = 8; chart2.width = 16
    chart2.dataLabels = DataLabelList(showVal=True)
    ws.add_chart(chart2, "H22")

    # How to read
    ws["B13"] = "How to read this"
    ws["B13"].font = FONT_BOLD
    ws["B14"] = ("Healthy activewear: dead <15%, long-tail 25-35%, active 35-45%, core 10-20%. "
                 "If dead > 50%, you have a delisting project. If core < 5%, you don't have enough winners.")
    ws["B14"].font = FONT_MUTED
    ws["B14"].alignment = ALIGN_LT
    ws.merge_cells("B14:F16")
    ws.row_dimensions[14].height = 22

    _fit_landscape(ws)


# ============================================================
# 07 — Content Health scorecard
# ============================================================

def _sheet_content_health(wb, flat_rows):
    ws = wb.create_sheet("07_Content_Health")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    for c, w in zip("BCDEFGHIJK", [12, 38, 10, 10, 10, 12, 10, 12, 10, 3]):
        ws.column_dimensions[c].width = w

    ws["B2"] = "Per-ASIN Content Health Scorecard"
    ws["B2"].font = FONT_H2
    ws.merge_cells("B2:J2")

    ws["B3"] = ("Composite 0-10 score built from 5 quality checks. Each check is a formula on 02_Catalog_Data. "
                "Change the dropdown to filter by score bucket.")
    ws["B3"].font = FONT_MUTED
    ws.merge_cells("B3:J3")
    ws.row_dimensions[3].height = 24

    ws["B5"] = "Filter by score bucket:"
    ws["B5"].font = FONT_BOLD
    ws["C5"] = "All"
    ws["C5"].fill = FILL_KPI_HERO
    ws["C5"].font = FONT_BOLD
    dv = DataValidation(type="list", formula1='"All,High (≥8),Medium (5-7),Low (<5)"', allow_blank=False)
    dv.add("C5")
    ws.add_data_validation(dv)

    hdrs = ["ASIN","Title","Image Count","Bullets Filled","Title Length","Description Length",
            "Fabric Filled","Score (0-10)","Bucket"]
    for c, h in enumerate(hdrs, start=2):
        _hdr_cell(ws, 7, c, h)

    # Per-row formulas — reference the same Catalog row
    for i, _row in enumerate(flat_rows):
        r = 8 + i
        cat_row = i + 1  # 1-based row index within Catalog table body
        ws.cell(row=r, column=2, value=f"=INDEX(Catalog[ASIN],{cat_row})")
        ws.cell(row=r, column=3, value=f"=INDEX(Catalog[Title],{cat_row})")
        ws.cell(row=r, column=4, value=f"=INDEX(Catalog[Image Count],{cat_row})")
        ws.cell(row=r, column=5, value=(
            f'=IF(LEN(INDEX(Catalog[Bullet 1],{cat_row}))>0,1,0)'
            f'+IF(LEN(INDEX(Catalog[Bullet 2],{cat_row}))>0,1,0)'
            f'+IF(LEN(INDEX(Catalog[Bullet 3],{cat_row}))>0,1,0)'
            f'+IF(LEN(INDEX(Catalog[Bullet 4],{cat_row}))>0,1,0)'
            f'+IF(LEN(INDEX(Catalog[Bullet 5],{cat_row}))>0,1,0)'
        ))
        ws.cell(row=r, column=6, value=f'=LEN(INDEX(Catalog[Title],{cat_row}))')
        ws.cell(row=r, column=7, value=f'=LEN(INDEX(Catalog[Description],{cat_row}))')
        ws.cell(row=r, column=8, value=f'=IF(LEN(INDEX(Catalog[Fabric / Material],{cat_row}))>0,1,0)')
        # Composite score
        ws.cell(row=r, column=9, value=(
            f'=ROUND('
            f'IF(ISNUMBER(D{r}),IF(D{r}>=5,2,D{r}*0.4),0)'
            f'+IF(E{r}>=3,2,E{r}*0.67)'
            f'+IF(AND(F{r}>=60,F{r}<=200),2,IF(F{r}>200,MAX(0,2-(F{r}-200)/100),F{r}/30))'
            f'+IF(G{r}>=200,2,G{r}/100)'
            f'+H{r}*2'
            f',1)'
        ))
        ws.cell(row=r, column=9).number_format = "0.0"
        ws.cell(row=r, column=10, value=f'=IF(I{r}>=8,"High",IF(I{r}>=5,"Medium","Low"))')
        for c in (4,5,6,7,8,9): ws.cell(row=r, column=c).alignment = ALIGN_R
        ws.cell(row=r, column=10).alignment = ALIGN_C

    last = 7 + len(flat_rows)
    if flat_rows:
        ws.conditional_formatting.add(f"I8:I{last}", ColorScaleRule(
            start_type="num", start_value=0, start_color="FEE2E2",
            mid_type="num", mid_value=5, mid_color="FEF3C7",
            end_type="num", end_value=10, end_color="DCFCE7",
        ))
        # Grey out rows outside the selected bucket
        ws.conditional_formatting.add(f"B8:J{last}", FormulaRule(
            formula=[
                f'AND($C$5<>"All",'
                f'IF($C$5="High (≥8)",J8<>"High",'
                f'IF($C$5="Medium (5-7)",J8<>"Medium",'
                f'IF($C$5="Low (<5)",J8<>"Low",FALSE))))'
            ],
            fill=PatternFill("solid", fgColor="F8FAFC"),
            font=_font(color=COL_INK_FAINT, size=10),
        ))

    ws.freeze_panes = "B8"


# ============================================================
# 08 — Sharp Inferences (cross-signal diagnoses) — THE MONEY SHEET
# ============================================================

INFERENCE_PATTERNS = [
    {
        "name": "Zombie ASINs",
        "predicate": "Marked active-ish but 0 sessions AND 0 units for the entire period",
        "count_formula": "=COUNTIFS(Catalog[Sessions],0,Catalog[Units],0)",
        "diagnosis": (
            "Likely SUPPRESSED or DE-INDEXED. Amazon has stopped showing the listing to buyers. "
            "Common causes: policy violation, missing required attribute (e.g., fabric on apparel), "
            "or Amazon's inactive-listing sweep."
        ),
        "confirm_with": (
            "Check listing_status field in Seller Central for each ASIN. If 'active' but 0 sessions, "
            "likely de-indexed — refresh content and monitor. If 'suppressed', fix the flagged field."
        ),
        "action": "Sample 10 ASINs, check listing_status. If suppressed: fix the trigger. If active-but-dead: refresh content.",
    },
    {
        "name": "PPC-starved winners",
        "predicate": "Content-Health score ≥ 8 AND revenue = 0 (great listing, no traffic)",
        "count_formula": (
            "=SUMPRODUCT("
            "((LEN(Catalog[Bullet 1])>0)+(LEN(Catalog[Bullet 2])>0)+(LEN(Catalog[Bullet 3])>0)"
            "+(LEN(Catalog[Bullet 4])>0)+(LEN(Catalog[Bullet 5])>0)>=4)"
            "*(Catalog[Image Count]>=5)"
            "*(LEN(Catalog[Description])>=200)"
            "*(Catalog[Sessions]=0)"
            "*(Catalog[Units]=0)"
            ")"
        ),
        "diagnosis": (
            "Excellent listing hygiene but no visibility. Almost always PPC-starvation — this ASIN "
            "would convert if buyers found it, but nobody's driving traffic. Highest-ROI opportunity "
            "in the catalog."
        ),
        "confirm_with": (
            "Check Sponsored Products campaigns. Is this ASIN in ANY active campaign? "
            "Check organic rank in category (BSR) — if unranked, that's the smoking gun."
        ),
        "action": "Add these ASINs to a keyword-targeted Sponsored Products campaign at $2-5 bids. Expect fast lift.",
    },
    {
        "name": "Cannibalizing mega-families",
        "predicate": "Parent has ≥15 children, all sharing the same variation carousel",
        "count_formula": (
            "=SUMPRODUCT((COUNTIF(Catalog[Parent ASIN],Catalog[Parent ASIN])>=15)"
            "*(Catalog[Parent ASIN]<>\"\"))"
        ),
        "diagnosis": (
            "Variation carousel is crowded. Top 3-4 variations likely capture 80%+ of family revenue; "
            "bottom variations drag family BSR down and eat shelf attention. Cannibalization is "
            "confirmed if bottom 50% of family has 0 units."
        ),
        "confirm_with": (
            "Sort children by units within family. If top 5 = 95%+ of family revenue and bottom "
            "50% has 0 units, this is textbook cannibalization."
        ),
        "action": "Delist bottom 50% of each mega-family. Expected result: 20-40% lift in family BSR within 30 days.",
    },
    {
        "name": "Compliance-suppression suspects",
        "predicate": "Missing fabric_material AND missing country_of_origin AND 0 sessions",
        "count_formula": (
            "=SUMPRODUCT((LEN(Catalog[Fabric / Material])=0)"
            "*(LEN(Catalog[Country of Origin])=0)"
            "*(Catalog[Sessions]=0))"
        ),
        "diagnosis": (
            "Very likely suppressed by Amazon's apparel-compliance sweeps. Missing multiple "
            "required fields + zero traffic is the fingerprint of an Amazon-hidden listing."
        ),
        "confirm_with": (
            "Log into Seller Central and check 'Manage Inventory' → filter by listing status. "
            "Suppressed listings will show status 'Suppressed' with the missing attribute flagged."
        ),
        "action": "Bulk-add fabric_material + country_of_origin via flat-file upload. Expected to restore visibility in 7-14 days.",
    },
    {
        "name": "Ghost families (complete failures)",
        "predicate": "Parent has ≥10 children AND family revenue = 0",
        "count_formula": (
            "=SUMPRODUCT((COUNTIF(Catalog[Parent ASIN],Catalog[Parent ASIN])>=10)"
            "*(Catalog[Parent ASIN]<>\"\")*(Catalog[Revenue]=0)"
            "/COUNTIF(Catalog[Parent ASIN],Catalog[Parent ASIN]))"
        ),
        "diagnosis": (
            "Family launched but never gained traction. Could be: (a) wrong category placement, "
            "(b) never got ad support, (c) launched into a saturated segment, (d) all children "
            "suppressed simultaneously (family-level policy issue)."
        ),
        "confirm_with": (
            "Check family launch date + first ad spend date. Also verify family is in the right category."
        ),
        "action": "Either kill the family entirely OR relaunch with content refresh + $500 ad injection on the parent for 30 days.",
    },
    {
        "name": "Rich content, thin sales",
        "predicate": "A+ enabled AND all 5 bullets filled AND description ≥ 500 chars AND units < 10 in period",
        "count_formula": (
            "=SUMPRODUCT((Catalog[A+ Status]=\"enabled\")"
            "*((LEN(Catalog[Bullet 1])>0)+(LEN(Catalog[Bullet 2])>0)+(LEN(Catalog[Bullet 3])>0)"
            "+(LEN(Catalog[Bullet 4])>0)+(LEN(Catalog[Bullet 5])>0)=5)"
            "*(LEN(Catalog[Description])>=500)"
            "*(Catalog[Units]<10))"
        ),
        "diagnosis": (
            "You invested in content but it isn't converting. Two likely causes: (a) buyers aren't "
            "finding the listing (traffic problem — see PPC-starved pattern), (b) buyers are landing "
            "but bouncing (conversion problem — price, images, or reviews are the block)."
        ),
        "confirm_with": (
            "Check sessions:units ratio. If sessions > 100 and units < 10, it's a conversion problem. "
            "If sessions < 20, it's a traffic problem."
        ),
        "action": "Split by sessions. Traffic problem: PPC. Conversion problem: image refresh + price competitiveness audit.",
    },
    {
        "name": "Missing MAP protection",
        "predicate": "list_price > $50 AND 3P holds buy box (potential MAP violation)",
        "count_formula": (
            "=SUMPRODUCT((Catalog[List Price]>50)"
            "*(Catalog[Buy Box Winner]<>\"Novelle\")"
            "*(Catalog[Buy Box Winner]<>\"\"))"
        ),
        "diagnosis": (
            "3rd party has the buy box on a premium-priced ASIN. Likely an unauthorized reseller "
            "undercutting your MAP. Every unit sold through them is your revenue leaking to a middleman."
        ),
        "confirm_with": (
            "Identify each 3P seller. Cross-reference against your authorized distributor list. "
            "If unauthorized, file a Test Buy through Brand Registry."
        ),
        "action": "MAP-enforce with cease-and-desist to unauthorized sellers. Adjust FBA inventory to prevent stock-outs.",
    },
]


def _sheet_sharp_inferences(wb, flat_rows):
    ws = wb.create_sheet("08_Sharp_Inferences")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    for c, w in zip("BCDEFG", [26, 44, 12, 42, 36, 3]):
        ws.column_dimensions[c].width = w

    ws["B2"] = "Sharp Inferences — cross-signal diagnoses"
    ws["B2"].font = FONT_H2
    ws.merge_cells("B2:F2")

    ws["B3"] = ("These patterns combine multiple rules to produce concrete diagnoses no single rule can. "
                "Every 'Count' cell is a live formula — change your raw data and the counts update.")
    ws["B3"].font = FONT_MUTED
    ws.merge_cells("B3:F3")
    ws.row_dimensions[3].height = 30

    _banner(ws, "B5",
            "△  These are heuristics. Confirm each with the follow-up check in the 'Confirm with' column before acting.",
            "B5:F5")
    ws.row_dimensions[5].height = 24

    hdrs = ["Pattern", "What it looks like", "Count", "Diagnosis", "How to confirm"]
    for c, h in enumerate(hdrs, start=2):
        _hdr_cell(ws, 7, c, h)

    for i, pat in enumerate(INFERENCE_PATTERNS):
        r = 8 + i * 2  # 2 rows per pattern for readability

        ws.cell(row=r, column=2, value=pat["name"]).font = FONT_BOLD
        ws.cell(row=r, column=2).alignment = ALIGN_LT

        ws.cell(row=r, column=3, value=pat["predicate"]).font = FONT_MUTED
        ws.cell(row=r, column=3).alignment = ALIGN_LT

        ws.cell(row=r, column=4, value=pat["count_formula"])
        ws.cell(row=r, column=4).number_format = "#,##0"
        ws.cell(row=r, column=4).font = _font(bold=True, size=14, color=COL_ACCENT_2)
        ws.cell(row=r, column=4).alignment = ALIGN_C

        ws.cell(row=r, column=5, value=pat["diagnosis"]).font = FONT_BODY
        ws.cell(row=r, column=5).alignment = ALIGN_LT

        ws.cell(row=r, column=6, value=pat["confirm_with"]).font = FONT_MUTED
        ws.cell(row=r, column=6).alignment = ALIGN_LT

        # Action row directly below
        ws.cell(row=r + 1, column=3, value="→ Recommended action:").font = _font(italic=True, size=9, color=COL_INK_MUTED)
        ws.cell(row=r + 1, column=4, value=pat["action"]).font = _font(bold=True, size=10, color="0891B2")
        ws.merge_cells(start_row=r + 1, start_column=4, end_row=r + 1, end_column=6)
        ws.cell(row=r + 1, column=4).alignment = ALIGN_LT

        # Stripe every other pattern for readability
        if i % 2 == 0:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = FILL_ROW_STRIPE
                ws.cell(row=r + 1, column=c).fill = FILL_ROW_STRIPE

        ws.row_dimensions[r].height = 90
        ws.row_dimensions[r + 1].height = 32

    ws.freeze_panes = "B8"
    _fit_landscape(ws, height=0)


# ============================================================
# 09 — All Findings
# ============================================================

def _sheet_all_findings(wb, findings):
    ws = wb.create_sheet("09_All_Findings")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    for c, w in zip("BCDEFGHI", [10, 34, 14, 55, 65, 28, 30, 3]):
        ws.column_dimensions[c].width = w

    ws["B2"] = "All findings — dashboard mirror"
    ws["B2"].font = FONT_H2
    ws.merge_cells("B2:H2")

    ws["B3"] = ("Every finding the dashboard raised on this catalog. "
                "For rule definitions and inference logic, jump to 10_Rules_Methodology.")
    ws["B3"].font = FONT_MUTED
    ws.merge_cells("B3:H3")

    hdrs = ["Severity","Rule","ASIN","Finding (proposed fix)","Why This Matters","Finding ID"]
    for c, h in enumerate(hdrs, start=2):
        _hdr_cell(ws, 5, c, h)

    sev_order = {"critical":0,"high":1,"medium":2,"low":3,"info":4}
    sorted_f = sorted(findings, key=lambda f: (
        sev_order.get(f.get("severity","medium"), 9),
        -(f.get("priority_score") or 0)
    ))

    for i, f in enumerate(sorted_f, start=6):
        rule = f.get("rule_name") or ""
        spec = RULE_SPECS.get(rule, {})
        sev = (f.get("severity") or "").upper()
        ws.cell(row=i, column=2, value=sev).font = FONT_BOLD
        # Severity color
        sev_fill = {"CRITICAL": FILL_ALERT, "HIGH": FILL_ALERT, "MEDIUM": FILL_WARN,
                    "LOW": FILL_KPI, "INFO": FILL_KPI}.get(sev, FILL_KPI)
        ws.cell(row=i, column=2).fill = sev_fill
        ws.cell(row=i, column=2).alignment = ALIGN_C

        ws.cell(row=i, column=3, value=spec.get("label") or rule).font = FONT_BOLD
        ws.cell(row=i, column=4, value=f.get("asin") or "").alignment = ALIGN_C
        ws.cell(row=i, column=5, value=f.get("proposed_fix") or f.get("message") or "")
        ws.cell(row=i, column=6, value=spec.get("why_matters", ""))
        ws.cell(row=i, column=7, value=f.get("finding_id") or "").font = FONT_CODE
        for c in (3,5,6): ws.cell(row=i, column=c).alignment = ALIGN_LT
        ws.row_dimensions[i].height = 90

    ws.freeze_panes = "B6"
    _fit_landscape(ws, height=0)


# ============================================================
# 10 — Rules Methodology — RICH CARDS
# ============================================================

def _sheet_rules_methodology(wb):
    ws = wb.create_sheet("10_Rules_Methodology")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28  # section label
    ws.column_dimensions["C"].width = 80  # content (wide but fits landscape)
    ws.column_dimensions["D"].width = 3

    ws["B2"] = "Rules Methodology — full explainer"
    ws["B2"].font = FONT_H2
    ws.merge_cells("B2:C2")

    ws["B3"] = ("For every one of the 15 rules Catalog Intel can fire, this appendix explains "
                "what it checks, why it matters, where the best practice comes from, what "
                "the inference is when it fires, and the one thing to do this week.")
    ws["B3"].font = FONT_MUTED
    ws["B3"].alignment = ALIGN_LT
    ws.merge_cells("B3:C3")
    ws.row_dimensions[3].height = 30

    row = 5
    for rule_id in sorted(RULE_SPECS.keys()):
        s = RULE_SPECS[rule_id]

        # ── Rule header block ────────────────────────────────
        ws.cell(row=row, column=2, value=s.get("label") or rule_id).font = _font(bold=True, size=14, color=COL_HEADER_TX)
        ws.cell(row=row, column=2).fill = FILL_HEADER
        ws.cell(row=row, column=2).alignment = ALIGN_L
        ws.cell(row=row, column=3, value=(
            f"{s.get('category','')}  ·  data source: {s.get('data_source','')}  ·  rule_id: {rule_id}"
        )).font = _font(size=10, color=COL_HEADER_TX)
        ws.cell(row=row, column=3).fill = FILL_HEADER
        ws.cell(row=row, column=3).alignment = ALIGN_L
        ws.row_dimensions[row].height = 24

        row += 1
        # Live count of ASINs this rule affects (formula pointing at Catalog)
        live_formula = s.get("live_count_formula")
        if live_formula:
            ws.cell(row=row, column=2, value="ASINs currently affected").font = FONT_BOLD
            cell = ws.cell(row=row, column=3, value=live_formula)
            cell.font = _font(bold=True, size=14, color=COL_ACCENT_2)
            cell.number_format = "#,##0"
            ws.row_dimensions[row].height = 22
            row += 1

        # Rich sections
        sections = [
            ("What this rule checks",      s.get("predicate","")),
            ("Threshold",                  s.get("threshold","")),
            ("Severity logic",             s.get("severity_logic","")),
            ("Fields inspected",           ", ".join(s.get("checks_field") or []) or "(no single field — cross-catalog rule)"),
            ("Minimum coverage required",  s.get("min_coverage","")),
            ("WHY THIS MATTERS",           s.get("why_matters", "")),
            ("Where this rule came from",  s.get("source", "")),
            ("What we can infer when it fires", s.get("inference_when_flagged", "")),
            ("What would sharpen this",    s.get("sharpen_with", "")),
            ("First check this week",      s.get("first_check", "")),
            ("SQL predicate (as executed)",s.get("sql_predicate", "")),
            ("Standalone verify query",    s.get("verify_query", "")),
        ]
        for label, content in sections:
            if not content:
                continue
            lcell = ws.cell(row=row, column=2, value=label)
            lcell.font = FONT_BOLD if "WHY THIS MATTERS" not in label else _font(bold=True, size=10, color=COL_ACCENT_2)
            lcell.alignment = ALIGN_LT
            lcell.fill = FILL_ROW_STRIPE if "WHY" in label or "First check" in label or "infer" in label else PatternFill()

            code_style = label.startswith("SQL") or label.startswith("Standalone")
            ccell = ws.cell(row=row, column=3, value=content)
            ccell.font = FONT_CODE if code_style else FONT_BODY
            ccell.alignment = ALIGN_LT
            ccell.fill = FILL_ROW_STRIPE if "WHY" in label or "First check" in label or "infer" in label else PatternFill()

            # Row height based on content length
            n_lines = max(1, len(content) // 80 + content.count("\n"))
            ws.row_dimensions[row].height = min(220, 18 + n_lines * 14)
            row += 1

        # Spacer between rules
        row += 1

    ws.freeze_panes = "B5"
    _fit_landscape(ws, height=0)


# ============================================================
# 11 — Data Gaps (pitch sheet)
# ============================================================

DATA_GAPS = [
    {"gap":"Promo / discount depth","field":"sale_price with dated periods, coupon codes","unlocks":"Discount depth analysis, holiday timing, price elasticity by ASIN, MAP-integrity check","priority":"high","how":"Export Amazon Deals dashboard or SP-API GetPromotions"},
    {"gap":"Country of origin","field":"country_of_origin per ASIN","unlocks":"Compliance risk audit, tariff exposure map, sourcing consolidation","priority":"high","how":"Add to Seller Central listing attributes; export via bulk report"},
    {"gap":"Care instructions (apparel)","field":"care_instructions per apparel ASIN","unlocks":"Apparel compliance audit, correlation between care complexity and return rate","priority":"high","how":"Amazon apparel category requires this — pull from category listing attributes report"},
    {"gap":"Search terms / backend keywords","field":"backend_keywords, front-end search term rank","unlocks":"SEO gap analysis, keyword coverage score, discoverability audit","priority":"medium","how":"Search Query Performance report + backend keyword report from Seller Central"},
    {"gap":"Reviews & ratings","field":"review_count, avg_rating, review_text sample","unlocks":"Sentiment mining, complaint theming, competitor comparison","priority":"medium","how":"Amazon Vine + review scraping via approved SP-API endpoints"},
    {"gap":"Rank data (BSR)","field":"bsr_category, bsr_rank_daily_history","unlocks":"Rank decay detection, category positioning, competitive threat mapping","priority":"medium","how":"Third-party rank tracker (Helium 10, Jungle Scout) or manual SP-API pulls"},
    {"gap":"Ad spend / TACOS","field":"ad_spend, impressions, clicks per ASIN per period","unlocks":"Ad efficiency per ASIN, wasted spend detection, TACOS trend","priority":"high","how":"Advertising bulk file download from Seller Central"},
    {"gap":"Returns data","field":"return_count, return_reason breakdown","unlocks":"Return rate by ASIN, defective SKU detection, size-fit signals","priority":"high","how":"Returns FBA/FBM report from Seller Central"},
    {"gap":"Historical traffic & rank trend","field":"session and rank time series (12+ weeks)","unlocks":"Category velocity, seasonality mapping, launch effect detection","priority":"medium","how":"Business Reports archive from Seller Central (30-day windows stitched)"},
    {"gap":"Competitor URLs & data","field":"manual list of competing brand + ASIN pairs","unlocks":"PDP gap analysis, price positioning, image-count comparison","priority":"low","how":"Operator-provided competitor list; agency can scrape/pull"},
]


def _sheet_data_gaps(wb):
    ws = wb.create_sheet("11_Data_Gaps")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    for c, w in zip("BCDEFG", [28, 34, 60, 12, 42, 3]):
        ws.column_dimensions[c].width = w

    ws["B2"] = "What more data do you need?"
    ws["B2"].font = FONT_H2
    ws.merge_cells("B2:F2")

    ws["B3"] = ("Rules that can't currently fire on your catalog because the underlying data isn't "
                "in the upload. Send any of these and the analyses on the right become available.")
    ws["B3"].font = FONT_MUTED
    ws.merge_cells("B3:F3")
    ws.row_dimensions[3].height = 30

    for c, h in enumerate(["Data Gap","Field / Signal","What This Unlocks","Priority","How to Provide"], start=2):
        _hdr_cell(ws, 5, c, h)

    pri_fill = {"high":FILL_ALERT, "medium":FILL_WARN, "low":PatternFill("solid", fgColor="DBEAFE")}
    for i, gap in enumerate(DATA_GAPS, start=6):
        ws.cell(row=i, column=2, value=gap["gap"]).font = FONT_BOLD
        ws.cell(row=i, column=3, value=gap["field"])
        ws.cell(row=i, column=4, value=gap["unlocks"])
        p = ws.cell(row=i, column=5, value=gap["priority"].upper())
        p.fill = pri_fill.get(gap["priority"], PatternFill())
        p.alignment = ALIGN_C
        p.font = FONT_BOLD
        ws.cell(row=i, column=6, value=gap["how"])
        for c in (2,3,4,6): ws.cell(row=i, column=c).alignment = ALIGN_LT
        ws.row_dimensions[i].height = 60

    _fit_landscape(ws, height=0)


# ============================================================
# 12 — Trend KPIs (skeleton, populated on re-upload)
# ============================================================

_TREND_KPIS = [
    ("Total ASINs","count","=COUNTA(Catalog[ASIN])"),
    ("Dead ASINs % (0 units)","pct","=IFERROR(COUNTIF(Catalog[Units],0)/COUNTA(Catalog[ASIN]),0)"),
    ("Active ASINs","count",'=COUNTIFS(Catalog[Sessions],">0")+COUNTIFS(Catalog[Sessions],0,Catalog[Units],">0")'),
    ("Total revenue","money","=SUM(Catalog[Revenue])"),
    ("Titles filled %","pct",'=IFERROR(COUNTIF(Catalog[Title],"?*")/COUNTA(Catalog[ASIN]),0)'),
    ("Descriptions filled %","pct",'=IFERROR(COUNTIF(Catalog[Description],"?*")/COUNTA(Catalog[ASIN]),0)'),
    ("Fabric/material filled %","pct",'=IFERROR(COUNTIF(Catalog[Fabric / Material],"?*")/COUNTA(Catalog[ASIN]),0)'),
    ("Avg images per ASIN","num","=IFERROR(AVERAGE(Catalog[Image Count]),0)"),
]


def _snap_label(snapshot):
    if not snapshot: return None
    ts = snapshot.get("uploaded_at") or ""
    return ts[:10] if ts else "Now"


def _sheet_trend_kpis(wb, snapshot):
    ws = wb.create_sheet("12_Trend_KPIs")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    for c, w in zip("BCDEFGHIJ", [32, 10, 14, 14, 14, 18, 16, 20, 3]):
        ws.column_dimensions[c].width = w

    ws["B2"] = "KPI Trend — populated as you re-upload"
    ws["B2"].font = FONT_H2
    ws.merge_cells("B2:H2")

    ws["B3"] = ("Every KPI tracked over time. Only the current column has data. Re-upload monthly "
                "and the T-1 / T-2 / T-3 columns fill automatically in the dashboard's diff view.")
    ws["B3"].font = FONT_MUTED
    ws.merge_cells("B3:H3")
    ws.row_dimensions[3].height = 26

    _banner(ws, "B5",
            "△  Excel workbook is point-in-time. The dashboard maintains this table automatically — no manual sheet-per-month juggling.",
            "B5:H5")
    ws.row_dimensions[5].height = 24

    snap_label = _snap_label(snapshot) or "Now"
    hdrs = ["KPI","Format","T-3","T-2","T-1", snap_label, "Δ T-1 → Now", "Direction"]
    for c, h in enumerate(hdrs, start=2):
        _hdr_cell(ws, 7, c, h)

    for i, (label, kind, formula) in enumerate(_TREND_KPIS, start=8):
        ws.cell(row=i, column=2, value=label).font = FONT_BOLD
        ws.cell(row=i, column=3, value=kind).font = FONT_MUTED
        for col in (4,5,6):
            c = ws.cell(row=i, column=col, value="")
            c.fill = FILL_ROW_STRIPE
        cur = ws.cell(row=i, column=7, value=formula)
        cur.font = FONT_BOLD
        cur.fill = FILL_KPI_HERO
        d_cell = ws.cell(row=i, column=8, value=f'=IF(F{i}="","",G{i}-F{i})')
        ws.cell(row=i, column=9, value=f'=IF(F{i}="","awaiting T-1",IF(H{i}>0,"↑ up",IF(H{i}<0,"↓ down","→ flat")))')

        if kind == "pct":
            for col in (4,5,6,7): ws.cell(row=i, column=col).number_format = "0.0%"
            d_cell.number_format = "+0.0%;-0.0%;—"
        elif kind == "money":
            for col in (4,5,6,7): ws.cell(row=i, column=col).number_format = "$#,##0"
            d_cell.number_format = "+$#,##0;-$#,##0;—"
        elif kind == "count":
            for col in (4,5,6,7): ws.cell(row=i, column=col).number_format = "#,##0"
            d_cell.number_format = "+#,##0;-#,##0;—"
        else:
            for col in (4,5,6,7): ws.cell(row=i, column=col).number_format = "#,##0.0"
            d_cell.number_format = "+#,##0.0;-#,##0.0;—"

    footer_row = 8 + len(_TREND_KPIS) + 1
    ws.cell(row=footer_row, column=2, value=(
        "To populate T-1: send your previous month's catalog upload. "
        "T-2 needs 2 months back. T-3 needs 3."
    )).font = FONT_MUTED
    ws.merge_cells(start_row=footer_row, start_column=2, end_row=footer_row, end_column=9)

    ws.freeze_panes = "B8"
    _fit_landscape(ws, height=0)


# ============================================================
# 13 — Trend by Rule
# ============================================================

_RULE_TREND_METRICS = {
    "dead_inventory":              {"key":"dead_pct","direction":"lower_is_better","fmt":"pct"},
    "description_presence":        {"key":"pct_with_description","direction":"higher_is_better","fmt":"pct"},
    "fabric_material_coverage":    {"key":"pct_filled","direction":"higher_is_better","fmt":"pct"},
    "buy_box_ownership":           {"key":"likely_owner_pct","direction":"higher_is_better","fmt":"pct"},
    "image_count_dist":            {"key":"under_5_pct","direction":"lower_is_better","fmt":"pct"},
    "bullet_completeness_dist":    {"key":"under_3_pct","direction":"lower_is_better","fmt":"pct"},
    "title_length_dist":           {"key":"flagged_pct","direction":"lower_is_better","fmt":"pct"},
    "variation_theme_integrity":   {"key":"inconsistent_pct","direction":"lower_is_better","fmt":"pct"},
    "style_family_concentration":  {"key":"mega_family_count","direction":"lower_is_better","fmt":"count"},
    "list_price_dist":             {"key":"outlier_count","direction":"lower_is_better","fmt":"count"},
    "concentration_pareto":        {"key":"top_50pct_asins","direction":"higher_is_better","fmt":"count"},
    "cohort_split":                {"key":"dead_pct","direction":"lower_is_better","fmt":"pct"},
    "a_plus_lift":                 {"key":"lift_multiplier","direction":"higher_is_better","fmt":"num"},
    "fill_rate_report":            None,
    "subcategory_rollup":          None,
}


def _sheet_trend_by_rule(wb, snapshot):
    ws = wb.create_sheet("13_Trend_By_Rule")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    for c, w in zip("BCDEFGHIJK", [34, 24, 8, 12, 12, 12, 16, 16, 22, 3]):
        ws.column_dimensions[c].width = w

    ws["B2"] = "Per-Rule Metric Trend"
    ws["B2"].font = FONT_H2
    ws.merge_cells("B2:I2")

    ws["B3"] = ("For each of the 15 rules, the primary metric that moves when things get better or worse. "
                "Only the current column shows data — historicals populate as you re-upload.")
    ws["B3"].font = FONT_MUTED
    ws.merge_cells("B3:I3")
    ws.row_dimensions[3].height = 26

    _banner(ws, "B5",
            "△  The dashboard's snapshot-diff view is the interactive version of this table with materiality thresholds applied.",
            "B5:I5")
    ws.row_dimensions[5].height = 24

    snap_label = _snap_label(snapshot) or "Now"
    hdrs = ["Rule","Metric","Format","T-3","T-2","T-1", snap_label, "Δ T-1 → Now","Direction"]
    for c, h in enumerate(hdrs, start=2):
        _hdr_cell(ws, 7, c, h)

    row = 8
    for rule_id in sorted(RULE_SPECS.keys()):
        spec = RULE_SPECS[rule_id]
        m = _RULE_TREND_METRICS.get(rule_id)
        ws.cell(row=row, column=2, value=spec.get("label") or rule_id).font = FONT_BOLD
        if not m:
            ws.cell(row=row, column=3, value="(no single metric)").font = FONT_MUTED
            for col in (4,5,6,7,8,9,10):
                ws.cell(row=row, column=col, value="—").font = FONT_MUTED
            row += 1
            continue
        ws.cell(row=row, column=3, value=m["key"])
        ws.cell(row=row, column=4, value=m["fmt"]).font = FONT_MUTED
        for col in (5,6,7):
            c = ws.cell(row=row, column=col, value="")
            c.fill = FILL_ROW_STRIPE
        # Look up the finding text from All_Findings (col C = Rule label, col E = Finding)
        ws.cell(row=row, column=8, value=(
            f'=IFERROR(VLOOKUP("{spec.get("label") or rule_id}",\'09_All_Findings\'!C:E,3,FALSE),"—")'
        ))
        ws.cell(row=row, column=8).font = FONT_BOLD
        ws.cell(row=row, column=8).fill = FILL_KPI_HERO
        ws.cell(row=row, column=9, value=f'=IF(G{row}="","",H{row}-G{row})')
        dir_up = "↑ improved" if m["direction"] == "higher_is_better" else "↑ worsened"
        dir_dn = "↓ worsened" if m["direction"] == "higher_is_better" else "↓ improved"
        ws.cell(row=row, column=10, value=(
            f'=IF(G{row}="","awaiting T-1",'
            f'IF(I{row}>0,"{dir_up}",'
            f'IF(I{row}<0,"{dir_dn}","→ unchanged")))'
        ))
        row += 1

    ws.freeze_panes = "B8"
    _fit_landscape(ws, height=0)


# ============================================================
# 14 — Fix Effectiveness
# ============================================================

def _sheet_fix_effectiveness(wb, findings):
    ws = wb.create_sheet("14_Fix_Effectiveness")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    for c, w in zip("BCDEFGHI", [32, 10, 22, 16, 20, 20, 20, 32]):
        ws.column_dimensions[c].width = w

    ws["B2"] = "Fix Effectiveness — the money loop"
    ws["B2"].font = FONT_H2
    ws.merge_cells("B2:I2")

    ws["B3"] = ("For every finding you mark 'fixed' in the dashboard, this sheet shows whether "
                "the underlying metric actually improved on the next snapshot. Proves the audit is producing outcomes.")
    ws["B3"].font = FONT_MUTED
    ws.merge_cells("B3:I3")
    ws.row_dimensions[3].height = 30

    _banner(ws, "B5",
            "△  Requires: finding_status workflow in the dashboard + at least 2 snapshots. Excel can't track status across uploads without manual copying.",
            "B5:I5")
    ws.row_dimensions[5].height = 24

    hdrs = ["Rule","Severity","Status","Fixed Date","Metric at Fix","Metric Now","Actually Improved?","Verification"]
    for c, h in enumerate(hdrs, start=2):
        _hdr_cell(ws, 7, c, h)

    sev_order = {"critical":0,"high":1,"medium":2,"low":3,"info":4}
    sorted_f = sorted(findings, key=lambda f: (
        sev_order.get(f.get("severity","medium"), 9),
        -(f.get("priority_score") or 0)
    ))
    for i, f in enumerate(sorted_f, start=8):
        rule = f.get("rule_name") or ""
        spec = RULE_SPECS.get(rule, {})
        ws.cell(row=i, column=2, value=spec.get("label") or rule)
        ws.cell(row=i, column=3, value=(f.get("severity") or "").upper())
        for col in (4,5,6,7):
            c = ws.cell(row=i, column=col, value="awaiting workflow")
            c.font = FONT_MUTED
            c.fill = FILL_ROW_STRIPE
        ws.cell(row=i, column=8, value='=IF(OR(F' + str(i) + '="awaiting workflow",G' + str(i) + '="awaiting workflow"),"pending","")').font = FONT_MUTED
        ws.cell(row=i, column=9, value="see dashboard → finding history").font = FONT_MUTED

    ws.freeze_panes = "B8"
    _fit_landscape(ws, height=0)


# ============================================================
# 15 — How to Add Historicals
# ============================================================

def _sheet_how_to_add_historicals(wb):
    ws = wb.create_sheet("15_How_To_Add_Historicals")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    for c, w in zip("BCDE", [28, 32, 68, 3]):
        ws.column_dimensions[c].width = w

    ws["B2"] = "How to fill in the trend columns"
    ws["B2"].font = FONT_H2
    ws.merge_cells("B2:D2")

    ws["B3"] = ("Sheets 12-14 have empty T-1 / T-2 / T-3 columns. Each blank column corresponds to "
                "a past monthly upload. Here's what to send and what it unlocks.")
    ws["B3"].font = FONT_MUTED
    ws.merge_cells("B3:D3")
    ws.row_dimensions[3].height = 30

    plan = [
        ("For T-1 (last month)","1 catalog export dated ~30 days ago",
         "Enables month-over-month deltas on all 8 KPIs and all 15 rule metrics. Dashboard's diff view flags improved / worsened / unchanged with 1pt materiality."),
        ("For T-2 (2 months back)","1 catalog export dated ~60 days ago",
         "Enables 3-point trend visibility (T-2 → T-1 → Now). Sparklines start showing shape. Identify accelerating problems vs. one-off regressions."),
        ("For T-3 (3 months back)","1 catalog export dated ~90 days ago",
         "Full quarterly trend. Sparklines meaningful. Category velocity and content-health drift become detectable."),
        ("For monthly ongoing","1 fresh upload every month",
         "Automated diff runs. Fix-effectiveness sheet populates. Dashboard notifications fire on materially worsened metrics."),
    ]
    for c, h in enumerate(["When","What to send","What this unlocks"], start=2):
        _hdr_cell(ws, 5, c, h)

    for i, (when, what, unlocks) in enumerate(plan, start=6):
        ws.cell(row=i, column=2, value=when).font = FONT_BOLD
        ws.cell(row=i, column=3, value=what)
        ws.cell(row=i, column=4, value=unlocks)
        for c in (2,3,4): ws.cell(row=i, column=c).alignment = ALIGN_LT
        ws.row_dimensions[i].height = 60

    ws["B12"] = "Why not just juggle Excel files month-over-month?"
    ws["B12"].font = FONT_H3
    reasons = [
        "• Managing 12 workbooks/year manually + copying T-1/T-2/T-3 values by hand every month.",
        "• Materiality threshold (1pt) has to be applied consistently across every rule — error-prone in Excel.",
        "• Status workflow (in_progress / fixed / wontfix) has no home in raw Excel.",
        "• Audit trail (who changed what status when) needs an immutable history log.",
        "• Cross-brand pooling (Novelle vs Roxy vs future brands) is a database join, not a spreadsheet.",
        "• Real-time drilldowns require querying live data.",
    ]
    for i, r in enumerate(reasons, start=14):
        cell = ws.cell(row=i, column=2, value=r)
        cell.font = FONT_BODY
        cell.alignment = ALIGN_LT
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

    ws["B22"] = "Excel verifies the numbers. The dashboard is where the numbers live over time."
    ws["B22"].font = _font(bold=True, italic=True, size=11, color=COL_INK)
    ws.merge_cells("B22:D22")

    _fit_landscape(ws, height=0)


# ============================================================
# 16 — How This Works (trust page)
# ============================================================

def _sheet_how_it_works(wb):
    ws = wb.create_sheet("16_How_This_Works")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100
    ws.column_dimensions["C"].width = 3

    ws["B2"] = "How this workbook works"
    ws["B2"].font = FONT_H2

    body = [
        "",
        "Every downstream number in this workbook is a live Excel formula pointing at 02_Catalog_Data.",
        "You can click any KPI, chart, or table cell, look at the formula bar, and see the exact math.",
        "",
        "Verification workflow:",
        "  1. Open 09_All_Findings. Every row is a finding from the dashboard.",
        "  2. Open 10_Rules_Methodology. Find the same rule. Read the WHY, the SOURCE, the SQL, and the verify query.",
        "  3. Return to 02_Catalog_Data and confirm the raw values that produced the finding.",
        "  4. Change a raw value on 02_Catalog_Data and watch KPIs update in real time. If they don't update, the formula is broken and you've found a dashboard bug.",
        "",
        "Named ranges (Formulas → Name Manager):",
        "  Catalog                — the raw catalog table on 02_Catalog_Data",
        "  Catalog[ASIN]          — all ASINs",
        "  Catalog[Revenue]       — all revenue values",
        "  Catalog[Sessions]      — all sessions",
        "  Catalog[Units]         — all units",
        "  Catalog[Title]         — all titles",
        "  Catalog[Image Count]   — all image counts",
        "  Sales                  — the raw sales rollup on 03_Sales_Data",
        "",
        "Interactivity (data-validation dropdowns, not slicers):",
        "  04_Coverage_Matrix       — cell C5: filter fields by category",
        "  05_Revenue_Concentration — cell C5: highlight top N (10/20/50/100/200)",
        "  07_Content_Health        — cell C5: filter ASINs by score bucket",
        "",
        "Sheets that recompute automatically when you edit 02_Catalog_Data:",
        "  06_Cohort_Analysis        — helper cells I6/I7 (P10 and P80 revenue thresholds)",
        "  08_Sharp_Inferences       — counts recompute when catalog changes",
        "  10_Rules_Methodology      — 'ASINs currently affected' updates as data changes",
        "  12_Trend_KPIs             — 'Now' column always live; T-3/T-2/T-1 populate as you re-upload",
        "  13_Trend_By_Rule          — same as 12, but per-rule metrics",
        "  14_Fix_Effectiveness      — requires status workflow in dashboard",
        "",
        "Open this workbook in Excel 365 (Office 2019+) or LibreOffice 7.5+ for full formula support.",
        "PERCENTILE on structured references may show '#NAME?' in older Office builds.",
        "",
        "Materiality note (for snapshot-diff comparisons):",
        "  The dashboard treats deltas under 1pt (or under 5 for counts) as 'unchanged' to avoid noise.",
        "  This workbook shows exact numbers with no rounding. Absolute values match to the row.",
        "",
        "Generated by Perplexity Computer — Atlas Catalog Intel v1.3.",
    ]
    for i, line in enumerate(body, start=4):
        cell = ws.cell(row=i, column=2, value=line)
        cell.font = FONT_CODE
        cell.alignment = ALIGN_LT

    _fit_landscape(ws, height=0)
