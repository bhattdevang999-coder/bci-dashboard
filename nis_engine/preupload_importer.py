"""Read a TLG-style Sage/Volcom 'Pre-Upload' .xlsx and convert each style row
into an NIS form-state dict that the rule engine can evaluate.

Designed to be schema-tolerant: column headers are matched by name (lowercased,
punctuation-insensitive) so R0 vs R1 column shuffles don't break it.
"""

import re
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    openpyxl = None


# Brand-aware vendor-code lookup. Add brands here as they sign on.
_BRAND_VENDOR_CODES = {
    "Sage Collective": "QT5G8",
}


# Color-code -> Amazon standardized color (last-resort fallback used only if the
# importer is called without a brand/PT context). The main app.py.normalize_color
# function is now PT-aware (snaps results to dropdown_cache/{PT}.json) and is the
# single source of truth used by the rule engine and template-writer. This local
# map exists for the rare case where preupload_importer is invoked standalone.
_COLOR_MAP = {
    "BEIGE": "Beige", "BLACK": "Black", "BROWN": "Brown", "BLUE": "Blue",
    "RED": "Red", "GREEN": "Green", "GRAY": "Grey", "GREY": "Grey",
    "WHITE": "White", "IVORY": "Off White", "NAVY": "Blue", "TAN": "Brown",
    "CAMEL": "Brown", "OLIVE": "Green", "CREAM": "Off White", "CHARCOAL": "Grey",
    "BURGUNDY": "Red", "WINE": "Red", "PINK": "Pink", "PURPLE": "Purple",
    "TAUPE": "Brown", "COGNAC": "Brown", "TRUFFLE": "Brown", "BEG": "Beige",
    "BLK": "Black", "NVY": "Blue",
}


# Sub-class label -> (product_category, product_subcategory, item_type_keyword, item_type_name)
# COAT cascades known to Amazon NIS — verified against engine bundles.
_SUBCLASS_TAXONOMY = {
    "Puffer":              ("Women's Outerwear", "Down and Parkas", "down-and-parkas-coats", "Puffer Coat"),
    "Faux Wool Outerwear": ("Women's Outerwear", "Wool",            "wool-outerwear-coats",  "Wool Coat"),
    "Wool Outerwear":      ("Women's Outerwear", "Wool",            "wool-outerwear-coats",  "Wool Coat"),
    "Anorak":              ("Women's Outerwear", "Lightweight Jackets and Windbreakers",
                                                                    "anorak-jackets",        "Anorak"),
    "Vest":                ("Women's Outerwear", "Lightweight Jackets and Windbreakers",
                                                                    "outerwear-vests",       "Vest"),
    "Parka":               ("Women's Outerwear", "Down and Parkas", "parkas",                "Parka"),
}


def _norm_header(s: Optional[str]) -> str:
    """Lowercase, strip non-alphanumerics. Used for tolerant header matching."""
    if not s:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(s).lower())


# Map of normalized header tokens to canonical field key. The list of aliases
# captures the variations we've seen across Sage/Volcom pre-upload files.
_HEADER_ALIASES = {
    "season":          ["seasoncode", "season"],
    "tlgdiv":          ["tlgdivname", "brandcode"],
    "division":        ["division"],
    "sub_class":       ["subclassname"],
    "sub_sub":         ["subsubclassname"],
    "style":           ["style", "styleno", "stylenumber"],
    "name":            ["basicstylename", "stylename", "tlgstylename"],
    "keywords":        ["relatedkeywords", "tlgstyledesc"],
    "color_code":      ["colorcode"],
    "color":           ["colorname"],
    "model":           ["modelnumber"],
    "size":            ["productsize", "size"],
    "upc":             ["upccode", "upc"],
    "asin":            ["childasin"],
    "sku":             ["sku"],
    "due_date":        ["duedateearliestshipdate", "duedate", "earliestshipdate", "shipdate"],
    "amazon_cost":     ["amazoncost", "amznwholesale", "wholesale", "cost", "costprice"],
    "list_price":      ["amazonlistprice", "amznretail", "retail", "listprice", "msrp"],
    "department":      ["department"],
    "type_jacket":     ["typeofjacket", "typeofjacket"],
    "coo":             ["coo", "countryoforigin"],
    "care":            ["careinstructions", "care"],
    "fabric":          ["fabriccontentpercentage", "fabric", "fabriccontent"],
    "closure":         ["closuretype", "closure", "closuretype1", "closure1"],
    # Pass 3 (agency R0 items 11, 15) — v2 template will ship explicit
    # Material 2/3 and Closure Type 2 columns. v1 template doesn't have them;
    # these aliases are no-ops on v1 ingest but unblock v2 once Pass 6 ships.
    "material_2":      ["material2", "materialcomposition2", "secondarymaterial"],
    "material_3":      ["material3", "materialcomposition3", "tertiarymaterial"],
    "closure_2":       ["closuretype2", "closure2", "secondaryclosure"],
    "length":          ["centerbacklengthcbl", "length", "cbl"],
    "pockets":         ["numberofpockets"],
    "hood":            ["removablehood"],
    # Pass 1 (agency R0 item 13) — Sleeve Type / Length, Neck, Fit, UPF all
    # exist in the Tahari template but were never mapped to state.
    "sleeve_type":     ["sleevetype"],
    "sleeve_length":   ["sleevelength", "sleevelengthdescription"],
    "neck":            ["necktype", "neckline", "necklinestyle"],
    "fit":             ["fittype", "fit"],
    "upf":             ["upfrating", "upf"],
    # Pass 6 (agency R0 strategic 2) — v2 template ships Amazon-attribute-
    # shaped columns directly instead of inferring from TLG-internal cols.
    # These aliases are no-op on v1 ingest, populated on v2 ingest.
    "age_range":             ["agerange", "agerangedescription"],
    "target_gender":         ["targetgender"],
    "item_length_description":["itemlengthdescription", "itemlength"],
    # Pass 1 (agency R0 item 5) — the 5 per-bullet columns were unmapped, so
    # the parser never read template-supplied bullets and the verdict reported
    # "Bullet Points missing". The fifth alias variant covers the Sage/Volcom
    # plus the Tahari header conventions.
    "bullet_1":        ["keyfeaturesbullet1", "bullet1", "keyfeatures1"],
    "bullet_2":        ["keyfeaturesbullet2", "bullet2", "keyfeatures2"],
    "bullet_3":        ["keyfeaturesbullet3", "bullet3", "keyfeatures3"],
    "bullet_4":        ["keyfeaturesbullet4", "bullet4", "keyfeatures4"],
    "bullet_5":        ["keyfeaturesbullet5", "bullet5", "keyfeatures5"],
    "addl":            ["additionaldetailsstandoutscalloutsfeatures",
                        "additionaldetails", "standouts"],
}


def _build_column_map(header_row: List[Any]) -> Dict[str, int]:
    """Given the first row of a sheet, return a map of canonical_key -> 1-based column index."""
    out: Dict[str, int] = {}
    for col_idx, val in enumerate(header_row, start=1):
        norm = _norm_header(val)
        for canonical, aliases in _HEADER_ALIASES.items():
            if norm in aliases:
                if canonical not in out:  # first match wins
                    out[canonical] = col_idx
                break
    return out


def parse_preupload(xlsx_path: str, brand_hint: Optional[str] = None) -> Dict[str, Any]:
    """Parse the pre-upload xlsx into a structured per-style dict.

    Returns:
        {
          "brand":   "Sage Collective" (inferred from TLGDIV NAME or brand_hint),
          "styles": {
              "107010295": {
                "style":       "107010295",
                "sub_class":   "Puffer",
                "department":  "Women's",
                "name":        "Long Stretch Quilted Puffer Coat ...",
                "raw":         {...all parsed columns...},
                "colors":      {"Black", "Truffle"},
                "sizes":       {"Small", "Medium", "Large", "X-Large"},
                "upcs":        ["199...", ...],
                "skus":        ["F26-...", ...]
              },
              ...
          },
          "errors": [...]
        }
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to parse pre-upload files")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # Find the UPC sheet — name varies; look for "Upload Template UPC" or similar
    target_ws = None
    for sn in wb.sheetnames:
        if "upload" in sn.lower() and "upc" in sn.lower():
            target_ws = wb[sn]
            break
    if target_ws is None:
        # Fallback: first sheet
        target_ws = wb[wb.sheetnames[0]]

    header = [target_ws.cell(row=1, column=c).value for c in range(1, target_ws.max_column + 1)]
    cmap = _build_column_map(header)

    if "style" not in cmap:
        return {"brand": None, "styles": {}, "errors": ["No 'STYLE#' column found in sheet"]}

    styles: Dict[str, Dict[str, Any]] = {}
    inferred_brand = brand_hint
    errors: List[str] = []

    # Pass 1 (agency R0 follow-on) — template's row 2 (REQUIRED/OPTIONAL labels)
    # and row 3 (column descriptions like "Style number — must be unique...")
    # were being parsed as real styles. Skip rows whose STYLE# is one of the
    # annotation sentinels or is clearly sentence text rather than an ID.
    _SENTINEL_STYLES = {"REQUIRED", "OPTIONAL", "IMPORTANT", "CONDITIONALLY REQUIRED", "RECOMMENDED"}

    for r in range(2, target_ws.max_row + 1):
        sn_raw = target_ws.cell(row=r, column=cmap["style"]).value
        if not sn_raw:
            continue
        sn = str(sn_raw).strip()
        if not sn:
            continue
        # Skip template-annotation rows
        if sn.upper() in _SENTINEL_STYLES:
            continue
        if " " in sn or len(sn) > 14 or len(sn) < 4:
            # Real style numbers are short alphanumeric IDs; descriptions are sentence-shaped
            continue

        # Extract every known field for this row
        row_data: Dict[str, Any] = {}
        for k, col in cmap.items():
            v = target_ws.cell(row=r, column=col).value
            if v is not None and v != "":
                row_data[k] = v

        # Infer brand from TLGDIV NAME the first time we see a non-empty one
        if not inferred_brand:
            tlgdiv = row_data.get("tlgdiv") or ""
            if isinstance(tlgdiv, str):
                upper = tlgdiv.upper()
                if "SAGE" in upper:
                    inferred_brand = "Sage Collective"
                elif "VOLCOM" in upper:
                    inferred_brand = "Volcom"
                elif "SPYDER" in upper:
                    inferred_brand = "Spyder"
                elif "STELLA" in upper:
                    inferred_brand = "Stella Parker"

        if sn not in styles:
            styles[sn] = {
                "style":      sn,
                "sub_class":  row_data.get("sub_class"),
                "sub_sub":    row_data.get("sub_sub"),
                "department": row_data.get("department"),
                "name":       row_data.get("name"),
                "tlgdiv":     row_data.get("tlgdiv"),
                "model":      row_data.get("model"),
                "coo":        row_data.get("coo"),
                "care":       row_data.get("care"),
                "fabric":     row_data.get("fabric"),
                "closure":    row_data.get("closure"),
                # Pass 3 (item 15) — multi-closure source columns. v1 template
                # doesn't have closure_2, so this will be None on Tahari/Volcom
                # ingests today; the operator's comma-encoded multi in 'closure'
                # is what _split_closures parses. Forward-compat for v2.
                "closure_2":  row_data.get("closure_2"),
                # Pass 3 (item 11b) — same idea for materials. v1 template
                # encodes multi-material via the single fabric percentage
                # string ('95% Polyester, 5% Spandex'); _split_fabric_into_materials
                # parses that. v2 will ship explicit Material 2/3 columns.
                "material_2": row_data.get("material_2"),
                "material_3": row_data.get("material_3"),
                # Pass 6 (strategic 2) — v2 template explicit Amazon attrs.
                "age_range":              row_data.get("age_range"),
                "target_gender_v2":       row_data.get("target_gender"),
                "item_length_description_v2": row_data.get("item_length_description"),
                "sleeve_type":   row_data.get("sleeve_type"),
                "sleeve_length": row_data.get("sleeve_length"),
                "neck":          row_data.get("neck"),
                "fit":           row_data.get("fit"),
                "upf":           row_data.get("upf"),
                "length":     row_data.get("length"),
                "list_price": row_data.get("list_price"),
                "amazon_cost":row_data.get("amazon_cost"),
                "addl":       row_data.get("addl"),
                "keywords":   row_data.get("keywords"),
                "due_date":   row_data.get("due_date"),
                "type_jacket":row_data.get("type_jacket"),
                # 5 bullets (each may be None if the operator didn't fill one)
                "bullets":    [
                    row_data.get("bullet_1"),
                    row_data.get("bullet_2"),
                    row_data.get("bullet_3"),
                    row_data.get("bullet_4"),
                    row_data.get("bullet_5"),
                ],
                "raw":        dict(row_data),
                "colors":     set(),
                "sizes":      set(),
                "upcs":       [],
                "skus":       [],
            }
        if "color" in row_data:
            styles[sn]["colors"].add(row_data["color"])
        if "size" in row_data:
            styles[sn]["sizes"].add(row_data["size"])
        if "upc" in row_data:
            styles[sn]["upcs"].append(str(row_data["upc"]))
        if "sku" in row_data:
            styles[sn]["skus"].append(str(row_data["sku"]))

    # Convert sets to sorted lists for JSON-friendliness
    for s in styles.values():
        s["colors"] = sorted(s["colors"])
        s["sizes"]  = sorted(s["sizes"])

    return {
        "brand":  inferred_brand,
        "styles": styles,
        "errors": errors,
    }


# Pass 6 (agency R0 strategic 2) — template version detection.
# v1 is the legacy TLG-internal template (Tahari/Volcom). v2 is the
# Amazon-attribute-shaped template that drops TLGDIV/MODEL/SKU and adds
# Department/Age Range/Target Gender/Item Length/Sleeve Length/Material 2/3/
# Closure Type 2. Both parse through the same code path — the importer's
# header aliases accept either set. detect_template_version is exposed for
# the UI so it can label which format is being ingested.
_V2_SIGNAL_HEADERS = {"material2", "materialcomposition2", "closuretype2",
                      "closure2", "agerange", "agerangedescription",
                      "targetgender"}
_V1_SIGNAL_HEADERS = {"tlgdivname", "tlgstyledesc", "modelnumber", "sku"}


def detect_template_version(header_row: List[Any]) -> str:
    """Return 'v2' if the header row carries any v2-only column,
    'v1' if it carries v1-only columns, 'unknown' otherwise.

    A header row with BOTH a v2 signal and v1 columns still resolves to 'v2'
    — the v2 ingest path is a superset of v1 (extra Amazon columns; legacy
    TLG columns are just ignored).
    """
    normed = {_norm_header(h) for h in header_row if h is not None}
    has_v2 = bool(normed & _V2_SIGNAL_HEADERS)
    has_v1 = bool(normed & _V1_SIGNAL_HEADERS)
    if has_v2:
        return "v2"
    if has_v1:
        return "v1"
    return "unknown"


def _split_closures(closure_raw: str, closure_2_raw: str = "") -> list[str]:
    """Pass 3 (agency R0 item 15): Closure Type can be multi-valued.

    Two sources:
      1. v2 template ships a dedicated 'Closure Type 2' column → closure_2_raw.
      2. v1 template only has one column — operators today encode multi by
         comma-separating in the single cell: 'Zipper, Snap', 'Zip/Hook & Eye'.
         Split on comma, slash, ampersand, '+'.

    Returns an ordered list of de-duped closure names (already title-cased
    where the original was). Up to 5 — the COAT bundle bundle only validates
    closure#1.type#1 and #2, so #3+ are kept for forward-compat.
    """
    out: list[str] = []
    raws = [closure_raw, closure_2_raw]
    for r in raws:
        s = (r or "").strip()
        if not s:
            continue
        # Split on common multi-value separators; preserve 'Hook & Eye' as a
        # single value by only splitting on ' & ' when surrounded by non-letter
        # context. The safe approach: split on comma, slash, semicolon, '+',
        # and the word ' and ' — NOT on '&', because closure names contain it.
        for chunk in re.split(r"[,/;+]\s*|\s+and\s+", s):
            name = chunk.strip(" .,;:")
            if name and name not in out:
                out.append(name)
    return out[:5]


def _split_fabric_into_materials(fabric: str) -> tuple[list[str], str]:
    """Split a fabric/composition string into (material_names, composition_string).

    Examples:
        "100% Polyester"             -> (["Polyester"], "100% Polyester")
        "95% Polyester, 5% Spandex"  -> (["Polyester", "Spandex"], "95% Polyester, 5% Spandex")
        "Polyester"                  -> (["Polyester"], "Polyester")
        ""                            -> ([], "")

    Always returns the original string as the composition; only the names list is parsed.
    """
    s = (fabric or "").strip()
    if not s:
        return [], ""
    # Pull out every "<num>% <Name>" group, fall back to bare names if no %
    names: list[str] = []
    for chunk in re.split(r"[,/&;|]\s*|\s+and\s+", s):
        chunk = chunk.strip()
        if not chunk:
            continue
        m = re.match(r"^\d+\.?\d*\s*%\s*(.+)$", chunk)
        name = m.group(1).strip() if m else chunk
        # Strip trailing punctuation and any leftover digits
        name = re.sub(r"^[\d.%\s]+", "", name).strip(" .,;:")
        if name and name not in names:
            names.append(name)
    return names, s


def style_to_form_state(
    style: Dict[str, Any],
    brand: str,
    scope_overrides: Optional[Dict[str, Any]] = None,
    batch_overrides: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Convert one parsed pre-upload style row into an NIS form_state dict.

    Pass 4 (agency R0 strategic 1) — scope-aware editing.
    `scope_overrides` is a flat dict of {field_key: value} that should win over
    any value derived from the template/brand defaults. Source is the operator's
    brand_always edits earlier in the session. Empty / None values do not
    override (would clobber valid template content).
    `batch_overrides` is the session-scoped equivalent — same precedence rules.
    Both default to empty for backward compatibility.
    """
    s = style
    dept_raw = (s.get("department") or "").strip()
    # Amazon validates titlecase 'Womens'/'Mens' for COAT/SHIRT/SWIMWEAR/PANTS dropdowns
    # (DRESS/SHORTS use lowercase). Use titlecase — the fuzzy-matcher in the writer will
    # snap to whichever case the active PT's dropdown expects.
    dept = "Womens" if "women" in dept_raw.lower() else (
           "Mens"   if "men"   in dept_raw.lower() else "Womens")
    gender = "Female" if dept.lower() == "womens" else "Male"

    # Pass 6 (strategic 2): v2 template ships Target Gender as an explicit
    # Amazon-validated column. When present, it wins over the inferred value.
    # Accepted Amazon values: Female / Male / Unisex.
    target_gender_v2 = (s.get("target_gender_v2") or "").strip()
    if target_gender_v2:
        if target_gender_v2.lower() in ("female", "male", "unisex"):
            gender = target_gender_v2.title()

    # v2 also ships Age Range. Default 'Adult' when absent (matches v1 behavior
    # since v1 had no age column and apparel_defaults filled it in).
    age_range = (s.get("age_range") or "").strip() or "Adult"

    primary_color = s["colors"][0] if s.get("colors") else ""
    color_value = primary_color.title()
    color_map = _COLOR_MAP.get((primary_color or "").upper(), color_value)

    # Pick a sensible variant size
    sizes = s.get("sizes") or []
    primary_size = "Medium" if "Medium" in sizes else (
                   "M"      if "M"      in sizes else (sizes[0] if sizes else "Medium"))

    sub_class = s.get("sub_class") or ""
    cat, subcat, itk, itn = _SUBCLASS_TAXONOMY.get(
        sub_class,
        ("Women's Outerwear", "Wool", "wool-outerwear-coats", "Coat")
    )
    if dept == "mens":
        cat = cat.replace("Women's", "Men's")

    vendor_code = _BRAND_VENDOR_CODES.get(brand, "")

    pockets_val = s.get("raw", {}).get("pockets")

    # Pass 1 (agency R0 item 11) — split fabric string into material names
    # (parsed list) vs. the original composition string. Material slot gets
    # the names; fabric_type slot keeps the percentage composition.
    material_names, fabric_composition = _split_fabric_into_materials(s.get("fabric") or "")
    # Pass 3 (item 11b): if v2 template supplied explicit Material 2/3 columns,
    # append them ahead of any duplicates already parsed from fabric. Source-of-
    # truth precedence: explicit columns win; the fabric percentage string fills
    # remaining slots.
    for explicit in (s.get("material_2"), s.get("material_3")):
        if explicit:
            ename = str(explicit).strip()
            if ename and ename not in material_names:
                material_names.append(ename)

    # Pass 3 (item 15): parse multi-closure. v1 source = comma-encoded in the
    # single Closure Type cell. v2 source = explicit Closure Type 2 column.
    closure_list = _split_closures(s.get("closure") or "", s.get("closure_2") or "")

    # Pass 1 (agency R0 item 5) — 5 per-bullet columns now flow through.
    # Fall back to legacy "addl" only when no per-bullet content was supplied.
    template_bullets = s.get("bullets") or []
    template_bullets = [str(b).strip() for b in template_bullets if b]
    addl_text = (s.get("addl") or "").strip()
    if not template_bullets and addl_text:
        # Legacy behavior: synthesize bullet#1 from addl when no bullet cols
        template_bullets = [addl_text[:200]]

    # Pass 1 (title gendering case bug) — gender_word's old comparison was
    # case-sensitive: `dept == "womens"` failed when dept was "Womens", so
    # titles came out as "<Brand> Men's ..." even for women's product. Use
    # case-insensitive comparison.
    gw = "Women's" if dept.lower() == "womens" else "Men's"

    state = {
        "rtip_vendor_code#1.value":      vendor_code,
        "vendor_sku#1.value":            str(s.get("model") or s.get("style") or ""),
        "product_type#1.value":          "COAT",
        "parentage_level#1.value":       "Parent",
        "item_name#1.value":             (
            f"{brand} {gw} {s.get('name') or ''}"
        ).strip(),
        "brand#1.value":                 brand,
        "external_product_id#1.type":    "UPC",
        "product_category#1.value":      cat,
        "product_subcategory#1.value":   subcat,
        "item_type_keyword#1.value":     itk,
        "item_type_name#1.value":        itn,
        "model_number#1.value":          str(s.get("model") or ""),
        "model_name#1.value":            s.get("name") or "",
        "style#1.value":                 s.get("name") or "",
        "department#1.value":            dept,
        "target_gender#1.value":         gender,
        "country_of_origin#1.value":     str(s.get("coo") or "").title(),
        "color#1.value":                 color_value,
        "color#1.standardized_values#1": color_map,
        "care_instructions#1.value":     s.get("care") or "",
        # Material now holds the parsed first material name; Fabric Type keeps
        # the full composition string. Pass 3 will extend Material to multi.
        "material#1.value":              material_names[0] if material_names else "",
        "fabric_type#1.value":           fabric_composition,
        # Pass 3 (item 15) — closure is now multi-valued. The COAT bundle
        # validates closure#1.type#1.value and closure#1.type#2.value. If the
        # operator only supplied one value, only the first slot is populated.
        "closure#1.type#1.value":        (closure_list[0] if closure_list else (s.get("closure") or "")),
        # Pass 1 (agency R0 item 13) — Sleeve Type / Length now mapped.
        "sleeve#1.type#1.value":         s.get("sleeve_type") or "",
        "sleeve#1.length_description#1.value": s.get("sleeve_length") or "",
        # Other newly-mapped optional structured attributes.
        "neck#1.value":                  s.get("neck") or "",
        "fit_type#1.value":              s.get("fit") or "Regular",
        "upf_rating#1.value":            s.get("upf") or "",
        "apparel_size#1.size":           primary_size,
        # Pass 1 (agency R0 item 5) — all 5 bullets mapped per column.
        "bullet_point#1.value":          (template_bullets[0] if len(template_bullets) > 0 else "")[:200],
        "bullet_point#2.value":          (template_bullets[1] if len(template_bullets) > 1 else "")[:200],
        "bullet_point#3.value":          (template_bullets[2] if len(template_bullets) > 2 else "")[:200],
        "bullet_point#4.value":          (template_bullets[3] if len(template_bullets) > 3 else "")[:200],
        "bullet_point#5.value":          (template_bullets[4] if len(template_bullets) > 4 else "")[:200],
        "rtip_product_description#1.value":
            ((s.get("addl") or "") + " " + (s.get("keywords") or "")).strip(),
        "list_price#1.value":            str(s.get("list_price") or ""),
        "list_price#1.currency":         "USD",
        # Pass 6 (strategic 2): v2 has an explicit Item Length Description
        # column with valid Amazon vocabulary ('Standard Length' etc). When
        # present, it wins over the v1 derivation (which used CBL inches).
        "item_length_description#1.value": (
            (s.get("item_length_description_v2") or "").strip() or
            (f"{s.get('length')}-inch" if s.get("length") else "")
        ),
        # Pass 6 (strategic 2): v2 ships Age Range as an explicit column.
        # Defaults to 'Adult' when absent.
        "age_range_description#1.value": age_range,
    }
    # Pass 3 (item 11b): expose secondary + tertiary materials. Engine bundle
    # for COAT validates material#1/2/3.value — all three slots flow now.
    for idx, name in enumerate(material_names[1:3], start=2):
        state[f"material#{idx}.value"] = name
    # Pass 3 (item 15): expose secondary closure when present.
    if len(closure_list) >= 2:
        state["closure#1.type#2.value"] = closure_list[1]
    # Wire cost price + ship/booking date from pre-upload (previously unmapped)
    due = s.get("due_date")
    if due is not None:
        try:
            date_str = due.strftime("%Y-%m-%dT%H:%M:%SZ") if hasattr(due, "strftime") else str(due)
            state["rtip_earliest_shipping_date#1.value"] = date_str
            state["merchant_release_date#1.value"] = date_str
            state["item_booking_date#1.value"] = date_str
        except Exception:
            pass
    cost = s.get("amazon_cost")
    if cost is not None:
        state["cost_price#1.value"] = str(cost)
    # Number of pockets — wire from pre-upload (Sage feedback: was missing on template)
    if pockets_val is not None and str(pockets_val).strip():
        try:
            state["number_of_pockets#1.value"] = int(pockets_val)
        except (ValueError, TypeError):
            state["number_of_pockets#1.value"] = str(pockets_val)
    # Pass 4 (strategic 1): apply scope overrides AFTER the template-derived
    # state is built. Precedence: brand_always wins over batch wins over
    # template/auto-derive. Empty/None overrides are ignored so they don't
    # clobber real template content.
    for ov_map in (batch_overrides, scope_overrides):
        if not ov_map:
            continue
        for k, v in ov_map.items():
            if v in (None, "", " "):
                continue
            state[k] = v

    # Clean blanks
    return {k: v for k, v in state.items() if v not in (None, "", " ")}


def gender_word(dept: str) -> str:
    return "Women's" if dept == "womens" else "Men's"
