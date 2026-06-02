"""Build the canonical Pre-Upload Template v2 fixture.

Pass 6 (agency R0 strategic 2) deliverable. Drops TLG-internal columns
(TLGDIV NAME, MODEL NAME, SKU, CHILD ASIN, TLG Style Desc) and ships
Amazon-attribute-shaped columns directly:
  - Department / Age Range / Target Gender as explicit columns
  - Item Length Description with Amazon vocabulary (not CBL inches)
  - Sleeve Length as a first-class column (was missing in v1)
  - Material 1 / 2 / 3 split out (instead of comma-encoded in one cell)
  - Closure Type / Closure Type 2 (instead of comma-encoded in one cell)

Output: data/fixtures/preupload_template_v2_2026_06_02.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "fixtures" / "preupload_template_v2_2026_06_02.xlsx"

# v2 column order. Numbers in comments = position in v1 for reference.
COLUMNS = [
    # ─── Identity ─────────────────────────────────────────────────
    ("STYLE#",                            "REQUIRED",     "Style number — must be unique within the upload"),
    ("STYLE NAME",                        "REQUIRED",     "Operator-facing style name (also used in item_name composition)"),
    ("Brand Code",                        "REQUIRED",     "Brand identifier — used to load brand_config and vendor_code"),
    # ─── Amazon classification (NEW IN V2) ────────────────────────
    ("Department",                        "REQUIRED",     "Womens / Mens / Boys / Girls / Baby / Unisex"),
    ("Age Range",                         "REQUIRED",     "Adult / Big Kid / Little Kid / Toddler / Infant / Newborn"),
    ("Target Gender",                     "REQUIRED",     "Female / Male / Unisex"),
    # ─── Sub-class & variant ──────────────────────────────────────
    ("Sub-Class Name",                    "REQUIRED",     "Drives PT inference + taxonomy resolution"),
    ("Sub-Sub-Class Name",                "OPTIONAL",     "Finer-grained classification when relevant"),
    ("Color Code",                        "REQUIRED",     "Internal color SKU code"),
    ("Color Name",                        "REQUIRED",     "Color label shown to customer"),
    ("Product Size",                      "REQUIRED",     "XS / S / M / L / XL / etc"),
    ("UPC Code",                          "REQUIRED",     "12-digit UPC for the SKU"),
    # ─── Compliance ───────────────────────────────────────────────
    ("Country of Origin",                 "REQUIRED",     "ISO country (US / CN / VN / BD / etc)"),
    ("Care Instructions",                 "REQUIRED",     "Machine wash cold / Dry clean only / etc"),
    # ─── Materials (NEW IN V2: split out) ─────────────────────────
    ("Material 1",                        "REQUIRED",     "Primary material name (Polyester / Cotton / Wool / etc)"),
    ("Material 2",                        "OPTIONAL",     "Secondary material (when blend has 2+ fibers)"),
    ("Material 3",                        "OPTIONAL",     "Tertiary material"),
    ("Fabric Content Percentage",         "REQUIRED",     "Full composition string: '95% Polyester, 5% Spandex'"),
    # ─── Closure (NEW IN V2: secondary) ───────────────────────────
    ("Closure Type",                      "RECOMMENDED",  "Primary closure: Zipper / Button / Snap / Hook & Eye / etc"),
    ("Closure Type 2",                    "OPTIONAL",     "Secondary closure when garment has two (Zipper + Snap)"),
    # ─── Sleeve & length (NEW IN V2: Sleeve Length as column) ─────
    ("Sleeve Type",                       "CONDITIONAL",  "Long Sleeve / Short Sleeve / Sleeveless / 3/4 Sleeve / etc"),
    ("Sleeve Length",                     "CONDITIONAL",  "Hip Length / Standard / Long / etc (was missing in v1)"),
    ("Item Length Description",           "CONDITIONAL",  "Amazon vocabulary: 'Standard Length' / 'Knee-Length' / 'Maxi' / etc"),
    # ─── Construction details ─────────────────────────────────────
    ("Number of Pockets",                 "OPTIONAL",     "Integer 0-6"),
    ("Removable Hood",                    "OPTIONAL",     "Yes / No"),
    # ─── Content ──────────────────────────────────────────────────
    ("Key Features Bullet 1",             "REQUIRED",     "First bullet — ALLCAPS HEADLINE: rest format works"),
    ("Key Features Bullet 2",             "REQUIRED",     "Second bullet"),
    ("Key Features Bullet 3",             "REQUIRED",     "Third bullet"),
    ("Key Features Bullet 4",             "REQUIRED",     "Fourth bullet"),
    ("Key Features Bullet 5",             "REQUIRED",     "Fifth bullet"),
    # ─── Pricing ──────────────────────────────────────────────────
    ("Amazon Cost",                       "REQUIRED",     "Wholesale cost to Amazon"),
    ("Amazon List Price",                 "REQUIRED",     "MAP price"),
    ("Due Date / Earliest Ship Date",     "REQUIRED",     "Date string"),
]


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Upload Template UPC"

    header_fill = PatternFill(start_color="FFE5E7EB", end_color="FFE5E7EB", fill_type="solid")
    req_fill    = PatternFill(start_color="FFFEE2E2", end_color="FFFEE2E2", fill_type="solid")
    opt_fill    = PatternFill(start_color="FFFEF3C7", end_color="FFFEF3C7", fill_type="solid")
    cond_fill   = PatternFill(start_color="FFDBEAFE", end_color="FFDBEAFE", fill_type="solid")
    rec_fill    = PatternFill(start_color="FFE0E7FF", end_color="FFE0E7FF", fill_type="solid")

    fills_by_level = {
        "REQUIRED": req_fill, "OPTIONAL": opt_fill,
        "CONDITIONAL": cond_fill, "RECOMMENDED": rec_fill,
    }

    bold = Font(bold=True, size=10)
    italic = Font(italic=True, size=9, color="FF6B7280")

    # Row 1: column names
    for c, (name, level, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = max(14, min(28, len(name) + 4))

    # Row 2: requirement level
    for c, (_, level, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=c, value=level)
        cell.font = italic
        cell.fill = fills_by_level.get(level, opt_fill)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: column descriptions
    for c, (_, _, desc) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=c, value=desc)
        cell.font = italic
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Add one example data row (Tahari-shaped, for parser smoke-testing)
    sample = {
        "STYLE#": "V2EXAMPLE1",
        "STYLE NAME": "Vera Quilted Long Puffer",
        "Brand Code": "TAHARI",
        "Department": "Womens",
        "Age Range": "Adult",
        "Target Gender": "Female",
        "Sub-Class Name": "Puffer",
        "Sub-Sub-Class Name": "Down and Parka",
        "Color Code": "BLK",
        "Color Name": "Black",
        "Product Size": "Medium",
        "UPC Code": "199555012345",
        "Country of Origin": "CN",
        "Care Instructions": "Machine wash cold; tumble dry low",
        "Material 1": "Polyester",
        "Material 2": "Down",
        "Material 3": "",
        "Fabric Content Percentage": "80% Polyester, 20% Down",
        "Closure Type": "Zipper",
        "Closure Type 2": "Snap",
        "Sleeve Type": "Long Sleeve",
        "Sleeve Length": "Standard Length",
        "Item Length Description": "Long Length",
        "Number of Pockets": 4,
        "Removable Hood": "Yes",
        "Key Features Bullet 1": "ELEVATED WARMTH: Down-fill construction traps body heat for sub-zero days.",
        "Key Features Bullet 2": "WEATHER-READY SHELL: Water-resistant outer with sealed seams.",
        "Key Features Bullet 3": "TAILORED FIT: Cut close to flatter without compromising mobility.",
        "Key Features Bullet 4": "FUNCTIONAL POCKETS: Four secure pockets including interior chest pocket.",
        "Key Features Bullet 5": "REMOVABLE HOOD: Snap-detach hood adapts to weather and styling.",
        "Amazon Cost": 88.00,
        "Amazon List Price": 198.00,
        "Due Date / Earliest Ship Date": "2026-09-15",
    }
    for c, (name, _, _) in enumerate(COLUMNS, start=1):
        ws.cell(row=4, column=c, value=sample.get(name, ""))

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 60
    ws.freeze_panes = "A4"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT))
    print(f"v2 template written: {OUT}")
    print(f"  columns: {len(COLUMNS)}")
    print(f"  example row: 1")


if __name__ == "__main__":
    build()
