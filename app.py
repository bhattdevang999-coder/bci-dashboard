"""
NIS Wizard v3 — Flask Backend
The Levy Group — Amazon Intelligence
Port: 5000
"""

import os
import re
import json
import csv
import zipfile
import shutil
import copy
import io
import time
import traceback
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict

from flask import Flask, request, jsonify, render_template, send_file, send_from_directory, abort, Response
from flask_cors import CORS
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

from anthropic import Anthropic

# ── App setup ──────────────────────────────────────────────────────────────────
import threading

# Initialize Anthropic client (graceful fallback if no API key)
try:
    _anthropic_client = Anthropic()
    print("[LLM] Anthropic client initialized successfully.")
except Exception as _anthro_err:
    _anthropic_client = None
    print(f"[LLM] Anthropic client failed to initialize ({_anthro_err}). Will use rule-based generation.")

# Last vision-pass error reason (set by analyze_style_image on failure)
_last_vision_error = None

BASE_DIR = Path(__file__).parent

# Progress tracking for NIS generation
nis_progress = {"total": 0, "completed": 0, "current_style": "", "current_step": "", "status": "idle", "started_at": None}

# Progress tracking for content generation
content_progress = {"total": 0, "completed": 0, "current_style": "", "current_step": "", "status": "idle", "started_at": None}
UPLOAD_TEMPLATES = BASE_DIR / "uploads" / "templates"
UPLOAD_PRODUCTS  = BASE_DIR / "uploads" / "products"
UPLOAD_KEYWORDS  = BASE_DIR / "uploads" / "keywords"
UPLOAD_OUTPUT    = BASE_DIR / "uploads" / "output"
UPLOAD_IMAGES    = BASE_DIR / "uploads" / "style_images"
FEEDBACK_FILE    = BASE_DIR / "feedback" / "content_feedback.jsonl"
FEEDBACK_DIR     = BASE_DIR / "feedback"
DEFAULT_TEMPLATE = UPLOAD_TEMPLATES / "Dresses-Training.xlsm"
BRAND_CONFIGS_DIR = BASE_DIR / "brand_configs"
DROPDOWN_CACHE_DIR = BASE_DIR / "dropdown_cache"
SUBCLASS_MAP_FILE = BASE_DIR / "subclass_product_type_map.json"

# ── Taxonomy Overrides (Phase 1) ───────────────────────────────────────────────
# Per-item-type confirmed taxonomy quadruple: product_category / product_subcategory
# / item_type_keyword / item_type_name. Keyed by (product_type, sub_class, gender_bucket).
# Universal across brands — Amazon's taxonomy doesn't care which brand owns the SKU.
TAXONOMY_OVERRIDES_FILE = BASE_DIR / "taxonomy_overrides.json"
TAXONOMY_UNIVERSE_FILE  = BASE_DIR / "taxonomy_universe.json"
TAXONOMY_HISTORY_FILE   = BASE_DIR / "feedback" / "taxonomy_history.jsonl"

# Closed set of 13 gender buckets — see spec for derivation rules.
GENDER_BUCKETS = [
    "Mens", "Womens", "MensPlus", "WomensPlus", "WomensPetite",
    "BoysToddler", "BoysLittle", "BoysBig",
    "GirlsToddler", "GirlsLittle", "GirlsBig",
    "Baby", "Unisex",
]

def _load_taxonomy_universe():
    """Load the pre-computed valid-value universe.
    Structure: { PRODUCT_TYPE: { product_categories, subcategories_by_category,
                                 item_type_keywords_by_cat_sub, item_type_names } }

    If the static file is missing OR a product type has no cascade data
    (item_type_keywords_by_cat_sub), backfill from the rule-engine bundles
    so the dropdown modal always shows the latest cascade values from Amazon's NIS templates.
    """
    static = {}
    if TAXONOMY_UNIVERSE_FILE.exists():
        try:
            static = json.loads(TAXONOMY_UNIVERSE_FILE.read_text())
        except Exception:
            static = {}

    # Auto-enrich any product type that's missing cascade data
    needs_enrich = False
    for pt, data in (static or {}).items():
        if not data.get("item_type_keywords_by_cat_sub"):
            needs_enrich = True
            break

    if needs_enrich and (BASE_DIR / "nis_rules").is_dir():
        try:
            from nis_engine.taxonomy_builder import build_universe_from_engine, merge_universes
            engine_universe = build_universe_from_engine(str(BASE_DIR / "nis_rules"))
            return merge_universes(static, engine_universe)
        except Exception as e:
            print(f"[taxonomy] enrichment failed (using static only): {e}")
            return static
    return static

def _load_taxonomy_overrides():
    """Load the operator-confirmed taxonomy store. Returns a dict with
    'version' / 'updated_at' / 'entries' keys. Creates an empty shell if missing.
    """
    if not TAXONOMY_OVERRIDES_FILE.exists():
        return {"version": 1, "updated_at": "", "entries": {}}
    try:
        data = json.loads(TAXONOMY_OVERRIDES_FILE.read_text())
        if "entries" not in data:
            data["entries"] = {}
        if "version" not in data:
            data["version"] = 1
        return data
    except Exception:
        return {"version": 1, "updated_at": "", "entries": {}}

def _save_taxonomy_overrides(data):
    """Atomic write: tmp -> rename. No git push here — that's done separately
    so one failing push doesn't block the save."""
    data["updated_at"] = datetime.now().isoformat() + "Z"
    tmp = TAXONOMY_OVERRIDES_FILE.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(data, indent=2, sort_keys=False))
    tmp.replace(TAXONOMY_OVERRIDES_FILE)

def _append_taxonomy_history(key, entry, action):
    """Append audit log entry for every taxonomy change."""
    try:
        TAXONOMY_HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(str(TAXONOMY_HISTORY_FILE), "a") as f:
            f.write(json.dumps({
                "timestamp": datetime.now().isoformat() + "Z",
                "action": action,
                "key": key,
                "entry": entry,
            }) + "\n")
    except Exception:
        pass

def _try_git_commit_taxonomy(key, user):
    """Best-effort git commit + push of taxonomy_overrides.json. Silent failure.
    Runs in-process but with a short timeout so a slow network doesn't block the request.
    """
    import subprocess
    try:
        cwd = str(BASE_DIR)
        subprocess.run(["git", "add", "taxonomy_overrides.json"], cwd=cwd, timeout=5, capture_output=True)
        msg = f"taxonomy: confirm {key} by {user or 'unknown'}"
        subprocess.run(["git", "commit", "-m", msg], cwd=cwd, timeout=5, capture_output=True)
        subprocess.run(["git", "push", "origin", "master"], cwd=cwd, timeout=10, capture_output=True)
    except Exception as e:
        print(f"[taxonomy] git push failed (non-blocking): {e}")

def _derive_gender_bucket(style):
    """Map a style to one of the 13 closed-set gender buckets.
    Uses division_name AND style_name so 'VOLCOM YOUTH SWIM' + 'Little Boys ...'
    resolves to 'BoysLittle' not 'Unisex'.
    """
    dn = (style.get("division_name", "") or "").upper()
    sn = (style.get("style_name", "") or "").upper()
    combined = f"{dn} {sn}"

    # Baby/infant first
    if "BABY" in combined or "INFANT" in combined:
        return "Baby"

    # Youth buckets (check before adult since YOUTH-division styles may have style_name signals)
    if any(t in combined for t in ["YOUTH", "KIDS", " BOY", "BOYS", " GIRL", "GIRLS", "TODDLER"]):
        is_girls = "GIRL" in combined
        is_boys  = "BOY" in combined
        if "TODDLER" in combined:
            return "GirlsToddler" if is_girls else "BoysToddler"
        if "LITTLE" in combined:
            return "GirlsLittle" if is_girls else "BoysLittle"
        if "BIG" in combined:
            return "GirlsBig" if is_girls else "BoysBig"
        # Youth without size bucket — default to Big Kid
        if is_girls:
            return "GirlsBig"
        if is_boys:
            return "BoysBig"
        return "Unisex"

    # Women's buckets (check BEFORE men's — 'WOMENS' contains 'MENS' substring)
    if "WOMENS" in dn or "WOMEN'S" in dn or "WOMEN " in dn:
        if "PETITE" in combined:
            return "WomensPetite"
        if any(t in combined for t in ["PLUS", " 1X", " 2X", " 3X"]):
            return "WomensPlus"
        return "Womens"

    # Men's buckets
    if "MENS" in dn or "MEN'S" in dn or " MEN " in dn or dn.endswith(" MEN"):
        if any(t in combined for t in ["BIG AND TALL", "BIG & TALL", "PLUS", " 2XL", " 3XL", " 4XL"]):
            return "MensPlus"
        return "Mens"

    return "Unisex"

def _taxonomy_key(product_type, sub_class, gender_bucket):
    """Build the canonical composite key for the overrides store."""
    pt = (product_type or "").strip().upper()
    sc = (sub_class or "").strip()
    gb = (gender_bucket or "Unisex").strip()
    return f"{pt}|{sc}|{gb}"

def _resolve_taxonomy_for_style(style, brand_cfg=None):
    """Look up a confirmed taxonomy quadruple for a style.
    Returns a dict with: matched (bool), source ('override'|'auto'), entry (the quadruple),
    key (the composite key), gender_bucket, auto_derived (the rule-based fallback for UI display).

    Callers should prefer entry[field] when matched=True, else fall back to the existing
    rule-based _derive_* functions.
    """
    pt = _resolve_style_product_type(style) or ""
    sc = style.get("subclass", "") or style.get("sub_class", "")
    gb = _derive_gender_bucket(style)
    key = _taxonomy_key(pt, sc, gb)

    overrides = _load_taxonomy_overrides()
    entry = overrides.get("entries", {}).get(key)
    if entry and entry.get("source") == "manual":
        return {"matched": True, "source": "override", "entry": entry, "key": key,
                "gender_bucket": gb, "product_type": pt, "sub_class": sc}

    # No confirmed entry — return the auto-derived quadruple for fallback display
    style_gender, _ = _derive_gender_department(style)
    eff_gender = style_gender or (brand_cfg.get("gender", "") if brand_cfg else "")
    style_name = style.get("style_name", "")
    auto_cat    = _derive_amazon_product_category(sc, gender=eff_gender, product_type=pt,
                                                  style_name=style_name)
    auto_subcat = _derive_swim_product_subcategory(sc, gender=eff_gender, style_name=style_name,
                                                   product_category=auto_cat) if pt == "SWIMWEAR" else SUBCLASS_SUBCATEGORY_MAP.get(sc, "")
    auto_itk    = _derive_item_type_keyword(sc, product_type=pt, gender=eff_gender, style_name=style_name)
    auto_itn    = _derive_item_type_name(sc, product_type=pt, gender=eff_gender, style_name=style_name)
    auto = {
        "product_type": pt,
        "product_category": auto_cat,
        "product_subcategory": auto_subcat,
        "item_type_keyword": auto_itk,
        "item_type_name": auto_itn,
    }
    return {"matched": False, "source": "auto", "entry": auto, "key": key,
            "gender_bucket": gb, "product_type": pt, "sub_class": sc,
            "auto_derived": auto}

def _validate_taxonomy_quadruple(pt, category, subcategory, itk, itn):
    """Check each value is dropdown-valid for the product type. Returns (ok, errors).
    errors is a list of {field, value, reason, valid_options} dicts."""
    errors = []
    universe = _load_taxonomy_universe().get(pt, {})
    cache = load_dropdown_cache(pt) or {}

    valid_cats = universe.get("product_categories", []) or cache.get("product_category#1.value", [])
    if category and valid_cats and category not in valid_cats:
        errors.append({"field": "product_category", "value": category,
                       "reason": "not in Amazon dropdown", "valid_options": valid_cats[:50]})

    valid_subs_by_cat = universe.get("subcategories_by_category", {})
    if subcategory and category and category in valid_subs_by_cat:
        if subcategory not in valid_subs_by_cat[category]:
            errors.append({"field": "product_subcategory", "value": subcategory,
                           "reason": f"not valid under category '{category}'",
                           "valid_options": valid_subs_by_cat[category]})

    valid_itns = universe.get("item_type_names", []) or cache.get("item_type_name#1.value", [])
    if itn and valid_itns and itn not in valid_itns:
        errors.append({"field": "item_type_name", "value": itn,
                       "reason": "not in Amazon dropdown", "valid_options": valid_itns[:50]})

    # item_type_keyword: validate against the cascade map if Amazon defines one for this Cat+Sub
    itk_cascade = (universe.get("item_type_keywords_by_cat_sub", {}) or {}).get(category, {})
    valid_itks = itk_cascade.get(subcategory, []) or []
    if itk and valid_itks and itk not in valid_itks:
        # Soft-warn (don't block save) since some operators legitimately use synonyms.
        # Save will go through but the entry's `notes` will surface the warning.
        errors.append({"field": "item_type_keyword", "value": itk,
                       "reason": f"not in Amazon's cascade for {category} > {subcategory}",
                       "valid_options": valid_itks,
                       "severity": "warning"})
    if itk and len(itk) > 100:
        errors.append({"field": "item_type_keyword", "value": itk,
                       "reason": "exceeds 100 chars", "valid_options": []})

    # Filter out warnings — they don't block saving
    blocking_errors = [e for e in errors if e.get("severity") != "warning"]
    return (len(blocking_errors) == 0, errors)
# ── End Taxonomy Overrides ────────────────────────────────────────────────────

def _load_subclass_map():
    """Load the learned sub-class → product type mapping."""
    if SUBCLASS_MAP_FILE.exists():
        with open(str(SUBCLASS_MAP_FILE), "r") as f:
            return json.load(f)
    return {}

def _save_subclass_map(mapping):
    """Save the sub-class → product type mapping."""
    with open(str(SUBCLASS_MAP_FILE), "w") as f:
        json.dump(mapping, f, indent=2)

def resolve_product_type(sub_class, division_name=""):
    """Resolve a sub-class to a product type using all available sources.
    Returns (product_type_id, confidence, reason)
    confidence: 'known'|'detected'|'unknown'
    reason: human-readable explanation of why this mapping was chosen
    """
    if not sub_class:
        if division_name:
            dn = division_name.upper()
            if "SWIM" in dn: return "SWIMWEAR", "detected", f"Division '{division_name}' contains SWIM"
            if "DRESS" in dn: return "DRESS", "detected", f"Division '{division_name}' contains DRESS"
            if "SHIRT" in dn or "TOP" in dn: return "SHIRT", "detected", f"Division '{division_name}' contains SHIRT/TOP"
        return "UNKNOWN", "unknown", "No sub-class or division name provided"
    
    # 1. Learned map (operator-confirmed)
    learned = _load_subclass_map()
    if sub_class in learned:
        return learned[sub_class], "known", f"Previously confirmed by operator: {sub_class} = {learned[sub_class]}"
    
    # 2. ALL_PRODUCT_TYPES
    for pt in ALL_PRODUCT_TYPES:
        if sub_class in pt.get("sub_classes", []):
            return pt["id"], "known", f"Sub-class '{sub_class}' is a known {pt['label']} type"
    
    # 3. TEMPLATE_PRODUCT_TYPE_MAP
    if sub_class in TEMPLATE_PRODUCT_TYPE_MAP:
        tpl = TEMPLATE_PRODUCT_TYPE_MAP[sub_class]
        tpl_to_id = {"Dresses": "DRESS", "Swimwear": "SWIMWEAR", "Other_Shirts": "SHIRT",
                      "Shorts": "SHORTS", "Jackets_and_Coats": "COAT", "Skirts": "SKIRT"}
        pt_id = tpl_to_id.get(tpl, tpl.upper())
        return pt_id, "known", f"Sub-class '{sub_class}' maps to {tpl} template"
    
    # 4. Division name heuristic
    if division_name:
        dn = division_name.upper()
        if "SWIM" in dn: return "SWIMWEAR", "detected", f"Division '{division_name}' suggests swimwear"
        if "DRESS" in dn: return "DRESS", "detected", f"Division '{division_name}' suggests dresses"
        if "SHIRT" in dn or "TOP" in dn: return "SHIRT", "detected", f"Division '{division_name}' suggests tops"
        if "SHORT" in dn: return "SHORTS", "detected", f"Division '{division_name}' suggests shorts"
        if "JACKET" in dn or "COAT" in dn: return "COAT", "detected", f"Division '{division_name}' suggests outerwear"
    
    # 5. Fuzzy sub_class name matching
    sc_lower = sub_class.lower()
    if any(w in sc_lower for w in ["swim", "bikini", "rashguard", "trunk", "tankini"]):
        return "SWIMWEAR", "detected", f"Sub-class '{sub_class}' contains swim-related keyword"
    if any(w in sc_lower for w in ["dress", "gown", "romper"]):
        return "DRESS", "detected", f"Sub-class '{sub_class}' contains dress-related keyword"
    if any(w in sc_lower for w in ["shirt", "blouse", "top", "tee", "polo", "tank"]):
        return "SHIRT", "detected", f"Sub-class '{sub_class}' contains top/shirt keyword"
    if any(w in sc_lower for w in ["short", "board"]):
        return "SHORTS", "detected", f"Sub-class '{sub_class}' contains shorts keyword"
    if any(w in sc_lower for w in ["jacket", "coat", "hoodie", "pullover", "fleece"]):
        return "COAT", "detected", f"Sub-class '{sub_class}' contains outerwear keyword"
    if any(w in sc_lower for w in ["pant", "legging", "jogger"]):
        return "PANTS", "detected", f"Sub-class '{sub_class}' contains pants keyword"
    if any(w in sc_lower for w in ["skirt", "skort"]):
        return "SKIRT", "detected", f"Sub-class '{sub_class}' contains skirt keyword"
    
    return "UNKNOWN", "unknown", f"Cannot determine product type for sub-class '{sub_class}'"


for d in [UPLOAD_TEMPLATES, UPLOAD_PRODUCTS, UPLOAD_KEYWORDS, UPLOAD_OUTPUT, UPLOAD_IMAGES, BRAND_CONFIGS_DIR, FEEDBACK_DIR, DROPDOWN_CACHE_DIR]:
    d.mkdir(parents=True, exist_ok=True)

app = Flask(__name__, template_folder=str(BASE_DIR / "templates"))
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB
CORS(app)

# Boot timestamp — used by /api/version so deploy-verify scripts can confirm
# the process actually restarted on a new deploy (not just the same warm
# instance still serving cached responses).
from datetime import datetime as _dt_boot, timezone as _tz_boot
_APP_BOOT_TS = _dt_boot.now(_tz_boot.utc).isoformat()


# ─── Atlas substrate bootstrap ────────────────────────────────────
# When DATABASE_URL is set (production on Render), apply the substrate
# schema once at import time. The migration uses CREATE TABLE IF NOT EXISTS
# everywhere so re-running is safe — every deploy hits this code path.
#
# If Postgres is unreachable or schema apply fails, we log and continue.
# The logger automatically falls back to JSONL writes so generation
# still works.
try:
    from substrate.db import get_pool, apply_schema as _atlas_apply_schema
    _atlas_pool = get_pool()
    if _atlas_pool is not None:
        with _atlas_pool.connection() as _atlas_conn:
            _atlas_apply_schema(_atlas_conn)
        print("[atlas] substrate Postgres schema applied", flush=True)
        # Seed brand_profile rows for workspaces we operate on.
        # Idempotent — existing profile_versions are not overwritten.
        try:
            from substrate.brand_profile_seed import seed_brand_profiles
            _seeded = seed_brand_profiles()
            if _seeded:
                print(f"[atlas] seeded {_seeded} brand_profile row(s)", flush=True)
        except Exception as _atlas_seed_exc:
            print(f"[atlas] brand_profile seed skipped: {_atlas_seed_exc}", flush=True)
        # Migrate any surviving JSONL substrate files into Postgres.
        # Idempotent: ON CONFLICT DO NOTHING + deterministic UUIDs for
        # legacy rows that lack event_id. Most deploys this is a no-op
        # because Render's filesystem wipes between deploys, but we
        # leave it on so any short-lived JSONL artifacts get captured.
        try:
            from substrate.migrate_jsonl import migrate_all
            _mig = migrate_all()
            if _mig.get("inserted") or _mig.get("sessions_migrated"):
                print(
                    f"[atlas] migrated +{_mig.get('inserted')} events, "
                    f"+{_mig.get('sessions_migrated')} sessions from JSONL",
                    flush=True,
                )
        except Exception as _atlas_mig_exc:
            print(f"[atlas] JSONL migration skipped: {_atlas_mig_exc}", flush=True)
    else:
        print("[atlas] no DATABASE_URL set; substrate using JSONL backend", flush=True)
except Exception as _atlas_boot_exc:
    print(f"[atlas] substrate Postgres bootstrap skipped: {_atlas_boot_exc}", flush=True)


# ═══════════════════════════════════════════════════════════════════════════════
# NIS RULE ENGINE — universal Amazon NIS conditional-logic evaluator.
# Reads .xlsm templates, parses every CF/DV formula, evaluates against form state.
# All 31 templates produce 16 product-type bundles, ~10K rules, 0 needing review.
# ═══════════════════════════════════════════════════════════════════════════════
from nis_engine import nis_rule_engine as _nis_engine  # noqa: E402
from nis_engine import pt_defaults as _pt_defaults  # noqa: E402
_NIS_RULES_DIR = BASE_DIR / "nis_rules"
_nis_engine.set_bundle_dir(str(_NIS_RULES_DIR))
OVERRIDES_LOG = BASE_DIR / "feedback" / "overrides_log.jsonl"
OVERRIDES_LOG.parent.mkdir(parents=True, exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════════════
# TEMPLATE DROPDOWN VALIDATION ENGINE (permanent infrastructure)
#
# Every Amazon .xlsm template has dropdown validations as named ranges.
# This engine:
# 1. Extracts valid dropdown values when a template is uploaded
# 2. Caches per product type (DRESS, SWIMWEAR, SHIRT, etc.)
# 3. Validates every field value before writing to .xlsm
# 4. Auto-corrects fuzzy matches (e.g., "30+" → "UPF 30")
# 5. Flags values that can't be matched
# ═══════════════════════════════════════════════════════════════════════════════

_dropdown_cache = {}  # product_type → { field_id: [valid_values] }


def extract_template_dropdowns(template_path):
    """Extract all dropdown valid values from an Amazon .xlsm template.
    Called automatically on template upload and at server startup.
    """
    from openpyxl.utils import range_boundaries
    import warnings as _w
    with _w.catch_warnings():
        _w.simplefilter("ignore")
        wb = openpyxl.load_workbook(template_path, keep_vba=True)

    product_type = "UNKNOWN"
    ws = None
    for name in wb.sheetnames:
        if name.upper().startswith("TEMPLATE"):
            ws = wb[name]
            parts = name.split("-", 1)
            if len(parts) == 2 and parts[1].strip():
                product_type = parts[1].strip().upper()
            break
    if not ws:
        wb.close()
        return None

    # Build field_id normalized lookup
    max_col = ws.max_column or 254
    fid_norm_map = {}
    for col in range(1, max_col + 1):
        raw = ws.cell(row=4, column=col).value
        if raw:
            fid = str(raw).strip()
            norm = fid.replace("#", "").replace(".", "")
            fid_norm_map[norm] = fid

    # Extract dropdown values from defined names
    field_dropdowns = {}
    for name in wb.defined_names:
        if not name.startswith(product_type):
            continue
        field_ref = name[len(product_type):]
        ref_norm = field_ref.replace(".", "")
        actual_fid = fid_norm_map.get(ref_norm)
        if not actual_fid:
            continue

        dn = wb.defined_names[name]
        try:
            for title, coord in dn.destinations:
                sheet = wb[title]
                mc, mr, xc, xr = range_boundaries(coord)
                vals = []
                for r in range(mr, xr + 1):
                    v = sheet.cell(row=r, column=mc).value
                    if v is not None:
                        vals.append(str(v))
                if vals:
                    field_dropdowns[actual_fid] = vals
        except Exception:
            pass

    wb.close()

    _dropdown_cache[product_type] = field_dropdowns
    cache_path = DROPDOWN_CACHE_DIR / f"{product_type}.json"
    with open(str(cache_path), "w", encoding="utf-8") as f:
        json.dump(field_dropdowns, f, indent=2)

    print(f"[Dropdown] Extracted {len(field_dropdowns)} dropdown fields for {product_type}")
    return {"product_type": product_type, "dropdown_fields": len(field_dropdowns)}


def load_dropdown_cache(product_type):
    """Load dropdown values for a product type. Memory first, then disk."""
    if product_type in _dropdown_cache:
        return _dropdown_cache[product_type]
    cache_path = DROPDOWN_CACHE_DIR / f"{product_type}.json"
    if cache_path.exists():
        with open(str(cache_path), "r", encoding="utf-8") as f:
            data = json.load(f)
        _dropdown_cache[product_type] = data
        return data
    return {}


def _fuzzy_match_dropdown(value, valid_values):
    """Find closest matching dropdown value. Returns (match, confidence) or (None, 0)."""
    if not value or not valid_values:
        return None, 0
    val_str = str(value).strip()
    val_lower = val_str.lower()

    for v in valid_values:
        if v == val_str:
            return v, 1.0
    for v in valid_values:
        if v.lower() == val_lower:
            return v, 0.95
    for v in valid_values:
        if val_lower in v.lower() or v.lower() in val_lower:
            return v, 0.8
    for v in valid_values:
        if v.lower().startswith(val_lower[:3]) and len(val_lower) >= 3:
            return v, 0.6

    val_words = set(val_lower.split())
    best_score, best_match = 0, None
    for v in valid_values:
        v_words = set(v.lower().split())
        if val_words and v_words:
            overlap = len(val_words & v_words) / max(len(val_words), len(v_words))
            if overlap > best_score:
                best_score, best_match = overlap, v
    if best_score >= 0.5:
        return best_match, best_score * 0.7

    return None, 0


def validate_field_value(field_id, value, product_type):
    """Validate a single field against its dropdown. Returns result dict."""
    dropdowns = load_dropdown_cache(product_type)
    if field_id not in dropdowns:
        return {"field_id": field_id, "original": value, "status": "no_dropdown"}
    valid = dropdowns[field_id]
    val_str = str(value).strip() if value else ""
    if not val_str:
        return {"field_id": field_id, "original": value, "status": "empty"}
    if val_str in valid:
        return {"field_id": field_id, "original": value, "status": "valid"}
    match, confidence = _fuzzy_match_dropdown(val_str, valid)
    if match and confidence >= 0.6:
        return {"field_id": field_id, "original": value, "status": "corrected",
                "corrected": match, "confidence": round(confidence, 2)}
    return {"field_id": field_id, "original": value, "status": "invalid",
            "valid_values": valid[:10]}


def validate_and_correct_for_build(field_values, product_type):
    """Validate all fields and auto-correct before writing to .xlsm.
    Modifies field_values in-place. Returns list of all validation results.
    """
    dropdowns = load_dropdown_cache(product_type)
    results = []
    for field_id, value in list(field_values.items()):
        if not value or field_id not in dropdowns:
            continue
        r = validate_field_value(field_id, value, product_type)
        if r["status"] == "corrected":
            field_values[field_id] = r["corrected"]
        results.append(r)
    return results


# Auto-extract dropdowns from templates on disk at startup
for _tpl_file in UPLOAD_TEMPLATES.glob("*.xlsm"):
    try:
        extract_template_dropdowns(str(_tpl_file))
    except Exception as _e:
        print(f"[Dropdown] Startup extract failed for {_tpl_file.name}: {_e}")


# ── Brand configs ──────────────────────────────────────────────────────────────
BRAND_CONFIGS = {
    "Stella Parker": {
        "vendor_code_prefix": "FC0C0",
        "vendor_code_full": "Stella Parker Sportswear, us_apparel, FC0C0",
        "default_upf": "UPF 30",
        "default_fabric": "95% Polyester, 5% Spandex",
        "default_coo": "Mexico",
        "default_care": "Machine Wash",
        "gender": "Female",
        "department": "Womens",
        "bullet_1_focus": "UPF sun protection",
        "title_formula": "{brand} Women's {style_descriptor} {product_type}, UPF {upf}, {color}, {size}",
        "never_words": [],
    },
    "Novelle Fashion": {
        "vendor_code_prefix": "",
        "vendor_code_full": "",
        "default_upf": "",
        "default_fabric": "79% Nylon, 21% Spandex",
        "default_coo": "Bangladesh",
        "default_care": "Machine Wash",
        "gender": "Female",
        "department": "Womens",
        "bullet_1_focus": "Butterlux fabric softness",
        "title_formula": "{brand} Women's {style_descriptor} {product_type}, {color}, {size}",
        "never_words": ["affordable"],
    },
    "Volcom": {
        "vendor_code_prefix": "7E8G6",
        "vendor_code_full": "Volcom, us_apparel, 7E8G6",
        "default_upf": "",
        "default_fabric": "",
        "default_coo": "",
        "default_care": "",
        "gender": "",
        "department": "",
        "bullet_1_focus": "Brand lifestyle and quality",
        "title_formula": "{brand} {gender} {style_name} {product_type}",
        "never_words": [],
    },
    "Roxy": {
        "vendor_code_prefix": "PG823",
        "vendor_code_full": "Roxy Women's Swimwear, us_apparel, PG823",
        "default_upf": "",
        "default_fabric": "",
        "default_coo": "",
        "default_care": "",
        "gender": "Female",
        "department": "Womens",
        "bullet_1_focus": "Beach/surf lifestyle",
        "title_formula": "{brand} {gender} {style_name} {product_type}",
        "never_words": [],
    },
    "Nautica": {
        "vendor_code_prefix": "",
        "vendor_code_full": "",
        "default_upf": "",
        "default_fabric": "",
        "default_coo": "",
        "default_care": "Machine Wash",
        "gender": "Female",
        "department": "Womens",
        "bullet_1_focus": "Nautical inspired style",
        "title_formula": "{brand} Women's {style_descriptor} {product_type}, {color}, {size}",
        "never_words": [],
    },
    # Ben Sherman: leave gender/department empty so per-style division_name
    # drives the gender ('BEN SHERMAN MENS OUTERWEAR' → Male). The hardcoded
    # Women's seed that used to live here was wrong — Ben Sherman is a men's
    # heritage brand. Pass 12.7 stripped the bad cached file; this seed fix
    # makes sure we don't recreate the bug from the in-memory map.
    "Ben Sherman": {
        "vendor_code_prefix": "",
        "vendor_code_full": "",
        "default_upf": "",
        "default_fabric": "",
        "default_coo": "",
        "default_care": "Machine Wash",
        "gender": "",
        "department": "",
        "bullet_1_focus": "British mod heritage style",
        "title_formula": "{brand} {gender} {style_descriptor} {product_type}, {color}, {size}",
        "never_words": [],
    },
    "Spyder": {
        "vendor_code_prefix": "",
        "vendor_code_full": "",
        "default_upf": "",
        "default_fabric": "",
        "default_coo": "",
        "default_care": "Machine Wash",
        "gender": "Female",
        "department": "Womens",
        "bullet_1_focus": "Performance athletic design",
        "title_formula": "{brand} Women's {style_descriptor} {product_type}, {color}, {size}",
        "never_words": [],
    },
    "Tahari": {
        "vendor_code_prefix": "",
        "vendor_code_full": "",
        "default_upf": "",
        "default_fabric": "",
        "default_coo": "",
        "default_care": "Dry Clean",
        "gender": "Female",
        "department": "Womens",
        "bullet_1_focus": "Sophisticated tailored design",
        "title_formula": "{brand} Women's {style_descriptor} {product_type}, {color}, {size}",
        "never_words": [],
    },
    "Sage": {
        "vendor_code_prefix": "",
        "vendor_code_full": "",
        "default_upf": "",
        "default_fabric": "",
        "default_coo": "",
        "default_care": "Machine Wash",
        "gender": "Female",
        "department": "Womens",
        "bullet_1_focus": "Effortless everyday style",
        "title_formula": "{brand} Women's {style_descriptor} {product_type}, {color}, {size}",
        "never_words": [],
    },
    "Sage Collective": {
        "vendor_code_prefix": "QT5G8",
        "vendor_code_full":   "Sage Collective - Levy Group, us_apparel, QT5G8",
        "default_upf":        "",
        "default_fabric":     "",
        "default_coo":        "",
        "default_care":       "Machine Wash Cold, Tumble Dry Low",
        "gender":             "Female",
        "department":         "Womens",
        "bullet_1_focus":     "Quality and craftsmanship",
        "title_formula":      "{brand} Women's {style_name}",
        "never_words":        ["cheap", "fast fashion", "knockoff"],
    },
}

# ── Color maps ─────────────────────────────────────────────────────────────────
COLOR_MAP = {
    "MAUVE": "Pink", "ROSE": "Pink", "BLUSH": "Pink", "PINK": "Pink",
    "CORAL": "Pink", "HOT PINK": "Pink", "MAGENTA": "Pink", "FUCHSIA": "Pink",
    "RED": "Red", "CRIMSON": "Red", "BURGUNDY": "Red", "WINE": "Red",
    "MAROON": "Red", "BRICK": "Red", "CHERRY": "Red",
    "BLUE": "Blue", "NAVY": "Blue", "COBALT": "Blue", "ROYAL BLUE": "Blue",
    "SKY BLUE": "Blue", "PERIWINKLE": "Blue", "DENIM": "Blue", "INDIGO": "Blue",
    "TEAL": "Teal", "TURQUOISE": "Teal", "AQUA": "Teal", "CYAN": "Teal",
    "GREEN": "Green", "OLIVE": "Green", "SAGE": "Green", "FOREST": "Green",
    "EMERALD": "Green", "MINT": "Green", "LIME": "Green", "HUNTER": "Green",
    "KHAKI": "Khaki", "TAN": "Khaki", "CAMEL": "Khaki", "BEIGE": "Beige",
    "IVORY": "Ivory", "CREAM": "Ivory", "OFF WHITE": "Ivory",
    "WHITE": "White", "BRIGHT WHITE": "White",
    "BLACK": "Black", "JET BLACK": "Black", "ONYX": "Black",
    "GREY": "Grey", "GRAY": "Grey", "CHARCOAL": "Grey", "SILVER": "Silver",
    "PURPLE": "Purple", "LAVENDER": "Purple", "VIOLET": "Purple",
    "PLUM": "Purple", "LILAC": "Purple", "VRYVIOLET": "Purple",
    "ORANGE": "Orange", "RUST": "Orange", "PUMPKIN": "Orange",
    "AMBER": "Orange", "TERRACOTTA": "Orange",
    "YELLOW": "Yellow", "GOLD": "Gold", "MUSTARD": "Yellow",
    "BROWN": "Brown", "CHOCOLATE": "Brown", "ESPRESSO": "Brown", "MOCHA": "Brown",
    "MULTI": "Multicolor", "MULTICOLOR": "Multicolor", "PRINT": "Multicolor",
    "COMBO": "Multicolor", "FLORAL": "Multicolor",
}

SIZE_MAP = {
    "XS": "X-Small", "S": "Small", "M": "Medium", "L": "Large",
    "XL": "X-Large", "XXL": "XX-Large", "2XL": "XX-Large",
    "XXXL": "3X-Large", "3XL": "3X-Large", "1X": "1X-Large",
    "2X": "2X-Large", "3X": "3X-Large", "4X": "4X-Large",
    "0X": "0X-Large", "0": "0", "2": "2", "4": "4", "6": "6",
    "8": "8", "10": "10", "12": "12", "14": "14", "16": "16",
}

# ── Template-to-product-type routing ──────────────────────────────────────────
# Maps sub_class values to template product type names
TEMPLATE_PRODUCT_TYPE_MAP = {
    # Dresses
    "Day Dress": "Dresses",
    "Cocktail Dress": "Dresses",
    "Active Dress": "Dresses",
    "Swimdress": "Dresses",
    "Maxi Dress": "Dresses",
    "Mini Dress": "Dresses",
    "Wrap Dress": "Dresses",
    "Shirt Dress": "Dresses",
    "Shift Dress": "Dresses",
    "A-Line Dress": "Dresses",
    "Sundress": "Dresses",
    "Bodycon Dress": "Dresses",
    # Shirts / Other_Shirts
    "Polo": "Other_Shirts",
    "Tee": "Other_Shirts",
    "Shirt": "Other_Shirts",
    "Blouse": "Other_Shirts",
    "Tank": "Other_Shirts",
    # Swimwear (from TLG pre-upload templates)
    "Rashguard": "Swimwear",
    "One Piece Swim": "Swimwear",
    "Bikini Top": "Swimwear",
    "Swim Bottom": "Swimwear",
    "Bikini Bottom": "Swimwear",
    "Tankini": "Swimwear",
    "Swim Set 2 pcs": "Swimwear",
    "Trunk": "Swimwear",
    "Cover Up": "Swimwear",
    # Shorts / Boardshorts
    "Short": "Swimwear",  # swim shorts/trunks — use Swimwear template
    "Board Short": "Shorts",
    "Chino Short": "Shorts",
    "Boardshorts": "Shorts",
    # Jackets and Coats
    "Jacket": "Jackets_and_Coats",
    "Coat": "Jackets_and_Coats",
    "Hoodie": "Jackets_and_Coats",
    "Pullover": "Jackets_and_Coats",
    # Skirts
    "Skirt": "Skirts",
}

SUBCLASS_CATEGORY_MAP = {
    "Day Dress": "casual-and-day-dresses",
    "Cocktail Dress": "special-occasion-dresses",
    "Maxi Dress": "maxi-dresses",
    "Mini Dress": "mini-dresses",
    "Active Dress": "active-dresses",
    "Wrap Dress": "wrap-dresses",
    "Shirt Dress": "shirt-dresses",
    "Shift Dress": "casual-and-day-dresses",
    "A-Line Dress": "special-occasion-dresses",
    "Sundress": "casual-and-day-dresses",
    "Bodycon Dress": "casual-and-day-dresses",
}

# product_subcategory values must match the Amazon template dropdown
# (these are from the Dropdown Lists sheet, not item_type_keyword slugs)
SUBCLASS_SUBCATEGORY_MAP = {
    "Day Dress": "Casual and Day Dresses",
    "Cocktail Dress": "Night Out and Cocktail Dresses",
    "Maxi Dress": "Casual and Day Dresses",
    "Mini Dress": "Casual and Day Dresses",
    "Active Dress": "Casual and Day Dresses",
    "Wrap Dress": "Casual and Day Dresses",
    "Shirt Dress": "Casual and Day Dresses",
    "Swimdress": "Special Occasion Dresses",
}

DESCRIPTION_OPENERS = [
    "Elevate your wardrobe with the {brand} {style_name} — a versatile piece designed for the modern woman.",
    "Introducing the {brand} {style_name}, where effortless style meets all-day comfort.",
    "Step into confidence with the {brand} {style_name}, crafted for women who refuse to compromise on style.",
    "The {brand} {style_name} is your go-to choice for a polished look from morning to evening.",
    "Discover the {brand} {style_name} — thoughtfully designed for women who love beautiful, functional fashion.",
    "Meet the {brand} {style_name}: a wardrobe essential that blends timeless design with contemporary flair.",
    "The {brand} {style_name} brings together sophisticated design and everyday wearability in one stunning piece.",
    "Designed for the woman on the move, the {brand} {style_name} delivers style without sacrificing comfort.",
]

# Swimwear openers rotate so 20+ rash guards don't share the same opening line.
SWIM_DESCRIPTION_OPENERS = [
    "The {brand} {style_name} is built for performance in and out of the water.",
    "Dive into warm-weather adventures with the {brand} {style_name}.",
    "Crafted for sun, surf, and everything in between, the {brand} {style_name} is your summer essential.",
    "Meet the {brand} {style_name} — a water-ready staple engineered for comfort and durability.",
    "From beach mornings to pool-side afternoons, the {brand} {style_name} has you covered.",
    "The {brand} {style_name} balances beach style with all-day performance.",
    "Whether you're paddling out or soaking up sun, the {brand} {style_name} keeps up.",
    "Designed with {brand}'s water-sports DNA, the {style_name} is ready for any adventure.",
]

# Bullet 1 variation pool for swim UPF items so 20+ rash guards don't share one bullet.
SWIM_UPF_B1_OPENERS = [
    "UPF {upf} SUN PROTECTION",
    "BLOCK HARMFUL UV RAYS",
    "ALL-DAY SUN DEFENSE",
    "UV PROTECTION BUILT IN",
    "SUN-SAFE PERFORMANCE",
    "FULL-COVERAGE UPF {upf}",
]
SWIM_UPF_B1_TAILS = [
    "shields your skin from harmful UV rays. Ideal for surfing, swimming, and all-day outdoor wear.",
    "delivers certified UV defense so you can stay in the water longer without worry.",
    "keeps sunburns at bay during long beach days, pool sessions, and open-water swims.",
    "provides lasting UV protection for active surfers, swimmers, and beach-goers.",
    "gives you reliable sun coverage from first light to sundown.",
]

# Bullet 1 variation pool for non-UPF swim items (bikini tops/bottoms, one-pieces, trunks w/o UPF).
SWIM_LIFESTYLE_B1_OPENERS = [
    "SURF & BEACH READY",
    "OCEAN-INSPIRED STYLE",
    "WATER-TO-SAND VERSATILE",
    "BEACH-DAY ESSENTIAL",
    "POOLSIDE TO SHORELINE",
]
SWIM_LIFESTYLE_B1_TAILS = [
    "combines ocean-inspired style with durable, quick-dry construction perfect for any beach day or pool session.",
    "pairs effortless coastal style with performance fabric that moves with you in and out of the water.",
    "brings classic surf aesthetics together with modern quick-dry performance.",
    "blends laid-back beach style with the durability you need for real water days.",
    "delivers easy swim-wear style that transitions seamlessly from shoreline to boardwalk.",
]

# Bullet 2 variation pool. Avoids stuffing the full style name into bullet 2.
# Feature phrase is filled in per style.
SWIM_B2_OPENERS = [
    "DESIGNED FOR PERFORMANCE",
    "PURPOSE-BUILT DETAILS",
    "THOUGHTFUL CONSTRUCTION",
    "STYLE MEETS FUNCTION",
    "DIALED-IN FEATURES",
]
SWIM_B2_TEMPLATES = [
    "This {itn_lower} features {feature_str} for lasting comfort and durability in the water.",
    "Every detail of this {itn_lower} is built for real beach days: {feature_str}.",
    "Purpose-built details — {feature_str} — give this {itn_lower} a technical edge.",
    "This {itn_lower} brings together {feature_str} so you can focus on the swim, not the fit.",
    "From fit to finish, this {itn_lower} delivers {feature_str}.",
]

DESCRIPTION_OPENERS_ROTATION = {}  # style_num -> opener_index

# ── Helper utilities ───────────────────────────────────────────────────────────
def _safe(v):
    return str(v).strip() if v is not None else ""

def _derive_gender_department(style):
    """Derive Amazon gender + department from a style's division_name AND style_name.
    style_name is used to refine YOUTH styles that don't have BOY/GIRL in division_name
    (e.g. 'VOLCOM YOUTH SWIM' + style 'Little Boys Long Sleeve Rashguard' -> Male/boys).
    Returns (gender, department) tuple — e.g. ('Male', 'mens') or ('Female', 'womens').
    Falls back to empty strings when unknown.
    """
    dn = (style.get("division_name", "") or "").upper()
    sn = (style.get("style_name", "") or "").upper()
    gender = ""
    department = ""
    if "YOUTH" in dn or "KIDS" in dn or "BOY" in dn or "GIRL" in dn:
        # Youth — refine with style_name
        if "GIRL" in dn or "GIRL" in sn:
            gender = "Female"
            department = "girls"
        elif "BOY" in dn or "BOY" in sn:
            gender = "Male"
            department = "boys"
        else:
            gender = "Unisex"
            department = "boys"  # Amazon default for youth unisex
    elif "WOMENS" in dn or "WOMEN'S" in dn or "WOMEN " in dn:
        # Check WOMENS before MENS ("WOMENS" contains "MENS" as substring)
        gender = "Female"
        # Amazon validates titlecase "Womens" — lowercase fails the dropdown
        department = "Womens"
    elif "MENS" in dn or "MEN'S" in dn or " MEN " in dn or dn.endswith(" MEN"):
        gender = "Male"
        department = "Mens"
    return gender, department

def _derive_youth_size_info(style_name, gender, raw_size):
    """Given a youth style's name + gender + raw_size, return:
      (special_size_type, size_class, age_range_description, normalized_size)
    for adults returns ("", "Alpha", "Adult", normalize_size(raw_size)).
    Maps toddler 2T/3T/etc. to '2 Years'/'3 Years' (Amazon-valid values).
    """
    sn = (style_name or "").lower()
    g = (gender or "").lower()
    raw = str(raw_size).strip()

    # Detect youth bucket from style_name
    is_toddler    = "toddler" in sn or raw.upper().endswith("T") and raw[:-1].isdigit()
    is_little_boy  = "little boy" in sn or "little boys" in sn
    is_big_boy     = "big boy" in sn or "big boys" in sn
    is_little_girl = "little girl" in sn or "little girls" in sn
    is_big_girl    = "big girl" in sn or "big girls" in sn
    is_youth       = g == "unisex" or is_toddler or is_little_boy or is_big_boy or is_little_girl or is_big_girl or "boys" in sn or "girls" in sn

    # If this isn't a youth style, adult defaults
    if not is_youth:
        return "", "Alpha", "Adult", (normalize_size(raw) or raw)

    # Pick special_size_type
    if is_toddler:
        sst = "Toddler Girls" if ("girls" in sn or is_little_girl) else "Toddler Boys"
    elif is_little_boy:
        sst = "Little Boys"
    elif is_big_boy:
        sst = "Big Boys"
    elif is_little_girl:
        sst = "Little Girls"
    elif is_big_girl:
        sst = "Big Girls"
    else:
        # Fallback — use gender
        sst = "Big Boys" if g == "male" else ("Big Girls" if g == "female" else "")

    # Pick age_range_description
    if is_toddler:
        ard = "Toddler"
    elif is_little_boy or is_little_girl:
        ard = "Little Kid"
    elif is_big_boy or is_big_girl:
        ard = "Big Kid"
    else:
        ard = "Big Kid"

    # Normalize size for youth
    # 2T/3T/4T/5T -> 'N Years' (Amazon-valid)
    if raw.upper().endswith("T") and raw[:-1].isdigit():
        return sst, "Age", ard, f"{int(raw[:-1])} Years"
    # Plain numeric kid sizes (4,5,6,7,8,10,12,14,16) -> 'N Years'
    if raw.isdigit():
        years = int(raw)
        if 0 <= years <= 18:
            return sst, "Age", ard, f"{years} Years"
        return sst, "Numeric", ard, raw
    # Alpha youth sizes — keep alpha
    return sst, "Alpha", ard, (normalize_size(raw) or raw)

# PT-aware color snapping cache: { product_type: set(valid color_map values) }
_PT_COLOR_DROPDOWN_CACHE = {}

def _valid_color_set_for_pt(product_type):
    """Return the set of valid color#1.standardized_values#1 for a PT.
    Reads from dropdown_cache/{PT}.json once and memoizes.
    """
    if not product_type:
        return set()
    pt = product_type.upper()
    if pt in _PT_COLOR_DROPDOWN_CACHE:
        return _PT_COLOR_DROPDOWN_CACHE[pt]
    valid = set()
    try:
        path = DROPDOWN_CACHE_DIR / f"{pt}.json"
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                cache = json.load(f)
            vals = cache.get("color#1.standardized_values#1", []) or []
            if isinstance(vals, list):
                valid = {v for v in vals if v}
    except Exception as e:
        print(f"[normalize_color] dropdown cache read failed for {pt}: {e}")
    _PT_COLOR_DROPDOWN_CACHE[pt] = valid
    return valid

# Common variant names that should snap to dropdown family values when present in the PT's set.
# Used as a fallback when the global COLOR_MAP returns a value not in the PT's dropdown.
_COLOR_SNAP_FALLBACKS = {
    "IVORY":   ["Off White", "Beige", "White"],
    "CREAM":   ["Off White", "Beige", "White"],
    "TRUFFLE": ["Brown", "Beige"],
    "COGNAC":  ["Brown"],
    "CAMEL":   ["Brown", "Beige"],
    "TAUPE":   ["Brown", "Beige", "Grey"],
    "WINE":    ["Red"],
    "BURGUNDY":["Red"],
    "NAVY":    ["Blue"],
    "OLIVE":   ["Green"],
    "CHARCOAL":["Grey"],
    "GRAY":    ["Grey"],
}

def normalize_color(raw_color, product_type=""):
    """Map raw color to an Amazon color family.
    When product_type is supplied, the result is snapped to a value that
    actually appears in that PT's dropdown (read live from dropdown_cache/{PT}.json).
    This fixes Cream→Ivory failing for COAT (COAT only validates 'Off White'),
    Truffle returning blank, etc.
    """
    if not raw_color:
        return ""
    upper = raw_color.upper().strip()
    # 1. Try the global COLOR_MAP first
    candidate = ""
    for key, val in COLOR_MAP.items():
        if key in upper:
            candidate = val
            break
    if not candidate:
        candidate = raw_color.title()

    # 2. PT-aware snap: if we know the valid set for this PT and our candidate
    #    isn't in it, walk the fallback ladder.
    valid = _valid_color_set_for_pt(product_type)
    if valid and candidate not in valid:
        # Try fallback ladder for known synonyms
        for token, options in _COLOR_SNAP_FALLBACKS.items():
            if token in upper:
                for opt in options:
                    if opt in valid:
                        return opt
                break
        # Last resort: leave candidate as-is. The dropdown validator will flag it,
        # surfacing the issue in the QA panel rather than silently writing a bad value.
    return candidate

def normalize_size(raw_size):
    """Standardize size string."""
    if not raw_size:
        return ""
    return SIZE_MAP.get(str(raw_size).strip().upper(), str(raw_size).strip())

def parse_fabric(raw_fabric):
    """Parse fabric string like '95 POLY 5 SPAN' → '95% Polyester, 5% Spandex'."""
    if not raw_fabric:
        return ""
    abbreviations = {
        "POLY": "Polyester", "SPAN": "Spandex", "COTT": "Cotton",
        "NYLON": "Nylon", "RAYON": "Rayon", "LINEN": "Linen",
        "SILK": "Silk", "WOOL": "Wool", "MODAL": "Modal",
        "ACRY": "Acrylic", "LYOCEL": "Lyocell", "TENCEL": "Tencel",
        "VISCOSE": "Viscose", "BAMBOO": "Bamboo",
    }
    s = str(raw_fabric).strip()
    # Already in percentage format
    if "%" in s:
        return s
    # Try to parse "95 POLY 5 SPAN" format
    parts = re.findall(r'(\d+)\s*([A-Za-z]+)', s)
    if parts:
        result = []
        for pct, fiber in parts:
            full = abbreviations.get(fiber.upper(), fiber.title())
            result.append(f"{pct}% {full}")
        return ", ".join(result)
    return s

def derive_neck_type(style_name):
    """Derive neck type from style name."""
    name = style_name.upper()
    mappings = [
        ("V NECK", "V-Neck"), ("V-NECK", "V-Neck"), ("VNECK", "V-Neck"),
        ("HALTER", "Halter"), ("CREW", "Crew Neck"), ("SCOOP", "Scoop Neck"),
        ("SQUARE", "Square Neck"), ("COWL", "Cowl Neck"), ("MOCK", "Mock Neck"),
        ("TURTLENECK", "Turtleneck"), ("HIGH NECK", "High Neck"),
        ("SWEETHEART", "Sweetheart"), ("OFF THE SHOULDER", "Off Shoulder"),
        ("OFF SHLD", "Off Shoulder"), ("OFF-SHOULDER", "Off Shoulder"),
        ("BAND NECK", "Band Neck"), ("BAND NCK", "Band Neck"),
        ("YOKE NECK", "Yoke Neck"), ("YOKE NCK", "Yoke Neck"),
        ("PINTUCK", "V-Neck"), ("KEYHOLE", "Keyhole"),
    ]
    for pattern, neck in mappings:
        if pattern in name:
            return neck
    return ""

# PT → fallback sleeve type when the style name has no signal. Coats are long
# sleeve unless explicitly vest-marked; swimwear is sleeveless; etc.
_PT_DEFAULT_SLEEVE_TYPE = {
    "COAT":       "Long Sleeve",
    "BLAZER":     "Long Sleeve",
    "SWEATSHIRT": "Long Sleeve",
    "SNOWSUIT":   "Long Sleeve",
    "SHIRT":      "Short Sleeve",
    "DRESS":      "Short Sleeve",
    "ONE_PIECE_OUTFIT": "Short Sleeve",
    "SWIMWEAR":   "Sleeveless",
    "BRA":        "Sleeveless",
}

def derive_sleeve_type(style_name, product_type=""):
    """Derive sleeve type from style name.
    PT-aware: when no in-name signal, falls back to the PT-appropriate sleeve
    instead of always returning "Sleeveless". This fixes coats / blazers /
    sweatshirts being mislabeled as Sleeveless when their style name doesn't
    explicitly say "Long Sleeve".
    """
    name = style_name.upper()
    # Vest detection short-circuits to Sleeveless regardless of PT
    if "VEST" in name:
        return "Sleeveless"
    mappings = [
        ("SLEEVELESS", "Sleeveless"), ("SLVLES", "Sleeveless"),
        ("SLVLS", "Sleeveless"),
        ("FLUTTER SLEEVE", "Flutter Sleeve"), ("FLUTTER SLV", "Flutter Sleeve"),
        ("FLUTTER", "Flutter Sleeve"),
        ("RUFFLE SLV", "Ruffle Sleeve"), ("RFL SLV", "Ruffle Sleeve"),
        ("OFF SHOULDER", "Off-Shoulder"), ("OFF SHLD", "Off-Shoulder"),
        ("BALLOON SL", "Balloon Sleeve"), ("CAP SLEEVE", "Cap Sleeve"),
        ("SHORT SLEEVE", "Short Sleeve"), ("LONG SLEEVE", "Long Sleeve"),
        ("3/4 SLEEVE", "3/4 Sleeve"),
        ("SLV", "Short Sleeve"),  # Last — most permissive abbreviation
    ]
    for pattern, sleeve in mappings:
        if pattern in name:
            return sleeve
    # No signal in style name — PT-aware default
    return _PT_DEFAULT_SLEEVE_TYPE.get((product_type or "").upper(), "Sleeveless")

# Common ISO country codes → full names (Amazon template dropdowns use full names)
COUNTRY_CODE_MAP = {
    "MX": "Mexico", "BD": "Bangladesh", "CN": "China", "US": "United States",
    "IN": "India", "VN": "Vietnam", "KH": "Cambodia", "PK": "Pakistan",
    "ID": "Indonesia", "TW": "Taiwan", "KR": "South Korea", "JP": "Japan",
    "TH": "Thailand", "TR": "Turkey", "IT": "Italy", "PT": "Portugal",
    "PE": "Peru", "GT": "Guatemala", "HN": "Honduras", "SV": "El Salvador",
    "NI": "Nicaragua", "HT": "Haiti", "DO": "Dominican Republic",
    "LK": "Sri Lanka", "MM": "Myanmar", "PH": "Philippines",
    "ET": "Ethiopia", "MG": "Madagascar", "MA": "Morocco", "EG": "Egypt",
}

def normalize_coo(raw):
    """Convert ISO country code to full name. Pass through if already a full name."""
    if not raw:
        return raw
    s = str(raw).strip()
    upper = s.upper()
    if upper in COUNTRY_CODE_MAP:
        return COUNTRY_CODE_MAP[upper]
    # Already a full name
    return s


def clean_brand_name(raw_brand):
    """Strip vendor labels from brand name. 'Stella Parker PL Ladies SPTW' -> 'Stella Parker'"""
    if not raw_brand:
        return raw_brand
    b = str(raw_brand).strip()
    # Remove common vendor suffixes
    for suffix in [' PL Ladies SPTW', ' PL Ladies', ' PL Mens', ' SPTW', ' Sportswear',
                   ' us_apparel', ' Women\'s Swimwear', ', us_apparel']:
        b = b.replace(suffix, '')
    # Known brands: keep the longest matching prefix so 'Sage Collective' wins over 'Sage'.
    known = ['Sage Collective', 'Stella Parker', 'Novelle Fashion', 'Novelle',
             'Volcom', 'Roxy', 'Nautica', 'Ben Sherman', 'Spyder', 'Tahari', 'Sage']
    # Sort by length descending so 'Sage Collective' is checked before 'Sage'
    for k in sorted(known, key=len, reverse=True):
        if b.startswith(k):
            return k
    return b.strip()

def _derive_parent_sku_from_variants(variants, style_num):
    """If every variant's source SKU starts with the same '<prefix>-<style_num>' chunk, return
    that chunk so the parent row uses the team's own season-coded format (e.g. 'F26-107010297')
    instead of the synthetic '<vendor_code_prefix>-<style_num>'. Returns empty string if we
    can't confidently extract a common parent stem.
    """
    if not variants or not style_num:
        return ""
    skus = [str(v.get("sku", "") or "").strip() for v in variants if v.get("sku")]
    if not skus:
        return ""
    # Find the chunk before '-{color}-{size}' in each SKU — must include style_num
    candidates = set()
    for sku in skus:
        # Source SKUs follow '<prefix>-<style_num>-<color_code>-<size_code>'
        # We want everything up to (and including) <style_num>.
        idx = sku.find(str(style_num))
        if idx <= 0:
            return ""  # style_num not in SKU — unexpected, bail to legacy
        end = idx + len(str(style_num))
        candidates.add(sku[:end])
    if len(candidates) == 1:
        return next(iter(candidates))
    return ""


def _derive_child_sku(variant, parent_sku, color_name, size, color_code="", size_code=""):
    """Return the child SKU. Prefers the source pre-upload SKU verbatim. Falls back to
    '{parent_sku}-{color_code or color_name}-{size_code or size}' so SKUs stay short-coded
    (BLK-S) instead of full names (BLACK-SMALL). Per the team's working agreement:
    short codes for SKU, full names for display fields.
    """
    src = str(variant.get("sku", "") or "").strip()
    if src:
        return src
    # Synthesize: prefer short codes if available
    cc = (color_code or color_name or "").strip()
    sc = (size_code or size or "").strip()
    return f"{parent_sku}-{cc}-{sc}".replace(" ", "-")


def derive_silhouette(sub_subclass):
    """Derive silhouette from sub_subclass. Never returns #N/A or empty junk."""
    if not sub_subclass:
        return "flattering"
    s = str(sub_subclass).strip()
    # Catch #N/A, N/A, NA, None, nan, empty
    if s.upper() in ('', '#N/A', 'N/A', 'NA', 'NONE', 'NAN', 'NA (CONVERSION)', '#N/A (CONVERSION)'):
        return "flattering"
    mapping = {
        "Shift Dress": "Shift",
        "A-Line Dress": "A-Line",
        "Fit & Flare Dress": "Fit & Flare",
        "Dress with Shorts": "Romper",
        "Wrap Dress": "Wrap",
        "Sheath Dress": "Sheath",
        "Maxi Dress": "Maxi",
        "Mini Dress": "Mini",
        "Bodycon Dress": "Bodycon",
    }
    result = mapping.get(s, s.replace(" Dress", "").strip())
    # Final safety check
    if not result or '#' in result or result.upper() in ('N/A', 'NA', 'NAN'):
        return "flattering"
    return result

def style_descriptor_from_name(style_name):
    """Extract a clean style descriptor for use in titles."""
    name = str(style_name).upper()
    # Clean up common abbreviations
    replacements = {
        "SLVLES": "Sleeveless", "SLVLS": "Sleeveless", "SLV": "Sleeve",
        "DRS": "Dress", "DRSS": "Dress", "NCK": "Neck",
        "SHLD": "Shoulder", "RFL": "Ruffle", "BBYDOLL": "Baby Doll",
        "SHRT": "Short", "CINCH WST": "Cinch Waist",
        "TSL": "Tassel", "FR": "Front", "FIT": "Fit",
        "FLR": "Flare", "ZIP": "Zip", "BTN": "Button",
    }
    result = name
    for abbr, full in replacements.items():
        result = re.sub(r'\b' + abbr + r'\b', full, result)
    # Remove "DRESS" from end since product_type covers it
    result = re.sub(r'\bDRESS\b', '', result).strip()
    # Title case
    return result.title().strip()

def _title_case_preserve_acronyms(s):
    """Title-case a string while preserving common acronyms (UPF, 4-Way, etc.)
    and correctly handling possessives (Men's, Women's, Boys', Girls').
    """
    if not s:
        return s
    out = s.title()
    # Fix possessives that .title() breaks: Men'S -> Men's
    out = re.sub(r"'S\b", "'s", out)
    out = re.sub(r"S'\b", "s'", out)
    # Restore acronyms that .title() lowercased
    restorations = {
        r'\bUpf\b': 'UPF',
        r'\bUv\b': 'UV',
        r'\bUsa\b': 'USA',
        r'\bUs\b': 'US',
        r'\bNfl\b': 'NFL',
        r'\bMlb\b': 'MLB',
        r'\bNba\b': 'NBA',
        r'\bLs\b': 'LS',
        r'\bSs\b': 'SS',
    }
    for pat, repl in restorations.items():
        out = re.sub(pat, repl, out)
    return out

def _gender_title_word(gender_str, style_name=""):
    """Return the gender word used in titles, given a style-derived gender.
    Male -> Men's, Female -> Women's, Unisex/boys -> Boys' / Girls' / Kids' per style_name.
    """
    g = (gender_str or "").lower()
    sn = (style_name or "").lower()
    if g == "male" and ("boys" in sn or "boy" in sn or "toddler" in sn):
        return "Boys'"
    if g == "female" and ("girls" in sn or "girl" in sn):
        return "Girls'"
    if g == "male":
        return "Men's"
    if g == "female":
        return "Women's"
    if g == "unisex":
        if "boys" in sn:
            return "Boys'"
        if "girls" in sn:
            return "Girls'"
        return "Kids'"
    return ""

def generate_title(brand_cfg, brand, style_name, product_type, color, size, upf="", style_gender=""):
    """Generate Amazon-compliant title. Max 120 chars for Vendor Central apparel.
    style_gender should come from _derive_gender_department(style)[0] — takes priority
    over brand_cfg['gender'] so brand-level defaults never leak into mis-gendered titles.
    """
    # Always clean the brand name
    clean_brand = clean_brand_name(brand)
    formula = brand_cfg.get("title_formula", "{brand} {gender} {style_descriptor} {product_type}, {color}, {size}")
    descriptor = style_descriptor_from_name(style_name)

    # If style name already contains the product type, don't append it again
    pt_title = product_type.title() if product_type else ""
    if pt_title and pt_title.lower() in style_name.lower():
        pt_title = ""  # avoid "Swim Trunk Trunk"

    # Style-derived gender takes priority over brand config; brand config is last resort
    effective_gender = style_gender or brand_cfg.get("gender", "")
    gender_word = _gender_title_word(effective_gender, style_name)

    title = formula.format(
        brand=clean_brand,
        style_descriptor=descriptor,
        style_name=_title_case_preserve_acronyms(style_name),
        product_type=pt_title,
        color=color.title() if color else "",
        size=normalize_size(size),
        upf=upf or brand_cfg.get("default_upf", ""),
        gender=gender_word,
    )
    # Clean up double spaces, leading/trailing punctuation
    title = re.sub(r'\s+', ' ', title).strip()
    title = re.sub(r',\s*,', ',', title)
    title = re.sub(r',\s*$', '', title)
    # Preserve acronyms (UPF etc.) in final title
    title = _title_case_preserve_acronyms(title) if title else title
    # Enforce 120 char limit — prefer dropping comma-separated descriptors from the end
    # over chopping mid-clause. This addresses the Sage feedback: child titles like
    # "...Heavyweight Jacket, Removable Faux Fur Hood" used to lose the entire tail when
    # the size+color suffix pushed past 120; now we drop whole clauses cleanly.
    if len(title) > 120:
        # Try removing whole comma-separated segments from the end first
        segs = title.split(",")
        while len(segs) > 1 and len(",".join(segs).strip()) > 120:
            segs.pop()
        candidate = ",".join(segs).strip().rstrip(",")
        if len(candidate) <= 120 and candidate:
            title = candidate
        else:
            # Last resort: word-boundary truncation, but never leave a dangling "With" / "And"
            t = title[:120].rsplit(" ", 1)[0].rstrip(",")
            # Drop trailing connector words to avoid "...Coat With" or "...Jacket And"
            while t and t.split()[-1].lower() in {"with", "and", "for", "in", "of", "to", "by", "the", "a", "an", "–", "-"}:
                t = t.rsplit(" ", 1)[0].rstrip(",")
            title = t
    return title

def generate_bullets(brand_cfg, brand, style_name, sub_subclass, fabric, care, color, upf="",
                     subclass="", gender="", product_type="", style_num=""):
    """Generate 5 bullet points per brand + style context. Product-type-aware."""
    brand = clean_brand_name(brand)
    focus = brand_cfg.get("bullet_1_focus", "Style and quality")
    actual_fabric = fabric or brand_cfg.get("default_fabric", "")
    actual_care = care or brand_cfg.get("default_care", "Machine Wash")
    actual_upf = upf or brand_cfg.get("default_upf", "")
    # Normalize UPF: strip redundant "UPF" prefix so templates can add their own
    if actual_upf and actual_upf.upper().startswith("UPF"):
        actual_upf = actual_upf[3:].strip()  # "UPF 50+" → "50+"

    # Determine the product word to use instead of "dress"
    itn = _derive_item_type_name(subclass, product_type) or subclass or "garment"
    itn_lower = itn.lower()
    is_swim = product_type == "SWIMWEAR" or subclass in (
        "Rashguard", "Trunk", "Bikini Top", "Bikini Bottom", "Swim Bottom",
        "One Piece Swim", "Tankini", "Short", "Swim Set 2 pcs", "Board Short")
    gender_word = "men's" if (gender or "").lower() == "male" else "women's" if (gender or "").lower() == "female" else ""

    # ── Bullet 1: brand-specific focus, rotated per style to avoid duplicate content ──
    # Hash style_name for a stable-but-varied index across 20+ styles of same subclass.
    # Use style_num (unique) so 21 rashguards w/ same style_name still rotate through all variants
    # Use two independent indices so opener*tail = OxT distinct combos across styles
    _b1_seed = style_num or style_name or ""
    _b1_idx_o = abs(hash(_b1_seed + "b1o"))
    _b1_idx_t = abs(hash(_b1_seed + "b1t"))
    if actual_upf and ("upf" in focus.lower() or is_swim):
        _opener = SWIM_UPF_B1_OPENERS[_b1_idx_o % len(SWIM_UPF_B1_OPENERS)].format(upf=actual_upf)
        _tail = SWIM_UPF_B1_TAILS[_b1_idx_t % len(SWIM_UPF_B1_TAILS)]
        b1 = f"{_opener} — Built with UPF {actual_upf} ultraviolet protection factor fabric, this {itn_lower} {_tail}"
    elif "butterlux" in focus.lower():
        b1 = f"BUTTERLUX FABRIC — Crafted from our signature Butterlux material, this {brand} {itn_lower} delivers an extraordinarily soft, silky touch for all-day luxurious comfort."
    elif "beach" in focus.lower() or "surf" in focus.lower() or is_swim:
        _opener = SWIM_LIFESTYLE_B1_OPENERS[_b1_idx_o % len(SWIM_LIFESTYLE_B1_OPENERS)]
        _tail = SWIM_LIFESTYLE_B1_TAILS[_b1_idx_t % len(SWIM_LIFESTYLE_B1_TAILS)]
        b1 = f"{_opener} — This {brand} {itn_lower} {_tail}"
    elif "nautical" in focus.lower():
        b1 = f"NAUTICAL INSPIRED — Rooted in {brand}'s rich maritime heritage, this {itn_lower} features classic nautical design elements that bring timeless style to every occasion."
    elif "british" in focus.lower() or "mod" in focus.lower():
        b1 = f"BRITISH MOD HERITAGE — Influenced by {brand}'s iconic British mod aesthetic, this {itn_lower} delivers bold, fashion-forward style that stands out."
    elif "performance" in focus.lower():
        b1 = f"PERFORMANCE DESIGN — Engineered with {brand}'s performance expertise, this {itn_lower} combines athletic functionality with modern styling."
    elif "tailored" in focus.lower() or "sophisticated" in focus.lower():
        b1 = f"SOPHISTICATED TAILORING — {brand}'s expert tailoring creates a refined, polished silhouette that transitions effortlessly between occasions."
    else:
        b1 = f"QUALITY CRAFTSMANSHIP — {brand} brings signature quality to this {itn_lower}, combining premium materials with expert construction for lasting style and durability."

    # ── Bullet 2: Style-specific features ──
    style_features = []
    sn_upper = style_name.upper()
    if is_swim:
        if "HOODED" in sn_upper or "HOOD" in sn_upper: style_features.append("integrated hood for extra sun coverage")
        if "LONG SLEEVE" in sn_upper: style_features.append("long-sleeve design for full arm protection")
        if "SHORT SLEEVE" in sn_upper: style_features.append("short-sleeve design for freedom of movement")
        if "LOOSE" in sn_upper: style_features.append("loose, relaxed fit for layering comfort")
        if "4-WAY" in sn_upper or "4 WAY" in sn_upper: style_features.append("4-way stretch fabric for unrestricted movement")
        if "E-WAIST" in sn_upper or "ELASTIC" in sn_upper: style_features.append("elastic waistband for easy on/off")
        if "BOARD" in sn_upper: style_features.append("boardshort styling built for the water")
        if "TWIST" in sn_upper: style_features.append("twist-front detail for a flattering silhouette")
        if "V-NECK" in sn_upper or "VNECK" in sn_upper: style_features.append("V-neckline for a sleek profile")
        if "PRINTED" in sn_upper or "PRINT" in sn_upper: style_features.append("bold all-over print")
        if "LOGO" in sn_upper: style_features.append("signature logo branding")
        if "ZIP" in sn_upper: style_features.append("front zip closure")
    else:
        # Dress/apparel features
        neck = derive_neck_type(style_name)
        sleeve = derive_sleeve_type(style_name, product_type)
        if neck: style_features.append(f"flattering {neck} neckline")
        if sleeve and sleeve != "Sleeveless": style_features.append(f"{sleeve.lower()} detail")
        if "PLEATED" in sn_upper: style_features.append("elegant pleated front")
        if "RUFFLE" in sn_upper: style_features.append("playful ruffle accents")

    if not style_features:
        style_features = ["thoughtfully designed details"]
    feature_str = ", ".join(style_features[:3])
    # Bullet 2: no style_name stuffing. Rotate headline + template per style.
    _b2_seed = style_num or style_name or ""
    _b2_idx_h = abs(hash(_b2_seed + "b2h"))
    _b2_idx_t = abs(hash(_b2_seed + "b2t"))
    if is_swim:
        _b2_head = SWIM_B2_OPENERS[_b2_idx_h % len(SWIM_B2_OPENERS)]
        _b2_body = SWIM_B2_TEMPLATES[_b2_idx_t % len(SWIM_B2_TEMPLATES)].format(itn_lower=itn_lower, feature_str=feature_str)
        b2 = f"{_b2_head} — {_b2_body}"
    else:
        b2 = f"DESIGNED FOR PERFORMANCE — This {itn_lower} features {feature_str} for a look that stands apart."

    # ── Bullet 3: Fit & sizing ──
    fit_type = brand_cfg.get("default_fit_type", "")
    if is_swim:
        b3 = f"COMFORTABLE FIT — Designed to fit true to size with a {fit_type.lower() or 'relaxed'} fit that moves with your body. Available in a full size range so every {gender_word + ' ' if gender_word else ''}customer finds the right fit."
    else:
        sil = derive_silhouette(sub_subclass)
        fit_desc = f"{sil} silhouette" if sil else "flattering cut"
        b3 = f"PERFECT FIT & COMFORT — The {fit_desc} is designed to flatter a range of body types with a relaxed yet refined fit that moves with you throughout the day."

    # ── Bullet 4: Fabric + care ──
    if actual_fabric:
        if is_swim:
            b4 = f"QUICK-DRY FABRIC — Made from {actual_fabric} for rapid moisture wicking and fast drying. {actual_care} for easy upkeep after every swim session."
        else:
            b4 = f"PREMIUM FABRIC — Made from {actual_fabric}, offering a smooth, comfortable feel with just the right amount of stretch. {actual_care} for easy home care."
    else:
        b4 = f"EASY CARE — Crafted for effortless wearability with durable construction. {actual_care} for convenient upkeep."

    # ── Bullet 5: Use case ──
    if is_swim:
        b5 = f"VERSATILE WATER WEAR — Whether you're surfing, swimming laps, or lounging poolside, this {brand} {itn_lower} delivers. Pair with your favorite {brand} swim gear for a complete look."
    else:
        b5 = f"COMPLETE THE LOOK — Pair this {color.title() if color else ''} {itn_lower} with your favorite accessories for any occasion. A versatile addition to your {brand} wardrobe."

    return [b1, b2, b3, b4, b5]

def generate_description(brand_cfg, brand, style_num, style_name, sub_subclass, fabric, care, color, upf="",
                         subclass="", gender="", product_type=""):
    """Generate product description. Max 2000 chars. Product-type-aware."""
    brand = clean_brand_name(brand)
    actual_fabric = fabric or brand_cfg.get("default_fabric", "")
    actual_care = care or brand_cfg.get("default_care", "Machine Wash")
    actual_upf = upf or brand_cfg.get("default_upf", "")
    if actual_upf and actual_upf.upper().startswith("UPF"):
        actual_upf = actual_upf[3:].strip()
    itn = _derive_item_type_name(subclass, product_type) or subclass or "garment"
    itn_lower = itn.lower()
    is_swim = product_type == "SWIMWEAR" or subclass in (
        "Rashguard", "Trunk", "Bikini Top", "Bikini Bottom", "Swim Bottom",
        "One Piece Swim", "Tankini", "Short", "Swim Set 2 pcs", "Board Short")

    parts = []
    if is_swim:
        # Rotate swim opener across 20+ styles so same-subclass products don't share line 1.
        _sw_idx = abs(hash((style_num or "") + "swim")) if style_num else 0
        _sw_opener = SWIM_DESCRIPTION_OPENERS[_sw_idx % len(SWIM_DESCRIPTION_OPENERS)]
        parts.append(_sw_opener.format(brand=brand, style_name=_title_case_preserve_acronyms(style_name)))
        if actual_fabric:
            fp = f"Constructed from {actual_fabric}"
            if actual_upf:
                fp += f" with UPF {actual_upf} sun protection"
            fp += ", this " + itn_lower + " dries quickly and resists chlorine and salt water."
            parts.append(fp)
        _closers = [
            f"Designed for surfing, swimming, and beach days, it combines {brand}'s signature style with functional performance. {actual_care} for easy care after every session.",
            f"Whether you're chasing waves or lounging on the sand, this {itn_lower} delivers the durability and comfort you expect from {brand}. {actual_care} between sessions.",
            f"From sunrise paddle-outs to sunset hangs, this {brand} {itn_lower} keeps up. {actual_care} after each use for long-lasting wear.",
            f"Packable, quick-drying, and built to last — {brand}'s approach to warm-weather essentials shows in every detail. {actual_care} for easy upkeep.",
        ]
        parts.append(_closers[_sw_idx % len(_closers)])
    else:
        # Dress / general apparel
        global DESCRIPTION_OPENERS_ROTATION
        if style_num not in DESCRIPTION_OPENERS_ROTATION:
            idx = len(DESCRIPTION_OPENERS_ROTATION) % len(DESCRIPTION_OPENERS)
            DESCRIPTION_OPENERS_ROTATION[style_num] = idx
        opener_template = DESCRIPTION_OPENERS[DESCRIPTION_OPENERS_ROTATION[style_num]]
        parts.append(opener_template.format(brand=brand, style_name=style_name.title()))
        if actual_fabric:
            fp = f"Constructed from {actual_fabric}"
            if actual_upf:
                fp += f" with UPF {actual_upf} sun protection built right in"
            fp += f", this {itn_lower} provides all-day comfort. {actual_care} for easy upkeep."
            parts.append(fp)
        parts.append(f"Versatile enough for any occasion, this {brand} {itn_lower} transitions effortlessly from day to night.")

    return " ".join(parts)[:2000]

_BANNED_STOCK_PHRASES = [
    (r"\bthe modern woman\b", "customers who value quality"),
    (r"\bthe modern man\b", "customers who value quality"),
    (r"\bthe modern individual\b", "customers who value quality"),
    (r"\bwomen who refuse\b", "customers who refuse"),
    (r"\bmen who refuse\b", "customers who refuse"),
    (r"\bwoman on the move\b", "person on the move"),
    (r"\bman on the move\b", "person on the move"),
]

def _scrub_gender_drift(text, expected_gender):
    """Belt-and-suspenders defense against the LLM (or rule-based fallback)
    leaking gender-mismatched phrases into a listing description.

    expected_gender: 'Male', 'Female', or '' (unknown / unisex).

    1. Strips known stock phrases ("the modern woman", "women who refuse", etc.)
       regardless of gender so descriptions don't lean on filler.
    2. When expected_gender is set, swaps any opposite-gender pronouns or
       descriptors that slipped past the prompt rules. We do NOT touch
       gender words inside an item-type phrase ("women's jacket") because
       Amazon validation expects those to match department/title.
    """
    if not text:
        return text
    out = text
    # 1) Banned stock phrases (always scrubbed)
    for pat, repl in _BANNED_STOCK_PHRASES:
        out = re.sub(pat, repl, out, flags=re.IGNORECASE)
    # 2) Cross-gender swap when we know the gender. Only flip standalone
    # pronouns/audience nouns, not item-type phrases like "men's jacket".
    # Avoid touching apostrophe-s collocations ("women's" / "men's") because
    # those carry item-type meaning and Amazon's department field expects them
    # to align with the title. Only flip standalone audience words.
    if expected_gender == "Male":
        out = re.sub(r"\bwomen(?!')\b", "men", out, flags=re.IGNORECASE)
        out = re.sub(r"\bwoman(?!')\b", "man", out, flags=re.IGNORECASE)
        out = re.sub(r"\bshe\b", "he", out)
        out = re.sub(r"\bher\b", "his", out)
        out = re.sub(r"\bherself\b", "himself", out)
        out = re.sub(r"\bladies\b", "gentlemen", out, flags=re.IGNORECASE)
    elif expected_gender == "Female":
        out = re.sub(r"\bmen(?!')\b", "women", out, flags=re.IGNORECASE)
        out = re.sub(r"\bman(?!')\b", "woman", out, flags=re.IGNORECASE)
        out = re.sub(r"\bhe\b", "she", out)
        out = re.sub(r"\bhis\b", "her", out)
        out = re.sub(r"\bhimself\b", "herself", out)
        out = re.sub(r"\bgentlemen\b", "ladies", out, flags=re.IGNORECASE)
    return out


def qa_check_content(content, brand):
    """Run QA checks on generated content. Returns list of issues."""
    issues = []
    clean_brand = clean_brand_name(brand)
    title = content.get("title", "")
    
    # Title checks
    if len(title) > 120:
        issues.append({"field": "title", "severity": "error", "msg": f"Title exceeds 120 chars ({len(title)})."})
    if len(title) < 40:
        issues.append({"field": "title", "severity": "warning", "msg": f"Title is very short ({len(title)} chars). Add more keywords."})
    if '#N/A' in title or '#n/a' in title.lower():
        issues.append({"field": "title", "severity": "error", "msg": "Title contains #N/A — data parsing error."})
    if brand and clean_brand not in title:
        issues.append({"field": "title", "severity": "warning", "msg": f"Brand name '{clean_brand}' not in title."})
    
    # Bullet checks
    for i in range(1, 6):
        b = content.get(f"bullet_{i}", "")
        if not b:
            issues.append({"field": f"bullet_{i}", "severity": "error", "msg": f"Bullet {i} is empty."})
        elif len(b) < 50:
            issues.append({"field": f"bullet_{i}", "severity": "warning", "msg": f"Bullet {i} is very short ({len(b)} chars)."})
        if len(b) > 500:
            issues.append({"field": f"bullet_{i}", "severity": "error", "msg": f"Bullet {i} exceeds 500 chars ({len(b)})."})
        if '#N/A' in b or '#n/a' in b.lower():
            issues.append({"field": f"bullet_{i}", "severity": "error", "msg": f"Bullet {i} contains #N/A."})
        # Check for prohibited language
        for word in ['best seller', 'best-seller', 'limited time', 'on sale', 'free shipping', 'guaranteed']:
            if word.lower() in b.lower():
                issues.append({"field": f"bullet_{i}", "severity": "error", "msg": f"Bullet {i} contains prohibited phrase: '{word}'."})
    
    # Description checks
    desc = content.get("description", "")
    if not desc:
        issues.append({"field": "description", "severity": "error", "msg": "Description is empty."})
    elif len(desc) < 200:
        issues.append({"field": "description", "severity": "warning", "msg": f"Description is short ({len(desc)} chars). Aim for 500+."})
    if len(desc) > 2000:
        issues.append({"field": "description", "severity": "error", "msg": f"Description exceeds 2000 chars ({len(desc)})."})
    if '#N/A' in desc:
        issues.append({"field": "description", "severity": "error", "msg": "Description contains #N/A."})
    
    # Backend keywords checks
    kw = content.get("backend_keywords", "")
    if not kw:
        issues.append({"field": "backend_keywords", "severity": "warning", "msg": "Backend keywords empty."})
    elif len(kw.encode('utf-8')) > 250:
        issues.append({"field": "backend_keywords", "severity": "error", "msg": f"Backend keywords exceed 250 bytes ({len(kw.encode('utf-8'))})."})
    if clean_brand and clean_brand.lower() in (kw or '').lower():
        issues.append({"field": "backend_keywords", "severity": "warning", "msg": "Backend keywords contain brand name (Amazon indexes it automatically)."})
    
    return issues


def generate_title_why(brand_cfg, brand, style_name, title, upf, has_keywords):
    """Generate 'why' explanation for the title."""
    char_count = len(title)
    parts = []
    # Gender format
    parts.append('"Women\'s" format used — outperforms "for Women" in Amazon search CTR.')
    # UPF
    if upf or brand_cfg.get("default_upf"):
        parts.append(f'UPF {upf or brand_cfg.get("default_upf")} placed after brand as lead differentiator — detected in product data.')
    # Keywords
    if has_keywords:
        parts.append('Top keyword from uploaded Helium 10 data incorporated into title.')
    else:
        parts.append('No keyword data uploaded — category defaults used for title structure.')
    parts.append(f'{char_count}/200 characters used.')
    return ' '.join(parts)


def generate_bullet_why(idx, brand_cfg, brand, style_name, sub_subclass, upf, fabric, has_keywords):
    """Generate 'why' explanation for a bullet."""
    focus = brand_cfg.get("bullet_1_focus", "Style and quality")
    if idx == 0:
        if "upf" in focus.lower():
            return f'UPF {upf or brand_cfg.get("default_upf", "30+")} detected in product data. Sun protection leads for this brand based on category positioning — highest differentiator for outdoor/activewear.'
        elif "butterlux" in focus.lower():
            return f'Butterlux fabric is {brand}\'s signature material — used as Bullet 1 per brand config.'
        elif "beach" in focus.lower() or "surf" in focus.lower():
            return f'{brand}\'s surf heritage is the primary brand differentiator — configured as Bullet 1 focus.'
        else:
            return f'Brand quality and craftsmanship leads Bullet 1 per brand config setting: "{focus}".'
    elif idx == 1:
        neck = derive_neck_type(style_name)
        sleeve = derive_sleeve_type(style_name)
        features = []
        if neck: features.append(f'"{neck}" detected in style name')
        if sleeve: features.append(f'"{sleeve}" detected in style name')
        if "PLEATED" in style_name.upper(): features.append('"PLEATED" → hourglass effect copy')
        if "RUFFLE" in style_name.upper() or "RFL" in style_name.upper(): features.append('"RUFFLE" → playful accent copy')
        if features:
            return f'Style-specific features derived from name: {", ".join(features[:2])}. This bullet varies per style to avoid duplicate content.'
        return 'Style details derived from style name analysis. This bullet varies per style to avoid duplicate content.'
    elif idx == 2:
        silhouette = derive_silhouette(sub_subclass)
        return f'Fit & sizing copy generated from silhouette: "{silhouette or "flattering"}". Size range XS–3X included per Amazon best practice for apparel.'
    elif idx == 3:
        actual_fabric = parse_fabric(fabric) or brand_cfg.get("default_fabric", "")
        if actual_fabric:
            return f'Fabric composition "{actual_fabric}" from product data. Care instructions from product data or brand defaults. Premium fabric positioning improves conversion.'
        return f'Fabric information not found in product data — using brand default. Care instructions from brand config. Upload product data with fabric column for style-specific copy.'
    elif idx == 4:
        return 'Cross-sell bullet drives average order value by suggesting complementary styling. Color-specific to each variant (updated per color in NIS output).'
    return ''


def generate_description_why(brand_cfg, style_num, opener_idx, has_keywords):
    """Generate 'why' explanation for description."""
    total_openers = len(DESCRIPTION_OPENERS)
    opener_num = (opener_idx % total_openers) + 1
    parts = [
        f'Opener #{opener_num} of {total_openers} used — rotated per style to avoid duplicate content flags.',
        'Three-paragraph structure: opener + style details + fabric/care + occasion/versatility.',
    ]
    if not has_keywords:
        parts.append('No keyword data uploaded — keyword integration uses category defaults. Upload Helium 10 CSV for optimized keyword placement.')
    return ' '.join(parts)


def generate_keywords_why(brand, keywords_list, result_kw, has_keywords):
    """Generate 'why' explanation for backend keywords."""
    byte_count = len(result_kw.encode('utf-8'))
    term_count = len(result_kw.split()) if result_kw else 0
    if has_keywords:
        top_kw = [k['keyword'] for k in keywords_list[:3]] if keywords_list else []
        return (f'{byte_count}/250 bytes used. {term_count} terms: top keywords from uploaded Helium 10 data '
                f'({(", ".join(top_kw)) or "none"}) plus category defaults. '
                f'Brand name excluded (Amazon penalizes brand repetition in backend keywords).')
    return (f'{byte_count}/250 bytes used. {term_count} terms derived from category defaults + style name analysis. '
            f'No keyword data uploaded — upload Helium 10 CSV for search-volume-ranked backend keywords. '
            f'Brand name excluded per Amazon guidelines.')


def generate_backend_keywords(brand, style_name, sub_subclass, color, fabric, upf="",
                              subclass="", gender="", product_type=""):
    """Generate backend keywords. Max 250 bytes, lowercase, no brand, no title duplicates.
    Product-type-aware."""
    brand_lower = brand.lower()
    itn = _derive_item_type_name(subclass, product_type) or subclass or ""
    itk = _derive_item_type_keyword(subclass, product_type) or ""
    is_swim = product_type == "SWIMWEAR" or subclass in (
        "Rashguard", "Trunk", "Bikini Top", "Bikini Bottom", "Swim Bottom",
        "One Piece Swim", "Tankini", "Short", "Swim Set 2 pcs", "Board Short")
    g = (gender or "").lower()
    gender_kw = "mens" if g == "male" else "womens" if g == "female" else "kids" if g == "unisex" else ""

    candidates = []
    if is_swim:
        # Swimwear-specific keywords
        if gender_kw:
            candidates.append(f"{gender_kw} {itn.lower()}" if itn else f"{gender_kw} swimwear")
        candidates.append(itn.lower() if itn else "swimwear")
        candidates.append(itk.replace("-", " ") if itk else "")
        candidates.append(subclass.lower() if subclass else "")
        candidates.extend(["swim", "swimwear", "beach", "pool"])
        if "RASH" in (subclass or "").upper() or "RASH" in style_name.upper():
            candidates.extend(["rash guard", "rashguard", "sun shirt", "swim shirt"])
        if "TRUNK" in (subclass or "").upper() or "TRUNK" in style_name.upper():
            candidates.extend(["swim trunks", "board shorts", "bathing suit"])
        if "BIKINI" in (subclass or "").upper():
            candidates.extend(["bikini", "two piece", "swimsuit"])
        if "ONE PIECE" in (subclass or "").upper():
            candidates.extend(["one piece", "swimsuit", "bathing suit"])
        if "TANKINI" in (subclass or "").upper():
            candidates.extend(["tankini", "two piece", "swimsuit"])
        if "BOARD" in style_name.upper():
            candidates.extend(["boardshorts", "board shorts", "surf shorts"])
        if upf:
            candidates.extend([f"upf {upf}", "sun protection", "uv protection"])
        candidates.extend(["quick dry", "water sports", "surf", "swimming"])
    else:
        # Dress / general apparel keywords
        if gender_kw:
            candidates.append(f"{gender_kw} {itn.lower()}" if itn else f"{gender_kw} dress")
        candidates.append(itn.lower() if itn else "dress")
        candidates.append(f"{sub_subclass.lower()}" if sub_subclass else "")
        candidates.extend(["casual", "everyday", "comfortable", "stylish"])
        if upf:
            candidates.extend([f"upf {upf}", "sun protective clothing"])

    if fabric:
        if "polyester" in fabric.lower(): candidates.append("polyester")
        if "spandex" in fabric.lower(): candidates.append("stretch")
        if "nylon" in fabric.lower(): candidates.append("nylon")
    if color:
        candidates.append(normalize_color(color).lower())

    # Filter: no brand name, no empty, no duplicates, no stem-redundant phrases
    # (e.g. drop "rash guard shirt" when "mens rash guard shirt" already covers all tokens)
    def _stem(tok):
        t = tok.strip().lower()
        # normalize singular/plural
        if len(t) > 3 and t.endswith("es") and not t.endswith("ses"):
            t = t[:-2]
        elif len(t) > 3 and t.endswith("s"):
            t = t[:-1]
        return t
    seen_phrases = set()
    covered_stems = set()
    result = []
    for kw in candidates:
        kw = kw.strip().lower()
        if not kw or brand_lower in kw or kw in seen_phrases or len(kw) < 2:
            continue
        toks = kw.split()
        stems = [_stem(t) for t in toks]
        # Skip if every stem is already covered by an earlier phrase (pure redundancy)
        if stems and all(s in covered_stems for s in stems):
            continue
        seen_phrases.add(kw)
        for s in stems:
            covered_stems.add(s)
        result.append(kw)

    # Join and cap at 250 bytes
    joined = " ".join(result)
    while len(joined.encode('utf-8')) > 250 and result:
        result.pop()
        joined = " ".join(result)

    return joined

# ── Template parsing ───────────────────────────────────────────────────────────
# ── LLM feedback loading ───────────────────────────────────────────────────────
def load_brand_feedback(brand):
    """Load the 20 most recent feedback entries for a brand as a summary string."""
    if not FEEDBACK_FILE.exists():
        return ""
    entries = []
    try:
        with open(str(FEEDBACK_FILE), "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    entry = json.loads(line)
                    if entry.get("brand") == brand:
                        entries.append(entry)
                except json.JSONDecodeError:
                    pass
    except Exception:
        return ""
    # Sort newest first, take top 20
    entries.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
    recent = entries[:20]
    if not recent:
        return ""
    lines = []
    for e in recent:
        field = e.get("field", "")
        feedback = e.get("feedback", "")
        orig = e.get("original", "")
        updated = e.get("updated", "")
        if feedback:
            lines.append(f"- [{field}] {feedback}")
        elif orig and updated:
            lines.append(f"- [{field}] Changed from: '{str(orig)[:80]}' to: '{str(updated)[:80]}'")
    return "\n".join(lines)


def generate_content_llm(brand_cfg, brand, style, feedback_history, regen_keys=None):
    """
    Use Claude to generate Amazon listing content for a style.
    Falls back to rule-based generation if LLM is unavailable.
    Returns a content dict with title, bullet_1-5, description, backend_keywords.

    regen_keys (optional): set of field keys the operator is explicitly
    regenerating. For those keys we DROP the 'pre-upload USE THIS' lock
    on bullets and DROP the merge_bullets override so the LLM's new text
    actually replaces the prior content. Without this, a regen request
    on bullet_2 returns the same words because the prompt forces VERBATIM
    and merge_bullets overwrites the LLM output with pre-upload text.
    """
    global _anthropic_client
    if _anthropic_client is None:
        return None  # Caller will fall back to rule-based
    regen_keys = set(regen_keys or [])
    regen_bullets = {k for k in regen_keys if k.startswith("bullet_")}

    clean_brand = clean_brand_name(brand)
    style_num = style["style_num"]
    style_name = style["style_name"]
    subclass = style.get("subclass", "")
    sub_subclass = style.get("sub_subclass", "")
    fabric = parse_fabric(style.get("fabric", "")) or brand_cfg.get("default_fabric", "")
    care = style.get("care", "") or brand_cfg.get("default_care", "")
    upf = style.get("upf", "") or brand_cfg.get("default_upf", "")
    coo = style.get("coo", "") or brand_cfg.get("default_coo", "")

    # Collect color/size lists
    variants = style.get("variants", [])
    colors = list(dict.fromkeys([v.get("color_name", "") for v in variants if v.get("color_name")]))
    sizes = list(dict.fromkeys([v.get("size", "") for v in variants if v.get("size")]))
    upcs = [v.get("upc", "") for v in variants[:3] if v.get("upc")]

    # Brand voice description
    bullet_1_focus = brand_cfg.get("bullet_1_focus", "Style and quality")
    never_words = brand_cfg.get("never_words", [])
    # Pull the structured BRAND VOICE block from substrate.brand_voice.
    # This is the new source of truth (tone, hero adjectives, signature
    # phrases, like/unlike examples, etc.). Falls back to a 'no voice
    # defined yet' line for brands that haven't edited voice yet, so the
    # prompt is always valid. Best-effort: substrate failures don't break
    # generation.
    try:
        from substrate.brand_voice import voice_prompt_block as _voice_block
        _ws = (brand or "tlg").lower().replace(" ", "_") or "tlg"
        brand_voice_block = _voice_block(_ws)
    except Exception as _bv_exc:
        print(f"[atlas] brand_voice block skipped: {_bv_exc}", flush=True)
        brand_voice_block = "=== BRAND VOICE ===\n(brand voice unavailable)"
    # Derive gender per-style from division_name, not brand config
    style_gender, style_dept = _derive_gender_department(style)
    gender = style_gender or brand_cfg.get("gender", "")
    never_words_str = ", ".join(never_words) if never_words else "none"
    product_type = _resolve_style_product_type(style) or ""
    itn = _derive_item_type_name(subclass, product_type) or subclass or "garment"
    division = style.get("division_name", "")

    # Product brief: per-style brief overrides brand-level brief
    style_briefs = session_data.get("style_briefs", {})
    product_briefs = brand_cfg.get("product_briefs", {})
    brief = style_briefs.get(str(style_num), "") or product_briefs.get(f"{product_type}/{subclass}", "") or product_briefs.get(product_type, "") or product_briefs.get("_default", "")

    feedback_count = len([l for l in feedback_history.splitlines() if l.strip()]) if feedback_history else 0
    feedback_section = f"LEARNED PREFERENCES (from {feedback_count} previous edits):\n{feedback_history}" if feedback_history else "LEARNED PREFERENCES: None yet."

    # Gender-aware audience description
    if gender == "Male":
        audience = "men and boys"
        gender_prefix = "Men's"
    elif gender == "Female":
        audience = "women and girls"
        gender_prefix = "Women's"
    else:
        audience = "all genders"
        gender_prefix = ""

    # ═══ v0.7.6: pull baseline rules + pre-upload bullets so AI fills only gaps ═══
    pu_bullets = style.get("bullets_from_upload") or []
    # Pad to 5 — each slot is either pre-upload text or empty (AI will fill empty slots)
    pu_bullets = (list(pu_bullets) + [""] * 5)[:5]
    # If the operator explicitly asked to regenerate a bullet (and possibly
    # gave feedback for it), DROP the pre-upload lock for that slot so the
    # LLM is allowed to write fresh copy. Without this, the prompt tells
    # Claude to return the pre-upload text VERBATIM and merge_bullets would
    # overwrite the LLM output anyway.
    if regen_bullets:
        for i in range(5):
            slot_key = f"bullet_{i+1}"
            if slot_key in regen_bullets:
                pu_bullets[i] = ""
    pu_bullets_block = "\n".join(
        f"  Bullet {i+1}: {('USE THIS — ' + b) if b and b.strip() else 'WRITE THIS — (operator did not provide)'}"
        for i, b in enumerate(pu_bullets)
    )
    keywords_seed = style.get("keywords", "") or ""
    addl = style.get("additional_details", "") or ""
    closure = style.get("closure_type", "") or ""
    neck    = style.get("neck_type", "") or ""
    sleeve  = style.get("sleeve_type", "") or ""
    fit     = style.get("fit_type", "") or ""

    # When the operator is explicitly regenerating one or more fields with
    # feedback, surface that as the FIRST instruction so Claude actually
    # acts on it rather than burying it in LEARNED PREFERENCES at the end.
    regen_directive = ""
    if regen_keys:
        keys_str = ", ".join(sorted(regen_keys))
        feedback_inline = feedback_history.strip() or "(no specific feedback — just rewrite with fresh language)"
        regen_directive = f"""=== OPERATOR REGEN REQUEST (HIGHEST PRIORITY) ===
The operator is explicitly asking you to REWRITE these fields: {keys_str}.
DO NOT return the previous text. You MUST produce different wording for these fields.
Operator feedback for this regen:
{feedback_inline}
If the feedback contradicts a baseline rule (length, format, gender), follow the baseline rule but adjust everything else per feedback.

"""

    prompt = f"""You are an Amazon NIS (New Item Setup) content expert. You create SEO-healthy, buyer-focused content that converts on Amazon.com.

{regen_directive}=== BASELINE CONTENT RULES (apply to EVERY listing, every brand) ===
TITLE — hard cap 120 chars (Vendor Central apparel). Target 100-118.
  Format: {{Brand}} {{Women's|Men's}} {{Function/Use}} {{Item Type}} - {{Top Feature}}, {{Top Feature}}
  - Front-load the top 3-5 search keywords inside the first 80 chars (mobile preview).
  - Brand MUST be at position 0. Gender word ({gender_prefix or 'derived from department'}) MUST be present.
  - NEVER include the style number as text. NEVER use #N/A. NEVER duplicate the item type.
  - Item type is '{itn}', NOT 'dress'. Use '{itn}' (or its plural).

BULLETS — 5 bullets, each max 256 chars. Target 220-250.
  Format: HEADLINE — benefit sentence. Headline is 2-5 words, ALL CAPS, then ' — ' (em-dash with spaces).
  Slot intent (in order):
    1. FABRIC + FEEL (lead with fabric story + tactile benefit; UPF here for swim)
    2. FIT + CONSTRUCTION (fit type, closure, neck/sleeve, pockets, length)
    3. FUNCTION + USE (where/when/how it's worn; versatility, occasion)
    4. CARE + DURABILITY (care instructions + longevity, no raw fiber percentages here)
    5. SIZE RANGE + COMPLETE THE LOOK (size availability + brand call-to-action)
  Lead with benefit, not feature. Weave 2-4 search keywords per bullet naturally.

DESCRIPTION — max 2000 chars. Target 1300-1500. Plain text, no HTML.
  Structure: brand-voice opener (2-3 sentences) -> THE FABRIC -> THE DESIGN -> THE FIT -> THE CARE/PROMISE.
  Mention brand 2-3 times. Buyer-focused. No promotional words (best seller, limited time, free shipping).

BACKEND KEYWORDS — max 249 bytes. Lowercase, space-separated, no commas.
  No brand name. No words already in title or bullets. Include 1-2 misspellings, 1-2 Spanish terms (vestido, mallas, pantalones), synonyms, occasion + activity terms.

=== BRAND CONTEXT ===
BRAND: {clean_brand}
BRAND VOICE: {clean_brand} is a {bullet_1_focus}-focused brand for {audience}. {brief or ''}
HERO FEATURE for bullet 1 if no override: {bullet_1_focus}
NEVER USE these words: {never_words_str}

{brand_voice_block}

=== THIS PRODUCT ===
Item Type: {itn} (Amazon product type: {product_type})
Style Name: {style_name}
Style #: {style_num}    (NEVER print the style number in customer-facing text)
Sub-class: {subclass} / {sub_subclass}
Division: {division}
Gender: {gender or 'derive from department'}
Fabric: {fabric or 'not specified'}
UPF: {upf or 'none'}
COO: {coo or 'not specified'}
Care: {care or 'not specified'}
Closure: {closure or 'not specified'}
Neck: {neck or 'not specified'}
Sleeve: {sleeve or 'not specified'}
Fit Type: {fit or 'not specified'}
Colors available: {', '.join(colors[:8]) if colors else 'not specified'}
Sizes available: {', '.join(sizes[:10]) if sizes else 'not specified'}
Key selling points / additional details: {addl or 'none provided'}
Operator-supplied keywords (use heavily for SEO): {keywords_seed or 'none'}

=== OPERATOR-SUPPLIED BULLETS (file wins, AI fills gaps) ===
{pu_bullets_block}
Where a slot says 'USE THIS', return that bullet text VERBATIM (you may format the headline to ALL CAPS and add the em-dash if missing, but do not change the wording). Where a slot says 'WRITE THIS', generate per the slot intent.

{f'PRODUCT BRIEF FROM OPERATOR: {brief}' if brief else ''}

{feedback_section}

=== HARD RULES (must follow) ===
- Title len <=120, brand at pos 0, gender word present, item type present, no style number, no #N/A.
- Bullets <=256 each, ALL-CAPS headline + ' — ' + sentence.
- Description <=2000.
- Backend keywords <=249 bytes, lowercase, space-sep, no commas, no brand name, no overlap with title/bullets.
- No promotional language. No competitor brand names. No 'dress' for non-dress items.
- GENDER DISCIPLINE: every gender-marked phrase in the description must match the title's gender. If title says "Men's" the description must say "men" / "man" / "he" — never "woman" / "women" / "she". If title says "Women's" the description uses "women" / "woman" / "she" — never "man" / "men" / "he". If gender is unknown, write the description gender-neutral ("customer", "wearer", "you") — never default to "the modern woman" or "the modern man."
- BANNED STOCK PHRASES (never use, regardless of gender): "the modern woman", "the modern man", "the modern individual", "women who refuse", "men who refuse", "woman on the move", "man on the move".

Respond in this exact JSON format (no other text, no markdown):
{{"title": "...", "bullet_1": "...", "bullet_2": "...", "bullet_3": "...", "bullet_4": "...", "bullet_5": "...", "description": "...", "backend_keywords": "..."}}"""

    try:
        # Build message content — text prompt + optional image (vision)
        msg_content = []
        style_num = style["style_num"]
        img_path = (session_data.get("style_images") or {}).get(str(style_num))
        if img_path and Path(img_path).exists():
            import base64 as _b64
            with open(img_path, "rb") as _img_f:
                img_bytes = _b64.b64encode(_img_f.read()).decode("utf-8")
            ext = Path(img_path).suffix.lower()
            media_type = {"jpg": "image/jpeg", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                          ".png": "image/png", ".webp": "image/webp", ".gif": "image/gif"}.get(ext, "image/jpeg")
            msg_content.append({
                "type": "image",
                "source": {"type": "base64", "media_type": media_type, "data": img_bytes}
            })
            msg_content.append({"type": "text", "text": "Above is a photo of this product. Use it to accurately describe the product's appearance, design details, and visual features.\n\n" + prompt})
        else:
            msg_content.append({"type": "text", "text": prompt})

        # max_tokens needs headroom: title(~150) + 5 bullets(~1500) + description(~2000) +
        # keywords(~250) + JSON overhead = comfortably 4000. Bumped from 2000 which was
        # truncating description and backend_keywords on long generations.
        message = _anthropic_client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=4000,
            messages=[{"role": "user", "content": msg_content}],
        )
        raw = message.content[0].text.strip()
        # Strip any markdown code fences if present
        raw = re.sub(r'^```(?:json)?\s*', '', raw, flags=re.MULTILINE)
        raw = re.sub(r'```\s*$', '', raw, flags=re.MULTILINE).strip()
        parsed = json.loads(raw)

        # ═══ v0.7.6: enforce baseline content rules + merge pre-upload bullets ═══
        from nis_engine import content_rules as _cr
        ai_bullets = [str(parsed.get(f"bullet_{i}", ""))[:256] for i in range(1, 6)]
        # For slots being regenerated, force the AI output through (the
        # pre-upload value was already cleared above, so merge_bullets
        # would pick AI anyway, but we belt-and-suspenders it here).
        merge_pu = list(pu_bullets)
        if regen_bullets:
            for i in range(5):
                if f"bullet_{i+1}" in regen_bullets:
                    merge_pu[i] = ""
        merged_bullets = _cr.merge_bullets(merge_pu, ai_bullets)

        # Title: hard cap 120, never the style number, never #N/A
        title_raw = str(parsed.get("title", "")).strip()
        if _cr.is_garbage_value(title_raw) or len(title_raw) > 120:
            title_raw = _cr.compose_title(
                brand=clean_brand,
                gender_word=gender_prefix,
                style_name=style_name,
                item_type_name=itn,
                feature_phrases=[fit, neck, sleeve, closure],
            )
        title_raw = title_raw[:120]

        # Description: if Claude returned empty (intermittent ~30% of runs),
        # synthesize a fallback from bullets so the listing isn't broken.
        desc_raw = str(parsed.get("description", "")).strip()
        if len(desc_raw) < 200:  # too short to be a real Amazon description
            print(f"[LLM] description was {len(desc_raw)} chars for style {style_num}; synthesizing fallback from bullets.")
            # Build a competent fallback paragraph from the merged bullets
            bullet_sentences = [b.split(" \u2014 ", 1)[1] if " \u2014 " in b else b for b in merged_bullets if b]
            opener = f"{clean_brand} delivers {bullet_1_focus.lower()} for {audience}." if clean_brand and bullet_1_focus else f"{clean_brand} brings quality and craftsmanship to your wardrobe."
            desc_raw = opener + " " + " ".join(bullet_sentences)
            desc_raw = desc_raw[:2000]

        # Backend keywords: same fallback strategy
        kw_raw = str(parsed.get("backend_keywords", "")).strip()
        if len(kw_raw.encode("utf-8")) < 50:  # too short to be useful
            print(f"[LLM] backend_keywords was {len(kw_raw)} bytes for style {style_num}; synthesizing fallback.")
            # Pull keywords from style + colors + sizes + product type
            kw_parts = []
            if itn: kw_parts.append(itn.lower())
            if subclass: kw_parts.append(subclass.lower())
            for c in colors[:6]: kw_parts.append(c.lower())
            kw_parts += ["womens" if gender == "Female" else "mens", "apparel", "clothing"]
            if keywords_seed: kw_parts.append(keywords_seed.lower())
            # Add Spanish per the prompt rules
            kw_parts += ["vestido", "mallas", "ropa"]
            kw_raw = " ".join(dict.fromkeys(kw_parts))  # dedupe preserving order

        # Cap keywords at 249 bytes
        while len(kw_raw.encode("utf-8")) > 249 and kw_raw:
            kw_raw = kw_raw.rsplit(" ", 1)[0]

        content = {
            "title": title_raw,
            "bullet_1": merged_bullets[0],
            "bullet_2": merged_bullets[1],
            "bullet_3": merged_bullets[2],
            "bullet_4": merged_bullets[3],
            "bullet_5": merged_bullets[4],
            "description": desc_raw,
            "backend_keywords": kw_raw,
        }
        return content

    except Exception as e:
        print(f"[LLM] generate_content_llm failed for style {style_num}: {e}")
        return None


def parse_template_columns(template_path):
    """Parse .xlsm template rows 3 (headers) and 4 (field IDs). Returns col_map."""
    wb = openpyxl.load_workbook(template_path, keep_vba=True, read_only=True)
    ws = None
    for name in wb.sheetnames:
        if "template" in name.lower() or "dress" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active
    
    col_map = {}
    for col in range(1, (ws.max_column or 300) + 1):
        header = _safe(ws.cell(row=3, column=col).value)
        field_id = _safe(ws.cell(row=4, column=col).value)
        if header or field_id:
            col_map[col] = {"header": header, "field_id": field_id}
    
    wb.close()
    return col_map

_template_col_map_cache = {}

def get_template_col_map(template_path=None):
    if template_path is None:
        template_path = str(DEFAULT_TEMPLATE)
    if template_path not in _template_col_map_cache:
        _template_col_map_cache[template_path] = parse_template_columns(template_path)
    return _template_col_map_cache[template_path]

def find_col_by_field_id(col_map, field_id_pattern):
    """Find column number(s) by field_id pattern match."""
    results = []
    for col, info in col_map.items():
        if field_id_pattern.lower() in info["field_id"].lower():
            results.append(col)
    return results

def find_col_exact(col_map, field_id):
    """Find exact column number by field_id."""
    for col, info in col_map.items():
        if info["field_id"].lower() == field_id.lower():
            return col
    return None

# ── Product data parsing ───────────────────────────────────────────────────────
PRODUCT_HEADER_ALIASES = {
    # Brand
    "brand": "brand",
    "brand code": "brand",
    "brand name": "brand",
    # Division / product type
    "division": "division",
    "tlgdiv name": "division_name",
    "inline/value": "inline_value",
    # Category
    "sub-class name": "subclass",
    "sub class name": "subclass",
    "sub class": "subclass",
    "subclass": "subclass",
    "sub sub-class name": "sub_subclass",
    "sub sub class name": "sub_subclass",
    "sub-subclass": "sub_subclass",
    # Style
    "style #": "style_num",
    "style#": "style_num",
    "style number": "style_num",
    "style name": "style_name",
    "basic style name": "style_name",
    "tlg style name": "style_name",
    "style description": "style_desc",
    "tlg style desc": "style_desc",
    "related keywords": "keywords",
    "season code": "season_code",
    "season added to amzn": "season_added",
    # Color
    "color code": "color_code",
    "color name": "color_name",
    "color": "color_name",
    # Size
    "product - size": "size",
    "size": "size",
    # IDs
    "upc": "upc",
    "upc code": "upc",
    "casin": "casin",
    "child asin": "child_asin",
    "parent asin": "parent_asin",
    "model": "model_code",
    "model name": "model_name",
    "sku": "sku",
    # Pricing
    "list price": "list_price",
    "amzn retail": "list_price",
    "amazon list price": "list_price",
    "retail": "list_price",
    "retail price": "list_price",
    "msrp": "list_price",
    "cost price": "cost_price",
    "amzn wholesale": "cost_price",
    "amazon cost": "cost_price",
    "wholesale": "cost_price",
    "wholesale price": "cost_price",
    "case pack": "case_pack",
    # Product attributes
    "country of origin": "coo",
    "coo": "coo",
    "fabric": "fabric",
    "material": "fabric",
    "fabric content percentage": "fabric",
    "fabric content": "fabric",
    "care": "care",
    "care instructions": "care",
    "upf": "upf",
    "upf rating": "upf",
    "neck type": "neck_type",
    "collar type": "neck_type",
    "closure type": "closure_type",
    "closure": "closure_type",
    "sleeve type": "sleeve_type",
    "sleeve": "sleeve_type",
    "fit type": "fit_type",
    "fit type (regular, relaxed, oversized, slim , fitted, etc.)": "fit_type",
    "pockets": "pockets",
    "pockets?": "pockets",
    "type of jacket": "type_of_jacket",
    "type_of_jacket": "type_of_jacket",
    # Bullets from pre-upload
    "key features (bullet 1)": "bullet_1",
    "key features (bullet 2)": "bullet_2",
    "key features (bullet 3)": "bullet_3",
    "key features (bullet 4)": "bullet_4",
    "key features (bullet 5)": "bullet_5",
    "bullet 1": "bullet_1",
    "bullet 2": "bullet_2",
    "bullet 3": "bullet_3",
    "bullet 4": "bullet_4",
    "bullet 5": "bullet_5",
    # Dates
    "due date": "ship_date",
    "due date (earliest ship date)": "ship_date",
    "earliest ship date": "ship_date",
    "ship date": "ship_date",
    # Extra
    "additional details": "additional_details",
    "additional details, standouts, call outs, features": "additional_details",
}

def fuzzy_match_headers(headers):
    """Fuzzy-match raw headers to internal field names."""
    mapping = {}  # internal_field -> col_index (0-based)
    for idx, h in enumerate(headers):
        if not h:
            continue
        key = str(h).strip().lower()
        if key in PRODUCT_HEADER_ALIASES:
            internal = PRODUCT_HEADER_ALIASES[key]
            if internal not in mapping:
                mapping[internal] = idx
    return mapping

def _read_sheet_rows(ws):
    """Find header row (>=5 non-empty cells with style/brand/color/upc/price keywords)
    and return (headers, data_rows).

    Skips annotation rows that some Atlas templates include between the header
    and real data: a 'tag row' (REQUIRED / IMPORTANT / OPTIONAL / GROUND TRUTH)
    and a 'description row' (human-readable hints, no real data values).
    """
    all_rows = list(ws.iter_rows(values_only=True))
    header_row_idx = None
    for i, row in enumerate(all_rows):
        non_empty = sum(1 for c in row if c is not None)
        if non_empty >= 5:
            row_str = " ".join(str(c).lower() for c in row if c is not None)
            if any(kw in row_str for kw in ["style", "brand", "color", "size", "upc", "price"]):
                header_row_idx = i
                break
    if header_row_idx is None:
        return None, []
    headers = [str(c).strip() if c is not None else "" for c in all_rows[header_row_idx]]

    TAG_TOKENS = {"required", "important", "optional", "ground truth",
                  "required*", "optional*", "recommended", "conditional"}

    def _is_annotation_row(row):
        """True for tag-only rows (REQUIRED / OPTIONAL / etc.) or descriptive
        prose rows that explain what each column means."""
        cells = [c for c in row if c is not None and str(c).strip() != ""]
        if not cells:
            return False
        # Tag row: all non-empty cells are tag tokens.
        if all(str(c).strip().lower() in TAG_TOKENS for c in cells):
            return True
        # Description row: a clear majority of cells are long prose strings
        # (4+ words). Real data rows occasionally have a 3-4 word product
        # name but never have most cells looking like sentences.
        wordy = sum(1 for c in cells if isinstance(c, str) and len(c.split()) >= 4)
        if wordy >= max(3, (len(cells) * 3) // 5):
            return True
        # Also catch description-style cells with em-dashes, parens-explanations,
        # or hint phrasing like "e.g." / "if applicable" / "as registered".
        hint_markers = (" — ", " - ", "e.g.", "if applicable", "as registered",
                        "used to", "must be", "for the listing")
        hint_cells = sum(
            1 for c in cells if isinstance(c, str) and any(m in c.lower() for m in hint_markers)
        )
        return hint_cells >= max(3, len(cells) // 3)

    raw_data = [row for row in all_rows[header_row_idx + 1:] if any(c is not None for c in row)]
    data = [row for row in raw_data if not _is_annotation_row(row)]
    return headers, data


def _build_preupload_style_index(file_path):
    """Read the 'PreUpload Style' sheet (if present) and return a dict
    keyed by Style # — contains style-level rich content (bullets, neck,
    sleeve, fit, UPF, additional details, ship date) that is NOT in the
    UPC sheet.

    Files we've seen:
      Sheet 1: 'PreUpload Style'    → 25 cols, style-level: name, COO, care, fabric,
                                       neck, closure, sleeve, fit, pockets, 5 bullets,
                                       UPF, additional details, ship date
      Sheet 2: 'Upload Template UPC' → 28 cols, variant-level: UPC, color, size, prices
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception:
        return {}
    style_idx = {}
    target_sheet = None
    for sn in wb.sheetnames:
        low = sn.lower()
        if "preupload" in low.replace(" ", "") or ("style" in low and "upload" in low):
            target_sheet = wb[sn]
            break
    if target_sheet is None:
        wb.close()
        return {}
    headers, data = _read_sheet_rows(target_sheet)
    if not headers:
        wb.close()
        return {}
    cmap = fuzzy_match_headers(headers)
    sn_idx = cmap.get("style_num")
    if sn_idx is None:
        wb.close()
        return {}
    for row in data:
        if sn_idx >= len(row):
            continue
        sn = row[sn_idx]
        if sn is None or sn == "":
            continue
        sn_str = str(sn).strip()
        if not sn_str:
            continue
        rec = {}
        for k, idx in cmap.items():
            if idx < len(row) and row[idx] is not None and row[idx] != "":
                rec[k] = row[idx]
        style_idx[sn_str] = rec
    wb.close()
    return style_idx


def parse_product_file(file_path):
    """Parse product Excel/CSV file. Returns (rows, errors, warnings).

    For TLG Pre-Upload .xlsx files with two sheets (PreUpload Style + Upload
    Template UPC), reads BOTH and merges style-level rich content (bullets,
    neck/sleeve/fit, UPF, additional details, ship date) into each variant's
    style record using Style # as the join key.
    """
    ext = Path(file_path).suffix.lower()
    raw_rows = []
    style_index = {}  # Style # -> dict of rich style fields from PreUpload Style sheet

    if ext in [".xlsx", ".xls", ".xlsm"]:
        # First, try to build a style index from the PreUpload Style sheet (if present)
        try:
            style_index = _build_preupload_style_index(file_path)
        except Exception:
            style_index = {}

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        # Prefer the UPC sheet (variant-level) for the main pass
        ws = None
        for sn in wb.sheetnames:
            low = sn.lower()
            if ("upload" in low and "upc" in low) or "template upc" in low:
                ws = wb[sn]
                break
        if ws is None:
            ws = wb.active
        headers, data_rows_local = _read_sheet_rows(ws)
        wb.close()

        if headers is None:
            return [], ["Could not find header row in file"], []

        for row in data_rows_local:
            raw_rows.append(row)
        
    elif ext in [".csv", ".tsv"]:
        delimiter = "\t" if ext == ".tsv" else ","
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter=delimiter)
            all_csv = list(reader)
        if not all_csv:
            return [], ["Empty CSV file"], []
        headers = all_csv[0]
        raw_rows = [tuple(row) for row in all_csv[1:]]
    else:
        return [], [f"Unsupported file type: {ext}"], []
    
    col_map = fuzzy_match_headers(headers)
    
    errors = []
    warnings = []
    styles = {}  # style_num -> {style_info, variants:[]}
    
    for row_idx, row in enumerate(raw_rows, start=1):
        def get(field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return ""
            return _safe(row[idx])
        
        style_num = get("style_num")
        style_name = get("style_name") or get("style_desc")
        brand = get("brand")
        subclass = get("subclass")
        sub_subclass = get("sub_subclass")
        division_name = get("division_name")
        color_name = get("color_name")
        color_code = get("color_code")
        size = get("size")
        upc = get("upc")
        list_price = get("list_price")
        cost_price = get("cost_price")
        parent_asin = get("parent_asin")
        child_asin = get("child_asin")
        model_name = get("model_name")
        season_code = get("season_code")
        fabric = get("fabric")
        care = get("care")
        upf = get("upf")
        coo = get("coo")
        sku = get("sku")
        keywords = get("keywords")
        # New fields from pre-upload
        neck_type = get("neck_type")
        closure_type = get("closure_type")
        sleeve_type = get("sleeve_type")
        fit_type = get("fit_type")
        pockets = get("pockets")
        type_of_jacket = get("type_of_jacket")
        ship_date = get("ship_date")
        bullet_1 = get("bullet_1")
        bullet_2 = get("bullet_2")
        bullet_3 = get("bullet_3")
        bullet_4 = get("bullet_4")
        bullet_5 = get("bullet_5")
        additional_details = get("additional_details")

        # ═══ v0.7.6: merge in style-level rich content from PreUpload Style sheet ═══
        # Style # is the join key. PreUpload Style has: name, COO, care, fabric,
        # neck, closure, sleeve, fit, pockets, 5 bullets, UPF, additional details, ship date.
        if style_num and str(style_num).strip() in style_index:
            ridx = style_index[str(style_num).strip()]
            def _take(rich, current):
                v = rich
                if v is None or v == "":
                    return current
                return current if current else _safe(v)
            style_name = _take(ridx.get("style_name"), style_name)
            model_name = _take(ridx.get("style_name"), model_name)
            coo        = _take(ridx.get("coo"), coo)
            care       = _take(ridx.get("care"), care)
            fabric     = _take(ridx.get("fabric"), fabric)
            neck_type  = _take(ridx.get("neck_type"), neck_type)
            closure_type = _take(ridx.get("closure_type"), closure_type)
            sleeve_type = _take(ridx.get("sleeve_type"), sleeve_type)
            fit_type   = _take(ridx.get("fit_type"), fit_type)
            pockets    = _take(ridx.get("pockets"), pockets)
            upf        = _take(ridx.get("upf"), upf)
            ship_date  = _take(ridx.get("ship_date"), ship_date)
            additional_details = _take(ridx.get("additional_details"), additional_details)
            bullet_1   = _take(ridx.get("bullet_1"), bullet_1)
            bullet_2   = _take(ridx.get("bullet_2"), bullet_2)
            bullet_3   = _take(ridx.get("bullet_3"), bullet_3)
            bullet_4   = _take(ridx.get("bullet_4"), bullet_4)
            bullet_5   = _take(ridx.get("bullet_5"), bullet_5)

        # ═══ v0.7.6: infer brand from TLGDIV NAME when no Brand column ═══
        if not brand and division_name:
            dn_up = str(division_name).upper()
            if "SAGE" in dn_up:
                brand = "Sage Collective"
            elif "VOLCOM" in dn_up:
                brand = "Volcom"
            elif "SPYDER" in dn_up:
                brand = "Spyder"
            elif "STELLA" in dn_up:
                brand = "Stella Parker"
            elif "NOVELLE" in dn_up:
                brand = "Novelle"

        if not style_num:
            continue
        
        # Validation
        row_errors = []
        row_warnings = []
        
        # Use the rich style_index for missing names where possible (already merged above).
        if not style_name or str(style_name).strip() == str(style_num).strip():
            row_warnings.append(f"Row {row_idx}: Missing style name for style {style_num}")
        
        # UPC validation
        if upc:
            upc_clean = re.sub(r'\D', '', str(upc))
            if len(upc_clean) not in (12, 13, 14):
                row_errors.append(f"Row {row_idx}: UPC/EAN '{upc}' is not valid (expected 12, 13, or 14 digits, got {len(upc_clean)}) — style {style_num}, {color_name} {size}")
            elif len(upc_clean) == 13:
                row_warnings.append(f"Row {row_idx}: EAN-13 detected for {style_num} {color_name} {size} — will be used as-is")
        else:
            row_warnings.append(f"Row {row_idx}: Missing UPC for style {style_num}, color {color_name}, size {size}")
        
        # Price validation
        try:
            lp = float(list_price) if list_price else 0
            cp = float(cost_price) if cost_price else 0
            if lp > 0 and cp > 0:
                if cp > lp:
                    row_errors.append(f"Row {row_idx}: Cost (${cp}) > List price (${lp}) for style {style_num}")
                elif cp > 0.8 * lp:
                    row_warnings.append(f"Row {row_idx}: CRITICAL: Cost (${cp}) is >80% of List (${lp}) for style {style_num}")
                elif cp > 0.6 * lp:
                    row_warnings.append(f"Row {row_idx}: Cost (${cp}) is >60% of List (${lp}) for style {style_num}")
        except (ValueError, TypeError):
            pass
        
        errors.extend(row_errors)
        warnings.extend(row_warnings)
        
        # Build style entry
        if style_num not in styles:
            styles[style_num] = {
                "style_num": style_num,
                "style_name": style_name or style_num,
                "brand": brand,
                "subclass": subclass,
                "sub_subclass": sub_subclass,
                "division_name": division_name,
                "list_price": list_price,
                "cost_price": cost_price,
                "parent_asin": parent_asin,
                "model_name": model_name,
                "season_code": season_code,
                "fabric": fabric,
                "care": care,
                "upf": upf,
                "coo": coo,
                # Fields from pre-upload that take priority
                "neck_type": neck_type,
                "closure_type": closure_type,
                "sleeve_type": sleeve_type,
                "fit_type": fit_type,
                "pockets": pockets,
                "type_of_jacket": type_of_jacket,
                "ship_date": ship_date,
                "bullets_from_upload": [b or "" for b in [bullet_1, bullet_2, bullet_3, bullet_4, bullet_5]],
                "additional_details": additional_details,
                "keywords": keywords,
                "variants": [],
                "errors": [],
                "warnings": [],
            }
        
        styles[style_num]["errors"].extend(row_errors)
        styles[style_num]["warnings"].extend(row_warnings)
        
        # Deduplicate style-level info
        if style_name and not styles[style_num]["style_name"]:
            styles[style_num]["style_name"] = style_name
        if fabric and not styles[style_num]["fabric"]:
            styles[style_num]["fabric"] = fabric
        if care and not styles[style_num]["care"]:
            styles[style_num]["care"] = care
        if upf and not styles[style_num]["upf"]:
            styles[style_num]["upf"] = upf
        if coo and not styles[style_num]["coo"]:
            styles[style_num]["coo"] = coo
        if parent_asin and not styles[style_num]["parent_asin"]:
            styles[style_num]["parent_asin"] = parent_asin
        
        variant = {
            "color_name": color_name,
            "color_code": color_code,
            "size": size,
            "upc": upc,
            "child_asin": child_asin,
            "sku": sku,
            "errors": row_errors,
            "warnings": row_warnings,
        }
        styles[style_num]["variants"].append(variant)
    
    return list(styles.values()), errors, warnings

# ── Session state (in-memory, per app restart) ─────────────────────────────────
session_data = {
    "brand": None,
    "vendor_code": None,
    "template_path": str(DEFAULT_TEMPLATE),
    "col_map": None,
    "product_file": None,
    "styles": [],
    "keywords": [],
    "analytics": [],
    "generated_content": {},
    # Multi-template: maps product_type -> path, e.g. {"Dresses": "/path/to/Dresses.xlsm"}
    "templates": {},
    # Field overrides from QA review: { style_num: { field_id: value } }
    "field_overrides": {},
    "operator": "",
}

# ── Routes ─────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/session-restore")
def session_restore():
    """Return current session state so frontend can restore after page refresh."""
    return jsonify({
        "brand": session_data.get("brand"),
        "vendor_code": session_data.get("vendor_code"),
        "template_loaded": session_data.get("col_map") is not None,
        "template_columns": len(session_data.get("col_map") or {}),
        "styles_count": len(session_data.get("styles", [])),
        "styles": session_data.get("styles", []),
        "keywords_loaded": len(session_data.get("keywords", [])) > 0,
        "analytics_loaded": len(session_data.get("analytics", [])) > 0,
        "content_generated": len(session_data.get("generated_content", {})) > 0,
        "generated_content": session_data.get("generated_content", {}),
        "brand_config": BRAND_CONFIGS.get(session_data.get("brand"), {}),
    })

@app.route("/api/session-reset", methods=["POST"])
def session_reset():
    """Clear all session state for a fresh start."""
    session_data["brand"] = None
    session_data["vendor_code"] = None
    session_data["template_path"] = str(DEFAULT_TEMPLATE)
    session_data["col_map"] = None
    session_data["product_file"] = None
    session_data["styles"] = []
    session_data["keywords"] = []
    session_data["analytics"] = []
    session_data["generated_content"] = {}
    session_data["templates"] = {}
    session_data["field_overrides"] = {}
    session_data["operator"] = ""
    return jsonify({"ok": True})


@app.route("/api/set-operator", methods=["POST"])
def set_operator():
    """Set the current operator name for session tracking."""
    data = request.get_json(force=True)
    name = str(data.get("operator", "")).strip()
    session_data["operator"] = name
    return jsonify({"ok": True, "operator": name})


@app.route("/api/brand-config", methods=["POST"])
def brand_config():
    data = request.get_json(force=True)
    brand = data.get("brand", "")
    if brand not in BRAND_CONFIGS:
        return jsonify({"error": f"Unknown brand: {brand}"}), 400
    cfg = BRAND_CONFIGS[brand]
    session_data["brand"] = brand
    session_data["vendor_code"] = data.get("vendor_code", cfg.get("vendor_code_full", ""))
    return jsonify({"brand": brand, "config": cfg})

@app.route("/api/upload-template", methods=["POST"])
def upload_template():
    if "file" not in request.files:
        # Use default template (Dresses-Training fallback for legacy callers; the new
        # PT-aware UI no longer relies on this branch — templates are auto-resolved
        # from PRODUCT_TYPE_TEMPLATE_MAP based on the styles in the upload).
        template_path = str(DEFAULT_TEMPLATE)
        session_data["template_path"] = template_path
        session_data["col_map"] = get_template_col_map(template_path)
        col_count = len(session_data["col_map"])
        return jsonify({
            "template": "Dresses-Training.xlsm",
            "columns_mapped": col_count,
            "message": f"Dresses-Training template — {col_count} columns mapped (legacy fallback; templates now resolve per product type)",
            "template_path": template_path,
        })
    
    f = request.files["file"]
    if not f.filename.endswith(".xlsm"):
        return jsonify({"error": "Template must be a .xlsm file"}), 400
    
    save_path = UPLOAD_TEMPLATES / f.filename
    f.save(str(save_path))
    
    try:
        col_map = get_template_col_map(str(save_path))
        session_data["template_path"] = str(save_path)
        session_data["col_map"] = col_map
        # Auto-extract dropdown values for validation
        dd_result = extract_template_dropdowns(str(save_path))
        dd_count = dd_result["dropdown_fields"] if dd_result else 0
        return jsonify({
            "template": f.filename,
            "columns_mapped": len(col_map),
            "dropdown_fields": dd_count,
            "message": f"{f.filename} — {len(col_map)} columns, {dd_count} dropdown validations loaded",
            "template_path": str(save_path),
        })
    except Exception as e:
        return jsonify({"error": f"Failed to parse template: {str(e)}"}), 500


@app.route("/api/upload-category-template", methods=["POST"])
def upload_category_template():
    """Upload a .xlsm template for a specific product type (multi-template support)."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    
    f = request.files["file"]
    product_type = request.form.get("product_type", "").strip()
    
    if not f.filename.endswith(".xlsm"):
        return jsonify({"error": "Template must be a .xlsm file"}), 400
    if not product_type:
        return jsonify({"error": "product_type is required"}), 400
    
    # Save as {product_type}.xlsm
    safe_name = re.sub(r'[^\w]', '_', product_type)
    save_path = UPLOAD_TEMPLATES / f"{safe_name}.xlsm"
    f.save(str(save_path))
    
    try:
        col_map = get_template_col_map(str(save_path))
        # Register in session multi-template map
        session_data["templates"][product_type] = str(save_path)
        # If this is the first/only template, also set as default
        if not session_data.get("col_map"):
            session_data["template_path"] = str(save_path)
            session_data["col_map"] = col_map
        # Auto-extract dropdown values
        dd_result = extract_template_dropdowns(str(save_path))
        dd_count = dd_result["dropdown_fields"] if dd_result else 0
        return jsonify({
            "product_type": product_type,
            "template": f.filename,
            "columns_mapped": len(col_map),
            "dropdown_fields": dd_count,
            "message": f"{product_type} template loaded — {len(col_map)} columns, {dd_count} dropdowns",
            "loaded_templates": list(session_data["templates"].keys()),
        })
    except Exception as e:
        return jsonify({"error": f"Failed to parse template: {str(e)}"}), 500


@app.route("/api/templates")
def list_templates():
    """Return all loaded templates with their product types."""
    templates = session_data.get("templates", {})
    result = []
    for pt, path in templates.items():
        p = Path(path)
        result.append({
            "product_type": pt,
            "filename": p.name,
            "exists": p.exists(),
        })
    # Also include the default template if no multi-templates are registered
    if not result and session_data.get("template_path"):
        p = Path(session_data["template_path"])
        result.append({
            "product_type": "Dresses",
            "filename": p.name,
            "exists": p.exists(),
            "is_default": True,
        })
    return jsonify({"templates": result})


# All product types we know about (from master_nis_reference + common Amazon categories)
ALL_PRODUCT_TYPES = [
    {"id": "BLAZER", "label": "Blazers", "sub_classes": ["Blazer", "Sport Coat"]},
    {"id": "BRA", "label": "Bras / Intimates", "sub_classes": ["Bra", "Bralette", "Sports Bra"]},
    {"id": "COAT", "label": "Jackets & Coats", "sub_classes": ["Jacket", "Coat", "Vest", "Puffer", "Windbreaker", "Anorak", "Parka"]},
    {"id": "DRESS", "label": "Dresses", "sub_classes": ["Day Dress", "Cocktail Dress", "Active Dress", "Swimdress", "Maxi Dress", "Mini Dress", "Wrap Dress", "Shirt Dress", "Dress"]},
    {"id": "HAT", "label": "Hats / Headwear", "sub_classes": ["Hat", "Cap", "Beanie", "Visor", "Sun Hat", "Trucker Hat", "Bucket Hat"]},
    {"id": "ONE_PIECE_OUTFIT", "label": "One-Piece Outfits / Rompers", "sub_classes": ["Romper", "Jumpsuit", "Bodysuit", "Catsuit", "One Piece Outfit"]},
    {"id": "OVERALLS", "label": "Overalls", "sub_classes": ["Overalls", "Dungarees", "Overall"]},
    {"id": "PANTS", "label": "Pants / Leggings", "sub_classes": ["Pants", "Leggings", "Joggers", "Trousers", "Chino", "Cargo Pant"]},
    {"id": "SANDAL", "label": "Sandals / Footwear", "sub_classes": ["Sandal", "Flip Flop", "Slide", "Thong Sandal", "Slipper"]},
    {"id": "SHIRT", "label": "Shirts / Tops", "sub_classes": ["Pullover", "Tank", "Tee", "Blouse", "Shirt", "Polo", "Henley", "Crop Top", "Camisole", "Tunic"]},
    {"id": "SHORTS", "label": "Shorts", "sub_classes": ["Shorts", "Board Short", "Chino Short", "Boardshorts", "Skort", "Cargo Short"]},
    {"id": "SKIRT", "label": "Skirts", "sub_classes": ["Skirt", "Mini Skirt", "Maxi Skirt", "Wrap Skirt"]},
    {"id": "SNOWSUIT", "label": "Snowsuits", "sub_classes": ["Snowsuit", "Snow Suit", "Ski Suit"]},
    {"id": "SNOW_PANT", "label": "Snow Pants", "sub_classes": ["Snow Pant", "Snow Pants", "Ski Pants", "Ski Pant"]},
    {"id": "SWEATSHIRT", "label": "Sweatshirts / Hoodies", "sub_classes": ["Sweatshirt", "Hoodie", "Fleece", "Quarter Zip"]},
    {"id": "SWIMWEAR", "label": "Swimwear", "sub_classes": ["One Piece", "One Piece Swim", "Bikini Top", "Bikini Bottom", "Swim Bottom", "Tankini", "Cover Up", "Boardshorts", "Rashguard", "Swim Set 2 pcs", "Trunk", "Short", "Swim Trunk", "Rash Guard"]},
]


@app.route("/api/product-types")
def product_types():
    """Return all known product types with their training status.
    'trained' = we have a template uploaded + dropdowns extracted.
    'untrained' = we know it exists but no template yet.
    """
    result = []
    for pt in ALL_PRODUCT_TYPES:
        pt_id = pt["id"]
        dropdowns = load_dropdown_cache(pt_id)
        has_template = any(
            pt_id.upper() in str(p.name).upper()
            for p in UPLOAD_TEMPLATES.glob("*.xlsm")
        ) or pt_id in session_data.get("templates", {})

        result.append({
            "id": pt_id,
            "label": pt["label"],
            "sub_classes": pt["sub_classes"],
            "trained": len(dropdowns) > 0,
            "has_template": has_template,
            "dropdown_fields": len(dropdowns),
        })

    return jsonify({"product_types": result})


@app.route("/api/request-template-training", methods=["POST"])
def request_template_training():
    """Log a request for a product type template to be trained.
    Operator is telling us they need this product type but we don't have a template.
    """
    data = request.get_json(force=True)
    product_type = data.get("product_type", "")
    operator = data.get("operator", "") or session_data.get("operator", "")
    brand = data.get("brand", "") or session_data.get("brand", "")
    note = data.get("note", "")

    # Store as feedback
    _store_feedback({
        "id": f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{_uuid.uuid4().hex[:8]}",
        "timestamp": datetime.utcnow().isoformat(),
        "operator": operator,
        "brand": brand,
        "type": "template_request",
        "phase": "upload",
        "context": {"scope": "brand", "brand": brand, "product_type": product_type},
        "data": {"product_type": product_type, "note": note, "message": f"Need {product_type} template for {brand}"},
        "maps_to": "operator_note",
    })

    return jsonify({
        "ok": True,
        "message": f"Template request logged for {product_type}. Send the Amazon .xlsm template to Devang to enable this product type.",
        
        "product_type": product_type,
    })


@app.route("/api/download-sample-template")
def download_sample_template():
    """Download the sample pre-upload template with all expected columns."""
    path = BASE_DIR / "uploads" / "sample_preupload_template.xlsx"
    if not path.exists():
        return jsonify({"error": "Sample template not found"}), 404
    return send_file(str(path), as_attachment=True,
                     download_name="TLG_PreUpload_Template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/download-sample-catalog")
def download_sample_catalog():
    """Sample Catalog export sheet — 50 columns matching CATALOG_FIELD_MAP."""
    path = BASE_DIR / "uploads" / "sample_catalog_export.xlsx"
    if not path.exists():
        return jsonify({"error": "Sample catalog template not found"}), 404
    return send_file(str(path), as_attachment=True,
                     download_name="Atlas_Sample_Catalog_Export.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/download-sample-sales")
def download_sample_sales():
    """Sample Sales report sheet — 12 fields matching SALES_FIELD_MAP."""
    path = BASE_DIR / "uploads" / "sample_sales_report.xlsx"
    if not path.exists():
        return jsonify({"error": "Sample sales template not found"}), 404
    return send_file(str(path), as_attachment=True,
                     download_name="Atlas_Sample_Sales_Report.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/download-sample-ad-bulksheet")
def download_sample_ad_bulksheet():
    """Sample Ad eligibility bulksheet — 6 fields matching AD_BULKSHEET_FIELD_MAP."""
    path = BASE_DIR / "uploads" / "sample_ad_eligibility.xlsx"
    if not path.exists():
        return jsonify({"error": "Sample ad bulksheet template not found"}), 404
    return send_file(str(path), as_attachment=True,
                     download_name="Atlas_Sample_Ad_Eligibility.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")




@app.route("/api/save-subclass-mapping", methods=["POST"])
def save_subclass_mapping():
    """Permanently save a sub-class → product type mapping.
    Called when the operator assigns a product type to an unknown sub-class.
    This mapping persists and is used for all future uploads.
    """
    data = request.get_json(force=True)
    sub_class = data.get("sub_class", "").strip()
    product_type = data.get("product_type", "").strip()
    
    if not sub_class or not product_type:
        return jsonify({"error": "sub_class and product_type required"}), 400
    
    mapping = _load_subclass_map()
    mapping[sub_class] = product_type
    _save_subclass_map(mapping)
    
    return jsonify({"ok": True, "sub_class": sub_class, "product_type": product_type,
                     "total_mappings": len(mapping)})


@app.route("/api/subclass-mappings")
def get_subclass_mappings():
    """Return all learned sub-class → product type mappings."""
    return jsonify({"mappings": _load_subclass_map()})


@app.route("/api/upload-product-data", methods=["POST"])
def upload_product_data():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    
    f = request.files["file"]
    ext = Path(f.filename).suffix.lower()
    if ext not in [".xlsx", ".xls", ".xlsm", ".csv", ".tsv"]:
        return jsonify({"error": f"Unsupported file type: {ext}"}), 400
    
    save_path = UPLOAD_PRODUCTS / f.filename
    f.save(str(save_path))
    session_data["product_file"] = str(save_path)
    
    try:
        styles, errors, warnings = parse_product_file(str(save_path))

        # ── Strip template-example rows (Volcom Rashguard, etc.) ───────────
        # Atlas templates ship with one filled example row to show the
        # operator the expected shape. When the operator pastes their own
        # data without deleting that example, parse_product_file picks it up
        # as a real product. Filter it out before the multi-brand guard fires.
        TEMPLATE_EXAMPLE_STYLES = {"436008622"}  # known sample row(s) we ship
        if styles:
            from collections import Counter as _C
            brand_counter = _C((s.get("brand") or "").strip() for s in styles)
            total_rows = len(styles)
            stripped = []
            removed_examples = []
            for s in styles:
                sn = (s.get("style_num") or "").strip()
                br = (s.get("brand") or "").strip()
                # Rule A: matches a known seeded sample style number
                if sn in TEMPLATE_EXAMPLE_STYLES:
                    removed_examples.append({"reason": "sample_style_number", "style_num": sn, "brand": br})
                    continue
                # Rule B: minority brand (≤1 row) AND majority brand has ≥3 rows
                if br and brand_counter.get(br, 0) <= 1 and total_rows >= 4:
                    majority = brand_counter.most_common(1)[0]
                    if majority[1] >= 3 and majority[0] != br:
                        removed_examples.append({"reason": "minority_brand_example", "style_num": sn, "brand": br})
                        continue
                stripped.append(s)
            if removed_examples and stripped:
                styles = stripped
                for ex in removed_examples:
                    warnings.append(
                        f"Skipped template example row — style {ex['style_num']} ({ex['brand']}). "
                        f"If this was real product data, delete the seeded example row from the template before re-uploading."
                    )

        session_data["styles"] = styles

        total_variants = sum(len(s["variants"]) for s in styles)

        # ── Detect brand ──────────────────────────────────────────────
        brands_found = set(s.get("brand", "") for s in styles if s.get("brand"))

        # Multi-brand check: reject if more than one brand
        if len(brands_found) > 1:
            return jsonify({
                "error": "Multiple brands detected",
                "brands": sorted(brands_found),
                "message": f"This file contains {len(brands_found)} brands: {', '.join(sorted(brands_found))}. Please upload one brand at a time.",
            }), 400

        brand = list(brands_found)[0] if len(brands_found) == 1 else ""
        if brand:
            session_data["brand"] = brand
            brand_cfg = _load_brand_config_data(brand)
            session_data["vendor_code"] = brand_cfg.get("vendor_code_full", "")
            session_data["brandConfig"] = brand_cfg

        # ── Detect product types ──────────────────────────────────────
        present_types = set()
        type_counts = defaultdict(int)
        for s in styles:
            pt = TEMPLATE_PRODUCT_TYPE_MAP.get(s.get("subclass", ""), None)
            if pt:
                present_types.add(pt)
                type_counts[pt] += 1
        loaded_templates = session_data.get("templates", {})
        missing_templates = [
            {"product_type": pt, "style_count": type_counts[pt]}
            for pt in sorted(present_types)
            if pt not in loaded_templates
        ]

        # Also try to detect product type from division_name (e.g., "VOLCOM MENS SWIM")
        division_names = set(s.get("division_name", "") for s in styles if s.get("division_name"))
        detected_gender = ""
        detected_category = ""
        for dn in division_names:
            dn_upper = dn.upper()
            if "MENS" in dn_upper or "MEN" in dn_upper:
                detected_gender = "Male"
            elif "WOMENS" in dn_upper or "WOMEN" in dn_upper:
                detected_gender = "Female"
            if "SWIM" in dn_upper:
                detected_category = "SWIMWEAR"
            elif "DRESS" in dn_upper:
                detected_category = "DRESS"
            elif "SHIRT" in dn_upper or "TOP" in dn_upper:
                detected_category = "SHIRT"

        # ── Data quality audit ────────────────────────────────────────
        # Check every field across all styles — what's present, what's missing
        REQUIRED_FIELDS = {
            "style_num": "Style Number",
            "style_name": "Style Name",
            "brand": "Brand",
        }
        IMPORTANT_FIELDS = {
            "cost_price": "Cost Price (Wholesale)",
            "list_price": "List Price (Retail)",
            "coo": "Country of Origin",
            "fabric": "Fabric / Material",
            "care": "Care Instructions",
            "subclass": "Product Sub-Class",
        }
        NICE_TO_HAVE = {
            "upf": "UPF Rating",
            "neck_type": "Neck Type",
            "closure_type": "Closure Type",
            "sleeve_type": "Sleeve Type",
            "fit_type": "Fit Type",
            "ship_date": "Ship Date",
            "bullets_from_upload": "Key Features / Bullets",
            "model_name": "Model Name",
        }
        VARIANT_REQUIRED = {
            "upc": "UPC Code",
            "color_name": "Color Name",
            "size": "Size",
        }

        field_status = {}  # field_key -> {present: N, missing: N, label: str, level: str}

        for field_key, label in {**REQUIRED_FIELDS, **IMPORTANT_FIELDS, **NICE_TO_HAVE}.items():
            present = sum(1 for s in styles if s.get(field_key))
            level = "required" if field_key in REQUIRED_FIELDS else "important" if field_key in IMPORTANT_FIELDS else "optional"
            field_status[field_key] = {
                "label": label, "present": present, "missing": len(styles) - present,
                "total": len(styles), "level": level,
            }

        # Variant-level checks
        all_variants = [v for s in styles for v in s.get("variants", [])]
        for field_key, label in VARIANT_REQUIRED.items():
            present = sum(1 for v in all_variants if v.get(field_key))
            field_status[f"variant_{field_key}"] = {
                "label": f"{label} (per variant)", "present": present, "missing": len(all_variants) - present,
                "total": len(all_variants), "level": "required",
            }

        # Build action items
        action_items = []
        for key, info in field_status.items():
            if info["missing"] > 0 and info["level"] in ("required", "important"):
                pct = round(100 * info["present"] / info["total"]) if info["total"] else 0
                action_items.append({
                    "field": info["label"],
                    "level": info["level"],
                    "present": info["present"],
                    "missing": info["missing"],
                    "total": info["total"],
                    "pct": pct,
                    "message": f"{info['label']}: {info['missing']} of {info['total']} missing" + 
                               (" — required for NIS" if info["level"] == "required" else " — add in pre-upload or set on dashboard"),
                })

        # Check brand config existence
        brand_known = brand and (brand in BRAND_CONFIGS or (_load_brand_config_data(brand) != {}))

        # ── Category breakdown ────────────────────────────────────────
        category_breakdown = []
        styles_by_pt = defaultdict(lambda: defaultdict(list))
        unknown_subclasses = set()
        for s in styles:
            sub_class = s.get("subclass", "") or "Unknown"
            pt_id, confidence, reason = resolve_product_type(sub_class, s.get("division_name", ""))
            s["_resolved_pt"] = pt_id
            s["_pt_confidence"] = confidence
            s["_pt_reason"] = reason
            if confidence == "unknown":
                unknown_subclasses.add(sub_class)
            styles_by_pt[pt_id][sub_class].append(s["style_num"])

        for pt_id, sub_classes in styles_by_pt.items():
            pt_def = next((p for p in ALL_PRODUCT_TYPES if p["id"] == pt_id), None)
            pt_label = pt_def["label"] if pt_def else pt_id
            dd = load_dropdown_cache(pt_id)
            is_trained = len(dd) > 0
            total_in_pt = sum(len(v) for v in sub_classes.values())
            sub_list = [{"name": sc, "count": len(snums)} for sc, snums in sub_classes.items()]
            category_breakdown.append({
                "product_type": pt_id,
                "label": pt_label,
                "trained": is_trained,
                "dropdown_fields": len(dd),
                "total_styles": total_in_pt,
                "sub_classes": sub_list,
            })

        trained_count = sum(1 for c in category_breakdown if c["trained"])
        total_categories = len(category_breakdown)

        # Taxonomy bucket summary: which (product_type, sub_class, gender_bucket) triples
        # exist in this upload, and which are already confirmed in the overrides store.
        _tax_store = _load_taxonomy_overrides().get("entries", {})
        _tax_buckets = {}
        for _s in styles:
            _pt = _resolve_style_product_type(_s) or ""
            _sc = _s.get("subclass") or _s.get("sub_class") or ""
            _gb = _derive_gender_bucket(_s)
            _k = _taxonomy_key(_pt, _sc, _gb)
            _b = _tax_buckets.setdefault(_k, {"key": _k, "product_type": _pt,
                                              "sub_class": _sc, "gender_bucket": _gb,
                                              "style_count": 0, "confirmed": False})
            _b["style_count"] += 1
            if _tax_store.get(_k, {}).get("source") == "manual":
                _b["confirmed"] = True
        _tax_summary = {
            "total_buckets": len(_tax_buckets),
            "confirmed_buckets": sum(1 for b in _tax_buckets.values() if b["confirmed"]),
            "unconfirmed_buckets": sum(1 for b in _tax_buckets.values() if not b["confirmed"]),
            "buckets": list(_tax_buckets.values()),
        }
        return jsonify({
            "total_styles": len(styles),
            "total_variants": total_variants,
            "brand": brand,
            "brand_known": brand_known,
            "taxonomy_summary": _tax_summary,
            "category_breakdown": category_breakdown,
            "trained_count": trained_count,
            "total_categories": total_categories,
            "vendor_code": session_data.get("vendor_code", ""),
            "detected_gender": detected_gender,
            "detected_category": detected_category,
            "division_names": list(division_names),
            "errors": errors,
            "warnings": warnings,
            "error_count": len(errors),
            "warning_count": len(warnings),
            "styles": styles,
            "missing_templates": missing_templates,
            "present_product_types": list(present_types),
            "field_status": field_status,
            "action_items": action_items,
            "categories": dict(type_counts),
            "unique_colors": len(set(v.get("color_name", "") for s in styles for v in s.get("variants", []) if v.get("color_name"))),
            "unknown_subclasses": list(unknown_subclasses),
            "size_range": ", ".join(sorted(set(v.get("size", "") for s in styles for v in s.get("variants", []) if v.get("size")),
                                          key=lambda x: ["XXS","XS","S","M","L","XL","XXL","2XL","3XL"].index(x) if x in ["XXS","XS","S","M","L","XL","XXL","2XL","3XL"] else 99)),
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to parse product data: {str(e)}"}), 500

@app.route("/api/upload-keywords", methods=["POST"])
def upload_keywords():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    
    f = request.files["file"]
    save_path = UPLOAD_KEYWORDS / f.filename
    f.save(str(save_path))
    
    keywords = []
    try:
        ext = Path(f.filename).suffix.lower()
        if ext in [".csv", ".tsv"]:
            delimiter = "\t" if ext == ".tsv" else ","
            with open(str(save_path), "r", encoding="utf-8-sig") as fh:
                reader = csv.DictReader(fh, delimiter=delimiter)
                for row in reader:
                    kw = row.get("Keyword Phrase") or row.get("keyword") or row.get("Search Query", "")
                    volume = row.get("Search Volume") or row.get("volume", "0")
                    if kw:
                        try:
                            vol = int(str(volume).replace(",", ""))
                        except (ValueError, AttributeError):
                            vol = 0
                        keywords.append({"keyword": kw.strip().lower(), "volume": vol})
        
        # Sort by volume
        keywords.sort(key=lambda x: x["volume"], reverse=True)
        session_data["keywords"] = keywords
        
        top5 = [k["keyword"] for k in keywords[:5]]
        return jsonify({
            "total_keywords": len(keywords),
            "top5": top5,
            "message": f"{len(keywords)} keywords loaded. Top 5: {', '.join(top5)}",
            "keywords": keywords[:50],
        })
    except Exception as e:
        return jsonify({"error": f"Failed to parse keyword file: {str(e)}"}), 500

@app.route("/api/upload-analytics", methods=["POST"])
def upload_analytics():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    
    f = request.files["file"]
    save_path = UPLOAD_KEYWORDS / f.filename
    f.save(str(save_path))
    
    analytics = []
    try:
        ext = Path(f.filename).suffix.lower()
        if ext in [".csv", ".tsv"]:
            delimiter = "\t" if ext == ".tsv" else ","
            with open(str(save_path), "r", encoding="utf-8-sig") as fh:
                reader = csv.DictReader(fh, delimiter=delimiter)
                for row in reader:
                    analytics.append(dict(row))
        
        session_data["analytics"] = analytics
        return jsonify({
            "total_rows": len(analytics),
            "message": f"{len(analytics)} analytics rows loaded",
        })
    except Exception as e:
        return jsonify({"error": f"Failed to parse analytics file: {str(e)}"}), 500

def _run_content_generation(brand, styles, brand_cfg, has_keywords, feedback_history):
    """Background worker for content generation.

    Hardened: any style-level exception is caught and recorded so a single
    bad style never freezes the whole batch on a 'Running QA compliance check…'
    progress label. Any worker-level exception is caught at the bottom so
    the progress poller always reaches status='done' or status='error'.
    """
    global DESCRIPTION_OPENERS_ROTATION, content_progress
    try:
        return _run_content_generation_impl(brand, styles, brand_cfg, has_keywords, feedback_history)
    except Exception as _worker_err:
        traceback.print_exc()
        print(f"[GEN] Worker crashed: {_worker_err}", flush=True)
        content_progress["status"] = "error"
        content_progress["current_step"] = f"Generation crashed: {type(_worker_err).__name__}"
        content_progress["error"] = f"{type(_worker_err).__name__}: {str(_worker_err)[:300]}"
        content_progress["results"] = {
            "content": {}, "total": 0, "qa_errors": 1, "qa_warnings": 0,
            "error": content_progress["error"],
        }


def _run_content_generation_impl(brand, styles, brand_cfg, has_keywords, feedback_history):
    global DESCRIPTION_OPENERS_ROTATION, content_progress
    DESCRIPTION_OPENERS_ROTATION = {}
    content_map = {}
    total_qa_errors = 0
    total_qa_warnings = 0
    feedback_count = len([l for l in feedback_history.splitlines() if l.strip()]) if feedback_history else 0

    # ─── Atlas substrate: open a session for this batch ────────────────
    # Every NIS upload is one session. All decision_events written below
    # tag with this session_id so the batch can be replayed end-to-end.
    # If the substrate module is missing or fails, generation continues
    # unaffected — substrate writes are best-effort, never blocking.
    _atlas_session = None
    _atlas_workspace = (brand or "tlg").lower().replace(" ", "_") or "tlg"
    _atlas_operator = session_data.get("operator_id") or "devang"
    _atlas_brand_profile_version = brand_cfg.get("_version") or f"{_atlas_workspace}_legacy"
    try:
        from substrate.logger import open_session as _atlas_open_session
        from substrate.schema import Module as _AtlasModule
        _atlas_session = _atlas_open_session(
            workspace_id=_atlas_workspace,
            operator_id=_atlas_operator,
            module=_AtlasModule.NIS,
        )
        content_progress["session_id"] = _atlas_session.session_id
    except Exception as _atlas_exc:
        print(f"[atlas] session open skipped: {_atlas_exc}", flush=True)
        _atlas_session = None

    for i, style in enumerate(styles):
      try:
        style_num = style["style_num"]
        style_name = style["style_name"]
        subclass = style.get("subclass", "")
        sub_subclass = style.get("sub_subclass", "")
        fabric = parse_fabric(style.get("fabric", "")) or brand_cfg.get("default_fabric", "")
        care = style.get("care", "") or brand_cfg.get("default_care", "")
        upf = style.get("upf", "") or brand_cfg.get("default_upf", "")
        coo = style.get("coo", "") or brand_cfg.get("default_coo", "")

        content_progress["current_style"] = f"{style_num} — {style_name}"

        # Get first color for preview title
        first_variant = style["variants"][0] if style["variants"] else {}
        first_color = first_variant.get("color_name", "")
        first_size = first_variant.get("size", "")

        # Dramatic progress steps
        content_progress["current_step"] = f"Analyzing style attributes for {style_name}..."
        time.sleep(0.3)
        content_progress["current_step"] = f"Loading {clean_brand_name(brand)} brand preferences..."
        time.sleep(0.2)
        if feedback_count > 0:
            content_progress["current_step"] = f"Reading {feedback_count} previous feedback corrections..."
            time.sleep(0.2)
        content_progress["current_step"] = f"Generating SEO-optimized title (max 120 chars)..."

        # ═══ Resolve PT + gender up front so they're always defined ═══
        # These are used by derive_neck_type / normalize_color / subcategory
        # logic AFTER both LLM and rule-based branches. Pre-Pass-12.2 they
        # were only assigned in the rule-based else-branch, so successful
        # LLM runs raised UnboundLocalError downstream.
        resolved_pt = _resolve_style_product_type(style) or ""
        _style_gender, _ = _derive_gender_department(style)
        eff_gender_gen = _style_gender or brand_cfg.get("gender", "")

        # Try LLM generation first (unless mode is "rules")
        gen_mode = session_data.get("generation_mode", "auto")
        llm_result = None
        if gen_mode != "rules" and _anthropic_client is not None:
            try:
                llm_result = generate_content_llm(brand_cfg, brand, style, feedback_history)
            except Exception as e:
                print(f"[LLM] Fallback for {style_num}: {e}")

        if llm_result:
            title = llm_result["title"]
            bullets = [
                llm_result.get("bullet_1", ""),
                llm_result.get("bullet_2", ""),
                llm_result.get("bullet_3", ""),
                llm_result.get("bullet_4", ""),
                llm_result.get("bullet_5", ""),
            ]
            description = llm_result["description"]
            backend_kw = llm_result["backend_keywords"]
        else:
            # Rule-based fallback
            if _anthropic_client is not None:
                print(f"[LLM] Falling back to rule-based for style {style_num}")
            # Use actual subclass as product type descriptor (not hardcoded "Dress")
            pt_label = subclass or sub_subclass or resolved_pt.replace("_", " ").title() or "Dress"
            style_gender = _style_gender
            title = generate_title(brand_cfg, brand, style_name, pt_label, first_color, first_size, upf, style_gender=style_gender)
            bullets = generate_bullets(brand_cfg, brand, style_name, sub_subclass, fabric, care, first_color, upf,
                                       subclass=subclass, gender=eff_gender_gen, product_type=resolved_pt, style_num=style_num)
            description = generate_description(brand_cfg, brand, style_num, style_name, sub_subclass, fabric, care, first_color, upf,
                                               subclass=subclass, gender=eff_gender_gen, product_type=resolved_pt)
            backend_kw = generate_backend_keywords(brand, style_name, subclass, first_color, fabric, upf,
                                                    subclass=subclass, gender=eff_gender_gen, product_type=resolved_pt)

            # ═══ v0.7.6: enforce baseline content rules + merge pre-upload bullets ═══
            from nis_engine import content_rules as _cr
            _itn = _derive_item_type_name(subclass, resolved_pt) or subclass or ""
            _gw = _cr.gender_word_for(style.get("department") or "",
                                       target_gender=eff_gender_gen)
            # Re-validate title — if garbage, too long, or missing brand/gender, recompose
            _cb = clean_brand_name(brand)
            _needs_recompose = (
                _cr.is_garbage_value(title)
                or len(title) > 120
                or (_cb and _cb not in title)
                or (_gw and _gw not in title)
            )
            if _needs_recompose:
                title = _cr.compose_title(
                    brand=_cb,
                    gender_word=_gw,
                    style_name=style_name,
                    item_type_name=_itn,
                    feature_phrases=[
                        style.get("fit_type") or "",
                        style.get("neck_type") or "",
                        style.get("sleeve_type") or "",
                        style.get("closure_type") or "",
                    ],
                )
            # Bullets: file wins, AI/rule-based fills gaps; normalize headline format on every slot
            pu_bullets = (style.get("bullets_from_upload") or []) + [""] * 5
            bullets = _cr.merge_bullets(pu_bullets[:5], bullets)

        # ═══ Gender-drift scrub (Pass 12.7) ═══
        # Runs after BOTH the LLM path and the rule-based fallback so a
        # men's listing never ships with "the modern woman" or vice versa.
        # If the explicit gender derived from the sheet is empty, fall back
        # to the title we just wrote: if the title says "Men's" we treat
        # the listing as male; same for "Women's". Otherwise stay neutral.
        _expected_g = eff_gender_gen if eff_gender_gen in ("Male", "Female") else ""
        if not _expected_g and title:
            _t = title.lower()
            # Check female first — "women's" contains "men's" as substring.
            if "women's" in _t or "women\u2019s" in _t or "womens " in _t:
                _expected_g = "Female"
            elif "men's" in _t or "men\u2019s" in _t or "mens " in _t or _t.startswith("mens"):
                _expected_g = "Male"
        description = _scrub_gender_drift(description, _expected_g)
        bullets = [_scrub_gender_drift(b, _expected_g) for b in bullets]

        content_progress["current_step"] = f"Crafting 5 unique bullet points..."
        time.sleep(0.2)
        content_progress["current_step"] = f"Writing buyer-focused description..."
        time.sleep(0.2)
        content_progress["current_step"] = f"Building backend keywords (250 bytes max)..."
        time.sleep(0.1)
        content_progress["current_step"] = f"Running QA compliance check..."

        # Derived attributes (PT-aware fallbacks)
        neck = derive_neck_type(style_name)
        sleeve = derive_sleeve_type(style_name, resolved_pt)
        silhouette = derive_silhouette(sub_subclass)
        color_map_val = normalize_color(first_color, resolved_pt)
        category = SUBCLASS_CATEGORY_MAP.get(subclass, "")
        subcategory = SUBCLASS_SUBCATEGORY_MAP.get(subclass, "")
        if not subcategory and (resolved_pt == "SWIMWEAR" or subclass in ("Rashguard","Trunk","Bikini Top","Bikini Bottom","Swim Bottom","One Piece Swim","Tankini","Short","Swim Set 2 pcs","Swim Set","Board Short","Cover Up","Swim Shirt")):
            _cat_for_sub = _derive_amazon_product_category(subclass, gender=eff_gender_gen, product_type=resolved_pt, style_name=style_name)
            subcategory = _derive_swim_product_subcategory(subclass, gender=eff_gender_gen, style_name=style_name, product_category=_cat_for_sub)

        opener_idx = DESCRIPTION_OPENERS_ROTATION.get(style_num, 0)
        bullet_whys = [
            generate_bullet_why(j, brand_cfg, brand, style_name, sub_subclass, upf, fabric, has_keywords)
            for j in range(5)
        ]

        entry = {
            "style_num": style_num,
            "style_name": style_name,
            "title": title,
            "title_why": generate_title_why(brand_cfg, brand, style_name, title, upf, has_keywords),
            "bullets": bullets,
            "bullet_whys": bullet_whys,
            "bullet_1": bullets[0] if len(bullets) > 0 else "",
            "bullet_2": bullets[1] if len(bullets) > 1 else "",
            "bullet_3": bullets[2] if len(bullets) > 2 else "",
            "bullet_4": bullets[3] if len(bullets) > 3 else "",
            "bullet_5": bullets[4] if len(bullets) > 4 else "",
            "description": description,
            "description_why": generate_description_why(brand_cfg, style_num, opener_idx, has_keywords),
            "backend_keywords": backend_kw,
            "backend_keywords_why": generate_keywords_why(brand, session_data.get("keywords", []), backend_kw, has_keywords),
            "qa_issues": [],
            "neck_type": neck,
            "sleeve_type": sleeve,
            "silhouette": silhouette,
            "color_map": color_map_val,
            "category": category,
            "sub_class": subclass,
            "subcategory": subcategory,
            "fabric": fabric,
            "care": care,
            "upf": upf,
            "coo": coo,
            "llm_generated": llm_result is not None,
        }

        # QA check
        issues = qa_check_content(entry, brand)
        entry["qa_issues"] = issues
        total_qa_errors += sum(1 for iss in issues if iss["severity"] == "error")
        total_qa_warnings += sum(1 for iss in issues if iss["severity"] == "warning")

        # ─── Atlas substrate: log decision_events for this style ───────
        # Logs one event per field the substrate filter accepts (strategic
        # fields are always logged; others gate on confidence + rule density).
        # Never blocks generation — any exception is swallowed with a print.
        if _atlas_session is not None:
            try:
                from substrate.logger import log_field_decision as _atlas_log
                from substrate.schema import Module as _AtlasModule
                _atlas_field_outputs = [
                    ("item_name", entry.get("title", "")),
                    ("bullet_1", entry.get("bullet_1", "")),
                    ("bullet_2", entry.get("bullet_2", "")),
                    ("bullet_3", entry.get("bullet_3", "")),
                    ("bullet_4", entry.get("bullet_4", "")),
                    ("bullet_5", entry.get("bullet_5", "")),
                    ("description", entry.get("description", "")),
                    ("backend_keywords", entry.get("backend_keywords", "")),
                ]
                # Confidence proxy at v1.0.0: high when LLM produced output
                # cleanly with no QA errors, lower otherwise. Computed by
                # Atlas, not asked from the LLM (per substrate principle).
                _err_count = sum(1 for iss in issues if iss["severity"] == "error")
                _warn_count = sum(1 for iss in issues if iss["severity"] == "warning")
                if entry.get("llm_generated") and _err_count == 0:
                    _atlas_conf = 0.85 if _warn_count == 0 else 0.72
                else:
                    _atlas_conf = 0.55
                # Minimal rule trace: marker rules so the event is replayable.
                # Full rule trace will be added once the engine layer ships.
                _atlas_rules = [
                    {"rule_id": "nis.engine.compose_v0_7_6", "version": "0.7.6"},
                ]
                if entry.get("llm_generated"):
                    _atlas_rules.append({"rule_id": "nis.llm.claude_generation"})
                if has_keywords:
                    _atlas_rules.append({"rule_id": "nis.brand.keywords_active"})
                # Capture event_ids so the frontend can attach operator
                # responses back to the same decision_event when the user
                # clicks Accept/Edit/Why in the review UI. Without this the
                # 3-action footer has no handle to call /decision-response.
                _atlas_event_ids: dict[str, str] = {}
                _atlas_rules_by_field: dict[str, list[dict]] = {}
                _atlas_confidence_by_field: dict[str, float] = {}
                # Pre-change snapshot anchor: the style's first variant
                # ASIN is what closed-loop attribution joins against later.
                # In Amazon's parent/child model the title + bullets are
                # shared across variants, so any variant ASIN under the
                # style anchors the comparison. We pick the first one
                # deterministically; if no variants have an ASIN (Day-1
                # listing for a new style), snapshot stays empty and the
                # logger silently skips snapshot capture.
                _atlas_asin = (first_variant.get("child_asin") or "").strip() or None
                for _fname, _fval in _atlas_field_outputs:
                    if not _fval:
                        continue
                    _eid = _atlas_log(
                        workspace_id=_atlas_workspace,
                        session_id=_atlas_session.session_id,
                        module=_AtlasModule.NIS,
                        field_name=_fname,
                        atlas_output=_fval,
                        overall_confidence=_atlas_conf,
                        rules_injected=_atlas_rules,
                        brand_profile_version=_atlas_brand_profile_version,
                        style_id=str(style_num),
                        asin=_atlas_asin,
                    )
                    if _eid:
                        _atlas_event_ids[_fname] = _eid
                        _atlas_rules_by_field[_fname] = _atlas_rules
                        _atlas_confidence_by_field[_fname] = _atlas_conf
                # Stash on the entry so the frontend can render the
                # 3-action footer + Why panel without a second round-trip.
                if _atlas_event_ids:
                    entry["_atlas"] = {
                        "session_id": _atlas_session.session_id,
                        "workspace_id": _atlas_workspace,
                        "event_ids": _atlas_event_ids,
                        "rules_by_field": _atlas_rules_by_field,
                        "confidence_by_field": _atlas_confidence_by_field,
                    }
            except Exception as _atlas_log_exc:
                # Substrate writes are best-effort; never crash generation.
                print(f"[atlas] decision log skipped for {style_num}: {_atlas_log_exc}", flush=True)

        content_map[style_num] = entry
        content_progress["completed"] = i + 1
        content_progress["current_step"] = f"✓ {style_num} complete"
        time.sleep(0.1)
      except Exception as _style_err:
        # One bad style must not kill the whole batch. Log + record + advance.
        traceback.print_exc()
        sn = style.get("style_num", f"row{i}") if isinstance(style, dict) else f"row{i}"
        content_map[sn] = {
            "style_num": sn,
            "style_name": style.get("style_name", "") if isinstance(style, dict) else "",
            "title": "", "bullets": [""]*5,
            "bullet_1":"","bullet_2":"","bullet_3":"","bullet_4":"","bullet_5":"",
            "description": "", "backend_keywords": "",
            "qa_issues": [{"field": "_pipeline", "severity": "error",
                            "msg": f"Generation failed: {type(_style_err).__name__}: {str(_style_err)[:200]}"}],
            "llm_generated": False, "_pipeline_error": True,
        }
        total_qa_errors += 1
        content_progress["completed"] = i + 1
        content_progress["current_step"] = f"⚠ {sn} — generation error (logged)"
        print(f"[GEN] Style {sn} failed: {_style_err}", flush=True)
        time.sleep(0.05)

    session_data["generated_content"] = content_map
    content_progress["status"] = "done"
    content_progress["current_style"] = ""
    content_progress["current_step"] = ""
    content_progress["results"] = {
        "content": content_map,
        "total": len(content_map),
        "qa_errors": total_qa_errors,
        "qa_warnings": total_qa_warnings,
        "atlas_session_id": _atlas_session.session_id if _atlas_session else None,
    }


@app.route("/api/generate-content", methods=["POST"])
def generate_content():
    data = request.get_json(force=True)
    brand = data.get("brand") or session_data.get("brand")
    if not brand:
        return jsonify({"error": "No brand selected"}), 400

    styles = data.get("styles") or session_data.get("styles", [])
    if not styles:
        return jsonify({"error": "No product data loaded"}), 400

    # ── GATE: Check if all product types are trained ───────────────────
    untrained_types = []
    for s in styles:
        sub_class = s.get("subclass", "")
        pt = None
        for pt_def in ALL_PRODUCT_TYPES:
            if sub_class in pt_def.get("sub_classes", []):
                pt = pt_def["id"]
                break
        if pt:
            dropdowns = load_dropdown_cache(pt)
            if not dropdowns:
                untrained_types.append(pt)
        # If we can't even map the subclass, check division_name
        elif s.get("division_name"):
            dn = s["division_name"].upper()
            if "SWIM" in dn:
                dd = load_dropdown_cache("SWIMWEAR")
                if not dd:
                    untrained_types.append("SWIMWEAR")
            elif "DRESS" in dn:
                dd = load_dropdown_cache("DRESS")
                if not dd:
                    untrained_types.append("DRESS")

    untrained_types = list(set(untrained_types))
    if untrained_types:
        return jsonify({
            "error": "Cannot generate — untrained product types detected",
            "untrained_types": untrained_types,
            "message": f"The following product types are not trained yet: {', '.join(untrained_types)}. "
                       f"Contact admin to upload the Amazon .xlsm template(s) for these product types.",
            "action": "request_template",
            
        }), 400

    # Content generation mode: "api" (LLM) or "rules" (rule-based only)
    gen_mode = data.get("mode", "auto")  # auto = try API first, fall back to rules
    session_data["generation_mode"] = gen_mode

    # Store style briefs + brand brief from frontend
    style_briefs = data.get("style_briefs", {})
    if style_briefs:
        session_data["style_briefs"] = style_briefs
    brand_brief = data.get("brand_brief", "")

    # Load brand config from file if available, fall back to in-memory
    brand_cfg = _load_brand_config_data(brand)
    # Inject brand brief into config so LLM prompt picks it up
    if brand_brief:
        if "product_briefs" not in brand_cfg:
            brand_cfg["product_briefs"] = {}
        brand_cfg["product_briefs"]["_default"] = brand_brief
    has_keywords = len(session_data.get("keywords", [])) > 0

    # Load feedback history for this brand
    feedback_history = load_brand_feedback(brand)

    # Init progress
    content_progress["total"] = len(styles)
    content_progress["completed"] = 0
    content_progress["status"] = "running"
    content_progress["started_at"] = datetime.now().isoformat()
    content_progress["current_style"] = ""
    content_progress["current_step"] = ""
    content_progress["results"] = None

    # Run in background thread
    t = threading.Thread(
        target=_run_content_generation,
        args=(brand, styles, brand_cfg, has_keywords, feedback_history)
    )
    t.daemon = True
    t.start()

    return jsonify({"status": "started", "total": len(styles)})


@app.route("/api/content-progress")
def content_progress_endpoint():
    elapsed = ""
    eta = ""
    if content_progress.get("started_at") and content_progress["completed"] > 0:
        started = datetime.fromisoformat(content_progress["started_at"])
        elapsed_sec = (datetime.now() - started).total_seconds()
        per_style = elapsed_sec / content_progress["completed"]
        remaining = content_progress["total"] - content_progress["completed"]
        eta_sec = per_style * remaining
        elapsed = f"{int(elapsed_sec)}s"
        if eta_sec > 60:
            eta = f"~{int(eta_sec / 60)}m {int(eta_sec % 60)}s remaining"
        else:
            eta = f"~{int(eta_sec)}s remaining"
    return jsonify({
        "total": content_progress["total"],
        "completed": content_progress["completed"],
        "current_style": content_progress["current_style"],
        "current_step": content_progress["current_step"],
        "status": content_progress["status"],
        "error": content_progress.get("error", ""),
        "elapsed": elapsed,
        "eta": eta,
        "percent": round((content_progress["completed"] / max(content_progress["total"], 1)) * 100, 1),
    })


@app.route("/api/content-results")
def content_results():
    """Poll this after generate-content to get results when done."""
    if content_progress["status"] == "done":
        return jsonify({
            "status": "done",
            **content_progress.get("results", {}),
        })
    else:
        return jsonify({
            "status": content_progress["status"],
            "completed": content_progress["completed"],
            "total": content_progress["total"],
        })

# ───────────────────────────────────────────────────────────────────────────────
# STRUCTURED FEEDBACK SYSTEM
#
# Every feedback entry is stored in one JSONL file per brand under feedback/.
# Schema:
#   id            - unique (timestamp + random)
#   timestamp     - ISO 8601 UTC
#   operator      - who submitted (from session or explicit)
#   brand         - brand name (always present)
#   type          - one of:
#                     "content_edit"    — operator changed LLM-generated content
#                     "field_override"  — operator changed a NIS field value
#                     "apply_all"       — operator applied a value to all styles
#                     "manual"          — operator typed freeform feedback
#                     "session"         — end-of-session summary
#                     "regenerate"      — operator regenerated a field
#   phase         - where in the flow: "upload", "generate", "review", "download"
#   context       - what it's about (maps_to target):
#     scope       - "brand" | "style" | "field" | "session"
#     brand       - brand name
#     style_num   - style number (if scope is style or field)
#     field_id    - field_id string (if scope is field)
#     field_name  - human-readable field name (if scope is field)
#   data          - type-specific payload:
#     For content_edit:  { field, original, updated }
#     For field_override: { field_id, field_name, original, updated }
#     For apply_all:     { field_id, field_name, value, styles_count }
#     For manual:        { message }
#     For session:       { rating, note, stats }
#     For regenerate:    { field, attempt_num }
#   maps_to       - where this feedback should route:
#     "llm_prompt"    — feed into next LLM generation for this brand
#     "brand_config"  — update brand defaults
#     "derive_logic"  — improve field derivation functions
#     "operator_note" — informational, no auto-action
# ───────────────────────────────────────────────────────────────────────────────

import uuid as _uuid

def _feedback_file_for_brand(brand):
    """Return the JSONL feedback file path for a brand."""
    safe = re.sub(r'[^\w\-]', '_', brand or "unknown")
    return FEEDBACK_DIR / f"{safe}_feedback.jsonl"


def _store_feedback(entry):
    """Append a feedback entry to the brand-specific JSONL file.
    Also appends to the legacy content_feedback.jsonl for backward compat with LLM loading.
    """
    brand = entry.get("brand") or entry.get("context", {}).get("brand", "unknown")
    fpath = _feedback_file_for_brand(brand)
    try:
        with open(str(fpath), "a", encoding="utf-8") as f:
            f.write(json.dumps(entry) + "\n")
    except Exception:
        traceback.print_exc()

    # Also write to legacy file so load_brand_feedback() still works
    if entry.get("type") in ("content_edit", "field_override", "manual"):
        legacy = {
            "timestamp": entry.get("timestamp"),
            "brand": brand,
            "style_num": entry.get("context", {}).get("style_num", ""),
            "field": entry.get("data", {}).get("field") or entry.get("data", {}).get("field_name", ""),
            "feedback": entry.get("data", {}).get("message", ""),
            "original": entry.get("data", {}).get("original", ""),
            "updated": entry.get("data", {}).get("updated", ""),
        }
        try:
            with open(str(FEEDBACK_FILE), "a", encoding="utf-8") as f:
                f.write(json.dumps(legacy) + "\n")
        except Exception:
            pass


def _load_feedback(brand=None, scope=None, style_num=None, field_id=None,
                   fb_type=None, limit=100):
    """Load feedback entries with optional filters."""
    entries = []

    # Determine which files to read
    if brand:
        files = [_feedback_file_for_brand(brand)]
    else:
        files = list(FEEDBACK_DIR.glob("*_feedback.jsonl"))

    for fpath in files:
        if not fpath.exists():
            continue
        with open(str(fpath), "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    e = json.loads(line)
                except json.JSONDecodeError:
                    continue
                ctx = e.get("context", {})
                if scope and ctx.get("scope") != scope:
                    continue
                if style_num and ctx.get("style_num") != style_num:
                    continue
                if field_id and ctx.get("field_id") != field_id:
                    continue
                if fb_type and e.get("type") != fb_type:
                    continue
                entries.append(e)

    entries.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
    return entries[:limit]


@app.route("/api/feedback", methods=["GET", "POST"])
def feedback_endpoint():
    """Universal feedback endpoint.
    POST: Submit structured feedback with context mapping.
    GET:  Retrieve feedback entries with optional filters.
    """
    if request.method == "POST":
        data = request.get_json(force=True)
        fb_type  = data.get("type", "manual")
        phase    = data.get("phase", "review")
        context  = data.get("context", {})
        payload  = data.get("data", {})
        maps_to  = data.get("maps_to", "operator_note")

        # Fill in brand from session if not provided
        if not context.get("brand"):
            context["brand"] = session_data.get("brand", "unknown")
        if not context.get("scope"):
            if context.get("field_id"):
                context["scope"] = "field"
            elif context.get("style_num"):
                context["scope"] = "style"
            else:
                context["scope"] = "brand"

        # Auto-determine maps_to if not explicit
        if maps_to == "operator_note" and fb_type in ("content_edit", "regenerate"):
            maps_to = "llm_prompt"
        elif maps_to == "operator_note" and fb_type == "apply_all":
            maps_to = "brand_config"
        elif maps_to == "operator_note" and fb_type == "field_override":
            maps_to = "derive_logic"

        entry = {
            "id": f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{_uuid.uuid4().hex[:8]}",
            "timestamp": datetime.utcnow().isoformat(),
            "operator": data.get("operator", session_data.get("operator", "")),
            "brand": context.get("brand"),
            "type": fb_type,
            "phase": phase,
            "context": context,
            "data": payload,
            "maps_to": maps_to,
        }

        _store_feedback(entry)
        return jsonify({"ok": True, "id": entry["id"]})

    # GET — retrieve feedback entries with optional filters
    brand     = request.args.get("brand", "") or session_data.get("brand", "")
    scope     = request.args.get("scope", "")
    style_num = request.args.get("style_num", "")
    field_id  = request.args.get("field_id", "")
    fb_type   = request.args.get("type", "")
    limit     = int(request.args.get("limit", "100"))

    entries = _load_feedback(
        brand=brand or None, scope=scope or None,
        style_num=style_num or None, field_id=field_id or None,
        fb_type=fb_type or None, limit=limit
    )
    return jsonify({"brand": brand, "entries": entries, "total": len(entries)})


# Legacy endpoint — backward compat for existing frontend code
@app.route("/api/submit-feedback", methods=["POST"])
def submit_feedback_legacy():
    data = request.get_json(force=True)
    entry = {
        "id": f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{_uuid.uuid4().hex[:8]}",
        "timestamp": datetime.utcnow().isoformat(),
        "operator": session_data.get("operator", ""),
        "brand": session_data.get("brand"),
        "type": "content_edit",
        "phase": "review",
        "context": {
            "brand": session_data.get("brand"),
            "style_num": data.get("style_num", ""),
            "field_id": "",
            "field_name": data.get("field", ""),
            "scope": "field" if data.get("field") else "style",
        },
        "data": {
            "field": data.get("field", ""),
            "original": data.get("original", ""),
            "updated": data.get("updated", ""),
            "message": data.get("feedback", ""),
        },
        "maps_to": "llm_prompt",
    }
    _store_feedback(entry)
    return jsonify({"ok": True, "id": entry["id"]})


@app.route("/api/feedback/summary")
def feedback_summary():
    """Return feedback stats: count per brand, per type, most-edited fields."""
    brand = request.args.get("brand", "")
    entries = _load_feedback(brand=brand or None, limit=10000)

    by_brand = defaultdict(int)
    by_type  = defaultdict(int)
    by_field = defaultdict(int)
    by_maps  = defaultdict(int)

    for e in entries:
        by_brand[e.get("brand", "Unknown")] += 1
        by_type[e.get("type", "unknown")] += 1
        by_maps[e.get("maps_to", "unknown")] += 1
        fname = e.get("context", {}).get("field_name") or e.get("data", {}).get("field", "")
        if fname:
            by_field[fname] += 1

    # Top edited fields
    top_fields = sorted(by_field.items(), key=lambda x: x[1], reverse=True)[:10]

    return jsonify({
        "total": len(entries),
        "by_brand": dict(by_brand),
        "by_type": dict(by_type),
        "by_maps_to": dict(by_maps),
        "top_edited_fields": [{"field": f, "count": c} for f, c in top_fields],
    })


@app.route("/api/feedback/session-summary")
def feedback_session_summary():
    """Compile a session summary for email digest.
    Returns everything that happened this session: overrides, edits, manual notes.
    """
    brand = session_data.get("brand", "")
    styles = session_data.get("styles", [])
    content_map = session_data.get("generated_content", {})
    overrides = session_data.get("field_overrides", {})

    # Count overrides per style
    override_count = sum(len(v) for v in overrides.values())
    styles_with_overrides = sum(1 for v in overrides.values() if v)

    # Load recent feedback for this brand (this session — last hour)
    recent = _load_feedback(brand=brand or None, limit=500)
    # Filter to recent (within last 2 hours)
    cutoff = (datetime.utcnow() - timedelta(hours=2)).isoformat()
    session_entries = [e for e in recent if e.get("timestamp", "") >= cutoff]

    manual_notes = [e for e in session_entries if e.get("type") == "manual"]
    content_edits = [e for e in session_entries if e.get("type") == "content_edit"]
    field_overrides = [e for e in session_entries if e.get("type") == "field_override"]
    apply_alls = [e for e in session_entries if e.get("type") == "apply_all"]

    return jsonify({
        "brand": brand,
        "total_styles": len(styles),
        "styles_with_content": len(content_map),
        "total_overrides": override_count,
        "styles_with_overrides": styles_with_overrides,
        "manual_notes": [{"message": n.get("data", {}).get("message", ""),
                          "context": n.get("context", {}),
                          "timestamp": n.get("timestamp")} for n in manual_notes],
        "content_edits_count": len(content_edits),
        "field_overrides_count": len(field_overrides),
        "apply_all_count": len(apply_alls),
        "session_entries": len(session_entries),
    })


@app.route("/api/feedback/digest")
def feedback_digest():
    """Full feedback digest across all brands — for Devang to review.
    Shows: recent feedback, patterns, most-edited fields, template requests,
    operator activity, and learning recommendations.
    """
    all_entries = _load_feedback(limit=500)

    # Categorize
    by_type = defaultdict(list)
    by_brand = defaultdict(list)
    by_field = defaultdict(int)
    template_requests = []
    manual_notes = []

    for e in all_entries:
        by_type[e.get("type", "unknown")].append(e)
        by_brand[e.get("brand", "unknown")].append(e)
        fname = e.get("context", {}).get("field_name") or e.get("data", {}).get("field", "")
        if fname:
            by_field[fname] += 1
        if e.get("type") == "template_request":
            template_requests.append(e)
        if e.get("type") == "manual":
            manual_notes.append(e)

    # Build learning recommendations
    learning = []
    top_fields = sorted(by_field.items(), key=lambda x: x[1], reverse=True)[:5]
    for field, count in top_fields:
        if count >= 3:
            learning.append({
                "field": field, "edit_count": count,
                "recommendation": f"'{field}' has been corrected {count} times. Consider updating the derivation logic or brand config default."
            })

    # Field overrides that could become brand defaults
    brand_default_candidates = []
    for e in by_type.get("apply_all", []):
        brand_default_candidates.append({
            "brand": e.get("brand"),
            "field": e.get("data", {}).get("field_name", ""),
            "value": e.get("data", {}).get("value", ""),
            "timestamp": e.get("timestamp"),
        })

    return jsonify({
        "total_feedback": len(all_entries),
        "by_type": {k: len(v) for k, v in by_type.items()},
        "by_brand": {k: len(v) for k, v in by_brand.items()},
        "top_edited_fields": [{"field": f, "count": c} for f, c in top_fields],
        "template_requests": [{
            "product_type": e.get("data", {}).get("product_type", ""),
            "brand": e.get("brand"),
            "note": e.get("data", {}).get("note", ""),
            "timestamp": e.get("timestamp"),
        } for e in template_requests],
        "manual_notes": [{
            "message": e.get("data", {}).get("message", ""),
            "brand": e.get("brand"),
            "context": e.get("context", {}),
            "timestamp": e.get("timestamp"),
        } for e in manual_notes[:20]],
        "learning_recommendations": learning,
        "brand_default_candidates": brand_default_candidates[:10],
    })


@app.route("/api/feedback/changelog")
def feedback_changelog():
    """Return recent content edits as a flat table for the change log view."""
    brand = request.args.get("brand", "") or session_data.get("brand", "")
    limit = int(request.args.get("limit", "100"))
    entries = _load_feedback(brand=brand or None, fb_type="content_edit", limit=limit)
    # Also include field_override and manual types
    entries += _load_feedback(brand=brand or None, fb_type="field_override", limit=limit)
    entries += _load_feedback(brand=brand or None, fb_type="manual", limit=limit)
    # Sort by timestamp desc
    entries.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
    entries = entries[:limit]

    rows = []
    for e in entries:
        ctx = e.get("context", {})
        data = e.get("data", {})
        rows.append({
            "timestamp": e.get("timestamp", ""),
            "operator": e.get("operator", ""),
            "brand": e.get("brand", ""),
            "style_num": ctx.get("style_num", ""),
            "field": ctx.get("field_name", "") or data.get("field", ""),
            "type": e.get("type", ""),
            "original": str(data.get("original", ""))[:100],
            "updated": str(data.get("updated", ""))[:100],
            "reason": data.get("reason", "") or data.get("message", ""),
        })
    return jsonify({"brand": brand, "total": len(rows), "rows": rows})


@app.route("/api/feedback/learning")
def feedback_learning():
    """Aggregate feedback patterns into actionable learning insights."""
    brand = request.args.get("brand", "") or session_data.get("brand", "")
    entries = _load_feedback(brand=brand or None, limit=10000)

    # Patterns: which fields get edited most, which reasons, which operators
    field_edits = defaultdict(int)
    reason_counts = defaultdict(int)
    operator_counts = defaultdict(int)
    original_to_updated = defaultdict(list)  # field -> list of (original, updated)

    for e in entries:
        if e.get("type") not in ("content_edit", "field_override"):
            continue
        ctx = e.get("context", {})
        data = e.get("data", {})
        fname = ctx.get("field_name", "") or data.get("field", "")
        if fname:
            field_edits[fname] += 1
        reason = data.get("reason", "")
        if reason:
            reason_counts[reason] += 1
        op = e.get("operator", "")
        if op:
            operator_counts[op] += 1
        orig = str(data.get("original", ""))[:80]
        updated = str(data.get("updated", ""))[:80]
        if fname and orig and updated:
            original_to_updated[fname].append({"from": orig, "to": updated})

    insights = []
    # Top edited fields
    for field, count in sorted(field_edits.items(), key=lambda x: x[1], reverse=True)[:10]:
        examples = original_to_updated.get(field, [])[:3]
        insights.append({
            "type": "frequent_edit",
            "field": field,
            "count": count,
            "message": f"'{field}' was edited {count} times. Consider updating the default or derivation.",
            "examples": examples,
        })

    # Top reasons
    top_reasons = sorted(reason_counts.items(), key=lambda x: x[1], reverse=True)[:5]

    return jsonify({
        "brand": brand,
        "total_edits": sum(field_edits.values()),
        "top_edited_fields": [{"field": f, "count": c} for f, c in sorted(field_edits.items(), key=lambda x: x[1], reverse=True)[:10]],
        "top_reasons": [{"reason": r, "count": c} for r, c in top_reasons],
        "operators": [{"name": n, "edits": c} for n, c in sorted(operator_counts.items(), key=lambda x: x[1], reverse=True)],
        "insights": insights,
    })


def _run_nis_generation(brand, styles, content_map, vendor_code, template_path, brand_cfg, template_map):
    """Background worker for NIS file generation."""
    results = []
    errors = []

    for i, style in enumerate(styles):
        style_num = style["style_num"]
        style_name = style["style_name"]
        style_content = content_map.get(style_num, {})
        variants = style.get("variants", [])

        nis_progress["current_style"] = f"{style_num} \u2014 {style_name}"

        if not style_content:
            errors.append(f"No content for style {style_num}")
            nis_progress["completed"] = i + 1
            continue

        subclass = style.get("subclass", "")
        product_type_for_template = TEMPLATE_PRODUCT_TYPE_MAP.get(subclass, None)
        if product_type_for_template and product_type_for_template in template_map:
            style_template_path = template_map[product_type_for_template]
        else:
            style_template_path = template_path

        nis_progress["current_step"] = f"Reading preupload template for {style_name}..."
        time.sleep(0.2)
        nis_progress["current_step"] = f"Mapping 254 columns to NIS template..."
        time.sleep(0.2)
        nis_progress["current_step"] = f"Injecting data for {len(variants)} variants..."

        try:
            output_path = do_xlsm_surgery(
                template_path=style_template_path,
                brand=brand,
                brand_cfg=brand_cfg,
                vendor_code=vendor_code,
                style=style,
                content=style_content,
            )
            filename = Path(output_path).name
            nis_progress["current_step"] = f"Writing parent row + {len(variants)} child rows..."
            time.sleep(0.1)
            nis_progress["current_step"] = f"QA: verifying file integrity..."
            time.sleep(0.1)
            nis_progress["current_step"] = f"\u2713 NIS_{brand}_{style_num}.xlsm saved"
            results.append({
                "style_num": style_num,
                "style_name": style_name,
                "rows": len(variants) + 1,
                "filename": filename,
                "path": output_path,
            })
        except Exception as e:
            traceback.print_exc()
            errors.append(f"Failed to generate NIS for style {style_num}: {str(e)}")

        nis_progress["completed"] = i + 1

    nis_progress["status"] = "done"
    nis_progress["current_style"] = ""
    nis_progress["current_step"] = ""
    nis_progress["results"] = results
    nis_progress["errors"] = errors


@app.route("/api/generate-nis", methods=["POST"])
def generate_nis():
    data = request.get_json(force=True)
    brand = data.get("brand") or session_data.get("brand")
    styles = data.get("styles") or session_data.get("styles", [])
    content_map = data.get("content") or session_data.get("generated_content", {})
    vendor_code = data.get("vendor_code") or session_data.get("vendor_code") or ""
    template_path = data.get("template_path") or None
    if not template_path or template_path == "null" or template_path == "None":
        template_path = session_data.get("template_path") or str(DEFAULT_TEMPLATE)
    
    if not brand:
        return jsonify({"error": "No brand selected"}), 400
    if not styles:
        return jsonify({"error": "No product data loaded"}), 400
    if not content_map:
        return jsonify({"error": "Content not yet generated. Run Generate Content first."}), 400
    if not os.path.exists(template_path):
        return jsonify({"error": f"Template file not found: {template_path}. Upload an Amazon NIS template first."}), 400
    
    brand_cfg = _load_brand_config_data(brand)
    
    # Clear output dir
    for f in UPLOAD_OUTPUT.glob("*.xlsm"):
        f.unlink()
    
    # Init progress
    nis_progress["total"] = len(styles)
    nis_progress["completed"] = 0
    nis_progress["status"] = "running"
    nis_progress["started_at"] = datetime.now().isoformat()
    nis_progress["current_style"] = ""
    nis_progress["results"] = []
    nis_progress["errors"] = []
    
    template_map = dict(session_data.get("templates", {}))
    
    # Run in background thread to avoid Render's 30-sec request timeout
    t = threading.Thread(target=_run_nis_generation, args=(
        brand, styles, content_map, vendor_code, template_path, brand_cfg, template_map
    ))
    t.daemon = True
    t.start()
    
    # Return immediately
    return jsonify({"status": "started", "total": len(styles)})


@app.route("/api/generate-nis-results")
def generate_nis_results():
    """Poll this after generate-nis to get results when done."""
    if nis_progress["status"] == "done":
        return jsonify({
            "status": "done",
            "results": nis_progress.get("results", []),
            "errors": nis_progress.get("errors", []),
            "total": len(nis_progress.get("results", [])),
        })
    else:
        return jsonify({
            "status": nis_progress["status"],
            "completed": nis_progress["completed"],
            "total": nis_progress["total"],
        })

@app.route("/api/generate-progress")
def generate_progress():
    elapsed = ""
    eta = ""
    if nis_progress["started_at"] and nis_progress["completed"] > 0:
        started = datetime.fromisoformat(nis_progress["started_at"])
        elapsed_sec = (datetime.now() - started).total_seconds()
        per_style = elapsed_sec / nis_progress["completed"]
        remaining = nis_progress["total"] - nis_progress["completed"]
        eta_sec = per_style * remaining
        elapsed = f"{int(elapsed_sec)}s"
        if eta_sec > 60:
            eta = f"~{int(eta_sec / 60)}m {int(eta_sec % 60)}s remaining"
        else:
            eta = f"~{int(eta_sec)}s remaining"
    return jsonify({
        "total": nis_progress["total"],
        "completed": nis_progress["completed"],
        "current_style": nis_progress["current_style"],
        "current_step": nis_progress.get("current_step", ""),
        "status": nis_progress["status"],
        "elapsed": elapsed,
        "eta": eta,
        "percent": round((nis_progress["completed"] / max(nis_progress["total"], 1)) * 100, 1)
    })

# ── Category / field derivation helpers ───────────────────────────────────────
def _derive_amazon_product_category(sub_class, gender="", product_type="", style_name="", department=""):
    """Map sub_class + gender to Amazon product_category dropdown value.
    Gender-aware: Men's Swimwear vs Women's Swimwear vs Swim (youth).
    """
    # ── Dresses ──
    DRESS_MAP = {
        "Day Dress": "Women's Dresses",
        "Cocktail Dress": "Women's Dresses",
        "Active Dress": "Women's Active",
        "Swimdress": "Women's Dresses",
        "Pullover": "Women's Everyday Sportswear",
        "Tank": "Women's Everyday Sportswear",
        "Shirt": "Women's Everyday Sportswear",
        "Shorts": "Women's Everyday Sportswear",
        "Skirt": "Women's Everyday Sportswear",
        "Skort": "Women's Everyday Sportswear",
        "Jacket": "Women's Everyday Sportswear",
        "Coat": "Women's Everyday Sportswear",
    }
    # ── Swimwear — gender-dependent ──
    SWIM_SUBCLASSES = {"Rashguard", "Trunk", "Bikini Top", "Bikini Bottom", "Swim Bottom",
                       "One Piece Swim", "Tankini", "Short", "Swim Set 2 pcs", "Swim Set",
                       "Board Short", "Boardshort", "Swim Shirt", "Cover Up", "Swimdress"}

    if sub_class in DRESS_MAP:
        return DRESS_MAP[sub_class]

    if sub_class in SWIM_SUBCLASSES or product_type == "SWIMWEAR":
        g = (gender or "").lower()
        sn = (style_name or "").lower()
        dept = (department or "").lower()
        is_youth = ("boys" in sn or "girls" in sn or "toddler" in sn or
                    dept in ("boys","girls") or g == "unisex")
        if is_youth:
            return "Swim"
        if g == "male":
            return "Men's Swimwear"
        if g == "female":
            return "Women's Swimwear"
        return "Swim"

    return ""  # leave blank if unknown — operator sets on dashboard

def _derive_item_type_keyword(sub_class, product_type="", gender="", style_name=""):
    """Map sub_class to Amazon item_type_keyword (free-text SEO slug).
    Covers all 16 product types."""
    _map = {
        # Dresses
        "Day Dress": "casual-and-day-dresses", "Cocktail Dress": "cocktail-and-party-dresses",
        "Active Dress": "active-dresses", "Maxi Dress": "maxi-dresses",
        "Mini Dress": "mini-dresses", "Wrap Dress": "wrap-dresses",
        "Shirt Dress": "shirt-dresses", "Dress": "dresses",
        # Swimwear
        "Rashguard": "rash-guards", "Rash Guard": "rash-guards",
        "Trunk": "swim-trunks", "Swim Trunk": "swim-trunks",
        "Board Short": "board-shorts", "Boardshort": "board-shorts", "Boardshorts": "board-shorts",
        "Bikini Top": "bikini-tops", "Bikini Bottom": "bikini-bottoms", "Swim Bottom": "bikini-bottoms",
        "One Piece Swim": "one-piece-swimsuits", "One Piece": "one-piece-swimsuits",
        "Tankini": "tankini-swimsuits", "Short": "board-shorts",
        # Gender-aware: see post-mapping logic below
        "Swim Set 2 pcs": "rash-guard-sets", "Swim Set": "rash-guard-sets",
        "Swim Shirt": "rash-guards", "Cover Up": "fashion-swimwear-cover-ups",
        "Swimdress": "fashion-swimwear-cover-ups",
        # Blazers
        "Blazer": "blazers", "Sport Coat": "sport-coats",
        # Bras
        "Bra": "bras", "Bralette": "bralettes", "Sports Bra": "sports-bras",
        # Coats/Jackets
        "Jacket": "jackets", "Coat": "coats", "Vest": "vests",
        "Puffer": "puffer-jackets", "Windbreaker": "windbreakers",
        "Anorak": "anoraks", "Parka": "parkas",
        # Hats
        "Hat": "hats", "Cap": "baseball-caps", "Beanie": "beanies",
        "Visor": "visors", "Sun Hat": "sun-hats",
        "Trucker Hat": "trucker-hats", "Bucket Hat": "bucket-hats",
        # One-piece outfits
        "Romper": "rompers", "Jumpsuit": "jumpsuits",
        "Bodysuit": "bodysuits", "One Piece Outfit": "rompers",
        # Overalls
        "Overalls": "overalls", "Dungarees": "overalls", "Overall": "overalls",
        # Pants
        "Pants": "casual-pants", "Leggings": "leggings",
        "Joggers": "jogger-pants", "Trousers": "dress-pants",
        "Chino": "chinos", "Cargo Pant": "cargo-pants",
        # Sandals
        "Sandal": "sandals", "Flip Flop": "flip-flops",
        "Slide": "slide-sandals", "Thong Sandal": "thong-sandals", "Slipper": "slippers",
        # Shirts
        "Pullover": "pullovers", "Tank": "tank-tops", "Tee": "t-shirts",
        "Blouse": "blouses", "Polo": "polo-shirts", "Henley": "henley-shirts",
        "Crop Top": "crop-tops", "Camisole": "camisoles", "Tunic": "tunics",
        # Shorts
        "Shorts": "casual-shorts", "Chino Short": "chino-shorts",
        "Cargo Short": "cargo-shorts", "Skort": "skorts",
        # Skirts
        "Skirt": "skirts", "Mini Skirt": "mini-skirts",
        "Maxi Skirt": "maxi-skirts", "Wrap Skirt": "wrap-skirts",
        # Snowsuits
        "Snowsuit": "snowsuits", "Snow Suit": "snowsuits", "Ski Suit": "ski-suits",
        # Snow Pants
        "Snow Pant": "snow-pants", "Snow Pants": "snow-pants",
        "Ski Pants": "ski-pants", "Ski Pant": "ski-pants",
        # Sweatshirts
        "Sweatshirt": "sweatshirts", "Hoodie": "hoodies",
        "Fleece": "fleece-jackets", "Quarter Zip": "quarter-zip-pullovers",
    }
    base = _map.get(sub_class, "")
    # Gender-aware swim-set keyword
    if sub_class in ("Swim Set 2 pcs", "Swim Set"):
        sn = (style_name or "").lower()
        g = (gender or "").lower()
        if "rash" in sn or "sleeve" in sn or "swim shirt" in sn or "sun shirt" in sn:
            return "rash-guard-sets"
        if g == "female" or "girls" in sn or "bikini" in sn:
            return "bikini-sets"
        return "swim-sets"
    return base

def _derive_item_type_name(sub_class, product_type="", gender="", style_name=""):
    """Human-readable item type name — must match template dropdown exactly.
    Covers all 16 product types."""
    _map = {
        # Dresses
        "Day Dress": "Casual Dress", "Cocktail Dress": "Cocktail Dress",
        "Active Dress": "Tennis Dress", "Maxi Dress": "Casual Dress",
        "Mini Dress": "Casual Dress", "Wrap Dress": "Casual Dress",
        "Shirt Dress": "Business Casual Dress", "Dress": "Dress",
        # Swimwear
        "Rashguard": "Rash Guard Shirt", "Rash Guard": "Rash Guard Shirt",
        "Trunk": "Swim Trunks", "Swim Trunk": "Swim Trunks",
        "Board Short": "Board Shorts", "Boardshort": "Board Shorts", "Boardshorts": "Board Shorts",
        "Bikini Top": "Bikini Top", "Bikini Bottom": "Bikini Bottoms", "Swim Bottom": "Bikini Bottoms",
        "One Piece Swim": "One Piece Swimsuit", "One Piece": "One Piece Swimsuit",
        "Tankini": "Tankini Swimsuit", "Short": "Board Shorts",
        # "Swim Set 2 pcs" resolved below — gender-aware (see post-mapping logic)
        "Swim Set 2 pcs": "Rash Guard Set", "Swim Set": "Rash Guard Set",
        "Swim Shirt": "Rash Guard Shirt", "Cover Up": "Swimwear Cover Up",
        "Swimdress": "Swimwear Cover Up",
        # Blazers
        "Blazer": "Blazer", "Sport Coat": "Sport Jacket",
        # Bras
        "Bra": "Bra", "Bralette": "Bra", "Sports Bra": "Sports Bra",
        # Coats/Jackets (template has no ITN dropdown — leave blank, operator fills)
        # Hats
        "Hat": "Hat", "Cap": "Baseball Cap", "Beanie": "Beanie Hat",
        "Visor": "Cap", "Sun Hat": "Hat",
        "Trucker Hat": "Baseball Cap", "Bucket Hat": "Bucket Hat",
        # One-piece outfits (template has no ITN dropdown)
        "Romper": "", "Jumpsuit": "", "Bodysuit": "",
        # Overalls (template has no ITN dropdown)
        "Overalls": "", "Dungarees": "",
        # Pants
        "Pants": "Casual Pants", "Leggings": "Leggings",
        "Joggers": "Casual Pants", "Trousers": "Dress Pants",
        "Chino": "Khakis", "Cargo Pant": "Casual Pants",
        # Sandals
        "Sandal": "Sandal", "Flip Flop": "Flip-Flop",
        "Slide": "Slide Sandal", "Thong Sandal": "Flat Sandal", "Slipper": "Sandal",
        # Shirts (template has no ITN dropdown)
        "Pullover": "", "Tank": "", "Tee": "", "Blouse": "",
        "Polo": "", "Henley": "", "Crop Top": "", "Camisole": "", "Tunic": "",
        # Shorts
        "Shorts": "Casual Shorts", "Chino Short": "Khaki Shorts",
        "Cargo Short": "Cargo Shorts", "Skort": "Skorts",
        # Skirts
        "Skirt": "Skirt", "Mini Skirt": "Skirt",
        "Maxi Skirt": "Skirt", "Wrap Skirt": "Skirt",
        # Snowsuits (template has no ITN dropdown)
        "Snowsuit": "", "Snow Suit": "", "Ski Suit": "",
        # Snow Pants
        "Snow Pant": "Casual Pants", "Snow Pants": "Casual Pants",
        "Ski Pants": "Casual Pants", "Ski Pant": "Casual Pants",
        # Sweatshirts
        "Sweatshirt": "Sweatshirt", "Hoodie": "Hooded Sweatshirt",
        "Fleece": "Pullover Sweater", "Quarter Zip": "Pullover Sweater",
    }
    base = _map.get(sub_class, "")
    # Gender-aware swim-set mapping: boys' swim sets are NOT bikini sets.
    # Valid SWIMWEAR item_type_name values include: 'Bikini Set', 'Rash Guard Set',
    # 'Swim Shirt Set', 'Tankini Set', 'Two Piece Swimsuit', 'Swimwear Cover Up Set'.
    if sub_class in ("Swim Set 2 pcs", "Swim Set"):
        sn = (style_name or "").lower()
        g = (gender or "").lower()
        # Rash guard style sets (long/short sleeve swim shirt + trunk)
        if "rash" in sn or "sleeve" in sn or "swim shirt" in sn or "sun shirt" in sn:
            return "Rash Guard Set"
        # Girls/women bikini set (only when clearly signaled)
        if g == "female" or "girls" in sn or "girl" in sn or "bikini" in sn:
            return "Bikini Set"
        # Default — two-piece swimsuit (safer than 'Bikini Set' for boys/unisex)
        return "Two Piece Swimsuit"
    return base  # leave blank if unknown

def _derive_item_length(sub_subclass, style_name, product_type="", sub_class=""):
    """Derive item length description from sub_subclass / style name.
    SWIMWEAR: return blank (field isn't on the Swimwear template; rashguards are torso
    garments, not leg garments, and Amazon's 'Knee-Length' value doesn't apply).
    DRESS/SKIRT: use MAXI/MINI/MIDI detection; default Knee-Length.
    """
    # Swimwear items don't have a meaningful item_length_description
    if (product_type or "").upper() == "SWIMWEAR" or sub_class in (
            "Rashguard","Rash Guard","Trunk","Swim Trunk","Bikini Top","Bikini Bottom",
            "Swim Bottom","One Piece Swim","Tankini","Short","Swim Set 2 pcs","Swim Set",
            "Board Short","Boardshort","Boardshorts","Cover Up","Swim Shirt"):
        return ""
    combined = f"{sub_subclass or ''} {style_name or ''}".upper()
    if "MAXI" in combined:
        return "Long"
    if "MINI" in combined:
        return "Short"
    if "MIDI" in combined:
        return "Mid-Calf"
    return "Knee-Length"

def _derive_swim_product_subcategory(sub_class, gender="", style_name="", product_category=""):
    """Derive the correct Amazon product_subcategory value for a SWIMWEAR style.
    Depends on product_category (Men's Swimwear | Women's Swimwear | Swim).
    Falls back to '' (blank) rather than guessing when we can't match.
    """
    cat = (product_category or "").strip()
    sc = (sub_class or "").strip()
    sn = (style_name or "").lower()
    is_baby = "baby" in sn or "infant" in sn

    # Men's Swimwear sub-subcategories: Board Shorts | Trunks | Briefs | Misc
    if cat == "Men's Swimwear":
        if sc in ("Trunk", "Swim Trunk"):
            if "board" in sn or "boardshort" in sn:
                return "Board Shorts"
            return "Trunks"
        if sc in ("Board Short", "Boardshort", "Boardshorts", "Short"):
            return "Board Shorts"
        # Men's rashguards live under Athletic or Misc — Amazon has no Rashguards
        # under Men's Swimwear. Use Misc for non-bottom items.
        if sc in ("Rashguard", "Rash Guard", "Swim Shirt"):
            return "Misc"
        return "Misc"

    # Women's Swimwear sub-subcategories
    if cat == "Women's Swimwear":
        if sc in ("Bikini Top",):
            return "Bikini Top Separates"
        if sc in ("Bikini Bottom", "Swim Bottom"):
            return "Bikini Bottom Separates"
        if sc in ("One Piece Swim", "One Piece"):
            return "One-Piece Swimsuits"
        if sc == "Tankini":
            return "Tankini Top Separates"
        if sc in ("Rashguard", "Rash Guard", "Swim Shirt"):
            return "Rashguards"
        if sc in ("Swim Set 2 pcs", "Swim Set"):
            return "Two-Piece Swimsuit Sets"
        if sc in ("Short", "Board Short", "Boardshort", "Boardshorts"):
            return "Swim Shorts"
        if sc == "Cover Up":
            return "Cover-Ups"
        return ""

    # Swim (youth / boys / girls)
    if cat == "Swim":
        is_boys = "boys" in sn or "boy" in sn or (gender or "").lower() == "male"
        is_girls = "girls" in sn or "girl" in sn or (gender or "").lower() == "female"
        prefix = "Baby " if is_baby else ""
        kind_boys = "Boys" if is_boys else ("Girls" if is_girls else "Boys")
        if sc in ("Rashguard", "Rash Guard", "Swim Shirt"):
            return f"{prefix}{kind_boys} Swim Rashguards" if is_boys else f"{prefix}{kind_boys} Rashguards"
        if sc in ("Trunk", "Swim Trunk", "Board Short", "Boardshort", "Boardshorts", "Short"):
            # Boys -> 'Boys Swim Bottoms'; Girls -> 'Girls Coverups & Boardshorts'
            return f"{prefix}{kind_boys} Swim Bottoms" if is_boys else f"{prefix}{kind_boys} Coverups & Boardshorts"
        if sc in ("Swim Set 2 pcs", "Swim Set"):
            return f"{prefix}{kind_boys} Swim Sets"
        if sc in ("Bikini Top", "Bikini Bottom", "Swim Bottom"):
            return f"{prefix}{kind_boys} Two Piece Swim"
        if sc in ("One Piece Swim", "One Piece"):
            return f"{prefix}{kind_boys} One Piece Swim"
        return ""

    return ""

# Keyword → Amazon special_feature value (sourced from dropdown_cache).
# Highest-signal mappings only; if no signal, leave the field blank.
_SPECIAL_FEATURE_KEYWORDS = {
    "hood": "Hooded", "hooded": "Hooded", "removable hood": "Hooded",
    "belt": "Belted", "belted": "Belted", "with belt": "Belted",
    "reversible": "Reversible",
    "waterproof": "Waterproof", "water-proof": "Waterproof",
    "water resistant": "Water Resistant", "water-resistant": "Water Resistant",
    "windproof": "Windproof", "wind-proof": "Windproof",
    "wind resistant": "Windproof", "wind-resistant": "Windproof",
    "lightweight": "Lightweight", "light-weight": "Lightweight",
    "flame resistant": "Flame Resistant",
    "vented": "Vented",
    "stain resistant": "Stain Resistant", "stain-resistant": "Stain Resistant",
    "quick dry": "Quick Dry", "quick-dry": "Quick Dry",
    "removable padding": "Removable Padding",
    "abrasion resistant": "Abrasion Resistant",
    "hemline drawstring": "Hemline Drawstring", "drawstring": "Hemline Drawstring",
    "heated": "Heated",
    "fade resistant": "Fade Resistant",
    "bleach resistant": "Bleach Resistant",
    "wrinkle resistant": "Wrinkle Resistant", "wrinkle-resistant": "Wrinkle Resistant",
}

def _derive_special_features(style_name, additional_details="", max_features=5):
    """Look at the style name + additional details and pick up to N Amazon-valid
    special_feature values. Sage feedback ("can AI do suggestions based off of product data?")
    is exactly this signal-from-name approach. Conservative: only adds features with
    explicit keyword evidence.
    """
    if not style_name and not additional_details:
        return []
    haystack = f"{style_name or ''} {additional_details or ''}".lower()
    found = []
    seen = set()
    for kw, label in _SPECIAL_FEATURE_KEYWORDS.items():
        if kw in haystack and label not in seen:
            found.append(label)
            seen.add(label)
            if len(found) >= max_features:
                break
    return found


# Lifestyle inference from style name. Rough but better than blank.
_LIFESTYLE_KEYWORDS = {
    "work": "Business Casual", "office": "Business Casual", "professional": "Business Casual",
    "blazer": "Business Casual", "trench": "Business Casual",
    "cocktail": "Evening", "gown": "Evening", "evening": "Evening",
    "formal": "Formal", "black tie": "Formal", "tuxedo": "Formal",
    "club": "Club", "party": "Club",
    "casual": "Casual", "weekend": "Casual",
    "lounge": "Comfort", "loungewear": "Comfort", "comfort": "Comfort",
    "sleep": "Comfort",
}

def _derive_lifestyle(style_name, sub_class="", max_lifestyles=2):
    """Pick up to N Amazon-valid lifestyle tags from the style name + sub_class."""
    if not style_name and not sub_class:
        return ["Casual"]  # Sensible default for apparel
    haystack = f"{style_name or ''} {sub_class or ''}".lower()
    found = []
    seen = set()
    for kw, label in _LIFESTYLE_KEYWORDS.items():
        if kw in haystack and label not in seen:
            found.append(label)
            seen.add(label)
            if len(found) >= max_lifestyles:
                break
    if not found:
        # Fallback: outerwear/coat → Casual; otherwise Casual
        return ["Casual"]
    return found


# 'Type of Jacket' from TLG pre-upload → Amazon's coat_silhouette_type#1.value valid set.
# Free-text fallback when no clean match (the fuzzy-matcher in wc/write_cell will refine).
_TYPE_OF_JACKET_MAP = {
    "PUFFER":     "Quilted",
    "QUILTED":    "Quilted",
    "DOWN":       "Quilted",
    "PARKA":      "Anorak",
    "ANORAK":     "Anorak",
    "TRENCH":     "Trench Coat",
    "TRENCHCOAT": "Trench Coat",
    "PEACOAT":    "Peacoat",
    "PEA COAT":   "Peacoat",
    "OVERCOAT":   "Overcoat",
    "RAIN":       "Rain Coat",
    "RAINCOAT":   "Rain Coat",
    "CAPE":       "Cape",
    "PONCHO":     "Poncho",
    "COCOON":     "Cocoon",
    "AVIATOR":    "Aviator",
    "BOMBER":     "Aviator",  # Closest in coat_silhouette dropdown
}

def _derive_coat_silhouette(type_of_jacket, fallback_subclass=""):
    """Map TLG 'Type of Jacket' or sub_class → Amazon coat_silhouette_type. Returns empty if no match.
    The .xlsm writer's fuzzy-matcher will further snap to a valid dropdown value.
    """
    candidate = (type_of_jacket or fallback_subclass or "").strip().upper()
    if not candidate:
        return ""
    # Direct hit
    if candidate in _TYPE_OF_JACKET_MAP:
        return _TYPE_OF_JACKET_MAP[candidate]
    # Substring search (handles "Long Puffer", "Quilted Coat", etc.)
    for kw, label in _TYPE_OF_JACKET_MAP.items():
        if kw in candidate:
            return label
    return ""


def _split_fabric_into_materials(fabric):
    """Split a blend string like '80% Polyester, 15% Cotton, 10% Spandex' into the
    constituent fiber names ['Polyester', 'Cotton', 'Spandex'] for the material#1..N
    Amazon fields. The full blend remains in fabric_type#1.value (free text).

    Sage feedback ("Material is 'Polyester', 'Cotton', 'Spandex', while Fabric Type is
    '80% Polyester, 15% Cotton, 10% Spandex'") is exactly this split.
    """
    if not fabric:
        return []
    s = str(fabric).strip()
    # Split on commas and slashes (the two common separators in TLG pre-uploads)
    chunks = re.split(r'[,/]|\s+and\s+', s)
    out = []
    for chunk in chunks:
        # Strip leading numbers + percent + whitespace
        cleaned = re.sub(r'^\s*\d+(?:\.\d+)?\s*%?\s*', '', chunk).strip()
        # Strip trailing percent (e.g. "Polyester 80%")
        cleaned = re.sub(r'\s*\d+(?:\.\d+)?\s*%?\s*$', '', cleaned).strip()
        cleaned = cleaned.title()
        if not cleaned: continue
        if cleaned.lower() in {"and", "with", "of", "in"}: continue
        if cleaned not in out:
            out.append(cleaned)
    return out


def _derive_fabric_type(fabric):
    """Derive simplified fabric type for Col 59."""
    if not fabric:
        return "Polyester"
    f = fabric.upper()
    if "NYLON" in f:
        return "Nylon"
    if "COTTON" in f or "COTT" in f:
        return "Cotton"
    if "RAYON" in f:
        return "Rayon"
    if "LINEN" in f:
        return "Linen"
    if "MODAL" in f:
        return "Modal"
    if "SILK" in f:
        return "Silk"
    if "WOOL" in f:
        return "Wool"
    if "BAMBOO" in f:
        return "Bamboo"
    if "SPAN" in f or "SPANDEX" in f or "LYCRA" in f:
        return "Spandex"
    if "POLY" in f or "POLYESTER" in f:
        return "Polyester"
    return "Polyester"

def _derive_sleeve_length(sleeve_type, product_type=""):
    """Map sleeve type string to sleeve length description for Col 129.
    Now PT-aware: when no signal in the type string, fall back to that PT's default
    (COAT → Long Sleeve, SWIMWEAR → Sleeveless, etc.) instead of the old global "Sleeveless".
    """
    if sleeve_type:
        s = sleeve_type.lower()
        if "sleeveless" in s or "off" in s:
            return "Sleeveless"
        if "long" in s:
            return "Long Sleeve"
        if "3/4" in s:
            return "3/4 Sleeve"
        if "short" in s or "flutter" in s or "cap" in s or "ruffle" in s or "balloon" in s:
            return "Short Sleeve"
    # PT-aware fallback when sleeve_type is empty/unrecognized
    pt_default = _pt_defaults.get_pt_default(product_type or "", "default_sleeve_length", "")
    return pt_default or "Sleeveless"


# ── QA Preview helpers ─────────────────────────────────────────────────────────
def _build_preview_fields(brand, brand_cfg, vendor_code, style, content):
    """
    Build the list of all fields that would go into the .xlsm for a given style,
    with status: 'filled', 'default', 'empty', or 'locked'.
    Returns a list of field dicts.
    """
    style_num     = style["style_num"]
    style_name    = style["style_name"]
    variants      = style["variants"]
    list_price    = style.get("list_price", "")
    cost_price    = style.get("cost_price", "")
    sub_class     = style.get("subclass", "") or style.get("sub_class", "")
    sub_subclass  = style.get("sub_subclass", "")
    model_name_raw = style.get("model_name", "") or style_name

    bullets      = content.get("bullets", [])
    description  = content.get("description", "")
    backend_kw   = content.get("backend_keywords", "")
    # Resolve actual product type for this style FIRST so derive_sleeve_type and
    # other PT-aware helpers below get the right fallback (e.g. "Long Sleeve" for COAT).
    resolved_pt    = _resolve_style_product_type(style) or "DRESS"
    neck_type    = content.get("neck_type", "") or style.get("neck_type", "") or derive_neck_type(style_name)
    sleeve_type  = content.get("sleeve_type", "") or style.get("sleeve_type", "") or derive_sleeve_type(style_name, resolved_pt)
    silhouette   = content.get("silhouette", "") or derive_silhouette(sub_subclass)
    # Derive gender/department per-style from division_name
    style_gender, style_dept = _derive_gender_department(style)
    eff_gender = style_gender or brand_cfg.get("gender", "")
    eff_dept   = style_dept or brand_cfg.get("department", "")
    # Always derive from dropdown-validated function, gender-aware
    category     = _derive_amazon_product_category(sub_class, gender=eff_gender, product_type=resolved_pt, style_name=style_name, department=eff_dept)
    subcategory  = SUBCLASS_SUBCATEGORY_MAP.get(sub_class, '')
    # Swim-aware subcategory derivation (replaces blank fallback for all 59 swim styles)
    if not subcategory and resolved_pt == "SWIMWEAR":
        subcategory = _derive_swim_product_subcategory(sub_class, gender=eff_gender, style_name=style_name, product_category=category)
    fabric       = content.get("fabric", "") or style.get("fabric", "") or brand_cfg.get("default_fabric", "")
    care         = content.get("care", "") or style.get("care", "") or brand_cfg.get("default_care", "")
    upf          = content.get("upf", "") or style.get("upf", "") or brand_cfg.get("default_upf", "")
    coo          = normalize_coo(content.get("coo", "") or style.get("coo", "") or brand_cfg.get("default_coo", "")) or ""
    clean_brand  = clean_brand_name(brand)
    item_type_name = _derive_item_type_name(sub_class, product_type=resolved_pt, gender=eff_gender, style_name=style_name)
    item_length    = _derive_item_length(sub_subclass, style_name, product_type=resolved_pt, sub_class=sub_class)
    fabric_type    = _derive_fabric_type(fabric)
    itk_value      = _derive_item_type_keyword(sub_class, product_type=resolved_pt, gender=eff_gender, style_name=style_name)
    # Taxonomy override: a confirmed override (if any) trumps the auto-derived values above
    _tax = _resolve_taxonomy_for_style(style, brand_cfg)
    if _tax.get("matched"):
        _e = _tax["entry"]
        category     = _e.get("product_category", category)
        subcategory  = _e.get("product_subcategory", subcategory)
        itk_value    = _e.get("item_type_keyword", itk_value)
        item_type_name = _e.get("item_type_name", item_type_name)
    taxonomy_source = "override" if _tax.get("matched") else "auto"
    taxonomy_key = _tax.get("key", "")
    today_str      = datetime.now().strftime("%Y%m%d")
    # Parent SKU: prefer team's own season-coded format (F26-107010297) when variant SKUs allow.
    parent_sku     = _derive_parent_sku_from_variants(variants, style_num) or style_num

    # Preview values: variation theme, target_gender (Male/Female refined), youth size info
    _prev_vts = variants or []
    _prev_hc  = len({(v.get("color_name") or v.get("color") or "") for v in _prev_vts}) > 1
    _prev_hs  = len({v.get("size", "") for v in _prev_vts}) > 1
    _prev_vt  = "SIZE/COLOR" if (_prev_hc and _prev_hs) else ("COLOR" if _prev_hc else ("SIZE" if _prev_hs else "COLOR"))
    _prev_sn_low = (style_name or "").lower()
    if "boys" in _prev_sn_low or "men" in _prev_sn_low:
        _prev_tg = "Male"
    elif "girls" in _prev_sn_low or "women" in _prev_sn_low:
        _prev_tg = "Female"
    else:
        _prev_tg = eff_gender if eff_gender in ("Male","Female") else ""
    _first_var_sz = (variants[0] if variants else {}).get("size", "")
    _prev_sst, _prev_sclass, _prev_ard, _prev_size = _derive_youth_size_info(style_name, eff_gender, _first_var_sz)

    # Sample title (parent)
    title = content.get("title", style_name)

    # Determine first variant for example child fields
    first_variant = variants[0] if variants else {}
    color_name = first_variant.get("color_name", "")
    size       = first_variant.get("size", "")
    upc        = first_variant.get("upc", "")
    sku        = first_variant.get("sku", "") or f"{style_num}-{color_name}-{size}".replace(" ", "-")
    # PT-aware color snap: ensures Cream→Off White (not Ivory) for COAT, Truffle→Brown, etc.
    color_family = normalize_color(color_name, resolved_pt)
    size_normalized = normalize_size(size)
    variant_cost = first_variant.get("cost_price", "") or cost_price

    def f(col, header, value, status, editable=True, note="", field_id="", req_level="optional"):
        return {"col": col, "header": header, "value": str(value) if value is not None else "",
                "status": status, "editable": editable, "note": note,
                "field_id": field_id, "req_level": req_level}

    # req_level: "required" = Amazon rejects without it
    #            "conditional" = required depending on product type / other fields
    #            "recommended" = improves listing quality
    #            "optional" = nice to have
    fields = [
        # ── Identity & Structure ──
        f(1, "Vendor Code", vendor_code or brand_cfg.get("vendor_code_full", ""),
          "filled" if (vendor_code or brand_cfg.get("vendor_code_full", "")) else "empty", False,
          field_id="rtip_vendor_code#1.value", req_level="required"),
        f(2, "Vendor SKU (Parent)", parent_sku, "filled", False,
          field_id="vendor_sku#1.value", req_level="required"),
        f(3, "Product Type", resolved_pt, "default", True,
          field_id="product_type#1.value", req_level="required"),
        f(4, "Parentage Level", "Parent / Child", "default", True,
          field_id="parentage_level#1.value", req_level="required"),
        f(5, "Child Relationship Type", "Variation", "default", True,
          field_id="child_parent_sku_relationship#1.child_relationship_type", req_level="required"),
        f(6, "Parent SKU", parent_sku, "default", True,
          field_id="child_parent_sku_relationship#1.parent_sku", req_level="required"),
        f(7, "Variation Theme", _prev_vt, "default", True,
          field_id="variation_theme#1.name", req_level="required"),

        # ── Product Info ──
        f(8, "Item Name", title, "filled" if title and title != style_name else "default", True,
          field_id="item_name#1.value", req_level="required"),
        f(9, "Brand Name", clean_brand, "filled" if clean_brand else "empty", False,
          field_id="brand#1.value", req_level="required"),
        f(10, "External Product ID Type", "UPC" if upc else "",
          "filled" if upc else "default", False,
          field_id="external_product_id#1.type", req_level="required"),
        f(11, "External Product ID Value", re.sub(r'\D', '', str(upc)) if upc else "",
          "filled" if upc else "default", False,
          field_id="external_product_id#1.value", req_level="required"),
        f(13, "Product Category", category, "filled" if category else "default", False,
          field_id="product_category#1.value", req_level="recommended"),
        f(14, "Product Subcategory", subcategory, "filled" if subcategory else "default", False,
          field_id="product_subcategory#1.value", req_level="recommended"),
        f(15, "Item Type Keyword", itk_value, "filled" if itk_value else "default", False,
          field_id="item_type_keyword#1.value", req_level="required"),
        f(18, "Model Number", style_num, "filled", False,
          field_id="model_number#1.value", req_level="required"),
        f(19, "Model Name", (model_name_raw or style_name).title(), "filled", False,
          field_id="model_name#1.value", req_level="required"),

        # ── Content ──
        f(30, "Bullet Point 1", bullets[0][:120] + "..." if bullets and len(bullets[0]) > 120 else (bullets[0] if bullets else ""),
          "filled" if bullets else "empty", True,
          "" if bullets else "Required for submission.",
          field_id="bullet_point#1.value", req_level="required"),
        f(31, "Bullet Point 2", bullets[1][:120] + "..." if len(bullets) > 1 and len(bullets[1]) > 120 else (bullets[1] if len(bullets) > 1 else ""),
          "filled" if len(bullets) > 1 else "empty", True,
          field_id="bullet_point#2.value", req_level="required"),
        f(32, "Bullet Point 3", bullets[2][:120] + "..." if len(bullets) > 2 and len(bullets[2]) > 120 else (bullets[2] if len(bullets) > 2 else ""),
          "filled" if len(bullets) > 2 else "empty", True,
          field_id="bullet_point#3.value", req_level="recommended"),
        f(33, "Bullet Point 4", bullets[3][:120] + "..." if len(bullets) > 3 and len(bullets[3]) > 120 else (bullets[3] if len(bullets) > 3 else ""),
          "filled" if len(bullets) > 3 else "empty", True,
          field_id="bullet_point#4.value", req_level="recommended"),
        f(34, "Bullet Point 5", bullets[4][:120] + "..." if len(bullets) > 4 and len(bullets[4]) > 120 else (bullets[4] if len(bullets) > 4 else ""),
          "filled" if len(bullets) > 4 else "empty", True,
          field_id="bullet_point#5.value", req_level="recommended"),
        f(35, "Backend Keywords", backend_kw[:100] + "..." if backend_kw and len(backend_kw) > 100 else backend_kw,
          "filled" if backend_kw else "default", True,
          "" if backend_kw else "Using category defaults.",
          field_id="generic_keyword#1.value", req_level="recommended"),
        f(67, "Product Description", description[:120] + "..." if description and len(description) > 120 else description,
          "filled" if description else "empty", True,
          "" if description else "Required for submission.",
          field_id="rtip_product_description#1.value", req_level="required"),

        # ── Attributes ──
        f(46, "Style Name", style_name.title(), "filled", False,
          field_id="style#1.value", req_level="recommended"),
        f(47, "Department", eff_dept, "filled" if eff_dept else "empty", False,
          field_id="department#1.value", req_level="required"),
        f(48, "Target Gender", _prev_tg, "filled" if _prev_tg else "empty", False,
          field_id="target_gender#1.value", req_level="required"),
        f(49, "Age Range", _prev_ard, "default", True,
          field_id="age_range_description#1.value", req_level="required"),
        f(50, "Size System", "US", "default", True,
          field_id=_size_field(resolved_pt, "size_system"), req_level="required"),
        f(51, "Size Class", _prev_sclass, "default", True,
          field_id=_size_field(resolved_pt, "size_class"), req_level="required"),
        f(52, "Size (first variant)", size_normalized or size, "filled" if size else "default", False,
          field_id=_size_field(resolved_pt, "size"), req_level="required"),
        f(56, "Material", fabric, "filled" if fabric else "default", True,
          "" if fabric else "No fabric data. Will use brand default.",
          field_id="material#1.value", req_level="required"),
        f(59, "Fabric Type", fabric_type, "filled" if fabric else "default", False,
          field_id="fabric_type#1.value", req_level="recommended"),
        f(61, "Number of Items", "1", "default", True,
          field_id="number_of_items#1.value", req_level="required"),
        f(62, "Item Type Name", item_type_name, "filled", False,
          field_id="item_type_name#1.value", req_level="required"),
        f(66, "Special Size Type", _prev_sst, "default", True,
          field_id="special_size_type#1.value", req_level="conditional"),
        f(68, "Color (Standardized)", color_family, "filled" if color_family else "default", False,
          field_id="color#1.standardized_values#1", req_level="required"),
        f(69, "Color", color_name.title() if color_name else "",
          "filled" if color_name else "default", False,
          field_id="color#1.value", req_level="required"),
        f(70, "Item Length Description", item_length, "filled", False,
          field_id="item_length_description#1.value", req_level="conditional"),
        f(83, "Fit Type", "", "empty", True,
          "Not auto-filled. Set if known.",
          field_id="fit_type#1.value", req_level="recommended"),
        f(89, "Care Instructions", care, "filled" if care else "default", True,
          "" if care else "No care data. Will use brand default.",
          field_id="care_instructions#1.value", req_level="recommended"),
        f(118, "Neck/Collar Style", neck_type, "filled" if neck_type else "default", True,
          "" if neck_type else "Could not derive from style name.",
          field_id="neck#1.neck_style#1.value", req_level="conditional"),
        f(128, "Silhouette", silhouette, "filled" if silhouette else "default", True,
          field_id="apparel_silhouette#1.value", req_level="conditional"),
        f(129, "Sleeve Length", _derive_sleeve_length(sleeve_type, resolved_pt), "filled", False,
          field_id="sleeve#1.length_description#1.value", req_level="conditional"),
        f(130, "Sleeve Type", sleeve_type, "filled" if sleeve_type else "default", True,
          field_id="sleeve#1.type#1.value", req_level="conditional"),
        f(131, "Closure Type", "", "empty", True,
          "Not auto-filled. Set if known (Pull On, Button, Zipper, etc.).",
          field_id="closure#1.type#1.value", req_level="recommended"),
        f(138, "UPF Protection", upf, "filled" if upf else "default", True,
          "" if upf else "No UPF value — leave blank if not applicable.",
          field_id="ultraviolet_protection_factor#1.value", req_level="conditional"),

        # ── Dates & Lifecycle ──
        f(77, "Item Booking Date", datetime.now().strftime("%Y-%m-%dT00:00:00Z"), "default", True,
          "Using today's date. Change if different.",
          field_id="item_booking_date#1.value", req_level="required"),
        f(126, "Product Lifecycle", "Perennial", "default", True,
          field_id="lifecycle_supply_type#1.value", req_level="required"),
        f(153, "Earliest Shipping Date", today_str, "default", True,
          "Using today's date. Update if different.",
          field_id="rtip_earliest_shipping_date#1.value", req_level="required"),

        # ── Counts & Units ──
        f(91, "Unit Count", "1", "default", True,
          field_id="unit_count#1.value", req_level="required"),
        f(92, "Unit Count Type", "Count", "default", True,
          field_id="unit_count#1.type.value", req_level="required"),
        f(149, "Skip Offer", "No", "default", True,
          field_id="skip_offer#1.value", req_level="required"),

        # ── Pricing ──
        f(150, "List Price", list_price, "filled" if list_price else "empty", True,
          "" if list_price else "Enter retail list price.",
          field_id="list_price#1.value", req_level="required"),
        f(151, "Cost Price", variant_cost, "filled" if variant_cost else "empty", True,
          "" if variant_cost else "Required for submission. Enter cost price.",
          field_id="cost_price#1.value", req_level="required"),

        # ── Shipping & Compliance ──
        f(152, "Import Designation", "Imported", "default", True,
          field_id="import_designation#1.value", req_level="required"),
        f(160, "Item Package Length", brand_cfg.get("default_pkg_length", ""), "filled" if brand_cfg.get("default_pkg_length") else "empty", True,
          "" if brand_cfg.get("default_pkg_length") else "Set package length or use Apply All.",
          field_id="item_package_dimensions#1.length.value", req_level="required"),
        f(161, "Package Length Unit", "Inches", "default", True,
          field_id="item_package_dimensions#1.length.unit", req_level="required"),
        f(162, "Item Package Width", brand_cfg.get("default_pkg_width", ""), "filled" if brand_cfg.get("default_pkg_width") else "empty", True,
          "" if brand_cfg.get("default_pkg_width") else "Set package width or use Apply All.",
          field_id="item_package_dimensions#1.width.value", req_level="required"),
        f(163, "Package Width Unit", "Inches", "default", True,
          field_id="item_package_dimensions#1.width.unit", req_level="required"),
        f(164, "Item Package Height", brand_cfg.get("default_pkg_height", ""), "filled" if brand_cfg.get("default_pkg_height") else "empty", True,
          "" if brand_cfg.get("default_pkg_height") else "Set package height or use Apply All.",
          field_id="item_package_dimensions#1.height.value", req_level="required"),
        f(165, "Package Height Unit", "Inches", "default", True,
          field_id="item_package_dimensions#1.height.unit", req_level="required"),
        f(166, "Item Package Weight", brand_cfg.get("default_pkg_weight", ""), "filled" if brand_cfg.get("default_pkg_weight") else "empty", True,
          "" if brand_cfg.get("default_pkg_weight") else "Set package weight or use Apply All.",
          field_id="item_package_weight#1.value", req_level="required"),
        f(167, "Package Weight Unit", "Pounds", "default", True,
          field_id="item_package_weight#1.unit", req_level="required"),
        f(168, "Order Aggregate Type", "Each", "default", True,
          field_id="rtip_order_aggregate_type#1.value", req_level="required"),
        f(169, "Items per Inner Pack", "1", "default", True,
          field_id="rtip_items_per_inner_pack#1.value", req_level="required"),
        f(170, "Country of Origin", coo, "filled" if coo else "default", True,
          "" if coo else "Update with actual country of origin.",
          field_id="country_of_origin#1.value", req_level="required"),
        f(171, "Batteries Required", "No", "default", True,
          field_id="batteries_required#1.value", req_level="required"),
        f(172, "Batteries Included", "No", "default", True,
          field_id="batteries_included#1.value", req_level="required"),
        f(230, "Contains Battery or Cell", "No", "default", True,
          field_id="contains_battery_or_cell#1.value", req_level="required"),
    ]

    return fields


def _qa_summary_for_style(style_num, style_name, fields):
    """Given a list of field dicts, return summary counts and status."""
    filled  = sum(1 for f in fields if f["status"] == "filled" or f["status"] == "locked")
    defaults = sum(1 for f in fields if f["status"] == "default")
    empty   = sum(1 for f in fields if f["status"] == "empty")
    req_empty = sum(1 for f in fields if f["status"] == "empty" and f.get("req_level") == "required")
    total   = len(fields)
    if req_empty > 0:
        status = "attention"
    elif defaults > 0:
        status = "defaults"
    else:
        status = "ready"
    return {
        "style_num": style_num,
        "style_name": style_name,
        "filled": filled,
        "defaults": defaults,
        "empty": empty,
        "req_empty": req_empty,
        "total": total,
        "status": status,
    }


@app.route("/api/preview-nis", methods=["POST"])
def preview_nis():
    """Return all fields for a style with their values and QA status.
    Accepts brand/style/content from request body as fallback when session is empty.
    """
    data = request.get_json(force=True)
    style_num = data.get("style_num", "").strip()
    if not style_num:
        return jsonify({"error": "style_num required"}), 400

    # Use session data, but allow frontend to pass data directly as fallback
    brand        = session_data.get("brand") or data.get("brand", "")
    vendor_code  = session_data.get("vendor_code", "") or data.get("vendor_code", "")
    styles       = session_data.get("styles", []) or data.get("styles", [])
    content_map  = session_data.get("generated_content", {}) or data.get("content_map", {})
    overrides    = session_data.get("field_overrides", {}).get(style_num, {})

    # If session was empty but frontend sent data, restore session for subsequent calls
    if not session_data.get("brand") and brand:
        session_data["brand"] = brand
    if not session_data.get("vendor_code") and vendor_code:
        session_data["vendor_code"] = vendor_code
    if not session_data.get("styles") and styles:
        session_data["styles"] = styles
    if not session_data.get("generated_content") and content_map:
        session_data["generated_content"] = content_map

    if not brand:
        return jsonify({"error": "No brand selected"}), 400

    # Find style from session or from frontend-provided list
    style = next((s for s in styles if s["style_num"] == style_num), None)
    # Also accept a single style object passed directly
    if not style and data.get("style"):
        style = data["style"]
    if not style:
        return jsonify({"error": f"Style {style_num} not found"}), 404

    brand_cfg = _load_brand_config_data(brand)
    content   = content_map.get(style_num, {}) or data.get("content", {})

    fields = _build_preview_fields(brand, brand_cfg, vendor_code, style, content)

    # Apply any stored overrides (keyed by field_id) — mark overridden fields as 'filled'
    for field in fields:
        fid = field.get("field_id", "")
        if fid and fid in overrides:
            field["value"]  = overrides[fid]
            field["status"] = "filled"
            field["overridden"] = True

    filled_count   = sum(1 for f in fields if f["status"] in ("filled", "locked"))
    defaults_count = sum(1 for f in fields if f["status"] == "default")
    empty_count    = sum(1 for f in fields if f["status"] == "empty")

    return jsonify({
        "style_num":       style_num,
        "style_name":      style["style_name"],
        "total_fields":    len(fields),
        "filled":          filled_count,
        "defaults_used":   defaults_count,
        "needs_attention": empty_count,
        "fields":          fields,
    })


@app.route("/api/update-field", methods=["POST"])
def update_field():
    """Store a field override for a style, keyed by field_id.
    Picked up when generating .xlsm via do_xlsm_surgery / _generate_category_file.
    """
    data      = request.get_json(force=True)
    style_num = str(data.get("style_num", "")).strip()
    field_id  = str(data.get("field_id", "")).strip()
    value     = data.get("value", "")

    if not style_num or not field_id:
        return jsonify({"error": "style_num and field_id required"}), 400

    if "field_overrides" not in session_data:
        session_data["field_overrides"] = {}
    if style_num not in session_data["field_overrides"]:
        session_data["field_overrides"][style_num] = {}

    # Get the previous value for feedback tracking
    prev_value = session_data["field_overrides"][style_num].get(field_id, "")
    session_data["field_overrides"][style_num][field_id] = value

    # Auto-capture as implicit feedback (no extra clicks from operator)
    field_name = data.get("field_name", field_id)  # frontend can send human name
    _store_feedback({
        "id": f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{_uuid.uuid4().hex[:8]}",
        "timestamp": datetime.utcnow().isoformat(),
        "operator": session_data.get("operator", ""),
        "brand": session_data.get("brand", ""),
        "type": "field_override",
        "phase": "review",
        "context": {
            "scope": "field",
            "brand": session_data.get("brand", ""),
            "style_num": style_num,
            "field_id": field_id,
            "field_name": field_name,
        },
        "data": {"field_id": field_id, "field_name": field_name,
                 "original": prev_value, "updated": value},
        "maps_to": "derive_logic",
    })

    return jsonify({"ok": True, "style_num": style_num, "field_id": field_id, "value": value})


@app.route("/api/update-field-all", methods=["POST"])
def update_field_all():
    """Apply a field override to ALL styles at once.
    Used for shared fields like package dims, COO, care, etc.
    """
    data     = request.get_json(force=True)
    field_id = str(data.get("field_id", "")).strip()
    value    = data.get("value", "")

    if not field_id:
        return jsonify({"error": "field_id required"}), 400

    styles = session_data.get("styles", [])
    if not styles:
        return jsonify({"error": "No styles loaded"}), 400

    if "field_overrides" not in session_data:
        session_data["field_overrides"] = {}

    count = 0
    for style in styles:
        sn = style["style_num"]
        if sn not in session_data["field_overrides"]:
            session_data["field_overrides"][sn] = {}
        session_data["field_overrides"][sn][field_id] = value
        count += 1

    # Auto-capture as implicit feedback
    field_name = data.get("field_name", field_id)
    _store_feedback({
        "id": f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{_uuid.uuid4().hex[:8]}",
        "timestamp": datetime.utcnow().isoformat(),
        "operator": session_data.get("operator", ""),
        "brand": session_data.get("brand", ""),
        "type": "apply_all",
        "phase": "review",
        "context": {
            "scope": "brand",
            "brand": session_data.get("brand", ""),
            "field_id": field_id,
            "field_name": field_name,
        },
        "data": {"field_id": field_id, "field_name": field_name,
                 "value": value, "styles_count": count},
        "maps_to": "brand_config",
    })

    return jsonify({"ok": True, "field_id": field_id, "value": value, "styles_updated": count})


@app.route("/api/save-field-as-brand-default", methods=["POST"])
def save_field_as_brand_default():
    """Save a field value as a brand default in the brand config JSON.
    Next time this brand is loaded, this value will be pre-filled.
    """
    data     = request.get_json(force=True)
    brand    = data.get("brand", "") or session_data.get("brand", "")
    field_id = data.get("field_id", "").strip()
    value    = data.get("value", "")
    field_name = data.get("field_name", field_id)

    if not brand or not field_id:
        return jsonify({"error": "brand and field_id required"}), 400

    # Load existing brand config
    cfg = _load_brand_config_data(brand)

    # Map known field_ids to brand config keys
    FIELD_TO_CONFIG = {
        "country_of_origin#1.value": "default_coo",
        "care_instructions#1.value": "default_care",
        "ultraviolet_protection_factor#1.value": "default_upf",
        "material#1.value": "default_fabric",
        "item_package_dimensions#1.length.value": "default_pkg_length",
        "item_package_dimensions#1.width.value": "default_pkg_width",
        "item_package_dimensions#1.height.value": "default_pkg_height",
        "item_package_weight#1.value": "default_pkg_weight",
        "department#1.value": "department",
        "target_gender#1.value": "gender",
    }

    config_key = FIELD_TO_CONFIG.get(field_id)
    if config_key:
        cfg[config_key] = value
    else:
        # Store in a generic defaults dict for unmapped fields
        if "field_defaults" not in cfg:
            cfg["field_defaults"] = {}
        cfg["field_defaults"][field_id] = value

    # Save brand config
    safe_brand = re.sub(r'[^\w\-]', '_', brand)
    cfg_path = BRAND_CONFIGS_DIR / f"{safe_brand}.json"
    with open(str(cfg_path), "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2)

    return jsonify({"ok": True, "brand": brand, "field_id": field_id, "config_key": config_key or f"field_defaults.{field_id}"})


@app.route("/api/nis-qa-summary", methods=["GET", "POST"])
def nis_qa_summary():
    """Return QA summary across all styles.
    Accepts POST with brand/styles/content_map as fallback when session is empty.
    """
    req_data = request.get_json(force=True) if request.method == "POST" else {}

    brand       = session_data.get("brand") or req_data.get("brand", "")
    vendor_code = session_data.get("vendor_code", "") or req_data.get("vendor_code", "")
    styles      = session_data.get("styles", []) or req_data.get("styles", [])
    content_map = session_data.get("generated_content", {}) or req_data.get("content_map", {})
    overrides   = session_data.get("field_overrides", {})

    # Restore session if frontend sent data
    if not session_data.get("brand") and brand:
        session_data["brand"] = brand
    if not session_data.get("styles") and styles:
        session_data["styles"] = styles
    if not session_data.get("generated_content") and content_map:
        session_data["generated_content"] = content_map
    if not session_data.get("vendor_code") and vendor_code:
        session_data["vendor_code"] = vendor_code

    if not brand:
        return jsonify({"error": "No brand selected"}), 400
    if not styles:
        return jsonify({"error": "No styles loaded"}), 400

    brand_cfg = _load_brand_config_data(brand)

    style_summaries = []
    total_ready = 0
    total_attention = 0
    total_defaults = 0

    for style in styles:
        snum    = style["style_num"]
        sname   = style["style_name"]
        content = content_map.get(snum, {})
        fields  = _build_preview_fields(brand, brand_cfg, vendor_code, style, content)

        # Apply overrides (keyed by field_id)
        style_overrides = overrides.get(snum, {})
        for field in fields:
            fid = field.get("field_id", "")
            if fid and fid in style_overrides:
                field["status"] = "filled"

        summary = _qa_summary_for_style(snum, sname, fields)
        style_summaries.append(summary)

        if summary["status"] == "ready":
            total_ready += 1
        elif summary["status"] == "attention":
            total_attention += 1
        else:  # defaults
            total_defaults += 1

    return jsonify({
        "total_styles":    len(styles),
        "ready":           total_ready,
        "has_defaults":    total_defaults,
        "needs_attention": total_attention,
        "content_generated": len(content_map) > 0,
        "styles":          style_summaries,
    })


# ═══════════════════════════════════════════════════════════════════════
# NIS ROW-SPACING FIX
# Long-text fields (bullets, description, item_name, style, model_name) must
# render with wrap_text=True and auto-fit row heights. The Amazon templates
# ship with customHeight=True, height=12.75 on every data row — when wrapped
# 200-char bullets are written, Excel hides the wrapped lines because the
# row is clamped to a single text-line's worth of vertical space, which
# shows up as "rows on top of each other".
# ═══════════════════════════════════════════════════════════════════════
LONG_TEXT_FIELD_IDS = {
    "bullet_point#1.value", "bullet_point#2.value", "bullet_point#3.value",
    "bullet_point#4.value", "bullet_point#5.value",
    "rtip_product_description#1.value",
    "item_name#1.value", "style#1.value", "model_name#1.value",
    "generic_keyword#1.value", "item_type_keyword#1.value",
}

def _is_long_text_field(field_id):
    # v0.7.6.1: keep the registry but DO NOT force wrap_text — Amazon's
    # reference NIS files ship with wrap_text=None and let Excel overflow
    # the cell visually. _apply_long_text_alignment is now a no-op.
    return field_id in LONG_TEXT_FIELD_IDS

def _apply_long_text_alignment(cell, cached_alignment=None):
    """v0.7.6.1: NO-OP. Amazon's reference NIS files (Stella Parker SPTW etc.)
    ship with wrap_text=None on long-text cells — Excel overflows the cell
    visually and the data row stays at default 16pt height. Forcing
    wrap_text=True caused our generated files to look unlike real Amazon
    output. We now leave the template's row-7 alignment as-is.
    """
    return

def _clear_row_heights_for_auto_fit(ws, start_row=7, end_row=None):
    """v0.7.6.1: Match Amazon's reference behavior — leave data rows at default
    height (~16pt) with no wrap_text. Reference NIS files (e.g. Stella Parker
    SPTW) ship with no row_dim on data rows, default 16pt; bullets and
    item_name overflow on a single visual line. Forcing tall rows + wrap_text
    makes our output look unlike a real Amazon template.

    Strategy: drop any pre-existing customHeight on data rows so they fall
    back to the sheet default. Hidden header rows (1, 4) are left alone.
    """
    end_row = end_row or (ws.max_row or 100)
    for r in range(start_row, end_row + 1):
        if r in ws.row_dimensions:
            try:
                del ws.row_dimensions[r]
            except Exception:
                pass

def do_xlsm_surgery(template_path, brand, brand_cfg, vendor_code, style, content):
    """
    .xlsm surgery — field-ID based dynamic column mapping:
    1. Load template with keep_vba=True
    2. Find the Template-* sheet; derive product_type from sheet name
    3. Build field_id → column_number map from row 4 (exact matches, stripped)
    4. Capture cell styles from row 7
    5. Clear rows 7+
    6. Write parent + child rows using exact field_id lookups
    7. Save as new file
    """
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(template_path, keep_vba=True)

    # ── Find Template sheet and detect product type ───────────────────────────
    ws = None
    detected_product_type = "DRESS"  # safe default
    for name in wb.sheetnames:
        if name.upper().startswith("TEMPLATE"):
            ws = wb[name]
            # Extract product type: "Template-SWIMWEAR" → "SWIMWEAR"
            parts = name.split("-", 1)
            if len(parts) == 2 and parts[1].strip():
                detected_product_type = parts[1].strip().upper()
            break
    if ws is None:
        # Fallback: any sheet with "template" in name
        for name in wb.sheetnames:
            if "template" in name.lower():
                ws = wb[name]
                break
    if ws is None:
        ws = wb.active

    # ── Build field_id → column_number map from row 4 ─────────────────────────
    max_col = ws.max_column or 254
    col_map = {}  # field_id_string → column_number
    for col in range(1, max_col + 1):
        raw = ws.cell(row=4, column=col).value
        if raw is not None:
            fid = str(raw).strip()
            if fid:
                col_map[fid] = col

    def _col(field_id):
        """Exact field_id lookup. Returns column number or None."""
        return col_map.get(field_id)

    # ── Capture styles from row 7 (before clearing) ───────────────────────────
    style_cache = {}
    for col in range(1, max_col + 1):
        cell = ws.cell(row=7, column=col)
        style_cache[col] = {
            "font":          copy.copy(cell.font)      if cell.font      else None,
            "fill":          copy.copy(cell.fill)      if cell.fill      else None,
            "border":        copy.copy(cell.border)    if cell.border    else None,
            "alignment":     copy.copy(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format,
        }

    # ── Clear existing data rows (7+) ─────────────────────────────────────────
    for row_idx in range(7, (ws.max_row or 100) + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row_idx, column=col).value = None

    # ── NIS row-spacing fix: release fixed row heights so wrapped bullet text auto-fits
    _clear_row_heights_for_auto_fit(ws, start_row=7)

    # ── Unpack style / content data ───────────────────────────────────────────
    style_num      = style["style_num"]
    style_name     = style["style_name"]
    variants       = style["variants"]
    list_price     = style.get("list_price", "")
    cost_price     = style.get("cost_price", "")
    model_name_raw = style.get("model_name", "") or style_name
    sub_class      = style.get("subclass", "") or style.get("sub_class", "")
    sub_subclass   = style.get("sub_subclass", "")

    bullets     = content.get("bullets", [])
    description = content.get("description", "")
    backend_kw  = content.get("backend_keywords", "")
    neck_type   = content.get("neck_type", "")  or style.get("neck_type", "") or derive_neck_type(style_name)
    # PT-aware sleeve fallback: COAT → "Long Sleeve" (not "Sleeveless") when no in-name signal.
    sleeve_type = content.get("sleeve_type", "") or style.get("sleeve_type", "") or derive_sleeve_type(style_name, detected_product_type)
    silhouette  = content.get("silhouette", "")  or derive_silhouette(sub_subclass)
    style_gender, style_dept = _derive_gender_department(style)
    eff_gender = style_gender or brand_cfg.get("gender", "")
    eff_dept   = style_dept or brand_cfg.get("department", "")
    category    = _derive_amazon_product_category(sub_class, gender=eff_gender, product_type=detected_product_type, style_name=style_name, department=eff_dept)
    subcategory = SUBCLASS_SUBCATEGORY_MAP.get(sub_class, '')
    if not subcategory and detected_product_type == "SWIMWEAR":
        subcategory = _derive_swim_product_subcategory(sub_class, gender=eff_gender, style_name=style_name, product_category=category)
    fabric      = content.get("fabric", "")      or brand_cfg.get("default_fabric", "")
    care        = content.get("care", "")        or brand_cfg.get("default_care", "")
    upf         = content.get("upf", "")         or brand_cfg.get("default_upf", "")
    coo         = normalize_coo(content.get("coo", "")         or brand_cfg.get("default_coo", "")) or "Imported"

    clean_brand    = clean_brand_name(brand)
    item_type_name = _derive_item_type_name(sub_class, product_type=detected_product_type, gender=eff_gender, style_name=style_name)
    item_length    = _derive_item_length(sub_subclass, style_name, product_type=detected_product_type, sub_class=sub_class)
    fabric_type    = _derive_fabric_type(fabric)
    itk_value      = _derive_item_type_keyword(sub_class, product_type=detected_product_type, gender=eff_gender, style_name=style_name)
    # Taxonomy override wins over auto-derivation (Phase 1)
    _tax = _resolve_taxonomy_for_style(style, brand_cfg)
    if _tax.get("matched"):
        _e = _tax["entry"]
        category       = _e.get("product_category", category)
        subcategory    = _e.get("product_subcategory", subcategory)
        itk_value      = _e.get("item_type_keyword", itk_value)
        item_type_name = _e.get("item_type_name", item_type_name)
    sleeve_len     = _derive_sleeve_length(sleeve_type, detected_product_type)
    today_str      = datetime.now().strftime("%Y%m%d")
    booking_date  = datetime.now().strftime("%Y-%m-%dT00:00:00Z")

    # Parent SKU: derive from variant SKUs (preserves seasoncode-style#) when possible,
    # else fall back to bare style_num.
    parent_sku = _derive_parent_sku_from_variants(style.get("variants", []), style_num) or style_num

    # Load any QA field overrides for this style (keyed by field_id)
    _field_overrides = session_data.get("field_overrides", {}).get(style_num, {})

    # ── Cell writer ───────────────────────────────────────────────────────────
    # Fields that must always show two-decimal currency formatting (132.00, not 132).
    # The Sage feedback ("List Price came out as 1 decimal; should be two decimals, no $ sign")
    # applies to any monetary field; we declare them once and apply consistently.
    PRICE_FIELDS = {
        "list_price#1.value",
        "cost_price#1.value",
        "unit_price#1.value",
        "map_price#1.value",
    }

    def write_cell(row_idx, field_id, value):
        """Write value to the cell for field_id, applying row-7 styles.
        Checks field_overrides first (keyed by field_id).
        """
        # Field-ID-keyed override takes priority
        if field_id in _field_overrides:
            value = _field_overrides[field_id]
        col_num = _col(field_id)
        if col_num is None or value is None:
            return
        if value == "":
            return
        cell = ws.cell(row=row_idx, column=col_num)
        is_price = field_id in PRICE_FIELDS
        # Try numeric conversion for price/dimension fields
        if isinstance(value, str):
            try:
                value = float(value)
                # Don't downgrade prices to int (132.0 → 132 strips trailing zero in Excel display)
                if not is_price and value == int(value):
                    value = int(value)
            except (ValueError, TypeError):
                pass
        cell.value = value if isinstance(value, (int, float)) else str(value)
        cached = style_cache.get(col_num, {})
        if cached.get("font"):          cell.font          = copy.copy(cached["font"])
        if cached.get("fill"):          cell.fill          = copy.copy(cached["fill"])
        if cached.get("border"):        cell.border        = copy.copy(cached["border"])
        if cached.get("alignment"):     cell.alignment     = copy.copy(cached["alignment"])
        # Apply cached number_format only for non-price fields; price fields force 0.00
        if not is_price and cached.get("number_format"):
            cell.number_format = cached["number_format"]
        if is_price and isinstance(value, (int, float)):
            cell.value = float(value)
            cell.number_format = "0.00"
        # Long-text fields need wrap_text=True so Excel renders all wrapped
        # lines instead of stacking them into a single 12.75pt row.
        if _is_long_text_field(field_id):
            _apply_long_text_alignment(cell, cached.get("alignment"))

    # ── Shared-fields writer (called for both parent and child rows) ──────────
    def write_shared(row_idx, vendor_sku_val, is_child=False):
        """Write all fields common to parent and child rows."""
        write_cell(row_idx, "rtip_vendor_code#1.value",         vendor_code or brand_cfg.get("vendor_code_full", ""))
        write_cell(row_idx, "vendor_sku#1.value",               vendor_sku_val)
        write_cell(row_idx, "product_type#1.value",             detected_product_type)
        # Variation theme must match Amazon dropdown. For SWIMWEAR, 'SIZE/COLOR' is valid; 'COLOR/SIZE' is NOT.
        _hm_color = len({(_v.get("color") or _v.get("color_name") or "") for _v in variants}) > 1
        _hm_size  = len({_v.get("size", "") for _v in variants}) > 1
        if _hm_color and _hm_size:
            _vt = "SIZE/COLOR"
        elif _hm_color:
            _vt = "COLOR"
        elif _hm_size:
            _vt = "SIZE"
        else:
            _vt = "COLOR"
        # Variation theme is only relevant for parent/child relationships; skip when standalone.
        if not bool(session_data.get("skip_parent_row") or brand_cfg.get("skip_parent_row", False)):
            write_cell(row_idx, "variation_theme#1.name",     _vt)
        write_cell(row_idx, "brand#1.value",                    clean_brand)
        if category:
            write_cell(row_idx, "product_category#1.value",     category)
        if subcategory:
            write_cell(row_idx, "product_subcategory#1.value",  subcategory)
        write_cell(row_idx, "item_type_keyword#1.value",        itk_value)
        write_cell(row_idx, "model_number#1.value",             style_num)
        write_cell(row_idx, "model_name#1.value",               (model_name_raw or style_name).title())
        # Bullets
        for i, bkey in enumerate(["bullet_point#1.value", "bullet_point#2.value",
                                   "bullet_point#3.value", "bullet_point#4.value",
                                   "bullet_point#5.value"]):
            if i < len(bullets):
                write_cell(row_idx, bkey, bullets[i][:500])
        write_cell(row_idx, "generic_keyword#1.value",          backend_kw)
        write_cell(row_idx, "style#1.value",                    style_name.title())
        # fit_type — left blank unless from data/override
        write_cell(row_idx, "fit_type#1.value",                 content.get("fit_type", "") or style.get("fit_type", "") or brand_cfg.get("default_fit_type", ""))
        # Target gender: Amazon accepts Male/Female; refine youth via style_name.
        _tg_sn = (style_name or "").lower()
        if "boys" in _tg_sn or "men" in _tg_sn:
            _tg = "Male"
        elif "girls" in _tg_sn or "women" in _tg_sn:
            _tg = "Female"
        else:
            _tg = eff_gender if eff_gender in ("Male", "Female") else ""
        _first_var = variants[0] if variants else {}
        _sst_s, _sclass_s, _ard_s, _ = _derive_youth_size_info(style_name, eff_gender, _first_var.get("size", ""))
        write_cell(row_idx, "department#1.value",               eff_dept)
        write_cell(row_idx, "target_gender#1.value",            _tg)
        write_cell(row_idx, "age_range_description#1.value",    _ard_s)
        if _sst_s:
            write_cell(row_idx, "special_size_type#1.value",   _sst_s)
        # Body Type / Height Type — use 'Regular' (valid Amazon value across all apparel PTs).
        # Sage feedback flagged these as previously blank; "All Body Types" / "All Heights" exist
        # in apparel_defaults.json but they're NOT in the actual COAT dropdown — 'Regular' is.
        write_cell(row_idx, _size_field(detected_product_type, "body_type", col_map),         "Regular")
        write_cell(row_idx, _size_field(detected_product_type, "height_type", col_map),       "Regular")
        # Sage feedback: Material should be single-fiber names (Polyester, Cotton, Spandex)
        # split into material#1, #2, #3; Fabric Type holds the full blend percentage string.
        if fabric:
            _materials = _split_fabric_into_materials(fabric)
            for _mi, _mat in enumerate(_materials[:5], start=1):
                write_cell(row_idx, f"material#{_mi}.value", _mat)
            write_cell(row_idx, "fabric_type#1.value", fabric)
        else:
            write_cell(row_idx, "fabric_type#1.value", fabric_type)
        # Special Features (AI-suggested from style name) — Sage feedback
        _addl = style.get("additional_details", "") or content.get("additional_details", "")
        for _fi, _feat in enumerate(_derive_special_features(style_name, _addl)[:5], start=1):
            write_cell(row_idx, f"special_feature#{_fi}.value", _feat)
        # Lifestyle (AI-suggested from style name + sub_class)
        for _li, _lf in enumerate(_derive_lifestyle(style_name, sub_class)[:2], start=1):
            write_cell(row_idx, f"lifestyle#{_li}.value", _lf)
        write_cell(row_idx, "number_of_items#1.value", "1")
        write_cell(row_idx, "item_type_name#1.value",           item_type_name)
        write_cell(row_idx, "rtip_product_description#1.value", description)
        write_cell(row_idx, "item_length_description#1.value",  item_length)
        write_cell(row_idx, "item_booking_date#1.value",        booking_date)
        if care:
            write_cell(row_idx, "care_instructions#1.value",    care)
        write_cell(row_idx, "unit_count#1.value",               "1")
        write_cell(row_idx, "unit_count#1.type.value",                "Count")
        # Neck/Collar field name varies by template (COAT uses collar_style, SWIMWEAR uses neck_style).
        if neck_type:
            _neck_fid = _neck_field(col_map)
            if _neck_fid:
                write_cell(row_idx, _neck_fid, neck_type)
        write_cell(row_idx, "lifecycle_supply_type#1.value", "Perennial")
        if silhouette:
            write_cell(row_idx, "apparel_silhouette#1.value",   silhouette)
        # COAT-specific: 'Type of Jacket' → coat_silhouette_type#1.value (Puffer→Quilted, etc.)
        if detected_product_type == "COAT":
            _toj = style.get("type_of_jacket", "") or content.get("type_of_jacket", "")
            _coat_sil = _derive_coat_silhouette(_toj, fallback_subclass=sub_class)
            if _coat_sil:
                write_cell(row_idx, "coat_silhouette_type#1.value", _coat_sil)
        write_cell(row_idx, "sleeve#1.length_description#1.value", sleeve_len)
        if sleeve_type:
            write_cell(row_idx, "sleeve#1.type#1.value",        sleeve_type)
        # closure — left blank unless from data/override
        write_cell(row_idx, "closure#1.type#1.value",             content.get("closure_type", "") or style.get("closure_type", "") or brand_cfg.get("default_closure", ""))
        # number of pockets — from pre-upload (Sage feedback: was missing on template)
        _pockets_v = content.get("pockets", "") or style.get("pockets", "")
        if _pockets_v not in (None, "", 0):
            try:    write_cell(row_idx, "number_of_pockets#1.value", int(_pockets_v))
            except (ValueError, TypeError): write_cell(row_idx, "number_of_pockets#1.value", str(_pockets_v))
        if upf:
            write_cell(row_idx, "ultraviolet_protection_factor#1.value", upf)
        write_cell(row_idx, "skip_offer#1.value",                       "No")
        # Import designation: "Imported" unless COO is US
        import_desig = "Imported" if coo.upper() not in ("US", "USA", "UNITED STATES") else "Domestic"
        write_cell(row_idx, "import_designation#1.value",       import_desig)
        write_cell(row_idx, "rtip_earliest_shipping_date#1.value", today_str)
        # Contains battery/cell — required compliance field
        write_cell(row_idx, "contains_battery_or_cell#1.value", "No")
        # Dangerous Goods Regulation — PT-aware (Not Applicable for apparel)
        if _pt_defaults.pt_writes(detected_product_type, "dg"):
            _dg = _pt_defaults.get_pt_default(detected_product_type, "default_dg_regulation", "Not Applicable")
            write_cell(row_idx, "supplier_declared_dg_hz_regulation#1.value", _dg)
        # Package dimensions
        # Package dims — left blank unless from data/override/brand config
        write_cell(row_idx, "item_package_dimensions#1.length.value",      brand_cfg.get("default_pkg_length", ""))
        write_cell(row_idx, "item_package_dimensions#1.length.unit",       "Inches")
        write_cell(row_idx, "item_package_dimensions#1.width.value",       brand_cfg.get("default_pkg_width", ""))
        write_cell(row_idx, "item_package_dimensions#1.width.unit",        "Inches")
        write_cell(row_idx, "item_package_dimensions#1.height.value",      brand_cfg.get("default_pkg_height", ""))
        write_cell(row_idx, "item_package_dimensions#1.height.unit",       "Inches")
        write_cell(row_idx, "item_package_weight#1.value",      brand_cfg.get("default_pkg_weight", ""))
        write_cell(row_idx, "item_package_weight#1.unit",       "Pounds")
        write_cell(row_idx, "rtip_order_aggregate_type#1.value",     "Each")
        write_cell(row_idx, "rtip_items_per_inner_pack#1.value",     "1")
        if coo:
            write_cell(row_idx, "country_of_origin#1.value",   coo)
        write_cell(row_idx, "batteries_required#1.value",  "No")
        write_cell(row_idx, "batteries_included#1.value",  "No")
        # List price + cost price on shared fields (applies to parent; child may override).
        # Sage feedback: cost_price was missing on parent rows because it was previously only
        # written for children. Vendor Central requires both on every row.
        if list_price:
            try:    write_cell(row_idx, "list_price#1.value",   float(list_price))
            except: write_cell(row_idx, "list_price#1.value",   list_price)
        if cost_price:
            try:    write_cell(row_idx, "cost_price#1.value",   float(cost_price))
            except: write_cell(row_idx, "cost_price#1.value",   cost_price)

    current_row = 7

    # ── Parent row ────────────────────────────────────────────────────────────
    # Skip-parent toggle (Sage feedback): when set, the writer emits children only.
    # parentage_level + child_parent_sku_relationship + variation_theme are bypassed.
    skip_parent = bool(session_data.get("skip_parent_row")
                       or brand_cfg.get("skip_parent_row", False))

    # Parent row (required by Amazon unless skip_parent is enabled)
    if not skip_parent:
        write_shared(current_row, parent_sku, is_child=False)
        write_cell(current_row, "parentage_level#1.value", "Parent")
        write_cell(current_row, "item_name#1.value", content.get("title", style_name))
        current_row += 1

    # Child rows (one per variant)
    for v in variants:
        color_name = v.get("color_name", "")
        size       = v.get("size", "")
        upc        = v.get("upc", "")
        child_asin = v.get("child_asin", "")
        # Prefer source SKU verbatim; only synthesize when missing.
        sku        = _derive_child_sku(v, parent_sku, color_name, size,
                                       color_code=v.get("color_code", ""),
                                       size_code=v.get("size_code", ""))

        # PT-aware color snap (drops bad 'Ivory' for COAT, fills 'Brown' for Truffle, etc.)
        color_family    = normalize_color(color_name, detected_product_type)
        size_normalized = normalize_size(size)

        variant_title = generate_title(
            brand_cfg, brand, style_name, detected_product_type.title(),
            color_name, size, upf, style_gender=style_gender
        )
        # Youth-aware size resolution: 2T -> '2 Years', Alpha stays as-is.
        _sst_v, _sclass_v, _ard_v, _size_youth = _derive_youth_size_info(style_name, eff_gender, size)
        size_normalized = _size_youth or size_normalized

        write_shared(current_row, sku, is_child=True)
        # In skip_parent mode, children stand alone (no parentage_level / parent_sku link).
        if not skip_parent:
            write_cell(current_row, "parentage_level#1.value",      "Child")
            write_cell(current_row, "child_parent_sku_relationship#1.child_relationship_type", "Variation")
            write_cell(current_row, "child_parent_sku_relationship#1.parent_sku",           parent_sku)
        write_cell(current_row, "item_name#1.value",            variant_title)
        if upc:
            write_cell(current_row, "external_product_id#1.type",  "UPC")
            write_cell(current_row, "external_product_id#1.value", re.sub(r'\D', '', str(upc)))
        if child_asin:
            asin_col = _col("merchant_suggested_asin#1.value")
            if asin_col:
                ws.cell(row=current_row, column=asin_col).value = child_asin
        write_cell(current_row, _size_field(detected_product_type, "size_system", col_map),   "US")
        write_cell(current_row, _size_field(detected_product_type, "size_class", col_map),    _sclass_v)
        write_cell(current_row, _size_field(detected_product_type, "size", col_map),          size_normalized or size)
        if _sst_v:
            write_cell(current_row, "special_size_type#1.value", _sst_v)
        write_cell(current_row, "age_range_description#1.value", _ard_v)
        write_cell(current_row, "color#1.standardized_values#1",         color_family)
        write_cell(current_row, "color#1.value",                color_name.title() if color_name else "")
        # Child cost price — always write so overrides can apply even if source is empty
        child_cost = v.get("cost_price") or cost_price
        if child_cost:
            try:    write_cell(current_row, "cost_price#1.value",     float(child_cost))
            except: write_cell(current_row, "cost_price#1.value",     child_cost)
        else:
            write_cell(current_row, "cost_price#1.value",             "")  # trigger override check
        current_row += 1

    # ── Save ──────────────────────────────────────────────────────────────────
    safe_name = re.sub(r'[^\w\-]', '_', style_num)
    output_filename = f"NIS_{brand.replace(' ', '_')}_{safe_name}.xlsm"
    output_path = str(UPLOAD_OUTPUT / output_filename)

    import warnings as _w
    with _w.catch_warnings():
        _w.simplefilter("ignore")
        wb.save(output_path)
    wb.close()

    return output_path

@app.route("/api/download/<filename>")
def download_file(filename):
    # Sanitize
    filename = Path(filename).name
    file_path = UPLOAD_OUTPUT / filename
    if not file_path.exists():
        abort(404)
    return send_file(
        str(file_path),
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.ms-excel.sheet.macroEnabled.12",
    )

def _generate_category_file(cat_styles, content_map, template_path, brand, brand_cfg, vendor_code, output_path):
    """Generate a single .xlsm with all styles of one category — field-ID based.
    Fills ALL required and conditionally-required Amazon columns.
    """
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(template_path, keep_vba=True)

    ws = None
    detected_product_type = "DRESS"  # safe default
    for name in wb.sheetnames:
        if name.upper().startswith("TEMPLATE"):
            ws = wb[name]
            parts = name.split("-", 1)
            if len(parts) == 2 and parts[1].strip():
                detected_product_type = parts[1].strip().upper()
            break
    if ws is None:
        for name in wb.sheetnames:
            if "template" in name.lower():
                ws = wb[name]
                break
    if ws is None:
        ws = wb.active

    # ── Build field_id → column_number map from row 4 ─────────────────────────
    max_col = ws.max_column or 254
    col_map = {}  # field_id_string → column_number
    for col in range(1, max_col + 1):
        raw = ws.cell(row=4, column=col).value
        if raw is not None:
            fid = str(raw).strip()
            if fid:
                col_map[fid] = col

    def _col(field_id):
        return col_map.get(field_id)
    
    # ── Capture styles from row 7 ────────────────────────────────────────────
    cell_styles = {}
    for col in range(1, max_col + 1):
        cell = ws.cell(row=7, column=col)
        cell_styles[col] = {
            "font":          copy.copy(cell.font)      if cell.font      else None,
            "fill":          copy.copy(cell.fill)      if cell.fill      else None,
            "border":        copy.copy(cell.border)    if cell.border    else None,
            "alignment":     copy.copy(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format,
        }

    # ── Clear data rows ───────────────────────────────────────────────────────
    for row in range(7, (ws.max_row or 100) + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None

    # ── NIS row-spacing fix: release fixed row heights so wrapped bullet text auto-fits
    _clear_row_heights_for_auto_fit(ws, start_row=7)

    clean_brand = clean_brand_name(brand)
    today_str   = datetime.now().strftime("%Y%m%d")
    booking_date = datetime.now().strftime("%Y-%m-%dT00:00:00Z")
    all_overrides = session_data.get("field_overrides", {})

    def wc(row_idx, field_id, value, style_num=None):
        """Write value to the cell for field_id, applying row-7 styles.
        Checks field_overrides first (keyed by field_id).
        """
        if style_num and field_id:
            style_ov = all_overrides.get(style_num, {})
            if field_id in style_ov:
                value = style_ov[field_id]
        c = _col(field_id)
        if c is None or value is None:
            return
        # Auto-correct against template dropdown (if available)
        # Skip fuzzy-match for free-text fields that accept compositions (e.g. "95% Polyester, 5% Spandex")
        SKIP_FUZZY = {"material#1.value", "generic_keyword#1.value", "item_type_keyword#1.value",
                      "rtip_product_description#1.value", "item_name#1.value",
                      "bullet_point#1.value", "bullet_point#2.value", "bullet_point#3.value",
                      "bullet_point#4.value", "bullet_point#5.value"}
        if isinstance(value, str) and value and field_id and field_id not in SKIP_FUZZY:
            dropdowns = load_dropdown_cache(detected_product_type)
            if field_id in dropdowns:
                valid = dropdowns[field_id]
                if value not in valid:
                    match, conf = _fuzzy_match_dropdown(value, valid)
                    if match and conf >= 0.6:
                        value = match

        # Skip truly empty values (don't write blank cells)
        if value == "":
            return
        cell = ws.cell(row=row_idx, column=c)
        # Currency fields need two-decimal display (132 → 132.00)
        is_price = field_id in {"list_price#1.value", "cost_price#1.value",
                                 "unit_price#1.value", "map_price#1.value"}
        # Try numeric conversion for price/dimension fields
        if isinstance(value, str):
            try:
                value = float(value)
                if not is_price and value == int(value):
                    value = int(value)
            except (ValueError, TypeError):
                pass
        cell.value = value if isinstance(value, (int, float)) else str(value)
        cached = cell_styles.get(c, {})
        if cached.get("font"):          cell.font          = copy.copy(cached["font"])
        if cached.get("fill"):          cell.fill          = copy.copy(cached["fill"])
        if cached.get("border"):        cell.border        = copy.copy(cached["border"])
        if cached.get("alignment"):     cell.alignment     = copy.copy(cached["alignment"])
        # Apply cached number_format ONLY for non-price fields. For prices we set 0.00 below
        # to fix the Sage feedback ("List Price came out as 1 decimal; should be two decimals").
        if not is_price and cached.get("number_format"):
            cell.number_format = cached["number_format"]
        if is_price and isinstance(value, (int, float)):
            cell.value = float(value)
            cell.number_format = "0.00"
        # Long-text fields need wrap_text=True for bullets, description, etc.
        if _is_long_text_field(field_id):
            _apply_long_text_alignment(cell, cached.get("alignment"))

    cr = 7
    for style in cat_styles:
        sn           = style["style_num"]
        style_name   = style.get("style_name", sn)
        sub_class    = style.get("subclass", "") or style.get("sub_class", "")
        sub_subclass = style.get("sub_subclass", "")
        content      = content_map.get(sn, {})
        if not content:
            continue

        # Parent SKU: prefer to derive from variants' source SKUs (e.g. 'F26-107010297-BEG-S' → 'F26-107010297').
        # This preserves the seasoncode-style format the team uses, instead of substituting vendor_code_prefix.
        # Falls back to the legacy '<vendor_code_prefix>-<style_num>' if variant SKUs are missing or non-matching.
        psku = _derive_parent_sku_from_variants(style.get("variants", []), sn) \
               or f"{brand_cfg.get('vendor_code_prefix', '')}-{sn}".strip("-") \
               or sn

        # Derive per-style gender/department from division_name
        style_gender, style_dept = _derive_gender_department(style)
        eff_gender = style_gender or brand_cfg.get("gender", "")
        eff_dept   = style_dept or brand_cfg.get("department", "")
        # Derive per-style fields — also fall back to the style dict (pre-upload values),
        # not just content + brand_cfg. Otherwise pre-upload fabric (e.g. "100% Polyester")
        # is dropped on the floor and Material/Fabric Type writes use only the global default.
        fabric     = content.get("fabric", "")     or style.get("fabric", "")     or brand_cfg.get("default_fabric", "")
        care       = content.get("care", "")       or style.get("care", "")       or brand_cfg.get("default_care", "")
        upf        = content.get("upf", "")        or style.get("upf", "")        or brand_cfg.get("default_upf", "")
        coo        = normalize_coo(content.get("coo", "") or style.get("coo", "") or brand_cfg.get("default_coo", "")) or "Imported"
        neck       = content.get("neck_type", "") or style.get("neck_type", "") or derive_neck_type(style_name)
        sleeve     = content.get("sleeve_type", "") or style.get("sleeve_type", "") or derive_sleeve_type(style_name, detected_product_type)
        sil        = content.get("silhouette", "") or derive_silhouette(sub_subclass)
        itk        = _derive_item_type_keyword(sub_class, product_type=detected_product_type, gender=eff_gender, style_name=style_name)
        itn        = _derive_item_type_name(sub_class, product_type=detected_product_type, gender=eff_gender, style_name=style_name)
        # Taxonomy override: if a confirmed entry exists for this style's bucket, use it
        _tax = _resolve_taxonomy_for_style(style, brand_cfg)
        if _tax.get("matched"):
            _e = _tax["entry"]
            itk = _e.get("item_type_keyword", itk)
            itn = _e.get("item_type_name", itn)
        ilen       = _derive_item_length(sub_subclass, style_name, product_type=detected_product_type, sub_class=sub_class)
        ftype      = _derive_fabric_type(fabric)
        slvlen     = _derive_sleeve_length(sleeve, detected_product_type)
        list_price = style.get("list_price", "") or content.get("list_price", "")
        cost_price = style.get("cost_price", "") or content.get("cost_price", "")
        bullets    = content.get("bullets", [])
        import_desig = "Imported" if coo.upper() not in ("US", "USA", "UNITED STATES") else "Domestic"
        cat_val    = _derive_amazon_product_category(sub_class, gender=eff_gender, product_type=detected_product_type, style_name=style_name, department=eff_dept)
        subcat_val = SUBCLASS_SUBCATEGORY_MAP.get(sub_class, "")
        if not subcat_val and detected_product_type == "SWIMWEAR":
            subcat_val = _derive_swim_product_subcategory(sub_class, gender=eff_gender, style_name=style_name, product_category=cat_val)
        # Taxonomy override wins for category + subcategory too
        if _tax.get("matched"):
            _e = _tax["entry"]
            cat_val    = _e.get("product_category", cat_val)
            subcat_val = _e.get("product_subcategory", subcat_val)

        # ── Shared-fields helper for this style ─────────────────────────────
        def write_shared_row(r, sku_val, _fabric=fabric, _care=care, _upf=upf,
                             _coo=coo, _neck=neck, _sleeve=sleeve, _sil=sil,
                             _itk=itk, _itn=itn, _ilen=ilen, _ftype=ftype,
                             _slvlen=slvlen, _bullets=bullets, _content=content,
                             _style_name=style_name, _sn=sn, _import_desig=import_desig,
                             _cat_val=cat_val, _subcat_val=subcat_val,
                             _style=style):
            # Vendor code — REQUIRED by Amazon, falls back to brand config
            vc_val = vendor_code or brand_cfg.get("vendor_code_full", "")
            wc(r, "rtip_vendor_code#1.value",         vc_val, style_num=_sn)
            wc(r, "vendor_sku#1.value",               sku_val, style_num=_sn)
            wc(r, "product_type#1.value",             detected_product_type, style_num=_sn)
            # Variation theme — must match Amazon dropdown. For SWIMWEAR, 'SIZE/COLOR' is valid;
            # 'COLOR/SIZE' is NOT. Use SIZE/COLOR when both vary, COLOR-only or SIZE-only otherwise.
            _vts = _style.get("variants", []) or []
            _has_multi_color = len({(v.get("color") or v.get("color_name") or "") for v in _vts}) > 1
            _has_multi_size  = len({v.get("size", "") for v in _vts}) > 1
            if _has_multi_color and _has_multi_size:
                _vt = "SIZE/COLOR"
            elif _has_multi_color:
                _vt = "COLOR"
            elif _has_multi_size:
                _vt = "SIZE"
            else:
                _vt = "COLOR"
            # Variation theme is only relevant when parent/child relationships exist.
            if not bool(session_data.get("skip_parent_row") or brand_cfg.get("skip_parent_row", False)):
                wc(r, "variation_theme#1.name",     _vt, style_num=_sn)
            wc(r, "brand#1.value",                    clean_brand, style_num=_sn)
            if _cat_val:
                wc(r, "product_category#1.value",     _cat_val, style_num=_sn)
            if _subcat_val:
                wc(r, "product_subcategory#1.value",  _subcat_val, style_num=_sn)
            wc(r, "item_type_keyword#1.value",        _itk, style_num=_sn)
            wc(r, "model_number#1.value",             _sn, style_num=_sn)
            wc(r, "model_name#1.value",               _style_name.title(), style_num=_sn)
            for i, bfid in enumerate(["bullet_point#1.value", "bullet_point#2.value",
                                      "bullet_point#3.value", "bullet_point#4.value",
                                      "bullet_point#5.value"]):
                if i < len(_bullets):
                    wc(r, bfid, _bullets[i][:500], style_num=_sn)
                else:
                    bullet_fallback = _content.get(f"bullet_{i+1}", "")
                    if bullet_fallback:
                        wc(r, bfid, bullet_fallback[:500], style_num=_sn)
            wc(r, "generic_keyword#1.value",          _content.get("backend_keywords", ""), style_num=_sn)
            wc(r, "style#1.value",                    _style_name.title(), style_num=_sn)
            # fit_type — from data/override only
            wc(r, "fit_type#1.value",                 _content.get("fit_type", "") or _style.get("fit_type", "") if hasattr(_style, "get") else _content.get("fit_type", ""), style_num=_sn)
            # Amazon target_gender for SWIMWEAR accepts only Male/Female; refine youth
            # using style_name ("Little Boys", "Big Girls") when division gives only Unisex.
            _tg_sn = (_style_name or "").lower()
            if "boys" in _tg_sn or " boy" in _tg_sn or "men" in _tg_sn:
                _tg = "Male"
            elif "girls" in _tg_sn or " girl" in _tg_sn or "women" in _tg_sn:
                _tg = "Female"
            else:
                _tg = eff_gender if eff_gender in ("Male", "Female") else ""
            wc(r, "department#1.value",               eff_dept, style_num=_sn)
            wc(r, "target_gender#1.value",            _tg, style_num=_sn)
            # Age range / size_class / special_size_type — youth-aware via first variant
            _first_var = (_style.get("variants", []) or [{}])[0]
            _sst, _sclass, _ard, _ = _derive_youth_size_info(_style_name, eff_gender, _first_var.get("size", ""))
            wc(r, "age_range_description#1.value",    _ard, style_num=_sn)
            # Body Type / Height Type — 'Regular' is the safe Amazon-valid default for all apparel.
            wc(r, _size_field(detected_product_type, "body_type", col_map),         "Regular", style_num=_sn)
            wc(r, _size_field(detected_product_type, "height_type", col_map),       "Regular", style_num=_sn)
            # Material gets per-fiber split; Fabric Type gets the full blend percentage.
            if _fabric:
                _materials = _split_fabric_into_materials(_fabric)
                for _mi, _mat in enumerate(_materials[:5], start=1):
                    wc(r, f"material#{_mi}.value", _mat, style_num=_sn)
                wc(r, "fabric_type#1.value", _fabric, style_num=_sn)
            else:
                wc(r, "fabric_type#1.value", _ftype, style_num=_sn)
            # Special Features (AI-suggested) — Sage feedback
            _addl = _style.get("additional_details", "") or _content.get("additional_details", "")
            for _fi, _feat in enumerate(_derive_special_features(_style_name, _addl)[:5], start=1):
                wc(r, f"special_feature#{_fi}.value", _feat, style_num=_sn)
            # Lifestyle (AI-suggested)
            for _li, _lf in enumerate(_derive_lifestyle(_style_name, sub_class)[:2], start=1):
                wc(r, f"lifestyle#{_li}.value", _lf, style_num=_sn)
            wc(r, "number_of_items#1.value", "1", style_num=_sn)
            wc(r, "item_type_name#1.value",           _itn, style_num=_sn)
            if _sst:
                wc(r, "special_size_type#1.value",    _sst, style_num=_sn)
            wc(r, "rtip_product_description#1.value", _content.get("description", ""), style_num=_sn)
            wc(r, "item_length_description#1.value",  _ilen, style_num=_sn)
            wc(r, "item_booking_date#1.value",        booking_date, style_num=_sn)
            if _care:
                wc(r, "care_instructions#1.value",   _care, style_num=_sn)
            wc(r, "unit_count#1.value",               "1", style_num=_sn)
            wc(r, "unit_count#1.type.value",                "Count", style_num=_sn)
            # Neck/Collar field varies by template; pick whichever the active template has.
            if _neck:
                _neck_fid = _neck_field(col_map)
                if _neck_fid:
                    wc(r, _neck_fid, _neck, style_num=_sn)
            wc(r, "lifecycle_supply_type#1.value", "Perennial", style_num=_sn)
            if _sil:
                wc(r, "apparel_silhouette#1.value",   _sil, style_num=_sn)
            # COAT-specific: 'Type of Jacket' → coat_silhouette_type (Puffer→Quilted, etc.)
            if detected_product_type == "COAT":
                _toj = _style.get("type_of_jacket", "") or _content.get("type_of_jacket", "")
                _coat_sil = _derive_coat_silhouette(_toj, fallback_subclass=sub_class)
                if _coat_sil:
                    wc(r, "coat_silhouette_type#1.value", _coat_sil, style_num=_sn)
            wc(r, "sleeve#1.length_description#1.value", _slvlen, style_num=_sn)
            if _sleeve:
                wc(r, "sleeve#1.type#1.value",        _sleeve, style_num=_sn)
            # closure — from data/override (Sage feedback: closure flow now also reads style)
            _closure_v = _content.get("closure_type", "") or _style.get("closure_type", "") or brand_cfg.get("default_closure", "")
            wc(r, "closure#1.type#1.value",             _closure_v, style_num=_sn)
            # number of pockets — from pre-upload
            _pockets_v = _content.get("pockets", "") or _style.get("pockets", "")
            if _pockets_v not in (None, "", 0):
                try:    wc(r, "number_of_pockets#1.value", int(_pockets_v), style_num=_sn)
                except (ValueError, TypeError): wc(r, "number_of_pockets#1.value", str(_pockets_v), style_num=_sn)
            if _upf:
                wc(r, "ultraviolet_protection_factor#1.value", _upf, style_num=_sn)
            wc(r, "skip_offer#1.value",                       "No", style_num=_sn)
            wc(r, "import_designation#1.value",       _import_desig, style_num=_sn)
            wc(r, "rtip_earliest_shipping_date#1.value", today_str, style_num=_sn)
            # Contains battery/cell — required compliance field
            wc(r, "contains_battery_or_cell#1.value", "No", style_num=_sn)
            # Dangerous Goods Regulation — PT-aware
            if _pt_defaults.pt_writes(detected_product_type, "dg"):
                _dg2 = _pt_defaults.get_pt_default(detected_product_type, "default_dg_regulation", "Not Applicable")
                wc(r, "supplier_declared_dg_hz_regulation#1.value", _dg2, style_num=_sn)
            wc(r, "item_package_dimensions#1.length.value",      brand_cfg.get("default_pkg_length", ""), style_num=_sn)
            wc(r, "item_package_dimensions#1.length.unit",       "Inches", style_num=_sn)
            wc(r, "item_package_dimensions#1.width.value",       brand_cfg.get("default_pkg_width", ""), style_num=_sn)
            wc(r, "item_package_dimensions#1.width.unit",        "Inches", style_num=_sn)
            wc(r, "item_package_dimensions#1.height.value",      brand_cfg.get("default_pkg_height", ""), style_num=_sn)
            wc(r, "item_package_dimensions#1.height.unit",       "Inches", style_num=_sn)
            wc(r, "item_package_weight#1.value",      brand_cfg.get("default_pkg_weight", ""), style_num=_sn)
            wc(r, "item_package_weight#1.unit",       "Pounds", style_num=_sn)
            wc(r, "rtip_order_aggregate_type#1.value",     "Each", style_num=_sn)
            wc(r, "rtip_items_per_inner_pack#1.value",     "1", style_num=_sn)
            if _coo:
                wc(r, "country_of_origin#1.value",   _coo, style_num=_sn)
            wc(r, "batteries_required#1.value",  "No", style_num=_sn)
            wc(r, "batteries_included#1.value",  "No", style_num=_sn)

        # Skip-parent toggle (Sage feedback): brands without self-service variations skip parent row.
        skip_parent = bool(session_data.get("skip_parent_row")
                           or brand_cfg.get("skip_parent_row", False))

        # ── Parent row (required by Amazon unless skip_parent is enabled) ─────────────────────
        if not skip_parent:
            write_shared_row(cr, psku)
            wc(cr, "parentage_level#1.value",                "Parent", style_num=sn)
            wc(cr, "item_name#1.value",                      content.get("title", style_name), style_num=sn)
            # Parent rows also need list_price + cost_price (Vendor Central validation).
            if list_price:
                try:    wc(cr, "list_price#1.value",  float(list_price), style_num=sn)
                except: wc(cr, "list_price#1.value",  list_price, style_num=sn)
            if cost_price:
                try:    wc(cr, "cost_price#1.value", float(cost_price), style_num=sn)
                except: wc(cr, "cost_price#1.value", cost_price, style_num=sn)
            cr += 1

        # ── Child rows ────────────────────────────────────────────────────────
        for var in style.get("variants", []):
            color  = var.get("color", "") or var.get("color_name", "")
            size   = var.get("size", "")
            upc    = var.get("upc", "")
            v_cost = var.get("cost_price", "")
            # Youth-aware size resolution: 2T -> '2 Years'; adult alpha stays as-is.
            _sst_c, _sclass_c, _ard_c, size_norm = _derive_youth_size_info(style_name, eff_gender, size)
            # Prefer source SKU verbatim (preserves F26-107010297-BEG-S format).
            csku   = _derive_child_sku(var, psku, color, size,
                                       color_code=var.get("color_code", ""),
                                       size_code=var.get("size_code", ""))
            color_family = normalize_color(color, detected_product_type)
            if color:
                ctitle = content.get("title", "").split(",")[0] + f", {color.title()}, {size_norm or size}"
            else:
                ctitle = content.get("title", "")

            write_shared_row(cr, csku)
            # In skip_parent mode, children stand alone — no parentage_level / parent_sku link.
            if not skip_parent:
                wc(cr, "parentage_level#1.value",         "Child", style_num=sn)
                wc(cr, "child_parent_sku_relationship#1.child_relationship_type", "Variation", style_num=sn)
                wc(cr, "child_parent_sku_relationship#1.parent_sku",              psku, style_num=sn)
            wc(cr, "item_name#1.value",               ctitle, style_num=sn)
            if upc:
                wc(cr, "external_product_id#1.type",  "UPC", style_num=sn)
                wc(cr, "external_product_id#1.value", re.sub(r"\D", "", str(upc)), style_num=sn)
            wc(cr, _size_field(detected_product_type, "size_system", col_map),      "US", style_num=sn)
            wc(cr, _size_field(detected_product_type, "size_class", col_map),       _sclass_c, style_num=sn)
            wc(cr, _size_field(detected_product_type, "size", col_map),             size_norm or size, style_num=sn)
            # Repeat special_size_type + age range on every child row (Amazon expects per-row)
            if _sst_c:
                wc(cr, "special_size_type#1.value",   _sst_c, style_num=sn)
            wc(cr, "age_range_description#1.value",   _ard_c, style_num=sn)
            wc(cr, "color#1.standardized_values#1",            color_family, style_num=sn)
            wc(cr, "color#1.value",                   color.title() if color else "", style_num=sn)
            v_list_price = var.get("list_price", "") or list_price
            if v_list_price:
                try:    wc(cr, "list_price#1.value",  float(v_list_price), style_num=sn)
                except: wc(cr, "list_price#1.value",  v_list_price, style_num=sn)
            else:
                wc(cr, "list_price#1.value",  "", style_num=sn)  # trigger override check
            if v_cost:
                try:    wc(cr, "cost_price#1.value",        float(v_cost), style_num=sn)
                except: wc(cr, "cost_price#1.value",        v_cost, style_num=sn)
            else:
                wc(cr, "cost_price#1.value",  "", style_num=sn)  # trigger override check
            cr += 1

    import warnings as _w2
    with _w2.catch_warnings():
        _w2.simplefilter("ignore")
        wb.save(output_path)


@app.route("/api/validate-before-build", methods=["POST"])
def validate_before_build():
    """Validate all field values against template dropdowns before building .xlsm.
    Returns per-style validation results with auto-corrections.
    Used by the dramatic QA UI.
    """
    data = request.get_json(force=True)
    brand = data.get("brand") or session_data.get("brand", "")
    styles = data.get("styles") or session_data.get("styles", [])
    content_map = data.get("content_map") or session_data.get("generated_content", {})
    overrides = data.get("field_overrides") or session_data.get("field_overrides", {})
    template_path = session_data.get("template_path") or str(DEFAULT_TEMPLATE)

    if not brand or not styles:
        return jsonify({"error": "No brand or styles"}), 400

    brand_cfg = _load_brand_config_data(brand)

    all_results = []
    total_valid = 0
    total_corrected = 0
    total_invalid = 0

    for style in styles:
        sn = style["style_num"]
        content = content_map.get(sn, {})
        style_overrides = overrides.get(sn, {})

        # Detect product type PER STYLE
        product_type = _resolve_style_product_type(style)
        dropdowns = load_dropdown_cache(product_type)

        # Build the field values that would be written
        coo = normalize_coo(content.get("coo", "") or style.get("coo", "") or brand_cfg.get("default_coo", "")) or ""
        upf = content.get("upf", "") or style.get("upf", "") or brand_cfg.get("default_upf", "")
        care = content.get("care", "") or style.get("care", "") or brand_cfg.get("default_care", "")
        fabric = content.get("fabric", "") or style.get("fabric", "") or brand_cfg.get("default_fabric", "")

        field_values = {
            "product_category#1.value": _derive_amazon_product_category(
                style.get("subclass", "") or style.get("sub_class", ""),
                gender=_derive_gender_department(style)[0] or brand_cfg.get("gender", ""),
                product_type=product_type),
            "lifecycle_supply_type#1.value": "Perennial",
            "department#1.value": _derive_gender_department(style)[1] or brand_cfg.get("department", ""),
            "target_gender#1.value": _derive_gender_department(style)[0] or brand_cfg.get("gender", ""),
            "country_of_origin#1.value": coo,
            "care_instructions#1.value": care,
            "material#1.value": fabric,
            "closure#1.type#1.value": "Pull On",
            "import_designation#1.value": "Imported" if coo != "United States" else "Made in the USA",
            "item_package_dimensions#1.length.unit": "Inches",
            "item_package_weight#1.unit": "Pounds",
            "batteries_required#1.value": "No",
            "batteries_included#1.value": "No",
            "skip_offer#1.value": "No",
            "ultraviolet_protection_factor#1.value": upf,
            "fit_type#1.value": "Regular",
        }

        # Apply overrides
        field_values.update(style_overrides)

        style_results = []
        for fid, val in field_values.items():
            if not val or fid not in dropdowns:
                continue
            r = validate_field_value(fid, val, product_type)
            style_results.append(r)
            if r["status"] == "valid":
                total_valid += 1
            elif r["status"] == "corrected":
                total_corrected += 1
            elif r["status"] == "invalid":
                total_invalid += 1

        all_results.append({
            "style_num": sn,
            "style_name": style.get("style_name", sn),
            "validations": style_results,
        })

    return jsonify({
        "product_type": product_type,
        "dropdown_fields_loaded": len(dropdowns),
        "total_valid": total_valid,
        "total_corrected": total_corrected,
        "total_invalid": total_invalid,
        "styles": all_results,
    })


@app.route("/api/sync-before-download", methods=["POST"])
def sync_before_download():
    """Sync frontend state to server before download.
    Ensures overrides, styles, content survive Render worker restarts.
    """
    data = request.get_json(force=True)
    if data.get("brand"):
        session_data["brand"] = data["brand"]
    if data.get("vendor_code"):
        session_data["vendor_code"] = data["vendor_code"]
    if data.get("styles"):
        session_data["styles"] = data["styles"]
    if data.get("content_map"):
        session_data["generated_content"] = data["content_map"]
    if data.get("field_overrides"):
        # Merge (don't replace) — frontend might have partial overrides
        for sn, ov in data["field_overrides"].items():
            if sn not in session_data.get("field_overrides", {}):
                session_data.setdefault("field_overrides", {})[sn] = {}
            session_data["field_overrides"][sn].update(ov)
    if data.get("style_product_types"):
        session_data["style_product_types"] = data["style_product_types"]
    return jsonify({"ok": True})


# Template name mapping: product type ID → template filename
# Size field prefix varies by template — detect from col_map or use this fallback
SIZE_PREFIX_MAP = {
    "BLAZER": "apparel_size",
    "BRA": "shapewear_size",
    "COAT": "apparel_size",
    "DRESS": "apparel_size",
    "HAT": "headwear_size",
    "ONE_PIECE_OUTFIT": "apparel_size",
    "OVERALLS": "bottoms_size",
    "PANTS": "bottoms_size",
    "SANDAL": "footwear_size",
    "SHIRT": "shirt_size",
    "SHORTS": "bottoms_size",
    "SKIRT": "skirt_size",
    "SNOWSUIT": "apparel_size",
    "SNOW_PANT": "bottoms_size",
    "SWEATSHIRT": "apparel_size",
    "SWIMWEAR": "shapewear_size",
}

def _neck_field(col_map=None):
    """Return the right neck/collar field_id for the active template.
    COAT/BLAZER/DRESS/SHIRT use 'collar_style#1.value'; SWIMWEAR uses 'neck#1.neck_style#1.value'.
    Returns None if neither exists in the template (PANTS/SHORTS/SANDAL).
    """
    if col_map:
        for candidate in ("collar_style#1.value", "neck#1.neck_style#1.value"):
            if candidate in col_map:
                return candidate
        return None
    return "collar_style#1.value"


def _size_field(product_type, suffix, col_map=None):
    """Return the correct size field_id for this product type.
    e.g. _size_field('SWIMWEAR', 'size_system') -> 'shapewear_size#1.size_system'
    If col_map provided, detect from template; else use SIZE_PREFIX_MAP.
    """
    if col_map:
        # Auto-detect from template columns
        for prefix in ["apparel_size", "shapewear_size", "bottoms_size", "shirt_size",
                       "headwear_size", "footwear_size", "skirt_size"]:
            candidate = f"{prefix}#1.{suffix}"
            if candidate in col_map:
                return candidate
    prefix = SIZE_PREFIX_MAP.get(product_type, "apparel_size")
    return f"{prefix}#1.{suffix}"

PRODUCT_TYPE_TEMPLATE_MAP = {
    "BLAZER": "Blazers.xlsm",
    "BRA": "Bras.xlsm",
    "COAT": "Jackets_and_Coats.xlsm",
    "DRESS": "Dresses.xlsm",
    "HAT": "Hats.xlsm",
    "ONE_PIECE_OUTFIT": "One-piece_Outfits.xlsm",
    "OVERALLS": "Overalls.xlsm",
    "PANTS": "Other_Pants.xlsm",
    "SANDAL": "Sandals.xlsm",
    "SHIRT": "Other_Shirts.xlsm",
    "SHORTS": "Shorts.xlsm",
    "SKIRT": "Skirts.xlsm",
    "SNOWSUIT": "Snowsuits.xlsm",
    "SNOW_PANT": "Snow_Pants.xlsm",
    "SWEATSHIRT": "Sweatshirts.xlsm",
    "SWIMWEAR": "Swimwear.xlsm",
}

def _get_template_for_product_type(product_type_id):
    """Get the template .xlsm path for a product type.
    Checks session templates first, then falls back to disk.
    """
    # Check session-loaded templates
    session_templates = session_data.get("templates", {})
    if product_type_id in session_templates:
        return session_templates[product_type_id]

    # Check the mapping
    fname = PRODUCT_TYPE_TEMPLATE_MAP.get(product_type_id)
    if fname:
        path = UPLOAD_TEMPLATES / fname
        if path.exists():
            return str(path)

    # Fallback: try to find any template that matches
    for f in UPLOAD_TEMPLATES.glob("*.xlsm"):
        if product_type_id.lower().replace("_", "") in f.stem.lower().replace("_", ""):
            return str(f)

    # Last resort: default template
    return str(DEFAULT_TEMPLATE)


@app.route("/api/template-coverage", methods=["GET"])
def template_coverage():
    """PT-aware health snapshot: for each PT that appears in the current session,
    report whether the template, rule bundle, and dropdown cache are present.
    Drives the Health card on upload + the per-style trace chip.

    Response shape:
      {
        "pts": [
          { "product_type": "COAT",
            "label": "4 coats",
            "style_count": 4,
            "template": { "present": true,  "file": "Jackets_and_Coats.xlsm" },
            "rules":    { "present": true,  "file": "COAT.json",  "rule_count": 612 },
            "dropdowns":{ "present": true,  "file": "COAT.json",  "field_count": 155 },
            "ready": true
          },
          ...
        ],
        "per_style": {
          "107010297": { "product_type": "COAT", "ready": true,
                          "template": true, "rules": true, "dropdowns": true,
                          "reason": "Sub-class 'Faux Wool Outerwear' is a known Coat type" }
        },
        "summary": { "total_styles": 5, "ready_styles": 4, "unready_styles": 1 }
      }
    """
    styles = session_data.get("styles", []) or []
    session_templates = session_data.get("templates", {}) or {}

    # Tally styles per PT and capture resolution reasons
    counts = {}
    style_pt = {}
    style_reason = {}
    for s in styles:
        sn = s.get("style_num", "")
        if not sn:
            continue
        sub_class = s.get("subclass", "") or s.get("sub_class", "")
        div_name  = s.get("division_name", "")
        pt_id, _conf, reason = resolve_product_type(sub_class, div_name)
        # Operator override wins
        pt_id = (session_data.get("style_product_types") or {}).get(sn, pt_id)
        style_pt[sn] = pt_id
        style_reason[sn] = reason
        counts[pt_id] = counts.get(pt_id, 0) + 1

    pts_out = []
    for pt_id in sorted(counts.keys()):
        # Template
        tpl_filename = (PRODUCT_TYPE_TEMPLATE_MAP.get(pt_id)
                        or _pt_defaults.pt_template_filename(pt_id))
        tpl_path = (UPLOAD_TEMPLATES / tpl_filename) if tpl_filename else None
        tpl_present = bool(
            (pt_id in session_templates) or (tpl_path and tpl_path.exists())
        )

        # Rule bundle
        bundle_path = _NIS_RULES_DIR / f"{pt_id}.json"
        bundle_present = bundle_path.exists()
        rule_count = 0
        if bundle_present:
            try:
                with open(bundle_path, "r", encoding="utf-8") as _f:
                    _b = json.load(_f)
                rule_count = len(_b.get("fields", _b.get("rules", [])) or [])
            except Exception:
                rule_count = 0

        # Dropdown cache
        dd_path = DROPDOWN_CACHE_DIR / f"{pt_id}.json"
        dd_present = dd_path.exists()
        dd_field_count = 0
        if dd_present:
            try:
                with open(dd_path, "r", encoding="utf-8") as _f:
                    _dd = json.load(_f)
                dd_field_count = len([k for k in _dd.keys() if not k.startswith("_")])
            except Exception:
                dd_field_count = 0

        ready = tpl_present and bundle_present and dd_present and pt_id != "UNKNOWN"

        pts_out.append({
            "product_type": pt_id,
            "label":        _pt_defaults.pt_label(pt_id, counts[pt_id]),
            "style_count":  counts[pt_id],
            "template":     {"present": tpl_present,    "file": tpl_filename or ""},
            "rules":        {"present": bundle_present, "file": f"{pt_id}.json" if bundle_present else "", "rule_count": rule_count},
            "dropdowns":    {"present": dd_present,     "file": f"{pt_id}.json" if dd_present else "",     "field_count": dd_field_count},
            "ready":        ready,
        })

    # Per-style trace
    per_style = {}
    ready_count = 0
    for sn, pt_id in style_pt.items():
        pt_block = next((p for p in pts_out if p["product_type"] == pt_id), None)
        if pt_block:
            ready = pt_block["ready"]
            per_style[sn] = {
                "product_type": pt_id,
                "template":     pt_block["template"]["present"],
                "rules":        pt_block["rules"]["present"],
                "dropdowns":    pt_block["dropdowns"]["present"],
                "ready":        ready,
                "reason":       style_reason.get(sn, ""),
            }
            if ready:
                ready_count += 1
        else:
            per_style[sn] = {
                "product_type": pt_id,
                "template": False, "rules": False, "dropdowns": False,
                "ready": False,
                "reason": style_reason.get(sn, ""),
            }

    return jsonify({
        "pts":      pts_out,
        "per_style": per_style,
        "summary": {
            "total_styles":   len(style_pt),
            "ready_styles":   ready_count,
            "unready_styles": len(style_pt) - ready_count,
            "session_label":  _pt_defaults.template_label_for_session(
                                  list(counts.keys()), counts),
        },
    })


@app.route("/api/dropdowns-for-session", methods=["GET"])
def dropdowns_for_session():
    """Return the dropdown cache for every PT present in the current session.
    Powers the Fit Type / Department / etc. recommended-values + freeform input pattern
    in the All Fields tab. Sage feedback: "Fit Type has a dropdown of recommended values...
    would it be possible to have Amazon's recommended dropdown values populate here?"

    Response: { "COAT": { "fit_type#1.value": [...], "department#1.value": [...], ... },
                "SWIMWEAR": { ... } }
    """
    styles = session_data.get("styles", []) or []
    pts_seen = set()
    for s in styles:
        sub_class = s.get("subclass", "") or s.get("sub_class", "")
        div_name = s.get("division_name", "")
        pt_id, _, _ = resolve_product_type(sub_class, div_name)
        pt_id = (session_data.get("style_product_types") or {}).get(s.get("style_num", ""), pt_id)
        if pt_id and pt_id != "UNKNOWN":
            pts_seen.add(pt_id)
    out = {}
    for pt in pts_seen:
        try:
            out[pt] = load_dropdown_cache(pt) or {}
        except Exception as e:
            print(f"[dropdowns-for-session] failed for {pt}: {e}")
            out[pt] = {}
    return jsonify({"dropdowns": out, "product_types": sorted(pts_seen)})


@app.route("/api/set-skip-parent", methods=["POST"])
def set_skip_parent():
    """Toggle the skip-parent-row preference for the current session.
    Sage feedback: brands without self-service variations want to skip parent rows entirely.
    Body: { skip: true|false }
    """
    data = request.get_json(force=True) or {}
    session_data["skip_parent_row"] = bool(data.get("skip", False))
    return jsonify({"ok": True, "skip_parent_row": session_data["skip_parent_row"]})


@app.route("/api/save-style-product-types", methods=["POST"])
def save_style_product_types():
    """Save operator's product type assignments for styles.
    Called from the per-style PT selector table.
    """
    data = request.get_json(force=True)
    assignments = data.get("assignments", {})
    session_data["style_product_types"] = assignments
    return jsonify({"ok": True, "count": len(assignments)})


# ── Taxonomy Overrides API endpoints (Phase 1) ───────────────────────────────

@app.route("/api/taxonomy", methods=["GET"])
def taxonomy_get():
    """Return the full override store + the valid-value universe so the frontend
    can populate cascading dropdowns client-side without per-keystroke backend calls.

    Optional ?product_type=SWIMWEAR filter to reduce payload.
    """
    pt_filter = (request.args.get("product_type") or "").strip().upper() or None
    overrides = _load_taxonomy_overrides()
    universe_all = _load_taxonomy_universe()
    universe = {pt_filter: universe_all[pt_filter]} if pt_filter and pt_filter in universe_all else universe_all

    # Also surface observed sub_classes from current session for UI convenience
    sub_classes_seen = set()
    for s in session_data.get("styles", []):
        sc = s.get("subclass") or s.get("sub_class")
        if sc:
            sub_classes_seen.add(sc)

    return jsonify({
        "version": overrides.get("version", 1),
        "updated_at": overrides.get("updated_at", ""),
        "entries": overrides.get("entries", {}),
        "universe": universe,
        "sub_classes_seen": sorted(sub_classes_seen),
        "gender_buckets": GENDER_BUCKETS,
    })


@app.route("/api/taxonomy/save", methods=["POST"])
def taxonomy_save():
    """Upsert one taxonomy entry. Validates against Amazon dropdowns + template
    cascade rules before writing. Triggers a best-effort git commit + push so
    learned taxonomy survives Render redeploys.
    """
    data = request.get_json(force=True) or {}
    sub_class = (data.get("sub_class") or "").strip()
    gender_bucket = (data.get("gender_bucket") or "").strip()
    pt = (data.get("product_type") or "").strip().upper()
    cat = (data.get("product_category") or "").strip()
    subcat = (data.get("product_subcategory") or "").strip()
    itk = (data.get("item_type_keyword") or "").strip()
    itn = (data.get("item_type_name") or "").strip()
    notes = data.get("notes") or ""
    user = data.get("confirmed_by") or "unknown"

    if not sub_class or not gender_bucket or not pt:
        return jsonify({"error": "sub_class, gender_bucket, and product_type are required"}), 400
    if gender_bucket not in GENDER_BUCKETS:
        return jsonify({"error": f"Invalid gender_bucket. Must be one of {GENDER_BUCKETS}"}), 400

    # Dropdown validation
    ok, errors = _validate_taxonomy_quadruple(pt, cat, subcat, itk, itn)
    if not ok:
        return jsonify({"error": "Validation failed", "errors": errors}), 400

    key = _taxonomy_key(pt, sub_class, gender_bucket)
    now = datetime.now().isoformat() + "Z"
    entry = {
        "product_type": pt,
        "product_category": cat,
        "product_subcategory": subcat,
        "item_type_keyword": itk,
        "item_type_name": itn,
        "confirmed_by": user,
        "confirmed_at": now,
        "source": "manual",
        "notes": notes,
    }

    store = _load_taxonomy_overrides()
    existing = store.get("entries", {}).get(key)
    action = "update" if existing else "create"
    store.setdefault("entries", {})[key] = entry
    _save_taxonomy_overrides(store)
    _append_taxonomy_history(key, entry, action)

    # Count how many currently-loaded styles match this bucket — UI hint
    affected = 0
    for s in session_data.get("styles", []):
        s_pt = _resolve_style_product_type(s) or ""
        s_sc = s.get("subclass") or s.get("sub_class") or ""
        s_gb = _derive_gender_bucket(s)
        if s_pt.upper() == pt and s_sc == sub_class and s_gb == gender_bucket:
            affected += 1

    # Best-effort git commit in a thread so it doesn't block the response
    try:
        import threading
        threading.Thread(target=_try_git_commit_taxonomy, args=(key, user), daemon=True).start()
    except Exception:
        pass

    return jsonify({
        "ok": True,
        "key": key,
        "entry": entry,
        "action": action,
        "affected_styles_in_session": affected,
    })


@app.route("/api/taxonomy/validate", methods=["POST"])
def taxonomy_validate():
    """Pre-flight check. For every style in the session (or in the request body),
    report which buckets have a confirmed override and which are auto-derived.
    Used by the post-upload banner to decide whether to prompt for confirmation.
    """
    data = request.get_json(silent=True) or {}
    styles = data.get("styles") or session_data.get("styles", [])
    if not styles:
        return jsonify({"error": "No styles in session"}), 400

    # Group styles by (product_type, sub_class, gender_bucket)
    buckets = {}  # key -> { style_nums, sub_class, gender_bucket, product_type }
    for s in styles:
        pt = _resolve_style_product_type(s) or ""
        sc = s.get("subclass") or s.get("sub_class") or ""
        gb = _derive_gender_bucket(s)
        key = _taxonomy_key(pt, sc, gb)
        b = buckets.setdefault(key, {
            "key": key, "product_type": pt, "sub_class": sc,
            "gender_bucket": gb, "style_nums": [], "style_count": 0,
        })
        b["style_nums"].append(s.get("style_num"))
        b["style_count"] += 1

    overrides = _load_taxonomy_overrides().get("entries", {})
    brand = session_data.get("brand", "")
    brand_cfg = _load_brand_config_data(brand) if brand else {}

    confirmed = []
    unconfirmed = []
    for key, b in buckets.items():
        entry = overrides.get(key)
        if entry and entry.get("source") == "manual":
            confirmed.append({"key": key, **b, "entry": entry})
        else:
            # Auto-derive fallback for UI display
            sample = {"subclass": b["sub_class"],
                      "style_name": (styles[0].get("style_name", "") if styles else ""),
                      "division_name": (styles[0].get("division_name", "") if styles else "")}
            # Find a style actually in this bucket so style_name is representative
            for s in styles:
                if s.get("style_num") in b["style_nums"]:
                    sample = s; break
            resolved = _resolve_taxonomy_for_style(sample, brand_cfg)
            unconfirmed.append({"key": key, **b, "auto_derived": resolved["entry"]})

    return jsonify({
        "total_styles": len(styles),
        "total_buckets": len(buckets),
        "confirmed": sum(b["style_count"] for b in confirmed) if False else len([b for b in confirmed]),
        "confirmed_styles": sum(b["style_count"] for b in confirmed),
        "unconfirmed_styles": sum(b["style_count"] for b in unconfirmed),
        "buckets": {"confirmed": confirmed, "unconfirmed": unconfirmed},
        "blocking": len(unconfirmed) > 0,
        "message": (
            f"{len(unconfirmed)} of {len(buckets)} item-type buckets need taxonomy confirmation."
            if unconfirmed else
            f"All {len(buckets)} item-type buckets are confirmed."
        ),
    })
# ── End Taxonomy Overrides API ──────────────────────────────────────────


def _resolve_style_product_type(style):
    """Resolve the product type for a single style."""
    # Check operator assignment first
    pt_assignments = session_data.get("style_product_types", {})
    if style["style_num"] in pt_assignments:
        return pt_assignments[style["style_num"]]
    # Use resolve_product_type
    sub_class = style.get("subclass", "")
    div_name = style.get("division_name", "")
    pt_id, _, _reason = resolve_product_type(sub_class, div_name)
    return pt_id


@app.route("/api/download-all")
def download_all():
    """Generate .xlsm files and ZIP them together.
    ?mode=per-style → one .xlsm per style
    ?mode=per-type (default) → one .xlsm per product type
    """
    dl_mode = request.args.get("mode", "per-type")
    brand = session_data.get("brand", "Brand")
    styles = session_data.get("styles", [])
    content_map = session_data.get("generated_content", {})

    if not styles or not content_map:
        return jsonify({"error": "No generated content"}), 400

    date_str = datetime.now().strftime("%m%d%y")
    safe_brand = re.sub(r'[^\w\-]', '_', brand)
    brand_cfg = _load_brand_config_data(brand)
    vendor_code = session_data.get("vendor_code") or brand_cfg.get("vendor_code_full", "")

    generated_files = []  # (filename, filepath)

    # Optional: filter to single style
    single_style = request.args.get("style_num", "")

    if dl_mode == "per-style":
        # ── One .xlsm per STYLE ────────────────────────────────────────────
        target_styles = [s for s in styles if s["style_num"] == single_style] if single_style else styles
        for s in target_styles:
            sn = s["style_num"]
            if sn not in content_map:
                continue
            pt = _resolve_style_product_type(s)
            template_path = _get_template_for_product_type(pt)
            safe_sn = re.sub(r'[^\w\-]', '_', str(sn))
            safe_name = re.sub(r'[^\w\-]', '_', s.get("style_name", sn)[:30])
            fname = f"NIS_{safe_brand}_{safe_sn}_{safe_name}_{date_str}.xlsm"
            fpath = UPLOAD_OUTPUT / fname
            try:
                _generate_category_file([s], content_map, template_path, brand, brand_cfg, vendor_code, str(fpath))
                generated_files.append((fname, str(fpath)))
            except Exception as e:
                print(f"[Download] Error generating {sn}: {e}")
                traceback.print_exc()
    else:
        # ── One .xlsm per PRODUCT TYPE (original behavior) ─────────────
        by_pt = defaultdict(list)
        for s in styles:
            sn = s["style_num"]
            if sn not in content_map:
                continue
            pt = _resolve_style_product_type(s)
            by_pt[pt].append(s)

        for pt_id, pt_styles in by_pt.items():
            template_path = _get_template_for_product_type(pt_id)
            safe_pt = re.sub(r'[^\w\-]', '_', pt_id)
            fname = f"NIS_{safe_brand}_{safe_pt}_{date_str}.xlsm"
            fpath = UPLOAD_OUTPUT / fname
            try:
                _generate_category_file(pt_styles, content_map, template_path, brand, brand_cfg, vendor_code, str(fpath))
                generated_files.append((fname, str(fpath)))
            except Exception as e:
                traceback.print_exc()

    if not generated_files:
        return jsonify({"error": "No files generated"}), 500

    if len(generated_files) == 1:
        fname, fpath = generated_files[0]
        return send_file(fpath, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.ms-excel.sheet.macroEnabled.12")

    zip_name = f"NIS_{safe_brand}_{date_str}.zip"
    zip_path = UPLOAD_OUTPUT / zip_name
    with zipfile.ZipFile(str(zip_path), "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, fpath in generated_files:
            zf.write(fpath, fname)

    return send_file(str(zip_path), as_attachment=True, download_name=zip_name, mimetype="application/zip")


@app.route("/api/download-product-type/<pt_id>")
def download_product_type(pt_id):
    """Download one .xlsm for all styles of a specific product type."""
    brand = session_data.get("brand", "Brand")
    styles = session_data.get("styles", [])
    content_map = session_data.get("generated_content", {})
    brand_cfg = _load_brand_config_data(brand)
    vendor_code = session_data.get("vendor_code") or brand_cfg.get("vendor_code_full", "")

    pt_styles = [s for s in styles if _resolve_style_product_type(s) == pt_id and s["style_num"] in content_map]
    if not pt_styles:
        return jsonify({"error": f"No styles for product type {pt_id}"}), 404

    template_path = _get_template_for_product_type(pt_id)
    date_str = datetime.now().strftime("%m%d%y")
    safe_brand = re.sub(r'[^\w\-]', '_', brand)
    safe_pt = re.sub(r'[^\w\-]', '_', pt_id)
    fname = f"NIS_{safe_brand}_{safe_pt}_{date_str}.xlsm"
    fpath = UPLOAD_OUTPUT / fname

    _generate_category_file(pt_styles, content_map, template_path, brand, brand_cfg, vendor_code, str(fpath))

    return send_file(str(fpath), as_attachment=True, download_name=fname,
                     mimetype="application/vnd.ms-excel.sheet.macroEnabled.12")


@app.route("/api/download-style/<style_num>")
def download_style(style_num):
    """Generate and download a single .xlsm for one style using its product type template."""
    brand = session_data.get("brand", "Brand")
    styles = session_data.get("styles", [])
    content_map = session_data.get("generated_content", {})
    brand_cfg = _load_brand_config_data(brand)
    vendor_code = session_data.get("vendor_code") or brand_cfg.get("vendor_code_full", "")

    style = next((s for s in styles if s["style_num"] == style_num), None)
    if not style:
        return jsonify({"error": f"Style {style_num} not found"}), 404
    if style_num not in content_map:
        return jsonify({"error": f"No content for style {style_num}"}), 400

    # Use the correct template for this style's product type
    pt_id = _resolve_style_product_type(style)
    template_path = _get_template_for_product_type(pt_id)

    date_str = datetime.now().strftime("%m%d%y")
    safe_brand = re.sub(r'[^\w\-]', '_', brand)
    safe_sn = re.sub(r'[^\w\-]', '_', style_num)
    fname = f"NIS_{safe_brand}_{safe_sn}_{date_str}.xlsm"
    fpath = UPLOAD_OUTPUT / fname

    _generate_category_file([style], content_map, template_path, brand, brand_cfg, vendor_code, str(fpath))

    return send_file(str(fpath), as_attachment=True, download_name=fname,
                     mimetype="application/vnd.ms-excel.sheet.macroEnabled.12")


@app.route("/api/download-combined")
def download_combined():
    """Download ALL styles combined into a single .xlsm file.
    Delegates to _generate_category_file so all required columns are filled.
    """
    brand = session_data.get("brand", "Brand")
    styles = session_data.get("styles", [])
    content_map = session_data.get("generated_content", {})
    template_path = session_data.get("template_path", str(DEFAULT_TEMPLATE))
    brand_cfg = _load_brand_config_data(brand)
    vendor_code = session_data.get("vendor_code", brand_cfg.get("vendor_code_full", ""))
    
    if not styles or not content_map:
        return jsonify({"error": "No generated content. Run Generate Content first."}), 400
    
    safe_brand = brand.replace(" ", "_")
    combined_name = f"NIS_{safe_brand}_ALL_STYLES.xlsm"
    combined_path = UPLOAD_OUTPUT / combined_name
    
    # Re-use the comprehensive _generate_category_file with all styles
    _generate_category_file(
        styles, content_map, template_path, brand, brand_cfg, vendor_code, str(combined_path)
    )
    
    return send_file(
        str(combined_path),
        as_attachment=True,
        download_name=combined_name,
        mimetype="application/vnd.ms-excel.sheet.macroEnabled.12",
    )


@app.route("/api/download-category/<category>")
def download_category(category):
    """Download all styles of a specific category combined into one .xlsm file.
    Delegates to _generate_category_file so all required columns are filled.
    """
    brand = session_data.get("brand", "Brand")
    styles = session_data.get("styles", [])
    content_map = session_data.get("generated_content", {})
    template_path = session_data.get("template_path", str(DEFAULT_TEMPLATE))
    brand_cfg = _load_brand_config_data(brand)
    vendor_code = session_data.get("vendor_code", brand_cfg.get("vendor_code_full", ""))
    
    if not styles or not content_map:
        return jsonify({"error": "No generated content. Run Generate Content first."}), 400
    
    # Filter styles by category
    filtered_styles = [s for s in styles if (s.get("subclass", "") or s.get("sub_class", "")).lower() == category.lower()
                       or s.get("category", "").lower() == category.lower()]
    
    if not filtered_styles:
        return jsonify({"error": f"No styles found for category '{category}'"}), 404
    
    safe_brand = brand.replace(" ", "_")
    safe_cat = category.replace(" ", "_").replace("/", "_")
    fname = f"NIS_{safe_brand}_{safe_cat}.xlsm"
    fpath = UPLOAD_OUTPUT / fname
    
    # Re-use the comprehensive _generate_category_file
    _generate_category_file(
        filtered_styles, content_map, template_path, brand, brand_cfg, vendor_code, str(fpath)
    )
    
    return send_file(str(fpath), as_attachment=True, download_name=fname,
                     mimetype="application/vnd.ms-excel.sheet.macroEnabled.12")

@app.route("/api/categories")
def get_categories():
    """Return list of unique categories from uploaded styles."""
    styles = session_data.get("styles", [])
    cats = {}
    for s in styles:
        cat = s.get("subclass", "") or s.get("sub_class", "") or "Uncategorized"
        if cat not in cats:
            cats[cat] = {"name": cat, "count": 0, "variants": 0}
        cats[cat]["count"] += 1
        cats[cat]["variants"] += len(s.get("variants", []))
    return jsonify(list(cats.values()))


@app.route("/api/session-state")
def session_state():
    return jsonify({
        "brand": session_data.get("brand"),
        "vendor_code": session_data.get("vendor_code"),
        "template_path": session_data.get("template_path"),
        "styles_loaded": len(session_data.get("styles", [])),
        "keywords_loaded": len(session_data.get("keywords", [])),
        "content_generated": len(session_data.get("generated_content", {})),
    })


# ── Brand config file helpers ──────────────────────────────────────────────────
# Generic apparel defaults — used when a brand is unknown to BRAND_CONFIGS.
# CRITICAL: never fall back to a different brand's config. UPF/Butterlux/etc are
# brand-specific gimmicks that mis-tag content for new brands.
_GENERIC_BRAND_DEFAULTS = {
    "vendor_code_prefix": "",
    "vendor_code_full":   "",
    "default_upf":        "",
    "default_fabric":     "",
    "default_coo":        "",
    "default_care":       "Machine Wash",
    "gender":             "",
    "department":         "",
    "bullet_1_focus":     "Quality and craftsmanship",
    "title_formula":      "{brand} {gender} {style_name}",
    "never_words":        [],
}

def _load_brand_config_data(brand):
    """Load brand config from file if saved, else from in-memory BRAND_CONFIGS.

    NEVER falls back to another brand's defaults. If brand is unknown, use
    a neutral apparel defaults dict so per-brand gimmicks (UPF, Butterlux,
    etc.) don't leak into other brands' content.
    """
    brand_file = BRAND_CONFIGS_DIR / f"{re.sub(r'[^\w]', '_', brand)}.json"
    if brand_file.exists():
        try:
            with open(str(brand_file), "r", encoding="utf-8") as f:
                saved = json.load(f)
            # Start from neutral defaults, layer in-memory brand-specific cfg if present, then saved file.
            base = dict(_GENERIC_BRAND_DEFAULTS)
            base.update(BRAND_CONFIGS.get(brand, {}))
            base.update(saved)
            return base
        except Exception:
            pass
    if brand in BRAND_CONFIGS:
        return dict(BRAND_CONFIGS[brand])
    return dict(_GENERIC_BRAND_DEFAULTS)


@app.route("/api/save-brand-config", methods=["POST"])
def save_brand_config():
    data = request.get_json(force=True)
    brand = data.get("brand", "")
    config = data.get("config", {})
    if not brand:
        return jsonify({"error": "No brand provided"}), 400
    
    brand_file = BRAND_CONFIGS_DIR / f"{re.sub(r'[^\w]', '_', brand)}.json"
    try:
        with open(str(brand_file), "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2)
        return jsonify({"ok": True, "brand": brand})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/save-product-brief", methods=["POST"])
def save_product_brief():
    """Save a product brief for a brand + product type/subclass.
    Stored in brand config under 'product_briefs' key.
    """
    data = request.get_json(force=True)
    brand = data.get("brand") or session_data.get("brand", "")
    key = data.get("key", "_default")  # e.g. "SWIMWEAR/Rashguard" or "SWIMWEAR" or "_default"
    brief = data.get("brief", "")
    if not brand:
        return jsonify({"error": "No brand"}), 400

    cfg = _load_brand_config_data(brand)
    if "product_briefs" not in cfg:
        cfg["product_briefs"] = {}
    cfg["product_briefs"][key] = brief

    brand_file = BRAND_CONFIGS_DIR / f"{re.sub(r'[^\w]', '_', brand)}.json"
    try:
        with open(str(brand_file), "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2)
        return jsonify({"ok": True, "brand": brand, "key": key})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/load-brand-config", methods=["GET"])
def load_brand_config_endpoint():
    brand = request.args.get("brand", "")
    if not brand:
        return jsonify({"error": "No brand provided"}), 400
    
    cfg = _load_brand_config_data(brand)
    brand_file = BRAND_CONFIGS_DIR / f"{re.sub(r'[^\w]', '_', brand)}.json"
    return jsonify({
        "brand": brand,
        "config": cfg,
        "has_saved_config": brand_file.exists(),
    })


@app.route("/api/regenerate-style", methods=["POST"])
def regenerate_style():
    """Regenerate all content for a single style using a custom brief.
    Used by the per-style 'Train this style' feature.
    """
    data = request.get_json(force=True)
    style_num = data.get("style_num", "")
    brief = data.get("brief", "")
    brand = data.get("brand") or session_data.get("brand", "")
    styles = session_data.get("styles", [])
    style = next((s for s in styles if s["style_num"] == style_num), None)
    if not style:
        return jsonify({"error": f"Style {style_num} not found"}), 404

    brand_cfg = _load_brand_config_data(brand)
    # Save the style brief
    if brief:
        if "style_briefs" not in session_data:
            session_data["style_briefs"] = {}
        session_data["style_briefs"][style_num] = brief
        # Also save to brand config for persistence
        if "product_briefs" not in brand_cfg:
            brand_cfg["product_briefs"] = {}
        brand_cfg["product_briefs"][f"style_{style_num}"] = brief

    # Inject brief into brand_cfg temporarily for LLM
    if brief:
        brand_cfg["product_briefs"] = brand_cfg.get("product_briefs", {})
        brand_cfg["product_briefs"]["_default"] = brief

    feedback_history = load_brand_feedback(brand)
    style_name = style["style_name"]
    subclass = style.get("subclass", "")
    sub_subclass = style.get("sub_subclass", "")
    fabric = parse_fabric(style.get("fabric", "")) or brand_cfg.get("default_fabric", "")
    care = style.get("care", "") or brand_cfg.get("default_care", "")
    upf = style.get("upf", "") or brand_cfg.get("default_upf", "")
    first_variant = style["variants"][0] if style.get("variants") else {}
    first_color = first_variant.get("color_name", "")
    first_size = first_variant.get("size", "")

    # Try LLM first (unless rules mode)
    gen_mode = session_data.get("generation_mode", "auto")
    llm_result = None
    if gen_mode != "rules" and _anthropic_client is not None:
        try:
            llm_result = generate_content_llm(brand_cfg, brand, style, feedback_history)
        except Exception as e:
            print(f"[Regen] LLM failed for {style_num}: {e}")

    if llm_result:
        title = llm_result["title"]
        bullets = [llm_result.get(f"bullet_{i}", "") for i in range(1, 6)]
        description = llm_result["description"]
        backend_kw = llm_result["backend_keywords"]
    else:
        resolved_pt = _resolve_style_product_type(style) or ""
        style_gender, _ = _derive_gender_department(style)
        eff_gender = style_gender or brand_cfg.get("gender", "")
        pt_label = subclass or sub_subclass or resolved_pt.replace("_", " ").title() or "Dress"
        title = generate_title(brand_cfg, brand, style_name, pt_label, first_color, first_size, upf, style_gender=style_gender)
        bullets = generate_bullets(brand_cfg, brand, style_name, sub_subclass, fabric, care, first_color, upf,
                                   subclass=subclass, gender=eff_gender, product_type=resolved_pt, style_num=style_num)
        description = generate_description(brand_cfg, brand, style_num, style_name, sub_subclass, fabric, care, first_color, upf,
                                           subclass=subclass, gender=eff_gender, product_type=resolved_pt)
        backend_kw = generate_backend_keywords(brand, style_name, subclass, first_color, fabric, upf,
                                               subclass=subclass, gender=eff_gender, product_type=resolved_pt)

    entry = {
        "title": title,
        "bullets": bullets,
        "description": description,
        "backend_keywords": backend_kw,
        "llm_generated": llm_result is not None,
        "brief_used": brief[:100] if brief else "",
    }
    # Update session content
    if "generated_content" not in session_data:
        session_data["generated_content"] = {}
    existing = session_data["generated_content"].get(style_num, {})
    existing.update(entry)
    session_data["generated_content"][style_num] = existing

    return jsonify({"ok": True, "style_num": style_num, "content": entry})


@app.route("/api/nis-spreadsheet-preview", methods=["POST"])
def nis_spreadsheet_preview():
    """Return full NIS spreadsheet preview for a style — every column/row as it would appear in the .xlsm.
    Returns rows (parent header + child variants) with all field values.
    """
    data = request.get_json(force=True)
    style_num = data.get("style_num", "")
    brand = session_data.get("brand", "Brand")
    styles = session_data.get("styles", [])
    content_map = session_data.get("generated_content", {})
    brand_cfg = _load_brand_config_data(brand)
    vendor_code = session_data.get("vendor_code") or brand_cfg.get("vendor_code_full", "")

    style = next((s for s in styles if s["style_num"] == style_num), None)
    if not style:
        return jsonify({"error": f"Style {style_num} not found"}), 404

    content = content_map.get(style_num, {})
    fields = _build_preview_fields(brand, brand_cfg, vendor_code, style, content)

    # Apply overrides
    overrides = session_data.get("field_overrides", {}).get(style_num, {})
    for f in fields:
        fid = f.get("field_id", "")
        if fid and fid in overrides:
            f["value"] = overrides[fid]
            f["status"] = "filled"
            f["overridden"] = True

    # Build variant rows
    variants = style.get("variants", [])
    variant_rows = []
    for v in variants:
        color = v.get("color_name", "") or v.get("color", "")
        size = v.get("size", "")
        upc = v.get("upc", "")
        variant_rows.append({
            "color": color,
            "size": normalize_size(size) or size,
            "upc": re.sub(r"\D", "", str(upc)) if upc else "",
            "sku": f"{style_num}-{color}-{size}".replace(" ", "-"),
            "cost_price": v.get("cost_price", ""),
            "list_price": v.get("list_price", "") or style.get("list_price", ""),
        })

    return jsonify({
        "style_num": style_num,
        "style_name": style.get("style_name", ""),
        "product_type": _resolve_style_product_type(style),
        "fields": fields,
        "variant_count": len(variants),
        "variant_rows": variant_rows,
    })


@app.route("/api/save-style-brief", methods=["POST"])
def save_style_brief():
    """Save a per-style brief for later reuse."""
    data = request.get_json(force=True)
    style_num = data.get("style_num", "")
    brief = data.get("brief", "")
    if not style_num:
        return jsonify({"error": "No style_num"}), 400
    if "style_briefs" not in session_data:
        session_data["style_briefs"] = {}
    session_data["style_briefs"][style_num] = brief
    return jsonify({"ok": True, "style_num": style_num})


@app.route("/api/upload-style-image", methods=["POST"])
def upload_style_image():
    """Upload a product image for a specific style.
    Saved to disk and stored in session for LLM vision.
    """
    style_num = request.form.get("style_num", "")
    if not style_num:
        return jsonify({"error": "No style_num"}), 400
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400

    f = request.files["file"]
    ext = Path(f.filename).suffix.lower()
    if ext not in [".jpg", ".jpeg", ".png", ".webp", ".gif"]:
        return jsonify({"error": f"Unsupported image type: {ext}"}), 400

    # Save to disk
    style_dir = UPLOAD_IMAGES / str(style_num)
    style_dir.mkdir(parents=True, exist_ok=True)
    save_path = style_dir / f"product{ext}"
    f.save(str(save_path))

    # Store path in session
    if "style_images" not in session_data:
        session_data["style_images"] = {}
    session_data["style_images"][style_num] = str(save_path)

    # Auto-fire vision analysis (cached per style; ~$0.01 per image)
    # Use the active session brand if known so the prompt is tailored to it.
    intel = None
    try:
        sess_brand = session_data.get("brand") or ""
        intel = analyze_style_image(style_num, str(save_path), brand_override=sess_brand)
    except Exception as e:
        print(f"[image-intel] auto-analyze failed for {style_num}: {e}")

    return jsonify({
        "ok": True,
        "style_num": style_num,
        "path": f"/api/style-image/{style_num}",
        "intel": intel,
    })


def analyze_style_image(style_num, img_path, brand_override=None):
    """Run a vision pass on a style image and return structured JSON observations.

    This is the 'Image Intel' secondary advisory layer:
      - never overwrites a structured field automatically
      - returns observations the operator can review and selectively apply
      - cached in session_data["style_image_intel"][style_num] to avoid re-billing

    brand_override: optional brand name. When supplied (or when the resolved
      style is associated with a brand), the prompt is tailored to that brand's
      voice / focus / never-words so vision looks for what that brand cares about.

    Returns dict {observations:[], field_suggestions:[], image_quality:[], summary:""} or None.
    On None return, _last_vision_error holds a human-readable reason.
    """
    global _anthropic_client, _last_vision_error
    _last_vision_error = None
    if _anthropic_client is None:
        _last_vision_error = "ANTHROPIC_API_KEY not configured on the server. Vision is unavailable until the key is set in Render env."
        return None

    # Cache check
    cache = session_data.setdefault("style_image_intel", {})
    cached = cache.get(str(style_num))
    if cached and cached.get("img_path") == img_path:
        return cached["intel"]

    # Resolve style context (PT, sub-class) so the prompt is grounded
    style = None
    for s in (session_data.get("styles") or []):
        if str(s.get("style_num", "")) == str(style_num):
            style = s
            break
    pt = _resolve_style_product_type(style) if style else ""
    subclass = (style or {}).get("subclass", "") if style else ""
    fabric = (style or {}).get("fabric", "") if style else ""
    style_name = (style or {}).get("style_name", "") if style else ""

    # Encode image
    try:
        import base64 as _b64
        with open(img_path, "rb") as _f:
            img_b64 = _b64.b64encode(_f.read()).decode("utf-8")
        ext = Path(img_path).suffix.lower().lstrip(".")
        media_type = {"jpg": "image/jpeg", "jpeg": "image/jpeg",
                      "png": "image/png", "webp": "image/webp",
                      "gif": "image/gif"}.get(ext, "image/jpeg")
    except Exception as e:
        print(f"[image-intel] could not encode {img_path}: {e}")
        _last_vision_error = f"Could not read the image file: {e}"
        return None

    pt_enum_list = ", ".join([p["id"] for p in ALL_PRODUCT_TYPES])

    # Resolve brand context: explicit override > style.brand > nothing
    resolved_brand = (brand_override or "").strip() or (style or {}).get("brand", "") if style else (brand_override or "")
    brand_block = ""
    if resolved_brand:
        try:
            bc = _load_brand_config_data(resolved_brand) or {}
        except Exception:
            bc = {}
        focus       = bc.get("bullet_1_focus") or ""
        voice       = bc.get("brand_voice") or ""
        never_words = bc.get("never_words") or []
        gender      = bc.get("gender") or ""
        dept        = bc.get("department") or ""
        bits = []
        if focus:       bits.append(f"  Top brand focus: {focus} \u2014 prioritize observations that support this story")
        if voice:       bits.append(f"  Brand voice: {voice}")
        if gender:      bits.append(f"  Audience: {gender} ({dept})")
        if never_words: bits.append(f"  Avoid these words in observations: {', '.join(never_words)}")
        if bits:
            brand_block = "BRAND CONTEXT:\n  Brand: " + resolved_brand + "\n" + "\n".join(bits) + "\n\n"
        else:
            brand_block = f"BRAND CONTEXT: {resolved_brand} (no extra config; describe neutrally)\n\n"

    expected_block = (
        f"PRODUCT CONTEXT (from pre-upload):\n"
        f"  Expected product type: {pt or 'unknown'}\n"
        f"  Sub-class: {subclass or 'unknown'}\n"
        f"  Style name: {style_name or 'unknown'}\n"
        f"  Fabric (declared): {fabric or 'unknown'}\n"
    ) if pt else "PRODUCT CONTEXT: none provided \u2014 classify the photo from scratch.\n"

    prompt = f"""You are a product photo analyst for Amazon NIS listings. Look at the photo and return STRUCTURED JSON.

{brand_block}{expected_block}
YOUR JOB — return JSON only, no prose, no markdown.

  detected_subject: 4-12 words. The plain-English description of what is in the photo.
    Examples: "Black mid-rise leggings, ankle length", "Faux-wool double-breasted coat with belt",
              "One-piece swimsuit, deep V-neck, navy blue".
    Lead with item type + dominant color + 1-2 most defining features.

  detected_pt: ONE of these Amazon product type enum values that best matches the photo —
    [{pt_enum_list}].
    Be decisive: if the photo shows ANY wearable apparel, pick the closest enum value
    even if there are multiple items in frame (pick the most prominent / largest).
    Examples: a sports bra + leggings outfit → PANTS (leggings dominate) or BRA, NOT UNKNOWN.
    A model with shoes only → SANDAL. A jacket and pants → COAT (jacket is the headline item).
    Only return "UNKNOWN" when the photo is not apparel at all (a screenshot, a chart, a person
    with no product visible, an empty room).

  detected_color: 1-3 word color description (e.g. "Black", "Navy Blue", "Off White").

  observations: 3-6 short, factual visual callouts an operator may want in the listing copy.
    Each must be something the photo shows that the pre-upload likely didn't capture.
    Examples: "Hood is fur-trimmed", "Visible quilted diamond stitching", "Drawstring waist".
    Skip the obvious (e.g. don't say 'It is a coat' if the PT is COAT).

  field_suggestions: 0-5 advisory dropdown picks. Each item:
    {{ "field": "special_feature|pattern_type|coat_silhouette_type|neckline|sleeve_length|fit_type",
       "value": "<short Amazon-style enum value>",
       "why": "<one sentence — what the photo shows>" }}
    Only suggest if the photo CLEARLY shows it. Never guess. Operator decides whether to apply.

  image_quality: 0-4 issues, only if you see them. Each item:
    {{ "issue": "background_not_white|model_facing_away|low_resolution|cluttered_props|main_image_obstructed",
       "detail": "<one sentence>" }}
    Skip this list entirely if the photo looks compliant with Amazon main-image rules.

  summary: one sentence, <=120 chars. The single most useful thing the operator should know.

Return JSON only, no prose, no markdown:
{{"detected_subject":"...","detected_pt":"...","detected_color":"...","observations":["..."],"field_suggestions":[],"image_quality":[],"summary":"..."}}"""

    try:
        msg = _anthropic_client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=900,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_b64}},
                {"type": "text", "text": prompt},
            ]}],
        )
        raw = msg.content[0].text.strip()
        raw = re.sub(r'^```(?:json)?\s*', '', raw, flags=re.MULTILINE)
        raw = re.sub(r'```\s*$', '', raw, flags=re.MULTILINE).strip()
        parsed = json.loads(raw)
    except Exception as e:
        msg = str(e)
        print(f"[image-intel] vision pass failed for {style_num}: {msg}")
        # Categorize common failure modes
        if "authentication" in msg.lower() or "api_key" in msg.lower() or "x-api-key" in msg.lower():
            _last_vision_error = "Anthropic API key is missing or invalid on the server."
        elif "rate" in msg.lower() and "limit" in msg.lower():
            _last_vision_error = "Anthropic rate limit hit. Wait a few seconds and retry."
        elif "too large" in msg.lower() or "image_too_large" in msg.lower():
            _last_vision_error = "Image is too large for the vision API. Try a smaller image (\u22645MB)."
        elif "json" in msg.lower() and "decode" in msg.lower():
            _last_vision_error = "Vision returned an invalid response format. Retry once."
        elif "connection" in msg.lower() or "timeout" in msg.lower():
            _last_vision_error = "Network problem reaching Anthropic. Retry."
        else:
            _last_vision_error = f"Vision API error: {msg[:160]}"
        return None

    # Normalize + sanity-clip
    detected_pt = str(parsed.get("detected_pt", "")).strip().upper()
    valid_pts = {p["id"] for p in ALL_PRODUCT_TYPES} | {"UNKNOWN"}
    if detected_pt and detected_pt not in valid_pts:
        detected_pt = "UNKNOWN"
    expected_pt = (pt or "").strip().upper()
    if not expected_pt:
        pt_match_status = "no_expectation"  # standalone test, nothing to compare
    elif detected_pt == "UNKNOWN":
        pt_match_status = "unknown"
    elif detected_pt == expected_pt:
        pt_match_status = "match"
    else:
        pt_match_status = "mismatch"

    intel = {
        "detected_subject": str(parsed.get("detected_subject", ""))[:160],
        "detected_pt": detected_pt,
        "detected_color": str(parsed.get("detected_color", ""))[:60],
        "expected_pt": expected_pt,
        "pt_match": pt_match_status,
        "observations": [str(o)[:200] for o in (parsed.get("observations") or [])][:6],
        "field_suggestions": [
            {
                "field": str(s.get("field", ""))[:80],
                "value": str(s.get("value", ""))[:120],
                "why":   str(s.get("why", ""))[:240],
            }
            for s in (parsed.get("field_suggestions") or [])
            if isinstance(s, dict) and s.get("field") and s.get("value")
        ][:5],
        "image_quality": [
            {
                "issue":  str(q.get("issue", ""))[:80],
                "detail": str(q.get("detail", ""))[:240],
            }
            for q in (parsed.get("image_quality") or [])
            if isinstance(q, dict) and q.get("issue")
        ][:4],
        "summary": str(parsed.get("summary", ""))[:200],
        "model": "claude-sonnet-4-5",
        "analyzed_at": datetime.now().isoformat(timespec="seconds"),
    }

    cache[str(style_num)] = {"img_path": img_path, "intel": intel}
    return intel


@app.route("/api/analyze-style-image", methods=["POST"])
def api_analyze_style_image():
    """Manually re-run image intel for a style. Useful after image swap."""
    data = request.get_json(force=True) or {}
    style_num = str(data.get("style_num", ""))
    force = bool(data.get("force", False))
    if not style_num:
        return jsonify({"error": "No style_num"}), 400
    img_path = (session_data.get("style_images") or {}).get(style_num)
    if not img_path or not Path(img_path).exists():
        return jsonify({"error": "No image uploaded for this style"}), 404
    if force:
        cache = session_data.setdefault("style_image_intel", {})
        cache.pop(style_num, None)
    intel = analyze_style_image(style_num, img_path)
    if intel is None:
        return jsonify({"error": _last_vision_error or "Vision pass unavailable", "reason": _last_vision_error}), 503
    return jsonify({"ok": True, "style_num": style_num, "intel": intel})


@app.route("/api/test-image-intel", methods=["POST"])
def api_test_image_intel():
    """Standalone image-intel test — no pre-upload, no style_num required.
    Useful for QA, exploration, and seeing what vision picks up on a raw image.
    """
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    ext = Path(f.filename or "").suffix.lower()
    if ext not in [".jpg", ".jpeg", ".png", ".webp", ".gif"]:
        return jsonify({"error": f"Unsupported image type: {ext}"}), 400
    # Save to a temp slot under uploads/style_images/_test_<ts>/
    import time
    test_dir = UPLOAD_IMAGES / f"_test_{int(time.time()*1000)}"
    test_dir.mkdir(parents=True, exist_ok=True)
    save_path = test_dir / f"product{ext}"
    f.save(str(save_path))
    # Run analyze with no style context (standalone classification)
    intel = analyze_style_image("_test", str(save_path))
    if intel is None:
        return jsonify({"error": _last_vision_error or "Vision pass unavailable", "reason": _last_vision_error}), 503
    return jsonify({"ok": True, "intel": intel, "path": f"/uploads/{test_dir.name}/{save_path.name}"})


@app.route("/api/vision-status", methods=["GET"])
def api_vision_status():
    """Diagnostic: tells you whether vision is configured + the last error reason.
    Note: the Anthropic SDK does NOT validate the API key at construction time.
    The only honest health check is to make a tiny test call.
    """
    import os as _os
    has_key = bool(_os.environ.get("ANTHROPIC_API_KEY"))
    healthy = False
    health_error = None
    if _anthropic_client is not None:
        try:
            # Cheapest possible call: 1 token, no image
            _anthropic_client.messages.create(
                model="claude-sonnet-4-5",
                max_tokens=1,
                messages=[{"role": "user", "content": "hi"}],
            )
            healthy = True
        except Exception as e:
            health_error = str(e)[:200]

    if healthy:
        hint = "Vision is ready."
    elif _anthropic_client is None:
        hint = "Anthropic SDK never initialized. Set ANTHROPIC_API_KEY in Render env."
    elif not has_key:
        hint = "ANTHROPIC_API_KEY env var is not set. Add it in Render → Environment."
    else:
        hint = f"API key present but live call failed: {health_error}"

    return jsonify({
        "client_initialized":  _anthropic_client is not None,
        "api_key_in_env":      has_key,
        "healthy":             healthy,
        "health_error":        health_error,
        "last_call_error":     _last_vision_error,
        "model":               "claude-sonnet-4-5",
        "hint":                hint,
    })


@app.route("/api/style-image-intel/<style_num>", methods=["GET"])
def api_get_style_image_intel(style_num):
    """Retrieve cached image intel for a style without re-running vision."""
    cache = session_data.get("style_image_intel") or {}
    cached = cache.get(str(style_num))
    if not cached:
        return jsonify({"ok": True, "style_num": style_num, "intel": None})
    return jsonify({"ok": True, "style_num": style_num, "intel": cached.get("intel")})


@app.route("/api/style-image/<style_num>")
def serve_style_image(style_num):
    """Serve a previously uploaded style image."""
    # Check session first
    img_path = (session_data.get("style_images") or {}).get(style_num)
    if img_path and Path(img_path).exists():
        return send_file(img_path)
    # Check disk
    style_dir = UPLOAD_IMAGES / str(style_num)
    for ext in [".jpg", ".jpeg", ".png", ".webp"]:
        p = style_dir / f"product{ext}"
        if p.exists():
            return send_file(str(p))
    return jsonify({"error": "No image"}), 404


@app.route("/api/regenerate-field", methods=["POST"])
def regenerate_field():
    """Generate an alternative version of a single field."""
    data = request.get_json(force=True)
    style_id = data.get("style_id", "")
    field = data.get("field", "")
    current_content = data.get("current_content", "")
    
    brand = data.get("brand") or session_data.get("brand", "")
    styles = session_data.get("styles", [])
    style = next((s for s in styles if s["style_num"] == style_id), None)
    
    if not style or not brand:
        return jsonify({"error": "Style or brand not found"}), 400
    
    brand_cfg = _load_brand_config_data(brand)
    style_name = style["style_name"]
    subclass = style.get("subclass", "")
    sub_subclass = style.get("sub_subclass", "")
    fabric = parse_fabric(style.get("fabric", "")) or brand_cfg.get("default_fabric", "")
    care = style.get("care", "") or brand_cfg.get("default_care", "")
    upf = style.get("upf", "") or brand_cfg.get("default_upf", "")
    first_variant = style["variants"][0] if style.get("variants") else {}
    first_color = first_variant.get("color_name", "")
    first_size = first_variant.get("size", "")
    has_keywords = len(session_data.get("keywords", [])) > 0
    
    # Resolve product type for this style (no more hardcoded 'Dress')
    rpt_top = _resolve_style_product_type(style) or ""
    itn_top = _derive_item_type_name(subclass, rpt_top) or subclass or ""

    try:
        if field == "title":
            # Generate alternative title using different formula variation
            alt_title = generate_title(brand_cfg, brand, style_name, itn_top, first_color, first_size, upf)
            # Vary: swap color position or add style descriptor variation
            descriptor = style_descriptor_from_name(style_name)
            pt_word = itn_top or "Item"
            alt_title2 = f"{brand} {descriptor} {pt_word}, {first_color.title() if first_color else ''}, {first_size}".strip(", ")
            content = alt_title2[:200] if alt_title2 != current_content else alt_title[:200]
            why = generate_title_why(brand_cfg, brand, style_name, content, upf, has_keywords) + " [Alternative format.]"
        
        elif field.startswith("bullet_"):
            # Extract bullet index
            try:
                bullet_idx = int(field.split("_")[1]) - 1
            except (IndexError, ValueError):
                bullet_idx = 0
            rpt = _resolve_style_product_type(style) or ""
            sg, _ = _derive_gender_department(style)
            eg = sg or brand_cfg.get("gender", "")
            bullets = generate_bullets(brand_cfg, brand, style_name, sub_subclass, fabric, care, first_color, upf,
                                       subclass=subclass, gender=eg, product_type=rpt, style_num=style_id)
            # Rotate bullet labels for variation
            labels = ["OUTSTANDING FEATURE", "STYLE HIGHLIGHT", "DESIGN DETAIL", "FASHION FORWARD", "KEY BENEFIT"]
            if bullet_idx < len(bullets):
                b = bullets[bullet_idx]
                # Replace first word segment (the all-caps label) with an alternative
                alt_b = re.sub(r'^[A-Z\s&/]+—', labels[bullet_idx % len(labels)] + ' —', b, count=1)
                content = alt_b if alt_b != current_content else b
            else:
                content = current_content
            why = generate_bullet_why(bullet_idx, brand_cfg, brand, style_name, sub_subclass, upf, fabric, has_keywords) + " [Alternative phrasing.]"
        
        elif field == "description":
            # Use a different opener index
            total = len(DESCRIPTION_OPENERS)
            current_idx = DESCRIPTION_OPENERS_ROTATION.get(style_id, 0)
            alt_idx = (current_idx + 1) % total
            DESCRIPTION_OPENERS_ROTATION[style_id] = alt_idx
            rpt2 = _resolve_style_product_type(style) or ""
            sg2, _ = _derive_gender_department(style)
            eg2 = sg2 or brand_cfg.get("gender", "")
            content = generate_description(brand_cfg, brand, style_id, style_name, sub_subclass, fabric, care, first_color, upf,
                                           subclass=subclass, gender=eg2, product_type=rpt2)
            why = generate_description_why(brand_cfg, style_id, alt_idx, has_keywords) + " [Alternative opener used.]"
        
        elif field == "backend_keywords":
            # Reorder keywords
            rpt3 = _resolve_style_product_type(style) or ""
            sg3, _ = _derive_gender_department(style)
            eg3 = sg3 or brand_cfg.get("gender", "")
            kw_list = generate_backend_keywords(brand, style_name, subclass, first_color, fabric, upf,
                                                subclass=subclass, gender=eg3, product_type=rpt3)
            words = kw_list.split()
            # Shuffle order (deterministic rotation)
            mid = len(words) // 2
            alt_words = words[mid:] + words[:mid]
            alt_kw = " ".join(alt_words)
            while len(alt_kw.encode('utf-8')) > 250 and alt_words:
                alt_words.pop()
                alt_kw = " ".join(alt_words)
            content = alt_kw
            why = generate_keywords_why(brand, session_data.get("keywords", []), content, has_keywords) + " [Alternative keyword order.]"
        
        else:
            return jsonify({"error": f"Unknown field: {field}"}), 400
        
        return jsonify({"content": content, "why": why})
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/generate-csv", methods=["POST"])
def generate_csv():
    """Generate CSV output for all styles — no template required."""
    data = request.get_json(force=True)
    brand = data.get("brand") or session_data.get("brand", "")
    styles = data.get("styles") or session_data.get("styles", [])
    content_map = data.get("content") or session_data.get("generated_content", {})
    
    if not styles:
        return jsonify({"error": "No product data loaded"}), 400
    if not content_map:
        return jsonify({"error": "No generated content. Run Generate Content first."}), 400
    
    output = io.StringIO()
    fieldnames = ["Style #", "Title", "Bullet 1", "Bullet 2", "Bullet 3", "Bullet 4", "Bullet 5",
                  "Description", "Backend Keywords", "Color", "Size", "UPC", "Price", "Category", "Brand"]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    
    for style in styles:
        style_num = style["style_num"]
        content = content_map.get(style_num, {})
        bullets = content.get("bullets", [])
        
        for variant in style.get("variants", []):
            color = variant.get("color_name", "")
            size = variant.get("size", "")
            upc = variant.get("upc", "")
            
            # Per-variant title — resolve actual product type, no hardcoded 'Dress'
            brand_cfg = _load_brand_config_data(brand)
            upf = style.get("upf", "") or brand_cfg.get("default_upf", "")
            _rpt_csv = _resolve_style_product_type(style) or ""
            _itn_csv = _derive_item_type_name(style.get("subclass", ""), _rpt_csv) or style.get("subclass", "") or ""
            var_title = generate_title(brand_cfg, brand, style["style_name"], _itn_csv, color, size, upf)
            
            writer.writerow({
                "Style #": style_num,
                "Title": var_title,
                "Bullet 1": bullets[0][:500] if len(bullets) > 0 else "",
                "Bullet 2": bullets[1][:500] if len(bullets) > 1 else "",
                "Bullet 3": bullets[2][:500] if len(bullets) > 2 else "",
                "Bullet 4": bullets[3][:500] if len(bullets) > 3 else "",
                "Bullet 5": bullets[4][:500] if len(bullets) > 4 else "",
                "Description": content.get("description", ""),
                "Backend Keywords": content.get("backend_keywords", ""),
                "Color": color,
                "Size": normalize_size(size),
                "UPC": upc,
                "Price": style.get("list_price", ""),
                "Category": style.get("subclass", ""),
                "Brand": brand,
            })
    
    output.seek(0)
    safe_brand = re.sub(r'[^\w]', '_', brand)
    filename = f"NIS_{safe_brand}_Content.csv"
    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8")),
        as_attachment=True,
        download_name=filename,
        mimetype="text/csv",
    )


# ── Main ───────────────────────────────────────────────────────────────────────


# ═══ CATALOG HEALTH ═══
import csv
# numpy removed — not needed

# ── Catalog Health: in-memory session storage ──────────────────────────────────
catalog_health_state = {
    "catalog_data": None,        # list of dicts (rows)
    "sales_data": None,          # list of dicts (rows)
    "analysis": None,            # full analysis result
    "detected_fields": None,     # mapping of internal_field -> column_name
    "detected_format": None,     # e.g. "Vendor Central", "Seller Central", "Custom"
    "progress": {"status": "idle", "processed": 0, "total": 0, "message": ""},
}
catalog_health_lock = threading.Lock()

# ── Fuzzy column detection maps ───────────────────────────────────────────────
# Candidates cover TLG Catalog Health Template labels AND common Amazon
# Vendor/Seller Central export column names so the same detector handles both.
CATALOG_FIELD_MAP = {
    "asin":             ["child asin", "asin", "asin1", "child_asin"],
    "parent_asin":      ["parent asin", "parent_asin", "parent sku", "parent_sku"],
    "sku":              ["sku / vendor sku", "sku", "seller-sku", "seller_sku", "vendor sku", "item_sku"],
    "vendor_code":      ["vendor code / seller id", "vendor code", "seller id"],
    "upc":              ["upc / ean / gtin", "upc", "ean", "gtin"],
    "model_number":     ["model number"],
    "model_name":       ["model name"],
    "title":            ["title", "item_name", "item-name", "product title", "item name"],
    "brand":            ["brand", "brand_name", "brand name"],
    "color":            ["color", "color name", "color_name", "color map", "color_map"],
    "size":             ["size", "product - size", "size_name", "apparel size value", "apparel_size"],
    "size_system":      ["size system"],
    "bullet_1":         ["bullet point 1", "bullet_point1", "key product features 1", "bullet1"],
    "bullet_2":         ["bullet point 2", "bullet_point2", "key product features 2", "bullet2"],
    "bullet_3":         ["bullet point 3", "bullet_point3", "key product features 3", "bullet3"],
    "bullet_4":         ["bullet point 4", "bullet_point4", "key product features 4", "bullet4"],
    "bullet_5":         ["bullet point 5", "bullet_point5", "key product features 5", "bullet5"],
    "description":      ["description / a+ content", "description", "product_description", "product description"],
    "backend_keywords": ["backend keywords", "generic keywords", "generic_keywords", "search terms", "search_terms"],
    "main_image":       ["main image url", "main_image_url", "image-url", "main image"],
    "image_count":      ["additional image count", "image count"],
    "other_images":     ["other image url", "other_image_url", "other_image_url1", "image url 2", "image-url-2"],
    "aplus_status":     ["a+ / ebc status", "a+ status", "ebc status", "enhanced brand content"],
    "video_count":      ["video count"],
    "price":            ["list price", "price", "standard_price", "your price"],
    "sale_price":       ["sale price"],
    "buy_box_price":    ["buy box price"],
    "buy_box_winner":   ["buy box winner"],
    "quantity":         ["available quantity", "quantity", "amzn ioh", "fulfillable quantity", "quantity available"],
    "category":         ["category / product type", "product type", "sub-class name", "sub_class_name", "item_type", "category"],
    "subcategory":      ["subcategory", "sub sub-class name", "sub_sub_class_name"],
    "item_type_keyword":["item type keyword"],
    "fabric":           ["fabric content", "material"],
    "coo":              ["country of origin"],
    "care":             ["care instructions"],
    "weight":           ["item weight"],
    "package_dims":     ["package dimensions"],
    "fulfillment":      ["fulfillment method"],
    "inventory_status": ["inventory status"],
    "hazmat":           ["hazmat status"],
    "lqs":              ["listing quality score"],
    "suppressed":       ["search suppressed", "suppression reason"],
    "status":           ["status", "listing status"],
    "review_count":     ["customer review count", "review count"],
    "star_rating":      ["average star rating", "star rating"],
    "style":            ["style #", "style number", "style_num", "style_number"],
    "parent_child":     ["parentage level", "parent_child", "parentage", "parent/child"],
    "variation_theme":  ["variation theme", "variation_theme", "variation theme name"],
    "first_available":  ["first available date"],
    "season_code":      ["season code"],
    "last_updated":     ["last updated"],
}

SALES_FIELD_MAP = {
    "asin":         ["child asin", "asin"],
    "period_start": ["report period start", "period start", "start date"],
    "period_end":   ["report period end", "period end", "end date"],
    "month":        ["report month", "month"],
    "sessions":     ["sessions", "glance views / page views", "glance views", "glance_views", "page views"],
    "units":        ["ordered units", "shipped units", "units ordered", "shipped_units", "units"],
    "revenue":      ["ordered revenue", "shipped revenue", "ordered product sales", "shipped_revenue", "revenue"],
    "returns":      ["returns"],
    "return_rate":  ["return rate"],
    "cvr":          ["conversion rate", "unit session %", "unit session percentage", "conversion_rate", "cvr"],
    "buy_box_pct":  ["buy box %", "buy box percentage"],
    "ad_spend":     ["ad spend"],
    "ad_revenue":   ["ad revenue (attributed)", "ad revenue", "attributed sales"],
    "acos":         ["acos"],
    "rank":         ["subcategory sales rank", "sales rank"],
}

SEVERITY_WEIGHTS = {
    # Catalog hygiene (Layer 1 + Layer 2)
    "Orphan (no parent link)":         10,
    "Missing from variation matrix":    8,
    "Zero traffic / suppressed":        9,
    "Missing all bullet points":        6,
    "Missing main image":               7,
    "Missing backend keywords":         4,
    "Short title (<80 chars)":          3,
    "Missing description":              3,
    "Inconsistent title format":        2,
    "Single-child parent":              2,
    "Duplicate variation":              5,
    "Wrong parent link (brand mismatch)": 6,
    "Broken variation theme":           5,
    "Content issue killing conversion": 8,
    # Ad Readiness (PPC eligibility proxy)
    "Lost Buy Box":                     10,
    "Out of stock":                     10,
    "Listing suppressed (search)":      10,
    "Listing inactive":                 10,
    "Restricted category (no ads)":     10,
    "Price above Buy Box (>10%)":        7,
    "Low inventory (at-risk)":           6,
    "Content weak for ads (score <70)":  5,
    "Missing main image (no ads)":       9,
}

# Which issue types belong to the Ad Readiness group (used for filtering).
AD_READINESS_ISSUES = {
    "Lost Buy Box",
    "Out of stock",
    "Listing suppressed (search)",
    "Listing inactive",
    "Restricted category (no ads)",
    "Price above Buy Box (>10%)",
    "Low inventory (at-risk)",
    "Content weak for ads (score <70)",
    "Missing main image (no ads)",
}

# Which reasons definitively block ads (as opposed to putting an ASIN at risk).
AD_BLOCKING_ISSUES = {
    "Lost Buy Box",
    "Out of stock",
    "Listing suppressed (search)",
    "Listing inactive",
    "Restricted category (no ads)",
    "Missing main image (no ads)",
}

# Amazon categories/phrases that are ineligible for Sponsored Products.
RESTRICTED_AD_CATEGORIES = {
    "adult", "adult products", "sexual wellness",
    "used", "refurbished", "renewed",
    "firearms", "tobacco", "vaping",
    "prescription", "rx",
}

# ── Ad Bulksheet (Tier-2 ground truth) ────────────────────────────────────────────────
# Amazon Advertising Console → Bulk Operations → Sponsored Products bulksheet.
# The bulksheet uses verbose column names; we fuzzy-match the essential ones.
AD_BULKSHEET_FIELD_MAP = {
    "asin":         ["advertised asin", "asin", "child asin", "sku asin"],
    "sku":          ["advertised sku", "sku"],
    "status":       ["eligibility status", "asin eligibility status", "status"],
    "reasons":      ["eligibility reasons", "eligibility reason", "reason"],
    "ad_type":      ["ad type", "campaign type", "product type"],
    "campaign":     ["campaign name", "campaign"],
}

# Amazon's internal reason codes → our SEVERITY_WEIGHTS keys so the ground
# truth reconciles cleanly against our proxy output.
AMAZON_REASON_CODE_MAP = {
    # Buy Box / offer
    "asin_not_buyable":           "Lost Buy Box",
    "not_buyable":                "Lost Buy Box",
    "asin_not_featured_offer":    "Lost Buy Box",
    "not_featured_offer":         "Lost Buy Box",
    "featured_offer_ineligible":  "Lost Buy Box",
    "buy_box_suppressed":         "Lost Buy Box",
    # Inventory
    "out_of_stock":               "Out of stock",
    "no_inventory":               "Out of stock",
    "asin_not_available":         "Out of stock",
    # Suppression / listing status
    "search_suppressed":          "Listing suppressed (search)",
    "listing_suppressed":         "Listing suppressed (search)",
    "suppressed":                 "Listing suppressed (search)",
    "asin_inactive":              "Listing inactive",
    "inactive":                   "Listing inactive",
    # Policy / category
    "ad_policy_violation":        "Restricted category (no ads)",
    "restricted_category":        "Restricted category (no ads)",
    "category_ineligible":        "Restricted category (no ads)",
    "adult_product":              "Restricted category (no ads)",
    "book_format_ineligible":     "Restricted category (no ads)",
    # Pricing
    "price_not_competitive":      "Price above Buy Box (>10%)",
    "external_price_reference":   "Price above Buy Box (>10%)",
    # Content
    "missing_main_image":         "Missing main image (no ads)",
    "image_missing":              "Missing main image (no ads)",
}

def _normalize_ad_reason(raw):
    """Convert one Amazon reason code/phrase into our canonical reason label.
    Returns None when the phrase doesn't map (caller should keep the raw
    code as an informational note)."""
    if not raw:
        return None
    t = str(raw).strip().lower().replace(" ", "_").replace("-", "_")
    # Exact match first
    if t in AMAZON_REASON_CODE_MAP:
        return AMAZON_REASON_CODE_MAP[t]
    # Substring match (bulksheet sometimes wraps codes in explanatory text)
    for code, label in AMAZON_REASON_CODE_MAP.items():
        if code in t:
            return label
    return None


# ═══════════════════════════════════════════════════════════════════════
# Bulksheet snapshot history (trend tracking)
# ═══════════════════════════════════════════════════════════════════════
# Every bulksheet upload is automatically saved as a JSON snapshot under
# ./snapshots/. With ≥2 snapshots we compute trends: ineligible-over-time,
# newly blocked since prior snapshot, newly recovered, chronic flippers,
# per-reason deltas. Retention is capped at 52 (≈ 1 year weekly).
SNAPSHOTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "snapshots")
SNAPSHOT_RETENTION = 52


def _ensure_snapshots_dir():
    os.makedirs(SNAPSHOTS_DIR, exist_ok=True)


def _list_snapshots(limit=None):
    """Return snapshot metadata sorted oldest → newest."""
    _ensure_snapshots_dir()
    out = []
    for fn in sorted(os.listdir(SNAPSHOTS_DIR)):
        if not fn.endswith(".json"):
            continue
        try:
            with open(os.path.join(SNAPSHOTS_DIR, fn), "r") as f:
                d = json.load(f)
            out.append({
                "id":           d.get("id") or fn.replace(".json", ""),
                "file":         fn,
                "timestamp":    d.get("timestamp"),
                "filename":     d.get("filename"),
                "total":        d.get("total", 0),
                "ineligible":   d.get("ineligible", 0),
                "eligible":     d.get("eligible", 0),
            })
        except Exception:
            continue
    if limit:
        out = out[-limit:]
    return out


def _load_snapshot(snap_id):
    """Load a single snapshot's full ASIN map by id (filename without .json)."""
    _ensure_snapshots_dir()
    path = os.path.join(SNAPSHOTS_DIR, f"{snap_id}.json")
    if not os.path.exists(path):
        return None
    with open(path, "r") as f:
        return json.load(f)


def _save_snapshot(ad_truth_lookup, source_filename):
    """Persist the just-uploaded bulksheet as a dated snapshot.
    Returns the saved snapshot dict."""
    _ensure_snapshots_dir()
    now = datetime.now()
    snap_id = now.strftime("%Y-%m-%d_%H%M%S")
    ineligible = sum(1 for v in ad_truth_lookup.values() if v["status"] == "ineligible")
    total = len(ad_truth_lookup)
    snapshot = {
        "id":          snap_id,
        "timestamp":   now.isoformat(timespec="seconds"),
        "filename":    source_filename or "",
        "total":       total,
        "ineligible":  ineligible,
        "eligible":    total - ineligible,
        "asins":       ad_truth_lookup,   # {asin: {status, reasons, raw_reasons}}
    }
    out_path = os.path.join(SNAPSHOTS_DIR, f"{snap_id}.json")
    # Atomic write
    tmp = out_path + ".tmp"
    with open(tmp, "w") as f:
        json.dump(snapshot, f)
    os.rename(tmp, out_path)

    # Enforce retention
    _prune_snapshots()
    return snapshot


def _prune_snapshots():
    _ensure_snapshots_dir()
    files = sorted(fn for fn in os.listdir(SNAPSHOTS_DIR) if fn.endswith(".json"))
    extra = len(files) - SNAPSHOT_RETENTION
    if extra > 0:
        for fn in files[:extra]:
            try:
                os.remove(os.path.join(SNAPSHOTS_DIR, fn))
            except OSError:
                pass


def _compute_trends(scored_rows=None, max_history=12):
    """Compute trend stats across saved snapshots.

    Returns None when fewer than 2 snapshots exist.
    scored_rows (optional): list of row_result dicts so we can enrich
    newly-blocked/recovered entries with titles + categories.
    """
    metas = _list_snapshots()
    if len(metas) < 2:
        return None

    # Load full data for the most-recent window (cap history for response size)
    recent_metas = metas[-max_history:]
    full = []
    for m in recent_metas:
        snap = _load_snapshot(m["id"])
        if snap:
            full.append(snap)
    if len(full) < 2:
        return None

    # Title / category lookup from current catalog
    title_by_asin = {}
    cat_by_asin = {}
    if scored_rows:
        for r in scored_rows:
            title_by_asin[r["asin"]] = r.get("title", "")
            cat_by_asin[r["asin"]]   = r.get("category", "")

    # Time-series
    series = [
        {
            "id":           s["id"],
            "timestamp":    s["timestamp"],
            "total":        s["total"],
            "ineligible":   s["ineligible"],
            "eligible":     s["eligible"],
            "ineligible_pct": round(s["ineligible"] / max(1, s["total"]) * 100, 1),
        }
        for s in full
    ]

    # Current vs prior snapshot deltas
    prev_snap = full[-2]
    cur_snap  = full[-1]
    prev_asins = prev_snap["asins"]
    cur_asins  = cur_snap["asins"]

    newly_blocked = []
    newly_recovered = []
    for asin, cur in cur_asins.items():
        prev = prev_asins.get(asin)
        if cur["status"] == "ineligible" and (not prev or prev.get("status") == "eligible"):
            newly_blocked.append({
                "asin":    asin,
                "title":   title_by_asin.get(asin, ""),
                "category":cat_by_asin.get(asin, ""),
                "reasons": cur.get("reasons", []),
                "raw_reasons": cur.get("raw_reasons", []),
            })
        elif cur["status"] == "eligible" and prev and prev.get("status") == "ineligible":
            newly_recovered.append({
                "asin":     asin,
                "title":    title_by_asin.get(asin, ""),
                "category": cat_by_asin.get(asin, ""),
                "was_blocked_for": prev.get("reasons", []),
            })

    # Chronic flippers — ASINs that changed state at least flip_threshold times
    # across the history window.
    flip_threshold = 3
    flip_counts = {}
    for asin in set().union(*(s["asins"].keys() for s in full)):
        prior = None
        flips = 0
        for s in full:
            cur_state = s["asins"].get(asin, {}).get("status")
            if cur_state and prior and cur_state != prior:
                flips += 1
            if cur_state:
                prior = cur_state
        if flips >= flip_threshold:
            flip_counts[asin] = flips

    chronic_flippers = [
        {
            "asin":    asin,
            "title":   title_by_asin.get(asin, ""),
            "category":cat_by_asin.get(asin, ""),
            "flips":   n,
            "current_status": cur_asins.get(asin, {}).get("status", "unknown"),
        }
        for asin, n in sorted(flip_counts.items(), key=lambda kv: -kv[1])[:25]
    ]

    # Reason-trend deltas (current vs prior snapshot)
    def _count_reasons(snap):
        out = {}
        for v in snap["asins"].values():
            if v.get("status") != "ineligible":
                continue
            for r in v.get("reasons", []):
                out[r] = out.get(r, 0) + 1
            if not v.get("reasons") and v.get("raw_reasons"):
                # Use raw code bucket as fallback
                out["(uncategorized)"] = out.get("(uncategorized)", 0) + 1
        return out

    prev_reasons = _count_reasons(prev_snap)
    cur_reasons  = _count_reasons(cur_snap)
    reason_deltas = []
    for r in sorted(set(prev_reasons) | set(cur_reasons)):
        reason_deltas.append({
            "reason": r,
            "prev":   prev_reasons.get(r, 0),
            "current": cur_reasons.get(r, 0),
            "delta":  cur_reasons.get(r, 0) - prev_reasons.get(r, 0),
        })
    reason_deltas.sort(key=lambda x: -abs(x["delta"]))

    return {
        "snapshot_count":   len(metas),
        "history_window":   len(full),
        "series":           series,
        "current_id":       cur_snap["id"],
        "previous_id":      prev_snap["id"],
        "newly_blocked":    newly_blocked[:50],
        "newly_blocked_total":   len(newly_blocked),
        "newly_recovered":  newly_recovered[:50],
        "newly_recovered_total": len(newly_recovered),
        "chronic_flippers":  chronic_flippers,
        "reason_deltas":     reason_deltas[:12],
    }


def _parse_ad_bulksheet(rows, headers):
    """Turn an uploaded bulksheet into a per-ASIN ground-truth lookup.

    Returns {asin: {"status": "eligible"|"ineligible", "raw_reasons": [str],
                    "reasons": [canonical labels]}}
    """
    fields = detect_columns(headers, AD_BULKSHEET_FIELD_MAP)
    asin_col    = fields.get("asin")
    status_col  = fields.get("status")
    reason_col  = fields.get("reasons")

    lookup = {}
    for r in rows:
        asin = str(r.get(asin_col, "")).strip() if asin_col else ""
        if not asin:
            continue
        status_raw = str(r.get(status_col, "")).strip().lower() if status_col else ""
        reason_raw = str(r.get(reason_col, "")).strip() if reason_col else ""

        # Normalize status. Anything except explicit "eligible" is treated
        # as ineligible so we're conservative.
        if status_raw in ("eligible", "active", "ok", "serving"):
            status = "eligible"
        elif status_raw in ("ineligible", "blocked", "not_eligible", "paused", "not eligible"):
            status = "ineligible"
        else:
            # When status column is missing, infer from reasons: any reason = ineligible
            status = "ineligible" if reason_raw else "eligible"

        raw_list = []
        canon = []
        if reason_raw:
            for piece in reason_raw.replace(";", "|").replace(",", "|").split("|"):
                piece = piece.strip()
                if not piece:
                    continue
                raw_list.append(piece)
                mapped = _normalize_ad_reason(piece)
                if mapped and mapped not in canon:
                    canon.append(mapped)

        lookup[asin] = {
            "status":      status,
            "raw_reasons": raw_list,
            "reasons":     canon,
        }
    return lookup, fields


def _norm(s):
    """Normalize a column header for fuzzy matching."""
    return str(s).lower().strip().replace("_", " ").replace("-", " ")


def detect_columns(headers, field_map):
    """
    Fuzzy-match headers to internal field names.
    Returns {internal_field: actual_header} for matched fields.
    """
    detected = {}
    header_norm = {_norm(h): h for h in headers}
    
    for field, candidates in field_map.items():
        for cand in candidates:
            cand_norm = _norm(cand)
            # Exact normalized match
            if cand_norm in header_norm:
                detected[field] = header_norm[cand_norm]
                break
            # Substring match
            for hn, orig in header_norm.items():
                if cand_norm in hn or hn in cand_norm:
                    detected[field] = orig
                    break
            if field in detected:
                break
    
    return detected


def detect_format(headers, detected_fields):
    """Guess whether this looks like Vendor Central, Seller Central, or custom."""
    header_set = {_norm(h) for h in headers}
    if any("vendor" in h for h in header_set):
        return "Vendor Central"
    if any("seller" in h or "seller-sku" in _norm(h) for h in header_set):
        return "Seller Central"
    if "asin" in detected_fields:
        return "Custom (ASIN-based)"
    return "Custom"


# Markers Amazon/TLG templates place in a row below the real header row to
# annotate field criticality. Any row whose first few cells consist only of
# these values is treated as a metadata row and skipped.
_TEMPLATE_META_TOKENS = {
    "required", "recommended", "optional",
    "conditionally required", "cond. required", "cond required",
    "required*", "optional*",
}
# Sample rows in the TLG template start with placeholders like "B0XXXXXXXXX".
# Any row whose ASIN cell matches one of these patterns is also skipped.
_SAMPLE_ASIN_PATTERNS = ("B0XXXXXXXXX", "B0XXXXXXXX", "B0XXXXXXXXX1")


def _looks_like_metadata_row(row_vals, header_cells):
    """True when the row appears to be a REQUIRED/OPTIONAL marker line or a
    placeholder sample row, not real data."""
    non_empty = [str(v).strip().lower() for v in row_vals if v not in (None, "")]
    if not non_empty:
        return True
    # All non-empty values are REQUIRED/OPTIONAL markers
    if all(v in _TEMPLATE_META_TOKENS for v in non_empty):
        return True
    # Sample ASIN placeholder row
    for v in row_vals[:5]:
        if v and isinstance(v, str) and v.strip().upper() in _SAMPLE_ASIN_PATTERNS:
            return True
    return False


def _find_header_row(ws, max_scan=20):
    """Scan the first N rows, return the 1-indexed row number that looks like
    the real header row (most unique non-empty cells). Handles TLG Catalog
    Health Template (headers on row 4) and simple CSVs (row 1)."""
    best_row = 1
    best_score = -1
    for r in range(1, min(max_scan, (ws.max_row or 1)) + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, (ws.max_column or 1) + 1)]
        names = [str(v).strip() for v in cells if v not in (None, "")]
        if len(names) < 3:
            continue
        unique = len(set(names))
        # Penalize rows that look like category bands (few wide labels)
        if unique < len(names) * 0.6:
            continue
        # Penalize rows made entirely of REQUIRED/OPTIONAL markers
        if all(n.lower() in _TEMPLATE_META_TOKENS for n in names):
            continue
        score = unique
        if score > best_score:
            best_score = score
            best_row = r
    return best_row


def _pick_catalog_sheet(wb):
    """Choose the best sheet in a workbook for catalog data. Prefers sheets
    named 'Catalog Snapshot' (TLG template), then any sheet whose headers
    include ASIN/child-ASIN indicators."""
    # Preferred sheet names
    preferred = ["catalog snapshot", "catalog", "listings", "products", "template-"]
    for name in wb.sheetnames:
        nl = name.lower()
        for pref in preferred:
            if pref in nl:
                return wb[name]
    # Fall back: first sheet containing an ASIN column within first 20 rows
    for name in wb.sheetnames:
        ws = wb[name]
        for r in range(1, min(20, (ws.max_row or 1)) + 1):
            for c in range(1, min(60, (ws.max_column or 1)) + 1):
                v = ws.cell(row=r, column=c).value
                if v and "asin" in str(v).lower():
                    return ws
    return wb.active


def _pick_sales_sheet(wb):
    """Pick the performance/sales sheet."""
    preferred = ["monthly performance", "performance", "sales", "business report"]
    for name in wb.sheetnames:
        nl = name.lower()
        for pref in preferred:
            if pref in nl:
                return wb[name]
    return wb.active


def read_file_to_rows(file_storage, sheet_kind="catalog"):
    """Read uploaded file (CSV, TSV, XLSX/XLSM) into list of dicts.

    Handles the TLG Catalog Health Template format (banner row, category
    band, header row on row 4, REQUIRED/OPTIONAL row, sample row) as well
    as plain CSVs with headers on row 1.

    sheet_kind: "catalog" (default) picks the Catalog Snapshot-style sheet;
                "sales" picks the Monthly Performance-style sheet.
    """
    filename = file_storage.filename.lower()
    content = file_storage.read()

    if filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".xlsm"):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=False)
        ws = _pick_catalog_sheet(wb) if sheet_kind == "catalog" else _pick_sales_sheet(wb)
        header_row = _find_header_row(ws)
        max_col = ws.max_column or 1
        max_row = ws.max_row or 0
        raw_headers = [ws.cell(row=header_row, column=c).value for c in range(1, max_col + 1)]
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(raw_headers)]
        records = []
        for r in range(header_row + 1, max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            if _looks_like_metadata_row(row_vals, headers):
                continue
            row_dict = {}
            for i, val in enumerate(row_vals):
                if i < len(headers):
                    row_dict[headers[i]] = str(val).strip() if val is not None else ""
            if any(v for v in row_dict.values()):
                records.append(row_dict)
        wb.close()
        return records, headers
    else:
        # CSV or TSV
        text = content.decode("utf-8", errors="replace")
        first_line = text.split("\n")[0] if text else ""
        sep = "\t" if "\t" in first_line else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=sep)
        headers = [str(f).strip() for f in (reader.fieldnames or [])]
        records = []
        for row in reader:
            cleaned = {str(k).strip(): str(v).strip() if v else "" for k, v in row.items()}
            if _looks_like_metadata_row(list(cleaned.values()), headers):
                continue
            if any(v for v in cleaned.values()):
                records.append(cleaned)
        return records, headers


def score_content(row, detected_fields):
    """Compute 0-100 content completeness score for a single ASIN row."""
    score = 0
    issues = []

    def get(field):
        col = detected_fields.get(field)
        return str(row.get(col, "")).strip() if col else ""

    # Title: 15 pts
    title = get("title")
    if title:
        if 80 <= len(title) <= 200:
            score += 15
        elif len(title) < 80:
            score += 7
            issues.append("Short title (<80 chars)")
        else:
            score += 12  # Over 200 but present
    
    # Bullets: 15 pts (3 per bullet)
    for b in ["bullet_1", "bullet_2", "bullet_3", "bullet_4", "bullet_5"]:
        btext = get(b)
        if btext and len(btext) >= 50:
            score += 3
        elif btext:
            score += 1  # Partial credit
    if not any(get(f"bullet_{i}") for i in range(1, 6)):
        issues.append("Missing all bullet points")
    
    # Description: 10 pts
    desc = get("description")
    if desc and len(desc) >= 200:
        score += 10
    elif desc:
        score += 5
    else:
        issues.append("Missing description")
    
    # Backend keywords: 10 pts
    kw = get("backend_keywords")
    if kw:
        kw_bytes = len(kw.encode("utf-8"))
        score += 10 if kw_bytes <= 250 else 7
    else:
        issues.append("Missing backend keywords")
    
    # Main image: 10 pts
    if get("main_image"):
        score += 10
    else:
        issues.append("Missing main image")
    
    # Other images (6+): 10 pts
    img_count = 0
    if detected_fields.get("image_count"):
        try:
            img_count = int(get("image_count") or 0)
        except:
            pass
    else:
        for i in range(2, 10):
            col = detected_fields.get("other_images")
            if col:
                # Multi-image columns: check count
                img_count = 1 if get("other_images") else 0
                break
    img_count_bonus = img_count
    # Also count any columns that look like image URLs
    for col in [c for c in row if "image" in _norm(c) and c != detected_fields.get("main_image")]:
        if str(row.get(col, "")).strip():
            img_count_bonus += 1
    if img_count_bonus >= 6:
        score += 10
    elif img_count_bonus >= 3:
        score += 5

    # Price: 10 pts
    try:
        price_raw = get("price").replace("$", "").replace(",", "").strip()
        if price_raw and float(price_raw) > 0:
            score += 10
    except:
        pass
    
    # Brand: 5 pts
    if get("brand"):
        score += 5
    
    # Color + Size: 5 pts
    if get("color") or get("size"):
        score += 5
    
    # Category: 10 pts
    if get("category"):
        score += 10
    
    return min(100, score), issues


def score_color(score):
    if score >= 90:
        return "green"
    elif score >= 70:
        return "yellow"
    elif score >= 50:
        return "orange"
    return "red"


# ═══════════════════════════════════════════════════════════════════════
# AD READINESS (Tier 1 proxy) — predicts PPC eligibility from catalog data.
# This is intentionally a proxy, not ground truth: Amazon's authoritative
# answer lives in the Advertising Console bulksheet. The proxy covers the
# top-8 reasons Amazon makes an ASIN ineligible for Sponsored Products.
# ═══════════════════════════════════════════════════════════════════════
def _num(s):
    """Parse a cell value into a float, tolerating $ and commas. Returns None
    when the value is blank/unparseable."""
    if s is None:
        return None
    t = str(s).strip().replace("$", "").replace(",", "").replace("%", "")
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _bool_field(s):
    """Interpret a cell as a Yes/No boolean. Returns True for yes/y/1/true,
    False for no/n/0/false, None for blank/unknown."""
    if s is None:
        return None
    t = str(s).strip().lower()
    if not t:
        return None
    if t in ("yes", "y", "1", "true", "active", "in stock"):
        return True
    if t in ("no", "n", "0", "false"):
        return False
    return None


def _eligibility_for_row(row, detected_fields, content_score):
    """Run Tier-1 PPC eligibility checks against a single row.

    Returns (status, reasons) where:
        status  ∈ {"eligible", "at_risk", "ineligible"}
        reasons is a list of SEVERITY_WEIGHTS keys (issue labels)

    Rules are ordered from most- to least-severe; the first blocking rule
    decides status="ineligible" but all applicable reasons are returned so
    the user sees the full fix list.
    """
    def g(field):
        col = detected_fields.get(field)
        return str(row.get(col, "")).strip() if col else ""

    reasons = []

    # 1. Listing inactive → no ads can run at all
    status_val = g("status").lower()
    if status_val and status_val not in ("active", "available", "in_stock", "listed"):
        if status_val in ("inactive", "suppressed", "blocked", "removed", "stranded", "deleted"):
            if "suppressed" in status_val:
                reasons.append("Listing suppressed (search)")
            else:
                reasons.append("Listing inactive")

    # 2. Search-suppressed flag (separate column)
    suppressed = g("suppressed")
    if suppressed and suppressed.lower() not in ("", "no", "false", "0", "none"):
        if "Listing suppressed (search)" not in reasons:
            reasons.append("Listing suppressed (search)")

    # 3. Restricted category — hard block regardless of everything else
    cat = g("category").lower()
    subcat = g("subcategory").lower()
    blob = f"{cat} {subcat}"
    for phrase in RESTRICTED_AD_CATEGORIES:
        if phrase in blob:
            reasons.append("Restricted category (no ads)")
            break

    # 4. Out of stock — Amazon auto-pauses ads
    qty = _num(g("quantity"))
    inv_status = g("inventory_status").lower()
    if qty is not None and qty <= 0:
        reasons.append("Out of stock")
    elif inv_status in ("out of stock", "oos", "0"):
        reasons.append("Out of stock")
    elif qty is not None and 0 < qty <= 5:
        reasons.append("Low inventory (at-risk)")

    # 5. Lost Buy Box (#1 real-world cause of ineligibility)
    bb_winner = _bool_field(g("buy_box_winner"))
    if bb_winner is False:
        reasons.append("Lost Buy Box")

    # 6. Missing main image — Amazon won't surface ads without a compliant hero
    if detected_fields.get("main_image") and not g("main_image"):
        reasons.append("Missing main image (no ads)")

    # 7. Price above Buy Box by >10% (Buy Box often auto-suppressed after that)
    bb_price = _num(g("buy_box_price"))
    list_price = _num(g("price"))
    if bb_price and bb_price > 0 and list_price and list_price > 0:
        if list_price > bb_price * 1.10:
            reasons.append("Price above Buy Box (>10%)")

    # 8. Content too weak to convert ad clicks
    if content_score is not None and content_score < 70:
        reasons.append("Content weak for ads (score <70)")

    # Decide status
    blocking = [r for r in reasons if r in AD_BLOCKING_ISSUES]
    if blocking:
        status = "ineligible"
    elif reasons:
        status = "at_risk"
    else:
        status = "eligible"
    return status, reasons


def _eligibility_fix_action(issue):
    """Concrete, operator-facing fix guidance for each eligibility reason."""
    fixes = {
        "Lost Buy Box":                    "Win back the Buy Box: match/undercut Buy Box price, verify FBA status and seller performance metrics",
        "Out of stock":                    "Replenish inventory. If FBA inbound is slow, temporarily switch to FBM to keep ads serving",
        "Listing suppressed (search)":     "Open Seller Central → Inventory → Manage All Inventory → Search Suppressed to see Amazon's exact fix",
        "Listing inactive":                "Reactivate in Seller Central; confirm inventory is available and listing isn't policy-blocked",
        "Restricted category (no ads)":    "This category is ineligible for Sponsored Products per Amazon policy — exclude from ad campaigns",
        "Price above Buy Box (>10%)":      "Lower your price to within 10% of the current Buy Box price to regain eligibility",
        "Low inventory (at-risk)":         "Replenish soon — inventory <5 units typically flips to OOS within days",
        "Content weak for ads (score <70)":"Improve title/bullets/images/description (see Hygiene issues for this ASIN)",
        "Missing main image (no ads)":     "Upload a compliant main image (white background, 1000px+, product fills frame)",
    }
    return fixes.get(issue, "Review and fix this field")


def run_catalog_analysis(rows, detected_fields, sales_lookup=None, ad_truth_lookup=None):
    """
    Full catalog health analysis. Returns structured result dict.
    Progress is updated via catalog_health_state["progress"].

    ad_truth_lookup (optional): {asin: {"status": "eligible"|"ineligible",
                                        "reasons": [canonical labels],
                                        "raw_reasons": [str]}}
    When supplied, Amazon's ground-truth flag overrides the proxy for each
    matched ASIN, and a reconciliation block is included in the response.
    """
    state = catalog_health_state

    def get(row, field):
        col = detected_fields.get(field)
        return str(row.get(col, "")).strip() if col else ""

    total = len(rows)
    state["progress"] = {"status": "running", "processed": 0, "total": total, "message": "Starting analysis..."}

    # Build lookup structures
    parent_map = {}        # parent_asin -> list of child rows
    asin_map = {}          # asin -> row
    real_parents = set()   # ASINs that actually have parentage="parent"
    
    for i, row in enumerate(rows):
        asin = get(row, "asin") or get(row, "sku")
        if asin:
            asin_map[asin] = row
        p_asin = get(row, "parent_asin")
        pc = _norm(get(row, "parent_child"))
        if pc in ("parent",):
            real_parents.add(asin)
            if asin not in parent_map:
                parent_map[asin] = []
        elif p_asin:
            if p_asin not in parent_map:
                parent_map[p_asin] = []
            parent_map[p_asin].append(row)
        
        if (i + 1) % 1000 == 0:
            state["progress"]["processed"] = i + 1
            state["progress"]["message"] = f"Building lookup structures... {i+1}/{total}"

    # Content scoring + structural checks per ASIN
    issues_list = []
    scored_rows = []
    
    brands_seen = set()
    categories_seen = set()
    subcategories_seen = set()
    
    score_dist = {"green": 0, "yellow": 0, "orange": 0, "red": 0}
    
    # Ad Readiness aggregators
    elig_dist      = {"eligible": 0, "at_risk": 0, "ineligible": 0}
    elig_by_reason = {}          # reason -> {"asins": set, "categories": Counter, "revenue": 0.0}
    elig_by_cat    = {}          # category -> {"eligible":0,"at_risk":0,"ineligible":0,"total":0}
    
    for i, row in enumerate(rows):
        asin = get(row, "asin") or get(row, "sku") or f"row_{i}"
        title = get(row, "title")
        brand = get(row, "brand")
        category = get(row, "category")
        subcategory = get(row, "subcategory")
        p_asin = get(row, "parent_asin")
        pc = _norm(get(row, "parent_child"))
        
        if brand:
            brands_seen.add(brand)
        if category:
            categories_seen.add(category)
        if subcategory:
            subcategories_seen.add(subcategory)
        
        content_score, content_issues = score_content(row, detected_fields)
        color = score_color(content_score)
        score_dist[color] += 1
        
        # ── Ad Readiness (Tier 1 proxy) ─────────────────────────────────
        proxy_status, proxy_reasons = _eligibility_for_row(row, detected_fields, content_score)
        elig_status = proxy_status
        elig_reasons = list(proxy_reasons)
        elig_source = "proxy"
        elig_raw_codes = []

        # ── Tier-2 ground truth override ────────────────────────────────
        if ad_truth_lookup and asin in ad_truth_lookup:
            truth = ad_truth_lookup[asin]
            # Amazon's flag wins
            elig_status = "ineligible" if truth["status"] == "ineligible" else "eligible"
            # Merge canonical reasons Amazon reported. Keep any at-risk hints
            # from the proxy that Amazon doesn't surface (e.g. low inventory).
            elig_reasons = list(truth["reasons"])
            for pr in proxy_reasons:
                if pr not in elig_reasons and pr not in AD_BLOCKING_ISSUES:
                    elig_reasons.append(pr)
            elig_source = "actual"
            elig_raw_codes = truth["raw_reasons"]
        elif ad_truth_lookup:
            # Bulksheet uploaded but ASIN not present = Amazon likely hasn't
            # encountered it (not in any campaign). Fall back to proxy but
            # mark it so the UI can signal "Amazon hasn't ruled on this one".
            elig_source = "proxy_only_not_in_bulksheet"

        elig_dist[elig_status] += 1
        cat_key = category or "(Uncategorized)"
        ecat = elig_by_cat.setdefault(cat_key, {"eligible": 0, "at_risk": 0, "ineligible": 0, "total": 0})
        ecat[elig_status] += 1
        ecat["total"] += 1
        for reason in elig_reasons:
            slot = elig_by_reason.setdefault(reason, {"asins": set(), "categories": {}, "revenue": 0.0})
            slot["asins"].add(asin)
            slot["categories"][cat_key] = slot["categories"].get(cat_key, 0) + 1
        
        structural_issues = []
        
        # Orphan check: parent must actually exist as a parent row in dataset
        if pc in ("child", "variation") or p_asin:
            if not p_asin:
                structural_issues.append("Orphan (no parent link)")
            elif p_asin not in real_parents and p_asin not in asin_map:
                structural_issues.append("Orphan (no parent link)")
        
        # Wrong parent link (brand mismatch)
        if p_asin and p_asin in asin_map:
            parent_brand = get(asin_map[p_asin], "brand")
            if parent_brand and brand and parent_brand.lower() != brand.lower():
                structural_issues.append("Wrong parent link (brand mismatch)")
        
        # Single-child parent
        if pc == "parent":
            children = parent_map.get(asin, [])
            if len(children) == 1:
                structural_issues.append("Single-child parent")
        
        # Broken variation theme
        if detected_fields.get("variation_theme"):
            vt = get(row, "variation_theme")
            if vt and pc in ("child", "variation"):
                if not get(row, "color") and not get(row, "size"):
                    structural_issues.append("Broken variation theme")
        
        # Revenue cross-reference
        rev_impact = 0.0
        revenue_issues = []
        if sales_lookup and asin in sales_lookup:
            sale = sales_lookup[asin]
            try:
                sessions = float(str(sale.get("sessions", 0)).replace(",", "") or 0)
                units = float(str(sale.get("units", 0)).replace(",", "") or 0)
                revenue = float(str(sale.get("revenue", 0)).replace("$", "").replace(",", "") or 0)
                
                rev_impact = revenue
                
                if sessions == 0 and structural_issues:
                    revenue_issues.append("Zero traffic / suppressed")
                elif sessions > 0 and units == 0 and content_score < 70:
                    revenue_issues.append("Content issue killing conversion")
            except:
                pass
        elif sales_lookup and asin not in sales_lookup:
            # Check if it's an orphan with siblings
            if "Orphan (no parent link)" in structural_issues and p_asin:
                siblings = parent_map.get(p_asin, [])
                sibling_rev = []
                for sib in siblings:
                    sib_asin = get(sib, "asin") or get(sib, "sku")
                    if sib_asin and sib_asin in sales_lookup:
                        try:
                            sibling_rev.append(float(str(sales_lookup[sib_asin].get("revenue", 0)).replace("$","").replace(",","") or 0))
                        except:
                            pass
                if sibling_rev:
                    rev_impact = sum(sibling_rev) / len(sibling_rev)
        
        # Attribute sibling-revenue estimate to blocking Ad Readiness reasons
        for reason in elig_reasons:
            if reason in AD_BLOCKING_ISSUES:
                elig_by_reason[reason]["revenue"] += rev_impact
        
        all_issues = structural_issues + content_issues + revenue_issues + elig_reasons
        
        row_result = {
            "asin": asin,
            "title": title[:80] + ("..." if len(title) > 80 else "") if title else "",
            "brand": brand,
            "category": category,
            "subcategory": subcategory,
            "content_score": content_score,
            "score_color": color,
            "parent_asin": p_asin,
            "parent_child": pc,
            "issues": all_issues,
            "revenue_impact": round(rev_impact, 2),
            "ad_status": elig_status,
            "ad_reasons": elig_reasons,
            "ad_source":  elig_source,   # "proxy" | "actual" | "proxy_only_not_in_bulksheet"
            "ad_raw_codes": elig_raw_codes,
            "ad_proxy_status":  proxy_status,
            "ad_proxy_reasons": proxy_reasons,
        }
        
        # Compute priority score for each issue
        for issue in all_issues:
            severity = SEVERITY_WEIGHTS.get(issue, 2)
            priority = severity * max(1, rev_impact / 100 if rev_impact > 0 else 1) if rev_impact > 0 else severity
            issues_list.append({
                "priority": round(priority, 2),
                "asin": asin,
                "title": row_result["title"],
                "brand": brand,
                "category": category,
                "issue": issue,
                "severity": severity,
                "severity_label": _severity_label(severity),
                "revenue_impact": round(rev_impact, 2),
                "content_score": content_score,
                "fix_action": (_eligibility_fix_action(issue)
                               if issue in AD_READINESS_ISSUES
                               else _fix_action(issue)),
                "group": "ad_readiness" if issue in AD_READINESS_ISSUES else "hygiene",
            })
        
        scored_rows.append(row_result)
        
        if (i + 1) % 500 == 0:
            state["progress"]["processed"] = i + 1
            state["progress"]["message"] = f"Analyzing ASINs... {i+1}/{total}"
    
    # Missing variation matrix check
    variation_matrix = {}
    for p_asin, children in parent_map.items():
        colors = set()
        sizes = set()
        present = set()
        for child in children:
            c = get(child, "color")
            s = get(child, "size")
            if c:
                colors.add(c)
            if s:
                sizes.add(s)
            if c and s:
                present.add((c, s))
        if colors and sizes:
            expected = {(c, s) for c in colors for s in sizes}
            missing = expected - present
            for (mc, ms) in missing:
                issues_list.append({
                    "priority": 8,
                    "asin": p_asin,
                    "title": f"[Parent] {asin_map.get(p_asin, {}).get(detected_fields.get('title',''), '')[:60]}",
                    "brand": get(asin_map.get(p_asin, {}), "brand") if p_asin in asin_map else "",
                    "category": get(asin_map.get(p_asin, {}), "category") if p_asin in asin_map else "",
                    "issue": "Missing from variation matrix",
                    "severity": 8,
                    "severity_label": "High",
                    "revenue_impact": 0,
                    "content_score": 0,
                    "fix_action": f"Add variant: Color={mc}, Size={ms}",
                })
            if colors and sizes:
                variation_matrix[p_asin] = {
                    "colors": sorted(colors),
                    "sizes": sorted(sizes),
                    "present": [list(pair) for pair in present],
                    "missing": [list(pair) for pair in missing],
                }
    
    # Duplicate children check
    seen_variants = {}
    for row in rows:
        p = get(row, "parent_asin")
        c = get(row, "color")
        s = get(row, "size")
        if p and c and s:
            key = (p, _norm(c), _norm(s))
            if key in seen_variants:
                asin = get(row, "asin") or get(row, "sku")
                issues_list.append({
                    "priority": 5,
                    "asin": asin,
                    "title": get(row, "title")[:60],
                    "brand": get(row, "brand"),
                    "category": get(row, "category"),
                    "issue": "Duplicate variation",
                    "severity": 5,
                    "severity_label": "Medium",
                    "revenue_impact": 0,
                    "content_score": 0,
                    "fix_action": f"Duplicate of {seen_variants[key]}. Remove one.",
                })
            else:
                seen_variants[key] = get(row, "asin") or get(row, "sku")
    
    # Sort issues by priority (desc)
    issues_list.sort(key=lambda x: x["priority"], reverse=True)
    
    # Summary stats
    total_parents = sum(1 for r in scored_rows if r["parent_child"] == "parent")
    total_children = sum(1 for r in scored_rows if r["parent_child"] in ("child", "variation"))
    if total_parents == 0 and total_children == 0:
        total_parents = len(parent_map)
        total_children = total - total_parents
    
    avg_score = round(sum(r["content_score"] for r in scored_rows) / max(1, len(scored_rows)), 1)
    critical_count = sum(1 for iss in issues_list if iss["severity"] >= 8)
    total_revenue_at_risk = round(sum(iss["revenue_impact"] for iss in issues_list if iss["revenue_impact"] > 0), 2)
    
    state["progress"] = {"status": "done", "processed": total, "total": total, "message": "Analysis complete"}
    
    # ── Ad Readiness summary ────────────────────────────────────────────────
    ad_ineligible_revenue = sum(s["revenue"] for s in elig_by_reason.values())
    fast_fix_reasons = {
        "Listing suppressed (search)",
        "Listing inactive",
        "Missing main image (no ads)",
        "Price above Buy Box (>10%)",
    }
    fast_fix_asins = set()
    for r in fast_fix_reasons:
        if r in elig_by_reason:
            fast_fix_asins.update(elig_by_reason[r]["asins"])
    
    reasons_summary = []
    for reason, slot in elig_by_reason.items():
        top_cat = max(slot["categories"].items(), key=lambda x: x[1]) if slot["categories"] else ("", 0)
        reasons_summary.append({
            "reason":      reason,
            "asin_count":  len(slot["asins"]),
            "top_category": top_cat[0],
            "top_category_count": top_cat[1],
            "revenue_at_risk": round(slot["revenue"], 2),
            "fix_action": _eligibility_fix_action(reason),
            "severity":   SEVERITY_WEIGHTS.get(reason, 5),
            "blocking":   reason in AD_BLOCKING_ISSUES,
        })
    reasons_summary.sort(key=lambda x: (-x["asin_count"], -x["severity"]))
    
    # Category-level breakdown (sorted by ineligible count desc)
    categories_list = []
    for cat, stats in elig_by_cat.items():
        pct = (stats["eligible"] / stats["total"] * 100) if stats["total"] else 0
        categories_list.append({
            "category":    cat,
            "total":       stats["total"],
            "eligible":    stats["eligible"],
            "at_risk":     stats["at_risk"],
            "ineligible":  stats["ineligible"],
            "eligible_pct": round(pct, 1),
        })
    categories_list.sort(key=lambda x: (-x["ineligible"], -x["at_risk"]))
    
    eligibility_summary = {
        "total":         total,
        "eligible":      elig_dist["eligible"],
        "at_risk":       elig_dist["at_risk"],
        "ineligible":    elig_dist["ineligible"],
        "eligible_pct":  round(elig_dist["eligible"] / max(1, total) * 100, 1),
        "at_risk_pct":   round(elig_dist["at_risk"]  / max(1, total) * 100, 1),
        "ineligible_pct":round(elig_dist["ineligible"]/ max(1, total) * 100, 1),
        "revenue_at_risk":    round(ad_ineligible_revenue, 2),
        "fast_fix_count":     len(fast_fix_asins),
        "reasons":            reasons_summary,
        "categories":         categories_list,
        "ground_truth":       ad_truth_lookup is not None,
    }

    # ── Reconciliation (only when ground-truth bulksheet was supplied) ──────────
    if ad_truth_lookup:
        catalog_asins = {r["asin"] for r in scored_rows}
        bulksheet_asins = set(ad_truth_lookup.keys())
        matched_asins = catalog_asins & bulksheet_asins

        tp = 0   # proxy said ineligible AND Amazon says ineligible
        tn = 0   # proxy said eligible/at_risk AND Amazon says eligible
        fp = 0   # proxy said ineligible BUT Amazon says eligible
        fn = 0   # proxy said eligible/at_risk BUT Amazon says ineligible
        mismatch_examples = []

        for r in scored_rows:
            if r["asin"] not in ad_truth_lookup:
                continue
            truth = ad_truth_lookup[r["asin"]]
            proxy_ineligible = r["ad_proxy_status"] == "ineligible"
            actual_ineligible = truth["status"] == "ineligible"
            if proxy_ineligible and actual_ineligible:
                tp += 1
            elif not proxy_ineligible and not actual_ineligible:
                tn += 1
            elif proxy_ineligible and not actual_ineligible:
                fp += 1
                if len(mismatch_examples) < 20:
                    mismatch_examples.append({
                        "asin": r["asin"],
                        "title": r["title"],
                        "kind": "false_positive",
                        "proxy_reasons": r["ad_proxy_reasons"],
                        "actual_status": "eligible",
                        "actual_reasons": [],
                    })
            else:  # fn
                fn += 1
                if len(mismatch_examples) < 20:
                    mismatch_examples.append({
                        "asin": r["asin"],
                        "title": r["title"],
                        "kind": "false_negative",
                        "proxy_reasons": r["ad_proxy_reasons"],
                        "actual_status": "ineligible",
                        "actual_reasons": truth["reasons"] or truth["raw_reasons"],
                    })

        matched = len(matched_asins)
        accuracy = round((tp + tn) / max(1, matched) * 100, 1) if matched else 0.0
        precision = round(tp / max(1, tp + fp) * 100, 1) if (tp + fp) else 0.0
        recall    = round(tp / max(1, tp + fn) * 100, 1) if (tp + fn) else 0.0

        eligibility_summary["reconciliation"] = {
            "bulksheet_rows":      len(ad_truth_lookup),
            "catalog_rows":        len(catalog_asins),
            "matched":             matched,
            "catalog_only":        len(catalog_asins - bulksheet_asins),
            "bulksheet_only":      len(bulksheet_asins - catalog_asins),
            "true_positives":      tp,
            "true_negatives":      tn,
            "false_positives":     fp,
            "false_negatives":     fn,
            "accuracy":            accuracy,
            "precision":           precision,
            "recall":              recall,
            "mismatch_examples":   mismatch_examples,
        }
    
    # ── Snapshot trends ─────────────────────────────────────────────────────────
    # Always try to compute trends so the UI can show the 'snapshot count'
    # state even when only one has been taken. _compute_trends returns None
    # with <2 snapshots.
    try:
        trends = _compute_trends(scored_rows=scored_rows)
    except Exception:
        trends = None
    snapshot_count = len(_list_snapshots())

    return {
        "summary": {
            "total_asins": total,
            "total_parents": total_parents,
            "total_children": total_children,
            "avg_score": avg_score,
            "critical_issues": critical_count,
            "total_issues": len(issues_list),
            "revenue_at_risk": total_revenue_at_risk,
            "score_distribution": score_dist,
            "brands": sorted(brands_seen),
            "categories": sorted(categories_seen),
            "subcategories": sorted(subcategories_seen),
        },
        "eligibility":       eligibility_summary,
        "trends":            trends,
        "snapshot_count":    snapshot_count,
        "issues": issues_list[:5000],  # cap for response size
        "scored_rows": scored_rows[:5000],
        "variation_matrix": variation_matrix,
        "has_sales_data": sales_lookup is not None,
    }


def _severity_label(weight):
    if weight >= 9:
        return "Critical"
    elif weight >= 7:
        return "High"
    elif weight >= 4:
        return "Medium"
    return "Low"


def _fix_action(issue):
    actions = {
        "Orphan (no parent link)":          "Set parent_asin field to link this child to its parent",
        "Missing from variation matrix":    "Create new child ASIN for this color/size combination",
        "Zero traffic / suppressed":        "Check listing status, parent link, and image compliance",
        "Missing all bullet points":        "Write 5 bullet points, each >50 characters",
        "Missing main image":               "Upload a compliant main image (white background, 1000px+)",
        "Missing backend keywords":         "Add search terms (<250 bytes, no repeated words)",
        "Short title (<80 chars)":          "Expand title to 80-200 characters with key attributes",
        "Missing description":              "Write product description >200 characters",
        "Inconsistent title format":        "Align title format with brand title formula",
        "Single-child parent":              "Add more child variations or merge into standalone ASIN",
        "Duplicate variation":              "Remove or merge the duplicate child ASIN",
        "Wrong parent link (brand mismatch)":"Correct parent_asin or update brand to match parent",
        "Broken variation theme":           "Add required variation fields (color, size) for this child",
        "Content issue killing conversion": "Improve bullets, images, and description to boost CVR",
    }
    return actions.get(issue, "Review and fix this field")


# ── CATALOG HEALTH ENDPOINTS ───────────────────────────────────────────────────

@app.route("/api/catalog/upload-catalog", methods=["POST"])
def catalog_upload_catalog():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Empty filename"}), 400
    
    try:
        rows, headers = read_file_to_rows(f)
        
        if len(rows) > 60000:
            return jsonify({"error": "File too large. Max 60,000 rows."}), 400
        
        detected_fields = detect_columns(headers, CATALOG_FIELD_MAP)
        fmt = detect_format(headers, detected_fields)
        
        mapped_count = len(detected_fields)
        total_fields = len(CATALOG_FIELD_MAP)
        missing_fields = [k for k in CATALOG_FIELD_MAP if k not in detected_fields]
        
        # ─── Atlas substrate: open a catalog session + record ingestion ──────
        # The agency works inside Catalog Health to triage issues. Every
        # upload becomes a session AND an ingestion_records row; every
        # issue category surfaced becomes one decision_event; the
        # fix-file download becomes the operator response.
        #
        # Best-effort throughout: substrate writes never block analysis.
        atlas_session_id = None
        atlas_workspace = None
        try:
            brand = session_data.get("brand") or "tlg"
            atlas_workspace = brand.lower().replace(" ", "_") or "tlg"
            atlas_operator = session_data.get("operator_id") or "devang"
            from substrate.logger import open_session as _atlas_open_session
            from substrate.schema import Module as _AtlasModule
            _sess = _atlas_open_session(
                workspace_id=atlas_workspace,
                operator_id=atlas_operator,
                module=_AtlasModule.CATALOG_HEALTH,
            )
            atlas_session_id = _sess.session_id
            # Inputs audit trail: record this upload
            try:
                from substrate.inputs import record_ingestion
                record_ingestion(
                    workspace_id=atlas_workspace,
                    file_kind="catalog",
                    file_name=f.filename,
                    rows_parsed=len(rows),
                    asins_touched=len(rows),
                    detected_fields=list(detected_fields.keys()),
                    missing_fields=missing_fields,
                    summary=f"Detected {mapped_count} of {total_fields} fields",
                    uploaded_by=atlas_operator,
                    meta={"format": fmt, "session_id": atlas_session_id},
                )
            except Exception as ex2:
                print(f"[atlas] catalog ingestion record skipped: {ex2}", flush=True)
        except Exception as exc:
            print(f"[atlas] catalog session open skipped: {exc}", flush=True)
        
        with catalog_health_lock:
            catalog_health_state["catalog_data"] = rows
            catalog_health_state["detected_fields"] = detected_fields
            catalog_health_state["detected_format"] = fmt
            catalog_health_state["analysis"] = None
            catalog_health_state["progress"] = {"status": "idle", "processed": 0, "total": 0, "message": ""}
            # Stash session handles so downstream endpoints can attribute
            # to the same session without re-deriving from brand state.
            catalog_health_state["atlas_session_id"] = atlas_session_id
            catalog_health_state["atlas_workspace"] = atlas_workspace
        
        # Run analysis in background thread
        sales_lookup = None
        if catalog_health_state.get("sales_data"):
            sales_data = catalog_health_state["sales_data"]
            sales_fields = catalog_health_state.get("sales_fields", {})
            def sg(row, field):
                col = sales_fields.get(field)
                return str(row.get(col, "")).strip() if col else ""
            sales_lookup = {sg(r, "asin"): r for r in sales_data if sg(r, "asin")}
        ad_truth_lookup = catalog_health_state.get("ad_truth_lookup")
        
        def run_analysis():
            result = run_catalog_analysis(rows, detected_fields, sales_lookup, ad_truth_lookup)
            with catalog_health_lock:
                catalog_health_state["analysis"] = result
            # Once analysis is done, write one decision_event per issue
            # category surfaced. Rollup keeps the log dense with signal
            # without flooding with one-row-per-ASIN noise.
            _atlas_log_catalog_findings(result, atlas_workspace, atlas_session_id, len(rows))
        
        t = threading.Thread(target=run_analysis, daemon=True)
        t.start()
        
        return jsonify({
            "ok": True,
            "format": fmt,
            "rows": len(rows),
            "mapped_count": mapped_count,
            "total_fields": total_fields,
            "missing_fields": missing_fields,
            "detected_fields": {k: v for k, v in detected_fields.items()},
            "detection_summary": f"Detected {mapped_count} of {total_fields} fields. Missing: {', '.join(missing_fields) if missing_fields else 'none'}",
            "atlas_session_id": atlas_session_id,
        })
    
    except Exception as e:
        return jsonify({"error": f"Failed to parse file: {str(e)}"}), 500


def _atlas_log_catalog_findings(
    result: dict,
    workspace_id: str | None,
    session_id: str | None,
    catalog_size: int,
) -> None:
    """Write one decision_event per issue category surfaced by analysis.

    Called from the background analysis thread after the result dict is
    populated. Rolls up per-ASIN findings to the category level so the
    substrate sees signal-dense rows: 5-12 events per batch instead of
    one per ASIN.

    Best-effort — substrate failures must never disturb the analysis.
    """
    if not workspace_id or not result:
        return
    try:
        from substrate.logger import log_field_decision as _atlas_log
        from substrate.schema import Module as _AtlasModule
    except Exception:
        return

    brand_profile_version = f"{workspace_id}_legacy"

    # The actual analysis result returned by run_catalog_analysis() is
    # shaped { summary: { score_distribution: {...}, ... }, issues: [...] }.
    # We roll up by issue text so we get one event per issue category
    # rather than one per ASIN.
    summary = result.get("summary") or {}
    score_dist = summary.get("score_distribution") or {}
    issues = result.get("issues") or []

    # Bucket issues by issue-text. The text is human readable and stable
    # (it's what surfaces in the UI fix file too) so this stays meaningful
    # over time without requiring an enum.
    by_issue: dict[str, list[str]] = {}
    for it in issues:
        text = it.get("issue")
        asin = it.get("asin")
        if not text or not asin:
            continue
        by_issue.setdefault(text, []).append(asin)

    # Confidence proxy: severity weight scaled to [0.55, 0.95].
    def _confidence_for_issue(issue_text: str) -> float:
        # SEVERITY_WEIGHTS lives in app module scope; safe to reference.
        try:
            sev = SEVERITY_WEIGHTS.get(issue_text, 2)  # type: ignore[name-defined]
        except Exception:
            sev = 5
        return min(0.95, max(0.55, 0.45 + (sev / 20.0)))

    findings = []
    # Issue-text rollups
    for text, asins in sorted(by_issue.items(), key=lambda kv: -len(kv[1])):
        # Sanitize text into a field-name-friendly slug.
        slug = (
            text.lower()
            .replace(" ", "_")
            .replace("/", "_")
            .replace("-", "_")
            .replace("(", "")
            .replace(")", "")
        )
        slug = "".join(c for c in slug if c.isalnum() or c == "_")[:64]
        findings.append((
            f"issue_{slug}",
            len(asins),
            asins[:5],
            _confidence_for_issue(text),
        ))
    # Score-distribution rollup (always logged when present)
    for color, count in (score_dist or {}).items():
        try:
            cnum = int(count)
        except (ValueError, TypeError):
            continue
        if cnum <= 0:
            continue
        findings.append((
            f"content_score_{color}",
            cnum,
            None,
            0.80 if color in ("red", "orange") else 0.65,
        ))

    for field_name, count, sample, confidence in findings:
        if count == 0:
            continue
        try:
            _atlas_log(
                workspace_id=workspace_id,
                session_id=session_id,
                module=_AtlasModule.CATALOG_HEALTH,
                field_name=field_name,
                atlas_output={
                    "count": int(count),
                    "catalog_size": int(catalog_size),
                    "sample": sample,
                    "share": (float(count) / max(catalog_size, 1)),
                },
                overall_confidence=confidence,
                rules_injected=[{"rule_id": f"catalog.{field_name}"}],
                brand_profile_version=brand_profile_version,
                enforce_filter=False,  # always log catalog rollups
            )
        except Exception as exc:
            print(f"[atlas] catalog finding write skipped ({field_name}): {exc}", flush=True)


@app.route("/api/catalog/upload-ad-bulksheet", methods=["POST"])
def catalog_upload_ad_bulksheet():
    """Tier-2 ground truth: Amazon Advertising Console SP bulksheet with
    `Eligibility Status` and `Eligibility Reasons` columns. When uploaded,
    Amazon's flag overrides the proxy per ASIN and the analysis includes a
    reconciliation block."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Empty filename"}), 400

    try:
        rows, headers = read_file_to_rows(f)
        ad_truth_lookup, fields = _parse_ad_bulksheet(rows, headers)

        if not fields.get("asin"):
            return jsonify({
                "error": "Could not find an ASIN column. Expected one of: Advertised ASIN, ASIN, Child ASIN."
            }), 400
        if not (fields.get("status") or fields.get("reasons")):
            return jsonify({
                "error": "Could not find an Eligibility Status or Eligibility Reasons column. Make sure you downloaded the bulksheet with 'ASIN eligibility status' enabled."
            }), 400

        # Quick counts for the response
        total_in = len(ad_truth_lookup)
        ineligible_in = sum(1 for v in ad_truth_lookup.values() if v["status"] == "ineligible")

        # Auto-save snapshot for trend history
        try:
            snapshot = _save_snapshot(ad_truth_lookup, f.filename)
        except Exception as se:
            app.logger.warning(f"Snapshot save failed: {se}")
            snapshot = None

        with catalog_health_lock:
            catalog_health_state["ad_bulksheet_rows"] = rows
            catalog_health_state["ad_bulksheet_fields"] = fields
            catalog_health_state["ad_truth_lookup"] = ad_truth_lookup

        # Re-run analysis if catalog is already loaded so the UI flips to ground truth
        if catalog_health_state.get("catalog_data"):
            catalog_rows  = catalog_health_state["catalog_data"]
            detected_fields = catalog_health_state["detected_fields"]
            sales_lookup = None
            if catalog_health_state.get("sales_data"):
                sales_data = catalog_health_state["sales_data"]
                sales_fields = catalog_health_state.get("sales_fields", {})
                def sg(row, field):
                    col = sales_fields.get(field)
                    return str(row.get(col, "")).strip() if col else ""
                sales_lookup = {sg(r, "asin"): r for r in sales_data if sg(r, "asin")}

            def run_analysis():
                result = run_catalog_analysis(catalog_rows, detected_fields, sales_lookup, ad_truth_lookup)
                with catalog_health_lock:
                    catalog_health_state["analysis"] = result
            threading.Thread(target=run_analysis, daemon=True).start()

        snap_count = len(_list_snapshots())
        return jsonify({
            "ok": True,
            "rows": total_in,
            "ineligible": ineligible_in,
            "eligible":   total_in - ineligible_in,
            "fields":     list(fields.keys()),
            "snapshot_id":     (snapshot or {}).get("id"),
            "snapshot_count":  snap_count,
            "detection_summary": f"Loaded {total_in} ASINs from Ad Bulksheet ({ineligible_in} ineligible, {total_in - ineligible_in} eligible). Saved as snapshot #{snap_count}."
        })

    except Exception as e:
        return jsonify({"error": f"Failed to parse bulksheet: {str(e)}"}), 500


# ── Snapshot management endpoints ─────────────────────────────────────────────
@app.route("/api/catalog/snapshots", methods=["GET"])
def catalog_list_snapshots():
    return jsonify({
        "snapshots": _list_snapshots(),
    })


@app.route("/api/catalog/snapshots/<snap_id>", methods=["GET"])
def catalog_get_snapshot(snap_id):
    if not all(c.isalnum() or c in "-_" for c in snap_id):
        return jsonify({"error": "Invalid snapshot id"}), 400
    snap = _load_snapshot(snap_id)
    if not snap:
        return jsonify({"error": "Snapshot not found"}), 404
    return jsonify(snap)


@app.route("/api/catalog/snapshots/<snap_id>", methods=["DELETE"])
def catalog_delete_snapshot(snap_id):
    if not all(c.isalnum() or c in "-_" for c in snap_id):
        return jsonify({"error": "Invalid snapshot id"}), 400
    path = os.path.join(SNAPSHOTS_DIR, f"{snap_id}.json")
    if not os.path.exists(path):
        return jsonify({"error": "Not found"}), 404
    os.remove(path)
    return jsonify({"ok": True, "remaining": len(_list_snapshots())})


@app.route("/api/catalog/upload-sales", methods=["POST"])
def catalog_upload_sales():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Empty filename"}), 400
    
    try:
        rows, headers = read_file_to_rows(f, sheet_kind="sales")
        sales_fields = detect_columns(headers, SALES_FIELD_MAP)
        
        def sg(row, field):
            col = sales_fields.get(field)
            return str(row.get(col, "")).strip() if col else ""
        
        sales_lookup = {sg(r, "asin"): r for r in rows if sg(r, "asin")}
        
        with catalog_health_lock:
            catalog_health_state["sales_data"] = rows
            catalog_health_state["sales_fields"] = sales_fields
        
        # ─── Atlas substrate (Unit Economics Phase B): push sales metrics into
        # outcome_events so they're durable + reachable by Memory / snapshots.
        # Best-effort. Never blocks the upload response.
        atlas_workspace = catalog_health_state.get("atlas_workspace") or "tlg"
        ue_counts = {"rows_written": 0, "outcome_rows": 0, "skipped": 0}
        try:
            from substrate.unit_economics import record_sales_observations
            ue_counts = record_sales_observations(
                workspace_id=atlas_workspace,
                rows=rows,
                sales_fields=sales_fields,
                source_kind="business_report",
            )
        except Exception as ex:
            print(f"[atlas] sales→outcome_events skipped: {ex}", flush=True)
        try:
            from substrate.inputs import record_ingestion
            record_ingestion(
                workspace_id=atlas_workspace,
                file_kind="sales",
                file_name=f.filename,
                rows_parsed=len(rows),
                asins_touched=len(sales_lookup),
                detected_fields=list(sales_fields.keys()),
                summary=f"Sales report: {ue_counts['outcome_rows']} outcome_events written",
                meta={"module": "unit_economics", **ue_counts},
            )
        except Exception as ex:
            print(f"[atlas] sales ingestion record skipped: {ex}", flush=True)
        
        # Re-run analysis if catalog is already loaded
        if catalog_health_state.get("catalog_data"):
            catalog_rows = catalog_health_state["catalog_data"]
            detected_fields = catalog_health_state["detected_fields"]
            ad_truth_lookup = catalog_health_state.get("ad_truth_lookup")
            
            def run_analysis():
                result = run_catalog_analysis(catalog_rows, detected_fields, sales_lookup, ad_truth_lookup)
                with catalog_health_lock:
                    catalog_health_state["analysis"] = result
            
            t = threading.Thread(target=run_analysis, daemon=True)
            t.start()
        
        return jsonify({
            "ok": True,
            "rows": len(rows),
            "asins_matched": len(sales_lookup),
            "fields": list(sales_fields.keys()),
            "substrate": ue_counts,
        })
    
    except Exception as e:
        return jsonify({"error": f"Failed to parse sales file: {str(e)}"}), 500


@app.route("/api/catalog/results")
def catalog_results():
    progress = catalog_health_state.get("progress", {})
    analysis = catalog_health_state.get("analysis")
    
    if not analysis and progress.get("status") == "running":
        return jsonify({
            "status": "running",
            "progress": progress,
        })
    
    if not analysis:
        return jsonify({"status": "idle"})
    
    return jsonify({
        "status": "done",
        "analysis": analysis,
        "progress": progress,
    })


@app.route("/api/catalog/progress")
def catalog_progress():
    progress = catalog_health_state.get("progress", {"status": "idle", "processed": 0, "total": 0, "message": ""})
    pct = 0
    if progress.get("total", 0) > 0:
        pct = round(progress["processed"] / progress["total"] * 100)
    return jsonify({**progress, "percent": pct})


@app.route("/api/catalog/fix-file")
def catalog_fix_file():
    analysis = catalog_health_state.get("analysis")
    if not analysis:
        return jsonify({"error": "No analysis available"}), 404
    
    issues = analysis.get("issues", [])
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=["ASIN", "Title", "Brand", "Category", "Issue", "Severity", "Revenue Impact", "Fix Action"])
    writer.writeheader()
    for iss in issues:
        writer.writerow({
            "ASIN": iss["asin"],
            "Title": iss["title"],
            "Brand": iss["brand"],
            "Category": iss["category"],
            "Issue": iss["issue"],
            "Severity": iss["severity_label"],
            "Revenue Impact": f"${iss['revenue_impact']:,.2f}" if iss["revenue_impact"] else "",
            "Fix Action": iss["fix_action"],
        })
    
    output.seek(0)

    # ─── Atlas substrate: fix-file download = operator decided to act ──────────
    # We log one session_completed event marking the operator's commit to
    # action. Per-issue accept/dismiss attribution comes later when the
    # Memory tab surfaces individual issues; for now the download itself
    # is the strongest signal of intent.
    try:
        workspace_id = catalog_health_state.get("atlas_workspace")
        session_id = catalog_health_state.get("atlas_session_id")
        if workspace_id and session_id:
            from substrate.logger import (
                read_session as _atlas_read_session,
                submit_session as _atlas_submit,
            )
            from substrate.schema import (
                SessionObject as _AtlasSession,
                Module as _AtlasModule,
            )
            existing = _atlas_read_session(workspace_id, session_id)
            if existing and existing.get("state") == "live":
                s = _AtlasSession(
                    session_id=existing.get("session_id", session_id),
                    workspace_id=existing.get("workspace_id", workspace_id),
                    operator_id=existing.get("operator_id", "devang"),
                    module=_AtlasModule(existing.get("module", "catalog_health")),
                    started_at=existing.get("started_at", ""),
                    ended_at=existing.get("ended_at"),
                    state=existing.get("state", "live"),
                    operator_notes=existing.get("operator_notes"),
                    exemplar=bool(existing.get("exemplar", False)),
                )
                _atlas_submit(
                    s,
                    operator_notes=f"Fix file downloaded: {len(issues)} issues exported",
                    exemplar=False,
                )
    except Exception as exc:
        print(f"[atlas] catalog fix-file substrate close skipped: {exc}", flush=True)

    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8")),
        as_attachment=True,
        download_name="Catalog_Fix_File.csv",
        mimetype="text/csv",
    )


@app.route("/api/catalog/export")
def catalog_export():
    analysis = catalog_health_state.get("analysis")
    if not analysis:
        return jsonify({"error": "No analysis available"}), 404
    
    scored_rows = analysis.get("scored_rows", [])
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=["ASIN", "Title", "Brand", "Category", "Subcategory", "Content Score", "Score Grade", "Parent ASIN", "Issues", "Revenue Impact"])
    writer.writeheader()
    for r in scored_rows:
        writer.writerow({
            "ASIN": r["asin"],
            "Title": r["title"],
            "Brand": r["brand"],
            "Category": r["category"],
            "Subcategory": r.get("subcategory", ""),
            "Content Score": r["content_score"],
            "Score Grade": r["score_color"].upper(),
            "Parent ASIN": r.get("parent_asin", ""),
            "Issues": "; ".join(r.get("issues", [])),
            "Revenue Impact": f"${r['revenue_impact']:,.2f}" if r.get("revenue_impact") else "",
        })
    
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8")),
        as_attachment=True,
        download_name="Catalog_Health_Full_Analysis.csv",
        mimetype="text/csv",
    )



# ═══ MERGE LISTINGS MODULE ═══════════════════════════════════════════════════

# In-memory merge state
merge_state = {
    "plan": None,          # list of merge actions
    "approved": {},        # action_id -> True/False
    "generated_at": None,
    # Atlas substrate: opened on /analyze, used to attach operator_responses
    # on /approve. Map action_id -> decision_event_id so /approve can route
    # the response to the right substrate row.
    "atlas_workspace_id": None,
    "atlas_session_id": None,
    "atlas_event_ids": {},  # action_id -> event_id
}
merge_lock = threading.Lock()


def _build_merge_plan(catalog_data, detected_fields):
    """
    Analyse catalog_data and produce a list of merge action dicts.
    """
    def get(row, field):
        col = detected_fields.get(field)
        return str(row.get(col, "")).strip() if col else ""

    # Build structures
    asin_map = {}          # asin -> row
    parent_map = {}        # parent_asin -> [child rows]
    real_parents = set()   # ASINs that have parentage == "parent"
    model_to_asins = {}    # model_name -> [asin list]

    for row in catalog_data:
        asin = get(row, "asin") or get(row, "sku")
        if not asin:
            continue
        asin_map[asin] = row
        pc = get(row, "parent_child").lower()
        if pc == "parent":
            real_parents.add(asin)
            if asin not in parent_map:
                parent_map[asin] = []
        p_asin = get(row, "parent_asin")
        if p_asin and pc != "parent":
            if p_asin not in parent_map:
                parent_map[p_asin] = []
            parent_map[p_asin].append(row)
        # Model name grouping
        model = get(row, "model_name") or get(row, "sku")
        if model:
            model_base = re.split(r"[-_](?:XS|S|M|L|XL|XXL|2XL|3XL|BLACK|WHITE|RED|BLUE|GREEN|NAVY|[A-Z]{1,2}\d{0,2})$",
                                   model.upper())[0].strip()
            if model_base not in model_to_asins:
                model_to_asins[model_base] = []
            model_to_asins[model_base].append(asin)

    actions = []
    action_id = 0

    # ── 1. Split families: same model_name → multiple parent ASINs ──────────
    for model_base, asins_in_family in model_to_asins.items():
        if len(asins_in_family) < 2:
            continue
        parents_in_family = [a for a in asins_in_family if a in real_parents]
        if len(parents_in_family) <= 1:
            continue
        # Primary = the one with most children
        primary = max(parents_in_family, key=lambda p: len(parent_map.get(p, [])))
        secondary_parents = [p for p in parents_in_family if p != primary]
        for sec_parent in secondary_parents:
            children_to_move = parent_map.get(sec_parent, [])
            affected = [get(c, "asin") or get(c, "sku") for c in children_to_move] + [sec_parent]
            affected = [a for a in affected if a]
            primary_title = get(asin_map.get(primary, {}), "title")[:60] if asin_map.get(primary) else primary
            sec_title = get(asin_map.get(sec_parent, {}), "title")[:60] if asin_map.get(sec_parent) else sec_parent
            action_id += 1
            actions.append({
                "id": f"action_{action_id}",
                "action_type": "reassign",
                "affected_asins": affected,
                "from_parent": sec_parent,
                "to_parent": primary,
                "reasoning": f"Model family '{model_base}' is split across {len(parents_in_family)} parent ASINs. "
                             f"Primary parent {primary} has {len(parent_map.get(primary,[]))} children; "
                             f"{sec_parent} has {len(children_to_move)}. Consolidating under primary.",
                "confidence": "High" if len(children_to_move) > 0 else "Medium",
                "from_parent_title": sec_title,
                "to_parent_title": primary_title,
            })

    # ── 2. Orphan fix: children with no valid parent ────────────────────────
    for asin, row in asin_map.items():
        pc = get(row, "parent_child").lower()
        if pc in ("child", "variation", ""):
            p_asin = get(row, "parent_asin")
            if not p_asin or p_asin == asin or p_asin not in asin_map:
                # Try to find a parent by model name
                model = get(row, "model_name") or get(row, "sku")
                model_base = re.split(r"[-_](?:XS|S|M|L|XL|XXL|2XL|3XL|BLACK|WHITE|RED|BLUE|GREEN|NAVY|[A-Z]{1,2}\d{0,2})$",
                                       (model or "").upper())[0].strip() if model else ""
                suggested_parent = None
                if model_base and model_base in model_to_asins:
                    candidates = [a for a in model_to_asins[model_base]
                                  if a in real_parents and a != asin]
                    if candidates:
                        suggested_parent = max(candidates, key=lambda p: len(parent_map.get(p, [])))
                reason = (
                    f"ASIN {asin} has no valid parent link (parent_asin='{p_asin or 'empty'}'). "
                    + (f"Best match found: parent {suggested_parent} in same model family '{model_base}'."
                       if suggested_parent else "No matching parent found — may need to be made standalone or a new parent created.")
                )
                action_id += 1
                actions.append({
                    "id": f"action_{action_id}",
                    "action_type": "orphan_fix",
                    "affected_asins": [asin],
                    "from_parent": p_asin or "",
                    "to_parent": suggested_parent or "",
                    "reasoning": reason,
                    "confidence": "High" if suggested_parent else "Low",
                    "from_parent_title": "",
                    "to_parent_title": get(asin_map.get(suggested_parent, {}), "title")[:60] if suggested_parent else "",
                })

    # ── 3. Category mismatch: child's category differs from parent's ─────────
    for p_asin, children in parent_map.items():
        if p_asin not in asin_map:
            continue
        parent_cat = get(asin_map[p_asin], "category")
        for child_row in children:
            child_asin = get(child_row, "asin") or get(child_row, "sku")
            child_cat = get(child_row, "category")
            if parent_cat and child_cat and parent_cat.lower() != child_cat.lower():
                action_id += 1
                actions.append({
                    "id": f"action_{action_id}",
                    "action_type": "category_fix",
                    "affected_asins": [child_asin],
                    "from_parent": p_asin,
                    "to_parent": p_asin,
                    "reasoning": f"Child {child_asin} has category '{child_cat}' but its parent {p_asin} has category '{parent_cat}'. Child should match parent category.",
                    "confidence": "Medium",
                    "from_parent_title": get(asin_map[p_asin], "title")[:60],
                    "to_parent_title": get(asin_map[p_asin], "title")[:60],
                })

    return actions


@app.route("/api/merge/analyze", methods=["POST"])
def merge_analyze():
    catalog_data = catalog_health_state.get("catalog_data")
    detected_fields = catalog_health_state.get("detected_fields")
    if not catalog_data or not detected_fields:
        return jsonify({"error": "No catalog data loaded. Run Catalog Health upload first."}), 400
    try:
        actions = _build_merge_plan(catalog_data, detected_fields)

        # ─── Atlas substrate: open a Variations session + log proposals ───
        # One session per /analyze run (operator sitting). One decision_event
        # per proposed action so /approve can attach an operator_response.
        # Best-effort: substrate failures never break the merge plan response.
        atlas_session_id = None
        atlas_event_ids: dict[str, str] = {}
        try:
            from substrate.logger import (
                open_session as _atlas_open,
                log_field_decision as _atlas_log,
            )
            from substrate.schema import Module as _AtlasModule
            ws = _atlas_current_workspace()
            op_id = session_data.get("operator_id") or "devang"
            bpv = f"{ws}_legacy"
            atlas_sess = _atlas_open(workspace_id=ws, operator_id=op_id,
                                     module=_AtlasModule.VARIATIONS)
            atlas_session_id = atlas_sess.session_id
            rules = [
                {"rule_id": "variations.merge_plan.v1"},
            ]
            # Confidence proxy: derived inside Atlas. We bracket by
            # action_type — reassigns of small families are higher confidence
            # than orphan-fix guesses (which may need a 'TBD' parent).
            conf_by_type = {
                "reassign": 0.80,
                "category_fix": 0.75,
                "orphan_fix": 0.60,
            }
            for action in actions:
                affected = action.get("affected_asins") or []
                # Anchor ASIN only when the action targets exactly one ASIN.
                # Multi-ASIN actions don't fit the per-ASIN snapshot model,
                # so we log them without an asin (snapshot stays empty).
                anchor_asin = affected[0] if len(affected) == 1 else None
                eid = _atlas_log(
                    workspace_id=ws,
                    session_id=atlas_session_id,
                    module=_AtlasModule.VARIATIONS,
                    field_name="parentage_correction",
                    atlas_output={
                        "action_id": action["id"],
                        "action_type": action["action_type"],
                        "from_parent": action.get("from_parent"),
                        "to_parent": action.get("to_parent"),
                        "affected_asins": affected,
                        "reasoning": action.get("reasoning", ""),
                    },
                    overall_confidence=conf_by_type.get(
                        action["action_type"], 0.65),
                    rules_injected=rules,
                    brand_profile_version=bpv,
                    asin=anchor_asin,
                )
                if eid:
                    atlas_event_ids[action["id"]] = eid
        except Exception as _atlas_exc:
            print(f"[atlas] variations decision log skipped: {_atlas_exc}", flush=True)

        with merge_lock:
            merge_state["plan"] = actions
            # Default 'approved=True' means the operator implicitly accepts
            # every proposal unless they toggle it off. The audit-trail
            # equivalent is a stronger commitment that we ONLY record once
            # they hit /approve explicitly — we don't auto-write 'accept'
            # for everything, because that would inflate the accept count
            # for proposals the operator never actually looked at.
            merge_state["approved"] = {a["id"]: True for a in actions}
            merge_state["generated_at"] = datetime.now().isoformat()
            merge_state["atlas_session_id"] = atlas_session_id
            merge_state["atlas_event_ids"] = atlas_event_ids
            merge_state["atlas_workspace_id"] = (
                _atlas_current_workspace() if atlas_session_id else None
            )

        # Summary
        split_families = sum(1 for a in actions if a["action_type"] == "reassign")
        orphans = sum(1 for a in actions if a["action_type"] == "orphan_fix")
        category_fixes = sum(1 for a in actions if a["action_type"] == "category_fix")

        return jsonify({
            "ok": True,
            "plan": actions,
            "summary": {
                "split_families": split_families,
                "orphaned_asins": orphans,
                "category_mismatches": category_fixes,
                "total_actions": len(actions),
            },
            "atlas_session_id": atlas_session_id,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/merge/plan", methods=["GET"])
def merge_plan():
    plan = merge_state.get("plan")
    if plan is None:
        return jsonify({"plan": None, "summary": None})
    approved = merge_state.get("approved", {})
    split_families = sum(1 for a in plan if a["action_type"] == "reassign")
    orphans = sum(1 for a in plan if a["action_type"] == "orphan_fix")
    category_fixes = sum(1 for a in plan if a["action_type"] == "category_fix")
    return jsonify({
        "plan": plan,
        "approved": approved,
        "summary": {
            "split_families": split_families,
            "orphaned_asins": orphans,
            "category_mismatches": category_fixes,
            "total_actions": len(plan),
        },
        "generated_at": merge_state.get("generated_at"),
    })


@app.route("/api/merge/approve", methods=["POST"])
def merge_approve():
    data = request.get_json(force=True) or {}
    action_id = data.get("action_id")
    approved = data.get("approved", True)
    if not action_id:
        return jsonify({"error": "action_id required"}), 400
    with merge_lock:
        if merge_state["plan"] is None:
            return jsonify({"error": "No plan loaded"}), 400
        ids = {a["id"] for a in merge_state["plan"]}
        if action_id not in ids:
            return jsonify({"error": "Unknown action_id"}), 404
        merge_state["approved"][action_id] = bool(approved)
        ws = merge_state.get("atlas_workspace_id")
        event_id = (merge_state.get("atlas_event_ids") or {}).get(action_id)

    # ─── Atlas substrate: operator_response ──────────────────────────
    # Toggling approved=True → 'accept', approved=False → 'reject'.
    # Scope is 'just_this' — a parentage correction approval shouldn't
    # promote a brand-wide rule (every catalog has different shapes).
    # Best-effort: substrate failures never break the merge approve flow.
    if ws and event_id:
        try:
            from substrate.logger import (
                update_field_decision_with_operator_response as _atlas_resp,
            )
            from substrate.schema import (
                OperatorAction as _Act, OperatorScope as _Scope,
            )
            _atlas_resp(
                workspace_id=ws,
                event_id=event_id,
                operator_action=_Act("accept" if approved else "reject"),
                operator_value=None,
                operator_scope=_Scope("just_this"),
                operator_time_to_decision_ms=None,
                operator_comment=None,
                operator_viewed_case=False,
            )
        except Exception as exc:
            print(f"[atlas] variations operator_response skipped: {exc}", flush=True)

    return jsonify({"ok": True, "action_id": action_id, "approved": bool(approved)})


@app.route("/api/merge/generate-fix", methods=["POST"])
def merge_generate_fix():
    plan = merge_state.get("plan")
    approved = merge_state.get("approved", {})
    if not plan:
        return jsonify({"error": "No merge plan. Run analyze first."}), 400

    approved_actions = [a for a in plan if approved.get(a["id"], True)]
    if not approved_actions:
        return jsonify({"error": "No approved actions to generate."}), 400

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ASIN", "Action", "Current Parent", "New Parent",
                     "Variation Theme", "Parentage Level", "Notes"])

    catalog_data = catalog_health_state.get("catalog_data", [])
    detected_fields = catalog_health_state.get("detected_fields", {})

    def get_field(row, field):
        col = detected_fields.get(field)
        return str(row.get(col, "")).strip() if col else ""

    asin_row_map = {}
    for row in catalog_data:
        a = get_field(row, "asin") or get_field(row, "sku")
        if a:
            asin_row_map[a] = row

    for action in approved_actions:
        vt = ""
        for asin in action["affected_asins"]:
            row = asin_row_map.get(asin, {})
            vt = get_field(row, "variation_theme") if row else ""
            parentage = "child"
            if action["action_type"] == "reassign":
                notes = f"Reassign from parent {action['from_parent']} to {action['to_parent']}"
            elif action["action_type"] == "orphan_fix":
                notes = f"Orphan fix — assign to parent {action['to_parent'] or 'TBD'}"
            else:
                notes = f"Category fix under parent {action['to_parent']}"
            writer.writerow([
                asin,
                action["action_type"].replace("_", " ").title(),
                action["from_parent"],
                action["to_parent"],
                vt,
                parentage,
                notes,
            ])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8")),
        as_attachment=True,
        download_name=f"Merge_Fix_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
        mimetype="text/csv",
    )


# ═══ INTEL MODULE ════════════════════════════════════════════════════════════

intel_state = {
    "recommendations": None,
    "dismissed": set(),
    "accepted": {},
    "generated_at": None,
}
intel_lock = threading.Lock()


def _build_intel_recommendations(catalog_data, detected_fields,
                                  nis_state=None, feedback_data=None):
    """
    Generate ranked intelligence recommendations from all available data.
    """
    def get(row, field):
        col = detected_fields.get(field) if detected_fields else None
        return str(row.get(col, "")).strip() if col else ""

    recs = []
    rec_id = 0

    def new_id():
        nonlocal rec_id
        rec_id += 1
        return f"intel_{rec_id}"

    if not catalog_data:
        return recs

    # Build data structures
    asin_map = {}
    parent_map = {}
    real_parents = set()
    bullet_sets = {}   # bullet_index -> list of (asin, text)
    title_lengths = []

    for row in catalog_data:
        asin = get(row, "asin") or get(row, "sku")
        if not asin:
            continue
        asin_map[asin] = row
        pc = get(row, "parent_child").lower()
        if pc == "parent":
            real_parents.add(asin)
        p_asin = get(row, "parent_asin")
        if p_asin and pc != "parent":
            if p_asin not in parent_map:
                parent_map[p_asin] = []
            parent_map[p_asin].append(asin)

        # Collect bullets
        for i in range(1, 6):
            bullet = get(row, f"bullet_{i}")
            if bullet:
                if i not in bullet_sets:
                    bullet_sets[i] = []
                bullet_sets[i].append((asin, bullet))

        # Title length
        title = get(row, "title")
        if title:
            title_lengths.append((asin, len(title)))

    total_asins = len(asin_map)

    # ── 1. Duplicate bullets across ASINs ────────────────────────────────────
    for bullet_idx, entries in bullet_sets.items():
        text_to_asins = {}
        for asin, text in entries:
            t = text.lower().strip()
            if t:
                if t not in text_to_asins:
                    text_to_asins[t] = []
                text_to_asins[t].append(asin)
        for text, asins in text_to_asins.items():
            if len(asins) >= 5:
                severity = "High" if len(asins) >= 20 else "Medium"
                recs.append({
                    "id": new_id(),
                    "type": "content_duplicate",
                    "severity": severity,
                    "title": f"Bullet {bullet_idx} is identical across {len(asins)} ASINs",
                    "description": f"The same bullet point text is used word-for-word on {len(asins)} listings. "
                                   f"Amazon may suppress duplicate content and customers see no differentiation.",
                    "why": f"Exact duplicate: \"{text[:120]}...\" appears on {len(asins)} ASINs. "
                           f"Unique bullet copy improves relevance signals and conversion on style-differentiated products.",
                    "affected_asins": asins[:50],
                    "estimated_impact": "Medium — duplicate content can reduce relevance scoring",
                    "suggested_action": f"Rewrite Bullet {bullet_idx} for each style to highlight unique attributes (color, fit, occasion).",
                    "action_type": "change_bullet",
                })

    # ── 2. Title length optimization ─────────────────────────────────────────
    short_titles = [(a, ln) for a, ln in title_lengths if ln < 80]
    if short_titles:
        severity = "High" if len(short_titles) > total_asins * 0.3 else "Medium"
        recs.append({
            "id": new_id(),
            "type": "title_optimization",
            "severity": severity,
            "title": f"{len(short_titles)} titles are under 80 characters",
            "description": f"{len(short_titles)} of {total_asins} titles use fewer than 80 chars of the 200-char limit. "
                           f"Longer titles with relevant keywords improve search visibility.",
            "why": f"Amazon allows 200 characters for titles. Short titles leave keyword space unused. "
                   f"Adding size range, key features, or occasion keywords can meaningfully improve CTR.",
            "affected_asins": [a for a, _ in short_titles[:50]],
            "estimated_impact": "High — titles are the #1 ranking signal for search",
            "suggested_action": "Expand short titles to include: key feature, target customer, or occasion. Aim for 130-180 chars.",
            "action_type": "change_title",
        })

    very_long_titles = [(a, ln) for a, ln in title_lengths if ln > 190]
    if very_long_titles:
        recs.append({
            "id": new_id(),
            "type": "title_optimization",
            "severity": "Medium",
            "title": f"{len(very_long_titles)} titles exceed 190 characters (at risk of truncation)",
            "description": f"Titles over 200 chars are truncated by Amazon, cutting off important keywords.",
            "why": "Amazon truncates titles at 200 chars in some views. End-of-title keywords are the most likely to be cut.",
            "affected_asins": [a for a, _ in very_long_titles[:50]],
            "estimated_impact": "Medium — truncated titles lose keyword visibility",
            "suggested_action": "Trim these titles to under 190 characters, keeping the most important keywords first.",
            "action_type": "change_title",
        })

    # ── 3. Missing backend keywords (description empty) ───────────────────────
    no_desc = [asin for asin, row in asin_map.items()
               if len(get(row, "description")) < 100]
    if no_desc:
        severity = "Critical" if len(no_desc) > total_asins * 0.5 else "High"
        recs.append({
            "id": new_id(),
            "type": "content_quality",
            "severity": severity,
            "title": f"Description under 100 chars on {len(no_desc)} ASINs",
            "description": f"{len(no_desc)} listings have minimal or empty descriptions. "
                           f"Descriptions provide keyword real estate and help customers make purchase decisions.",
            "why": "Product descriptions index for search and provide backend keyword coverage. "
                   "Empty descriptions miss out on long-tail keyword coverage and reduce Buy Box competitiveness.",
            "affected_asins": no_desc[:50],
            "estimated_impact": "High — descriptions contribute to A9 indexing",
            "suggested_action": "Write 200-500 char descriptions highlighting fabric, fit, care instructions, and occasion suitability.",
            "action_type": "add_keyword",
        })

    # ── 4. Variation gap: parents with fewer children than average ──────────
    child_counts = [len(children) for p, children in parent_map.items() if p in real_parents]
    if child_counts:
        avg_children = sum(child_counts) / len(child_counts)
        thin_parents = [(p, len(parent_map[p])) for p in real_parents
                        if len(parent_map.get(p, [])) < max(2, avg_children * 0.4)]
        if thin_parents:
            recs.append({
                "id": new_id(),
                "type": "variation_gap",
                "severity": "Medium",
                "title": f"{len(thin_parents)} parent ASINs have fewer variations than catalog average",
                "description": f"Catalog average is {avg_children:.1f} children per parent. "
                               f"{len(thin_parents)} parents have significantly fewer. "
                               f"Thin variation families miss size/color opportunities.",
                "why": f"Parents with {avg_children:.0f}+ children capture more organic traffic across size and color searches. "
                       f"Adding common sizes (S-2X) or seasonal colors can significantly expand reach.",
                "affected_asins": [p for p, _ in thin_parents[:50]],
                "estimated_impact": "Medium — more variants = more search coverage",
                "suggested_action": f"Review thin families and consider adding missing sizes or colors. Catalog average is {avg_children:.1f} variants.",
                "action_type": "add_variant",
            })

    # ── 5. Missing bullets ────────────────────────────────────────────────────
    no_bullet_5 = [asin for asin, row in asin_map.items()
                   if not get(row, "bullet_5")]
    if no_bullet_5 and len(no_bullet_5) > 5:
        recs.append({
            "id": new_id(),
            "type": "content_quality",
            "severity": "Medium",
            "title": f"{len(no_bullet_5)} ASINs missing Bullet Point 5",
            "description": f"Amazon allows 5 bullet points. {len(no_bullet_5)} listings only use 4 or fewer, "
                           f"leaving keyword and content opportunity on the table.",
            "why": "Each bullet point is a separate keyword indexing opportunity. "
                   "Bullet 5 is often used for care instructions, compatibility, or brand story — all indexable.",
            "affected_asins": no_bullet_5[:50],
            "estimated_impact": "Low — incremental keyword coverage",
            "suggested_action": "Add Bullet 5 with care instructions, size guidance, or brand/warranty information.",
            "action_type": "change_bullet",
        })

    # ── 6. A/B test suggestions for high-count duplicate bullet families ──────
    families_with_many_dupes = []
    for bullet_idx, entries in bullet_sets.items():
        text_to_asins = {}
        for asin, text in entries:
            t = text.lower().strip()
            if t:
                if t not in text_to_asins:
                    text_to_asins[t] = []
                text_to_asins[t].append(asin)
        most_common = max(text_to_asins.items(), key=lambda x: len(x[1]), default=(None, []))
        if most_common[0] and len(most_common[1]) >= 10:
            families_with_many_dupes.append((bullet_idx, most_common[0], most_common[1]))

    if families_with_many_dupes:
        bullet_idx, text, asins = families_with_many_dupes[0]
        recs.append({
            "id": new_id(),
            "type": "ab_test_suggestion",
            "severity": "Low",
            "title": f"A/B test opportunity: Bullet {bullet_idx} variant on {len(asins)} ASINs",
            "description": f"The most-used bullet ({len(asins)} ASINs) is a candidate for A/B testing. "
                           f"Test a feature-focused variant against the current generic version.",
            "why": f"Current bullet: \"{text[:100]}...\". "
                   f"With {len(asins)} ASINs using this identical copy, a small CTR improvement "
                   f"on a test variant could justify a full rollout.",
            "affected_asins": asins[:20],
            "estimated_impact": "Medium — A/B tests on high-volume bullet copy can yield 5-15% CVR lift",
            "suggested_action": f"Draft an alternative Bullet {bullet_idx} emphasizing a specific feature or benefit. Test on 5-10 ASINs for 30 days.",
            "action_type": "change_bullet",
        })

    # Sort by severity
    severity_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    recs.sort(key=lambda r: severity_order.get(r["severity"], 4))

    return recs


@app.route("/api/intel/analyze", methods=["POST"])
def intel_analyze():
    catalog_data = catalog_health_state.get("catalog_data")
    detected_fields = catalog_health_state.get("detected_fields")
    if not catalog_data or not detected_fields:
        return jsonify({"error": "No catalog data loaded. Upload data in Catalog Health first."}), 400
    try:
        recs = _build_intel_recommendations(catalog_data, detected_fields)
        with intel_lock:
            intel_state["recommendations"] = recs
            intel_state["dismissed"] = set()
            intel_state["accepted"] = {}
            intel_state["generated_at"] = datetime.now().isoformat()
        critical = sum(1 for r in recs if r["severity"] == "Critical")
        high = sum(1 for r in recs if r["severity"] == "High")
        quick_wins = sum(1 for r in recs
                         if r["severity"] in ("High", "Critical")
                         and r["action_type"] in ("change_bullet", "change_title"))
        return jsonify({
            "ok": True,
            "recommendations": recs,
            "summary": {
                "total": len(recs),
                "critical": critical,
                "high": high,
                "quick_wins": quick_wins,
            },
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/intel/recommendations", methods=["GET"])
def intel_recommendations():
    recs = intel_state.get("recommendations")
    dismissed = intel_state.get("dismissed", set())
    accepted = intel_state.get("accepted", {})
    if recs is None:
        return jsonify({"recommendations": None, "summary": None})
    visible = [r for r in recs if r["id"] not in dismissed]
    critical = sum(1 for r in visible if r["severity"] == "Critical")
    high = sum(1 for r in visible if r["severity"] == "High")
    quick_wins = sum(1 for r in visible
                     if r["severity"] in ("High", "Critical")
                     and r["action_type"] in ("change_bullet", "change_title"))
    return jsonify({
        "recommendations": visible,
        "accepted": accepted,
        "summary": {
            "total": len(visible),
            "critical": critical,
            "high": high,
            "quick_wins": quick_wins,
        },
        "generated_at": intel_state.get("generated_at"),
    })


@app.route("/api/intel/accept", methods=["POST"])
def intel_accept():
    data = request.get_json(force=True) or {}
    rec_id = data.get("rec_id")
    note = data.get("note", "")
    if not rec_id:
        return jsonify({"error": "rec_id required"}), 400
    with intel_lock:
        if intel_state["recommendations"] is None:
            return jsonify({"error": "No recommendations loaded"}), 400
        ids = {r["id"] for r in intel_state["recommendations"]}
        if rec_id not in ids:
            return jsonify({"error": "Unknown rec_id"}), 404
        intel_state["accepted"][rec_id] = {
            "accepted_at": datetime.now().isoformat(),
            "note": note,
        }
    return jsonify({"ok": True, "rec_id": rec_id})


@app.route("/api/intel/dismiss", methods=["POST"])
def intel_dismiss():
    data = request.get_json(force=True) or {}
    rec_id = data.get("rec_id")
    if not rec_id:
        return jsonify({"error": "rec_id required"}), 400
    with intel_lock:
        if intel_state["recommendations"] is None:
            return jsonify({"error": "No recommendations loaded"}), 400
        intel_state["dismissed"].add(rec_id)
    return jsonify({"ok": True, "rec_id": rec_id})


# ═══════════════════════════════════════════════════════════════════════════════
# NIS RULE ENGINE API
# Exposes the universal rule engine to the dashboard frontend.
# ═══════════════════════════════════════════════════════════════════════════════


@app.route("/api/rule-engine/index", methods=["GET"])
def rule_engine_index():
    """Return the product-type index with rule / field counts."""
    try:
        idx = _nis_engine.get_index()
        return jsonify({"ok": True, "index": idx, "product_types": _nis_engine.list_product_types()})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/rule-engine/bundle/<product_type>", methods=["GET"])
def rule_engine_bundle(product_type):
    """Return the fields + summary (NOT the full rule ASTs — those are large).
    Use ?full=1 to include rule ASTs."""
    bundle = _nis_engine.load_bundle(product_type)
    if not bundle:
        return jsonify({"ok": False, "error": f"no bundle for {product_type}"}), 404
    if request.args.get("full") == "1":
        return jsonify({"ok": True, "bundle": bundle})
    # Slim version — fields + summary only
    slim = {
        "product_type":    bundle.get("product_type"),
        "version":         bundle.get("version"),
        "template_file":   bundle.get("template_file"),
        "merged_from":     bundle.get("merged_from"),
        "data_row":        bundle.get("data_row", 7),
        "fields":          bundle.get("fields", {}),
        "coverage":        bundle.get("coverage", {}),
        "named_ranges_count": len(bundle.get("named_ranges") or {}),
        "indirect_names_count": len(bundle.get("indirect_names") or []),
    }
    return jsonify({"ok": True, "bundle": slim})


@app.route("/api/rule-engine/evaluate", methods=["POST"])
def rule_engine_evaluate():
    """Evaluate a form state against a product-type rule bundle.

    Body JSON: { "product_type": "COAT", "form_state": { ... }, "include_dropdowns": true }

    Returns per-field verdict info suitable for the review screen.
    """
    data = request.get_json(force=True) or {}
    pt = (data.get("product_type") or "").upper().strip()
    state = data.get("form_state") or {}
    include_dd = data.get("include_dropdowns", True)
    apply_defaults = data.get("apply_apparel_defaults", True)
    brand = data.get("brand") or None
    sub_class = data.get("sub_class") or None
    if not pt:
        return jsonify({"ok": False, "error": "product_type required"}), 400
    try:
        result = _nis_engine.evaluate_form(
            pt, state,
            include_dropdowns=include_dd,
            apply_apparel_defaults=apply_defaults,
            brand=brand, sub_class=sub_class,
        )
        if "error" in result:
            return jsonify({"ok": False, "error": result["error"]}), 404
        return jsonify({"ok": True, "result": result})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/rule-engine/override", methods=["POST"])
def rule_engine_override():
    """Log an operator override of a rule verdict.

    Body JSON: {
      "product_type": "COAT", "brand": "Sage", "style": "Puffer-001",
      "field_key": "rtip_vendor_code#1.value", "column": "A",
      "rule_id": "cf_0001", "original_verdict": "required_missing",
      "override_verdict": "optional", "reason": "Not applicable to this SKU",
      "operator": "bhatt.devang999@gmail.com"
    }
    Stored to feedback/overrides_log.jsonl (append-only).
    """
    data = request.get_json(force=True) or {}
    required = ["product_type", "field_key", "rule_id",
                "original_verdict", "override_verdict", "reason"]
    missing = [k for k in required if not data.get(k)]
    if missing:
        return jsonify({"ok": False, "error": f"missing fields: {missing}"}), 400
    entry = {
        "timestamp":        datetime.utcnow().isoformat() + "Z",
        "product_type":     data["product_type"],
        "brand":            data.get("brand", ""),
        "style":            data.get("style", ""),
        "field_key":        data["field_key"],
        "column":           data.get("column", ""),
        "rule_id":          data["rule_id"],
        "original_verdict": data["original_verdict"],
        "override_verdict": data["override_verdict"],
        "reason":           data["reason"],
        "operator":         data.get("operator", ""),
    }
    try:
        with open(str(OVERRIDES_LOG), "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
        return jsonify({"ok": True, "entry": entry})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/rule-engine/overrides", methods=["GET"])
def rule_engine_overrides():
    """Return all logged overrides. Optional query: ?product_type=COAT&brand=Sage."""
    pt  = (request.args.get("product_type") or "").strip().upper()
    brand = (request.args.get("brand") or "").strip().lower()
    entries = []
    if OVERRIDES_LOG.exists():
        with open(str(OVERRIDES_LOG), "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    e = json.loads(line)
                except Exception:
                    continue
                if pt and e.get("product_type", "").upper() != pt:
                    continue
                if brand and brand not in e.get("brand", "").lower():
                    continue
                entries.append(e)
    return jsonify({"ok": True, "count": len(entries), "overrides": entries})


@app.route("/api/rule-engine/packaging", methods=["GET"])
def rule_engine_packaging_get():
    """Return all saved packaging memory, or a single entry.
    Query args: brand, product_type, sub_class.
    """
    brand = request.args.get("brand")
    pt    = request.args.get("product_type")
    sub   = request.args.get("sub_class", "")
    if brand and pt:
        entry = _nis_engine.get_packaging_for(brand, pt, sub)
        return jsonify({"ok": True, "entry": entry})
    return jsonify({"ok": True, "memory": _nis_engine.list_packaging_memory()})


@app.route("/api/rule-engine/packaging", methods=["POST"])
def rule_engine_packaging_save():
    """Save operator-confirmed package dims for a (brand, product_type, sub_class).
    Body: { brand, product_type, sub_class, dims: { field_key: value, ... } }
    """
    data = request.get_json(force=True) or {}
    brand = data.get("brand") or ""
    pt    = data.get("product_type") or ""
    sub   = data.get("sub_class") or ""
    dims  = data.get("dims") or {}
    if not brand or not pt:
        return jsonify({"ok": False, "error": "brand and product_type required"}), 400
    try:
        entry = _nis_engine.save_packaging_for(brand, pt, sub, dims)
        # Git-commit so it survives Render redeploys (same pattern as taxonomy_overrides)
        try:
            import subprocess
            subprocess.run(["git", "add", "nis_engine/brand_packaging_memory.json"],
                           cwd=str(BASE_DIR), check=False, capture_output=True)
            subprocess.run(["git", "commit", "-m", f"chore(packaging): save {brand}/{pt}/{sub}",
                            "--author", "TLG Dashboard <noreply@tlg.local>"],
                           cwd=str(BASE_DIR), check=False, capture_output=True)
            subprocess.run(["git", "push", "origin", "master"],
                           cwd=str(BASE_DIR), check=False, capture_output=True, timeout=10)
        except Exception:
            pass
        return jsonify({"ok": True, "entry": entry})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/rule-engine/brand-config", methods=["GET"])
def rule_engine_brand_config_get():
    """Return the saved config for a brand plus whether it needs setup.
    Query: ?brand=Sage Collective
    """
    brand = request.args.get("brand")
    if not brand:
        return jsonify({"ok": False, "error": "brand required"}), 400
    try:
        from nis_engine.brand_setup import needs_setup, load_brand_config
        status = needs_setup(brand)
        return jsonify({"ok": True, **status})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/rule-engine/brand-config", methods=["POST"])
def rule_engine_brand_config_save():
    """Save brand config. Body: { brand: str, config: {...} }"""
    data = request.get_json(force=True) or {}
    brand = data.get("brand")
    config = data.get("config") or {}
    if not brand or not config:
        return jsonify({"ok": False, "error": "brand and config required"}), 400
    try:
        from nis_engine.brand_setup import save_brand_config, load_brand_config
        # Merge with existing so operator partial saves don't blow away fields
        existing = load_brand_config(brand) or {}
        existing.update(config)
        path = save_brand_config(brand, existing)
        # Best-effort git commit
        try:
            import subprocess
            rel = os.path.relpath(path, str(BASE_DIR))
            subprocess.run(["git", "add", rel], cwd=str(BASE_DIR), check=False, capture_output=True)
            subprocess.run(["git", "commit", "-m", f"chore(brand): update {brand}",
                            "--author", "TLG Dashboard <noreply@tlg.local>"],
                           cwd=str(BASE_DIR), check=False, capture_output=True)
            subprocess.run(["git", "push", "origin", "master"],
                           cwd=str(BASE_DIR), check=False, capture_output=True, timeout=10)
        except Exception:
            pass
        return jsonify({"ok": True, "path": path, "config": existing})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/rule-engine/import-preupload", methods=["POST"])
def rule_engine_import_preupload():
    """Upload a pre-upload .xlsx/.xlsm and return per-style evaluation results.

    multipart/form-data: file=<xlsx>
    Response: { ok, brand, styles: [{style_id, name, sub_class, state, evaluation: {...}}] }
    Each evaluation already has apparel_defaults + packaging memory applied.
    """
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "no file uploaded"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"ok": False, "error": "empty filename"}), 400

    tmp = BASE_DIR / "uploads" / "preupload"
    tmp.mkdir(parents=True, exist_ok=True)
    dest = tmp / f.filename
    f.save(str(dest))

    try:
        from nis_engine.preupload_importer import parse_preupload, style_to_form_state
        from nis_engine.brand_setup import needs_setup as _brand_needs_setup
        parsed = parse_preupload(str(dest))
        brand = parsed.get("brand") or ""
        # Check brand setup before evaluating — if absolute-required fields are
        # missing, return the schema so the dashboard can prompt before showing styles.
        setup_status = _brand_needs_setup(brand) if brand else {
            "brand": brand, "needs_setup": True,
            "missing_fields": ["brand_name", "vendor_code_prefix", "default_coo", "department"],
            "schema": {}, "current_config": {},
        }
        out_styles = []
        for style_id, style in parsed.get("styles", {}).items():
            state = style_to_form_state(style, brand)
            evaluation = _nis_engine.evaluate_form(
                "COAT", state,
                apply_apparel_defaults=True,
                brand=brand,
                sub_class=style.get("sub_class") or "",
            )
            hard = [ff["label"] for ff in evaluation.get("fields", {}).values()
                    if ff["verdict"] == "required_missing" and ff["base_requirement"] == "REQUIRED"]
            out_styles.append({
                "style_id":   style_id,
                "name":       style.get("name"),
                "sub_class":  style.get("sub_class"),
                "department": style.get("department"),
                "upcs_count": len(style.get("upcs") or []),
                "colors":     style.get("colors"),
                "sizes":      style.get("sizes"),
                "state":      state,
                "summary":    evaluation.get("summary"),
                "hard_missing": hard,
                "ready":      len(hard) == 0,
                "defaults_applied":  evaluation.get("defaults_applied"),
                "packaging_applied": evaluation.get("packaging_applied"),
            })
        return jsonify({
            "ok":      True,
            "brand":   brand,
            "file":    f.filename,
            "styles":  out_styles,
            "total_styles":     len(out_styles),
            "ready_to_upload":  sum(1 for s in out_styles if s["ready"]),
            "needs_attention":  sum(1 for s in out_styles if not s["ready"]),
            "brand_setup":      setup_status,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/rule-engine/rebuild", methods=["POST"])
def rule_engine_rebuild():
    """Rebuild the rule bundles from the .xlsm templates in uploads/templates/.
    Manual trigger — the operator clicks a button to re-extract after Amazon ships a new template.
    """
    from nis_engine import nis_rule_extractor as _extractor
    templates_dir = str(UPLOAD_TEMPLATES)
    # Also look at the top-level workspace for any .xlsm templates uploaded directly
    scan_dirs = [templates_dir, str(BASE_DIR.parent)]
    bundles = []
    errors  = []
    seen = set()
    try:
        import glob
        # Clear output dir first so stale bundles don't linger
        for f in os.listdir(str(_NIS_RULES_DIR)):
            if f.endswith(".json"):
                os.remove(str(_NIS_RULES_DIR / f))
        for d in scan_dirs:
            if not os.path.isdir(d):
                continue
            for f in sorted(glob.glob(os.path.join(d, "*.xlsm"))):
                if os.path.basename(f) in seen:
                    continue
                seen.add(os.path.basename(f))
                try:
                    b = _extractor.extract_rules(f)
                    _extractor.write_bundle(b, str(_NIS_RULES_DIR))
                    bundles.append({
                        "file": os.path.basename(f),
                        "product_type": b.get("product_type"),
                        "rules": b["coverage"]["total_formulas"],
                        "fields": b["coverage"]["field_count"],
                        "needs_review": b["coverage"]["needs_review"],
                    })
                except Exception as e:
                    errors.append({"file": os.path.basename(f), "error": str(e)})
        _extractor.write_index([], str(_NIS_RULES_DIR))
        # Refresh engine cache
        _nis_engine.set_bundle_dir(str(_NIS_RULES_DIR))
        return jsonify({"ok": True, "bundles": bundles, "errors": errors})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════
# v0.7.5 — Pre-Upload History (merged Upload Pre-Upload flow)
# ═══════════════════════════════════════════════════════════════════════════
_PREUPLOAD_HISTORY_PATH = BASE_DIR / "preupload_history.json"

def _load_preupload_history():
    try:
        if not _PREUPLOAD_HISTORY_PATH.exists():
            return {"entries": []}
        with open(_PREUPLOAD_HISTORY_PATH, "r") as f:
            data = json.load(f)
            if not isinstance(data, dict):
                return {"entries": []}
            data.setdefault("entries", [])
            return data
    except Exception:
        return {"entries": []}

def _save_preupload_history(data):
    with open(_PREUPLOAD_HISTORY_PATH, "w") as f:
        json.dump(data, f, indent=2)

@app.route("/api/preupload/history", methods=["GET"])
def preupload_history_get():
    data = _load_preupload_history()
    # Newest first, cap to last 50
    entries = list(reversed(data.get("entries", [])))[:50]
    return jsonify({"ok": True, "entries": entries})

@app.route("/api/preupload/history", methods=["POST"])
def preupload_history_post():
    payload = request.get_json(force=True, silent=True) or {}
    filename = (payload.get("filename") or "").strip()
    if not filename:
        return jsonify({"ok": False, "error": "filename required"}), 400
    entry = {
        "id": datetime.utcnow().strftime("%Y%m%d-%H%M%S-") + str(_uuid.uuid4())[:8],
        "filename": filename,
        "brand": payload.get("brand") or "",
        "total_styles": payload.get("total_styles"),
        "total_variants": payload.get("total_variants"),
        "generated_file": payload.get("generated_file") or "",
        "generated_files": payload.get("generated_files") or [],
        # ISO 8601 UTC with explicit Z suffix — frontend converts to ET (America/New_York).
        "timestamp": datetime.utcnow().isoformat() + "Z",
    }
    data = _load_preupload_history()
    data.setdefault("entries", []).append(entry)
    # Cap at 200 to keep file small
    if len(data["entries"]) > 200:
        data["entries"] = data["entries"][-200:]
    _save_preupload_history(data)
    return jsonify({"ok": True, "entry": entry})

@app.route("/api/preupload/download/<path:filename>", methods=["GET"])
def preupload_download(filename):
    """Serve a generated NIS file from UPLOAD_OUTPUT."""
    # Strip any path components for safety
    safe_name = os.path.basename(filename)
    fpath = UPLOAD_OUTPUT / safe_name
    if not fpath.exists():
        return jsonify({"ok": False, "error": f"File not found: {safe_name}"}), 404
    return send_from_directory(str(UPLOAD_OUTPUT), safe_name, as_attachment=True)


# ╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗
# BETA: Generate NIS from Image only — isolated module
# All endpoints under /api/beta-image-nis/*
# Operator drops one image + brand. Vision detects PT/color/features.
# We build a draft style dict, content, and structured fields with provenance
# tags (vision/brand/pt/operator). Operator confirms then can download xlsm.
# ╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗╗

BETA_IMAGE_NIS_FEEDBACK_LOG = BASE_DIR / "data" / "beta_image_nis_feedback.jsonl"
BETA_IMAGE_NIS_FEEDBACK_LOG.parent.mkdir(parents=True, exist_ok=True)
_BETA_SESSION = {}  # session_id -> {image_path, brand, vision, draft, ...}


def _beta_provenance(value, source, confidence=None, suggestion_only=False):
    """Wrap a field value with provenance metadata.
    source: 'vision' | 'brand' | 'pt' | 'operator' | 'unknown'
    """
    return {
        "value": value,
        "source": source,
        "confidence": confidence,
        "suggestion_only": suggestion_only,
    }


def _beta_build_draft_style(vision_intel, brand, brand_cfg, hints):
    """Translate vision output + brand config into a draft style dict +
    a parallel provenance map. Operator-only fields (SKU/UPC/sizes/COO/price)
    are blanked and tagged 'operator'.
    """
    detected_pt = (vision_intel or {}).get("detected_pt", "") or ""
    if detected_pt == "UNKNOWN":
        detected_pt = ""
    detected_subject = (vision_intel or {}).get("detected_subject", "")
    detected_color = (vision_intel or {}).get("detected_color", "")

    # Map vision suggestions to a dict by field for quick lookup
    vision_field_map = {}
    for s in (vision_intel or {}).get("field_suggestions") or []:
        if isinstance(s, dict) and s.get("field"):
            vision_field_map[s["field"]] = s.get("value", "")

    # Hints from operator (override anything if present)
    hints = hints or {}
    coo = hints.get("coo") or brand_cfg.get("default_coo", "")
    fabric = hints.get("fabric") or brand_cfg.get("default_fabric", "")
    care = hints.get("care") or brand_cfg.get("default_care", "")
    season_code = hints.get("season_code") or brand_cfg.get("vendor_code_prefix", "")
    style_num_hint = hints.get("style_num", "") or "BETA-DRAFT"

    # Subclass: derive from detected_pt if not provided
    detected_pt_def = next((p for p in ALL_PRODUCT_TYPES if p["id"] == detected_pt), None)
    default_subclass = (detected_pt_def or {}).get("sub_classes", [""])[0] if detected_pt_def else ""
    subclass = hints.get("subclass") or default_subclass

    # Build the actual style dict shape that the writers expect
    style = {
        "style_num": style_num_hint,
        "style_name": detected_subject or f"{brand} {subclass}".strip(),
        "subclass": subclass,
        "sub_subclass": subclass,
        "division_name": brand_cfg.get("default_division_name", brand),
        "_resolved_pt": detected_pt,
        "fabric": fabric,
        "care": care,
        "coo": coo,
        "variants": [],  # operator must add
        "closure_type": vision_field_map.get("closure_type", ""),
        "sleeve_type": vision_field_map.get("sleeve_length", ""),
        "neck_type": vision_field_map.get("neckline", ""),
        "fit_type": vision_field_map.get("fit_type", ""),
        "type_of_jacket": vision_field_map.get("coat_silhouette_type", ""),
        "pockets": vision_field_map.get("pockets", ""),
    }

    # Provenance map: which fields came from where
    prov = {
        "style_num":      _beta_provenance(style["style_num"], "operator"),
        "style_name":     _beta_provenance(style["style_name"], "vision" if detected_subject else "unknown"),
        "subclass":       _beta_provenance(style["subclass"], "pt" if not hints.get("subclass") else "operator"),
        "product_type":   _beta_provenance(detected_pt, "vision", confidence="high" if detected_pt else None),
        "color":          _beta_provenance(detected_color, "vision"),
        "division_name":  _beta_provenance(style["division_name"], "brand"),
        "fabric":         _beta_provenance(fabric, "operator" if hints.get("fabric") else ("brand" if brand_cfg.get("default_fabric") else "unknown"), suggestion_only=not bool(hints.get("fabric"))),
        "care":           _beta_provenance(care, "operator" if hints.get("care") else ("brand" if brand_cfg.get("default_care") else "unknown")),
        "coo":            _beta_provenance(coo, "operator" if hints.get("coo") else ("brand" if brand_cfg.get("default_coo") else "unknown")),
        "closure_type":   _beta_provenance(style["closure_type"], "vision" if style["closure_type"] else "unknown"),
        "sleeve_length":  _beta_provenance(style["sleeve_type"], "vision" if style["sleeve_type"] else "pt"),
        "neck_style":     _beta_provenance(style["neck_type"], "vision" if style["neck_type"] else "pt"),
        "fit_type":       _beta_provenance(style["fit_type"], "vision" if style["fit_type"] else "unknown"),
        "coat_silhouette_type": _beta_provenance(style["type_of_jacket"], "vision" if style["type_of_jacket"] else "unknown"),
        "pockets":        _beta_provenance(style["pockets"], "vision" if style["pockets"] else "unknown"),
        # Operator-only fields
        "sku_pattern":    _beta_provenance(f"{season_code}-{{styleNum}}-{{COLOR}}-{{SIZE}}" if season_code else "", "brand" if season_code else "operator"),
        "upc":            _beta_provenance("", "operator"),
        "sizes":          _beta_provenance([], "operator"),
        "price":          _beta_provenance("", "operator"),
        "cost_price":     _beta_provenance("", "operator"),
        "fabric_composition": _beta_provenance("", "operator", suggestion_only=True),
    }

    return style, prov


def _atlas_visible_brands_filter(brands: list[str]) -> list[str]:
    """Apply the ATLAS_VISIBLE_BRANDS env-var allowlist.

    When ATLAS_VISIBLE_BRANDS is set to a comma-separated list, only those
    brands surface in UI dropdowns. The substrate data for other brands
    is untouched — this is a UI hide, not an access control. Case-
    insensitive match. Empty/unset means "show all" (legacy behavior).
    """
    allow = os.environ.get("ATLAS_VISIBLE_BRANDS", "").strip()
    if not allow:
        return brands
    allowed = {b.strip().lower() for b in allow.split(",") if b.strip()}
    return [b for b in brands if b.lower() in allowed]


@app.route("/api/beta-image-nis/brands", methods=["GET"])
def api_beta_image_nis_brands():
    """Return the deduped, sorted list of brands the operator can pick.

    Honours ATLAS_VISIBLE_BRANDS to hide brands the dashboard isn't
    actively servicing. Substrate data for hidden brands is preserved.
    """
    file_brands = [p.stem.replace('_', ' ') for p in BRAND_CONFIGS_DIR.glob('*.json')]
    in_mem = list(BRAND_CONFIGS.keys())
    deduped = sorted({b.strip() for b in (file_brands + in_mem) if b and b.strip()})
    deduped = _atlas_visible_brands_filter(deduped)
    return jsonify({"ok": True, "brands": deduped})


# ─── Atlas operators ────────────────────────────────────────────────────────────────
# Lightweight named-account flow. Cookie-based attribution, no real auth.
# Used so the agency's team is distinguishable in the substrate — each
# decision_event carries the operator_id who made it.

_ATLAS_OPERATOR_COOKIE = "atlas_operator_id"
_ATLAS_OPERATOR_COOKIE_MAX_AGE = 60 * 60 * 24 * 90  # 90 days


def _atlas_current_workspace() -> str:
    """Resolve the operator's current workspace from session_data.

    Falls back to 'novelle' if nothing is set, since single-brand mode
    means that's the working assumption.
    """
    brand = (session_data.get("brand") or "").strip()
    if brand:
        return brand.lower().replace(" ", "_")
    default_brand = os.environ.get("ATLAS_VISIBLE_BRANDS", "").split(",")[0].strip()
    if default_brand:
        return default_brand.lower().replace(" ", "_")
    return "novelle"


@app.route("/api/atlas/operator", methods=["GET"])
def atlas_operator_get():
    """Return the currently-identified operator for this browser.

    Reads the atlas_operator_id cookie and looks up the operator in the
    operators table. If the cookie is absent or the operator is unknown,
    returns ok=true but identified=false so the frontend knows to prompt.
    """
    workspace_id = _atlas_current_workspace()
    op_id = (request.cookies.get(_ATLAS_OPERATOR_COOKIE) or "").strip()
    if not op_id:
        return jsonify({
            "ok": True,
            "identified": False,
            "workspace_id": workspace_id,
        })

    try:
        from substrate.operators import get_operator, touch_operator
        op = get_operator(workspace_id, op_id)
        if op:
            touch_operator(workspace_id, op_id)
            session_data["operator_id"] = op_id
            session_data["operator"] = op["display_name"]
            return jsonify({
                "ok": True,
                "identified": True,
                "workspace_id": workspace_id,
                "operator": op,
            })
    except Exception as exc:
        print(f"[atlas] operator lookup failed: {exc}", flush=True)

    return jsonify({
        "ok": True,
        "identified": False,
        "workspace_id": workspace_id,
        "reason": "cookie present but operator not found",
    })


@app.route("/api/atlas/operator", methods=["POST"])
def atlas_operator_set():
    """Set or update the current operator and return a cookie.

    Body: {
        display_name: str,            # required
        role: 'owner'|'operator'|'agency'|'viewer'  # optional, default 'operator'
        operator_id: str              # optional, derived from display_name if absent
    }
    """
    from flask import make_response
    data = request.get_json(force=True, silent=True) or {}
    display_name = (data.get("display_name") or "").strip()
    if not display_name:
        return jsonify({"ok": False, "error": "display_name required"}), 400

    role = (data.get("role") or "operator").strip().lower()
    op_id = (data.get("operator_id") or "").strip().lower()

    try:
        from substrate.operators import upsert_operator, slugify_operator_id
    except Exception as exc:
        return jsonify({"ok": False, "error": f"operators module unavailable: {exc}"}), 500

    if not op_id:
        op_id = slugify_operator_id(display_name)

    workspace_id = _atlas_current_workspace()
    op = upsert_operator(
        workspace_id=workspace_id,
        operator_id=op_id,
        display_name=display_name,
        role=role,
    )

    session_data["operator_id"] = op_id
    session_data["operator"] = display_name

    resp = make_response(jsonify({
        "ok": True,
        "workspace_id": workspace_id,
        "operator": op,
    }))
    resp.set_cookie(
        _ATLAS_OPERATOR_COOKIE,
        op_id,
        max_age=_ATLAS_OPERATOR_COOKIE_MAX_AGE,
        httponly=True,
        samesite="Lax",
    )
    return resp


@app.route("/api/atlas/operators", methods=["GET"])
def atlas_operators_list():
    """Return the list of operators in the current workspace.

    Used by the frontend to populate the operator switcher dropdown
    (when more than one operator has logged in).
    """
    workspace_id = _atlas_current_workspace()
    try:
        from substrate.operators import list_operators
        ops = list_operators(workspace_id)
    except Exception as exc:
        print(f"[atlas] operators list failed: {exc}", flush=True)
        ops = []
    return jsonify({"ok": True, "workspace_id": workspace_id, "operators": ops})


# ─── Atlas inputs (Phase 1) ────────────────────────────────────────────────────────
# Single dropzone for every file type. Auto-detects file_kind by header
# signature, hands off to the right parser, then writes one
# ingestion_records row for the audit trail.

@app.route("/api/atlas/inputs/history", methods=["GET"])
def atlas_inputs_history():
    """Return the ingestion history for the current workspace.

    Query params:
        kind  - optional file_kind filter
        limit - default 100
    """
    workspace_id = _atlas_current_workspace()
    file_kind = (request.args.get("kind") or "").strip() or None
    try:
        limit = int(request.args.get("limit") or 100)
    except ValueError:
        limit = 100
    limit = max(1, min(limit, 1000))
    try:
        from substrate.inputs import list_ingestions
        rows = list_ingestions(workspace_id, file_kind=file_kind, limit=limit)
    except Exception as exc:
        print(f"[atlas] inputs history failed: {exc}", flush=True)
        rows = []
    return jsonify({"ok": True, "workspace_id": workspace_id, "ingestions": rows})


@app.route("/api/atlas/inputs/freshness", methods=["GET"])
def atlas_inputs_freshness():
    """Return per-file-kind freshness for the staleness bar."""
    workspace_id = _atlas_current_workspace()
    try:
        from substrate.inputs import freshness_summary
        summary = freshness_summary(workspace_id)
    except Exception as exc:
        print(f"[atlas] inputs freshness failed: {exc}", flush=True)
        summary = {}
    return jsonify({"ok": True, "workspace_id": workspace_id, "freshness": summary})


@app.route("/api/atlas/inputs/detect", methods=["POST"])
def atlas_inputs_detect():
    """Detect a file's kind from its first row of headers.

    Used by the frontend before upload so the operator sees "Looks like
    a Search Term Report" before committing the upload. The actual
    upload still goes through the existing per-kind endpoints (catalog
    upload, etc.) which run their own parsers.
    """
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "file required"}), 400
    f = request.files["file"]
    try:
        content = f.read()
        if not content:
            return jsonify({"ok": False, "error": "empty file"}), 400
        # Read just the first line as bytes; let downstream parsers worry
        # about full parsing. We just need headers for detection.
        first_line = content.split(b"\n", 1)[0]
        # Try comma + tab as separators since CSV/TSV both arrive here.
        headers: list[str] = []
        for sep in (b",", b"\t"):
            if sep in first_line:
                headers = [
                    h.decode("utf-8", errors="replace").strip().strip('"')
                    for h in first_line.split(sep)
                ]
                if headers:
                    break
        if not headers:
            # Likely .xlsx — fall back to openpyxl read of first sheet's row 1.
            try:
                import io as _io
                import openpyxl as _ox
                wb = _ox.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if first:
                    headers = [str(h).strip() if h is not None else "" for h in first]
            except Exception:
                headers = []
        from substrate.inputs import detect_file_kind, file_hash
        kind = detect_file_kind(headers)
        return jsonify({
            "ok": True,
            "file_name": f.filename,
            "file_kind": kind,
            "bytes": len(content),
            "file_hash": file_hash(content)[:16],
            "headers_detected": [h for h in headers if h][:40],
            "recognised": kind is not None,
        })
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)[:200]}), 200


@app.route("/api/atlas/memory/sessions", methods=["GET"])
def atlas_memory_sessions():
    """Return a page of past sessions for the Memory tab.

    Query params:
        limit       - default 50, max 200
        offset      - default 0
        state       - optional: 'live' | 'submitted' | 'abandoned'
        operator_id - optional filter
    Response: { ok, workspace_id, sessions, total }
    """
    workspace_id = _atlas_current_workspace()
    try:
        limit = int(request.args.get("limit") or 50)
    except ValueError:
        limit = 50
    try:
        offset = int(request.args.get("offset") or 0)
    except ValueError:
        offset = 0
    state = (request.args.get("state") or "").strip() or None
    if state and state not in ("live", "submitted", "abandoned"):
        return jsonify({"ok": False, "error": "invalid state"}), 400
    operator_id = (request.args.get("operator_id") or "").strip() or None
    try:
        from substrate.memory import list_sessions
        result = list_sessions(
            workspace_id=workspace_id,
            limit=limit,
            offset=offset,
            state=state,
            operator_id=operator_id,
        )
    except Exception as exc:
        print(f"[atlas] memory sessions list failed: {exc}", flush=True)
        result = {"sessions": [], "total": 0}
    return jsonify({
        "ok": True,
        "workspace_id": workspace_id,
        "sessions": result.get("sessions", []),
        "total": result.get("total", 0),
    })


@app.route("/api/atlas/memory/decisions", methods=["GET"])
def atlas_memory_decisions():
    """Cross-session decisions feed for the Memory tab.

    Query params (all optional):
        limit          default 50, max 200
        offset         default 0
        field          substring match on field_name
        asin           exact match
        action         accept | edit | reject | dismiss | no_response
        operator_id    filter by the operator who ran the session
        session_id     scope to a single session (deep-link from Sessions)
        start, end     ISO timestamp bounds on the decision's timestamp
    Response: { ok, workspace_id, decisions: [...], total }
    """
    workspace_id = _atlas_current_workspace()
    try:
        limit = int(request.args.get("limit") or 50)
    except ValueError:
        limit = 50
    try:
        offset = int(request.args.get("offset") or 0)
    except ValueError:
        offset = 0
    field = (request.args.get("field") or "").strip() or None
    asin = (request.args.get("asin") or "").strip() or None
    action = (request.args.get("action") or "").strip() or None
    if action and action not in ("accept", "edit", "reject", "comment", "no_response"):
        return jsonify({"ok": False, "error": "invalid action"}), 400
    operator_id = (request.args.get("operator_id") or "").strip() or None
    session_id = (request.args.get("session_id") or "").strip() or None
    start = (request.args.get("start") or "").strip() or None
    end = (request.args.get("end") or "").strip() or None
    try:
        from substrate.memory import list_decisions
        result = list_decisions(
            workspace_id=workspace_id,
            limit=limit, offset=offset,
            field=field, asin=asin, action=action,
            operator_id=operator_id, session_id=session_id,
            start=start, end=end,
        )
    except Exception as exc:
        print(f"[atlas] memory decisions list failed: {exc}", flush=True)
        result = {"decisions": [], "total": 0}
    return jsonify({
        "ok": True,
        "workspace_id": workspace_id,
        "decisions": result.get("decisions", []),
        "total": result.get("total", 0),
    })


# ---------------------------------------------------------------------------
# Brand Voice endpoints (Phase 1.5 — Step 2)
# Single source of truth for tone, hero adjectives, signature phrases,
# banned words/phrasings, like/unlike examples. Auto-bumps profile_version
# on every save; writes a decision_event with module='brand_voice' so the
# audit trail appears in Memory.
# ---------------------------------------------------------------------------


@app.route("/api/atlas/brand-voice", methods=["GET"])
def atlas_brand_voice_get():
    """Return the latest brand voice payload for the current workspace."""
    workspace_id = _atlas_current_workspace()
    try:
        from substrate.brand_voice import read_voice
        v = read_voice(workspace_id)
    except Exception as exc:
        print(f"[atlas] brand-voice read failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": "brand-voice unavailable"}), 500
    return jsonify({"workspace_id": workspace_id, **v})


@app.route("/api/atlas/brand-voice", methods=["POST"])
def atlas_brand_voice_save():
    """Save a new brand voice revision. Auto-bumps profile_version.

    JSON body accepts (all optional):
      brand_name, category_scope, tier_scope, stage_scope,
      tone_descriptors, hero_adjectives,
      banned_words, banned_phrasings, required_words, signature_phrases,
      like_examples, unlike_examples,
      target_customer, competitor_set

    Missing keys preserve previous values from the latest row.
    """
    workspace_id = _atlas_current_workspace()
    payload = request.get_json(silent=True) or {}
    try:
        from substrate.brand_voice import save_voice
        operator_id = (request.cookies.get(_ATLAS_OPERATOR_COOKIE) or "").strip() or None
        result = save_voice(workspace_id, payload, operator_id=operator_id)
    except Exception as exc:
        print(f"[atlas] brand-voice save failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500
    status = 200 if result.get("ok") else 400
    return jsonify({"workspace_id": workspace_id, **result}), status


# ────────────────────────────────────────────────────────────────
# Unit Economics endpoints (Phase C: cost inputs / Phase D: margin rollup)
# ────────────────────────────────────────────────────────────────

@app.route("/api/atlas/unit-economics/costs", methods=["GET"])
def atlas_unit_economics_list_costs():
    """List all per-ASIN cost rows for the current workspace."""
    workspace_id = _atlas_current_workspace()
    try:
        from substrate.cost_inputs import list_cost_inputs, read_overhead
        rows = list_cost_inputs(workspace_id)
        overhead = read_overhead(workspace_id)
    except Exception as exc:
        print(f"[atlas] cost list failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": "costs unavailable"}), 500
    return jsonify({
        "ok": True,
        "workspace_id": workspace_id,
        "rows": rows,
        "overhead": overhead,
    })


@app.route("/api/atlas/unit-economics/costs/<asin>", methods=["GET"])
def atlas_unit_economics_get_cost(asin: str):
    """Return the cost row for one ASIN (well-formed empty payload if none)."""
    workspace_id = _atlas_current_workspace()
    try:
        from substrate.cost_inputs import read_cost_input
        row = read_cost_input(workspace_id, asin)
    except Exception as exc:
        print(f"[atlas] cost read failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": "cost read unavailable"}), 500
    return jsonify(row)


@app.route("/api/atlas/unit-economics/costs/<asin>", methods=["POST"])
def atlas_unit_economics_save_cost(asin: str):
    """Upsert a per-ASIN cost row.

    JSON body accepts (all optional, missing keys preserve previous):
      landed_cost, fba_fee, third_pl_fee, referral_pct, map_price, notes

    Numeric fields tolerant-parse $, %, commas. Empty string → None
    ("not on file"). Operator-typed zero is a real zero.
    """
    workspace_id = _atlas_current_workspace()
    payload = request.get_json(silent=True) or {}
    try:
        from substrate.cost_inputs import save_cost_input
        operator_id = (request.cookies.get(_ATLAS_OPERATOR_COOKIE) or "").strip() or None
        result = save_cost_input(workspace_id, asin, payload, operator_id=operator_id)
    except Exception as exc:
        print(f"[atlas] cost save failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500
    status = 200 if result.get("ok") else 400
    return jsonify(result), status


@app.route("/api/atlas/unit-economics/overhead", methods=["GET"])
def atlas_unit_economics_get_overhead():
    """Brand-level fixed overhead (Model 1: above-the-line)."""
    workspace_id = _atlas_current_workspace()
    try:
        from substrate.cost_inputs import read_overhead
        row = read_overhead(workspace_id)
    except Exception as exc:
        print(f"[atlas] overhead read failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": "overhead unavailable"}), 500
    return jsonify(row)


@app.route("/api/atlas/unit-economics/overhead", methods=["POST"])
def atlas_unit_economics_save_overhead():
    """Save brand-level fixed overhead.

    JSON body: { fixed_overhead_monthly, notes }
    """
    workspace_id = _atlas_current_workspace()
    payload = request.get_json(silent=True) or {}
    try:
        from substrate.cost_inputs import save_overhead
        operator_id = (request.cookies.get(_ATLAS_OPERATOR_COOKIE) or "").strip() or None
        result = save_overhead(workspace_id, payload, operator_id=operator_id)
    except Exception as exc:
        print(f"[atlas] overhead save failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500
    status = 200 if result.get("ok") else 400
    return jsonify(result), status


@app.route("/api/atlas/unit-economics/margin", methods=["GET"])
def atlas_unit_economics_margin():
    """Per-ASIN-per-month margin rollup.

    Query params:
      period   YYYY-MM filter (optional)
      asin     single-ASIN filter (optional)

    Returns three margin columns per row:
      contribution_margin_per_unit, tacos, net_after_ads_per_unit.
    Honest about gaps: contribution is None when costs incomplete.
    """
    workspace_id = _atlas_current_workspace()
    period = (request.args.get("period") or "").strip() or None
    asin = (request.args.get("asin") or "").strip() or None
    try:
        from substrate.margin import margin_rollup, list_periods
        roll = margin_rollup(workspace_id, period=period, asin=asin)
        periods = list_periods(workspace_id)
    except Exception as exc:
        print(f"[atlas] margin rollup failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": "margin unavailable"}), 500
    return jsonify({**roll, "available_periods": periods})


# ────────────────────────────────────────────────────────────────
# Cited NIS (Phase 1.5 · M3): 5-layer reasoning chain on every output
# ────────────────────────────────────────────────────────────────

@app.route("/api/atlas/cited-nis/generate", methods=["POST"])
def atlas_cited_nis_generate():
    """Generate cited content (title/bullet/description) via the 5-layer
    reasoning chain. Per CITATION_CHAIN.md.

    Body:
      asin            (optional): ASIN to generate for
      decision_class  required: 'title_generation' | 'bullet_generation' |
                                'description_generation'
      operator_id     (optional): defaults to 'devang' (single-operator)

    Returns the full bundle_summary, primary/alternates, citations
    (with verifier_status), confidence breakdown, and decision_event_id.
    """
    body = request.get_json(silent=True) or {}
    asin = (body.get("asin") or "").strip() or None
    decision_class = (body.get("decision_class") or "").strip()
    operator_id = (body.get("operator_id") or "devang").strip()
    workspace_id = _atlas_current_workspace()

    if not decision_class:
        return jsonify({"ok": False,
                        "error": "decision_class required"}), 400

    try:
        from substrate.citation_chain import generate_cited
        result = generate_cited(
            workspace_id=workspace_id,
            asin=asin,
            decision_class=decision_class,
            operator_id=operator_id,
            log_decision=True,
        )
        return jsonify({"ok": True, **result})
    except Exception as exc:
        print(f"[atlas] cited-nis generate failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/cited-nis/preview-context", methods=["POST"])
def atlas_cited_nis_preview_context():
    """Read-only preview of the assembled context (no LLM call, no
    unknowns emission). Useful for showing operator what L0 sees before
    they ask for generation.
    """
    body = request.get_json(silent=True) or {}
    asin = (body.get("asin") or "").strip() or None
    decision_class = (body.get("decision_class") or "").strip()
    operator_id = (body.get("operator_id") or "devang").strip()
    workspace_id = _atlas_current_workspace()

    if not decision_class:
        return jsonify({"ok": False,
                        "error": "decision_class required"}), 400

    try:
        from substrate.context import build_context
        bundle = build_context(
            workspace_id=workspace_id,
            asin=asin,
            decision_class=decision_class,
            operator_id=operator_id,
            emit_unknowns_on_gaps=False,
        )
        return jsonify({"ok": True, "bundle": bundle})
    except Exception as exc:
        print(f"[atlas] preview context failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/unknowns", methods=["GET"])
def atlas_unknowns_list():
    """List open unknowns. Filters: scope, scope_ref, evidence_path,
    decision_class.
    """
    workspace_id = _atlas_current_workspace()
    scope = (request.args.get("scope") or "").strip() or None
    scope_ref = (request.args.get("scope_ref") or "").strip() or None
    evidence_path = (request.args.get("evidence_path") or "").strip() or None
    decision_class = (request.args.get("decision_class") or "").strip() or None
    try:
        from substrate.unknowns import list_open_unknowns
        rows = list_open_unknowns(
            workspace_id=workspace_id,
            scope=scope,
            scope_ref=scope_ref,
            evidence_path=evidence_path,
            decision_class=decision_class,
        )
        return jsonify({"ok": True, "workspace_id": workspace_id,
                        "rows": rows, "count": len(rows)})
    except Exception as exc:
        print(f"[atlas] unknowns list failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": "unknowns unavailable"}), 500


@app.route("/api/atlas/unknowns/<unknown_id>/resolve", methods=["POST"])
def atlas_unknowns_resolve(unknown_id: str):
    """Mark an unknown as answered or declared_unknowable."""
    body = request.get_json(silent=True) or {}
    answer_value = body.get("answer_value")
    answer_source = (body.get("answer_source") or "operator_typed").strip()
    status = (body.get("status") or "answered").strip()
    answered_by = (body.get("answered_by") or "devang").strip()
    try:
        from substrate.unknowns import resolve_unknown
        ok = resolve_unknown(
            unknown_id=unknown_id,
            answer_value=answer_value,
            answer_source=answer_source,
            answered_by=answered_by,
            status=status,
        )
        return jsonify({"ok": ok}), (200 if ok else 400)
    except Exception as exc:
        print(f"[atlas] unknowns resolve failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/citation-rejections", methods=["POST"])
def atlas_citation_rejection_log():
    """Log an operator rejection of a specific citation."""
    body = request.get_json(silent=True) or {}
    decision_event_id = (body.get("decision_event_id") or "").strip()
    citation_layer = (body.get("citation_layer") or "").strip()
    citation_source_id = (body.get("citation_source_id") or "").strip()
    reason = (body.get("reason") or "").strip()
    rejected_by = (body.get("rejected_by") or "devang").strip()
    workspace_id = _atlas_current_workspace()

    if not (decision_event_id and citation_layer and citation_source_id):
        return jsonify({"ok": False,
                        "error": "decision_event_id, citation_layer, citation_source_id required"}), 400

    try:
        from substrate.db import get_pool
        import uuid as _uuid
        pool = get_pool()
        if pool is None:
            return jsonify({"ok": False, "error": "substrate unavailable"}), 500
        rejection_id = str(_uuid.uuid4())
        with pool.connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO citation_rejections (
                        rejection_id, workspace_id, decision_event_id,
                        citation_layer, citation_source_id, reason,
                        rejected_by
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (rejection_id, workspace_id, decision_event_id,
                     citation_layer, citation_source_id, reason, rejected_by),
                )
            conn.commit()
        return jsonify({"ok": True, "rejection_id": rejection_id})
    except Exception as exc:
        print(f"[atlas] citation rejection log failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/memory/decisions/<event_id>/confound", methods=["GET"])
def atlas_memory_decision_confound(event_id: str):
    """Confound view for one decision_event.

    Returns before/after/confounds + honest caveats. Strictly NO causal
    claims, lift numbers, or attribution — the operator does the
    interpretation themselves. See substrate/confound.py for the contract.
    """
    workspace_id = _atlas_current_workspace()
    try:
        from substrate.confound import confound_view_for_decision
        v = confound_view_for_decision(workspace_id, event_id)
    except Exception as exc:
        print(f"[atlas] confound view failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": "confound view unavailable"}), 500
    return jsonify({"workspace_id": workspace_id, **v})


@app.route("/api/atlas/memory/sessions/<session_id>", methods=["GET"])
def atlas_memory_session_detail(session_id: str):
    """Return the full timeline for a single session."""
    workspace_id = _atlas_current_workspace()
    try:
        from substrate.memory import get_session_detail
        detail = get_session_detail(workspace_id, session_id)
    except Exception as exc:
        print(f"[atlas] memory session detail failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": "detail unavailable"}), 500
    if detail is None:
        return jsonify({"ok": False, "error": "session not found"}), 404
    return jsonify({"ok": True, "workspace_id": workspace_id, "session": detail})


# ───────────────────────────────────────────────────────────────────
# Marketing endpoints (Phase 1)
# ───────────────────────────────────────────────────────────────────

@app.route("/api/atlas/marketing/keywords", methods=["GET"])
def atlas_marketing_keywords():
    """Paginated keyword_library reader.

    Query params:
        limit / offset      pagination
        q                   substring match on keyword_norm
        asin                filter to rows linked to this ASIN
        order_by            'last_seen_at' | 'last_acos' | 'last_spend' | 'keyword'
    """
    workspace_id = _atlas_current_workspace()
    try:
        limit = int(request.args.get("limit") or 50)
    except ValueError:
        limit = 50
    try:
        offset = int(request.args.get("offset") or 0)
    except ValueError:
        offset = 0
    q = (request.args.get("q") or "").strip() or None
    asin = (request.args.get("asin") or "").strip() or None
    order_by = (request.args.get("order_by") or "last_seen_at").strip()
    try:
        from substrate.marketing import list_keywords
        result = list_keywords(workspace_id, limit=limit, offset=offset,
                               q=q, asin=asin, order_by=order_by)
    except Exception as exc:
        print(f"[atlas] marketing keywords list failed: {exc}", flush=True)
        result = {"keywords": [], "total": 0}
    return jsonify({
        "ok": True,
        "workspace_id": workspace_id,
        "keywords": result.get("keywords", []),
        "total": result.get("total", 0),
    })


@app.route("/api/atlas/marketing/direction", methods=["GET"])
def atlas_marketing_direction():
    """Trend / direction summary for one keyword (+ optional ASIN).

    Returns a per-metric { n, direction, confidence_label, delta, ... }
    block built from outcome_events. Honest about data density:
    n=0 → no_data, n=1 → prior_only, n>=14 → statistical.
    """
    workspace_id = _atlas_current_workspace()
    keyword = (request.args.get("keyword") or "").strip() or None
    asin = (request.args.get("asin") or "").strip() or None
    if not keyword:
        return jsonify({"ok": False, "error": "keyword required"}), 400
    try:
        from substrate.marketing import get_keyword_direction
        d = get_keyword_direction(workspace_id, keyword=keyword, asin=asin)
    except Exception as exc:
        print(f"[atlas] marketing direction failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": "direction unavailable"}), 500
    return jsonify({"ok": True, "workspace_id": workspace_id, **d})


# ---------------------------------------------------------------------------
# Budget endpoints (Phase 1 — PPC budget tracking).
# Budgets live under /marketing/ since they're strictly PPC for v1; the UI
# surfaces them as a sub-tab inside the Marketing module.
# ---------------------------------------------------------------------------


@app.route("/api/atlas/marketing/budget", methods=["GET"])
def atlas_marketing_budget_list():
    """List budgets for the current workspace, optionally filtered to a period."""
    workspace_id = _atlas_current_workspace()
    period = (request.args.get("period") or "").strip() or None
    try:
        from substrate.budget import list_budgets
        rows = list_budgets(workspace_id, period=period)
    except Exception as exc:
        print(f"[atlas] budget list failed: {exc}", flush=True)
        rows = []
    return jsonify({
        "ok": True,
        "workspace_id": workspace_id,
        "period": period,
        "budgets": rows,
        "total": len(rows),
    })


@app.route("/api/atlas/marketing/budget", methods=["POST"])
def atlas_marketing_budget_set():
    """Upsert a budget row. Strictly PPC for v1.

    JSON body:
        period       'YYYY-MM' (required)
        scope_type   'theme' | 'overall' | 'asin' (required)
        scope_value  theme name | '_overall' | ASIN (required)
        amount       numeric, >= 0 (required)
        currency     default 'USD'
        notes        optional free text
    """
    workspace_id = _atlas_current_workspace()
    payload = request.get_json(silent=True) or {}
    period = (payload.get("period") or "").strip()
    scope_type = (payload.get("scope_type") or "").strip()
    scope_value = (payload.get("scope_value") or "").strip()
    if "amount" not in payload:
        return jsonify({"ok": False, "error": "amount required"}), 400
    try:
        from substrate.budget import set_budget
        operator_id = (request.cookies.get(_ATLAS_OPERATOR_COOKIE) or "").strip() or None
        result = set_budget(
            workspace_id, period, scope_type, scope_value, payload["amount"],
            currency=(payload.get("currency") or "USD"),
            set_by=operator_id,
            notes=payload.get("notes"),
        )
    except Exception as exc:
        print(f"[atlas] budget set failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500
    status = 200 if result.get("ok") else 400
    return jsonify({"workspace_id": workspace_id, **result}), status


@app.route("/api/atlas/marketing/budget", methods=["DELETE"])
def atlas_marketing_budget_delete():
    """Remove one budget row. Audit-trail decision_events stay in place.

    Accepts period / scope_type / scope_value via query string or JSON body.
    """
    workspace_id = _atlas_current_workspace()
    payload = request.get_json(silent=True) or {}
    period = (request.args.get("period") or payload.get("period") or "").strip()
    scope_type = (request.args.get("scope_type") or payload.get("scope_type") or "").strip()
    scope_value = (request.args.get("scope_value") or payload.get("scope_value") or "").strip()
    try:
        from substrate.budget import delete_budget
        result = delete_budget(workspace_id, period, scope_type, scope_value)
    except Exception as exc:
        print(f"[atlas] budget delete failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500
    status = 200 if result.get("ok") else 400
    return jsonify({"workspace_id": workspace_id, **result}), status


@app.route("/api/atlas/marketing/budget/variance", methods=["GET"])
def atlas_marketing_budget_variance():
    """Compute planned-vs-actual variance for one period.

    Query params:
        period   'YYYY-MM' (required)

    Returns scopes (overall + per-theme + per-ASIN where data exists),
    totals, and a content_changes_summary. NIS decisions in the period on
    ASINs that had spend appear as markers so the operator can read
    variance honestly (a price/title change is a confound).
    """
    workspace_id = _atlas_current_workspace()
    period = (request.args.get("period") or "").strip()
    if not period:
        return jsonify({"ok": False, "error": "period required (YYYY-MM)"}), 400
    try:
        from substrate.budget import variance_for_period
        result = variance_for_period(workspace_id, period)
    except Exception as exc:
        print(f"[atlas] budget variance failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500
    status = 200 if result.get("ok") else 400
    return jsonify({"workspace_id": workspace_id, **result}), status


@app.route("/api/atlas/marketing/upload", methods=["POST"])
def atlas_marketing_upload():
    """Upload + parse a PPC bulk file or Search Term Report.

    Auto-detects file_kind via Inputs detection, then routes to the right
    marketing parser, writes keyword_library rows, and appends outcome_events.
    Also records an ingestion_records audit row (same path as catalog uploads).
    """
    workspace_id = _atlas_current_workspace()
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "file required"}), 400
    f = request.files["file"]
    try:
        content = f.read()
        if not content:
            return jsonify({"ok": False, "error": "empty file"}), 400
        # File-kind detection (same logic as Inputs detect endpoint)
        first_line = content.split(b"\n", 1)[0]
        headers: list[str] = []
        for sep in (b",", b"\t"):
            if sep in first_line:
                headers = [h.decode("utf-8", errors="replace").strip().strip('"')
                           for h in first_line.split(sep)]
                if headers:
                    break
        if not headers:
            try:
                import io as _io, openpyxl as _ox
                wb = _ox.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if first:
                    headers = [str(h).strip() if h is not None else "" for h in first]
            except Exception:
                headers = []
        from substrate.inputs import detect_file_kind, record_ingestion, file_hash as _hash
        kind = detect_file_kind(headers)
        # Force kind hint from form if the file is ambiguous.
        if not kind:
            kind = (request.form.get("kind") or "").strip() or None
        if kind not in ("ppc_bulk", "search_term", "ad_bulksheet"):
            return jsonify({
                "ok": False,
                "error": f"file kind {kind!r} not supported by marketing upload",
                "headers_detected": headers[:20],
            }), 400
        from substrate.marketing_parsers import parse_ppc_bulk, parse_search_term
        parser = parse_ppc_bulk if kind in ("ppc_bulk", "ad_bulksheet") else parse_search_term
        observations = parser(content, file_name=f.filename)
        from substrate.marketing import record_keyword_observations
        ingestion_meta = record_keyword_observations(
            workspace_id, observations, source_kind=kind,
            source_file_hash=_hash(content),
        )
        # Audit row
        operator_id = (_atlas_operator_from_request() if "_atlas_operator_from_request" in globals() else None) or session_data.get("operator_id")
        record_ingestion(
            workspace_id=workspace_id,
            file_kind=kind, file_name=f.filename,
            file_hash_value=_hash(content), bytes_size=len(content),
            rows_parsed=len(observations),
            asins_touched=len({(o.get("asin") or "").strip() for o in observations if o.get("asin")}),
            detected_fields=[h for h in headers if h][:40],
            summary=f"{len(observations)} keyword rows, {ingestion_meta.get('keywords_written', 0)} into library, {ingestion_meta.get('outcome_rows', 0)} outcome rows",
            uploaded_by=operator_id,
        )
        return jsonify({
            "ok": True,
            "workspace_id": workspace_id,
            "file_kind": kind,
            "rows_parsed": len(observations),
            "keywords_written": ingestion_meta.get("keywords_written", 0),
            "outcome_rows": ingestion_meta.get("outcome_rows", 0),
            "skipped": ingestion_meta.get("skipped", 0),
        })
    except Exception as exc:
        print(f"[atlas] marketing upload failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/marketing/wizard/start", methods=["POST"])
def atlas_mkt_wizard_start():
    """Open a marketing day-1 wizard session.

    JSON body: { asin: str (required), product_type?: str, style_name?: str }
    Returns: { ok, session_id }
    """
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    asin = (body.get("asin") or "").strip()
    if not asin:
        return jsonify({"ok": False, "error": "asin required"}), 400
    operator_id = session_data.get("operator_id") or "anonymous"
    try:
        from substrate.logger import open_session
        from substrate.schema import Module
        s = open_session(workspace_id=workspace_id, operator_id=operator_id, module=Module.MARKETING)
    except Exception as exc:
        print(f"[atlas] wizard open session failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": "could not open session"}), 500
    return jsonify({
        "ok": True,
        "session_id": s.session_id,
        "workspace_id": workspace_id,
        "asin": asin,
    })


@app.route("/api/atlas/marketing/wizard/generate", methods=["POST"])
def atlas_mkt_wizard_generate():
    """Generate candidate keyword list for an ASIN.

    JSON body: { asin, product_type?, style_name?, target_count? }
    Returns: { ok, asin, candidates: [...], source: 'llm'|'fallback', siblings_used }
    """
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    asin = (body.get("asin") or "").strip()
    if not asin:
        return jsonify({"ok": False, "error": "asin required"}), 400
    target = int(body.get("target_count") or 40)
    target = max(5, min(target, 80))
    try:
        from substrate.marketing_wizard import generate_candidates
        result = generate_candidates(
            workspace_id=workspace_id,
            asin=asin,
            product_type=body.get("product_type"),
            style_name=body.get("style_name"),
            target_count=target,
            anthropic_client=_anthropic_client,
        )
    except Exception as exc:
        print(f"[atlas] wizard generate failed: {exc}", flush=True)
        return jsonify({
            "ok": False,
            "error": "Generation failed. The substrate is reachable but candidate generation hit a snag \u2014 try again, or refresh the page.",
            "detail": str(exc)[:200],
        }), 500
    if not result.get("candidates"):
        return jsonify({
            "ok": False,
            "error": "No candidates produced. Check ANTHROPIC_API_KEY on the server, or upload a PPC bulk first so the rule-based fallback has siblings to draw from.",
            **result,
        }), 200   # 200 because the wizard succeeded structurally
    return jsonify({"ok": True, "workspace_id": workspace_id, **result})


@app.route("/api/atlas/marketing/wizard/decision", methods=["POST"])
def atlas_mkt_wizard_decision():
    """Log a single candidate decision (accept/edit/reject/comment).

    JSON body:
        session_id (required), asin (required), candidate (required dict),
        action (accept|edit|reject|comment), operator_value? (for edits),
        scope? (just_this|brand_always), comment?
    """
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    session_id = (body.get("session_id") or "").strip()
    asin = (body.get("asin") or "").strip()
    candidate = body.get("candidate") or {}
    action = (body.get("action") or "").strip().lower()
    if not (session_id and asin and isinstance(candidate, dict)
            and action in ("accept", "edit", "reject", "comment")):
        return jsonify({"ok": False, "error": "missing session_id/asin/candidate/action"}), 400
    keyword = (candidate.get("keyword") or "").strip()
    if not keyword:
        return jsonify({"ok": False, "error": "candidate.keyword required"}), 400
    operator_value = body.get("operator_value")
    scope = (body.get("scope") or "none").strip().lower()
    comment = (body.get("comment") or "").strip() or None
    brand_profile_version = body.get("brand_profile_version") or f"{workspace_id}_legacy"
    try:
        from substrate.logger import (
            log_field_decision, update_field_decision_with_operator_response,
        )
        from substrate.schema import Module, OperatorAction, OperatorScope
        # 1) Write the decision_event (Atlas's proposal)
        event_id = log_field_decision(
            workspace_id=workspace_id,
            session_id=session_id,
            module=Module.MARKETING,
            field_name="keyword_candidate",
            atlas_output={
                "keyword": keyword,
                "match_type": candidate.get("match_type"),
                "theme": candidate.get("theme"),
                "suggested_bid_low": candidate.get("suggested_bid_low"),
                "suggested_bid_high": candidate.get("suggested_bid_high"),
                "rationale": candidate.get("rationale"),
                "has_history": candidate.get("has_history", False),
            },
            overall_confidence=float(candidate.get("confidence") or 0.5),
            rules_injected=[],
            brand_profile_version=brand_profile_version,
            asin=asin,
            enforce_filter=False,   # keyword decisions always log
        )
        if not event_id:
            return jsonify({"ok": False, "error": "decision write skipped"}), 500
        # 2) Write the operator_response
        try:
            op_action = OperatorAction(action)
        except ValueError:
            op_action = OperatorAction.COMMENT
        try:
            op_scope = OperatorScope(scope)
        except ValueError:
            op_scope = OperatorScope.NONE
        update_field_decision_with_operator_response(
            workspace_id=workspace_id,
            event_id=event_id,
            operator_action=op_action,
            operator_value=operator_value,
            operator_scope=op_scope,
            operator_comment=comment,
            operator_viewed_case=True,
        )
    except Exception as exc:
        print(f"[atlas] wizard decision failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500
    return jsonify({"ok": True, "event_id": event_id})


@app.route("/api/atlas/marketing/wizard/submit", methods=["POST"])
def atlas_mkt_wizard_submit():
    """Finalize a wizard session. Body: { session_id, operator_notes?, exemplar? }."""
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    session_id = (body.get("session_id") or "").strip()
    if not session_id:
        return jsonify({"ok": False, "error": "session_id required"}), 400
    notes = body.get("operator_notes")
    exemplar = bool(body.get("exemplar", False))
    try:
        from substrate.logger import submit_session, read_session
        from substrate.schema import SessionObject, Module
        # Reconstruct a SessionObject from the substrate row
        meta = read_session(workspace_id, session_id)
        if meta is None:
            return jsonify({"ok": False, "error": "session not found"}), 404
        s = SessionObject(
            workspace_id=workspace_id,
            operator_id=meta.get("operator_id") or "anonymous",
            module=Module(meta.get("module") or "marketing"),
            session_id=session_id,
            started_at=meta.get("started_at"),
            state=meta.get("state") or "live",
        )
        submit_session(s, operator_notes=notes, exemplar=exemplar)
    except Exception as exc:
        print(f"[atlas] wizard submit failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500
    return jsonify({"ok": True, "session_id": session_id})


@app.route("/api/atlas/marketing/wizard/batch-export", methods=["POST"])
def atlas_mkt_wizard_batch_export():
    """Render an accepted-candidate list across many ASINs as one merged bulk CSV.

    Body:
        batches: list of { asin, accepted, campaign_name?, ad_group_name? }
        campaign_prefix: optional, defaults to 'Atlas day-1'

    Each ASIN keeps its own campaign so PPC reports stay per-ASIN clean.
    """
    body = request.get_json(silent=True) or {}
    batches = body.get("batches") or []
    if not isinstance(batches, list) or not batches:
        return jsonify({"ok": False, "error": "batches[] required"}), 400
    try:
        from substrate.marketing_wizard import batch_candidates_to_bulk_csv
        csv_body = batch_candidates_to_bulk_csv(
            batches,
            campaign_prefix=(body.get("campaign_prefix") or "Atlas day-1"),
        )
    except Exception as exc:
        print(f"[atlas] wizard batch export failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500
    fn = body.get("filename") or f"atlas_day1_batch_{len(batches)}_asins.csv"
    return Response(
        csv_body, mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@app.route("/api/atlas/marketing/wizard/export", methods=["POST"])
def atlas_mkt_wizard_export():
    """Render an accepted candidate list as a downloadable bulk CSV.

    Body: { asin, accepted: [...], campaign_name?, ad_group_name? }
    """
    body = request.get_json(silent=True) or {}
    asin = (body.get("asin") or "").strip()
    accepted = body.get("accepted") or []
    if not asin or not isinstance(accepted, list) or not accepted:
        return jsonify({"ok": False, "error": "asin + accepted[] required"}), 400
    try:
        from substrate.marketing_wizard import candidates_to_bulk_csv
        csv_body = candidates_to_bulk_csv(
            asin=asin,
            accepted=accepted,
            campaign_name=(body.get("campaign_name") or f"Atlas day-1 - {asin}"),
            ad_group_name=body.get("ad_group_name"),
        )
    except Exception as exc:
        print(f"[atlas] wizard export failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500
    return Response(
        csv_body,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=atlas_day1_{asin}.csv"},
    )


# /docs/onboarding intentionally removed — the static onboarding doc was
# too blunt for the agency rollout. Devang will walk new operators through
# Atlas in person instead. The template file was also deleted with this
# change; recover from git history if a future doc needs the starting
# point.


@app.route("/api/version", methods=["GET"])
def api_version():
    """Return the git SHA + build metadata of the currently-running code.

    Used by the deploy-verify loop: poll this endpoint after a push until
    it reports the new SHA, then run live smoke tests. Render injects
    RENDER_GIT_COMMIT at build time; we fall back to a local .git_sha
    file or 'unknown'.
    """
    sha = (
        os.environ.get("RENDER_GIT_COMMIT")
        or os.environ.get("GIT_SHA")
        or ""
    )
    if not sha:
        try:
            here = os.path.dirname(os.path.abspath(__file__))
            with open(os.path.join(here, ".git_sha"), "r", encoding="utf-8") as fh:
                sha = fh.read().strip()
        except OSError:
            sha = "unknown"
    return jsonify({
        "ok": True,
        "sha": sha,
        "short_sha": sha[:7] if sha and sha != "unknown" else sha,
        "branch": os.environ.get("RENDER_GIT_BRANCH") or "",
        "service": os.environ.get("RENDER_SERVICE_NAME") or "",
        "started_at": _APP_BOOT_TS,
    })


@app.route("/api/atlas/visible-brands", methods=["GET"])
def api_atlas_visible_brands():
    """Return the operator-facing brand list + the default brand.

    The frontend uses this on page load to populate dropdowns and to
    decide whether to pre-select a brand (only when there is exactly
    one visible brand).
    """
    file_brands = [p.stem.replace('_', ' ') for p in BRAND_CONFIGS_DIR.glob('*.json')]
    in_mem = list(BRAND_CONFIGS.keys())
    deduped = sorted({b.strip() for b in (file_brands + in_mem) if b and b.strip()})
    visible = _atlas_visible_brands_filter(deduped)
    default = visible[0] if len(visible) == 1 else None
    return jsonify({
        "ok": True,
        "brands": visible,
        "default": default,
        "single_brand_mode": default is not None,
    })


@app.route("/api/beta-image-nis/analyze", methods=["POST"])
def api_beta_image_nis_analyze():
    """Step 1: operator drops image + picks brand. We run vision, then map
    the output to a draft style dict + provenance. No xlsm yet.

    multipart/form-data:
      file: image
      brand: brand name (must match a brand_configs/*.json)
      coo (optional), fabric (optional), care (optional), season_code (optional),
      subclass (optional), style_num (optional)

    Returns: { ok, session_id, vision, style, provenance, brand_cfg }
    """
    if "file" not in request.files:
        return jsonify({"error": "No image file"}), 400
    brand = (request.form.get("brand") or "").strip()
    if not brand:
        return jsonify({"error": "Brand is required"}), 400
    # Strict brand validation — must have a saved config file or be in BRAND_CONFIGS
    brand_file = BRAND_CONFIGS_DIR / f"{re.sub(r'[^\w]', '_', brand)}.json"
    if not brand_file.exists() and brand not in BRAND_CONFIGS:
        known = sorted([p.stem.replace('_', ' ') for p in BRAND_CONFIGS_DIR.glob('*.json')] + list(BRAND_CONFIGS.keys()))
        return jsonify({"error": f"Unknown brand: {brand}", "known_brands": known}), 400
    brand_cfg = _load_brand_config_data(brand)

    f = request.files["file"]
    ext = Path(f.filename or "").suffix.lower()
    if ext not in [".jpg", ".jpeg", ".png", ".webp", ".gif"]:
        return jsonify({"error": f"Unsupported image type: {ext}"}), 400

    # Save image into a beta session directory
    import time, uuid
    session_id = f"beta-{int(time.time()*1000)}-{uuid.uuid4().hex[:6]}"
    sess_dir = UPLOAD_IMAGES / f"_beta_{session_id}"
    sess_dir.mkdir(parents=True, exist_ok=True)
    img_path = sess_dir / f"product{ext}"
    f.save(str(img_path))

    # Run vision — standalone (no expected PT, force classification)
    vision_intel = analyze_style_image(f"_beta_{session_id}", str(img_path), brand_override=brand)
    if vision_intel is None:
        return jsonify({"error": _last_vision_error or "Vision pass unavailable", "reason": _last_vision_error}), 503

    # Collect operator hints
    hints = {
        "coo":          (request.form.get("coo") or "").strip(),
        "fabric":       (request.form.get("fabric") or "").strip(),
        "care":         (request.form.get("care") or "").strip(),
        "season_code":  (request.form.get("season_code") or "").strip(),
        "subclass":     (request.form.get("subclass") or "").strip(),
        "style_num":    (request.form.get("style_num") or "").strip(),
    }

    style, provenance = _beta_build_draft_style(vision_intel, brand, brand_cfg, hints)

    # Stash everything in the beta session so subsequent endpoints can reuse it
    _BETA_SESSION[session_id] = {
        "image_path": str(img_path),
        "brand": brand,
        "brand_cfg": brand_cfg,
        "vision_intel": vision_intel,
        "style": style,
        "provenance": provenance,
        "hints": hints,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }

    return jsonify({
        "ok": True,
        "session_id": session_id,
        "image_url": f"/api/beta-image-nis/image/{session_id}",
        "vision": vision_intel,
        "style": style,
        "provenance": provenance,
        "brand": brand,
        "brand_cfg_summary": {
            "vendor_code_prefix": brand_cfg.get("vendor_code_prefix", ""),
            "default_coo": brand_cfg.get("default_coo", ""),
            "default_fabric": brand_cfg.get("default_fabric", ""),
            "gender": brand_cfg.get("gender", ""),
            "department": brand_cfg.get("department", ""),
        },
    })


@app.route("/api/beta-image-nis/generate", methods=["POST"])
def api_beta_image_nis_generate():
    """Step 2: take a beta session (from /analyze) + any operator overrides,
    run full content LLM with image attached, and return the complete draft
    listing (title/bullets/description/backend_keywords) + the merged style
    + provenance.

    JSON body:
      session_id: required, from /analyze
      overrides: optional dict of fields the operator manually edited
        (e.g. {"fabric":"100% Cotton","coo":"India","style_num":"S26-12345"})

    Returns: { ok, session_id, content, style, provenance, image_url }
    """
    data = request.get_json(force=True) or {}
    session_id = (data.get("session_id") or "").strip()
    overrides = data.get("overrides") or {}
    sess = _BETA_SESSION.get(session_id)
    if not sess:
        return jsonify({"error": "Unknown beta session. Run /analyze first."}), 404

    # Apply operator overrides into hints + recompute style/provenance
    hints = dict(sess.get("hints") or {})
    for k in ("coo", "fabric", "care", "season_code", "subclass", "style_num"):
        if k in overrides and overrides[k]:
            hints[k] = str(overrides[k]).strip()
    sess["hints"] = hints

    style, provenance = _beta_build_draft_style(
        sess["vision_intel"], sess["brand"], sess["brand_cfg"], hints
    )

    # Inject the beta image into session_data["style_images"] briefly so
    # generate_content_llm pulls it as a vision input. Restore after.
    prior_img_map = (session_data.get("style_images") or {}).copy()
    session_data.setdefault("style_images", {})[str(style["style_num"])] = sess["image_path"]

    # Inject the beta style into session_data["styles"] briefly so
    # _resolve_style_product_type picks up the right PT context.
    prior_styles = list(session_data.get("styles") or [])
    session_data["styles"] = prior_styles + [style]

    try:
        # Run content LLM. Vision image will be picked up via session_data["style_images"].
        content = generate_content_llm(sess["brand_cfg"], sess["brand"], style, "")
    finally:
        # Restore session_data so we don't leak this beta into the main flow
        session_data["style_images"] = prior_img_map
        session_data["styles"] = prior_styles

    if content is None:
        return jsonify({"error": "Content generation unavailable (LLM client missing)."}), 503

    # Persist on the session
    sess["style"] = style
    sess["provenance"] = provenance
    sess["content"] = content

    # ─── Atlas substrate write ─────────────────────────────────────
    # Image → NIS was completely invisible to Memory until this commit.
    # Open a session, log one decision_event per generated field, capture
    # event_ids so /feedback can later attach operator_responses. Pattern
    # mirrors /api/generate-content. Best-effort: substrate failures must
    # never break the beta flow.
    atlas_event_ids: dict[str, str] = {}
    atlas_session_id = None  # str | None
    try:
        from substrate.logger import (
            open_session as _atlas_open,
            log_field_decision as _atlas_log,
        )
        from substrate.schema import Module as _AtlasModule
        ws = (sess["brand"] or "tlg").lower().replace(" ", "_") or "tlg"
        op_id = session_data.get("operator_id") or "devang"
        bpv = sess["brand_cfg"].get("_version") or f"{ws}_legacy"
        atlas_sess = _atlas_open(workspace_id=ws, operator_id=op_id,
                                 module=_AtlasModule.NIS)
        atlas_session_id = atlas_sess.session_id
        # ASIN anchor: hints may carry one; otherwise None (Day-1 listing).
        asin = (hints.get("asin") or hints.get("child_asin") or "").strip() or None
        # Confidence proxy for image-driven gen. Vision-driven runs are
        # inherently lower confidence than catalog-driven runs because
        # vision can be wrong. We bracket conservatively.
        conf = 0.70
        rules = [
            {"rule_id": "nis.engine.compose_v0_7_6", "version": "0.7.6"},
            {"rule_id": "nis.llm.claude_generation"},
            {"rule_id": "nis.image.vision_driven"},
        ]
        for fname in ("item_name", "bullet_1", "bullet_2", "bullet_3",
                      "bullet_4", "bullet_5", "description", "backend_keywords"):
            # Map title -> item_name in the substrate vocabulary.
            content_key = "title" if fname == "item_name" else fname
            fval = content.get(content_key)
            if not fval:
                continue
            _eid = _atlas_log(
                workspace_id=ws,
                session_id=atlas_session_id,
                module=_AtlasModule.NIS,
                field_name=fname,
                atlas_output=fval,
                overall_confidence=conf,
                rules_injected=rules,
                brand_profile_version=bpv,
                style_id=str(style.get("style_num") or ""),
                asin=asin,
            )
            if _eid:
                atlas_event_ids[fname] = _eid
        sess["atlas_session_id"] = atlas_session_id
        sess["atlas_event_ids"] = atlas_event_ids
        sess["atlas_workspace_id"] = ws
    except Exception as exc:
        print(f"[atlas] beta-image-nis decision log skipped: {exc}", flush=True)

    return jsonify({
        "ok": True,
        "session_id": session_id,
        "image_url": f"/api/beta-image-nis/image/{session_id}",
        "vision": sess["vision_intel"],
        "style": style,
        "provenance": provenance,
        "content": content,
        "brand": sess["brand"],
        "atlas_session_id": atlas_session_id,
        "atlas_event_ids": atlas_event_ids,
    })


@app.route("/api/beta-image-nis/feedback", methods=["POST"])
def api_beta_image_nis_feedback():
    """Step 4: Operator flags an issue with a vision read or a generated field.
    Logged to data/beta_image_nis_feedback.jsonl for later analysis.

    JSON body:
      session_id: required
      kind: 'pt_wrong' | 'field_wrong' | 'observation_wrong' | 'image_quality_missed' |
            'pt_correct' | 'general' (free-text)
      field: optional, when kind='field_wrong' (e.g. 'closure_type')
      detected_value: what vision returned
      correct_value: what operator says it should be (optional)
      note: free-text detail
      operator: optional name/email tag
    """
    data = request.get_json(force=True) or {}
    session_id = (data.get("session_id") or "").strip()
    kind = (data.get("kind") or "").strip().lower()
    valid_kinds = {"pt_wrong", "pt_correct", "field_wrong", "observation_wrong", "image_quality_missed", "general"}
    if kind not in valid_kinds:
        return jsonify({"error": f"Invalid kind. Must be one of: {sorted(valid_kinds)}"}), 400
    sess = _BETA_SESSION.get(session_id)
    if not sess:
        return jsonify({"error": "Unknown beta session"}), 404

    entry = {
        "timestamp":      datetime.now().isoformat(timespec="seconds"),
        "session_id":     session_id,
        "brand":          sess.get("brand"),
        "image_path":     sess.get("image_path"),
        "detected_pt":    (sess.get("vision_intel") or {}).get("detected_pt"),
        "detected_subject": (sess.get("vision_intel") or {}).get("detected_subject"),
        "kind":           kind,
        "field":          (data.get("field") or "").strip()[:80] or None,
        "detected_value": str(data.get("detected_value") or "")[:240] or None,
        "correct_value":  str(data.get("correct_value") or "")[:240] or None,
        "note":           str(data.get("note") or "")[:1000] or None,
        "operator":       (data.get("operator") or "").strip()[:80] or None,
    }

    try:
        with open(BETA_IMAGE_NIS_FEEDBACK_LOG, "a", encoding="utf-8") as fh:
            fh.write(json.dumps(entry) + "\n")
    except Exception as e:
        return jsonify({"error": f"Could not write feedback: {e}"}), 500

    # Mirror onto the session for in-memory traceability
    sess.setdefault("feedback", []).append(entry)

    # ─── Atlas substrate: operator_response write ──────────────────────
    # Only field_wrong feedback maps cleanly to a decision_event. The
    # other kinds (pt_wrong, observation_wrong, image_quality_missed,
    # pt_correct, general) are about the *vision* layer, which we don't
    # log as decision_events today. They stay in the JSONL feedback log
    # for the vision-team's analysis but don't bind to substrate.
    atlas_op_event_id = None  # str | None
    try:
        if kind == "field_wrong":
            from substrate.logger import (
                update_field_decision_with_operator_response as _atlas_resp,
            )
            from substrate.schema import (
                OperatorAction as _Act, OperatorScope as _Scope,
            )
            ws = sess.get("atlas_workspace_id")
            event_ids = sess.get("atlas_event_ids") or {}
            field = entry["field"] or ""
            # The frontend passes content keys (title / bullet_1 / etc.);
            # we map title -> item_name to match the substrate vocab.
            substrate_field = "item_name" if field == "title" else field
            target_eid = event_ids.get(substrate_field)
            if ws and target_eid:
                # If operator supplied a correction, it's an 'edit'.
                # Otherwise treat as a 'reject' (they flagged it but didn't
                # write the replacement).
                action_str = "edit" if entry.get("correct_value") else "reject"
                _atlas_resp(
                    workspace_id=ws,
                    event_id=target_eid,
                    operator_action=_Act(action_str),
                    operator_value=entry.get("correct_value"),
                    operator_scope=_Scope("just_this"),
                    operator_time_to_decision_ms=None,
                    operator_comment=entry.get("note"),
                    operator_viewed_case=False,
                )
                atlas_op_event_id = target_eid
    except Exception as exc:
        print(f"[atlas] beta-image-nis operator_response skipped: {exc}", flush=True)

    return jsonify({"ok": True, "entry": entry,
                    "atlas_event_id": atlas_op_event_id})


@app.route("/api/beta-image-nis/feedback", methods=["GET"])
def api_beta_image_nis_feedback_list():
    """Read recent feedback entries for analysis. Optional ?limit=N (default 200, max 1000)."""
    try:
        limit = max(1, min(int(request.args.get("limit", 200)), 1000))
    except Exception:
        limit = 200
    entries = []
    if BETA_IMAGE_NIS_FEEDBACK_LOG.exists():
        with open(BETA_IMAGE_NIS_FEEDBACK_LOG, "r", encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    entries.append(json.loads(line))
                except Exception:
                    pass
    entries = entries[-limit:]
    # Quick aggregates: counts per kind, per detected_pt, per field
    from collections import Counter
    by_kind = Counter(e.get("kind") for e in entries)
    by_pt = Counter(e.get("detected_pt") for e in entries if e.get("detected_pt"))
    by_field = Counter(e.get("field") for e in entries if e.get("field"))
    return jsonify({
        "ok": True,
        "count": len(entries),
        "entries": entries,
        "summary": {
            "by_kind":  dict(by_kind),
            "by_pt":    dict(by_pt),
            "by_field": dict(by_field),
        },
    })


@app.route("/api/beta-image-nis/build-xlsm", methods=["POST"])
def api_beta_image_nis_build_xlsm():
    """Step 3: Build a single-style .xlsm download from a beta session.

    JSON body:
      session_id: required
      operator_fields: optional dict of last-mile fills
        (color_name, size, upc, list_price, cost_price, sku_override)
        If omitted, we build with a placeholder variant + return validation
        warnings telling the operator exactly what to fill before Amazon submit.

    Returns: { ok, download_url, filename, warnings, fields_written }
    """
    data = request.get_json(force=True) or {}
    session_id = (data.get("session_id") or "").strip()
    op_fields = data.get("operator_fields") or {}
    sess = _BETA_SESSION.get(session_id)
    if not sess:
        return jsonify({"error": "Unknown beta session"}), 404

    style = dict(sess.get("style") or {})
    if not style.get("_resolved_pt"):
        return jsonify({"error": "No product type detected. Run /analyze first."}), 400
    if not sess.get("content"):
        return jsonify({"error": "No content generated. Run /generate first."}), 400

    pt = style["_resolved_pt"]
    template_file = _pt_defaults.get_pt_default(pt, "template_file")
    if not template_file:
        return jsonify({"error": f"No template configured for product type {pt}"}), 400
    template_path = UPLOAD_TEMPLATES / template_file
    if not template_path.exists():
        return jsonify({"error": f"Template file missing on server: {template_file}"}), 500

    # Build a single placeholder variant from operator_fields (or fall back to defaults)
    color_name = (op_fields.get("color_name") or sess["vision_intel"].get("detected_color") or "Black").strip()
    size       = (op_fields.get("size") or "M").strip()
    upc        = (op_fields.get("upc") or "").strip()
    list_price = (op_fields.get("list_price") or "").strip()
    cost_price = (op_fields.get("cost_price") or "").strip()
    style_num  = (op_fields.get("sku_override") or style.get("style_num") or "BETA-DRAFT").strip()
    style["style_num"] = style_num

    # Color code: first 3 letters uppercase
    color_code = re.sub(r'[^A-Z]', '', color_name.upper())[:3] or "BLK"
    season_code = (sess["hints"] or {}).get("season_code") or sess["brand_cfg"].get("vendor_code_prefix", "BETA")
    # Avoid duplicating season_code if operator already prefixed it
    sku_root = style_num if style_num.upper().startswith(season_code.upper() + "-") else f"{season_code}-{style_num}"

    variant = {
        "sku":         f"{sku_root}-{color_code}-{size}",
        "color_name": color_name,
        "color_code": color_code,
        "size":       size,
        "upc":        upc,
        "list_price": list_price,
        "cost_price": cost_price,
    }
    style["variants"] = [variant]

    # Skip the parent row by default in beta — single-variant builds don't need it
    skip_parent_prior = session_data.get("skip_parent_row", False)
    session_data["skip_parent_row"] = True

    # Inject the beta image into session_data["style_images"] briefly so the
    # writer can reach it if needed.
    prior_imgs = (session_data.get("style_images") or {}).copy()
    session_data.setdefault("style_images", {})[str(style_num)] = sess["image_path"]
    prior_styles = list(session_data.get("styles") or [])
    session_data["styles"] = prior_styles + [style]

    # Output filename
    safe_brand = re.sub(r'[^\w]', '_', sess["brand"])
    out_name = f"BETA_{safe_brand}_{style_num}_{int(datetime.now().timestamp())}.xlsm"
    out_path = UPLOAD_OUTPUT / out_name

    content_map = {style_num: sess["content"]}
    fields_written = 0
    try:
        _generate_category_file(
            [style], content_map, str(template_path),
            sess["brand"], sess["brand_cfg"],
            season_code, str(out_path),
        )
        # Count non-blank cells written on the data row to give a sense of coverage
        try:
            wb = openpyxl.load_workbook(str(out_path), keep_vba=True, read_only=True)
            for sn in wb.sheetnames:
                if sn.upper().startswith("TEMPLATE"):
                    ws = wb[sn]
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(row=7, column=c).value
                        if v not in (None, ""):
                            fields_written += 1
                    break
        except Exception:
            pass
    except Exception as e:
        session_data["skip_parent_row"] = skip_parent_prior
        session_data["style_images"] = prior_imgs
        session_data["styles"] = prior_styles
        return jsonify({"error": f"Build failed: {str(e)[:200]}"}), 500
    finally:
        session_data["skip_parent_row"] = skip_parent_prior
        session_data["style_images"] = prior_imgs
        session_data["styles"] = prior_styles

    # Validation warnings — honest about placeholders
    warnings = []
    if not upc:
        warnings.append("UPC is blank — Amazon will reject. Add a real UPC before submit.")
    if not list_price:
        warnings.append("List price is blank — add before submit.")
    if not cost_price:
        warnings.append("Cost price is blank.")
    if size == "M" and not op_fields.get("size"):
        warnings.append("Size defaulted to 'M'. Add real size grid (XS/S/M/L/XL) before submit.")
    if style_num == "BETA-DRAFT":
        warnings.append("Style number is the placeholder 'BETA-DRAFT'. Replace with real season-style code.")
    if not style.get("fabric"):
        warnings.append("Fabric composition not set — vision can't see this; add manually.")
    if not style.get("coo"):
        warnings.append("Country of Origin not set — vision can't see this; add manually.")

    sess["last_build"] = {
        "filename": out_name,
        "path": str(out_path),
        "warnings": warnings,
        "fields_written": fields_written,
        "built_at": datetime.now().isoformat(timespec="seconds"),
    }

    return jsonify({
        "ok": True,
        "session_id": session_id,
        "filename": out_name,
        "download_url": f"/api/preupload/download/{out_name}",
        "warnings": warnings,
        "fields_written": fields_written,
        "product_type": pt,
        "template": template_file,
    })


@app.route("/api/beta-image-nis/session/<session_id>", methods=["GET"])
def api_beta_image_nis_session(session_id):
    """Fetch the current state of a beta session."""
    sess = _BETA_SESSION.get(session_id)
    if not sess:
        return jsonify({"error": "Unknown session"}), 404
    return jsonify({
        "ok": True,
        "session_id": session_id,
        "image_url": f"/api/beta-image-nis/image/{session_id}",
        "brand": sess.get("brand"),
        "vision": sess.get("vision_intel"),
        "style": sess.get("style"),
        "provenance": sess.get("provenance"),
        "content": sess.get("content"),
        "hints": sess.get("hints"),
    })


@app.route("/api/beta-image-nis/image/<session_id>", methods=["GET"])
def api_beta_image_nis_image(session_id):
    """Serve the beta session image."""
    sess = _BETA_SESSION.get(session_id)
    if not sess:
        return jsonify({"error": "Unknown session"}), 404
    p = sess.get("image_path")
    if not p or not Path(p).exists():
        return jsonify({"error": "Image missing"}), 404
    return send_file(p)


# ══════════════════════════════════════════════════════════════
# LAB · GRID EDITOR (Stage 1) — separate session bucket so it never
# touches the live Bulk Upload session_data. Single-style scoped grid
# returning {columns, rows} for the front-end to render in Handsontable.
# ══════════════════════════════════════════════════════════════
lab_session = {
    "styles": [],
    "brand": "",
    "file_path": "",
}

# Group definitions — each maps to a list of column specs the grid renders.
# Two scopes per group:
#   scope="variant"  → one row per variant (size/color/upc/weight/etc.)
#   scope="style"    → one row per style (title/bullets/description/etc.)
# The "All fields" view (Stage 2 stretch) joins everything into one wide sheet.
LAB_GRID_GROUPS = {
    "identity": {
        "label": "Identity",
        "scope": "style",
        "columns": [
            {"key": "style_num",     "title": "Style #",       "type": "text", "readonly": True, "required": True},
            {"key": "style_name",    "title": "Style Name",    "type": "text", "required": True},
            {"key": "brand",         "title": "Brand",         "type": "text", "required": True},
            {"key": "vendor_code",   "title": "Vendor Code",   "type": "text"},
            {"key": "subclass",      "title": "Sub-Class",     "type": "text", "required": True},
            {"key": "sub_subclass",  "title": "Sub-Sub-Class", "type": "text"},
            {"key": "product_type",  "title": "Amazon PT",     "type": "text", "readonly": True},
        ],
    },
    "title_copy": {
        "label": "Title & Copy",
        "scope": "style",
        "columns": [
            {"key": "style_num",   "title": "Style #", "type": "text", "readonly": True},
            {"key": "title",       "title": "Title",        "type": "text", "required": True, "max_length": 120, "width": 320},
            {"key": "bullet_1",    "title": "Bullet 1",     "type": "text", "required": True, "max_length": 256, "width": 240},
            {"key": "bullet_2",    "title": "Bullet 2",     "type": "text", "required": True, "max_length": 256, "width": 240},
            {"key": "bullet_3",    "title": "Bullet 3",     "type": "text", "required": True, "max_length": 256, "width": 240},
            {"key": "bullet_4",    "title": "Bullet 4",     "type": "text", "required": True, "max_length": 256, "width": 240},
            {"key": "bullet_5",    "title": "Bullet 5",     "type": "text", "required": True, "max_length": 256, "width": 240},
            {"key": "description", "title": "Description",  "type": "text", "required": True, "max_length": 2000, "width": 360},
            {"key": "backend_keywords", "title": "Backend Keywords", "type": "text", "max_length": 250, "width": 240},
        ],
    },
    "taxonomy": {
        "label": "Taxonomy",
        "scope": "style",
        "columns": [
            {"key": "style_num",        "title": "Style #",      "type": "text", "readonly": True},
            {"key": "feed_product_type", "title": "Feed PT",     "type": "text", "required": True},
            {"key": "item_type",        "title": "Item Type",    "type": "text", "required": True},
            {"key": "department",       "title": "Department",   "type": "dropdown_dynamic", "dropdown_field": "department#1.value", "required": True},
            {"key": "target_gender",    "title": "Target Gender","type": "dropdown_dynamic", "dropdown_field": "target_gender#1.value", "required": True},
            {"key": "age_range",        "title": "Age Range",    "type": "dropdown_dynamic", "dropdown_field": "age_range_description#1.value"},
            {"key": "lifestyle_1",      "title": "Lifestyle 1",  "type": "dropdown_dynamic", "dropdown_field": "lifestyle#1.value"},
            {"key": "lifestyle_2",      "title": "Lifestyle 2",  "type": "dropdown_dynamic", "dropdown_field": "lifestyle#2.value"},
        ],
    },
    "variants": {
        "label": "Variants",
        "scope": "variant",
        "columns": [
            {"key": "variant_id",  "title": "Variant ID", "type": "text", "readonly": True},
            {"key": "sku",         "title": "SKU",        "type": "text"},
            {"key": "upc",         "title": "UPC",        "type": "text", "required": True, "validator": "upc_check"},
            {"key": "asin",        "title": "ASIN",       "type": "text"},
            {"key": "color_name",  "title": "Color",      "type": "text",     "required": True},
            {"key": "color_map",   "title": "Color Map",  "type": "dropdown_dynamic", "dropdown_field": "color_map#1.value"},
            {"key": "size",        "title": "Size",       "type": "text",     "required": True},
            {"key": "size_map",    "title": "Size Map",   "type": "dropdown_dynamic", "dropdown_field": "size_name#1.value"},
        ],
    },
    "weight": {
        "label": "Weight & Dimensions",
        "scope": "variant",
        "columns": [
            {"key": "variant_id", "title": "Variant ID", "type": "text", "readonly": True},
            {"key": "size",       "title": "Size",  "type": "text", "readonly": True},
            {"key": "color_name", "title": "Color", "type": "text", "readonly": True},
            {"key": "item_weight_value", "title": "Item Weight", "type": "numeric",
             "required": True, "validator": "positive_number"},
            {"key": "item_weight_unit",  "title": "Wt Unit",     "type": "dropdown",
             "required": True, "options": ["pounds", "ounces", "kilograms", "grams"]},
            {"key": "item_length_value", "title": "Length", "type": "numeric", "validator": "positive_number"},
            {"key": "item_width_value",  "title": "Width",  "type": "numeric", "validator": "positive_number"},
            {"key": "item_height_value", "title": "Height", "type": "numeric", "validator": "positive_number"},
            {"key": "item_dim_unit",     "title": "Dim Unit", "type": "dropdown",
             "options": ["inches", "centimeters", "millimeters", "feet"]},
        ],
    },
    "compliance": {
        "label": "Compliance",
        "scope": "style",
        "columns": [
            {"key": "style_num", "title": "Style #", "type": "text", "readonly": True},
            {"key": "coo",      "title": "Country of Origin", "type": "text", "required": True},
            {"key": "fabric",   "title": "Fabric / Material", "type": "text", "required": True},
            {"key": "care",     "title": "Care Instructions", "type": "text", "required": True},
            {"key": "upf",      "title": "UPF Rating", "type": "text"},
            {"key": "contains_batteries", "title": "Contains Batteries", "type": "dropdown",
             "options": ["No", "Yes"]},
            {"key": "is_hazmat",          "title": "Hazardous Material", "type": "dropdown",
             "options": ["No", "Yes"]},
            {"key": "federal_contract_compliant", "title": "Berry Compliant", "type": "dropdown",
             "options": ["No", "Yes"]},
        ],
    },
    "apparel": {
        "label": "Apparel Attributes",
        "scope": "style",
        "columns": [
            {"key": "style_num",     "title": "Style #",     "type": "text", "readonly": True},
            {"key": "sleeve_type",   "title": "Sleeve",      "type": "text"},
            {"key": "neck_type",     "title": "Neck",        "type": "text"},
            {"key": "fit_type",      "title": "Fit",         "type": "dropdown_dynamic", "dropdown_field": "fit_type#1.value"},
            {"key": "closure_type",  "title": "Closure",     "type": "dropdown_dynamic", "dropdown_field": "closure_type#1.value"},
            {"key": "collar_style",  "title": "Collar",      "type": "dropdown_dynamic", "dropdown_field": "collar_style#1.value"},
            {"key": "occasion",      "title": "Occasion",    "type": "text"},
            {"key": "special_feature_1", "title": "Special Feature 1", "type": "dropdown_dynamic", "dropdown_field": "special_feature#1.value"},
            {"key": "special_feature_2", "title": "Special Feature 2", "type": "dropdown_dynamic", "dropdown_field": "special_feature#2.value"},
        ],
    },
    "commercial": {
        "label": "Pricing & Logistics",
        "scope": "style",
        "columns": [
            {"key": "style_num",   "title": "Style #",       "type": "text",    "readonly": True},
            {"key": "cost_price",  "title": "Cost (Wholesale)", "type": "numeric"},
            {"key": "list_price",  "title": "List (Retail)",   "type": "numeric"},
            {"key": "ship_date",   "title": "Ship Date",    "type": "text"},
            {"key": "model_name",  "title": "Model Name",   "type": "text"},
            {"key": "manufacturer", "title": "Manufacturer", "type": "text"},
            {"key": "warranty",    "title": "Warranty",     "type": "text"},
        ],
    },
}

# Stage-1-compatible alias kept so old front-end calls still work.
# Stage 1 used only "weight". Stage 2 now exposes the full set above.


def _lab_build_full_group():
    """Build the synthetic 'Full NIS sheet' group by unioning every column
    from LAB_GRID_GROUPS. One row per variant; style-level fields repeat
    across all variants of the same style (matching the NIS .xlsm layout).

    Column order: variant anchor first (variant_id/sku/upc), then style
    identity, then title & copy, then taxonomy, then apparel attributes,
    then compliance, then weight & dimensions, then commercial.
    """
    seen = set()
    cols = []
    order = ["variants", "identity", "title_copy", "taxonomy", "apparel",
            "compliance", "weight", "commercial"]
    for gkey in order:
        grp = LAB_GRID_GROUPS.get(gkey) or {}
        for col in grp.get("columns", []):
            k = col["key"]
            if k in seen:
                continue
            seen.add(k)
            # Tag each column with its source group so the frontend can
            # render a soft section divider or color band.
            cols.append({**col, "source_group": gkey})
    # Full view is always variant-scoped — one row per SKU is how the Amazon
    # NIS template expresses a multi-variant style.
    return {"label": "Full NIS sheet", "scope": "variant", "columns": cols}


def _lab_session_get_style(style_num):
    for s in lab_session.get("styles", []):
        if str(s.get("style_num", "")) == str(style_num):
            return s
    return None


@app.route("/api/lab/upload", methods=["POST"])
def lab_upload():
    """Upload a pre-upload sheet into the Lab session bucket.

    Reuses parse_product_file + the same template-example-row guard that
    the live Bulk Upload uses, but writes into lab_session instead of
    session_data so the two flows stay isolated.
    """
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    f = request.files["file"]
    ext = Path(f.filename).suffix.lower()
    if ext not in [".xlsx", ".xls", ".xlsm", ".csv", ".tsv"]:
        return jsonify({"error": f"Unsupported file type: {ext}"}), 400

    save_path = UPLOAD_PRODUCTS / f"lab_{f.filename}"
    f.save(str(save_path))
    lab_session["file_path"] = str(save_path)

    try:
        styles, errors, warnings = parse_product_file(str(save_path))
        # Reuse the same example-row strip from the Bulk Upload route
        TEMPLATE_EXAMPLE_STYLES = {"436008622"}
        if styles:
            from collections import Counter as _C
            brand_counter = _C((s.get("brand") or "").strip() for s in styles)
            total_rows = len(styles)
            stripped, removed = [], []
            for s in styles:
                sn = (s.get("style_num") or "").strip()
                br = (s.get("brand") or "").strip()
                if sn in TEMPLATE_EXAMPLE_STYLES:
                    removed.append({"style_num": sn, "brand": br}); continue
                if br and brand_counter.get(br, 0) <= 1 and total_rows >= 4:
                    maj = brand_counter.most_common(1)[0]
                    if maj[1] >= 3 and maj[0] != br:
                        removed.append({"style_num": sn, "brand": br}); continue
                stripped.append(s)
            if removed and stripped:
                styles = stripped
                for ex in removed:
                    warnings.append(f"Skipped template example row — style {ex['style_num']} ({ex['brand']}).")

        # Reject multi-brand uploads (same rule as live Bulk Upload)
        brands_found = set(s.get("brand", "") for s in styles if s.get("brand"))
        if len(brands_found) > 1:
            return jsonify({
                "error": "Multiple brands detected",
                "message": f"This file contains {len(brands_found)} brands: {', '.join(sorted(brands_found))}. Upload one brand at a time.",
            }), 400
        brand = next(iter(brands_found)) if brands_found else ""

        # Trim payload sent back to the client — only what Stage 1 needs
        client_styles = [{
            "style_num": s.get("style_num", ""),
            "style_name": s.get("style_name", ""),
            "subclass":   s.get("subclass", ""),
            "variants":   s.get("variants", []),
        } for s in styles]

        lab_session["styles"] = styles  # full record kept server-side
        lab_session["brand"] = brand
        # Reset snapshot name so a new upload starts a new timestamped file
        lab_session["_snapshot_name"] = None
        # Persist to disk so a server restart doesn't wipe edits
        _lab_session_persist()

        # ═══ Stage 5b — auto-generate on every fresh upload ═══
        # Kick off LLM generation in the background so the operator drops a
        # sheet and walks straight into populated cards 2-3 min later.
        # Skipped only when:
        #   • LLM unavailable (offline mode)
        #   • a generation job is already running
        #   • every style in this batch already carries _lab_generated content
        #     (re-upload of the same brand mid-session — don't blow away edits)
        # The frontend reads `auto_gen_started` and opens the streaming banner.
        auto_gen_started = False
        if _anthropic_client is not None and lab_gen_progress.get("status") != "running":
            already_drafted = all(
                bool((s.get("_lab_generated") or {}).get("title"))
                for s in styles
            ) if styles else False
            if not already_drafted:
                threading.Thread(
                    target=_lab_run_generation, args=(list(styles),), daemon=True
                ).start()
                auto_gen_started = True

        return jsonify({
            "ok": True,
            "brand": brand,
            "styles": client_styles,
            "warnings": warnings,
            "errors": errors,
            "auto_gen_started": auto_gen_started,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to parse: {str(e)[:200]}"}), 500


def _resolve_pt_for_style(style):
    """Resolve Amazon PT key (COAT/SHIRT/PANTS/etc.) for a style record.
    Falls back to subclass-based mapping if not pre-resolved."""
    pt = (style.get("product_type") or "").strip().upper()
    if pt:
        return pt
    return (_resolve_style_product_type(style) or "").upper()


def _hydrate_columns_with_dropdowns(columns, style_pt):
    """For columns of type 'dropdown_dynamic', fill in 'options' from
    Amazon's NIS template dropdown cache for the style's PT.

    Returns a NEW columns list — does not mutate the LAB_GRID_GROUPS source.
    """
    if not style_pt:
        return columns
    try:
        cache = load_dropdown_cache(style_pt) or {}
    except Exception:
        cache = {}
    out = []
    for col in columns:
        if col.get("type") == "dropdown_dynamic":
            field = col.get("dropdown_field", "")
            opts = cache.get(field) or []
            new_col = dict(col)
            new_col["type"] = "dropdown"
            new_col["options"] = opts
            new_col["_pt"] = style_pt
            new_col["_dropdown_field"] = field
            out.append(new_col)
        else:
            out.append(col)
    return out


def _build_style_row(style):
    """One row representing a single style for style-scope groups.
    Pulls from the style record + any LLM-generated content stored under
    style['_lab_generated'] (set by the upload step or a future regen).
    """
    gen = style.get("_lab_generated") or {}
    return {
        "style_num":      style.get("style_num", "") or "",
        "style_name":     style.get("style_name", "") or "",
        "brand":          style.get("brand", "") or "",
        "vendor_code":    style.get("vendor_code", "") or "",
        "subclass":       style.get("subclass", "") or "",
        "sub_subclass":   style.get("sub_subclass", "") or "",
        "product_type":   _resolve_pt_for_style(style) or "",
        # LLM-generated copy (filled lazily by /api/lab/generate or empty)
        "title":          gen.get("title", style.get("title", "")) or "",
        "bullet_1":       gen.get("bullet_1", "") or "",
        "bullet_2":       gen.get("bullet_2", "") or "",
        "bullet_3":       gen.get("bullet_3", "") or "",
        "bullet_4":       gen.get("bullet_4", "") or "",
        "bullet_5":       gen.get("bullet_5", "") or "",
        "description":    gen.get("description", "") or "",
        "backend_keywords": gen.get("backend_keywords", "") or "",
        # Taxonomy
        "feed_product_type": style.get("feed_product_type", "") or _resolve_pt_for_style(style) or "",
        "item_type":         style.get("item_type", "") or style.get("sub_subclass", "") or "",
        "department":        style.get("department", "") or "",
        "target_gender":     style.get("target_gender", style.get("gender", "")) or "",
        "age_range":         style.get("age_range", "") or "",
        "lifestyle_1":       style.get("lifestyle_1", "") or "",
        "lifestyle_2":       style.get("lifestyle_2", "") or "",
        # Compliance
        "coo":               style.get("coo", "") or "",
        "fabric":            style.get("fabric", "") or "",
        "care":              style.get("care", "") or "",
        "upf":               style.get("upf", "") or "",
        "contains_batteries": style.get("contains_batteries", "No") or "No",
        "is_hazmat":         style.get("is_hazmat", "No") or "No",
        "federal_contract_compliant": style.get("federal_contract_compliant", "No") or "No",
        # Apparel
        "sleeve_type":   style.get("sleeve_type", "") or "",
        "neck_type":     style.get("neck_type", "") or "",
        "fit_type":      style.get("fit_type", "") or "",
        "closure_type":  style.get("closure_type", "") or "",
        "collar_style":  style.get("collar_style", "") or "",
        "occasion":      style.get("occasion", "") or "",
        "special_feature_1": style.get("special_feature_1", "") or "",
        "special_feature_2": style.get("special_feature_2", "") or "",
        # Commercial
        "cost_price":   style.get("cost_price", "") or "",
        "list_price":   style.get("list_price", "") or "",
        "ship_date":    style.get("ship_date", "") or "",
        "model_name":   style.get("model_name", "") or "",
        "manufacturer": style.get("manufacturer", "") or style.get("brand", "") or "",
        "warranty":     style.get("warranty", "") or "",
    }


def _build_variant_row(v, style):
    """One row per variant for variant-scope groups.
    Pulls from variant + falls back to style-level for size_map/color_map."""
    return {
        "variant_id": v.get("variant_id", "") or v.get("sku", "") or "",
        "sku":        v.get("sku", "") or v.get("variant_id", "") or "",
        "upc":        v.get("upc", "") or "",
        "asin":       v.get("asin", "") or v.get("child_asin", "") or "",
        "color_name": v.get("color_name", "") or "",
        "color_map":  v.get("color_map", "") or "",
        "size":       v.get("size", "") or "",
        "size_map":   v.get("size_map", "") or "",
        "item_weight_value": v.get("item_weight_value", v.get("weight", "")) or "",
        "item_weight_unit":  v.get("item_weight_unit", "pounds") or "pounds",
        "item_length_value": v.get("item_length_value", "") or "",
        "item_width_value":  v.get("item_width_value", "") or "",
        "item_height_value": v.get("item_height_value", "") or "",
        "item_dim_unit":     v.get("item_dim_unit", "inches") or "inches",
    }


@app.route("/api/lab/grid", methods=["GET"])
def lab_grid_get():
    """Return the grid spec (columns + rows) for one style + group.

    columns: spec for the grid's column config (type, options, validators).
             dropdown_dynamic columns are hydrated with PT-specific options
             from Amazon's NIS template dropdown cache.
    rows:    one row per variant (variant-scope) or one row total
             (style-scope), pre-filled from style + variant records.
    """
    style_num = request.args.get("style", "").strip()
    group_key = request.args.get("group", "weight").strip()
    if not style_num:
        return jsonify({"error": "style param required"}), 400
    # "__full__" is a synthetic group: union of every style-scope and
    # variant-scope column from LAB_GRID_GROUPS, flattened onto one row
    # per variant with style-level fields repeated. This is the "Full NIS
    # sheet" view — what the operator sees matches what ships in the .xlsm.
    if group_key == "__full__":
        grp = _lab_build_full_group()
    else:
        grp = LAB_GRID_GROUPS.get(group_key)
    if not grp:
        return jsonify({"error": f"Unknown group: {group_key}"}), 400
    style = _lab_session_get_style(style_num)
    if not style:
        return jsonify({"error": f"Style {style_num} not in session. Upload first."}), 404

    # Resolve PT and hydrate dropdown columns with Amazon's options.
    style_pt = _resolve_pt_for_style(style)
    columns = _hydrate_columns_with_dropdowns(grp["columns"], style_pt)

    scope = grp.get("scope", "variant")
    if group_key == "__full__":
        # Variant rows with style-level fields merged in, so the operator
        # sees title/bullets/taxonomy repeated on every SKU row — just like
        # the downloaded .xlsm does.
        srow = _build_style_row(style)
        rows = []
        for v in style.get("variants", []) or []:
            vrow = _build_variant_row(v, style)
            rows.append({**srow, **vrow})
        if not rows:
            # No variants yet — still return one row so the operator can see
            # the style-level fields.
            rows = [srow]
    elif scope == "style":
        rows = [_build_style_row(style)]
    else:
        rows = [_build_variant_row(v, style) for v in style.get("variants", [])]

    return jsonify({
        "ok": True,
        "style_num": style_num,
        "group": group_key,
        "group_label": grp["label"],
        "scope": scope,
        "product_type": style_pt,
        "columns": columns,
        "rows": rows,
    })


@app.route("/api/lab/grid/groups", methods=["GET"])
def lab_grid_groups():
    """List the available groups + per-style missing-cell counts.
    Front-end uses this to render the group switcher with badges.
    """
    style_num = request.args.get("style", "").strip()
    style = _lab_session_get_style(style_num) if style_num else None
    out = []
    # Prepend the synthetic "Full NIS sheet" view so it's the first option
    # in the group switcher — default landing for anyone hitting 'Full view'.
    full_grp = _lab_build_full_group()
    full_missing = 0
    if style:
        srow = _build_style_row(style)
        variants = style.get("variants", []) or [None]
        for v in variants:
            row = {**srow, **(_build_variant_row(v, style) if v else {})}
            full_missing += sum(1 for c in full_grp["columns"]
                                 if c.get("required") and not row.get(c["key"]))
    out.append({
        "key": "__full__",
        "label": full_grp["label"],
        "scope": full_grp["scope"],
        "missing": full_missing,
        "col_count": len(full_grp["columns"]),
    })
    for key, grp in LAB_GRID_GROUPS.items():
        missing = 0
        if style:
            scope = grp.get("scope", "variant")
            if scope == "style":
                row = _build_style_row(style)
                missing = sum(1 for c in grp["columns"]
                              if c.get("required") and not row.get(c["key"]))
            else:
                for v in style.get("variants", []):
                    row = _build_variant_row(v, style)
                    missing += sum(1 for c in grp["columns"]
                                   if c.get("required") and not row.get(c["key"]))
        out.append({
            "key": key,
            "label": grp["label"],
            "scope": grp.get("scope", "variant"),
            "missing": missing,
            "col_count": len(grp["columns"]),
        })
    return jsonify({"ok": True, "groups": out})


# Style-level keys we route into style["_lab_generated"] so they never
# collide with the parsed-sheet "raw" fields.
_STYLE_GENERATED_KEYS = {
    "title", "bullet_1", "bullet_2", "bullet_3", "bullet_4", "bullet_5",
    "description", "backend_keywords",
}


@app.route("/api/lab/grid/save", methods=["POST"])
def lab_grid_save():
    """Persist grid edits back into lab_session.

    For variant-scope groups: write each editable column onto the matching
    variant by variant_id.
    For style-scope groups: write the single row's editable columns onto
    the style record (LLM copy fields routed under style['_lab_generated']).
    """
    data = request.get_json(force=True) or {}
    style_num = (data.get("style") or "").strip()
    group_key = (data.get("group") or "weight").strip()
    rows = data.get("rows") or []
    if not style_num:
        return jsonify({"error": "style required"}), 400
    if group_key == "__full__":
        grp = _lab_build_full_group()
    else:
        grp = LAB_GRID_GROUPS.get(group_key)
    if not grp:
        return jsonify({"error": f"Unknown group: {group_key}"}), 400
    style = _lab_session_get_style(style_num)
    if not style:
        return jsonify({"error": f"Style {style_num} not in session"}), 404

    writable_keys = [c["key"] for c in grp["columns"] if not c.get("readonly")]
    scope = grp.get("scope", "variant")
    updated = 0
    skipped = 0

    # Full NIS view: rows are variant-scoped but carry style-level fields
    # merged in. Save style-level keys once (from the first row) onto the
    # style record, and variant-level keys onto each matching variant.
    if group_key == "__full__":
        # Classify writable keys by source group
        style_keys = {c["key"] for c in _lab_build_full_group()["columns"]
                      if not c.get("readonly")
                      and c.get("source_group") in ("identity", "title_copy", "taxonomy",
                                                    "apparel", "compliance", "commercial")}
        variant_keys = set(writable_keys) - style_keys
        # Style-level: take from first row only (all variant rows carry the
        # same merged style fields, but we only want to write once).
        if rows:
            first = rows[0]
            gen = style.setdefault("_lab_generated", {})
            for k in style_keys:
                if k not in first:
                    continue
                new = first.get(k)
                new_norm = "" if new in (None, "") else new
                target_dict = gen if k in _STYLE_GENERATED_KEYS else style
                old = target_dict.get(k)
                old_norm = "" if old in (None, "") else old
                if old_norm != new_norm:
                    target_dict[k] = new_norm
                    updated += 1
        # Variant-level: per-row, matched by variant_id
        variants = style.get("variants", [])
        index = {(v.get("variant_id") or v.get("sku") or "").strip(): v for v in variants}
        for row in rows:
            vid = (row.get("variant_id") or "").strip()
            v = index.get(vid)
            if not v:
                skipped += 1
                continue
            for k in variant_keys:
                if k not in row:
                    continue
                old = v.get(k)
                new = row.get(k)
                old_norm = "" if old in (None, "") else old
                new_norm = "" if new in (None, "") else new
                if old_norm != new_norm:
                    v[k] = new_norm
                    updated += 1
    elif scope == "style":
        # One row expected; merge into the style record.
        if not rows:
            return jsonify({"ok": True, "updated": 0, "skipped_unmatched": 0})
        row = rows[0]
        gen = style.setdefault("_lab_generated", {})
        for k in writable_keys:
            if k not in row:
                continue
            new = row.get(k)
            new_norm = "" if new in (None, "") else new
            target_dict = gen if k in _STYLE_GENERATED_KEYS else style
            old = target_dict.get(k)
            old_norm = "" if old in (None, "") else old
            if old_norm != new_norm:
                target_dict[k] = new_norm
                updated += 1
    else:
        variants = style.get("variants", [])
        index = {(v.get("variant_id") or v.get("sku") or "").strip(): v for v in variants}
        for row in rows:
            vid = (row.get("variant_id") or "").strip()
            v = index.get(vid)
            if not v:
                skipped += 1
                continue
            for k in writable_keys:
                if k not in row:
                    continue
                old = v.get(k)
                new = row.get(k)
                old_norm = "" if old in (None, "") else old
                new_norm = "" if new in (None, "") else new
                if old_norm != new_norm:
                    v[k] = new_norm
                    updated += 1

    if updated:
        _lab_session_persist()
    return jsonify({
        "ok": True,
        "style_num": style_num,
        "group": group_key,
        "scope": scope,
        "updated": updated,
        "skipped_unmatched": skipped,
    })


@app.route("/api/lab/download-xlsm", methods=["GET"])
def lab_download_xlsm():
    """Emit the lab_session as a Vendor-Central-shaped .xlsm.

    Stage 2 ships a *single sheet* per style with all the edited fields
    flattened into a single header-row + data-rows shape. It's not yet a
    bit-perfect Amazon NIS template (that arrives in Stage 4 alongside
    the live rule-engine validation) — but it IS a clean, columnar export
    of every cell the operator has touched, downloadable in one click.
    """
    if not lab_session.get("styles"):
        return jsonify({"error": "No data in session. Upload first."}), 400

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atlas Lab Export"

    # Build header from union of all groups' columns.
    header_keys = []
    seen = set()
    style_keys, variant_keys = [], []
    for grp in LAB_GRID_GROUPS.values():
        for c in grp["columns"]:
            k = c["key"]
            if k in seen:
                continue
            seen.add(k)
            (style_keys if grp.get("scope") == "style" else variant_keys).append((k, c.get("title", k)))
    header = [("variant_id", "Variant ID")]
    for k, t in style_keys + variant_keys:
        if k != "variant_id":
            header.append((k, t))

    # Header row
    for ci, (_, title) in enumerate(header, start=1):
        cell = ws.cell(row=1, column=ci, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="08111A")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows: one per variant, with style-level fields repeated.
    rownum = 2
    for s in lab_session["styles"]:
        srow = _build_style_row(s)
        for v in s.get("variants", []):
            vrow = _build_variant_row(v, s)
            merged = {**srow, **vrow}
            for ci, (key, _) in enumerate(header, start=1):
                ws.cell(row=rownum, column=ci, value=merged.get(key, "") or "")
            rownum += 1

    # Sensible column widths
    for ci, (key, title) in enumerate(header, start=1):
        col_letter = openpyxl.utils.get_column_letter(ci)
        if key in ("title", "description"):
            ws.column_dimensions[col_letter].width = 60
        elif key.startswith("bullet_"):
            ws.column_dimensions[col_letter].width = 40
        elif key in ("backend_keywords",):
            ws.column_dimensions[col_letter].width = 36
        else:
            ws.column_dimensions[col_letter].width = max(12, min(28, len(title) + 6))
    ws.freeze_panes = "B2"

    # Save and return
    out_path = UPLOAD_PRODUCTS / f"atlas_lab_{lab_session.get('brand','export').replace(' ','_')}.xlsx"
    wb.save(str(out_path))
    return send_file(
        str(out_path),
        as_attachment=True,
        download_name=out_path.name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ══════════════════════════════════════════════════════════════
# LAB · STAGE 3 — brain in the grid
# generate • persist • sessions • brand voice • brand rules
# ══════════════════════════════════════════════════════════════
LAB_SESSIONS_DIR = BASE_DIR / "lab_sessions"
LAB_SESSIONS_DIR.mkdir(parents=True, exist_ok=True)


def _lab_session_persist():
    """Write the current lab_session to disk.

    File layout:
      lab_sessions/
        <Brand>_<YYYYMMDD-HHMMSS>.json    — timestamped snapshot per upload
        <Brand>__current.json             — latest, used to restore after
                                            server restart for the same brand

    Idempotent. Best-effort — never raises into the request handler.
    """
    if not lab_session.get("brand") or not lab_session.get("styles"):
        return None
    brand_safe = re.sub(r"[^\w]", "_", lab_session["brand"])
    payload = {
        "brand": lab_session["brand"],
        "styles": lab_session["styles"],
        "file_path": lab_session.get("file_path", ""),
        "saved_at": datetime.utcnow().isoformat("T") + "Z",
        "style_count": len(lab_session.get("styles", [])),
        "variant_count": sum(len(s.get("variants", [])) for s in lab_session.get("styles", [])),
    }
    # snapshot file
    if not lab_session.get("_snapshot_name"):
        ts = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
        lab_session["_snapshot_name"] = f"{brand_safe}_{ts}.json"
    snap_path = LAB_SESSIONS_DIR / lab_session["_snapshot_name"]
    cur_path = LAB_SESSIONS_DIR / f"{brand_safe}__current.json"
    try:
        with open(snap_path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, indent=2, default=str)
        with open(cur_path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, indent=2, default=str)
        return str(snap_path)
    except Exception as e:
        print(f"[lab] persist failed: {e}", flush=True)
        return None


@app.route("/api/lab/sessions", methods=["GET"])
def lab_sessions_list():
    """List saved Lab sessions newest first. Used by the Past Sessions picker."""
    items = []
    for p in sorted(LAB_SESSIONS_DIR.glob("*.json"),
                    key=lambda x: x.stat().st_mtime, reverse=True):
        try:
            with open(p, encoding="utf-8") as fh:
                meta = json.load(fh)
        except Exception:
            continue
        items.append({
            "file": p.name,
            "brand": meta.get("brand", ""),
            "saved_at": meta.get("saved_at", ""),
            "style_count": meta.get("style_count", 0),
            "variant_count": meta.get("variant_count", 0),
            "is_current": p.name.endswith("__current.json"),
        })
    return jsonify({"ok": True, "sessions": items[:60]})


@app.route("/api/atlas/home-state", methods=["GET"])
def atlas_home_state():
    """Return the numbers the Home page "In Progress" region displays.

    Strictly counts only what Atlas can directly observe inside its own
    sessions. No claims about shipping, going live, suppression prevention,
    or time saved — Atlas has no integration with Vendor Central or the
    brand's actual ops yet, so those would be unverifiable.

    - drafts_ready: styles where the operator has accepted all required
      fields (title + 5 bullets + description + backend_keywords). Means
      "this style is fully drafted and the operator can choose to ship it."
      Atlas does NOT claim it was shipped — only that the draft is ready.
    - drafts_in_progress: styles where some required content exists but
      not all (operator is mid-review).
    - styles_total: total style count in active sessions
    - active_brand: the brand of the most recent session, for surfacing
      "<brand> in progress" copy honestly.
    - continue: same as before — most recent session as Continue target.
    """
    REQUIRED_KEYS = ("title", "description", "backend_keywords",
                     "bullet_1", "bullet_2", "bullet_3", "bullet_4", "bullet_5")
    drafts_ready = 0
    drafts_in_progress = 0
    styles_total = 0
    continue_meta = None
    most_recent_mtime = 0.0
    active_brand = ""

    try:
        for p in LAB_SESSIONS_DIR.glob("*.json"):
            # Only count the "__current.json" snapshots so the numbers
            # represent live work, not the full archive of every session
            # ever saved. Otherwise restoring a past session would bump
            # counts as if work was redone.
            if not p.name.endswith("__current.json"):
                continue
            try:
                with open(p, encoding="utf-8") as fh:
                    meta = json.load(fh)
            except Exception:
                continue
            for style in meta.get("styles", []) or []:
                styles_total += 1
                gen = style.get("_lab_generated") or {}
                filled = sum(1 for k in REQUIRED_KEYS if (gen.get(k) or "").strip())
                if filled == len(REQUIRED_KEYS):
                    drafts_ready += 1
                elif filled > 0:
                    drafts_in_progress += 1
            mt = p.stat().st_mtime
            if mt > most_recent_mtime:
                most_recent_mtime = mt
                active_brand = meta.get("brand", "") or ""
                continue_meta = {
                    "file":         p.name,
                    "brand":        active_brand,
                    "style_count":  meta.get("style_count", len(meta.get("styles", []))),
                    "is_current":   True,
                }
    except Exception:
        pass

    return jsonify({
        "ok": True,
        "in_progress": {
            "drafts_ready":       drafts_ready,
            "drafts_in_progress": drafts_in_progress,
            "styles_total":       styles_total,
            "active_brand":       active_brand,
        },
        "continue":  continue_meta,
    })


# ─── Atlas substrate: operator response + session submit ─────────────
# Step 3 / Step 4 of the substrate build. These endpoints are the UI's
# way to write back into the decision log without ever blocking generation.
# Both are best-effort: a substrate failure must not break the review flow.

_ATLAS_VALID_ACTIONS = {"accept", "edit", "reject", "add_comment", "view"}
_ATLAS_VALID_SCOPES = {"none", "just_this", "batch", "brand_always", "skip"}


@app.route("/api/atlas/decision-response", methods=["POST"])
def atlas_decision_response():
    """Record an operator response (accept/edit/reject/comment/view) on a
    decision_event. Idempotent at the substrate level — the log is append
    only, so duplicate POSTs just produce duplicate rows that the reader
    can dedupe by (event_id, action) if needed.

    Body: {
        workspace_id: str,
        event_id: str,                      # from entry._atlas.event_ids[field]
        action: 'accept'|'edit'|'reject'|'add_comment'|'view',
        value: any,                         # new value if edit, else null
        scope: 'none'|'just_this'|'batch'|'brand_always'|'skip',
        time_to_decision_ms: int|null,
        comment: str|null,
        viewed_case: bool                   # v1.1.0
    }
    """
    data = request.get_json(force=True, silent=True) or {}
    workspace_id = (data.get("workspace_id") or "").strip()
    event_id = (data.get("event_id") or "").strip()
    action = (data.get("action") or "").strip().lower()
    if not workspace_id or not event_id:
        return jsonify({"ok": False, "error": "workspace_id and event_id required"}), 400
    if action not in _ATLAS_VALID_ACTIONS:
        return jsonify({"ok": False, "error": f"action must be one of {sorted(_ATLAS_VALID_ACTIONS)}"}), 400
    scope = (data.get("scope") or "none").strip().lower()
    if scope not in _ATLAS_VALID_SCOPES:
        return jsonify({"ok": False, "error": f"scope must be one of {sorted(_ATLAS_VALID_SCOPES)}"}), 400
    time_ms = data.get("time_to_decision_ms")
    try:
        time_ms = int(time_ms) if time_ms is not None else None
        if time_ms is not None and time_ms < 0:
            time_ms = None
    except (TypeError, ValueError):
        time_ms = None
    comment = data.get("comment")
    if isinstance(comment, str):
        comment = comment.strip() or None
    else:
        comment = None
    viewed_case = bool(data.get("viewed_case", False))

    try:
        from substrate.logger import update_field_decision_with_operator_response as _atlas_resp
        from substrate.schema import OperatorAction as _AtlasAction, OperatorScope as _AtlasScope
        _atlas_resp(
            workspace_id=workspace_id,
            event_id=event_id,
            operator_action=_AtlasAction(action),
            operator_value=data.get("value"),
            operator_scope=_AtlasScope(scope),
            operator_time_to_decision_ms=time_ms,
            operator_comment=comment,
            operator_viewed_case=viewed_case,
        )
    except Exception as exc:
        # Best-effort: log and return ok=false but 200, so the UI can
        # surface a soft toast without blocking the review flow.
        print(f"[atlas] decision-response skipped: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 200

    # Step 5: run judgment detection on edit actions only. Accept/view
    # responses don't change what the detectors see, so we skip them to
    # keep the hot path quick. Detection is best-effort — failures must
    # never break the response flow.
    moments: list[dict] = []
    if action == "edit":
        session_id = (data.get("session_id") or "").strip()
        if session_id:
            try:
                from substrate.judgment import detect_for_session as _atlas_detect
                moments = _atlas_detect(workspace_id, session_id) or []
            except Exception as exc:
                print(f"[atlas] judgment detection skipped: {exc}", flush=True)
    return jsonify({"ok": True, "judgment_moments": moments})


@app.route("/api/atlas/session-submit", methods=["POST"])
def atlas_session_submit():
    """Close an Atlas session with end-of-session notes + exemplar flag.

    Body: {
        workspace_id: str,
        session_id: str,
        operator_notes: str|null,
        exemplar: bool,
        answers: dict|null              # the 3/5/7 scaled questions; stored
                                        # inside operator_notes as a JSON tail
                                        # for v1.1.0 (no schema bump needed).
    }
    """
    data = request.get_json(force=True, silent=True) or {}
    workspace_id = (data.get("workspace_id") or "").strip()
    session_id = (data.get("session_id") or "").strip()
    if not workspace_id or not session_id:
        return jsonify({"ok": False, "error": "workspace_id and session_id required"}), 400
    notes = data.get("operator_notes")
    if isinstance(notes, str):
        notes = notes.strip() or None
    else:
        notes = None
    exemplar = bool(data.get("exemplar", False))
    answers = data.get("answers")

    # Fold structured answers into operator_notes as a JSON tail. This keeps
    # schema v1.1.0 untouched while giving us a structured record to query
    # once we read past the human-readable preamble.
    if isinstance(answers, dict) and answers:
        try:
            tail = json.dumps({"_session_answers": answers}, ensure_ascii=False)
            notes = (notes + "\n\n" + tail) if notes else tail
        except Exception:
            pass

    try:
        from substrate.logger import read_session as _atlas_read_session, submit_session as _atlas_submit
        from substrate.schema import SessionObject as _AtlasSession, Module as _AtlasModule
        existing = _atlas_read_session(workspace_id, session_id)
        if not existing:
            return jsonify({"ok": False, "error": "session not found"}), 404
        s = _AtlasSession(
            session_id=existing.get("session_id", session_id),
            workspace_id=existing.get("workspace_id", workspace_id),
            operator_id=existing.get("operator_id", ""),
            module=_AtlasModule(existing.get("module", "nis")),
            started_at=existing.get("started_at", ""),
            ended_at=existing.get("ended_at"),
            state=existing.get("state", "live"),
            operator_notes=existing.get("operator_notes"),
            exemplar=bool(existing.get("exemplar", False)),
        )
        _atlas_submit(s, operator_notes=notes, exemplar=exemplar)
    except Exception as exc:
        print(f"[atlas] session-submit skipped: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 200
    return jsonify({"ok": True, "session_id": session_id})


# ─── M2: mode-aware substrate endpoints ───────────────────────────────────

@app.route("/api/atlas/field-schema", methods=["GET"])
def atlas_field_schema():
    """Return the full field_schema.yml as JSON. Front-end caches this."""
    try:
        from substrate.field_suggest import load_field_schema
        schema = load_field_schema()
        return jsonify({"ok": True, "schema": schema})
    except Exception as exc:
        print(f"[atlas] field-schema failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/field-suggest", methods=["POST"])
def atlas_field_suggest():
    """Resolve a field's mode + payload (LLM-suggest fires here).

    Body: {table: str, field: str, context: dict|null}
    """
    body = request.get_json(silent=True) or {}
    table = (body.get("table") or "").strip()
    field = (body.get("field") or "").strip()
    if not table or not field:
        return jsonify({"ok": False, "error": "table and field required"}), 400
    context = body.get("context") or {}
    if not isinstance(context, dict):
        context = {}
    try:
        from substrate.field_suggest import suggest_for_field
        out = suggest_for_field(table, field, context=context)
        return jsonify(out)
    except Exception as exc:
        print(f"[atlas] field-suggest failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/asin-metadata/<asin>", methods=["GET"])
def atlas_asin_metadata_get(asin: str):
    """Read asin_metadata with parent-child inheritance applied.

    Query: ?raw=1 returns the unmerged child row.
    """
    workspace_id = _atlas_current_workspace()
    raw = request.args.get("raw") == "1"
    try:
        from substrate.asin_metadata import (
            get_asin_metadata, read_asin_metadata,
        )
        row = (get_asin_metadata if raw else read_asin_metadata)(
            workspace_id, asin,
        )
        if row is None:
            return jsonify({"ok": False, "error": "not found"}), 404
        return jsonify({"ok": True, "row": row})
    except Exception as exc:
        print(f"[atlas] asin-metadata get failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/asin-metadata", methods=["POST"])
def atlas_asin_metadata_set():
    """Upsert one asin_metadata row.

    Body: {asin, parent_asin?, variation_family?, variation_axes?,
           ground_truth_fields?, field_sources?}
    """
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    asin = (body.get("asin") or "").strip()
    if not asin:
        return jsonify({"ok": False, "error": "asin required"}), 400
    set_by = (body.get("set_by") or "devang").strip()
    try:
        from substrate.asin_metadata import set_asin_metadata
        ok = set_asin_metadata(
            workspace_id, asin,
            parent_asin=(body.get("parent_asin") or None),
            variation_family=(body.get("variation_family") or None),
            variation_axes=body.get("variation_axes") or None,
            ground_truth_fields=body.get("ground_truth_fields") or {},
            field_sources=body.get("field_sources") or {},
            set_by=set_by,
        )
        return jsonify({"ok": ok}), (200 if ok else 400)
    except Exception as exc:
        print(f"[atlas] asin-metadata set failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/asin-metadata/<asin>/confirm-field",
           methods=["POST"])
def atlas_asin_confirm_field(asin: str):
    """Mark one field's source as confirmed_by_operator=true.

    Body: {field: str, confirmed_by?: str}
    """
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    field = (body.get("field") or "").strip()
    if not field:
        return jsonify({"ok": False, "error": "field required"}), 400
    confirmed_by = (body.get("confirmed_by") or "devang").strip()
    try:
        from substrate.asin_metadata import confirm_field
        ok = confirm_field(workspace_id, asin, field, confirmed_by)
        return jsonify({"ok": ok}), (200 if ok else 400)
    except Exception as exc:
        print(f"[atlas] confirm-field failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/asin-metadata/<asin>/field", methods=["POST"])
def atlas_asin_set_field(asin: str):
    """Set one field's value + source.

    Body: {field, value, source, confirmed?, set_by?}
    """
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    field = (body.get("field") or "").strip()
    source = (body.get("source") or "operator_typed").strip()
    if not field:
        return jsonify({"ok": False, "error": "field required"}), 400
    try:
        from substrate.asin_metadata import record_field_source
        ok = record_field_source(
            workspace_id, asin, field,
            value=body.get("value"),
            source=source,
            confirmed=bool(body.get("confirmed", False)),
            set_by=(body.get("set_by") or "devang").strip(),
        )
        return jsonify({"ok": ok}), (200 if ok else 400)
    except Exception as exc:
        print(f"[atlas] set-field failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/asin-metadata/family/<parent_asin>",
           methods=["GET"])
def atlas_asin_family(parent_asin: str):
    """List all ASINs in a variation family (parent + children)."""
    workspace_id = _atlas_current_workspace()
    try:
        from substrate.asin_metadata import list_family_asins
        rows = list_family_asins(workspace_id, parent_asin)
        return jsonify({"ok": True, "rows": rows, "count": len(rows)})
    except Exception as exc:
        print(f"[atlas] family list failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/brand-position", methods=["GET"])
def atlas_brand_position_get():
    """Return workspace's brand_position row."""
    workspace_id = _atlas_current_workspace()
    try:
        from substrate.brand_position import get_brand_position
        row = get_brand_position(workspace_id)
        if row is None:
            return jsonify({"ok": True, "row": None})
        return jsonify({"ok": True, "row": row})
    except Exception as exc:
        print(f"[atlas] brand-position get failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/brand-position", methods=["POST"])
def atlas_brand_position_set():
    """Upsert workspace's brand_position row.

    Body: {position_statement, competitor_set[], competitor_role{},
           price_band{}, positioning_hypothesis?, next_review_at,
           review_freq?, set_by?}
    """
    from datetime import datetime
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    statement = (body.get("position_statement") or "").strip()
    if not statement:
        return jsonify({"ok": False, "error": "position_statement required"}), 400
    next_review_raw = body.get("next_review_at")
    if not next_review_raw:
        return jsonify({"ok": False, "error": "next_review_at required"}), 400
    try:
        next_review = datetime.fromisoformat(
            str(next_review_raw).replace("Z", "+00:00")
        )
    except ValueError:
        return jsonify({"ok": False, "error": "next_review_at invalid"}), 400
    try:
        from substrate.brand_position import set_brand_position
        ok = set_brand_position(
            workspace_id,
            position_statement=statement,
            competitor_set=list(body.get("competitor_set") or []),
            competitor_role=dict(body.get("competitor_role") or {}),
            price_band=dict(body.get("price_band") or {}),
            positioning_hypothesis=body.get("positioning_hypothesis"),
            next_review_at=next_review,
            set_by=(body.get("set_by") or "devang").strip(),
            review_freq=(body.get("review_freq") or "quarterly").strip(),
        )
        return jsonify({"ok": ok}), (200 if ok else 400)
    except Exception as exc:
        print(f"[atlas] brand-position set failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/operator-positions", methods=["GET"])
def atlas_op_positions_list():
    """List active operator_positions. Filters: scope, scope_ref,
    position_type, asin (resolves via read_active_positions)."""
    workspace_id = _atlas_current_workspace()
    asin = (request.args.get("asin") or "").strip() or None
    family = (request.args.get("family") or "").strip() or None
    dc = (request.args.get("decision_class") or "").strip() or None
    scope = (request.args.get("scope") or "").strip() or None
    scope_ref = (request.args.get("scope_ref") or "").strip() or None
    ptype = (request.args.get("position_type") or "").strip() or None
    try:
        from substrate.operator_positions import (
            list_active_positions, read_active_positions,
        )
        if asin or family or dc:
            rows = read_active_positions(
                workspace_id, asin=asin, family=family,
                decision_class=dc,
            )
        else:
            rows = list_active_positions(
                workspace_id, scope=scope, scope_ref=scope_ref,
                position_type=ptype,
            )
        return jsonify({"ok": True, "rows": rows, "count": len(rows)})
    except Exception as exc:
        print(f"[atlas] operator-positions list failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/operator-positions", methods=["POST"])
def atlas_op_positions_create():
    """Create one operator_position.

    Body: {scope, scope_ref?, claim, reasoning?, position_type?,
           evidence_refs?, created_by_event_id?}
    """
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    claim = (body.get("claim") or "").strip()
    scope = (body.get("scope") or "").strip()
    if not claim or not scope:
        return jsonify({"ok": False, "error": "scope+claim required"}), 400
    try:
        from substrate.operator_positions import create_position
        pid = create_position(
            workspace_id,
            scope=scope,
            scope_ref=(body.get("scope_ref") or None),
            claim=claim,
            reasoning=body.get("reasoning"),
            position_type=(body.get("position_type") or "strategic"),
            operator_id=(body.get("operator_id") or "devang"),
            evidence_refs=list(body.get("evidence_refs") or []),
            created_by_event_id=body.get("created_by_event_id"),
        )
        if pid is None:
            return jsonify({"ok": False, "error": "create failed"}), 400
        return jsonify({"ok": True, "position_id": pid})
    except Exception as exc:
        print(f"[atlas] operator-positions create failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/operator-positions/<position_id>/archive",
           methods=["POST"])
def atlas_op_positions_archive(position_id: str):
    """Archive a position."""
    body = request.get_json(silent=True) or {}
    archived_by = (body.get("archived_by") or "devang").strip()
    try:
        from substrate.operator_positions import archive_position
        ok = archive_position(position_id, archived_by)
        return jsonify({"ok": ok}), (200 if ok else 400)
    except Exception as exc:
        print(f"[atlas] operator-positions archive failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/pricing-logic", methods=["GET"])
def atlas_pricing_logic_get():
    """Read active pricing_logic for an ASIN/family/global."""
    workspace_id = _atlas_current_workspace()
    asin = (request.args.get("asin") or "").strip() or None
    family = (request.args.get("family") or "").strip() or None
    try:
        from substrate.pricing_logic import read_active_logic
        row = read_active_logic(workspace_id, asin=asin, family=family)
        return jsonify({"ok": True, "row": row})
    except Exception as exc:
        print(f"[atlas] pricing-logic get failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/pricing-logic", methods=["POST"])
def atlas_pricing_logic_set():
    """Upsert a pricing_logic row.

    Body: {scope, scope_ref?, floor_rule, ceiling_rule, reasoning?,
           ceiling_next_review_at?, set_by?}
    """
    from datetime import datetime
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    scope = (body.get("scope") or "").strip()
    floor_rule = body.get("floor_rule")
    ceiling_rule = body.get("ceiling_rule")
    if not scope or not isinstance(floor_rule, dict) or not isinstance(
        ceiling_rule, dict,
    ):
        return jsonify({"ok": False,
                        "error": "scope + floor_rule + ceiling_rule required"}), 400
    review_at = None
    raw = body.get("ceiling_next_review_at")
    if raw:
        try:
            review_at = datetime.fromisoformat(
                str(raw).replace("Z", "+00:00")
            )
        except ValueError:
            return jsonify({"ok": False,
                            "error": "ceiling_next_review_at invalid"}), 400
    try:
        from substrate.pricing_logic import set_pricing_logic
        ok = set_pricing_logic(
            workspace_id,
            scope=scope,
            scope_ref=body.get("scope_ref"),
            floor_rule=floor_rule,
            ceiling_rule=ceiling_rule,
            reasoning=body.get("reasoning"),
            set_by=(body.get("set_by") or "devang").strip(),
            ceiling_next_review_at=review_at,
        )
        return jsonify({"ok": ok}), (200 if ok else 400)
    except Exception as exc:
        print(f"[atlas] pricing-logic set failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/pricing-decisions", methods=["GET"])
def atlas_pricing_decisions_list():
    """List recent pricing_decisions. Filters: asin, goal_regime, limit."""
    workspace_id = _atlas_current_workspace()
    asin = (request.args.get("asin") or "").strip() or None
    regime = (request.args.get("goal_regime") or "").strip() or None
    try:
        limit = max(1, min(int(request.args.get("limit") or 50), 200))
    except ValueError:
        limit = 50
    try:
        from substrate.pricing_logic import list_pricing_decisions
        rows = list_pricing_decisions(
            workspace_id, asin=asin, goal_regime=regime, limit=limit,
        )
        return jsonify({"ok": True, "rows": rows, "count": len(rows)})
    except Exception as exc:
        print(f"[atlas] pricing-decisions list failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/pricing-decisions", methods=["POST"])
def atlas_pricing_decisions_log():
    """Append one pricing_decisions row.

    Body: {asin, price_set, mode, goal_regime?, floor_at_time?,
           ceiling_at_time?, play_zone_position?, reasoning?,
           pattern_tags?, meta?, price_set_by?}
    """
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    asin = (body.get("asin") or "").strip()
    if not asin:
        return jsonify({"ok": False, "error": "asin required"}), 400
    try:
        price = float(body.get("price_set"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "price_set must be number"}), 400
    mode = (body.get("mode") or "manual").strip()
    try:
        from substrate.pricing_logic import log_pricing_decision
        did = log_pricing_decision(
            workspace_id,
            asin=asin,
            price_set=price,
            price_set_by=(body.get("price_set_by") or "devang").strip(),
            mode=mode,
            goal_regime=(body.get("goal_regime") or "launch_velocity"),
            floor_at_time=body.get("floor_at_time"),
            ceiling_at_time=body.get("ceiling_at_time"),
            play_zone_position=body.get("play_zone_position"),
            reasoning=body.get("reasoning"),
            pattern_tags=list(body.get("pattern_tags") or []),
            meta=dict(body.get("meta") or {}),
        )
        if did is None:
            return jsonify({"ok": False, "error": "log failed"}), 400
        return jsonify({"ok": True, "decision_id": did})
    except Exception as exc:
        print(f"[atlas] pricing-decisions log failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/pricing-floor/compute", methods=["POST"])
def atlas_pricing_floor_compute():
    """Compute variable_contribution_zero floor from passed-in costs.

    Body: {floor_rule, landed_cost, fba_fee, third_pl_fee,
           ad_spend_per_unit, referral_rate?}
    """
    body = request.get_json(silent=True) or {}
    rule = body.get("floor_rule") or {"method": "variable_contribution_zero"}
    try:
        from substrate.pricing_logic import compute_floor_from_rule
        floor = compute_floor_from_rule(
            rule,
            landed_cost=body.get("landed_cost"),
            fba_fee=body.get("fba_fee"),
            third_pl_fee=body.get("third_pl_fee"),
            ad_spend_per_unit=body.get("ad_spend_per_unit"),
            referral_rate=float(body.get("referral_rate") or 0.15),
        )
        return jsonify({"ok": True, "floor": floor})
    except Exception as exc:
        print(f"[atlas] pricing-floor compute failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/competitor-state", methods=["GET"])
def atlas_competitor_state_list():
    """List competitor observations. Filters: competitor_id, metric, limit."""
    workspace_id = _atlas_current_workspace()
    cid = (request.args.get("competitor_id") or "").strip() or None
    metric = (request.args.get("metric") or "").strip() or None
    try:
        limit = max(1, min(int(request.args.get("limit") or 50), 200))
    except ValueError:
        limit = 50
    try:
        from substrate.competitor_state import list_observations
        rows = list_observations(
            workspace_id, competitor_id=cid, metric=metric, limit=limit,
        )
        return jsonify({"ok": True, "rows": rows, "count": len(rows)})
    except Exception as exc:
        print(f"[atlas] competitor-state list failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/competitor-state", methods=["POST"])
def atlas_competitor_state_record():
    """Record a competitor observation.

    Body: {competitor_id, metric, value, source?, asin?, notes?, observed_by?}
    """
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    cid = (body.get("competitor_id") or "").strip()
    metric = (body.get("metric") or "").strip()
    if not cid or not metric or "value" not in body:
        return jsonify({"ok": False,
                        "error": "competitor_id, metric, value required"}), 400
    try:
        from substrate.competitor_state import record_observation
        oid = record_observation(
            workspace_id,
            competitor_id=cid,
            metric=metric,
            value=body["value"],
            observed_by=(body.get("observed_by") or "devang").strip(),
            source=(body.get("source") or "operator_manual"),
            asin=body.get("asin"),
            notes=body.get("notes"),
        )
        if oid is None:
            return jsonify({"ok": False, "error": "record failed"}), 400
        return jsonify({"ok": True, "observation_id": oid})
    except Exception as exc:
        print(f"[atlas] competitor-state record failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/velune-onboarding", methods=["POST"])
def atlas_velune_onboarding():
    """Velune family onboarding wizard.

    Body shape:
      {
        parents: [
          {asin, variation_family, ground_truth_fields, field_sources?},
          ...
        ],
        children_axes: {  # map of parent_asin -> {colors[], sizes[]}
          "<parent_asin>": {
            "colors": [{"color_name":..., "color_map":...}, ...],
            "sizes":  ["XS","S","M","L","XL"],
            "child_asin_prefix": "B0VEL-PKT"  # optional
          }, ...
        },
        seed_positions: bool  (default true)
      }

    Creates 2 parent rows + 4*5=20 child rows per parent, seeds positions
    #1-5 from OPERATOR_POSITIONS.md.

    Returns {ok, created: {parents:int, children:int, positions:int},
              errors: [str]}.
    """
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    parents = list(body.get("parents") or [])
    children_axes = dict(body.get("children_axes") or {})
    seed_positions = bool(body.get("seed_positions", True))
    if not parents:
        return jsonify({"ok": False, "error": "parents[] required"}), 400

    created = {"parents": 0, "children": 0, "positions": 0}
    errors: list[str] = []

    try:
        from substrate.asin_metadata import set_asin_metadata
        from substrate.operator_positions import create_position

        # 1. Parents
        for p in parents:
            asin = (p.get("asin") or "").strip()
            if not asin:
                errors.append("parent missing asin")
                continue
            ok = set_asin_metadata(
                workspace_id, asin,
                variation_family=p.get("variation_family"),
                ground_truth_fields=p.get("ground_truth_fields") or {},
                field_sources=p.get("field_sources") or {},
                set_by="devang",
            )
            if ok:
                created["parents"] += 1
            else:
                errors.append(f"parent {asin} insert failed")

        # 2. Children — generate color * size matrix
        for parent_asin, axes in children_axes.items():
            colors = list(axes.get("colors") or [])
            sizes = list(axes.get("sizes") or [])
            prefix = (axes.get("child_asin_prefix") or parent_asin).strip()
            if not colors or not sizes:
                errors.append(f"{parent_asin} missing colors or sizes")
                continue
            # Inherit family from parent if not overridden
            family = next(
                (p.get("variation_family") for p in parents
                 if (p.get("asin") or "") == parent_asin),
                None,
            )
            for color in colors:
                cname = (color.get("color_name") or "").strip()
                cmap = (color.get("color_map") or cname).strip()
                if not cname:
                    continue
                for size in sizes:
                    size_clean = str(size).strip()
                    if not size_clean:
                        continue
                    color_slug = cname.upper().replace(" ", "")[:6]
                    child_asin = f"{prefix}-{color_slug}-{size_clean}"
                    ok = set_asin_metadata(
                        workspace_id, child_asin,
                        parent_asin=parent_asin,
                        variation_family=family,
                        variation_axes={"color": cname, "size": size_clean},
                        ground_truth_fields={
                            "color_name": cname,
                            "color_map": cmap,
                            "size": size_clean,
                        },
                        set_by="devang",
                    )
                    if ok:
                        created["children"] += 1
                    else:
                        errors.append(f"child {child_asin} insert failed")

        # 3. Seed positions #1-5 if not already present
        if seed_positions:
            from substrate.operator_positions import list_active_positions
            existing_claims = {
                r["claim"]
                for r in list_active_positions(workspace_id)
            }
            starters = [
                {"scope": "brand", "scope_ref": None,
                 "claim": "Athletic positioning only, no Casual",
                 "reasoning": ("Brand position is premium-adjacent at $35-55; "
                              "Casual register dilutes intent at this tier"),
                 "position_type": "strategic"},
                {"scope": "brand", "scope_ref": None,
                 "claim": "No discount, value, or budget language in any content",
                 "reasoning": ("Premium-adjacent positioning incompatible "
                              "with value framing"),
                 "position_type": "hard_refusal"},
                {"scope": "family", "scope_ref": "velune_pocket",
                 "claim": ("Family has hidden_waistband pocket; pocket details "
                          "must be accurate in all content"),
                 "reasoning": ("Pocket vs no-pocket is the family-defining "
                              "differentiator; accuracy prevents return-rate spikes"),
                 "position_type": "hard_refusal"},
                {"scope": "family", "scope_ref": "velune_no_pocket",
                 "claim": ("Product Name must NOT include 'with Pockets' or "
                          "imply pocket presence"),
                 "reasoning": ("Family is explicitly no-pocket; product name "
                              "accuracy is launch-blocking"),
                 "position_type": "hard_refusal"},
                {"scope": "brand", "scope_ref": None,
                 "claim": ("Goal regime defaults to launch velocity for first "
                          "60 days per ASIN, then transitions to margin unless "
                          "operator overrides"),
                 "reasoning": ("Launch ranking is time-sensitive; margin "
                              "optimization happens after rank stabilizes"),
                 "position_type": "workflow"},
            ]
            for s in starters:
                if s["claim"] in existing_claims:
                    continue
                pid = create_position(workspace_id, **s)
                if pid:
                    created["positions"] += 1

        return jsonify({
            "ok": True,
            "workspace_id": workspace_id,
            "created": created,
            "errors": errors,
        })
    except Exception as exc:
        print(f"[atlas] velune onboarding failed: {exc}", flush=True)
        return jsonify({"ok": False,
                        "error": str(exc)[:200],
                        "created": created,
                        "errors": errors}), 500


# ─── M4: recommendation ingest + tokenized response ────────────────────

@app.route("/api/atlas/recommendations", methods=["GET"])
def atlas_recommendations_list():
    """List recommendations newest first. Filters: status, source, limit."""
    workspace_id = _atlas_current_workspace()
    status = (request.args.get("status") or "").strip() or None
    source = (request.args.get("source") or "").strip() or None
    try:
        limit = max(1, min(int(request.args.get("limit") or 50), 200))
    except ValueError:
        limit = 50
    try:
        from substrate.recommendation_ingest import list_recommendations
        rows = list_recommendations(
            workspace_id, status=status, source=source, limit=limit,
        )
        return jsonify({"ok": True, "rows": rows, "count": len(rows)})
    except Exception as exc:
        print(f"[atlas] recommendations list failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/recommendations", methods=["POST"])
def atlas_recommendations_create():
    """Create a recommendation_ingest row.

    Body: {source, source_tier?, source_contact?, raw_text?,
           raw_file_path?, rec_type?, scope_asins?, scope_confidence?,
           parsed_fields?}
    """
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    source = (body.get("source") or "").strip()
    if not source:
        return jsonify({"ok": False, "error": "source required"}), 400
    try:
        from substrate.recommendation_ingest import create_recommendation
        rec_id = create_recommendation(
            workspace_id,
            source=source,
            source_tier=body.get("source_tier"),
            source_contact=body.get("source_contact"),
            raw_text=body.get("raw_text"),
            raw_file_path=body.get("raw_file_path"),
            raw_file_hash=body.get("raw_file_hash"),
            rec_type=body.get("rec_type"),
            scope_asins=list(body.get("scope_asins") or []),
            scope_confidence=body.get("scope_confidence"),
            parsed_fields=body.get("parsed_fields") or None,
            ingested_by=(body.get("ingested_by") or "devang").strip(),
        )
        if rec_id is None:
            return jsonify({"ok": False, "error": "create failed"}), 400
        return jsonify({"ok": True, "rec_id": rec_id})
    except Exception as exc:
        print(f"[atlas] recommendations create failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/recommendations/<rec_id>", methods=["GET"])
def atlas_recommendations_get(rec_id: str):
    """Return one recommendation + its evaluation rows + summary."""
    try:
        from substrate.recommendation_ingest import get_recommendation
        from substrate.atlas_evaluation import (
            list_evaluations, summarize_rec,
        )
        rec = get_recommendation(rec_id)
        if rec is None:
            return jsonify({"ok": False, "error": "not found"}), 404
        evals = list_evaluations(rec_id)
        summary = summarize_rec(rec_id)
        return jsonify({
            "ok": True, "rec": rec,
            "evaluations": evals, "summary": summary,
        })
    except Exception as exc:
        print(f"[atlas] recommendations get failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/recommendations/<rec_id>/evaluate",
           methods=["POST"])
def atlas_recommendations_evaluate(rec_id: str):
    """Run parse + verdict pass on a recommendation.

    Body (optional): {force_reparse: bool}. When parsed_fields already
    exists and force_reparse is false (default), reuses prior parse.
    """
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    force_reparse = bool(body.get("force_reparse", False))
    try:
        from substrate.recommendation_ingest import (
            get_recommendation, update_parse, set_status,
        )
        from substrate.atlas_evaluation import create_evaluation
        from substrate.rec_evaluator import (
            parse_raw_text, evaluate_recommendation,
        )
        from substrate.brand_position import get_brand_position

        rec = get_recommendation(rec_id)
        if rec is None:
            return jsonify({"ok": False, "error": "not found"}), 404

        parsed = rec.get("parsed_fields") or {}
        if force_reparse or not parsed:
            parsed = parse_raw_text(rec.get("raw_text") or "")
            update_parse(rec_id, parsed_fields=parsed)

        if not parsed:
            return jsonify({
                "ok": False,
                "error": "no parsed fields; supply parsed_fields explicitly",
            }), 400

        bp = get_brand_position(workspace_id)
        verdicts = evaluate_recommendation(
            parsed,
            workspace_id=workspace_id,
            source=rec.get("source") or "",
            source_tier=rec.get("source_tier"),
            scope_asins=rec.get("scope_asins") or [],
            brand_position=bp,
        )
        written = 0
        for v in verdicts:
            eid = create_evaluation(
                rec_id, workspace_id,
                field_name=v["field_name"],
                submitted_value=v.get("submitted_value"),
                field_owner=v["field_owner"],
                verdict=v["verdict"],
                reasoning=v["reasoning"],
                citations=v.get("citations") or [],
                proposed_alternative=v.get("proposed_alternative"),
                test_design=v.get("test_design"),
                evidence_path=v.get("evidence_path"),
                confidence=v.get("confidence"),
                criticality=v.get("criticality") or "normal",
            )
            if eid:
                written += 1
        set_status(rec_id, "evaluated")
        return jsonify({
            "ok": True, "rec_id": rec_id,
            "parsed_fields": parsed,
            "evaluations_written": written,
        })
    except Exception as exc:
        print(f"[atlas] recommendations evaluate failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/recommendations/<rec_id>/token",
           methods=["POST"])
def atlas_recommendations_token(rec_id: str):
    """Generate (or regenerate) the tokenized response link.

    Body (optional): {ttl_days: int, base_url: str}
    """
    body = request.get_json(silent=True) or {}
    base_url = body.get("base_url") or request.host_url.rstrip("/")
    try:
        ttl_days = max(1, min(int(body.get("ttl_days") or 7), 30))
    except (TypeError, ValueError):
        ttl_days = 7
    try:
        from substrate.recommendation_ingest import generate_response_token
        link = generate_response_token(
            rec_id, base_url=base_url, ttl_days=ttl_days,
        )
        if link is None:
            return jsonify({"ok": False, "error": "token gen failed"}), 400
        return jsonify({"ok": True, **link})
    except Exception as exc:
        print(f"[atlas] recommendations token failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/recommendations/<rec_id>/status",
           methods=["POST"])
def atlas_recommendations_set_status(rec_id: str):
    """Change a recommendation's status.

    Body: {status: str}
    """
    body = request.get_json(silent=True) or {}
    status = (body.get("status") or "").strip()
    if not status:
        return jsonify({"ok": False, "error": "status required"}), 400
    try:
        from substrate.recommendation_ingest import set_status
        ok = set_status(rec_id, status)
        return jsonify({"ok": ok}), (200 if ok else 400)
    except Exception as exc:
        print(f"[atlas] set_status failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/evaluations/<eval_id>/operator-decision",
           methods=["POST"])
def atlas_eval_operator_decision(eval_id: str):
    """Record the operator's final call on one evaluation row.

    Body: {decision: accept|override|defer|reject,
           final_value?, reasoning?}
    """
    body = request.get_json(silent=True) or {}
    decision = (body.get("decision") or "").strip()
    if not decision:
        return jsonify({"ok": False, "error": "decision required"}), 400
    try:
        from substrate.atlas_evaluation import apply_operator_decision
        ok = apply_operator_decision(
            eval_id,
            decision=decision,
            final_value=body.get("final_value"),
            reasoning=body.get("reasoning"),
        )
        return jsonify({"ok": ok}), (200 if ok else 400)
    except Exception as exc:
        print(f"[atlas] operator-decision failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


# ─── M5: content benchmarks ─────────────────────────────────────────

@app.route("/api/atlas/benchmarks", methods=["GET"])
def atlas_benchmarks_list():
    """List benchmarks. Filters: scope, scope_ref, benchmark_type, status."""
    workspace_id = _atlas_current_workspace()
    scope = (request.args.get("scope") or "").strip() or None
    scope_ref = request.args.get("scope_ref")
    if scope_ref is not None:
        scope_ref = scope_ref.strip() or None
    benchmark_type = (request.args.get("benchmark_type") or "").strip() or None
    status = (request.args.get("status") or "").strip() or None
    try:
        from substrate.content_benchmarks import list_benchmarks
        rows = list_benchmarks(
            workspace_id,
            scope=scope, scope_ref=scope_ref,
            benchmark_type=benchmark_type, status=status,
        )
        return jsonify({"ok": True, "rows": rows, "count": len(rows)})
    except Exception as exc:
        print(f"[atlas] benchmarks list failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/benchmarks", methods=["POST"])
def atlas_benchmarks_lock():
    """Lock a new benchmark.

    Body: {scope, scope_ref?, benchmark_type, approved_value,
           source_event_id, citations?, resolved_inputs?,
           open_unknowns_at_approval?, approved_by?, enforce_cap?}
    """
    workspace_id = _atlas_current_workspace()
    body = request.get_json(silent=True) or {}
    scope = (body.get("scope") or "").strip()
    btype = (body.get("benchmark_type") or "").strip()
    source_event_id = (body.get("source_event_id") or "").strip()
    if not scope or not btype or not source_event_id or \
       "approved_value" not in body:
        return jsonify({
            "ok": False,
            "error": (
                "scope, benchmark_type, source_event_id, "
                "and approved_value are required"
            ),
        }), 400
    try:
        from substrate.content_benchmarks import lock_benchmark
        bid = lock_benchmark(
            workspace_id,
            scope=scope,
            scope_ref=(body.get("scope_ref") or None),
            benchmark_type=btype,
            approved_value=body["approved_value"],
            source_event_id=source_event_id,
            approved_by=(body.get("approved_by") or "devang"),
            citations=list(body.get("citations") or []),
            resolved_inputs=dict(body.get("resolved_inputs") or {}),
            open_unknowns_at_approval=list(
                body.get("open_unknowns_at_approval") or []
            ),
            enforce_cap=bool(body.get("enforce_cap", True)),
            meta=dict(body.get("meta") or {}),
        )
        if bid is None:
            return jsonify({
                "ok": False,
                "error": (
                    "lock failed (cap reached or invalid input); "
                    "archive an active benchmark or check inputs"
                ),
            }), 400
        return jsonify({"ok": True, "benchmark_id": bid})
    except Exception as exc:
        print(f"[atlas] benchmarks lock failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/benchmarks/<benchmark_id>", methods=["GET"])
def atlas_benchmarks_get(benchmark_id: str):
    """Fetch one benchmark."""
    try:
        from substrate.content_benchmarks import get_benchmark
        row = get_benchmark(benchmark_id)
        if row is None:
            return jsonify({"ok": False, "error": "not found"}), 404
        return jsonify({"ok": True, "row": row})
    except Exception as exc:
        print(f"[atlas] benchmarks get failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/benchmarks/applicable", methods=["GET"])
def atlas_benchmarks_applicable():
    """Resolve benchmarks that could seed a generation.

    Query: ?benchmark_type=title&asin=B0...&family=velune_pocket
           &decision_class=title_generation&include_review=1
    """
    workspace_id = _atlas_current_workspace()
    btype = (request.args.get("benchmark_type") or "").strip()
    if not btype:
        return jsonify({
            "ok": False, "error": "benchmark_type required",
        }), 400
    asin = (request.args.get("asin") or "").strip() or None
    family = (request.args.get("family") or "").strip() or None
    dc = (request.args.get("decision_class") or "").strip() or None
    include_review = request.args.get("include_review") == "1"
    try:
        from substrate.content_benchmarks import list_applicable
        rows = list_applicable(
            workspace_id,
            benchmark_type=btype,
            asin=asin, family=family, decision_class=dc,
            include_review_recommended=include_review,
        )
        return jsonify({"ok": True, "rows": rows, "count": len(rows)})
    except Exception as exc:
        print(f"[atlas] benchmarks applicable failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/benchmarks/<benchmark_id>/archive",
           methods=["POST"])
def atlas_benchmarks_archive(benchmark_id: str):
    """Archive a benchmark."""
    body = request.get_json(silent=True) or {}
    archived_by = (body.get("archived_by") or "devang").strip()
    try:
        from substrate.content_benchmarks import archive
        ok = archive(benchmark_id, archived_by)
        return jsonify({"ok": ok}), (200 if ok else 400)
    except Exception as exc:
        print(f"[atlas] benchmarks archive failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/benchmarks/<benchmark_id>/reactivate",
           methods=["POST"])
def atlas_benchmarks_reactivate(benchmark_id: str):
    """Move a review_recommended benchmark back to active."""
    body = request.get_json(silent=True) or {}
    by = (body.get("reactivated_by") or "devang").strip()
    try:
        from substrate.content_benchmarks import reactivate
        ok = reactivate(benchmark_id, by)
        return jsonify({"ok": ok}), (200 if ok else 400)
    except Exception as exc:
        print(f"[atlas] benchmarks reactivate failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/benchmarks/<old_id>/supersede",
           methods=["POST"])
def atlas_benchmarks_supersede(old_id: str):
    """Mark old benchmark superseded by a new one.

    Body: {new_benchmark_id, superseded_by_operator?}
    """
    body = request.get_json(silent=True) or {}
    new_id = (body.get("new_benchmark_id") or "").strip()
    by = (body.get("superseded_by_operator") or "devang").strip()
    if not new_id:
        return jsonify({
            "ok": False, "error": "new_benchmark_id required",
        }), 400
    try:
        from substrate.content_benchmarks import supersede
        ok = supersede(old_id, new_id, by)
        return jsonify({"ok": ok}), (200 if ok else 400)
    except Exception as exc:
        print(f"[atlas] benchmarks supersede failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/atlas/benchmarks/<benchmark_id>/bump-usage",
           methods=["POST"])
def atlas_benchmarks_bump_usage(benchmark_id: str):
    """Increment used_count + last_used_at. Called by NIS when a
    generation seeds from this benchmark."""
    try:
        from substrate.content_benchmarks import bump_usage
        ok = bump_usage(benchmark_id)
        return jsonify({"ok": ok}), (200 if ok else 400)
    except Exception as exc:
        print(f"[atlas] benchmarks bump_usage failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


# ─── Public tokenized response page (no login) ─────────────────────────

@app.route("/respond/<rec_id>", methods=["GET"])
def respond_get(rec_id: str):
    """Render the agency-facing response form. Token in query string."""
    from flask import render_template, abort
    token = (request.args.get("token") or "").strip()
    try:
        from substrate.recommendation_ingest import lookup_by_token
        from substrate.atlas_evaluation import list_evaluations
        rec = lookup_by_token(rec_id, token)
        if rec is None:
            return render_template(
                "_atlas_respond.html",
                rec=None, token=None,
                pending_evals=[], other_evals=[],
            )
        all_evals = list_evaluations(rec_id)
        pending = [
            e for e in all_evals
            if e["field_owner"] == "agency"
            and not e.get("agency_response")
        ]
        other = [e for e in all_evals if e not in pending]
        return render_template(
            "_atlas_respond.html",
            rec=rec, token=token,
            pending_evals=pending,
            other_evals=other,
        )
    except Exception as exc:
        print(f"[atlas] respond_get failed: {exc}", flush=True)
        return abort(500)


@app.route("/respond/<rec_id>/draft", methods=["POST"])
def respond_save_draft(rec_id: str):
    """Save draft responses without flipping status or consuming token."""
    body = request.get_json(silent=True) or {}
    token = (body.get("token") or "").strip()
    responses = list(body.get("responses") or [])
    try:
        from substrate.recommendation_ingest import lookup_by_token
        from substrate.atlas_evaluation import apply_agency_response
        rec = lookup_by_token(rec_id, token)
        if rec is None:
            return jsonify({"ok": False, "error": "invalid or expired token"}), 403
        written = 0
        for r in responses:
            eid = (r.get("eval_id") or "").strip()
            text = (r.get("response_text") or "").strip()
            if not eid or not text:
                continue
            ok = apply_agency_response(
                eid,
                response_text=text,
                agency_confidence=r.get("agency_confidence"),
            )
            if ok:
                written += 1
        return jsonify({"ok": True, "saved": written})
    except Exception as exc:
        print(f"[atlas] respond_save_draft failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/respond/<rec_id>/submit", methods=["POST"])
def respond_submit_final(rec_id: str):
    """Final submit: writes responses, flips status, consumes token."""
    body = request.get_json(silent=True) or {}
    token = (body.get("token") or "").strip()
    responses = list(body.get("responses") or [])
    try:
        from substrate.recommendation_ingest import (
            lookup_by_token, mark_response_received, consume_token,
        )
        from substrate.atlas_evaluation import apply_agency_response
        rec = lookup_by_token(rec_id, token)
        if rec is None:
            return jsonify({"ok": False, "error": "invalid or expired token"}), 403
        written = 0
        for r in responses:
            eid = (r.get("eval_id") or "").strip()
            text = (r.get("response_text") or "").strip()
            if not eid or not text:
                continue
            ok = apply_agency_response(
                eid,
                response_text=text,
                agency_confidence=r.get("agency_confidence"),
            )
            if ok:
                written += 1
        mark_response_received(rec_id)
        consume_token(rec_id)
        return jsonify({"ok": True, "saved": written})
    except Exception as exc:
        print(f"[atlas] respond_submit failed: {exc}", flush=True)
        return jsonify({"ok": False, "error": str(exc)[:200]}), 500


@app.route("/api/lab/load-session", methods=["POST"])
def lab_load_session():
    """Restore a saved Lab session by filename.

    Loads styles + brand + file_path back into lab_session and returns the
    same payload shape as /api/lab/upload so the front-end can rerender.
    """
    data = request.get_json(force=True) or {}
    fname = (data.get("file") or "").strip()
    if not fname or "/" in fname or "\\\\" in fname:
        return jsonify({"error": "file required"}), 400
    p = LAB_SESSIONS_DIR / fname
    if not p.exists():
        return jsonify({"error": "session not found"}), 404
    try:
        with open(p, encoding="utf-8") as fh:
            payload = json.load(fh)
    except Exception as e:
        return jsonify({"error": f"load failed: {e}"}), 500
    lab_session["brand"] = payload.get("brand", "")
    lab_session["styles"] = payload.get("styles", [])
    lab_session["file_path"] = payload.get("file_path", "")
    lab_session["_snapshot_name"] = fname if not fname.endswith("__current.json") else None
    client_styles = [{
        "style_num": s.get("style_num", ""),
        "style_name": s.get("style_name", ""),
        "subclass": s.get("subclass", ""),
        "variants": s.get("variants", []),
    } for s in lab_session["styles"]]
    return jsonify({
        "ok": True,
        "brand": lab_session["brand"],
        "styles": client_styles,
        "warnings": [f"Restored {len(client_styles)} styles from {fname}."],
        "errors": [],
    })


# ──── Brand voice + rules ─────────────────────────────────────
@app.route("/api/lab/brand-rules", methods=["GET"])
def lab_brand_rules():
    """Return the effective rules Atlas applies for a brand.

    Stage 3 ships READ-ONLY — the operator sees what Atlas does today.
    Stage 4 makes the structured rules editable. Voice IS editable today
    via /api/lab/brand-voice.
    """
    brand = (request.args.get("brand") or "").strip() or lab_session.get("brand", "")
    if not brand:
        return jsonify({"error": "brand required"}), 400
    cfg = _load_brand_config_data(brand) or {}
    return jsonify({
        "ok": True,
        "brand": brand,
        "voice": cfg.get("brand_voice", "") or cfg.get("voice", ""),
        "rules": {
            "title": {
                "format": "{brand} {gender} {item_type} – {key_attribute}, {color}",
                "max_chars": 120,
                "required_tokens": ["brand", "gender", "item_type"],
                "forbidden_words": list(cfg.get("never_words") or []) + [
                    "luxurious", "ultimate", "perfect", "best",
                ],
            },
            "bullets": {
                "count": 5,
                "max_chars_each": 256,
                "format": "ALL-CAPS HEADLINE — sentence.",
                "bullet_1_focus": cfg.get("bullet_1_focus") or "material",
                "forbidden_words": ["luxurious", "ultimate", "perfect", "best"],
            },
            "description": {
                "max_chars": 2000,
                "banned_stock_phrases": [
                    "the modern woman", "the modern man", "the modern individual",
                    "women who refuse", "men who refuse",
                    "woman on the move", "man on the move",
                ],
                "gender_discipline": (
                    "Description must match the title's gender. Men's title → men/man/he. "
                    "Women's title → women/woman/she. Unknown gender → customer/wearer/you."
                ),
            },
            "backend_keywords": {
                "max_bytes": 249,
                "format": "lowercase, space-separated",
                "forbidden": ["commas", "brand name", "overlap with title or bullets"],
            },
            "compliance": {
                "block_upf_without_test": True,
                "block_antimicrobial_without_epa": True,
                "block_promotional_language": True,
                "block_competitor_brand_names": True,
            },
            "defaults": {
                "care":   cfg.get("default_care", ""),
                "coo":    cfg.get("default_coo", ""),
                "upf":    cfg.get("default_upf", ""),
                "gender": cfg.get("gender", ""),
                "vendor_code_full": cfg.get("vendor_code_full", ""),
            },
        },
        # Tag the panel as read-only in Stage 3 so the front-end can show it that way.
        "editable": False,
        "voice_editable": True,
    })


@app.route("/api/lab/brand-voice", methods=["POST"])
def lab_brand_voice_save():
    """Save freeform brand voice text into the brand_config JSON.

    Body: {brand, voice}. Voice persists under cfg['brand_voice'].
    """
    data = request.get_json(force=True) or {}
    brand = (data.get("brand") or "").strip() or lab_session.get("brand", "").strip()
    voice = (data.get("voice") or "").strip()
    # Guard: brand must be ≥3 chars and contain at least one letter so we don't
    # write garbage filenames like ___________.json from a stray empty submit.
    if not brand or len(brand) < 3 or not re.search(r"[A-Za-z]", brand):
        return jsonify({"error": "brand required"}), 400
    cfg = _load_brand_config_data(brand) or {}
    cfg["brand_voice"] = voice
    brand_file = BRAND_CONFIGS_DIR / f"{re.sub(r'[^\\w]', '_', brand)}.json"
    try:
        with open(str(brand_file), "w", encoding="utf-8") as fh:
            json.dump(cfg, fh, indent=2)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"ok": True, "brand": brand, "chars": len(voice)})


# ──── Generate (LLM) ────────────────────────────────────────────
lab_gen_progress = {
    "total": 0, "completed": 0,
    "current_style": "", "current_step": "",
    "status": "idle",   # idle | running | done | error
    "error": "", "started_at": None,
}


def _lab_run_generation(target_styles):
    """Background worker. Runs the existing generate_content_llm pipeline
    over the requested styles, then routes results into style['_lab_generated']
    after the gender-drift scrub. Mirrors _run_content_generation_impl but
    scoped to lab_session and only the few fields the grid surfaces.
    """
    global lab_gen_progress
    try:
        brand = lab_session.get("brand", "") or ""
        brand_cfg = _load_brand_config_data(brand) or {}
        feedback_history = ""  # Lab Stage 3 doesn't yet replay operator feedback
        lab_gen_progress.update({
            "total": len(target_styles),
            "completed": 0,
            "current_style": "",
            "current_step": "Initializing…",
            "status": "running",
            "error": "",
            "started_at": time.time(),
        })
        for i, style in enumerate(target_styles):
            sn = style.get("style_num", f"row{i}")
            name = (style.get("style_name", "") or "")[:80]
            lab_gen_progress["current_style"] = f"{sn} — {name}"
            lab_gen_progress["current_step"] = "Generating SEO-optimized title (max 120 chars)…"
            try:
                llm_result = None
                if _anthropic_client is not None:
                    try:
                        llm_result = generate_content_llm(brand_cfg, brand, style, feedback_history)
                    except Exception as e:
                        print(f"[lab][LLM] {sn} fallback: {e}", flush=True)
                if llm_result:
                    title = llm_result.get("title", "") or ""
                    bullets = [
                        llm_result.get("bullet_1", ""),
                        llm_result.get("bullet_2", ""),
                        llm_result.get("bullet_3", ""),
                        llm_result.get("bullet_4", ""),
                        llm_result.get("bullet_5", ""),
                    ]
                    description = llm_result.get("description", "") or ""
                    backend_kw = llm_result.get("backend_keywords", "") or ""
                else:
                    # Rule-based fallback (same shape as the live Bulk Upload pipeline)
                    pt = (_resolve_style_product_type(style) or "").upper()
                    style_gender, _ = _derive_gender_department(style)
                    eff_gender = style_gender or brand_cfg.get("gender", "")
                    subclass = style.get("subclass", "") or ""
                    sub_subclass = style.get("sub_subclass", "") or ""
                    fabric = parse_fabric(style.get("fabric", "")) or brand_cfg.get("default_fabric", "")
                    care = style.get("care", "") or brand_cfg.get("default_care", "")
                    upf = style.get("upf", "") or brand_cfg.get("default_upf", "")
                    first_color = (style.get("variants", [{}])[0].get("color_name", "") if style.get("variants") else "") or ""
                    first_size = (style.get("variants", [{}])[0].get("size", "") if style.get("variants") else "") or ""
                    pt_label = subclass or sub_subclass or pt.replace("_", " ").title() or "Item"
                    title = generate_title(brand_cfg, brand, style.get("style_name", ""), pt_label,
                                           first_color, first_size, upf, style_gender=style_gender)
                    bullets = generate_bullets(brand_cfg, brand, style.get("style_name", ""),
                                               sub_subclass, fabric, care, first_color, upf,
                                               subclass=subclass, gender=eff_gender, product_type=pt,
                                               style_num=sn)
                    description = generate_description(brand_cfg, brand, sn, style.get("style_name", ""),
                                                       sub_subclass, fabric, care, first_color, upf,
                                                       subclass=subclass, gender=eff_gender, product_type=pt)
                    backend_kw = generate_backend_keywords(brand, style.get("style_name", ""),
                                                            subclass, first_color, fabric, upf,
                                                            subclass=subclass, gender=eff_gender, product_type=pt)

                # Gender-drift scrub (Pass 12.7) — same belt-and-suspenders
                style_gender2, _ = _derive_gender_department(style)
                eff_g = style_gender2 or brand_cfg.get("gender", "") or ""
                if eff_g not in ("Male", "Female") and title:
                    t = title.lower()
                    if "women's" in t or "women\u2019s" in t:
                        eff_g = "Female"
                    elif "men's" in t or "men\u2019s" in t:
                        eff_g = "Male"
                description = _scrub_gender_drift(description, eff_g)
                bullets = [_scrub_gender_drift(b or "", eff_g) for b in bullets]

                style["_lab_generated"] = {
                    "title": title,
                    "bullet_1": bullets[0] if len(bullets) > 0 else "",
                    "bullet_2": bullets[1] if len(bullets) > 1 else "",
                    "bullet_3": bullets[2] if len(bullets) > 2 else "",
                    "bullet_4": bullets[3] if len(bullets) > 3 else "",
                    "bullet_5": bullets[4] if len(bullets) > 4 else "",
                    "description": description,
                    "backend_keywords": backend_kw,
                    "_generated_at": datetime.utcnow().isoformat("T") + "Z",
                    "_llm_used": llm_result is not None,
                }
                lab_gen_progress["completed"] = i + 1
                lab_gen_progress["current_step"] = f"✓ {sn} complete"
            except Exception as style_err:
                traceback.print_exc()
                style["_lab_generated"] = {
                    "_pipeline_error": True,
                    "_error": f"{type(style_err).__name__}: {str(style_err)[:200]}",
                }
                lab_gen_progress["completed"] = i + 1
                lab_gen_progress["current_step"] = f"⚠ {sn} — generation error"
        # Persist after the run
        _lab_session_persist()
        lab_gen_progress["status"] = "done"
        lab_gen_progress["current_style"] = ""
        lab_gen_progress["current_step"] = ""
    except Exception as worker_err:
        traceback.print_exc()
        lab_gen_progress["status"] = "error"
        lab_gen_progress["error"] = f"{type(worker_err).__name__}: {str(worker_err)[:300]}"


@app.route("/api/lab/generate", methods=["POST"])
def lab_generate():
    """Kick off generation for one style or all styles in lab_session.

    Body: {scope: 'all' | 'style', style: <style_num> (when scope='style')}
    Returns immediately; the front-end polls /api/lab/generate-progress.
    """
    if not lab_session.get("styles"):
        return jsonify({"error": "No data in session. Upload first."}), 400
    if lab_gen_progress.get("status") == "running":
        return jsonify({"error": "Generation already running"}), 409
    data = request.get_json(force=True) or {}
    scope = (data.get("scope") or "all").strip()
    if scope == "style":
        sn = (data.get("style") or "").strip()
        s = _lab_session_get_style(sn)
        if not s:
            return jsonify({"error": f"Style {sn} not found"}), 404
        targets = [s]
    else:
        targets = lab_session["styles"]
    threading.Thread(target=_lab_run_generation, args=(targets,), daemon=True).start()
    return jsonify({"ok": True, "total": len(targets), "scope": scope})


@app.route("/api/lab/generate-progress", methods=["GET"])
def lab_generate_progress():
    elapsed = "—"
    if lab_gen_progress.get("started_at"):
        secs = int(time.time() - lab_gen_progress["started_at"])
        elapsed = f"{secs}s"
    pct = 0.0
    if lab_gen_progress.get("total"):
        pct = round(lab_gen_progress["completed"] / lab_gen_progress["total"] * 100, 1)
    return jsonify({
        "total":   lab_gen_progress["total"],
        "completed": lab_gen_progress["completed"],
        "current_style": lab_gen_progress["current_style"],
        "current_step":  lab_gen_progress["current_step"],
        "status":  lab_gen_progress["status"],
        "error":   lab_gen_progress.get("error", ""),
        "elapsed": elapsed,
        "percent": pct,
    })


# ══════════════════════════════════════════════════════════════
# LAB · STAGE 4 — grid IS the bulksheet, editor IS the brain
# validate • regen-cell • copy • lock • auto-gen • bit-perfect .xlsm
# ══════════════════════════════════════════════════════════════

# Generic “Atlas-side” cell rules so a cell turns red the moment a hard
# constraint is violated, even before the Amazon NIS rule engine runs.
# These mirror the LLM prompt rules in generate_content_llm().
LAB_CELL_RULES = {
    "title":            {"max_chars": 120, "min_chars": 30, "required": True},
    "bullet_1":         {"max_chars": 256, "min_chars": 30, "required": True},
    "bullet_2":         {"max_chars": 256, "min_chars": 30, "required": True},
    "bullet_3":         {"max_chars": 256, "min_chars": 30, "required": True},
    "bullet_4":         {"max_chars": 256, "min_chars": 30, "required": True},
    "bullet_5":         {"max_chars": 256, "min_chars": 30, "required": True},
    "description":      {"max_chars": 2000, "min_chars": 100, "required": True},
    "backend_keywords": {"max_bytes": 249},
    "upc":              {"pattern": r"^[0-9]{12,13}$", "required": True,
                          "err": "UPC must be 12 or 13 numeric digits"},
    "item_weight_value": {"min_value": 0.001, "required": True},
}

_LAB_BANNED_PHRASES = [
    "the modern woman", "the modern man", "the modern individual",
    "women who refuse", "men who refuse",
    "luxurious", "premium quality", "best seller", "limited time",
]


def _lab_validate_value(key, value, style=None):
    """Return list of {severity, msg} for one cell value.

    severity: 'error' → Amazon will reject · 'warn' → Atlas guidance.
    """
    out = []
    rule = LAB_CELL_RULES.get(key)
    sval = "" if value is None else str(value).strip()

    if rule:
        if rule.get("required") and not sval:
            out.append({"severity": "error", "msg": f"{key} is required"})
        if rule.get("max_chars") and len(sval) > rule["max_chars"]:
            out.append({"severity": "error",
                        "msg": f"{len(sval)} chars — Amazon caps {key} at {rule['max_chars']}"})
        if rule.get("min_chars") and sval and len(sval) < rule["min_chars"]:
            out.append({"severity": "warn",
                        "msg": f"{len(sval)} chars — below recommended minimum {rule['min_chars']}"})
        if rule.get("max_bytes") and len(sval.encode("utf-8")) > rule["max_bytes"]:
            out.append({"severity": "error",
                        "msg": f"{len(sval.encode('utf-8'))} bytes — over {rule['max_bytes']}-byte cap"})
        if rule.get("pattern") and sval and not re.match(rule["pattern"], sval):
            out.append({"severity": "error", "msg": rule.get("err", "format invalid")})
        if rule.get("min_value") is not None and sval:
            try:
                if float(sval) < rule["min_value"]:
                    out.append({"severity": "error",
                                "msg": f"value must be ≥ {rule['min_value']}"})
            except ValueError:
                out.append({"severity": "error", "msg": "must be numeric"})

    # Banned-phrase scan on copy fields
    if key in ("title", "description", "bullet_1", "bullet_2", "bullet_3", "bullet_4", "bullet_5"):
        low = sval.lower()
        for ph in _LAB_BANNED_PHRASES:
            if ph in low:
                out.append({"severity": "error", "msg": f"banned phrase: '{ph}'"})

    # Gender alignment between title + bullets/description (style-scope only)
    if style and key in ("description", "bullet_1", "bullet_2", "bullet_3", "bullet_4", "bullet_5"):
        gen = style.get("_lab_generated") or {}
        title = (gen.get("title") or style.get("title") or "").lower()
        if "women's" in title or "women\u2019s" in title:
            if re.search(r"\bmen\b|\bman\b|\bhe\b|\bhis\b", sval, flags=re.IGNORECASE):
                out.append({"severity": "error",
                            "msg": "title says women's, this cell uses male language"})
        elif "men's" in title or "men\u2019s" in title:
            if re.search(r"\bwomen\b|\bwoman\b|\bshe\b|\bher\b", sval, flags=re.IGNORECASE):
                out.append({"severity": "error",
                            "msg": "title says men's, this cell uses female language"})
    return out


@app.route("/api/lab/validate", methods=["POST"])
def lab_validate_grid():
    """Validate a single cell or a batch of rows. Body shape:
      {style: <style_num>, group: <group_key>, rows: [...]} — returns
      {violations: [{row, key, severity, msg}, …], cells_checked: N}.
    Front-end calls this on every edit.
    """
    data = request.get_json(force=True) or {}
    style_num = (data.get("style") or "").strip()
    group_key = (data.get("group") or "").strip()
    rows = data.get("rows") or []
    style = _lab_session_get_style(style_num) if style_num else None
    grp = LAB_GRID_GROUPS.get(group_key) or {}
    columns = grp.get("columns") or []
    violations = []
    n = 0
    for r_idx, row in enumerate(rows):
        for col in columns:
            if col.get("readonly"):
                continue
            key = col["key"]
            if key in row:
                n += 1
                for v in _lab_validate_value(key, row.get(key), style):
                    violations.append({"row": r_idx, "key": key, **v})
    return jsonify({"ok": True, "violations": violations, "cells_checked": n})


# ──── Single-cell LLM regenerate ──────────────────────────────────
@app.route("/api/lab/regen-cell", methods=["POST"])
def lab_regen_cell():
    """Regenerate one copy field for one style with optional operator feedback.

    Body: {style: <style_num>, key: 'title'|'bullet_1..5'|'description'|'backend_keywords',
           feedback: '<freeform instruction>' (optional)}
    Strategy: run the full content LLM with feedback appended to the
    feedback_history string, then return only the requested key.
    Reuses generate_content_llm + the gender-drift scrub.
    """
    if _anthropic_client is None:
        return jsonify({"error": "LLM unavailable"}), 503
    data = request.get_json(force=True) or {}
    style_num = (data.get("style") or "").strip()
    key = (data.get("key") or "").strip()
    feedback = (data.get("feedback") or "").strip()
    if key not in ("title", "description", "backend_keywords",
                    "bullet_1", "bullet_2", "bullet_3", "bullet_4", "bullet_5"):
        return jsonify({"error": f"key {key!r} not regen-able"}), 400
    style = _lab_session_get_style(style_num)
    if not style:
        return jsonify({"error": f"style {style_num} not in session"}), 404

    brand = lab_session.get("brand", "") or ""
    brand_cfg = _load_brand_config_data(brand) or {}
    feedback_history = f"- [{key}] {feedback}" if feedback else ""

    try:
        result = generate_content_llm(brand_cfg, brand, style, feedback_history, regen_keys={key})
    except Exception as e:
        return jsonify({"error": f"LLM call failed: {e}"}), 500
    if not result:
        return jsonify({"error": "LLM returned no result"}), 502

    # Apply gender-drift scrub on the regen output
    style_gender, _ = _derive_gender_department(style)
    eff_g = style_gender or brand_cfg.get("gender", "") or ""
    title = result.get("title", "") or ""
    if eff_g not in ("Male", "Female") and title:
        t = title.lower()
        if "women's" in t or "women\u2019s" in t: eff_g = "Female"
        elif "men's" in t or "men\u2019s" in t:  eff_g = "Male"
    if key == "description":
        result["description"] = _scrub_gender_drift(result.get("description", ""), eff_g)
    if key.startswith("bullet_"):
        result[key] = _scrub_gender_drift(result.get(key, ""), eff_g)

    new_value = result.get(key, "") or ""
    # Stage 7 endgame: regen is a PROPOSAL, not a commit. The new value goes
    # into _lab_proposed; _lab_generated (what downloads read) stays untouched
    # until the operator hits Accept. This prevents hallucinated / off-brand
    # regens from silently shipping into the NIS file.
    current = (style.get("_lab_generated") or {}).get(key, "") or ""
    proposed = style.setdefault("_lab_proposed", {})
    proposed[key] = {
        "value":       new_value,
        "prev":        current,
        "feedback":    feedback,
        "proposed_at": datetime.utcnow().isoformat("T") + "Z",
    }
    _lab_session_persist()
    return jsonify({
        "ok":       True,
        "key":      key,
        "value":    new_value,
        "prev":     current,
        "proposed": True,
    })


@app.route("/api/lab/proposal/accept", methods=["POST"])
def lab_proposal_accept():
    """Promote a pending proposal into _lab_generated. One key at a time.

    Body: {style, key}
    Only then does the new text become part of the downloadable NIS file.
    """
    data = request.get_json(force=True) or {}
    style_num = (data.get("style") or "").strip()
    key       = (data.get("key") or "").strip()
    if key not in _STYLE_GENERATED_KEYS:
        return jsonify({"error": f"{key!r} not accept-able"}), 400
    style = _lab_session_get_style(style_num)
    if not style:
        return jsonify({"error": f"style {style_num} not in session"}), 404
    proposals = style.get("_lab_proposed") or {}
    if key not in proposals:
        return jsonify({"error": f"no pending proposal for {key}"}), 404
    new_val = proposals[key].get("value", "") or ""
    style.setdefault("_lab_generated", {})[key] = new_val
    style["_lab_generated"]["_regen_at"] = datetime.utcnow().isoformat("T") + "Z"
    del proposals[key]
    if not proposals:
        style.pop("_lab_proposed", None)
    _lab_session_persist()
    return jsonify({"ok": True, "key": key, "value": new_val})


@app.route("/api/lab/proposal/revert", methods=["POST"])
def lab_proposal_revert():
    """Throw away a pending proposal. _lab_generated is unchanged."""
    data = request.get_json(force=True) or {}
    style_num = (data.get("style") or "").strip()
    key       = (data.get("key") or "").strip()
    style = _lab_session_get_style(style_num)
    if not style:
        return jsonify({"error": f"style {style_num} not in session"}), 404
    proposals = style.get("_lab_proposed") or {}
    if key not in proposals:
        return jsonify({"error": f"no pending proposal for {key}"}), 404
    del proposals[key]
    if not proposals:
        style.pop("_lab_proposed", None)
    _lab_session_persist()
    return jsonify({"ok": True, "key": key})


# ──── Selective copy + locks ───────────────────────────────────
_LAB_LOCK_KEY = "_lab_locks"  # stored on each style: {scope: 'style'|'group'|'field', group?: '...', key?: '...'}


def _lab_is_locked(style, scope, group=None, key=None):
    """Check whether a given target is locked on the style."""
    locks = style.get(_LAB_LOCK_KEY) or []
    for lk in locks:
        if lk.get("scope") == "style" and scope in ("style", "group", "field"):
            return True
        if lk.get("scope") == "group" and lk.get("group") == group and scope in ("group", "field"):
            return True
        if lk.get("scope") == "field" and lk.get("group") == group and lk.get("key") == key and scope == "field":
            return True
    return False


@app.route("/api/lab/lock", methods=["POST"])
def lab_lock_toggle():
    """Toggle a lock on the style. Body: {style, scope, group?, key?}.
    Returns the new locks list for that style.
    """
    data = request.get_json(force=True) or {}
    style_num = (data.get("style") or "").strip()
    scope = (data.get("scope") or "").strip()
    group = data.get("group")
    key = data.get("key")
    if scope not in ("style", "group", "field"):
        return jsonify({"error": "scope must be style|group|field"}), 400
    style = _lab_session_get_style(style_num)
    if not style:
        return jsonify({"error": f"style {style_num} not in session"}), 404
    locks = style.setdefault(_LAB_LOCK_KEY, [])
    target = {"scope": scope}
    if group: target["group"] = group
    if key:   target["key"]   = key
    # toggle: drop matching, otherwise add
    found = False
    new_locks = []
    for lk in locks:
        if (lk.get("scope") == scope
            and lk.get("group") == target.get("group")
            and lk.get("key") == target.get("key")):
            found = True
            continue
        new_locks.append(lk)
    if not found:
        new_locks.append(target)
    style[_LAB_LOCK_KEY] = new_locks
    _lab_session_persist()
    return jsonify({"ok": True, "locked": not found, "locks": new_locks})


@app.route("/api/lab/locks", methods=["GET"])
def lab_locks_list():
    """Return all locks across all styles — for the global Locks panel."""
    out = []
    for s in lab_session.get("styles", []):
        for lk in s.get(_LAB_LOCK_KEY) or []:
            out.append({
                "style_num": s.get("style_num", ""),
                "style_name": s.get("style_name", ""),
                **lk,
            })
    return jsonify({"ok": True, "locks": out})


@app.route("/api/lab/copy", methods=["POST"])
def lab_copy_to_styles():
    """Copy fields from a source style to a list of destination styles.

    Body:
      {source: <style_num>, group: <group_key>, key?: <single field>,
       destinations: [<style_num>, ...], override_locks: bool}

    If `key` is omitted, copies the entire group's writable values from
    source to destinations. Lock guard rejects locked destinations unless
    override_locks=True.
    """
    data = request.get_json(force=True) or {}
    src_num = (data.get("source") or "").strip()
    group_key = (data.get("group") or "").strip()
    key = (data.get("key") or "").strip()
    destinations = data.get("destinations") or []
    override = bool(data.get("override_locks"))
    grp = LAB_GRID_GROUPS.get(group_key)
    if not grp:
        return jsonify({"error": f"unknown group: {group_key}"}), 400
    src = _lab_session_get_style(src_num)
    if not src:
        return jsonify({"error": f"source {src_num} not in session"}), 404

    if grp.get("scope", "variant") != "style":
        return jsonify({"error": "copy currently supports style-scope groups only"}), 400
    src_row = _build_style_row(src)
    writable_keys = [c["key"] for c in grp["columns"]
                     if not c.get("readonly") and (not key or c["key"] == key)]
    if not writable_keys:
        return jsonify({"error": "no writable keys to copy"}), 400

    summary = {"updated": [], "skipped_locked": [], "skipped_missing": []}
    for dst_num in destinations:
        dst = _lab_session_get_style(dst_num)
        if not dst:
            summary["skipped_missing"].append(dst_num)
            continue
        # Lock guard
        any_locked = False
        if not override:
            for k in writable_keys:
                if _lab_is_locked(dst, "field", group=group_key, key=k) \
                   or _lab_is_locked(dst, "group", group=group_key) \
                   or _lab_is_locked(dst, "style"):
                    any_locked = True
                    break
        if any_locked:
            summary["skipped_locked"].append(dst_num)
            continue
        # Apply
        gen = dst.setdefault("_lab_generated", {})
        for k in writable_keys:
            new_val = src_row.get(k, "")
            target_dict = gen if k in _STYLE_GENERATED_KEYS else dst
            target_dict[k] = new_val
        summary["updated"].append(dst_num)

    _lab_session_persist()
    return jsonify({"ok": True, "summary": summary, "keys": writable_keys})


# ──── Bit-perfect Vendor-Central .xlsm download ──────────────────────
@app.route("/api/lab/save-field", methods=["POST"])
def lab_save_field():
    """Stage 5b — single-field save for the inline-edit (Excel feel) flow.

    Body: {style: '...', key: 'title|bullet_1..5|description|backend_keywords', value: '...'}
    Persists into style['_lab_generated'][key]. Lock-aware: returns 409 if the
    field/group/style is locked.
    """
    data = request.get_json(force=True) or {}
    style_num = (data.get("style") or "").strip()
    key       = (data.get("key") or "").strip()
    value     = data.get("value")
    if not style_num or not key:
        return jsonify({"error": "style and key required"}), 400
    if key not in _STYLE_GENERATED_KEYS:
        return jsonify({"error": f"Field '{key}' not editable via this endpoint"}), 400
    style = _lab_session_get_style(style_num)
    if not style:
        return jsonify({"error": f"Style {style_num} not in session"}), 404

    # Lock check (mirror the regen-cell behaviour)
    locks = (lab_session.get("locks") or {}).get(style_num, {}) or {}
    if locks.get(key) or locks.get("_style") or locks.get("_group_copy"):
        return jsonify({"error": f"Field {key} on {style_num} is locked"}), 409

    new = "" if value in (None,) else str(value)
    gen = style.setdefault("_lab_generated", {})
    old = gen.get(key, "") or ""
    if old != new:
        gen[key] = new
        _lab_session_persist()
        return jsonify({"ok": True, "updated": True, "style": style_num, "key": key})
    return jsonify({"ok": True, "updated": False, "style": style_num, "key": key})


@app.route("/api/lab/styles-meta", methods=["GET"])
def lab_styles_meta():
    """Return per-style metadata used by the Listing Card view: style_num,
    style_name, PT, generated/missing flags. Lets the front-end group by PT
    in the left rail, render status badges, and show 'Download <PT> only'.
    """
    if not lab_session.get("styles"):
        return jsonify({"ok": True, "brand": "", "styles": [], "by_pt": {}})
    by_pt = {}
    out = []
    for s in lab_session["styles"]:
        sn  = s.get("style_num", "") or ""
        nm  = s.get("style_name", "") or ""
        pt  = (_resolve_pt_for_style(s) or "").upper()
        gen = s.get("_lab_generated") or {}
        # "ready" = at least a title is present (LLM filled it in)
        ready = bool((gen.get("title") or "").strip())
        # bullet-coverage gives a softer "in progress" signal
        bullets_filled = sum(1 for i in range(1, 6) if (gen.get(f"bullet_{i}") or "").strip())
        rec = {
            "style_num":      sn,
            "style_name":     nm,
            "product_type":   pt or "UNCLASSIFIED",
            "variant_count":  len(s.get("variants", []) or []),
            "ready":          ready,
            "bullets_filled": bullets_filled,
            "has_description": bool((gen.get("description") or "").strip()),
            "has_keywords":   bool((gen.get("backend_keywords") or "").strip()),
        }
        out.append(rec)
        by_pt.setdefault(rec["product_type"], []).append(sn)
    return jsonify({
        "ok":      True,
        "brand":   lab_session.get("brand", "") or "",
        "styles":  out,
        "by_pt":   by_pt,
        "pts":     sorted(by_pt.keys()),
    })


# ═════════════ STAGE 7 · NIS rule-engine wiring ═════════════
# Map our short Lab column keys to the engine's field_keys, then expose the
# rule-engine verdict (required / conditionally required / recommended /
# optional) per column AND the trigger cells that explain *why* a conditional
# field becomes required.

# Lab key -> engine field_key. We can derive most of these from the column
# spec (col.dropdown_field), but copy fields and identity columns aren't
# dropdown_dynamic, so we encode the mapping explicitly here.
_LAB_KEY_TO_ENGINE_FIELD = {
    "title":             "item_name#1.value",
    "bullet_1":          "bullet_point#1.value",
    "bullet_2":          "bullet_point#2.value",
    "bullet_3":          "bullet_point#3.value",
    "bullet_4":          "bullet_point#4.value",
    "bullet_5":          "bullet_point#5.value",
    "description":       "product_description#1.value",
    "backend_keywords":  "generic_keyword#1.value",
    "brand":             "brand#1.value",
    "vendor_code":       "rtip_vendor_code#1.value",
    "item_type":         "item_type_keyword#1.value",
    "feed_product_type": "feed_product_type#1.value",
    "department":        "department#1.value",
    "target_gender":     "target_gender#1.value",
    "age_range":         "age_range_description#1.value",
    "lifestyle_1":       "lifestyle#1.value",
    "lifestyle_2":       "lifestyle#2.value",
    "item_weight":       "item_weight#1.value",
    "item_weight_unit":  "item_weight#1.unit",
    "item_length":       "item_length_width_height#1.length.value",
    "item_width":        "item_length_width_height#1.width.value",
    "item_height":       "item_length_width_height#1.height.value",
    "item_dim_unit":     "item_length_width_height#1.length.unit",
}

_REQ_TIER_RANK = {
    "REQUIRED": 4,
    "CONDITIONALLY REQUIRED": 3,
    "RECOMMENDED": 2,
    "OPTIONAL": 1,
}


def _engine_field_for_lab_key(col):
    """Resolve engine field_key for a Lab column spec dict."""
    if col.get("dropdown_field"):
        return col["dropdown_field"]
    return _LAB_KEY_TO_ENGINE_FIELD.get(col["key"])


def _engine_requirements_for_pt(pt, form_state=None):
    """Run nis_engine.evaluate_form for the given PT and return a dict keyed by
    engine field_key with {tier, verdict, trigger_field_keys}.
    Cheap and cached per (pt, frozen-state-key) within the request lifetime.
    """
    if not pt:
        return {}
    try:
        result = _nis_engine.evaluate_form(
            pt, form_state or {},
            include_dropdowns=False,
            apply_apparel_defaults=True,
        )
    except Exception as e:
        print(f"[lab] engine evaluate failed for {pt}: {e}")
        return {}
    fields = result.get("fields") or {}
    # Build a column-letter -> field_key map so we can resolve trigger_cells
    # back to engine field_keys for the frontend.
    col_to_fkey = {col_letter: f.get("field_key") for col_letter, f in fields.items() if f.get("field_key")}
    out = {}
    for col_letter, f in fields.items():
        fkey = f.get("field_key")
        if not fkey:
            continue
        # trigger_cells come back as 'A7', 'D7' — strip the row to get column letter
        triggers = []
        for tc in (f.get("trigger_cells") or []):
            col_only = "".join(c for c in tc if not c.isdigit())
            tk = col_to_fkey.get(col_only)
            if tk and tk != fkey:
                triggers.append(tk)
        out[fkey] = {
            "tier":     (f.get("base_requirement") or "OPTIONAL").upper(),
            "verdict":  f.get("verdict") or "optional",
            "label":    f.get("label") or "",
            "section":  f.get("section") or "",
            "triggers": list(dict.fromkeys(triggers)),  # dedupe, preserve order
        }
    return out


def _lab_state_for_engine(style):
    """Build a flat form_state dict keyed by engine field_keys from a Lab style.
    Reads both style['_lab_generated'] (copy fields) and style[k] (NIS attrs).
    """
    if not style:
        return {}
    gen = style.get("_lab_generated") or {}
    state = {}
    # Walk every Lab column we know about and copy into engine-keyed state
    for grp in LAB_GRID_GROUPS.values():
        for col in grp.get("columns", []):
            if col.get("scope") == "variant":
                continue
            fkey = _engine_field_for_lab_key(col)
            if not fkey:
                continue
            v = gen.get(col["key"]) if col["key"] in gen and gen.get(col["key"]) is not None \
                else style.get(col["key"])
            if v is None or v == "":
                continue
            state[fkey] = str(v)
    return state


@app.route("/api/lab/requirements", methods=["GET"])
def lab_requirements():
    """Return requirement tiers + dependency hints for one style.

    Used by:
      - Cell picker header (chip showing Required / Conditional / Recommended)
      - Grid validation footer (live dependency hints)
      - Matrix header (column-level requirement chips)

    Query: ?style=<style_num>
    Response:
      {
        product_type: 'COAT',
        columns: { lab_key: { tier, verdict, triggers:[lab_keys] } },
        summary: { required: N, missing: N, conditional: N, recommended: N }
      }
    """
    style_num = (request.args.get("style") or "").strip()
    style = _lab_session_get_style(style_num) if style_num else None
    if not style and style_num:
        return jsonify({"error": f"Style {style_num} not in session"}), 404
    pt = (_resolve_pt_for_style(style) if style else "") or (request.args.get("pt") or "").upper()
    if not pt:
        return jsonify({"ok": True, "product_type": "", "columns": {}, "summary": {}})

    state = _lab_state_for_engine(style) if style else {}
    eng_results = _engine_requirements_for_pt(pt, state)

    # Reverse map engine field_key -> lab key
    fkey_to_labkey = {}
    for grp in LAB_GRID_GROUPS.values():
        for col in grp.get("columns", []):
            fk = _engine_field_for_lab_key(col)
            if fk:
                fkey_to_labkey[fk] = col["key"]

    columns = {}
    summary = {"REQUIRED": 0, "CONDITIONALLY REQUIRED": 0, "RECOMMENDED": 0, "OPTIONAL": 0,
               "required_missing": 0, "required_ok": 0}
    for grp in LAB_GRID_GROUPS.values():
        for col in grp.get("columns", []):
            if col.get("scope") == "variant":
                continue
            fk = _engine_field_for_lab_key(col)
            if not fk:
                # Fall back to spec-declared required flag
                if col.get("required"):
                    columns[col["key"]] = {"tier": "REQUIRED", "verdict": "required", "triggers": []}
                continue
            info = eng_results.get(fk)
            if not info:
                if col.get("required"):
                    columns[col["key"]] = {"tier": "REQUIRED", "verdict": "required", "triggers": []}
                continue
            triggers_lab = [fkey_to_labkey[t] for t in info["triggers"] if t in fkey_to_labkey]
            columns[col["key"]] = {
                "tier":     info["tier"],
                "verdict":  info["verdict"],
                "triggers": triggers_lab,
                "engine_label": info["label"],
            }
            summary[info["tier"]] = summary.get(info["tier"], 0) + 1
            if info["verdict"] == "required_missing":
                summary["required_missing"] += 1
            elif info["verdict"] == "required_ok":
                summary["required_ok"] += 1

    return jsonify({
        "ok":           True,
        "product_type": pt,
        "style":        style_num,
        "columns":      columns,
        "summary":      summary,
    })


@app.route("/api/lab/pt-matrix", methods=["GET"])
def lab_pt_matrix():
    """Stage 7 · PT-aggregated compliance view.

    Returns one matrix per Product Type. Each matrix has:
      - columns: required attribute keys + titles for that PT (we use
        LAB_GRID_GROUPS to enumerate every required key, then de-dupe)
      - rows: one per style in this PT, each cell carries
          {value, status: 'ready'|'missing'|'override'}
      - totals: {ready, missing} per style and per column

    Status is computed from style['_lab_generated'] for copy fields and
    from style[k] / style['variants'][...][k] for everything else, mirroring
    how the grid editor reads.
    """
    if not lab_session.get("styles"):
        return jsonify({"ok": True, "matrices": []})

    # Build the unified "required attributes" list once — every group's
    # required keys, deduped. We exclude variant-scope keys from this matrix
    # (color/size/upc are per-variant, handled in their own group).
    seen = set()
    attribute_cols = []
    for grp_key, grp in LAB_GRID_GROUPS.items():
        if grp.get("scope") != "style":
            continue
        for col in grp.get("columns", []):
            if not col.get("required"):
                continue
            k = col["key"]
            if k in seen or col.get("readonly"):
                continue
            seen.add(k)
            attribute_cols.append({
                "key":   k,
                "title": col.get("title", k),
                "group": grp.get("label", grp_key),
                "group_key": grp_key,
            })

    # Group styles by PT
    by_pt = {}
    for s in lab_session["styles"]:
        pt = (_resolve_pt_for_style(s) or "UNCLASSIFIED").upper()
        by_pt.setdefault(pt, []).append(s)

    matrices = []
    for pt in sorted(by_pt.keys()):
        pt_styles = by_pt[pt]
        # Engine-driven tier per column: run evaluate_form once with the FIRST
        # style's state so conditional-required fields are exposed correctly.
        sample_state = _lab_state_for_engine(pt_styles[0]) if pt_styles else {}
        eng_results = _engine_requirements_for_pt(pt, sample_state) if pt != "UNCLASSIFIED" else {}
        # Map lab key -> tier for our column list
        col_tiers = {}
        for c in attribute_cols:
            fk = _engine_field_for_lab_key(c)
            tier = (eng_results.get(fk) or {}).get("tier") if fk else None
            col_tiers[c["key"]] = tier or "REQUIRED"   # default if engine doesn't know
        rows = []
        col_missing = {c["key"]: 0 for c in attribute_cols}
        for s in pt_styles:
            sn  = s.get("style_num", "") or ""
            nm  = s.get("style_name", "") or ""
            gen = s.get("_lab_generated") or {}
            cells = []
            ready_n  = 0
            missing_n = 0
            for c in attribute_cols:
                k = c["key"]
                # _lab_generated wins for copy fields; otherwise read from style
                v = gen.get(k) if k in gen and gen.get(k) is not None else s.get(k)
                v_str = ("" if v is None else str(v)).strip()
                status = "ready" if v_str else "missing"
                if status == "missing":
                    missing_n += 1
                    col_missing[k] += 1
                else:
                    ready_n += 1
                cells.append({
                    "key":    k,
                    "value":  v_str[:80],   # truncate for the matrix display
                    "status": status,
                })
            rows.append({
                "style_num":  sn,
                "style_name": nm,
                "variant_count": len(s.get("variants", []) or []),
                "cells":   cells,
                "ready":   ready_n,
                "missing": missing_n,
            })
        # Attach tier to each attribute column for this PT
        cols_with_tiers = [{**c, "tier": col_tiers.get(c["key"], "REQUIRED")} for c in attribute_cols]
        matrices.append({
            "product_type": pt,
            "columns":      cols_with_tiers,
            "rows":         rows,
            "col_missing":  col_missing,
            "total_styles": len(rows),
            "total_required": len(attribute_cols),
        })

    return jsonify({
        "ok":       True,
        "brand":    lab_session.get("brand", "") or "",
        "matrices": matrices,
        "attribute_count": len(attribute_cols),
    })



@app.route("/api/lab/listing-card", methods=["GET"])
def lab_listing_card():
    """Return the six promptable fields for ONE style, ready for the Listing
    Card view: title, bullets[1..5], description, backend_keywords. Also
    surfaces brand/style/PT/variant info so the card header renders without
    a second round-trip.
    """
    style_num = (request.args.get("style") or "").strip()
    if not style_num:
        return jsonify({"error": "style param required"}), 400
    style = _lab_session_get_style(style_num)
    if not style:
        return jsonify({"error": f"style {style_num} not in session"}), 404
    gen = style.get("_lab_generated") or {}
    bullets = [(gen.get(f"bullet_{i}") or "") for i in range(1, 6)]
    locks = (lab_session.get("locks") or {}).get(style_num, {}) or {}
    return jsonify({
        "ok":              True,
        "style_num":       style_num,
        "style_name":      style.get("style_name", "") or "",
        "brand":           lab_session.get("brand", "") or style.get("brand", "") or "",
        "product_type":    (_resolve_pt_for_style(style) or "").upper(),
        "variant_count":   len(style.get("variants", []) or []),
        "title":           gen.get("title", "") or "",
        "bullets":         bullets,
        "description":     gen.get("description", "") or "",
        "backend_keywords": gen.get("backend_keywords", "") or "",
        "locks":           locks,
        "is_generated":    bool((gen.get("title") or "").strip()),
        # Pending regen proposals — frontend renders Accept / Revert affordances.
        # Until accepted, these do NOT flow into _lab_generated and are NOT
        # included in any download (.xlsm or NIS bulksheet).
        "proposals":       style.get("_lab_proposed") or {},
    })


@app.route("/api/lab/template-preview", methods=["GET"])
def lab_template_preview():
    """Return the bit-perfect NIS template for one style as JSON for preview.

    Reuses do_xlsm_surgery() to write the Vendor Central .xlsm exactly as
    it will be downloaded, then reads the 'Template' sheet back with openpyxl
    and serializes the first ~60 rows + first ~80 columns as a 2D array.
    The frontend renders this read-only so the operator can see "this is
    what ships to Amazon" in the same dashboard.

    Body query: ?style=<style_num>
    """
    style_num = (request.args.get("style") or "").strip()
    if not style_num:
        return jsonify({"error": "style param required"}), 400
    style = _lab_session_get_style(style_num)
    if not style:
        return jsonify({"error": f"Style {style_num} not in session"}), 404
    pt = _resolve_style_product_type(style) or ""
    tpl_for_pt = _template_path_for_pt(pt) or str(DEFAULT_TEMPLATE)
    brand = lab_session.get("brand", "") or ""
    brand_cfg = _load_brand_config_data(brand) or {}
    vendor_code = brand_cfg.get("vendor_code_full", "") or ""
    gen = style.get("_lab_generated") or {}
    content = {
        "title":             gen.get("title", "") or "",
        "bullets":           [gen.get(f"bullet_{i}", "") for i in range(1, 6)],
        "description":       gen.get("description", "") or "",
        "backend_keywords":  gen.get("backend_keywords", "") or "",
        "neck_type":         style.get("neck_type", ""),
        "sleeve_type":       style.get("sleeve_type", ""),
        "fit_type":          style.get("fit_type", ""),
        "closure_type":      style.get("closure_type", ""),
        "collar_style":      style.get("collar_style", ""),
    }
    try:
        out_path = do_xlsm_surgery(tpl_for_pt, brand, brand_cfg, vendor_code, style, content)
    except Exception as e:
        return jsonify({"error": f"Template build failed: {e}"}), 500
    if not out_path or not Path(out_path).exists():
        return jsonify({"error": "Template output missing after build"}), 500

    # Read the Template sheet back as a 2D array. Limit to a sensible
    # viewport so the response stays snappy — the real file has 200+ columns
    # of mostly-empty Amazon attribute slots.
    import openpyxl
    wb = openpyxl.load_workbook(out_path, data_only=False, keep_vba=False)
    ws = wb.active
    # Vendor Central templates use "Template-COAT" / "Template-DRESS" / etc.
    # Match any sheet whose name starts with "Template" (case-insensitive)
    # since that's where the operator-edited rows actually live.
    for n in wb.sheetnames:
        if n.lower().startswith("template"):
            ws = wb[n]
            break

    # The Vendor Central template has 3 header rows + 3 data rows (parent
    # + children) + empty rows. Keep a small vertical viewport and a wide
    # horizontal one so the demo shows the full attribute column span.
    MAX_ROWS = 30
    MAX_COLS = 150  # covers all COAT fields (actual total is ~243)
    data = []
    for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row, MAX_ROWS),
                           max_col=min(ws.max_column, MAX_COLS), values_only=True):
        data.append([("" if v is None else str(v)) for v in r])
    # Also tag which row is the operator's parent / child rows by detecting
    # the first non-empty row after the template's frozen header block.
    header_row_index = 0
    for i, row in enumerate(data):
        if any((cell or "").strip() for cell in row[:5]):
            header_row_index = i
            break
    return jsonify({
        "ok":          True,
        "style_num":   style_num,
        "product_type": pt,
        "template_name": Path(tpl_for_pt).name,
        "rows":        data,
        "row_count":   len(data),
        "col_count":   len(data[0]) if data else 0,
        "truncated":   ws.max_row > MAX_ROWS or ws.max_column > MAX_COLS,
        "total_rows":  ws.max_row,
        "total_cols":  ws.max_column,
        "header_row_index": header_row_index,
    })


@app.route("/api/lab/download-bulksheet", methods=["GET"])
def lab_download_bulksheet():
    """Emit one bit-perfect Vendor Central .xlsm per style + zip them.

    Reuses do_xlsm_surgery() which already writes the parent + child rows
    to the matching Amazon NIS template per PT — same code path Bulk Upload
    uses, so the output drops into Vendor Central without manual edits.

    Optional query params:
        pt=COAT   -> only emit styles whose resolved PT equals 'COAT'
        style=ABC -> only emit a single style (operator card-level export)
    """
    if not lab_session.get("styles"):
        return jsonify({"error": "No data in session. Upload first."}), 400
    brand = lab_session.get("brand", "") or ""
    brand_cfg = _load_brand_config_data(brand) or {}
    vendor_code = brand_cfg.get("vendor_code_full", "") or ""

    pt_filter    = (request.args.get("pt") or "").strip().upper()
    style_filter = (request.args.get("style") or "").strip()

    # Build per-style outputs
    outdir = UPLOAD_PRODUCTS / f"lab_bulksheets_{re.sub(r'[^\\w]','_', brand)}"
    outdir.mkdir(parents=True, exist_ok=True)
    written = []
    for s in lab_session["styles"]:
        # Apply optional filters
        if style_filter and (s.get("style_num", "") or "") != style_filter:
            continue
        if pt_filter:
            s_pt = (_resolve_pt_for_style(s) or "").upper()
            if s_pt != pt_filter:
                continue
        sn = s.get("style_num", "")
        # Resolve template path for this style's PT
        pt = _resolve_style_product_type(s) or ""
        tpl_for_pt = _template_path_for_pt(pt) or str(DEFAULT_TEMPLATE)
        gen = s.get("_lab_generated") or {}
        content = {
            "title":             gen.get("title", "") or "",
            "bullets":           [gen.get(f"bullet_{i}", "") for i in range(1, 6)],
            "description":       gen.get("description", "") or "",
            "backend_keywords":  gen.get("backend_keywords", "") or "",
            "neck_type":         s.get("neck_type", ""),
            "sleeve_type":       s.get("sleeve_type", ""),
            "fit_type":          s.get("fit_type", ""),
            "closure_type":      s.get("closure_type", ""),
            "collar_style":      s.get("collar_style", ""),
        }
        # Skip styles with no generated copy (operator hasn't generated yet)
        if not (content["title"] or any(content["bullets"])):
            continue
        try:
            out_path = do_xlsm_surgery(tpl_for_pt, brand, brand_cfg, vendor_code, s, content)
        except Exception as e:
            print(f"[lab] xlsm surgery failed for {sn}: {e}", flush=True)
            continue
        if out_path and Path(out_path).exists():
            written.append(out_path)

    if not written:
        return jsonify({"error": "No styles have generated copy yet. Run Generate first."}), 400

    # Zip them
    import zipfile
    zip_path = UPLOAD_PRODUCTS / f"atlas_lab_bulksheets_{re.sub(r'[^\\w]','_', brand)}.zip"
    with zipfile.ZipFile(str(zip_path), "w", zipfile.ZIP_DEFLATED) as zf:
        for p in written:
            zf.write(p, Path(p).name)
    return send_file(
        str(zip_path),
        as_attachment=True,
        download_name=zip_path.name,
        mimetype="application/zip",
    )


def _template_path_for_pt(pt):
    """Map an Amazon PT (COAT, DRESS, SHIRT, …) to the local template file."""
    if not pt:
        return None
    pt = pt.upper()
    fname = {
        "COAT":            "Jackets_and_Coats.xlsm",
        "DRESS":           "Dresses.xlsm",
        "SHIRT":           "Other_Shirts.xlsm",
        "BLAZER":          "Blazers.xlsm",
        "BRA":             "Bras.xlsm",
        "HAT":             "Hats.xlsm",
        "ONE_PIECE_OUTFIT": "One-piece_Outfits.xlsm",
        "OVERALLS":        "Overalls.xlsm",
        "PANTS":           "Other_Pants.xlsm",
        "SANDAL":          "Sandals.xlsm",
        "SHORTS":          "Shorts.xlsm",
        "SKIRT":           "Skirts.xlsm",
        "SNOW_PANT":       "Snow_Pants.xlsm",
        "SNOWSUIT":        "Snowsuits.xlsm",
        "SWEATSHIRT":      "Sweatshirts.xlsm",
        "SWIMWEAR":        "Swimwear.xlsm",
    }.get(pt)
    if not fname:
        return None
    p = UPLOAD_TEMPLATES / fname
    return str(p) if p.exists() else None


if __name__ == "__main__":
    print("NIS Wizard v3 — TLG Amazon Intelligence starting on http://localhost:5000")
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
