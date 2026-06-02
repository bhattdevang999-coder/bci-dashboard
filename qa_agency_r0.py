"""qa_agency_r0.py — Tahari R0 regression harness for Sheik's agency feedback.

This is the running regression set that grows across Passes 1-6 of the
agency-feedback sprint. Every pass that claims to fix an item must add an
assertion here that proves the fix on the Tahari R0 fixture.

Fixture:
    data/fixtures/tahari_r0_preupload_template_2026_05_14.xlsx
    data/fixtures/tahari_r0_feedback_2026_05_14.xlsx

Items tracked:
    See docs/agency_r0_tracker.md for the 15 items, their root causes, and
    the pass each is fixed in.

Pass status:
    Pass 0 — fixture loads, 0 items fixed
    Pass 1 — items 5, 8, 9, 11, 13 (parser + verdict)
    Pass 2 — items 1, 2, 3, 10, 14 (taxonomy)
    Pass 3 — items 11, 15 (multi-value)
    Pass 4 — Strategic 1 (scope-aware editing)
    Pass 5 — items 6, 7, 12 (cosmetic)
    Pass 6 — Strategic 2 (template v2)

Run:
    python qa_agency_r0.py
    python qa_agency_r0.py --pass 1     # gate assertions to a single pass
    python qa_agency_r0.py --verbose
"""
from __future__ import annotations

import argparse
import os
import sys
import traceback
from pathlib import Path

ROOT = Path(__file__).resolve().parent
FIXTURES = ROOT / "data" / "fixtures"
TEMPLATE_FIXTURE = FIXTURES / "tahari_r0_preupload_template_2026_05_14.xlsx"
FEEDBACK_FIXTURE = FIXTURES / "tahari_r0_feedback_2026_05_14.xlsx"

# Make repo importable
sys.path.insert(0, str(ROOT))


# ============================================================================
# Assertion framework
# ============================================================================

_results: list[tuple[str, str, str, str | None]] = []
# (pass_label, item_id, name, error_message_or_None)


def assert_eq(actual, expected, msg: str = ""):
    if actual != expected:
        raise AssertionError(f"{msg or 'expected =='} → expected={expected!r}, actual={actual!r}")


def assert_in(needle, haystack, msg: str = ""):
    if needle not in haystack:
        raise AssertionError(f"{msg or 'expected in'} → {needle!r} not in {haystack!r}")


def assert_truthy(v, msg: str = ""):
    if not v:
        raise AssertionError(f"{msg or 'expected truthy'} → got {v!r}")


def assert_not_in(needle, haystack, msg: str = ""):
    if needle in haystack:
        raise AssertionError(f"{msg or 'expected not in'} → {needle!r} present in {haystack!r}")


def case(pass_label: str, item_id: str, name: str):
    """Decorator: register a test case under a pass + item."""
    def wrap(fn):
        def runner():
            try:
                fn()
                _results.append((pass_label, item_id, name, None))
                return True
            except AssertionError as e:
                _results.append((pass_label, item_id, name, str(e)))
                return False
            except Exception as e:
                _results.append((pass_label, item_id, name, f"UNEXPECTED: {type(e).__name__}: {e}"))
                return False
        runner.pass_label = pass_label
        runner.item_id = item_id
        runner.name = name
        return runner
    return wrap


# ============================================================================
# PASS 0 — fixture sanity (the only assertions today)
# ============================================================================

@case("0", "fixture", "Tahari preupload template fixture present")
def test_fixture_template_present():
    assert_truthy(TEMPLATE_FIXTURE.exists(), f"missing fixture: {TEMPLATE_FIXTURE}")


@case("0", "fixture", "Tahari feedback fixture present")
def test_fixture_feedback_present():
    assert_truthy(FEEDBACK_FIXTURE.exists(), f"missing fixture: {FEEDBACK_FIXTURE}")


@case("0", "fixture", "Preupload template parses with openpyxl")
def test_fixture_template_loads():
    from openpyxl import load_workbook
    wb = load_workbook(str(TEMPLATE_FIXTURE), data_only=True)
    ws = wb["Pre-Upload Template"]
    # First row should be the header row
    headers = [c.value for c in ws[1]]
    assert_in("Brand Code", headers)
    assert_in("STYLE#", headers)
    assert_in("Closure Type", headers)
    assert_in("Fabric Content Percentage", headers)
    # Should have >= 20 real data rows (Tahari had 23 styles)
    nonempty = sum(1 for row in ws.iter_rows(min_row=4, values_only=True) if row and row[0])
    if nonempty < 20:
        raise AssertionError(f"expected ≥20 data rows, got {nonempty}")


@case("0", "fixture", "Feedback fixture has the 15 numbered items + final thoughts")
def test_fixture_feedback_loads():
    from openpyxl import load_workbook
    wb = load_workbook(str(FEEDBACK_FIXTURE), data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    # Issue # column is the first; items 1..16
    issue_nums = [r[0] for r in rows if r[0] is not None]
    nums_seen = set()
    for n in issue_nums:
        try:
            nums_seen.add(int(n))
        except (ValueError, TypeError):
            pass
    expected = set(range(1, 17))
    missing = expected - nums_seen
    if missing:
        raise AssertionError(f"feedback missing items: {sorted(missing)}")


# ============================================================================
# PASS 1 — placeholders (each becomes a real assertion when Pass 1 ships)
# ============================================================================

def _load_tahari_sample():
    """Parse the fixture and return (sample_style_dict, state, evaluation).

    Picks the first real (non-sentinel) Tahari style with populated fabric+closure.
    Cached at module level on first call.
    """
    if hasattr(_load_tahari_sample, "_cache"):
        return _load_tahari_sample._cache
    from nis_engine.preupload_importer import parse_preupload, style_to_form_state
    from nis_engine import nis_rule_engine as _nis_engine
    _NIS_RULES_DIR = ROOT / "nis_rules"
    if _NIS_RULES_DIR.exists():
        _nis_engine.set_bundle_dir(str(_NIS_RULES_DIR))
    parsed = parse_preupload(str(TEMPLATE_FIXTURE))
    styles = parsed.get("styles", {})
    sample = None
    for sid, s in styles.items():
        if "TAHARI" in (s.get("tlgdiv") or "").upper() and s.get("fabric") and s.get("closure"):
            sample = s
            break
    if sample is None:
        # Fall back to any first style
        sample = next(iter(styles.values()))
    brand = "Tahari" if "TAHARI" in (sample.get("tlgdiv") or "").upper() else (parsed.get("brand") or "")
    state = style_to_form_state(sample, brand)
    evaluation = _nis_engine.evaluate_form(
        "COAT", state,
        apply_apparel_defaults=True,
        brand=brand,
        sub_class=sample.get("sub_class") or "",
    )
    _load_tahari_sample._cache = (sample, state, evaluation)
    return sample, state, evaluation


@case("1", "item_5", "Engine Verdict reports zero false-positive blockers on Tahari R0")
def test_pass1_item5_no_false_blockers():
    _, state, evaluation = _load_tahari_sample()
    summary = evaluation.get("summary", {})
    missing = summary.get("required_missing", 0)
    assert_eq(missing, 0, f"expected 0 required_missing, got {missing}")
    # Also confirm all 5 bullets are in state (the root cause)
    for n in range(1, 6):
        v = state.get(f"bullet_point#{n}.value")
        assert_truthy(v, f"bullet_point#{n}.value missing in state")


@case("1", "item_8", "Department + Target Gender flow into state correctly")
def test_pass1_item8_dept_gender_flow():
    _, state, _ = _load_tahari_sample()
    assert_eq(state.get("department#1.value"), "Womens")
    assert_eq(state.get("target_gender#1.value"), "Female")
    # Also confirm item_name has the right gender word (no "Tahari Men's" bug)
    name = state.get("item_name#1.value") or ""
    assert_in("Women's", name, "item_name should include 'Women's', not 'Men's'")
    assert_not_in("Men's", name, "item_name should NOT include 'Men's' for women's product")


@case("1", "item_9", "Closure Type from template surfaces in state under bundle field_key")
def test_pass1_item9_closure_flow():
    _, state, evaluation = _load_tahari_sample()
    closure_val = state.get("closure#1.type#1.value")
    assert_truthy(closure_val, "closure#1.type#1.value should be populated from template")
    assert_eq(closure_val, "Zipper", f"expected 'Zipper' from Tahari Vera puffer, got {closure_val!r}")
    # Also verify evaluation can find it (the engine view renders from this)
    fields = evaluation.get("fields", {})
    closure_field = None
    for col, f in fields.items():
        if (f.get("label") or "").lower() == "closure type":
            closure_field = f
            break
    assert_truthy(closure_field, "evaluation must surface a Closure Type field")
    assert_eq(closure_field.get("value"), "Zipper")


@case("1", "item_11_a", "Material field receives parsed material names (not percentage string)")
def test_pass1_item11_material_parse():
    _, state, _ = _load_tahari_sample()
    material = state.get("material#1.value")
    fabric_type = state.get("fabric_type#1.value")
    assert_eq(material, "Polyester", f"material#1 should be 'Polyester' (parsed name), got {material!r}")
    assert_eq(fabric_type, "100% Polyester", f"fabric_type#1 should be '100% Polyester' (composition), got {fabric_type!r}")
    # Sanity-test the parser on a multi-material string
    from nis_engine.preupload_importer import _split_fabric_into_materials
    names, comp = _split_fabric_into_materials("95% Polyester, 5% Spandex")
    assert_eq(names, ["Polyester", "Spandex"])
    assert_eq(comp, "95% Polyester, 5% Spandex")


@case("1", "item_13", "Sleeve Type parses from template into state under bundle field_key")
def test_pass1_item13_sleeve_length():
    _, state, _ = _load_tahari_sample()
    sleeve_type = state.get("sleeve#1.type#1.value")
    assert_eq(sleeve_type, "Long Sleeve", f"expected 'Long Sleeve', got {sleeve_type!r}")
    # Verify the sleeve length code path exists: when a fixture row HAS a sleeve
    # length value, it flows to the right field key. The v1 template doesn't have
    # a sleeve_length column today, so we test with a synthetic row.
    from nis_engine.preupload_importer import style_to_form_state
    fake_row = {"style": "TEST1", "name": "X", "sub_class": "Puffer", "sleeve_length": "Hip Length"}
    synth_state = style_to_form_state(fake_row, "Tahari")
    assert_eq(synth_state.get("sleeve#1.length_description#1.value"), "Hip Length",
              "sleeve_length should flow when source has a value")


# ============================================================================
# PASS 2 — taxonomy unsticking (placeholders)
# ============================================================================

_INDEX_HTML_CACHE = None
def _load_index_html():
    global _INDEX_HTML_CACHE
    if _INDEX_HTML_CACHE is None:
        _INDEX_HTML_CACHE = (ROOT / "templates" / "index.html").read_text(encoding="utf-8")
    return _INDEX_HTML_CACHE


@case("2", "item_1", "ITK datalist is PT-aware; no hardcoded swimwear list")
def test_pass2_item1_itk_dropdown_reset():
    # Indirect frontend check: the previously hardcoded swimwear array
    # ['rash-guards','swim-trunks',...] must no longer appear inside
    # taxonomyEnsureITKDatalist. The replacement reads from
    # taxonomyState.universe[pt].item_type_keywords_by_cat_sub at call time.
    html = _load_index_html()
    # The hardcoded list literal is removed
    assert_not_in("'rash-guards','swim-trunks','board-shorts'", html,
                  "hardcoded swim ITK list should be removed from taxonomyEnsureITKDatalist")
    # The new function takes a PT arg and reads from the universe
    assert_in("function taxonomyEnsureITKDatalist(pt)", html,
              "taxonomyEnsureITKDatalist should accept a PT argument")
    assert_in("item_type_keywords_by_cat_sub", html,
              "datalist builder should pull from item_type_keywords_by_cat_sub")
    # Caller passes PT through
    assert_in("taxonomyEnsureITKDatalist(pt);", html,
              "taxonomyRenderForStyle should call taxonomyEnsureITKDatalist(pt)")


@case("2", "item_2", "Manual PT selection on UNKNOWN row repopulates Cat/Subcat/ITK cascade")
def test_pass2_item2_unknown_recovery():
    # wsStylePTChanged must now update taxonomyState.styleMeta[styleNum] and
    # trigger taxonomyRenderForStyle so the cascade rebuilds from universe[newPT].
    html = _load_index_html()
    # Find the function body
    start = html.find("function wsStylePTChanged(styleNum, newPT)")
    assert_truthy(start >= 0, "wsStylePTChanged function not found")
    body = html[start:start + 3000]
    assert_in("taxonomyState.styleMeta[styleNum]", body,
              "wsStylePTChanged should sync taxonomyState.styleMeta")
    assert_in("taxonomyRenderForStyle(styleNum)", body,
              "wsStylePTChanged should re-render the taxonomy panel")
    assert_in("meta.product_type = newPT", body,
              "wsStylePTChanged should overwrite meta.product_type with newPT")


@case("2", "item_3", "Bulk-taxonomy view: UNKNOWN PT manual choice flows taxonomy options")
def test_pass2_item3_bulk_taxonomy_unknown():
    # Bulk modal must expose an inline PT picker for UNKNOWN buckets via
    # taxonomyBulkOnPTChange that propagates to wsState + per-style meta.
    html = _load_index_html()
    assert_in("function taxonomyBulkOnPTChange(i, newPT)", html,
              "bulk modal needs a taxonomyBulkOnPTChange handler")
    start = html.find("function taxonomyBulkOnPTChange(i, newPT)")
    body = html[start:start + 2500]
    assert_in("wsState.styleProductTypes", body,
              "taxonomyBulkOnPTChange should propagate PT to wsState")
    assert_in("taxonomyBuildBulkModalHTML", body,
              "taxonomyBulkOnPTChange should re-render the bulk modal")
    # And the row renderer should expose the picker when PT is UNKNOWN
    assert_in("isUnknownPT", html,
              "bulk modal row renderer should branch on UNKNOWN PT")


@case("2", "item_10", "Item Length Description default reflects current PT, not previous PT")
def test_pass2_item10_item_length_pt_aware():
    # Direct backend assertion: _derive_item_length must return PT-appropriate
    # vocabulary, not 'Knee-Length' for COAT.
    sys.path.insert(0, str(ROOT))
    import importlib
    # app.py is huge; import lazily and only what we need
    import app as _app
    importlib.reload(_app) if "_derive_item_length" not in dir(_app) else None
    fn = _app._derive_item_length
    # COAT with no length hint → Standard Length, NOT Knee-Length (the old bug)
    assert_eq(fn("", "Vera Quilted Puffer", product_type="COAT"),
              "Standard Length",
              "COAT default should be 'Standard Length' (was leaking 'Knee-Length' from DRESSES)")
    # PANTS/SHIRT/BLAZER/SHORTS all use the same 'X Length' vocabulary
    assert_eq(fn("", "Slim Stretch Pant", product_type="PANTS"), "Standard Length")
    assert_eq(fn("", "Oxford Shirt", product_type="SHIRT"), "Standard Length")
    # DRESS keeps the old adjective vocabulary
    assert_eq(fn("", "Linen Maxi Dress", product_type="DRESS"), "Maxi")
    assert_eq(fn("", "Midi Wrap Dress", product_type="DRESS"), "Midi")
    assert_eq(fn("", "Solid Sheath Dress", product_type="DRESS"), "Knee-Length")
    # SWIMWEAR stays blank
    assert_eq(fn("", "Rashguard Set", product_type="SWIMWEAR"), "")
    # UNKNOWN PT → blank (don't guess across vocabularies)
    assert_eq(fn("", "Mystery Item", product_type=""), "")
    # Cropped hint on COAT → Short Length
    assert_eq(fn("", "Cropped Bomber", product_type="COAT"), "Short Length")


@case("2", "item_14", "UNKNOWN taxonomy editable from style-level AND bulk views")
def test_pass2_item14_unknown_editable():
    # Per-style panel must show a helper instead of locked empty selects
    # when PT is UNKNOWN. Bulk modal must offer the PT recovery picker.
    html = _load_index_html()
    # Find the taxonomyRenderForStyle function
    start = html.find("function taxonomyRenderForStyle(styleNum)")
    assert_truthy(start >= 0, "taxonomyRenderForStyle not found")
    body = html[start:start + 6000]
    assert_in("pt === 'UNKNOWN'", body,
              "taxonomyRenderForStyle should branch on UNKNOWN PT")
    assert_in("Set Product Type first", body,
              "taxonomyRenderForStyle should show an honest UNKNOWN-PT helper")
    # Bulk-modal side: covered by item 3 picker (already asserted there)
    assert_in("function taxonomyBulkOnPTChange(i, newPT)", html,
              "bulk modal needs its own PT recovery path")


# ============================================================================
# PASS 3 — multi-value (placeholders)
# ============================================================================

@case("3", "item_11_b", "Material parses to ordered array; multi-material round-trips")
def test_pass3_item11_material_multi():
    from nis_engine.preupload_importer import style_to_form_state, _split_fabric_into_materials
    # 1. Single fabric cell with three materials → material#1/2/3 all populated
    fake_row = {
        "style": "M3TEST", "name": "Tri-blend Puffer",
        "sub_class": "Puffer",
        "fabric": "80% Polyester, 15% Cotton, 5% Spandex",
        "closure": "Zipper",
    }
    state = style_to_form_state(fake_row, "Tahari")
    assert_eq(state.get("material#1.value"), "Polyester",
              "material#1 should be first parsed name")
    assert_eq(state.get("material#2.value"), "Cotton",
              "material#2 should be second parsed name")
    assert_eq(state.get("material#3.value"), "Spandex",
              "material#3 should be third parsed name")
    assert_eq(state.get("fabric_type#1.value"), "80% Polyester, 15% Cotton, 5% Spandex",
              "fabric_type#1 should keep the original composition string")

    # 2. v2 template with explicit Material 2/3 columns — explicit wins
    fake_v2 = {
        "style": "M3V2", "name": "v2 Test", "sub_class": "Puffer",
        "fabric": "100% Polyester",
        "material_2": "Wool",
        "material_3": "Cashmere",
        "closure": "Zipper",
    }
    state_v2 = style_to_form_state(fake_v2, "Tahari")
    assert_eq(state_v2.get("material#1.value"), "Polyester")
    assert_eq(state_v2.get("material#2.value"), "Wool",
              "v2 explicit Material 2 column should populate material#2")
    assert_eq(state_v2.get("material#3.value"), "Cashmere",
              "v2 explicit Material 3 column should populate material#3")

    # 3. Parser direct test on tougher inputs
    names, comp = _split_fabric_into_materials("65%Cotton/30%Polyester/5%Spandex")
    assert_eq(names, ["Cotton", "Polyester", "Spandex"],
              "slash-separated multi should parse")
    assert_eq(comp, "65%Cotton/30%Polyester/5%Spandex")


@case("3", "item_15", "Closure Type accepts and persists multiple values")
def test_pass3_item15_closure_multi():
    from nis_engine.preupload_importer import style_to_form_state, _split_closures

    # 1. v1 template encoding: comma-separated multi in single Closure Type cell
    fake_v1 = {
        "style": "C2TEST", "name": "Multi-Closure Puffer",
        "sub_class": "Puffer",
        "fabric": "100% Polyester",
        "closure": "Zipper, Snap",
    }
    state = style_to_form_state(fake_v1, "Tahari")
    assert_eq(state.get("closure#1.type#1.value"), "Zipper",
              "comma-encoded multi should put first value into type#1")
    assert_eq(state.get("closure#1.type#2.value"), "Snap",
              "comma-encoded multi should put second value into type#2")

    # 2. v2 template with explicit Closure Type 2 column
    fake_v2 = {
        "style": "C2V2", "name": "v2 Closure Test",
        "sub_class": "Puffer",
        "fabric": "100% Polyester",
        "closure": "Zipper",
        "closure_2": "Hook & Eye",
    }
    state_v2 = style_to_form_state(fake_v2, "Tahari")
    assert_eq(state_v2.get("closure#1.type#1.value"), "Zipper")
    assert_eq(state_v2.get("closure#1.type#2.value"), "Hook & Eye",
              "v2 explicit Closure Type 2 should populate type#2")

    # 3. Hook & Eye must NOT split on '&' — it's a single Amazon value
    out = _split_closures("Hook & Eye, Snap")
    assert_eq(out, ["Hook & Eye", "Snap"],
              "_split_closures must preserve 'Hook & Eye' as one value")

    # 4. Single value stays single
    fake_single = {"style": "C1", "name": "X", "sub_class": "Puffer",
                   "fabric": "100% Polyester", "closure": "Zipper"}
    s_single = style_to_form_state(fake_single, "Tahari")
    assert_eq(s_single.get("closure#1.type#1.value"), "Zipper")
    assert_truthy(s_single.get("closure#1.type#2.value") in (None, ""),
                  "closure#1.type#2 must not be set when source has one value")


# ============================================================================
# PASS 4 — scope-aware editing
# ============================================================================

@case("4", "strategic_1", "scope=brand_always edit on style #1 propagates to style #2")
def test_pass4_scope_brand_always_reads_through():
    raise AssertionError("PENDING — Pass 4 not started")


# ============================================================================
# PASS 5 — cosmetic
# ============================================================================

@case("5", "item_6", "Style blocks render without hover-only blackout")
def test_pass5_item6_style_block_visible():
    raise AssertionError("PENDING — Pass 5 not started")


@case("5", "item_7", "Bullet formatter does not add ALLCAPS+dash when ALLCAPS+colon present")
def test_pass5_item7_bullet_colon_separator():
    # When this lands, the assertion will look like:
    #   from nis_engine.content_rules import format_bullet
    #   out = format_bullet("ELEVATED WARMTH: The quilted puffer construction...")
    #   assert_truthy(out.startswith("ELEVATED WARMTH:"))
    #   assert_not_in("ELEVATED WARMTH: THE —", out)
    #   assert_not_in("ELEVATED WARMTH —", out)
    raise AssertionError("PENDING — Pass 5 not started")


@case("5", "item_12", "All Fields > Content reads operator-edited value, not original (or removed)")
def test_pass5_item12_content_tab_sync():
    raise AssertionError("PENDING — Pass 5 not started")


# ============================================================================
# PASS 6 — template v2
# ============================================================================

@case("6", "strategic_2", "v2 template parses cleanly with new Amazon-attribute columns")
def test_pass6_template_v2_parse():
    raise AssertionError("PENDING — Pass 6 not started")


@case("6", "strategic_2", "v1 template still parses via mapping layer with no data loss")
def test_pass6_template_v1_mapping_lossless():
    raise AssertionError("PENDING — Pass 6 not started")


# ============================================================================
# Runner
# ============================================================================

ALL_CASES = [
    test_fixture_template_present,
    test_fixture_feedback_present,
    test_fixture_template_loads,
    test_fixture_feedback_loads,
    test_pass1_item5_no_false_blockers,
    test_pass1_item8_dept_gender_flow,
    test_pass1_item9_closure_flow,
    test_pass1_item11_material_parse,
    test_pass1_item13_sleeve_length,
    test_pass2_item1_itk_dropdown_reset,
    test_pass2_item2_unknown_recovery,
    test_pass2_item3_bulk_taxonomy_unknown,
    test_pass2_item10_item_length_pt_aware,
    test_pass2_item14_unknown_editable,
    test_pass3_item11_material_multi,
    test_pass3_item15_closure_multi,
    test_pass4_scope_brand_always_reads_through,
    test_pass5_item6_style_block_visible,
    test_pass5_item7_bullet_colon_separator,
    test_pass5_item12_content_tab_sync,
    test_pass6_template_v2_parse,
    test_pass6_template_v1_mapping_lossless,
]


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--pass", dest="pass_filter", default=None,
                   help="Only run cases for this pass label (e.g. 0, 1, 2). Default: all.")
    p.add_argument("--verbose", "-v", action="store_true")
    args = p.parse_args(argv)

    cases = [c for c in ALL_CASES
             if args.pass_filter is None or c.pass_label == args.pass_filter]

    print(f"\nqa_agency_r0.py — Tahari R0 regression harness")
    print(f"running {len(cases)} cases" + (f" (pass={args.pass_filter})" if args.pass_filter else ""))
    print("=" * 72)

    for c in cases:
        c()

    # Group results by pass
    by_pass: dict[str, list] = {}
    for entry in _results:
        by_pass.setdefault(entry[0], []).append(entry)

    pending_count = 0
    pass_count = 0
    fail_count = 0
    for pass_label in sorted(by_pass.keys()):
        print(f"\nPass {pass_label}:")
        for _, item, name, err in by_pass[pass_label]:
            if err is None:
                status = "PASS"
                pass_count += 1
            elif err.startswith("PENDING"):
                status = "PEND"
                pending_count += 1
            else:
                status = "FAIL"
                fail_count += 1
            print(f"  [{status}] {item:<14s} {name}")
            if err and args.verbose:
                print(f"           └─ {err}")

    print("\n" + "=" * 72)
    print(f"summary: {pass_count} pass, {fail_count} fail, {pending_count} pending")
    # Pending isn't a failure — passes 1-6 will turn pending into pass
    return 1 if fail_count else 0


if __name__ == "__main__":
    sys.exit(main())
